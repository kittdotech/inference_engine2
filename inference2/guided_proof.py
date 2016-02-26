from openpyxl import load_workbook

import timeit
import copy

cond_r = unichr(8835)
top = unichr(8868)
bottom = unichr(8869)
neg = unichr(172)
iff = unichr(8801)
mini_c = unichr(8658)
mini_e = unichr(8703)
implies = unichr(8866)
conditional = unichr(8594)
nonseq = unichr(8876)
xorr = unichr(8891)
idisj = unichr(8744)
cj = unichr(8896)
disj = unichr(8855)
equi = unichr(8660)
#variable initiated by rajiv
rel_conj = []
sn=0
############

pp = 7
#
# >> 8835
# ta^ 8868
# co^ 8869
# ; 172
# <> 8801
# c^ 8658
# # 8703
# i^ 8866
# > 8594
# nf^ 8876
# ed^ 8891
# + 8744
# && 8896
# @ 8855
# if^ 8660



def remove_outer_paren(str1):

    if str1 == "":
        return ""
    elif str1.count(")") == 0:
        return str1

    j = 0
    if str1[0] != "(" and str1[-1] != ")":
        return str1

    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1
        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
    return str1


def remove_redundant_paren(str1):
    j = 0
    str2 = str1[:2]
    str3 = str1[2:]
    if str2 == '((' and str3 == '))':

        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 1 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
    return str1


def mainconn(str1):
    # str_main = str_main.replace(implies,"&")
    # str_main = str_main.replace(nonseq, "&")
    # str_main = str_main.replace("~","")
    # str_main = str_main.replace(" ", "")

    ostring = copy.copy(str1)
    if os(str1):
        return ["", 0]
    if str1.find("&") < -1 and str1.find(idisj) < -1 and str1.find(iff) < -1 and str1.find(conditional) < -1 and \
            str1.find(implies) < -1 and str1.find(nonseq) < -1 and str1.find(xorr) < -1:
        return ["", 0]

    str3 = str1
    bool1 = False

    if str1[0] == "~":
        str1 = str1[1:]
        bool1 = True

    # list1 = tilde_removal2(str1)
    # if list1[1] == "~":
    #     str1 = str1[1:]
    #     bool1 = True

    j = 0
    bool2 = False
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
            bool2 = True


    j = -1

    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == conditional:
            f = -1
        if str2 == idisj or str2 == "&" or str2 == iff or str2 == implies or \
                str2 == nonseq or str2 == conditional or str2 == xorr:
            if str3 != str2 and str3 != "":
                j = j + 1

            str3 = str2

    if j == -1:
        i = ostring.find(str3)
        return [str3, i]
    k = -1
    j = -1
    while True:
        k += 1
        if k > 150:
            break
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]

            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1

            if j == -1 and (str2 == idisj or str2 == "&" or str2 == iff or str2 == implies
                            or str2 == nonseq or str2 == conditional or str2 == xorr):
                if bool1 and bool2:
                    return [str2, i+2]
                elif bool2 or bool1:
                    return [str2, i+1]
                else:
                    return [str2, i]
        else:
            str1 = str1[1:-1]

def isvariable(str3):
    bool2 = True
    if str3 == 'a':
        return False

    if str3 != "":
        str3 = str3.replace("'", "")
        str3 = str3.replace(neg, "")
        if len(str3) == 1 and str3.islower():
            bool2 = True
        else:
            bool2 = False
    return bool2


def os(str1):
    cnx = [xorr, iff, idisj, conditional, implies, nonseq, "&"]
    for i in range(0, len(cnx)):
        if str1.find(cnx[i]) > -1:
            os = False
            return os
    os = True
    return os

def enclose(str1):

    i = -1
    while i < len(str1) -1:
        i += 1
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        if str2.islower():
            if str3 == "~":
                str1 = str1[:i-1] + "(~" + str2 + ")" + str1[i+1:]
            else:
                str1 = str1[:i] + "(" + str2 + ")" + str1[i+1:]
            i += 3
    return str1

def find_sentences(instring, cut_skel = False):

    g = instring.count('(')
    h = instring.count(')')
    if g > h:
        easygui.msgbox('there are more open parentheses than closed')
    elif g < h:
        easygui.msgbox('there are more closed parentheses than closed')
    marker = False
    il = -1
    total = -1
    c = -1
    neg_value = []
    str1 = ""
    sent1 = []
    sent_type2 = []
    wneg = []
    output = [None] * 7
    # the skel name list names each single sentence after a greek letter, even if
    # the same sentence appears twice it obtains a different name on the second
    # appearance
    skel_nam = []
    sent_num = []
    if instring.find("~(") > -1:
        instring = instring.replace("~(", "(!")
    if instring.find(implies) > -1:
        str2 = implies
    elif instring.find(nonseq) > -1:
        str2 = nonseq
    str3 = mainconn(instring)
    str4 = str3[0]
    f = str3[1]
    id_num = []
    id_num.append(["1",str4,f])
    sent_num.append([1, '1', instring, str4,f])
    prtnum = 1
    skel_string = instring
    p = 947
    connectives = ["&", idisj, iff, conditional, nonseq, implies]
    arr1 = []
    mini_c2 = mini_c + neg
    instring2 = copy.copy(instring)
    instring = instring.strip()
    prt = copy.copy(instring)
    list1 = mainconn(instring)
    grandparent_type = list1[0]

    if os(instring) == False:
        temp_string = mainconn(instring)

        if instring.find(implies) > -1:
            str1 = implies
        elif instring.find(nonseq) > -1:
            str1 = nonseq
        else:
            if temp_string == iff:
                str1 = "bicond"
            elif temp_string == conditional:
                str1 = "cond"
            elif temp_string == "&":
                str1 = "cj"

        sent1.append(instring)
        neg_value.append("")
        sent_type2.append(str1)
        wneg.append(instring)
        skel_nam.append(None)

    j = 0
    n = 0
    for i in range(0, len(instring)):
        str1 = instring[i:(i + 1)]
        for o in connectives:
            if str1 == o:
                j += 1

    while n < j + 1:

        il += 1
        if il > 15:
            break

        e = 0
        l = len(instring)
        x = -1
        while x < l - 1:
            x += 1
            temp_string = instring[x: x + 1]
            if instring[x: x + 1] == "(":

                if marker == False:
                    z = x
                    marker = True

                total += 1
            elif instring[x: x + 1] == ")":
                total -= 1
                if total == -1:
                    marker = False
                    e += 1
                    c += 1

                    temp_sent = instring[z: x + 1]
                    otemp_sent = copy.copy(temp_sent)

                    if (len(instring) - len(temp_sent)) > 2:

                        if temp_sent in prt:
                            prtnum = findinlist(prt,sent_num,2,1,False)
                            numb = str(prtnum) + "1"
                        else:
                            prtnum = ""
                            for bb in range(len(sent_num)-1,-1,-1):
                                str3 = sent_num[bb][2]

                                if temp_sent in sent_num[bb][2]:
                                    prtnum = sent_num[bb][1]
                                    break
                            # if prtnum == "":
                            #     easygui.msgbox('your sentences are not numbered properly')
                            g = len(prtnum) + 1
                            f = 0

                            for bb in range(len(sent_num)-1,-1,-1):
                                temp_sn = sent_num[bb][1]
                                h = temp_sn[:g-1]
                                hh = sent_num[bb][0]
                                if g > sent_num[bb][0]:
                                    break
                                if sent_num[bb][0] == g and temp_sn[:g-1] == prtnum:
                                    f += 1

                            f += 1
                            numb = prtnum + str(f)

                        if n == 7:
                            pp = 7

                        prt = temp_sent
                        temp_mc = mainconn(temp_sent)
                        mc = temp_mc[0]
                        str3 = temp_mc[0]
                        g = temp_mc[1]
                        mc_num = temp_mc[1]
                        sent_num.append([len(numb), numb, temp_sent, str3,g])


                        if os(temp_sent):
                            # n counts the number of single sentences
                            n += 1
                            if temp_sent.find("~") > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace("~", "")

                            elif temp_sent.find(mini_c2) > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace(mini_c2, mini_c)

                            elif temp_sent.find(mini_c2) == -1 and temp_sent.find(mini_c) > -1:
                                neg_value.append("")
                            elif temp_sent.find("~") == -1:
                                neg_value.append("")
                            else:
                                break  # stop
                        else:
                            neg_value.append("")

                        sent1.append(temp_sent)
                        wneg.append(otemp_sent)
                        id_num.append([numb,mc,mc_num])

                        if os(otemp_sent):
                            p += 1
                            skel_string = skel_string.replace(
                                otemp_sent, unichr(p))
                            q = [otemp_sent, unichr(p)]
                            skel_nam.append(q)
                            # skel_nam2[otemp_sent] = unichr(p)
                        else:
                            skel_nam.append(None)

                        # if k > 0:
                        #     m += 1
                            # children[k][m][0] = temp_sent
                            # children[k][m][1] = neg_value[c]

                    else:
                        instring = instring[1:len(instring) - 1]
                        l = len(instring)
                        x = -1
                        c -= - 1
                        e -= - 1

        total = -1
        marker = False
        w = -1

        if n < j + 1:
            if len(sent1) > w:
                while w + 1 < len(sent1):
                    if w == 13:
                        pp = 7
                    w += 1
                    str21 = sent1[w]
                    if not os(str21) and w != 0:
                        if str21 not in arr1:

                            instring = str21
                            arr1.append(instring)
                            break

    for i in range(len(sent1)):
        # if os(sent1[i]) == False:
        #     str1 = wneg[i]
        #     str2 = str1.replace("~", "")
        #     str2 = str2.repalce(mini_c2, mini_c)
        #     wneg[i] = str2

        temp_string = sent1[i]
        if temp_string.find("(!") > -1:
            sent1[i] = sent1[i].replace("(!", "~(")
            wneg[i] = wneg[i].replace("(!", "~(")

    if cut_skel == False:
        b = skel_string.find('(')
        skel_string = skel_string[b + 1:]
        skel_string.lstrip()
        skel_string = skel_string[0:-1]
    if skel_string.find("(!") > -1:
        skel_string = skel_string.replace("(!","~(")

    output[0] = sent1
    output[1] = neg_value
    output[2] = sent_type2
    output[3] = wneg
    output[4] = id_num
    output[5] = skel_string
    output[6] = skel_nam

    return output


def word_sub(def_var, dv_nam, tot_sent, all_sent):

    global m
    bool1 = False
    pronouns = ['ii','you','he','she','we','they']
    lemma = ['plural_counterpart', 'adjective_counterpart', 'verbal_counterpart']
    num = [5, 14, 18, 22, 26, 30, 34]
    global sn
    if nonconjunct:
        str4 = build_sent(all_sent[m])
        str4v = name_sent(str4)

    for i in num:
        str2 = all_sent[m][i]
        if str2 != None and str2 not in pronouns and str2 not in lemma:
            if isvariable(all_sent[m][i]) == False:
                bool1 = True
                telist7 = [def_var[0], all_sent[m][i]]
                all_sent[m][i] = def_var[0]
                del def_var[0]
                dv_nam.append(telist7)

    if bool1:
        anc1 = str(all_sent[m][36])
        rule = 'isub'
        sn += 1
        all_sent[m][36] = sn
        str1 = build_sent(all_sent[m])
        str1v = name_sent(str1)
        str3 = '1'
        if nonconjunct:
            str1 = str4 + ' ' + iff + ' ' + str1
            str1v = str4v + ' ' + iff + ' ' + str1v
            str3 = ''
            anc1 = ''
        tot_sent.append([sn, str1, str1v, rule, anc1,str3])

def assigned_var(str1, dv_nam, def_var):

    bool1 = False
    for i in range(len(dv_nam)):
        if dv_nam[i][1] == str1:
            bool1 = True
            return dv_nam[i][0]

    if bool1 == False:
        str2 = def_var[0]
        list1 = [str2, str1]
        dv_nam.append(list1)
        del def_var[0]
        return str2

def df_pronoun(tot_sent, all_sent, def_var, idf_var, dv_nam):

    global m

    num = [5,14,18,22,26,30,34]
    pronouns = ['ii','you','he','she','we','they']
    list2 = [None] * 55
    list3 = [None] * 55
    list1 = copy.deepcopy(all_sent[m])
    str2 = assigned_var('person',dv_nam, def_var)
    for i in num:
        str1 = all_sent[m][i]
        if str1 in pronouns:
            rsent = build_sent(all_sent[m])
            rsentv = name_sent(rsent)
            kind = 1
            list1[i] = def_var[0]
            str5 = def_var[0]
            list2 = member_sent(str5, str2, list2)
            all_sent[m][i] = str5
            del def_var[0]

            if str1 == 'he':
                male = assigned_var('male',dv_nam, def_var)
                kind = 2
                list3 = property_sent(str5, male, list3)
            elif str1 == 'she':
                female = assigned_var('female',dv_nam, def_var)
                kind = 2
                list3 = property_sent(str5, female, list3)
            elif str1 == 'we':
                bool1 = assign_var('we', str5, dv_nam)
                kind = 2
                list3 = member_sent(idf_var[0], str5, list3)
                del idf_var[0]
            elif str1 == 'they':
                bool1 = assign_var('they', str5, dv_nam)
                kind = 2
                list3 = member_sent(idf_var[0], str5, list3)
                del idf_var[0]
            elif str1 == 'ii':
                bool1 = assign_var('ii',str5,dv_nam)
                kind = 1
            elif str1 == 'you':
                bool1 = assign_var('you',str5,dv_nam)
                kind = 1

            rule = 'df ' + str1
            if kind == 1:
                list3 = []
                str3 = new_sent(list2, list3, list4, rule, all_sent, tot_sent,2, rsent, rsentv, False)
            else:
                str3 = new_sent(list2, list3, list4, rule, all_sent, tot_sent,3, rsent, rsentv, False)
            del all_sent[m][46][0]
            list3 = [None] * 55

def name_sent(str1):
    global propositions
    global prop_name
    global prop_var

    if str1.find('~') > -1:
        str1 = str1.replace("~","")
        ng = '~'
    else:
        ng = ''

    if str1 in propositions:
        for i in range(len(prop_name)):
            if str1 == prop_name[i][1]:
                return ng + prop_name[i][0]
    else:
        propositions.append(str1)
        prop_name.append([prop_var[0], str1])
        str2 = prop_var[0]
        del prop_var[0]
        return ng + str2

def assign_var(str1, str2, dv_nam):

    list1 = [str2, str1]
    dv_nam.append(list1)

def member_sent(str1, str2, list1):
    list1[5] = str1
    list1[9] = 'IG'
    list1[14] = str2
    return list1

def property_sent(str1, str2, list1):
    list1[5] = str1
    list1[9] = 'IA'
    list1[14] = str2
    return list1

def new_sent(list2, list3, list4, rule, all_sent, tot_sent, r, rsent, rsentv, bool1):
    # this function only works if the 1st sentence is complex and the 2nd and 3rd are in variable form
    # right now list4 has not function
    # the final bool1 is for whether or not &E should be performed
    global sn
    global osent
    sn += 1
    q = copy.copy(sn)
    list1 = copy.deepcopy(all_sent[m])
    str2 = build_sent(all_sent[m])
    list1[0] = str2
    str2v = name_sent(str2)
    str3 = build_sent(list2)
    list2[0] = str3
    if not check_dimension(osent,0, str2) and osentlist(list1):
        osent.append(list1)
    if not check_dimension(osent,0, str3) and osentlist(list2):
        osent.append(list2)
    str3v = name_sent(str3)
    anc1 = str(all_sent[m][36])
    if r == 2:
        if nonconjunct == False:
            if bool1:
                str4 = str3 + ' ' + conditional + ' ' + str2
                str4v = str3v + ' ' + conditional + ' ' + str2v
            else:
                str4 = str2 + " & " + str3
                str4v = str2v + " & " + str3v
        else:
            anc1 = " "
            if bool1:
                str4 = rsent + ' ' + iff + ' (' + str3 + ' ' + conditional + ' ' + str2 + ')'
                str4v = rsentv + ' ' + iff + ' (' + str3v + ' ' + conditional + ' ' + str2v + ')'
            else:
                str4 = rsent + ' ' + iff + ' (' + str2 + " & " + str3 + ')'
                str4v = rsentv + ' ' + iff + ' (' + str2v + " & " + str3v + ')'
    elif r == 3:
        str5 = build_sent(list3)
        list3[0] = str5
        if not check_dimension(osent, 0, str5) and osentlist(list3):
            osent.append(list3)
        str5v = name_sent(str5)
        if nonconjunct == False:
            str4 = str2 + " & " + str3 + " & " + str5
            str4v = str2v + " & " + str3v + " & " + str5v
        else:
            anc1 = " "
            str4 = rsent + " " + iff + " (" + str2 + " & " + str3 + " & " + str5 + ")"
            str4v = rsentv + " " + iff + " (" + str2v + " & " + str3v + " & " + str5v + ")"

    list7 = [sn, str4, str4v, rule, anc1]
    tot_sent.append(list7)
    if nonconjunct == False and bool1 == False:
        sn += 1
        all_sent[m][36] = sn
        rule = "&E"
        list1 = [sn, str2, str2v, rule, str(q)]
        tot_sent.append(list1)
        sn += 1
        rule = "&E"
        list1 = [sn, str3,str3v, rule, str(q)]
        tot_sent.append(list1)
    if r == 3 and nonconjunct == False and bool1 == False:
        sn += 1
        rule = "&E"
        list1 = [sn, str5, str5v, rule, str(q)]
        tot_sent.append(list1)

def determ(idf_var, def_var, all_sent, tot_sent):

    global det
    global m

    list2 = [None] * 55
    num = [3, 10, 16, 20, 24, 28, 32]
    for i in num:
        cond = False
        str8 = all_sent[m][i]
        if all_sent[m][i] in det:
            if nonconjunct:
                rsent = build_sent(all_sent[m])
                rsentv = name_sent(rsent)
            else:
                rsent = ''
                rsentv = ''
            if str8 == 'a' or str8 == 'any' or str8 == 'no':
                rule = 'df ' + str8
                str1 = idf_var[0]
                del idf_var[0]
                if str8 == 'any' or str8 == 'no':
                    cond = True
            elif all_sent[m][i] == 'the':
                rule = 'df the'
                str1 = def_var[0]
                del def_var[0]
            if i == 10:
                j = 14
            else:
                j = i + 2
            all_sent[m][i] = None
            str2 = all_sent[m][j]
            all_sent[m][j] = str1
            if str8 == 'no':
                if all_sent[m][9] == 'IG' or all_sent[m][9] == 'IA':
                    all_sent[m][12] = '~'
                else:
                    all_sent[m][8] = '~'

            list2 = member_sent(str1, str2, list2)
            list2[45] = all_sent[m][45]
            list3 = []
            list4 = []
            bool1 = new_sent(list2,list3,list4,rule,all_sent,tot_sent, 2, rsent, \
                             rsentv, cond)
            del all_sent[m][46][0]
            list2 = [None] * 55


def insert_space(str, integer):
    return str[0:integer] + ' ' + str[integer:]


def space_words(str1):
    # this function place a space between variables and relations
    # so as to make it easier to categorize words in a sentence
    str1 = str1.replace("~", " ~ ")
    str1 = str1.replace("=", " = ")
    i = 0
    while i + 1 < len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        nxt_str = str1[(i + 1):(i + 2)]
        if nxt_str.isupper() == True and temp_str.islower() == True:
            str1 = insert_space(str1, i + 1)
        elif nxt_str.islower() == True and temp_str.isupper() == True:
            str1 = insert_space(str1, i + 1)
        elif temp_str == "'" and nxt_str.isupper() == True:
            str1 = insert_space(str1, i + 1)
    return str1


def id_def(str1):
    # this function picks out that variables in the id sentences of the
    # definition

    b = str1.find('((')
    # str4 = str1[b + 1:]
    # str4.lstrip()
    # str4 = str4[0:-1]
    str1 = str1[0:b - 1]
    str1.rstrip()
    str1 = str1.replace(' ','')
    dv_nam2 = []
    beg = 1
    i = 1
    while i <= len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        if temp_str == "=" or i == len(str1) - 1:
            str2 = str1[beg:i]
            beg = i + 1
            while temp_str != ")":
                i += 1
                temp_str = str1[i:(i + 1)]
                if temp_str == ")":
                    str3 = str1[beg:i]
                    dv_nam2.append([str2,str3])
                    i += 2
                    beg = i + 1
                    break

    return dv_nam2


def is_atomic(list1, atomic_relations):

    relat = [9,15,19,23,27,31]
    must_be_blank = [2,3,4,6,7,10,11,12,16,17,20,21,24,25,28,29,32,33]
    noun = [5,14,18,26,30,34]
    for i in range(len(list1)):
        str1 = list1[i]
        if str1 != None:
            if i in relat:
                if str1 in atomic_relations:
                    pass
                else:
                    pass
            elif i in must_be_blank:
                return False
            elif i in noun:
                if not isvariable(str1):
                    return False
    return True

def osentlist(list1):

    # this list is for those sentences which can be defined.  relational particles
    # belong to this list because this list is also used when performing isub
    global atomic_relations
    global atomic_relata

    must_be_blank = [2,3,4,6,7,10,11,12,16,17,20,21,24,25,28,29,32,33,15,19,23,27,31]
    noun = [5,14,18,26,30,34]
    for i in range(1, len(list1)):
        str1 = list1[i]
        if str1 != None:
            if i ==9:
                if str1 == 'IG' or str1 == 'IA' or str1 == '=':
                    pass
                elif str1 in atomic_relations:
                    return False
                else:
                    pass
            elif i in must_be_blank:
                return False
            elif i in noun:
                if not isvariable(str1):
                    return False
    return True

def isdefineable(list1):

    global atomic_relations
    global atomic_relata

    must_be_blank = [2,3,4,6,7,10,11,12,16,17,20,21,24,25,28,29,32,33,15,19,23,27,31]
    noun = [5,14,18,26,30,34]
    for i in range(1, len(list1)):
        str1 = list1[i]
        if str1 != None:
            if i ==9:
                if str1 == 'IG' or str1 == 'IA' or str1 == '=':
                    return True
                elif str1 in atomic_relations:
                    return False
                else:
                    pass
            elif i in must_be_blank:
                return False
            elif i in noun:
                if not isvariable(str1):
                    return False
    return True

def build_sent(list1):

    str1 = "("
    for i in range(1, 34):
        temp_str = list1[i]
        if temp_str != None:
            if str1 == "(":
                str1 = str1 + temp_str
            else:
                str1 = str1 + " " + temp_str

    return str1 + ")"

def build_sent_list(list1):
    # this list builder does not have the ~ separated from the sentence
    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i]
        else:
            str2 = str2 + ' & ' + list1[i]
    return str2

def build_sent_list2(list1):
    # this list builder has the ~ separated from the sentence
    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i][1] + list1[i][0]
        else:
            str2 = str2 + ' & ' + list1[i][1] + list1[i][0]
    return str2

def id_sent(dv_nam):
    # this function turns the dv_nam into a string of conjuncts
    str2 = None
    for i in range(len(dv_nam)):
        str1 = '(' + dv_nam[i][0] + '=' + dv_nam[i][1] + ')'
        if str2 == None:
            str2 = str1
        else:
            str2 = str2 + ' & ' + str1
    return str2

def def_rn(defined, definiendum,e):
    # this function renames the variables in a definition
#     defined = u'''z'=moment, y'=individual | (bIGz') ''' + iff + ''' ((cAb) & (bAd) & (bIGy') & (((eAd) & \
# (cAe)) ''' + conditional + ''' (e=b)))'''

    # the dv list keeps track of all the definite variables used in a definition
    global tot_sent
    global sn
    global osent
    global dv_nam
    global def_var
    global idf_var

    used_idf = []
    new_idf = []
    rule = 'df ' + definiendum
    sn += 1
    str1 = copy.copy(defined)
    dv = id_def(str1)
    tot_sent.append([sn, defined + ' !', None, rule])
    def_info = find_sentences(defined)
    skel_string = def_info[5]
    # the following loop informs python which sentence to start with since we are not
    # going to worry about the id sentences
    for i in range(len(def_info[0])):
        if def_info[2][i] == 'bic1':
            start = i
            break
    else:
        easygui.msgbox('there is no biconditional in your definition')
    def_sent = []
    match_dv = []
    rename = []

    for i in range(start, len(def_info[0])):
        if os(def_info[3][i]) == True:
            temp_str = space_words(def_info[3][i])
            temp_str = temp_str.replace("(","")
            temp_str = temp_str.replace(")","")
            telist7 = temp_str.split(' ')
            telist7 = categorize_words(telist7)
            str1 = build_sent(telist7)
            telist7[0] = str1
            for j in range(0,3):
                telist7.append(None)
            def_sent.append(telist7)


    # we now must match the definite variables in the definition to the definite variables
    # already assigned
    list1 = []
    for i in range(len(dv)):
        temp_str = dv[i][1]
        for j in range(len(dv_nam)):
            dvn_temp = dv_nam[j][1]
            if dvn_temp == temp_str:
                telist7 = [dv[i][0],dv_nam[j][0]]
                if telist7 not in match_dv:
                    match_dv.append(telist7)
                    break
        else:
            if dv[i][0] not in def_var:
                telist7 = [dv[i][0], def_var[-1]]
                match_dv.append(telist7)
                list1.append([def_var[-1], temp_str])
                del def_var[-1]
            else:
                telist7 = [dv[i][0], temp_str]
                list1.append(telist7)
                def_var.remove(dv[i][0])

    # the purpose of this is that the subject of the definiendum must match the subject
    # of the osent to be defined
    if osent[e][5].find("'") > - 1:
        match_dv.append([def_sent[0][5],osent[e][5]])
    else:
        new_idf.append([def_sent[0][5],osent[e][5]])
        str2 = '(' + def_sent[0][5] + mini_c + osent[e][5] + ')'
        rename.append(str2)

    dv_nam = dv_nam + list1
    num = [5,14]
    for i in range(len(def_sent)):
        for j in num:
            temp_str = def_sent[i][j]
            if temp_str.find("'") > -1:
                for k in range(len(match_dv)):
                    if temp_str == match_dv[k][0]:
                        if def_sent[i][j] != match_dv[k][1]:
                            str2 = '(' + def_sent[i][j] + mini_c + match_dv[k][1] + ')'
                            def_sent[i][j] = match_dv[k][1]
                        if str2 not in rename and str2 != "":
                            rename.append(str2)
                        str2 = ""
                        break
            else:
                bool1 = check_dimension(new_idf,0,temp_str)
                if bool1:
                    str1 = findinlist(temp_str,new_idf,0,1)
                    def_sent[i][j] = str1
                elif temp_str in used_idf:
                    pass
                elif temp_str not in idf_var:
                    k = 0
                    while k + 1 <= len(new_idf):
                        if temp_str == new_idf[k][0]:
                            def_sent[i][j] = new_idf[k][1]
                            break
                        k += 1
                    else:
                        telist7 = [def_sent[i][j], idf_var[-1]]
                        new_idf.append(telist7)
                        str2 = "(" + def_sent[i][j] + mini_c + idf_var[-1] + ")"
                        def_sent[i][j] = idf_var[-1]
                        del idf_var[-1]
                        rename.append(str2)
                        str2 = ""
                elif temp_str in idf_var:
                    idf_var.remove(temp_str)
                    used_idf.append(temp_str)

    # for i in range(start, len(def_sent)):
    #     for j in num:
    #         temp_str = def_sent[i][j]
    #         if temp_str == telist7[1] and temp_str.find("*") == -1:
    #             def_sent[i][j] = telist7[0] + "*"


    j = 0
    for i in range(0, len(def_info[6])):
        if def_info[2][i] == 'bic1':
            m = i - 1
            for k in range(m, len(def_info[6])):
                if def_info[6][k] != None:
                    str2 = build_sent(def_sent[j])
                    skel_string = skel_string.replace(def_info[6][k][1], str2)
                    j += 1

    str2 = build_sent_list(rename)
    str2 = str2 + ' !'
    sn += 1
    tot_sent.append([sn, str2, None, 'RN'])
    sn += 1
    tot_sent.append([sn, skel_string, None, 'esub'])

    # right now we're simply deleting the first sentence of the def_sent since
    # we have already defined that but in the future when we work with more
    # complex definienda we will have to change this

    del def_sent[0]
    for i in range(len(def_sent)):
        str3 = build_sent(def_sent[i])
        def_sent[i][0] = str3
        bool1 = check_dimension(osent,0,str3)
        if bool1 == False:
            osent.append(def_sent[i])

def categorize_words(list1):

    global ind_num
    global words
    global ind_words
    global redundant
    word_types = []
    list2 = []

    for i in list1:
        str9 = i
        if isvariable(str9) == True:
            word_types.append([str9, 'n'])
        elif str9 == ' ':
            pass
        else:
            for j in range(len(words)):
                if j == 5:
                    pp = 7
                if i in words[j]:
                    if j == 0:
                        word_types.append([str9, 'a'])
                    elif j == 1:
                        word_types.append([str9, 'c'])
                    elif j == 2:
                        word_types.append([str9, 'd'])
                    elif j == 3:
                        word_types.append([str9, 'e'])
                    elif j == 4:
                        word_types.append([str9, 'l'])
                    elif j == 5:
                        word_types.append([str9, 'n'])
                    elif j == 6:
                        word_types.append([str9, 'r'])
                    elif j == 7:
                        word_types.append([str9, 's'])
                    elif j == 8:
                        word_types.append([str9, 't'])
                    elif j == 9:
                        word_types.append([str9, 'u'])
                    elif j == 10:
                        word_types.append([str9, 'b'])
                    elif j == 11:
                        word_types.append([str9, 'm'])
                    elif j == 12:
                        word_types.append([str9, 'q'])
                    if str9 in ind_words:
                        for k in range(len(ind_num)):
                            if str9 == ind_num[k][0]:
                                list2.append(ind_num[k][1])
                                break


    # a, c, d, e, l, n, r, s, t, u, b, m, q
    list1_cat = [None] * 55
    relation_type = None

    for i in range(0, len(word_types)):
        word = word_types[i][0]
        pos = word_types[i][1]
        if word in redundant:
            list2.append(3)
        elif word == '':
            pass
        elif pos == 'd' and relation_type == None:
            list1_cat[3] = word
        elif pos == 'a' and relation_type == None:
            list1_cat[4] = word
        elif pos == 'n' and relation_type == None:
            list1_cat[5] = word
        elif pos == 'r' and relation_type == None:
            list1_cat[9] = word
            relation_type = 'r'
        elif pos == 'd' and relation_type == 'r':
            list1_cat[10] = word
        # this line of code must be first because if the word is an adjective
        # and the relation is IA then it must go in slot 14
        elif pos == 'a' and relation_type == 'r' and list1_cat[9] == 'IA':
            list1_cat[14] = word
        elif pos == 'a' and relation_type == 'r':
            list1_cat[13] = word
        elif pos == 'n' and relation_type == 'r':
            list1_cat[14] = word
        elif pos == 'r' and relation_type == 'r':
            list1_cat[15] = word
            relation_type = 'r2'
        elif pos == 'd' and relation_type == 'r2':
            list1_cat[16] = word
            relation_type = 'r2'
        elif pos == 'a' and relation_type == 'r2':
            list1_cat[17] = word
            relation_type = 'r2'
        elif pos == 'n' and relation_type == 'r2':
            list1_cat[18] = word
            relation_type = 'r2'
        elif pos == 's':
            relation_type = 's'
            list1_cat[19] = word
        elif pos == 'd' and relation_type == 's':
            list1_cat[20] = word
        elif pos == 'a' and relation_type == 's':
            list1_cat[21] = word
        elif pos == 'n' and relation_type == 's':
            list1_cat[22] = word
        elif pos == 't':
            relation_type = 't'
            list1_cat[27] = word
        elif pos == 'd' and relation_type == 't':
            list1_cat[28] = word
        elif pos == 'a' and relation_type == 't':
            list1_cat[29] = word
        elif pos == 'n' and relation_type == 't':
            list1_cat[30] = word
        elif pos == 'b':
            list1_cat[7] = word
        elif pos == 'm':
            if relation_type == None:
                list1_cat[8] = word
            elif relation_type == 'r':
                list1_cat[13] = word
        else:
            easygui.msgbox('you did not categorize words properly')


    # we now need to specify if the sentence is subject to the rules of relation division
    if list1_cat[15] != None:
        list2.append(8)
    if list1_cat[9] == '=':
        list2.append(9)
    lemmas = ['plural_counterpart', 'adjective_counterpart', 'verbal_counterpart']
    if list1_cat[15] == 'OF' and list1_cat[9] == '=':
        if list1_cat[14] in lemmas:
            list2.append(4)

    list2.sort()
    list1_cat[46] = list2


    return list1_cat

def build_sent_name(prop_name):
    str1 = ''
    str2 = ''
    list1 = []
    for i in range(len(prop_name)):
        if i == 5:
            pp = 7
        str1 = '(' + prop_name[i][0] + mini_e + prop_name[i][1] + ')'
        if len(str2) == 0 and len(str1) > 57:
            list1.append(str1)
        elif (len(str2) + len(str1)) > 57:
            list1.append(str2)
            str2 = str1
            if i + 1 == len(prop_name):
                list1.append(str2)
        elif (len(str2) + len(str1)) < 57:
            if len(str2) == 0:
                str2 = str1
            else:
                str2 = str2 + ' & ' + str1
            if i + 1 == len(prop_name):
                list1.append(str2)
    return list1



def syn(tot_sent, all_sent):

    global m
    global synon
    global syn_pairs
    global sn
    global ind_words
    global ind_num
    global nonconjunct
    for i in range(3,34):
        str1 = all_sent[m][i]
        if str1 in synon:
            for j in range(len(syn_pairs)):
                if str1 == syn_pairs[j][0]:
                    if nonconjunct:
                        str3 = build_sent(all_sent[m])
                        str3v = name_sent(str3)
                    anc1 = copy.copy(all_sent[m][36])
                    sn += 1
                    anc2 = copy.copy(sn)
                    rule = 'df ' + str1
                    str5 = syn_pairs[j][2]
                    all_sent[m][i] = syn_pairs[j][1]
                    if str5 not in tot_sent:
                        str5v = name_sent(syn_pairs[j][2])
                        tot_sent.append([sn, str5, str5v, rule])
                        sn += 1
                    else:
                        pp = 7
                    str2 = build_sent(all_sent[m])
                    str2v = name_sent(str2)
                    if nonconjunct:
                        str2 = str3 + ' ' + iff + ' ' + str2
                        str2v = str3v + ' ' + iff + ' ' + str2v
                    all_sent[m][36] = sn
                    list3 = copy.deepcopy(all_sent[m][46])
                    if syn_pairs[j][1] in ind_words:
                        for k in range(len(ind_words)):
                            if syn_pairs[j][1] == ind_num[k][0]:
                                list3.append(ind_num[k][1])
                                list3.sort()
                                break
                    all_sent[m][46] = list3
                    rule = 'isub'
                    tot_sent.append([sn, str2, str2v, rule, str(anc1), str(anc2)])

def divide_sent(str1a):
    global impl
    list4 = []
    list5 = []
    j = 1
    if str1a[-1:] == "." and str1a.find(implies) == -1 and \
                    str1a.find(nonseq) == -1 and str1a.count(".") > 1:
        list5 = str1a.split('. ')
        for k in range(len(list5)):
            j += 1
            str1a = list5[k]
            str1a = str1a.strip()
            if str1a[-1:] == ".":
                str1a = str1a[0:-1]
            list4.append([str1a,'a', j])
    elif str1a.find(implies) > -1:
        list4 = str1a.split(' ' + implies + ' ')
        impl = implies
        list4 = [[list4[0],'a',2], [list4[1],'c',3]]
    elif str1a.find(nonseq) > -1:
        list4 = str1a.split(' ' + nonseq + ' ')
        impl = nonseq
        list4 = [[list4[0],'a',2], [list4[1],'c',3]]
    elif str1a[-1:] == ".":
        str1a = str1a[0:-1]
        list4.append([str1a,'a',2])
    else:
        list4.append([str1a,'a',2])

    return list4

def det_nouns(tot_sent, all_sent):

    global m
    global dnoun
    global det_pairs
    num = [5, 14, 18, 22, 26, 30, 34]
    for i in num:
        if all_sent[m][i] in dnoun:
            for j in range(len(det_pairs)):
                if det_pairs[j][2] == all_sent[m][i]:

                    rule = 'df ' + det_pairs[j][2]
                    if i == 14:
                        all_sent[m][10] = det_pairs[j][0]
                        all_sent[m][14] = det_pairs[j][1]
                    else:
                        all_sent[m][i-2] = det_pairs[j][0]
                        all_sent[m][i] = det_pairs[j][1]
                    bool1 = one_new_sent(all_sent,tot_sent, rule)

def findinlist(str1, list1, i, j, bool1 = False):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):

        if str1 == list1[d][i]:
            str2 = list1[d][j]
            if bool1:
                return d
            else:
                return str2
    if bool1:
        return -1
    else:
        return None

def redundant_words(all_sent, tot_sent):

    global m
    global redundant
    rule = "RD "
    num = [3,5,6,7,10,13,14,16,17,18,20,21,22,24,25,26,28,29,30,32,33,34]
    for i in num:
        if all_sent[m][i] in redundant:
            rule = rule + all_sent[m][i]
            all_sent[m][i] = None
            del all_sent[m][46][0]
            bool1 = one_new_sent(all_sent,tot_sent,rule)
            if all_sent[m][46] == []:
                break
            elif all_sent[m][46][0] != 3:
                break


def id_relat(tot_sent,all_sent):

    global m
    global osent
    subj = all_sent[m][5]
    obj = all_sent[m][14]
    rule = 'isub'
    num = [5,14]
    bool1 = False
    list3 = []

    for i in range(len(osent)):
        if osent[i][9] != '=':
            for j in num:
                if subj == osent[i][j]:
                    list2 = copy.deepcopy(osent[i])
                    list2[j] = obj
                    bool1 = True
                elif obj == osent[i][j]:
                    list2 = copy.deepcopy(osent[i])
                    list2[j] = subj
                    bool1 = True
                if bool1:
                    list2 = one_new_sent(all_sent,tot_sent, rule, list2, \
                                         True, osent[i][0])
                    list3.append(list2)
                    bool1 = False
    osent = osent + list3
    del all_sent[m][46][0]

def lemma1_3(all_sent, tot_sent):

    global m
    str1 = all_sent[m][14]
    if str1 == 'plural_counterpart':
        str2 = "PCP"
        str3 = '1L'
    elif str1 == 'adjective_counterpart':
        str2 = "ACP"
        str3 = '2L'
    elif str1 == 'verbal_counterpart':
        str2 = "VCP"
        str3 = '3L'
    all_sent[m][10] = None
    all_sent[m][14] = all_sent[m][18]
    all_sent[m][18] = None
    all_sent[m][15] = None
    all_sent[m][9] = str2
    rule = 'Lemma ' + str3
    all_sent[m][46] = []
    one_new_sent(all_sent, tot_sent, rule)


def one_new_sent(all_sent, tot_sent, rule, list2 = [], \
                 bool1 = False, str5 = ""):

    global sn
    global nonconjunct
    global m

    sn += 1
    if bool1:
        str1 = build_sent(list2)
        list2[0] = str1
    else:
        str1 = build_sent(all_sent[m])
    str1v = name_sent(str1)
    if nonconjunct:
        str4 = build_sent(all_sent[m])
        str4v = name_sent(str4)
        str1 = str4 + ' ' + iff + ' ' + str1
        str1v = str4v + ' ' + iff + ' ' + str1v

    if nonconjunct == False:
        if bool1:
            anc1 = findinlist(str5,tot_sent,1,0)
            anc2 = all_sent[m][36]
        else:
            anc1 = str(all_sent[m][36])
            anc2 = None
    else:
        anc1 = None
        anc2 = None
    if bool1 == False:
        all_sent[m][36] = sn
    tot_sent.append([sn, str1, str1v, rule, anc1, anc2])
    if bool1:
        return list2

def check_dimension(list1, i, str1):

    for j in range(len(list1)):
        if list1[j][i] == str1:
            return True
    return False



def reldiva(tot_sent, all_sent):

    global m
    list1 = [None] * 55
    list1[5] = all_sent[m][5]
    list1[9] = all_sent[m][15]
    all_sent[m][15] = None
    list1[18] = all_sent[m][18]
    all_sent[m][18] = None
    rule = 'RDA'
    list3 = []
    list4 = []
    if nonconjunct:
        rsent = build_sent(all_sent[m])
        rsentv = name_sent(rsent)
    else:
        rsent = ''
        rsentv = ''
    list1[45] = all_sent[m][45]
    bool1 = new_sent(list1, list3, list4, rule, all_sent, tot_sent, 2, rsent,rsentv,False)
    del all_sent[m][46][0]

def replace_relations(list1, relations, relations2):

    str1 = '('
    for i in range(len(list1)):
        if list1[i] in relations2:
            for j in range(len(relations)):
                if list1[i] == relations[j][0]:
                    list1[i] = relations[j][1]
                    str1 = str1 + " " + relations[j][1]
        elif list1[i] == 'not':
            list1[i] = '~'
            str1 = str1 + ' ~'
        else:
            if str1 == "(":
                str1 = str1 + list1[i]
            else:
                str1 = str1 + " " + list1[i]
    str1 = str1 + ")"

    return [list1, str1]

#################################################

## The following functions are for statement logic


def name_conditional(list1,nat_logic = True):

    global prop_var
    global prop_name
    skel_string = list1[5]
    list2 = []

    for i in range(1, len(list1[6])):
        if list1[6][i] != None and list1[6][i] != "":
            str3 = list1[0][i]
            str3 = str3.replace(" ","")
            str1 = findinlist(str3,prop_name,1,0)
            if str1 != None:
                str2 = list1[1][i] + str1
                list2.append(str1)
                skel_string = skel_string.replace(list1[6][i][1], str2)

            else:
                str2 = prop_var[0]
                list2.append(str2)
                del prop_var[0]
                list1[0][i] = list1[0][i].replace(" ","")
                prop_name.append([str2, list1[0][i]])
                str2 = list1[1][i] + str2
                skel_string = skel_string.replace(list1[6][i][1], str2)

    return [skel_string, list2]

def tilde_removal(str1):


    str4 = ""
    if str1.find("~") > -1:
        str4 = "~"
        str1 = str1.replace("~","")
    return [str1,str4]

def tilde_removal2(str1):

    j = 0
    if str1[:2] == "~(":
        for i in range(len(str1)):
            str2 = str1[i:i+1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i > 0:
                if i == len(str1) -1:
                    str1 = str1[1:]
                    str4 = "~"
                    return [str1, str4]
                else:
                    return [str1,""]
    elif str1[0] == "~" and str1[1].islower() and os(str1):
        str1 = str1[1:]
        str4 = "~"
    else:
        str4 = ""
    return [str1,str4]

def simple_sent_name(str1):
    # if False, then the tilde has already been removed
    # if true then we're doing statement logic

    str2 = findinlist(str1,prop_name,1,0)
    if str2 == None:
        str2 = prop_var[0]
        del prop_var[0]
    prop_name.append([str2, str1])
    return str2

def is_conjunction(str1):

    if str1.count(idisj) == 0 and str1.count(iff) == 0 and \
                    str1.count(conditional) == 0 and str1.count('&') > 0:
        return True
    else:
        return False


def get_conjuncts (str1, bool1 = False):
    # remove outparent if true
    arr1 = []
    if bool1:
        str1 = remove_outer_paren(str1)

    j = 0
    k = 1
    for i in range (len(str1)):
        str2 = str1[i:i+1]
        if i > 0:
            str4 = str1[i-1:i]
        else:
            str4 = ""

        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 1 and str2 == "(" and str4 != "~":
            k = i
        elif j == 1 and str4 == "~" and str2 == "(":
            k = i - 1

        if (j==0 and str2 == ")") or (j == 0 and str2.islower()):
            if str2 == ")":
                str3 = str1[k:i+1]
            else:
                if i>=1:
                    if str1[i-1:i] == "~":
                        str3 = "~" + str2
                    else:
                        str3 = str2

                    if str1[i:i+1] == "'":
                        str3 = str3 + "'"

                else:
                    if str1[i:i+1] == "'":
                        str3 = str2 + "'"

                    else:
                        str3 = str2


            k = 1
            str3 = str3.strip()
            arr1.append(str3)



    return arr1


def prepare_iff_elim(str2, mainc, s):

    global sn
    list7 = [""] * 39
    list7[2] = sn + 1
    list7[4] = str2
    list7[5] = ""
    j = 0
    str2 = remove_outer_paren(str2)

    if mainc == iff:
        list7[3] = "e"
    else:
        list7[3] = "c"

    str8 = str2[: s - 1]
    str8 = str8.strip()
    str9 = str2[s+1:]
    str9 = str9.strip()
    bool1 = True
    bool2 = True
    list8 = mainconn(str8)
    list2 = tilde_removal2(str8)
    if list8[0] != "&" or (list2[1] == "~" and list8[0] != ""):
        list7[0] = [list2[0],list2[1]]
        bool1 = False
    list8 = mainconn(str9)
    list2 = tilde_removal2(str9)
    if list8[0] != "&" or (list2[1] == "~" and list8[0] != ""):
        list7[1] = [list2[0],list2[1]]
        bool2 = False
    list4 = [bool1, bool2]
    for j in range(0, 2):
        if list4[j]:
            if j == 0:
                str7 = str8
            else:
                str7 = str9
            list2 = mainconn(str7)
            if list2[0] == '&':
                list3 = get_conjuncts(str7, True)
                list6 = []
                for k in range(len(list3)):
                    list5 = tilde_removal2(list3[k])
                    list6.append([list5[0],list5[1]])
                if j == 0:
                    list7[0] = list6
                    list7[6] = str8
                else:
                    list7[1] = list6
                    list7[7] = str9

    return list7

def islist(list1):

    if list1[1] == '~' or list1[1] == "":
        return False
    else:
        return True

def new_prop(prop_sent, str1, ng, asp, anc1, anc2, anc3 = None, anc4 = None, \
             is_premise = False):

    if ng == None:
        ng = ""
    global sn
    list1 = [None] * 15
    str1 = remove_outer_paren(str1)
    str2 = findinlist(str1,prop_sent,1,2)
    if is_premise:
        sn -= 1
    if str2 == ng:
        return True
    elif str2 == None:
        sn += 1
        prop_sent.append([sn, str1, ng, asp, anc1, anc2, None, None, None, None, None, \
        None, None, None, None])
        return True
    elif str2 != ng:
        sn += 1
        prop_sent.append([sn, str1, ng, asp, anc1, anc2, None, None, None, None, None, \
        None, None, None, None])
        anc2 = findinlist(str1,prop_sent,1,0)
        sn += 1
        if not os(str1):
            str1 = "(" + str1 + ")"
        str1 = str1 + " & " + "~" + str1
        prop_sent.append([sn, str1, "", "&I", sn-1, anc2, None, None, None, None, None, \
        None, None, None, None])
        str1 = bottom
        sn += 1
        prop_sent.append([sn, str1, "", bottom + "I", sn-1, None, None, None, None, None, None, \
        None, None, None, None])
        return False

def many_cond(candd, conditionals, kind, asp, anc2, f, g, r):

    list1 = conditionals[g][f]
    del list1[0]
    m = len(list1)
    j = -1
    if f == 1:
        h = 7
    else:
        h = 6
    list2 = []
    list3 = []
    list2.append([candd[r][1],candd[r][2]])
    list3.append(candd[r][0])
    while j < m - 1:
        j += 1
        str2 = list1[j][0]
        neg2 = list1[j][1]
        for i in range(len(candd)):
            if i != r:
                str1 = candd[i][1]
                ng = candd[i][2]
                anc1 = candd[i][0]
                if str1 == str2 and ng == neg2:
                    del list1[j]
                    j -= 1
                    list2.append([str1, ng])
                    list3.append(anc1)
                    if list1 == []:
                        cjct = conditionals[g][h]
                        dummy = new_prop_sent("", kind, asp, "",anc2, \
                                    conditionals,g, list3, cjct)
                        if not dummy:
                            return False
                        else:
                            del conditionals[g]
                            return True
                    else:
                        break
                elif str1 == str2 and ng != neg2:
            # the point of having blank returns is because if it returns true
            # then we need to subtract the conditional counter, here g, by 1
                    return ""
    return ""

def modus_ponens(conditionals, candd, prop_sent):

    global sn
    r = -1
    while r < len(candd) -1:
        if conditionals == []:
            return True
        r += 1
        str1 = candd[r][1]
        str8 = candd[r][2]
        anc1 = candd[r][0]
        temp1 = copy.copy(str1)
        temp1 = temp1.replace(" ","")
        temp1 = remove_outer_paren(temp1)
        g = -1
        bool1 = False
        while g < len(conditionals) -1:
            g += 1
            if g == 9:
                bb = 7
            str12 = conditionals[g][3]
            if str12 != 'd':
                anc2 = conditionals[g][2]
                if conditionals[g][6] == "":
                    aconjunction = ""
                    temp_ant = conditionals[g][0][0]
                    temp_nega = conditionals[g][0][1]
                else:
                    aconjunction = conditionals[g][6]
                    temp_ant = conditionals[g][0][0][0]
                    temp_nega = conditionals[g][0][0][1]
                if conditionals[g][7] == "":
                    cconjunction = ""
                    temp_con = conditionals[g][1][0]
                    temp_negc = conditionals[g][1][1]
                else:
                    cconjunction = conditionals[g][7]
                    temp_con = conditionals[g][1][0][0]
                    temp_negc = conditionals[g][1][0][1]
                temp_ant = temp_ant.replace(" ","")
                temp_con = temp_con.replace(" ","")
                temp_ant = remove_outer_paren(temp_ant)
                temp_con = remove_outer_paren(temp_con)

                for f in range(0,2):
                    if f == 0 and temp1 == temp_ant:
                        if str8 == temp_nega:
                            if str12 == 'c':
                                str13 = "MP"
                            else:
                                str13 = "EE"
                            if aconjunction != "" :
                                dummy = many_cond(candd, conditionals, "con", str13, \
                                                  anc2, f, g, r)
                                if not dummy:
                                    return False
                                elif dummy:
                                    g -= 1
                            else:
                 # con indicates that the consequent of the conditional is to be detached
                                dummy = new_prop_sent("", "con", \
                                    str13, anc1, anc2,conditionals,g)
                                if not dummy:
                                    return False
                                del conditionals[g]
                                g -= 1
                                break
                        elif str8 != temp_nega and str12 == 'e':
                            dummy = new_prop_sent("~", "con", \
                                        "EN", anc1, anc2, conditionals,g)
                            if not dummy:
                                return False
                            del conditionals[g]
                            g -= 1
                            break
                    elif f == 1 and temp1 == temp_con:
                        if str8 == temp_negc and str12 == 'e':
                            if cconjunction == "":
                                dummy = new_prop_sent("", "ant", "EE", \
                                    anc1, anc2, conditionals,g)
                                if not dummy:
                                    return False
                                del conditionals[g]
                                g -= 1
                                break
                            else:
                                dummy = many_cond(candd, conditionals, "ant", "EE", \
                                                  anc2, f, g, r)
                                if not dummy:
                                    return False
                                elif dummy:
                                    g -= 1
                        elif str8 != temp_negc:
                            if str12 == 'c':
                                str13 = "MT"
                            else:
                                str13 = "EN"
                            dummy = new_prop_sent("~", "ant", \
                                        str13, anc1, anc2, conditionals,g)
                            if not dummy:
                                    return False
                            del conditionals[g]
                            g -= 1
                            break
                    elif f == 0 and temp1 != temp_ant and \
                                    aconjunction != "" and str12 == 'e':
                        s = -1
                        if conditionals != []:
                            while s < len(conditionals[g][0]) -1:
                                s += 1
                                if temp1 == conditionals[g][0][s][0] and \
                                    str8 != conditionals[g][0][s][1]:
                                    dummy = new_prop_sent("~", "con", \
                                    "EN", anc1, anc2, conditionals,g)
                                    if not dummy:
                                        return False
                                    del conditionals[g]
                                    g -= 1
                                    break

                    elif f == 1 and temp1 != temp_con and cconjunction != "":
                        s = -1
                        if conditionals != []:
                            while s < len(conditionals[g][1]) -1:
                                s += 1
                                if temp1 == conditionals[g][1][s][0] and \
                                    str8 != conditionals[g][1][s][1]:
                                    if str12 == 'e':
                                        str13 = "EN"
                                    else:
                                        str13 = "MT"
                                    dummy = new_prop_sent("~", "ant", \
                                    str13, anc1, anc2, conditionals,g)
                                    if not dummy:
                                        return False
                                    del conditionals[g]
                                    g -= 1
                                    break
    return True

def disjunction_heirarchy(conditionals, str5,d,new_disj = False):

    global prop_name
    global sn

    str5 = enclose(str5)
    list1 = find_sentences(str5, False)
    for i in range(len(list1[0])):
        list2 = unenclose(list1[0][i])
        list1[0][i] = list2[0]

    list2 = [""] * 39
    n = 7
    list2[3] = 'd'
    if new_disj:
        list2[2] = sn
    else:
        list2[2] = conditionals[d][2]
    list2[5] = ""
    list2[4] = list1[0][0] # fix this

    for i in range(len(list1[0])):
        if os(list1[0][i]):
            siblings = []
            list3 = [None] * 9
            n += 1
            # str1 = findinlist(list1[0][i],prop_name,1,0)
            str2 = list1[4][i][0][:-1]
            g = findinlist(str2,list1[4],0,1,True)
            parent = list1[0][g]
            if list1[4][g][1] == "&":
                list3[2] = 'c'
            else:
                list3[2] = 'd'
            if len(str2) > 1:
                str3 = list1[4][i][0][:-2]
                g = findinlist(str3,list1[4],0,1,True)
                gparent = list1[0][g]
            else:
                gparent = parent
            list3[1] = list1[4][i][0]
            list3[5] = parent
            list3[6] = gparent
            list3[0] = [list1[0][i], list1[1][i]]
            b = parent.count(idisj)
            if b == 1:
                list3[7] = 1
            else:
                list3[7] = 2
            sent_num = list1[4][i][0]
            m = len(sent_num)
            for j in range(len(list1[4])):
                if len(list1[4][j][0]) == m and list1[4][j][0][:-1] == str2 \
                    and j != i:
                    siblings.append([list1[0][j],list1[1][j]]) # fix this
            list3[4] = siblings
            list2[n] = list3
    conditionals[d] = list2

def proper_spacing(str1):

    str1 = str1.replace(" ","")
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional," " + conditional + " ")
    str1 = str1.replace(idisj," " + idisj + " ")
    str1 = str1.replace("&"," & ")
    return str1

def iff_elim(prop_sent, conditionals):

    new_sent = False
    no_contr = True

    for d in range(len(conditionals)):
        list1 = ["",""]
        ng = conditionals[d][5]
        if conditionals[d][3] == "d" or conditionals[d][3] == 'n':
            list1 = [conditionals[d][4], ""]
        else:
            if conditionals[d][6] == "":
                list1[0] = conditionals[d][0][0]
            else:
                list1[0] = conditionals[d][6]
            if conditionals[d][7] == "":
                list1[1] = conditionals[d][1][0]
            else:
                list1[1] = conditionals[d][7]
        for s in range(0,2):
            if list1[s].find(iff) > -1:
                new_sent = True
                str1 = list1[s]
                anc1 = conditionals[d][2]
                old_str = copy.copy(str1)
                str1 = str1.replace(" ","")
                str1 = str1.replace("&"," & ")
                str1 = str1.replace(iff , " " + iff + " ")
                str1 = remove_outer_paren(str1)
                list5 = mainconn(str1)
                mc = list5[0]
                bool4 = False
                if str1[:2] == "~(":
                    bool4 = True
                    str1 = str1[2:len(str1) - 1]
                r = str1.count(iff)
                i = -1
                for o in range(0,r):
                    k = 0
                    while i < len(str1) - 1:
                        i += 1
                        str2 = str1[i:i+1]
                        if str2 == iff:
                            p = copy.copy(i)
                            t = copy.copy(i)
                            bool5 = False
                            while k != -1:
                                i -= 1
                                str3 = str1[i:i+1]
                                if str3 == "(":
                                    k -= 1
                                elif str3 == ")":
                                    k += 1
                                    bool5 = True
                                if k == -1:
                                    m = i + 1
                                elif k == 0 and bool5:
                                    m = i
                                    break
                            bool3 = False
                            k = 0
                            while p < len(str1):
                                p += 1
                                str3 = str1[p:p+1]
                                if str3 == "(":
                                    bool3 = True
                                    k -= 1
                                elif str3 == ")":
                                    k += 1
                                if bool3 and k == 0:
                                    p += 1
                                    break
                                elif k == 1:
                                    break
                            ant = str1[m:t]
                            ant = ant.strip()
                            con = str1[t+1:p]
                            con = con.strip()
                            new1 = "(" + ant + " " + conditional + " " + con + ")"
                            new2 = "(" + con + " " + conditional + " " + \
                                   ant + ")"
                            if mc != "&":
                                if r > 1 and o + 1 != r:
                                    new3 = "(" + new1 + ' & ' + new2 + ")*"
                                else:
                                    new3 = "(" + new1 + ' & ' + new2 + ")"
                            else:
                                if r > 1 and o + 1 != r:
                                    new3 = new1 + ' & ' + new2 + '*'
                                else:
                                    new3 = new1 + ' & ' + new2
                            if m != 0 or p != len(str1):
                                replacee = str1[m-1:p+1]
                                new3 = str1.replace(replacee, new3)
                            str1 = new3
                            if r > 1 and o + 1 != r:
                                i = str1.find("*")
                                str1 = str1.replace("*","")
                            break
                new3 = str1
                new3 = proper_spacing(new3)
                if bool4:
                    new3 = "~(" + new3 + ")"
                else:
                    new3 = "(" + new3 + ")"
                conditionals[d][4] = conditionals[d][4].replace(old_str,new3)
                if conditionals[d][3] == 'c' or conditionals[d][3] == 'e':
                    if s == 0:
                        conditionals[d][0] = None
                        conditionals[d][6] = new3
                    else:
                        conditionals[d][1] = None
                        conditionals[d][7] = new3

        if conditionals[d][3] == "e":
            conditionals[d][3] = "c"
            anc1 = conditionals[d][2]
            list1 = [""] * 39
            list1[5] = ""
            if conditionals[d][6] == "":
                str1 = conditionals[d][0][0]
                ng1 = conditionals[d][0][1]
            else:
                str1 = conditionals[d][6]
                ng1 = ""
            if conditionals[d][7] == "":
                str4 = conditionals[d][1][0]
                ng4 = conditionals[d][1][1]
            else:
                str4 = conditionals[d][7]
                ng4 = ""

            list1[0] = [str4,ng4]
            list1[1] = [str1, ng1]
            str5 =  ng1 + str1 + " " + conditional + " " + ng4 + str4
            str6 = ng4 + str4 + " " + conditional + " " + ng1 + str1
            g = copy.copy(sn+1)
            if conditionals[d][5] == "" or conditionals[d][5] == None:
                str3 = "(" + str5 + ") & (" + str6 + ")"
                str7 = iff + "E"
                no_contr = new_prop(prop_sent, str3, ng, str7, anc1, None, None, None)

                if not no_contr:
                    return False
                conditionals[d][2] = sn+1
                conditionals[d][0] = [str1, ng1]
                conditionals[d][1] = [str4, ng4]
                no_contr = new_prop(prop_sent,str5,"","&E",g,"")
                if not no_contr:
                    return False
                list1[2] = sn + 1
                no_contr = new_prop(prop_sent,str6,"","&E",g,"")
                if not no_contr:
                    return False
                conditionals[d][4] = str5
                list1[4] = str6
                conditionals.append(list1)
            else:
                str3 = "(" + str5 + ") & (" + str6 + ")"
                no_contr = new_prop(prop_sent,str3,"~",iff + "E",g,"")
                if not no_contr:
                    return False
                conditionals[d][3] = 'd'
                conditionals[d][4] = str3
        elif new_sent:
            no_contr = new_prop(prop_sent,conditionals[d][4],"",iff + "E",anc1,"")
            if not no_contr:
                return False
    return True

def material_implication(prop_sent, conditionals):

    for d in range(len(conditionals)):
        str1 = conditionals[d][4]
        ng = conditionals[d][5]
        if str1.find(conditional) > -1:
            anc1 = conditionals[d][2]
            i = -1
            q = -1
            s = 0
            r = str1.count(conditional)
            while s < r:
                i += 1
                q += 1
                str2 = str1[i:i+1]
                if str2 == conditional:
                    s += 1
                    j = i
                    k = 0
                    m = 0
                    bool1 = False
                    while j != 0:
                        j -= 1
                        m += 1
                        if m > 100:
                            easygui.msgbox("in the material implication function \
                            you are caught in an infinite loop")
                            return
                        str3= str1[j:j+1]
                        if str3.islower() and k == 0:
                            bool1 = True
                        elif str3 == ")":
                            bool1 = True
                            k += 1
                        elif j == 0 and str3 != "(":
                            k = 0
                            bool1 = True
                        elif str3 == "(":
                            bool1 = True
                            k -= 1

                        if bool1 and k == 0:
                            str4 = str1[0:j]
                            str5 = str1[i+2:]
                            str7 = str1[j:i]
                            str1 = str4 + "~" + str7 + idisj + " " + str5
                            i += 1
                            break

            str1 = bad_paren(str1)
            no_contr = new_prop(prop_sent, str1, ng, conditional + "E", anc1,"")
            if not no_contr:
                return False
            if str1.find("~~") > -1:
                str1 = str1.replace("~~","")
                g = copy.copy(sn)
                no_contr = new_prop(prop_sent,str1, ng, "~~E", g,"")
                if not no_contr:
                    return False
            conditionals[d][2] = sn
            conditionals[d][4] = str1
            conditionals[d][3] = "d"
            conditionals[d][5] = ng

    return True

            # search for double negatives

def bad_paren(str1):

    str1 = enclose(str1)
    list1 = find_sentences(str1, False)
    for i in range(len(list1[3])):
        list2 = unenclose(list1[3][i])
        list1[3][i] = list2[0]
    mstr = list1[3][0]
    for i in range(1, len(list1[3])):
        if list1[4][i][1] != "":
            mc = list1[4][i][1]
            ostr = list1[3][i]
            str2 = list1[4][i][0][:-1]
            prcnt = findinlist(str2,list1[4],0,1,False)
            if mc == prcnt:
                nstr = remove_outer_paren(ostr)
                nstr = remove_outer_paren(nstr)
                mstr = mstr.replace(ostr, nstr)
    return mstr

def demorgan(prop_sent, conditionals, candd):

    d = -1
    temp_bool = True
    rop = False
    while d < len(conditionals) -1:
        d += 1
        str1 = conditionals[d][5] + conditionals[d][4]
        if conditionals[d][5] == "~":
            rop = True
        if str1.find("~(") > -1:
            anc1 = conditionals[d][2]
            r = str1.count("~(")
            s = 0
            i = -1
            while s < r:
                i += 1
                if i > 90:
                    easygui.msgbox("you are caught in an infinite loop in the \
                    demorgan function")
                    return
                bool2 = False
                str2 = str1[i:i+2]
                if str2 == "~(":
                    s += 1
                    bool2 = True
                    if i > 0:
                        str4 = str1[0:i]
                    else:
                        str4 = ""
                    str5 = str1[i+1:len(str1)+1]
                    str1 = str4 + str5
                    j = i
                    m = 0
                    k = 0
                    bool3 = False
                    while j < len(str1):
                        m += 1
                        if m > 90:
                            break
                        bool1 = False
                        str3 = str1[j:j+1]
                        if j-1 > 0:
                            str6 = str1[j-1:j+1]
                        else:
                            str6 = ""

                        if str6 == "~(" and bool3 == False and bool2 == False:
                            k = 0
                            k += 1
                            s += 1
                            # bool3 means that python is searching only for those letters and connectives
                            # which are on the same level of parentheses
                            bool3 = True
                            str4 = str1[0:j]
                            str5 = str1[j:]
                            str1 = str4 + "~" + str5
                            j += 2
                            str3 = str1[j:j+1]


                        bool2 = False

                        if str3.islower() and bool3 == False:
                            str4= str1[0:j]
                            str5 = str1[j+1:(len(str1)+1)]
                            str1 = str4 + "~" + str3 + str5
                            j += 2
                            bool1 = True
                        elif bool3 == False:
                            if str3 == "(":
                                k += 1
                                j += 1
                                bool1 = True
                            elif str3 == ")":
                                k -= 1
                                j += 1
                                bool1 = True
                            elif str3 == "&":
                                str4 = str1[0:j]
                                str5 = str1[(j + 1):(len(str1) + 1)]
                                str1 = str4 + idisj+ str5
                                j += 1
                            elif str3 == idisj:
                                str4 = str1[0:j]
                                str5 = str1[(j+1):(len(str1)+1)]
                                str1 = str4 + "&" + str5
                                j += 1
                            else:
                                j += 1

                        elif str3 == ")":
                            j += 1
                            k -= 1
                        elif str3 == "(":
                            k += 1
                            j += 1
                        else:
                            j += 1

                        if k == 0 and bool3:
                            bool3 = False
                        if k == 0 and bool1 == True:
                            i = j
                            break

        # rop means remove outer paren, this is necessary if the original sentence was negated
            if rop:
                str1 = remove_outer_paren(str1)
            list1 = [None] * 15
            if str1.find("~~") > -1:
                str2 = copy.copy(str1)
                str2 = str2.replace("~~","")
                str2 = bad_paren(str2)
            else:
                str1 = bad_paren(str1)
            no_contr = new_prop(prop_sent,str1,"","~(E",anc1,"")
            if not no_contr:
                return False
            if str1.find("~~") > -1:
                str1 = str2
                anc1 = copy.copy(sn)
                no_contr = new_prop(prop_sent,str2,"","~~E",anc1,"")
                if not no_contr:
                    return False
            list2 = mainconn(str1)

            if list2[0] == "&":
                del conditionals[d]
                list3 = get_conjuncts(str1,True)
                anc1 = copy.copy(sn)
                for i in range(len(list3)):
                    list4 = tilde_removal2(list3[i])
                    no_contr = new_prop(prop_sent,list4[0],list4[1],"&E",anc1,"")
                    if not no_contr:
                        return False
                    list2 = mainconn(list3[i])
                    if list2[0] == idisj:
                        # add in more nones if it turns out that I need them
                        list5 = [""] * 39
                        list5[2] = sn
                        list5[4] = list4[0]
                        list5[5] = list4[1]
                        conditionals.append(list5)
                    else:
                        candd.append([sn,list4[0],list4[1]])
            else:
                conditionals[d][2] = sn
                conditionals[d][4] = str1
    return True

def unenclose(str1):

    i = -1
    list1 = []
    while i < len(str1)-1:
        i += 1
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        if str2.islower() and str3 != "~":
            str1 = str1[:i-1] + str2 + str1[i+2:]
            list1.append(str2)
        elif str2.islower() and str3 == "~":
            str1 = str1[:i-2] + str3 + str2 + str1[i+2:]
            list1.append(str2)
    return [str1,list1]

def new_disjunct(str1, ng, n, prop_sent, conditionals, candd, anc1, anc2, \
            anc3 = None, anc4=None, kind = 0):

    global sn
    list2 = mainconn(str1)
    if kind == 1:
        del conditionals[n]
        dummy = new_prop(prop_sent, str1, ng, "&I", \
            anc1, anc2, anc3, anc4)
        return dummy
    elif kind == 2:
        dummy = new_prop(prop_sent, str1, ng, "&I", \
            anc1, anc2, anc3, anc4)
        return dummy
    else:
        if os(str1):
            del conditionals[n]
            list1 = tilde_removal2(str1)
            candd.append([sn+1,list1[0],list1[1]])
            dummy = new_prop(prop_sent, list1[0], list1[1], idisj + "E", \
            anc1, anc2)
            return dummy
        elif list2[0] == "&":
            del conditionals[n]
            str1 = remove_outer_paren(str1)
            dummy = new_prop(prop_sent, str1, ng, idisj + "E", \
            anc1, anc2)
            g = copy.copy(sn)
            list3 = get_conjuncts(str1)
            for i in range(len(list3)):
                list4 = tilde_removal(list3[i])
                dummy = new_prop(prop_sent, list4[0], list4[1], "&E", g,"")
                if dummy == False:
                    return dummy
            return True
        else:
            dummy = new_prop(prop_sent, str1, ng, idisj + "E", \
            anc1, anc2)
            if dummy == False:
                return dummy
            if ng == "~":
                str1 = ng + str1
            else:
                str1 = remove_outer_paren(str1)
            dummy = disjunction_heirarchy(conditionals, str1,n, True)
            return True

def disjunction_elimination(prop_sent, conditionals, candd):

    bool1 = False
    bool2 = False

    for i in range(len(conditionals)):
        dummy = disjunction_heirarchy(conditionals, conditionals[i][4],i)
    i = -1
    while i < len(candd) -1:
        i += 1
        if not os(candd[i][1]):
            del candd[i]
            i -= 1
    d = -1
    while d < len(candd) -1:
        d += 1
        str2 = candd[d][2]
        conj = candd[d][1]
        anc1 = candd[d][0]
        n = -1
        while n < len(conditionals) -1:
            if bool1:
                bool1 = False
                d = -1
                break
            n += 1
            i = 7
            while conditionals != []:
                if bool2:
                    bool2 = False
                    break
                i += 1
                if conditionals[n][i] == "":
                    break

                whole_d = conditionals[n][4]
                anc2 = conditionals[n][2]
                str3 = conditionals[n][i][0][0]
                # 'pos or neg'
                str4 = conditionals[n][i][0][1]
                # 'disjunct or conjunct
                str5 = conditionals[n][i][2]
                # 'disjunct number
                str6 = str(conditionals[n][i][1])

                if conj == str3:
                    grandparent = conditionals[n][i][6]
                    parent = conditionals[n][i][5]
                    parent2 = copy.copy(parent)
                    parent3 = copy.copy(parent)
                    str7 = " " + idisj + " "
                    if str2 == str4 and str5 == "d" :
                        # 'if the disjuncts are not embedded within a conjunct then the disjunction
                        # is simply deleted
                        del conditionals[n]
                        if len(str6) != 2:
                            conj = str2 + conj
                            str8 = whole_d.replace(parent, conj)
                            dummy = disjunction_heirarchy(conditionals, str8,n)
                        bool1 = True
                        n = -1
                        break

                    elif str2 == str4 and str5 == "c":
                        list2 = []
                        list2.append([conj,str2])
                        anc3 = ""
                        anc4 = ""
                        list1 = conditionals[n][i][4]
                        f = -1
                        while f < len(list1) -1:
                            f += 1
                            for e in range(len(candd)):
                                anc5 = candd[e][0]
                                if candd[e][1] == list1[f][0]:
                                    if candd[e][2] == list1[f][1]:
                                        list2.append([list1[f][0],list1[f][1]])
                                        if len(list2) == 2:
                                            anc3 = anc5
                                        elif len(list2) == 3:
                                            anc4 = anc5
                                        del list1[f]

                                        if list1 == []:
                                            str3 = build_sent_list2(list2)
                                    # if the conjunct is not embedded within another conjunct
                                    # then the disjunct is simply deleted
                                            if whole_d == grandparent:
                                                dummy = new_disjunct(str3,"",n, prop_sent,\
                                                conditionals,candd, anc1, anc3, anc4, anc5, 1)
                                            else:
                                                str8 = whole_d.replace(grandparent, parent2)
                                                dummy = new_disjunct(str3,"",n, prop_sent,\
                                                conditionals,candd, anc1, "", anc3, anc4, 2)
                                                dummy = new_disjunct(str8,"",n, prop_sent,\
                                                conditionals,candd, sn-1, anc2)
                                            bool1 = True
                                            bool2 = True
                                            n = 0
                                            break
                                        else:
                                            f -= 1
                                            break

                                    elif candd[e][2] != list1[f][1]:
                                        r = grandparent.find(parent)
                                        if r > 1:
                                            parent = str7 + parent
                                        else:
                                            parent = parent + str7
                                        anc1 = candd[e][0]
                                        str9 = grandparent.replace(parent, "")
                                        str8 = whole_d.replace(grandparent, str9)
                                        dummy = new_disjunct(str8,"",n, prop_sent,\
                                                conditionals,candd, anc1, anc2)
                                        if not dummy:
                                            return False
                                        else:
                                            bool1 = True
                                            bool2 = True
                                            n = 0
                                            break

                    elif str2 != str4 and str5 == "c":

                        str6 = str7 + parent
                        if grandparent.find(str6) > -1:
                            parent = str6
                        else:
                            parent = parent + str7
                        str9 = grandparent.replace(parent, "")
                        str8 = whole_d.replace(grandparent, str9)
                        dummy = new_disjunct(str8,"",n,prop_sent, conditionals, candd,anc1, anc2)
                        if not dummy:
                            return False
                        bool1 = True
                        n = -1
                        break

                    elif str2 != str4 and str5 == "d":
                        # if the disjunct is a triple disjunct then enter below
                        if conditionals[n][i][7] > 1:

                            str6 = str4 + str3 + " " + idisj + " "
                            if parent.find(str6) > -1:
                                str5 = str6
                            else:
                                str5 = " " + idisj + " " + str4 + str3
                            str9 = parent.replace(str5, "")
                            str8 = whole_d.replace(parent, str9)
                            dummy = new_disjunct(str8,"",n,prop_sent,conditionals,\
                                candd, anc1,anc2)
                            if not dummy:
                                return False
                            bool1 = True
                            n = -1
                            break

                        else:
                            str3 = conditionals[n][i][4][0][0]
                            str4 = conditionals[n][i][4][0][1]
                            str5 = str4 + str3
                            str8 = whole_d.replace(parent, str5)
                            dummy = new_disjunct(str8,"",n,prop_sent,conditionals,\
                                candd, anc1,anc2)
                            if not dummy:
                                return False
                            bool1 = True
                            n = -1
                            break

def statement_logic(prop_sent, conditionals, candd):

    global sn
    no_contr = modus_ponens(conditionals, candd, prop_sent)
    if not no_contr:
        return False
    no_contr = iff_elim(prop_sent,conditionals)
    if not no_contr:
        return False
    no_contr = material_implication(prop_sent, conditionals)
    if not no_contr:
        return False
    no_contr = demorgan(prop_sent, conditionals, candd)
    if not no_contr:
        return False
    no_contr = disjunction_elimination(prop_sent,conditionals,candd)
    if not no_contr:
        return False

    pp = 7

    return


def new_prop_sent(ng, kind, asp, anc1, anc2, conditionals,g,list3 =[], cjct = ""):

    global prop_sent
    global candd
    global sn

    if kind == 'con':
        h = 1
        e = 7
    else:
        h = 0
        e = 6
    bool1 = False
    if list3 != []:
        sn += 1
        ancc = [4,5,6,7]
        list4 = [None] * 15
        for i in range(len(list3)):
            list4[ancc[i]] = list3[i]
        list4[0] = sn
        list4[1] = cjct
        list4[3] = "&I"
        anc1 = sn
        prop_sent.append(list4)

    if conditionals[g][e] == "":
        str1 = conditionals[g][h][0]
        list2 = mainconn(str1)

        # here we take care of double negatives
        if ng == "~" and conditionals[g][h][1] == "~":
            sn += 1
            prop_sent.append([sn, str1, "~~", asp, anc1, anc2, None, None, None, None, None,None, None, None, None])
            dummy = new_prop(prop_sent,str1,"","~~E",sn-1, None)
            if not dummy:
                return False
            ng = ""
            bool1 = True
        elif (ng == "" and conditionals[g][h][1] == "~") or \
                (ng == "~" and conditionals[g][h][1] == ""):
            ng = "~"
        if list2[0] != "" and ng != "~":
            str1 = remove_outer_paren(str1)
        if bool1 == False:
            dummy = new_prop(prop_sent,str1,ng,asp,anc1, anc2)
            if not dummy:
                return False

        candd.append([sn, str1, ng])
        if (list2[0] == iff or list2[0] == conditional) and ng != "~":
            list3 = prepare_iff_elim(str1, list2[0],list2[1])
            conditionals.append(list3)
        elif list2[0] == iff or list2[0] == conditional or list2[0] == idisj:
            list5 = [""] * 39
            list5[2] = sn
            list5[4] = str1
            list5[5] = ng
            if list2[0] == iff:
                list5[3] = 'e'
            elif list2[0] == conditional:
                list5[3] = 'c'
            else:
                list5[3] = 'd'
            conditionals.append(list5)
    else:
        str1 = conditionals[g][e]
        if ng == "~" and conditionals[g][h][1] == "~":
            positive = True
        elif ng == "~" or conditionals[g][h][1] == "~":
            positive = False
        else:
            positive = True
        if not positive:
            list5 = [""] * 39
            list5[2] = sn
            list5[4] = str1
            list5[5] = "~"
            list5[3] = 'd'
            conditionals.append(list5)
            dummy = new_prop(prop_sent, str1, ng, asp, acn1, anc2)
            if not dummy:
                return False
        else:
            list1 = conditionals[g][h]
            str1 = remove_outer_paren(str1)
            dummy = new_prop(prop_sent, str1, ng, asp, anc1, anc2)
            if not dummy:
                return False
            anc1 = copy.copy(sn)
            for i in range(len(list1)):
                list2 = mainconn(list1[i][0])
                dummy = new_prop(prop_sent, list1[i][0], list1[i][1], "&E", anc1, "")
                if not dummy:
                    return False
                candd.append([sn, list1[i][0], list1[i][1]])
                if (list2[0] == conditional or list2[0] == iff) and list1[i][1] != "~":
                    list4 = prepare_iff_elim(list1[i][0], list2[0],list2[1])
                    conditionals.append(list4)
                elif list2[0] != "":
                    list5 = [""] * 39
                    if list2[0] == conditional:
                        list5[3] = 'd'
                    elif list2[0] == iff:
                        list5[3] = 'e'
                    else:
                        list5[3] = 'c'
                    list5[2] = sn
                    list5[4] = list1[i][0]
                    list5[5] = list1[i][1]
                    conditionals.append(list5)


    return True












def plan(sent, prop_sent, candd, conditionals, prop_name):

    global conc
    global sn
    global rel_conj
    conj_elim = []
    temp_conditionals = []
    list4 = []
    str3 = ""
    str1 = ""
    qq = 0
    nat_logic = False
    # if the first sentence is just one letter then we're using statement logic
    if sent[0][1].find("(") > -1:
        for j in range(len(sent[0][1])):
            str2 = sent[0][1][j:j+1]
            str3 = sent[0][1][j+1:j+2]
            if (str2.islower() or str2.isupper()) and (str3.islower() or str3.isupper()):
                nat_logic = True
            if j > 8:
                break

    for i in range(len(sent)):

        if nat_logic == False:
            sent[i][1] = enclose(sent[i][1])

        # if sent[i][0][-1] == "c":
        #     sent[i][0] = 1

        if sent[i][1].find("!") > -1:
            qq += 1
        else:
            if sent[i][1].count("(") != sent[i][1].count(")"):
                easygui.msgbox("line " + sent[i][0] + " does not have the right number \
                                             of parentheses" )
            sent[i][1] = remove_outer_paren(sent[i][1])
            sent[i][1] = proper_spacing(sent[i][1])
            if sent[i][1].find(implies) > -1 or sent[i][1].find(nonseq) > -1:

                if sent[i][1].find(implies) > -1:
                    impl = implies
                else:
                    impl = nonseq

                str1 = sent[i][1][2:]
                if str1.find("&") > -1:
                    list1 = str1.split("&")
                    list4 = find_sentences(str1, True)
                    list6 = name_conditional(list4)
                    str5 = list6[0]
                    str2 = impl + ' ' + str5
                    for i in range(len(list1)):
                        list3 = simple_sent_name(list1[i], True)
                        conc.append([sent[i][0], list3[0], list3[1]])
                else:
                    list3 = simple_sent_name(str1, True)
                    str2 = impl + ' ' + list3[1] + list3[0]
                    ng = ""
                    conc.append([sent[i][0], list3[0], list3[1]])

            elif os(sent[i][1]):

                list3 = tilde_removal(sent[i][1])
                if nat_logic:
                    list3[0] = simple_sent_name(list3[0], True)
                str2 = list3[0]
                ng = list3[1]
                candd.append([sent[i][0], list3[0], list3[1]])
                prop_sent.append([sent[i][0],str2,ng,""])
            else:
                list1 = find_sentences(sent[i][1], True)
                if not nat_logic:
                    list6 = unenclose(list1[0][0])
                    sent[i][1] = list6[0]
                else:
                    list6 = name_conditional(list1, nat_logic)
                str2 = list6[0]
                list4 = tilde_removal2(str2)
                ng = list4[1]
                str2 = list4[0]
                list2 = mainconn(str2)

                if list2[0] == idisj:
                    candd.append([sent[i][0], str2,ng])
                    rel_conj = rel_conj + list6[1]
                    list5 = [""] * 39
                    list5[3] = 'd'
                    list5[4] = str2
                    list5[5] = ng
                    conditionals.append(list5)
                elif list1[4][0][1] != "&":
                    str3 = ""
                    rel_conj = rel_conj + list6[1]
                    if list1[4][0][1] != idisj and ng == "":
                        list7 = prepare_iff_elim(str2, list2[0], list2[1])
                    else:
                        list7 = [""] * 39
                        list7[2] = sn
                        list7[3] = 'd'
                        list7[4] = str2
                        list7[5] = ng
                    conditionals.append(list7)
                    candd.append([sent[i][0], str2,ng])
                else:
                    list3 = get_conjuncts(list6[0])
                    for j in range(len(list3)):
                        list5 = tilde_removal2(list3[j])
                        if os(list3[j]):
                            conj_elim.append([sent[i][0],list5[0],list5[1]])
                        else:
                            temp_conditionals.append([sent[i][0], list5[0],list5[1]])

            no_contr = new_prop(prop_sent,str2,ng,"",None,None,None,None,True)
            if not no_contr:
                return False

    if conj_elim != []:
        for i in range(len(conj_elim)):
            no_contr = new_prop(prop_sent,conj_elim[i][1],conj_elim[i][2],"&E",\
                        conj_elim[i][0],None)
            if not no_contr:
                return False
            candd.append([sn, conj_elim[i][1], conj_elim[i][2]])

    if temp_conditionals != []:
        for i in range(len(temp_conditionals)):

            list1 = find_sentences(temp_conditionals[i][1], True)
            list6 = name_conditional(list1)
            str2 = list6[0]
            ng = temp_conditionals[i][2]
            list2 = mainconn(str2)

            if list2[0] != idisj and ng == "" and list2[0] != "&":
                list7 = prepare_iff_elim(str2, list2[0], list2[1])
            else:
                list7 = [""] * 39
                list7[2] = sn + 1
                list7[4] = str2
                list7[3] = 'd'
            if ng == "~":
                list7[5] = "~"
            rel_conj = rel_conj + list6[1]
            conditionals.append(list7)
            candd.append([sn+1, str2,ng])
            no_contr = new_prop(prop_sent, str2,ng,"&E",temp_conditionals[i][0], None)
            if not no_contr:
                return False


    # we now weed out the irrelevant conjuncts from the relevant
    # fix this later
    # i = -1
    # if len(candd) > 0:
    #     while i < len(candd) -1:
    #         i += 1
    #         if candd[i][1] not in rel_conj:
    #             del candd[i]
    #             i -= 1

    consistent = statement_logic(prop_sent, conditionals, candd)

    pp = 7


def get_result(post_data):
    #### rajiv - it starts here, everything above this is a function
    ## from row 1 to 15, is the input and it stops collecting input when it hits
    # the word 'stop'
    # a button the clears the output, but not the input

    # wb7 = load_workbook('define2.xlsx')
    # ws = wb7.worksheets[0]

    #wb4 = load_workbook('test7.xlsx')
    #w4 = wb4.worksheets[0]


    sn = 1
    conditionals = []
    candd = []
    rel_conj = []
    conc = []
    prop_sent = []
    sent = []
    prop_name = []
    prop_var = []
    prop_var = [unichr(97 + t) for t in range(26)]
    prop_var2 = []
    prop_var2 = [unichr(97 + t) + "'" for t in range(26)]
    prop_var = prop_var + prop_var2
    test_sent=[]
    p = 0


    # for row in w4.rows:
    #     p += 1
    #     if row[2].value == "stop":
    #         break
    #     elif row[1].value != None:
    #         sent.append([row[1].value,row[2].value])
    #     else:
    #         test_sent.append(sent)
    #         sent = []

    ######### for rajiv - the following will have to changed
    p = 0
    while  True:
        row = [post_data["text_"+str(p)+"_1"],post_data["text_"+str(p)+"_2"],post_data["text_"+str(p)+"_3"]]

        if row[1] == "stop":
            test_sent.append(sent)
            sent = []
            break
        elif row[1]!= None:
            sent.append([row[0], row[1], row[2]])
        elif row[1] == None:
            test_sent.append(sent)
            sent = []
        p += 1



    tot_prop_name = []
    tot_sent = []
    output = []
    # timeit.timeit('"-".join(str(n) for n in range(100))', number=10000)

    p += 3

    #begin
    for i in range(0, len(test_sent)):
    #for i in range(0, len(sent)):
        print i
        dummy = plan(test_sent[i], prop_sent, candd, conditionals, prop_name)
        # tot_sent.append(prop_sent)
        # tot_prop_name.append(prop_name)
        output.append(prop_sent)
        prop_sent = []
        candd = []
        conditionals = []
        sn = 1
        # prop_name = []
        # prop_var = []
        # prop_var = [unichr(97 + t) for t in range(26)]
        # prop_var2 = []
        # prop_var2 = [unichr(97 + t) + "'" for t in range(26)]
        # prop_var = prop_var + prop_var2




    for i in range(len(output)):
        for j in range(len(output[i])):
            if output[i][j][2] == None:
                output[i][j][2] = ""

            post_data["text_"+str(p)+"_1"] = output[i][j][0]
            if not os(output[i][j][1]) and output[i][j][2] == "~":
                str1 = output[i][j][2] + "(" + output[i][j][1] + ")"
            else:
                str1 = output[i][j][2] + output[i][j][1]
            post_data["text_"+str(p)+"_2"] = str1
            if output[i][j][3] != None and output[i][j][3] != "":
                if output[i][j][5] != None:
                    str2 = output[i][j][3] + " " + str(output[i][j][4]) + "," + \
                        str(output[i][j][5])
                else:
                    str2 = output[i][j][3] + " " + str(output[i][j][4])
                post_data["text_"+str(p)+"_3"] = str2
            p += 1
        p += 2



    # for i in range(len(prop_sent)):
    #     if prop_sent[i][2] == None:
    #         prop_sent[i][2] = ""
    #
    #     w4.cell(row=p,column=2).value = prop_sent[i][0]
    #     if not os(prop_sent[i][1]) and prop_sent[i][2] == "~":
    #         str1 = prop_sent[i][2] + "(" + prop_sent[i][1] + ")"
    #     else:
    #         str1 = prop_sent[i][2] + prop_sent[i][1]
    #     w4.cell(row=p,column=3).value = str1
    #     if prop_sent[i][3] != None and prop_sent[i][3] != "":
    #         if prop_sent[i][5] != None:
    #             str2 = prop_sent[i][3] + " " + str(prop_sent[i][4]) + "," + \
    #                 str(prop_sent[i][5])
    #         else:
    #             str2 = prop_sent[i][3] + " " + str(prop_sent[i][4])
    #         w4.cell(row=p,column=4).value = str2
    #     p += 1
    #wb4.save('test7.xlsx')

    pp = 7




    # for i in range(len(test_sent)):
    #     for j in range(0,2):
    #         w4.cell(row=p,column=2).value = test_sent[i][j][0]
    #         w4.cell(row=p,column=3).value = test_sent[i][j][1]
    #         p += 1
    #     p += 1
    #     list1 = build_sent_name(tot_prop_name[i])
    #
    #     for j in range(len(list1)):
    #         w4.cell(row=p,column=3).value = list1[j]
    #         p += 1
    #     p += 1
    #
    #     for j in range(len(tot_sent[i])):
    #         w4.cell(row=p,column=2).value = tot_sent[i][j][0]
    #         w4.cell(row=p,column=3).value = tot_sent[i][j][2] + tot_sent[i][j][1]
    #         w4.cell(row=p,column=4).value = str(tot_sent[i][j][3]) + " " + str(tot_sent[i][j][4]) \
    #         + "," + str(tot_sent[i][j][5])
    #         p += 1
    #     p += 2
    # wb4.save('test6.xlsx')
    return post_data
