import json
import os
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render
from django.core.servers.basehttp import FileWrapper
from django.http import HttpResponse
from django.conf import settings
from django.http import StreamingHttpResponse
from django_tools.middlewares import ThreadLocal
from django.views.decorators.csrf import csrf_exempt
import time
from django.db import transaction
from models import Output
import os
import importlib
# from inference2.Proofs #import 5.17.16.py
from inference2.models import Input

from models import Define3, Archives

# from Proofs import 5_17_16.py


def save_result(post_data):
    Output.objects.all().delete()
    Rows = []
    for idx in xrange(15000 - 1):
        c1 = post_data.get("text_" + str(idx) + "_1", '')
        c2 = post_data.get("text_" + str(idx) + "_2", '')
        c3 = post_data.get("text_" + str(idx) + "_3", '')
        if type(c1) == type([]):
            c1 = c1[0]
        if type(c2) == type([]):
            c2 = c2[0]
        if type(c3) == type([]):
            c3 = c3[0]
        R = Output(col1=c1,
                   col2=c2,
                   col3=c3
                   )
        Rows.append(R)
    Output.objects.bulk_create(Rows)


# Create your views here.

def current_archive():
    archive = Archives.objects.latest('archives_date')
    return archive


def index(request, archive=None):
    progressbar_send(request, 1, 100, 1)
    url_path = ''
    archive_date = ''
    if not archive:
        archive = current_archive()
        url_path = '/'
    else:
        url_path = '/archives/{}/'.format(archive.id)
        archive_date = archive.archives_date
    input = Input.objects.filter(archives_id=archive.id)
    result = {}
    output = []
    output = Output.objects.all()
    if request.method == 'POST':
        post_data = request.POST.copy()
        prove_algorithm = importlib.import_module(
            '.' + archive.algorithm, package='inference2.Proofs')
        post_data = prove_algorithm.get_result(
            request.POST.copy(), archive.id, request)
        if post_data:
            post_data["type"] = "prove"
            result = json.dumps(post_data, cls=DjangoJSONEncoder)

            save_result(post_data)
        output = Output.objects.all()

    #rows = json.dumps(rows,cls=DjangoJSONEncoder)

    template_args = {'result': result, 'input': input,
                     'url_path': url_path, 'archive_date': archive_date,
                     'output': output
                     }
    return render(request, "inference2/index.html", template_args)

def export_xlsx(request):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"
    queryset = Output.objects.all()
    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.col1,
            obj.col2,
            obj.col3,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

def stream_response_generator(request):
    for x in range(1, 11):
        yield "%s\n" % x  # Returns a chunk of the response to the browser
        request.session['idx'] = x
        request.session.modified = True

        request.session.save()
        time.sleep(1)


def prove(request, archive=None):
    progressbar_send(request, 1, 100, 1)
    if not archive:
        archive = current_archive()
    result = {}
    if request.method == 'POST':
        post_data = request.POST.copy()
        prove_algorithm = importlib.import_module(
            '.' + archive.algorithm, package='inference2.Proofs')
        post_data = prove_algorithm.get_result(
            request.POST.copy(), archive.id, request)
        result = json.dumps(post_data, cls=DjangoJSONEncoder)

    #rows = json.dumps(rows,cls=DjangoJSONEncoder)
    return result


def dictionary(request, archive=None):
    url_path = '/archives/'
    if not archive:
        archive = current_archive()
        url_path = '/'
    else:
        url_path = '/archives/{}/'.format(archive.id)
    dict = Define3.objects.filter(archives_id=archive.id)
    return render(request, "inference2/dict.html", {'result': dict, 'url_path': url_path})


def archives(request):
    dict = Archives.objects.all()
    return render(request, "inference2/archives.html", {'result': dict})


def assign_archives(request, num=-1, type=None):
    if num == -1:
        return
    archive = Archives.objects.get(pk=num)
    if type:
        return globals()[type](request, archive)
    else:
        return index(request, archive)


def manual(request):
    filename = os.path.join(settings.MANUAL_PATH)
    wrapper = FileWrapper(file(filename))
    response = HttpResponse(
        wrapper, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Length'] = os.path.getsize(filename)
    response['Content-Disposition'] = 'attachment; filename=' + \
        os.path.basename(filename)
    return response


def getdict(request, archive=None):
    if not archive:
        archive = current_archive()

    filename = os.path.join(settings.DICT_DIRS, archive.algorithm + ".csv")
    if not os.path.exists(filename):
        return HttpResponse("No CSV found for Algorithm %s" % archive.algorithm)
    wrapper = FileWrapper(file(filename))
    response = HttpResponse(
        wrapper, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Length'] = os.path.getsize(filename)
    response['Content-Disposition'] = 'attachment; filename=' + \
        os.path.basename(filename)
    return response


def progress(request):
    contxt = {"K": request.session['idx']}
    if request.session.get('status', 0) == 2:
        progressbar_send(request, 1, 100, 1)
    return HttpResponse(json.dumps(contxt), content_type="application/json")


def progressbar_send(request, strt, stp, k, status=0):
    if request is not None:
        request.session.modified = True
        request.session['strt'] = strt
        request.session['stp'] = stp
        request.session['idx'] = [strt, stp, k]
        request.session['status'] = status
        request.session.modified = True
        request.session.save()
