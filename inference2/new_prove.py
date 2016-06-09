from openpyxl import load_workbook
import timeit
import copy
import time
import operator
excel = False
debug = False

if not excel:
    from models import Define3
if debug:
    import easygui

anaphoric_relations = []
prop_name = []
prop_var = []
plural_c = []
anaphora = ""
impl = ""
definite = []
psent = []
definite2 = []
ant_cond = []
never_used = []
conditionals = []
candd = []
rel_conj = []
ind_var = []
gen_var = []
conc = []
prop_sent = []
tagged_nouns = []
tagged_nouns2 = []
dv_nam = []
basic_objects = []
result_data = {}
cond_r = unichr(8835)
top = unichr(8868)
bottom = unichr(8869)
neg = unichr(172)
iff = unichr(8801)
mini_c = unichr(8658)
mini_e = unichr(8703)
implies = unichr(8866)
conditional = unichr(8594)
nonseq = unichr(8876)
xorr = unichr(8891)
idisj = unichr(8744)
cj = unichr(8896)
disj = unichr(8855)
equi = unichr(8660)
sn = 1
id_num=0

ne = u"\u2260"
l1 = u"\u2081"
l2 = u"\u2082"
l3 = u"\u2083"
l7 = u"\u2087"
l8 = u"\u2088"
ua = u"\u1d43"
ub = u"\u1d47"
uc = u"\u1d9c"
ud = u"\u1d48"
ue = u"\u1d49"
uf = u"\u1da0"
ug = u"\u1d4d"
ui = u"\u2071"
uk = u"\u1d4f"
um = u"\u1d50"
un = u"\u207f"
uo = u"\u1d52"
up = u"\u1d56"
ut = u"\u1d57"
uv = u"\u1d5b"
uu = u"\u1d58"
uw = u"\u02b7"
uy = u"\u02b8"
uj = u"\u02B2"
ul = u"\u02E1"
ur = u"\u02b3"
us = u"\u02e2"
uh = u"\u02b0"

tot_prop_name = []
tot_prop_sent = []
prop_var4 = [unichr(97 + t) for t in range(26)]
prop_var2 = [unichr(97 + t) + u"\u2081" for t in range(26)]
prop_var3 = [unichr(97 + t) + u"\u2082" for t in range(26)]
prop_var5 = [unichr(97 + t) + u"\u2083" for t in range(26)]
prop_var6 = [unichr(97 + t) + u"\u2084" for t in range(26)]
prop_var4 = prop_var4 + prop_var2 + prop_var3 + prop_var5 + prop_var6
idf_var2 = [unichr(122 - t) for t in range(25)]
idf_var3 = [unichr(122 - t) + l1 for t in range(25)]
idf_var4 = [unichr(122 - t) + l2 for t in range(25)]
idf_var2 = idf_var2 + idf_var3 + idf_var4
p = 1
subscripts = [l1,l2,l7]
if excel:
    wb4 = load_workbook('inference engine.xlsx')
    wb5 = load_workbook('dictionary.xlsx')
    w4 = wb4.worksheets[0]
    ws = wb5.worksheets[0]
else:
    ws = Define3.objects.all()

#
# >> 8835
# ta^ 8868
# co^ 8869
# ; 172
# <> 8801
# c^ 8658
# # 8703
# i^ 8866
# > 8594
# nf^ 8876
# ed^ 8891
# + 8744
# && 8896
# @ 8855
# if^ 8660


def remove_outer_paren(str1):

    if str1 == "":
        return ""
    elif str1.count(")") == 0:
        return str1

    j = 0
    # on very rare occasions we will encounter strings of the following form ((p))
    if str1[0] != "(" and str1[-1] != ")":
        return str1
    if str1[:2] == "((" and str1[-2:] == "))":
        d = 2
    else:
        d = 1

    for k in range(0, d):
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]

    return str1


def remove_redundant_paren(str1):
    j = 0
    str2 = str1[:2]
    str3 = str1[2:]
    if str2 == '((' and str3 == '))':

        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 1 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
    return str1

def mainconn(str1):

    ostring = copy.copy(str1)
    if os(str1):
        return ["", 0]
    if str1.find("&") < -1 and str1.find(idisj) < -1 and str1.find(iff) < -1 and str1.find(conditional) < -1 and \
            str1.find(implies) < -1 and str1.find(nonseq) < -1 and str1.find(xorr) < -1:
        return ["", 0]

    str3 = str1
    bool1 = False

    if str1[0] == "~":
        str1 = str1[1:]
        bool1 = True

    j = 0
    bool2 = False
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
            bool2 = True

    j = -1
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == conditional:
            f = -1
        if str2 == idisj or str2 == "&" or str2 == iff or str2 == implies or \
                str2 == nonseq or str2 == conditional or str2 == xorr:
            if str3 != str2 and str3 != "":
                j = j + 1

            str3 = str2

    if j == -1:
        i = ostring.find(str3)
        return [str3, i]
    k = -1
    j = -1
    while True:
        k += 1
        if k > 150:
            break
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]

            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1

            if j == -1 and (str2 == idisj or str2 == "&" or str2 == iff or str2 == implies
                            or str2 == nonseq or str2 == conditional or str2 == xorr):
                if bool1 and bool2:
                    return [str2, i+2]
                elif bool2 or bool1:
                    return [str2, i+1]
                else:
                    return [str2, i]
        else:
            str1 = str1[1:-1]

def isvariable(str3,kind=""):

    bool2 = True
    if str3 == None:
        return False

    if str3 == 'a':
        return False
    try:
        if str3 != "":
            str3 = str3.replace(l1, "")
            str3 = str3.replace(l2, "")
            str3 = str3.replace(neg, "")
            if len(str3) == 1 and str3.islower():
                bool2 = True
            else:
                bool2 = False
    except AttributeError:
        easygui.msgbox('error in isvariable function')
    if kind == "":
        return bool2
    else:
        return isidef

def os(str1):

    cnx = [xorr, iff, idisj, conditional, implies, nonseq, "&"]
    for i in range(0, len(cnx)):
        if str1.find(cnx[i]) > -1:
            os = False
            return os
    os = True
    return os

def enclose(str1):

    i = -1
    global subscripts
    while i < len(str1) -1:
        i += 1
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        str4 = str1[i+1:i+2]
        if str2.islower() and str4 in subscripts:
            if str3 == "~":
                str1 = str1[:i-1] + "(~" + str2 + str4 + ")" + str1[i+2:]
            else:
                str1 = str1[:i] + "(" + str2 + str4 + ")" + str1[i+2:]
            i += 4
        elif str2.islower():
            if str3 == "~":
                str1 = str1[:i-1] + "(~" + str2 + ")" + str1[i+1:]
            else:
                str1 = str1[:i] + "(" + str2 + ")" + str1[i+1:]
            i += 3
    return str1


def find_sentences(instring, cut_skel = False):

    if instring == None:
        return

    g = instring.count('(')
    h = instring.count(')')
    if g != h:
        easygui.msgbox('wrong number of parentheses in sentence:' + instring)
        return "stop"
    marker = False
    il = -1
    total = -1
    c = -1
    neg_value = []
    str1 = ""
    sent1 = []
    sent_type2 = []
    wneg = []
    output = [None] * 9
    # the skel name list names each single sentence after a greek letter, even if
    # the same sentence appears twice it obtains a different name on the second
    # appearance
    skel_nam = []
    sent_num = []
    if instring.find("~(") > -1:
        instring = instring.replace("~(", "(!")
    if instring.find(implies) > -1:
        str2 = implies
    elif instring.find(nonseq) > -1:
        str2 = nonseq
    str3 = mainconn(instring)
    str4 = str3[0]
    f = str3[1]
    id_num = []
    id_num.append(["1",str4,f])
    sent_num.append([1, '1', instring, str4,f])
    prtnum = 1
    skel_string = instring
    p = 947
    connectives = ["&", idisj, iff, conditional, nonseq, implies,xorr]
    arr1 = []
    mini_c2 = mini_c + neg
    instring2 = copy.copy(instring)
    instring = instring.strip()
    prt = copy.copy(instring)
    list1 = mainconn(instring)
    grandparent_type = list1[0]
    more_num = [unichr(945 + x) for x in range(24)]
    # prop_var = [unichr(97 + t) for t in range(26)]

    if os(instring) == False:
        temp_string = mainconn(instring)

        if instring.find(implies) > -1:
            str1 = implies
        elif instring.find(nonseq) > -1:
            str1 = nonseq
        else:
            if temp_string == iff:
                str1 = "bicond"
            elif temp_string == conditional:
                str1 = "cond"
            elif temp_string == "&":
                str1 = "cj"

        sent1.append(instring)
        neg_value.append("")
        sent_type2.append(str1)
        wneg.append(instring)
        skel_nam.append(None)

    j = 0
    n = 0
    for i in range(0, len(instring)):
        str1 = instring[i:(i + 1)]
        for o in connectives:
            if str1 == o:
                j += 1

    while n < j + 1:

        il += 1
        if il > 15:
            break

        e = 0
        l = len(instring)
        x = -1
        while x < l - 1:
            x += 1
            temp_string = instring[x: x + 1]
            if instring[x: x + 1] == "(":

                if marker == False:
                    z = x
                    marker = True

                total += 1
            elif instring[x: x + 1] == ")":
                total -= 1
                if total == -1:
                    marker = False
                    e += 1
                    c += 1

                    temp_sent = instring[z: x + 1]
                    otemp_sent = copy.copy(temp_sent)

                    if (len(instring) - len(temp_sent)) > 2:
                        if temp_sent in prt:
                            prtnum = findinlist(prt,sent_num,2,1,False)
                            numb = prtnum + "1"
                        else:
                            prtnum = ""
                            for bb in range(len(sent_num)-1,-1,-1):
                                str3 = sent_num[bb][2]

                                if temp_sent in sent_num[bb][2]:
                                    prtnum = sent_num[bb][1]
                                    break
                            # if prtnum == "":
                            #     easygui.msgbox('your sentences are not numbered properly')
                            g = len(prtnum) + 1
                            f = 0

                            for bb in range(len(sent_num)-1,-1,-1):
                                temp_sn = sent_num[bb][1]
                                h = temp_sn[:g-1]
                                hh = sent_num[bb][0]
                                if g > sent_num[bb][0]:
                                    break
                                if sent_num[bb][0] == g and temp_sn[:g-1] == prtnum:
                                    f += 1

                            f += 1
                            if f < 10:
                                numb = prtnum + str(f)
                            else:
                                numb = prtnum + more_num[f-10]

                        if n == 7:
                            pp = 7

                        prt = temp_sent
                        temp_mc = mainconn(temp_sent)
                        mc = temp_mc[0]
                        str3 = temp_mc[0]
                        g = temp_mc[1]
                        mc_num = temp_mc[1]
                        sent_num.append([len(numb), numb, temp_sent, str3,g])


                        if os(temp_sent):
                            # n counts the number of single sentences
                            n += 1
                            if temp_sent.find("~") > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace("~", "")

                            elif temp_sent.find(mini_c2) > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace(mini_c2, mini_c)

                            elif temp_sent.find(mini_c2) == -1 and temp_sent.find(mini_c) > -1:
                                neg_value.append("")
                            elif temp_sent.find("~") == -1:
                                neg_value.append("")
                            else:
                                break  # stop
                        else:
                            neg_value.append("")

                        sent1.append(temp_sent)
                        wneg.append(otemp_sent)
                        id_num.append([numb,mc,mc_num])

                        if os(otemp_sent):
                            p += 1
                            skel_nam.append([otemp_sent, unichr(p)])
                        else:
                            skel_nam.append(None)
                    else:
                        instring = instring[1:len(instring) - 1]
                        l = len(instring)
                        x = -1
                        c -= - 1
                        e -= - 1

        total = -1
        marker = False
        w = -1

        if n < j + 1:
            if len(sent1) > w:
                while w + 1 < len(sent1):
                    if w == 13:
                        pp = 7
                    w += 1
                    str21 = sent1[w]
                    if not os(str21) and w != 0:
                        if str21 not in arr1:
                            instring = str21
                            arr1.append(instring)
                            break

    for i in range(len(sent1)):
        temp_string = sent1[i]
        if temp_string.find("(!") > -1:
            sent1[i] = sent1[i].replace("(!", "~(")
            wneg[i] = wneg[i].replace("(!", "~(")
    m = 0
    bool1 = False
    wid = copy.copy(instring2)
    for i in range(len(id_num)):
        if (id_num[i][1] == iff or id_num[i][1] == conditional) \
                and not bool1:
            str1 = wneg[i]
            bool1 = True
            m = i
            break
    for i in range(len(wneg)):
        if os(wneg[i]):
            str1 = str1.replace(skel_nam[i][0],skel_nam[i][1])
            wid = wid.replace(skel_nam[i][0],skel_nam[i][1])

    skel_wid = wid
    skel_string = str1
    skel_string = remove_outer_paren(skel_string)
    if skel_string.find("(!") > -1:
        skel_string = skel_string.replace("(!","~(")
        wid = wid.replace("(!","~(")

    output[0] = sent1
    output[1] = neg_value
    output[2] = sent_type2
    output[3] = wneg
    output[4] = id_num
    output[5] = skel_string
    output[6] = skel_nam
    # m is the first sentence to be used when changing variables in a definition
    output[7] = m
    output[8] = skel_wid

    return output

def add_to_dv(dv_nam,all_sent,m,k,idf_var,str2):


    if isvariable(str2) == False:
        str3 = findinlist(str2, dv_nam,1,0)
        if str3 == None:
            telist7 = [idf_var[0], str2]
            all_sent[m][k] = idf_var[0]
            del idf_var[0]
            dv_nam.append(telist7)
        else:
            all_sent[m][k] = str3


def word_sub(idf_var, dv_nam, tot_sent, all_sent, words,id_num):

    all_sent = remove_duplicates(all_sent,0)
    relations = words[18]
    relations2 = words[19]
    pronouns = words[24]
    num = [4, 5, 13, 14, 17, 18, 22, 26, 30,34,35,36,51,52,63,64,65,67]
    # num2 = [9,15,19,23,27,31,49]
    global sn
    m = -1
    while m < len(all_sent) -1:
        m += 1
        if all_sent[m][47] == "no word sub":
            return

        bool1 = False
        list4 = copy.deepcopy(all_sent[m][46])
        old_sent = all_sent[m][0]
        oldp = all_sent[m][42]
        # old_list = copy.deepcopy(all_sent[m])
        for i in range(len(all_sent[m][46])):
            k = all_sent[m][46][i]
            str2 = all_sent[m][k]
            if (k == 8 or k == 12) and str2 != None:
                bool1 = True
                str5 = findinlist(str2,words[16],0,1)
                all_sent[m][8] = str5
                all_sent[m][12] = None
            if k in num and all_sent[m][45] != k:
                bool1 = True
                if str2 != None and str2 not in pronouns and str2 != 'there':
                    dummy = add_to_dv(dv_nam,all_sent,m,k,idf_var,str2)
                    list4.remove(k)

        if bool1:
            new_sent = build_sent(all_sent[m])
            newp = name_sent(new_sent)
            all_sent[m][0] = new_sent
            all_sent[m][42] = newp
            dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"SUB",id_num)
            all_sent[m][46] = list4
            bool1 = False
    return

def assigned_var(str1, dv_nam, idf_var):

    bool1 = False
    for i in range(len(dv_nam)):
        if dv_nam[i][1] == str1:
            bool1 = True
            return dv_nam[i][0]

    if bool1 == False:
        str2 = idf_var[0]
        list1 = [str2, str1]
        dv_nam.append(list1)
        del idf_var[0]
        return str2

def there(all_sent,m,tot_sent,def_sent):

    global sn
    old_sent = all_sent[m][0]
    oldp = all_sent[m][42]
    all_sent[m][5] = all_sent[m][14]
    all_sent[m][3] = all_sent[m][10]
    all_sent[m][4] = all_sent[m][13]
    all_sent[m][14] = None
    all_sent[m][10] = None
    all_sent[m][13] = None
    new_sent = build_sent(all_sent[m])
    newp = name_sent(new_sent)
    all_sent[m][0] = new_sent
    def_sent.append(new_sent)
    all_sent[m][42] = newp
    bool1 = check_dimension(tot_sent,1,new_sent)
    if not bool1:
        dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"DE there")


def define(tot_sent, all_sent, idf_var, dv_nam,words,rep_rel,identities):

    all_sent = remove_duplicates(all_sent,0)
    num = [3,5,10,14,16,18,20,22,24,26,28,30,32,34]
    pronouns2 = copy.deepcopy(words[24])
    if "it" in pronouns2:
        pronouns2.remove("it")
    pronouns = pronouns2
    determinative = words[2]
    definitions = words[16]
    posp = words[28]
    atomic_relations = words[22]
    atomic_relata = words[23]
    def_relat = ['IA','IG','=','H']
    global sn
    global anaphora
    global gen_var
    bool1 = False
    def_sent = []
    defined = []
    universal = ['all','no','any','no'+us,"a","many"+un,"a"+ua]
    universal2 = ['no','no'+us]
    indefinite = ['a',"many"+un, 'no' + us,"a"+ua]
    universal3 = ['all','any']

    for i in range(len(dv_nam)):
        if i == 3:
            bb = 7
        g = findposinlist(dv_nam[i][1],definitions,0)
        if g > -1:
            list1 = [None] * 70
            list1[5] = dv_nam[i][0]
            list1[9] = '='
            list1[14] = dv_nam[i][1]
            str1 = build_sent(list1)
            str2 = name_sent(str1)
            list1[0] = str1
            list1[42] = str2
            list1[41] = 1
            all_sent.append(list1)

    m = -1
    g = (len(all_sent))
    while m < g - 1:
        m += 1
        if all_sent[m][45] > 2:
            all_sent.append(all_sent[m])
            del all_sent[m]
            m -= 1
            g -= 1
    for d in range(0,8):

        if d == 1:
            num = [0]
        elif d == 2:
            num = [3,10,16,20,24]
        elif d == 5:
            num = [5,14]
        elif d == 6:
            num = [0]
        elif d == 7:
            num = [9,14,48]
        m = -1
        while m < len(all_sent) -1:
            m += 1
            bool2 = False
            if m == 13 and d == 2:
                bb = 7
            if m > 100:
                easygui.msgbox('in the define function you are caught in an infinite loop')
            for i in num:
                if d == 0 and m == 1 and i == 5:
                    bb = 7
                str1 = all_sent[m][i]
                relat = all_sent[m][9]
                old_sent = all_sent[m][0]
                oldp = all_sent[m][42]
                if d == 0:
                    if str1 == 'there' and i == 5:
                        dummy = there(all_sent,m,tot_sent,def_sent)
                        m -= 1
                        break
            #this is for those sentences whose noun was once part of a relative pronoun
                    elif i == all_sent[m][45]:
                        str3 = findinlist(all_sent[m][i],tagged_nouns2,1,0)
                        if str3 == None:
                            all_sent.append(all_sent[m])
                            del all_sent[m]
                            m -= 1
                            break
                        all_sent[m][i] = str3
                        new_sent = build_sent(all_sent[m])
                        newp = name_sent(new_sent)
                        dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"rel pro")
                        all_sent[m][0] = new_sent
                        all_sent[m][42] = newp
                        all_sent[m][45] = None
                        m -= 1
                    elif (str1 in pronouns or str1 in determinative) and str1 not in universal:
                        bool1 = True
                        if str1 in pronouns:
                            str2 = "pronoun"
                        else:
                            str2 = "determinative"
                        definition = findinlist(str1,definitions,0,1)
                        if all_sent[m][0] not in def_sent:
                            dummy = def_rn(defined,definition, str1,0, tot_sent, \
                                dv_nam, idf_var,words, all_sent,m,rep_rel,str2,i)

                            def_sent.append(all_sent[m][0])
                            del all_sent[m]
                            m -= 1
                            break
                elif d == 1 and m == 0:
                    dummy = division(tot_sent,all_sent,words,0)
                    m = len(all_sent)
                    break
                elif d == 2 and all_sent[m][i] in universal2:
                    definition = findinlist(str1,definitions,0,1)
                    dummy = def_rn(defined,definition, str1,0, tot_sent, \
                        dv_nam, idf_var,words, all_sent,m,rep_rel,"determinative",i)
                    def_sent.append(all_sent[m][0])
                    del all_sent[m]
                    m -= 1
                    break

                elif d == 3 and all_sent[m][i] in indefinite:
                    definition = findinlist(str1,definitions,0,1)
                    dummy = def_rn(defined,definition, str1,0, tot_sent, \
                        dv_nam, idf_var,words, all_sent,m,rep_rel,"determinative",i)
                    def_sent.append(all_sent[m][0])
                    if all_sent[m][57] != None:
                        if all_sent[m][57] == i:
                            if all_sent[m][3] == "":
                                anaphora = all_sent[m][5]
                                all_sent[m][57] = None
                    del all_sent[m]
                    m -= 1
                    break

                elif d == 4 and all_sent[m][i] in universal3:
                    definition = findinlist(str1,definitions,0,1)
                    dummy = def_rn(defined,definition, str1,0, tot_sent, \
                        dv_nam, idf_var,words, all_sent,m,rep_rel,"determinative",i)
                    def_sent.append(all_sent[m][0])
                    del all_sent[m]
                    m -= 1
                    break

                elif d == 5 and all_sent[m][i] == 'it':
                    list2 = copy.deepcopy(all_sent[m])
                    list2[i] = anaphora
                    list2[0] = None
                    dummy = new_sentence(tot_sent,all_sent[m],list2,"","",1,"DE it")
                    all_sent.append(list2)
                    del all_sent[m]
                    m -= 1
                    break
                elif d == 6 and m == 0:
                    dummy = division(tot_sent,all_sent,words,1)
                    m = len(all_sent)
                    break
                elif d == 7:
                    adverb = False
                    id = False
                    kind = ""
                    if all_sent[m][43] != 'cc':
                        if relat in def_relat and i == 14:
                            definiendum = findinlist(str1,dv_nam,0,1)
                            if definiendum in atomic_relata:
                                bool2 = False
                            else:
                                bool2 = True
                        elif i == 48 and all_sent[m][48] != None:
                            definiendum = str1
                            bool2 = True
                            adverb = True
                            kind = 'R'
                        elif i == 9 and relat == 'AS':
                            kind = 'AS'
                            definiendum = str1
                            bool2 = True
                        elif relat == '=' and all_sent[m][41] == 1:
                            id = True
                            definiendum = all_sent[m][14]
                            all_sent[m][41] = None
                        elif relat == "=" and all_sent[m][41] != 1:
                            bool2 = False
                            identities.append([[all_sent[m][5],all_sent[m][14]],"","",""])
                            del all_sent[m]
                            m -= 1
                        elif i == 9 and relat != 'IA' and relat != 'IG' and relat != '=' \
                            and str1 not in atomic_relations:
                            definiendum = str1
                            bool2 = True
                            kind = 'R'
                        if (bool2 and isdefineable(all_sent[m]) and definiendum != None and \
                                definiendum != '') or id:
                            g = findposinlist(definiendum,definitions,0)
                            definition = definitions[g][1]
                            if definition == 'natural':
                                definition = "(c'=" + definiendum + ") & (d'=natural_whole) & ((bIGc') " + conditional \
                                + " (bIGd'))"
                            pos = definitions[g][2]
                            circ = definitions[g][3]
                            circ2 = all_sent[m][43]
                            basic_molecule = definitions[g][4]
                #this prevents us from getting caught in an infinite loop.
                            if basic_molecule == 'b' and all_sent[m][9] == 'IG':
                                break
                            if circ2 == 'c':
                                circ += circ2
                            if (relat == 'IA' and pos == 'a') or (relat == 'IG' and pos == 'n') \
                                or pos == 'r' or pos == 'e' or pos == 's' or (relat== '=' and pos == 'n') or adverb or id:
                                if definition != None and all_sent[m][0] not in def_sent:
                                    def_sent.append(all_sent[m][0])
                                    if definiendum == 'any':
                                        bb = 8
                                    if id:
                                        if 'definite' not in definition:
                                            break
                                    dummy = def_rn(defined,definition, definiendum,0,tot_sent,dv_nam, idf_var,\
                                        words,all_sent,m,rep_rel,kind,i,circ)
                                    break

# if we state that something is not a concept then we need to falisfy that
    dummy = concept(all_sent,tot_sent,dv_nam,definitions,posp)
    return

def concept(all_sent,tot_sent,dv_nam,definitions,posp):

    global sn
    str1 = ""
    list2 = []
    list1 = [None] * 70
    for i in range(len(dv_nam)):
        if dv_nam[i][1] == 'concept'+un or dv_nam[i][1] == 'concept'+ua:
            str1 = dv_nam[i][0]
        if str1 != "":
            for j in range(len(all_sent)):
                if all_sent[j][9] == 'IG' and all_sent[j][14] == str1:
                    str2 = all_sent[j][5]
                    con = findinlist(str2,dv_nam,0,1)
                    pos = findinlist(con,posp,0,1)
                    if pos == None:
                        bb = 7
                    if con != None:
                        if pos == 'a':
                            str4 = 'IA'
                        elif pos == 'n':
                            str4 = 'IG'
                        b = 0
                        bool1 = False
                        for k in range(len(all_sent)):
                            if all_sent[k][9] == str4 and all_sent[k][14] == str2 and \
                                str1 != str2 and str2 not in list2:
                                bool1 = True
                                str6 = all_sent[k][5]
                                list2.append(str2)
                                b += 1
                        if b > 1:
                            easygui.msgbox('you have not coded for multiple concepts')
                        olda = "(" + "b" + ' = ' + con + ")"
                        oldc = "(" + "c " +str4 + " b" + ")"
                        rn1 = ""
                        if str2 != "b":
                            rn1 = "(" + "b" + mini_c + str2 + ") & (" + "c" + mini_c + str6 + ")"
                        else:
                            rn1 = "(" + "c" + mini_c + str6 + ")"
                        newa = "(" + str2 + ' = ' + con + ")"
                        newc = "(" + str6 + " " + str4 + " " + str2 + ")"
                        oldcon = olda + " " + conditional + " " + oldc
                        sn += 1
                        tot_sent.append([sn,oldcon,"","","NC concept " + con,"",""])
                        sn += 1
                        tot_sent.append([sn,rn1,"","","RN",""])
                        oldp = name_sent(newa)
                        newp = name_sent(newc)
                        if not bool1:
                            list1[0] = newc
                            list1[5] = str6
                            list1[9] = str4
                            list1[14] = str2
                            list1[40] = False
                            list1[42] = newp
                            all_sent.append(list1)
                        str3 = newa + " " + conditional + " " + newc
                        str3p = oldp + " " + conditional + " " + newp
                        sn += 1
                        tot_sent.append([sn,str3,str3p,"","SUB",sn-2,sn-1])
                        return

def name_sent(str1,bool2 = False,str4 = ""):
    global prop_name
    global prop_var

    no_space = copy.copy(str1)
    if str1.find('~') > -1:
        no_space = str1.replace("~","")
        ng = '~'
    else:
        ng = ''

    no_space = remove_outer_paren(no_space)
    no_space = no_space.replace(" ","")

    if bool2:
        if str4 == 'something':
            no_space = no_space.replace("something","some thing")
        elif str4 == 'anything':
            no_space = no_space.replace("anything","any thing")
        elif str4 == 'everything':
            no_space = no_space.replace("everything","every thing")

    h = findinlist(no_space,prop_name,1,0)
    if h != None:
        return ng + h
    else:
        prop_name.append([prop_var[0], no_space, str1])
        str2 = prop_var[0]
        del prop_var[0]
        return ng + str2

def assign_var(str1, str2, dv_nam):

    list1 = [str2, str1]
    dv_nam.append(list1)

def determ(idf_var, all_sent, tot_sent,words,dv_nam,m):

    global sn
    # def_det contains the determinatives and their definitions
    def_det = words[15]
    detm = words[2]
    # num = [3, 10, 16, 20, 24, 28, 32]
    list1 = copy.deepcopy(all_sent[m][46])
    num = [3,10]
    for i in range(len(all_sent[m][46])):
        k = all_sent[m][46][i]
        if k in num:
            str1 = all_sent[m][k]
            if str1 in detm:
                defin = findinlist(str1,def_det,0,2)
                all_sent[m][k] = None
                list1.remove(k)
                if i == 10:
                    if "z~Rc" in defin:
                        defin = defin.replace("z~Rc","c~Rz")
                    else:
                        defin = defin.replace("zRc","cRz")
                dummy = def_rn(defin,str1,tot_sent, dv_nam,idf_var)
    all_sent[m][46] = list1


def insert_space(str, integer):
    return str[0:integer] + ' ' + str[integer:]

def space_words(str1):
    # this function place a space between variables and relations
    # so as to make it easier to categorize words in a sentence
    str1 = str1.replace("~", " ~ ")
    str1 = str1.replace("=", " = ")
    i = 0
    while i + 1 < len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        nxt_str = str1[(i + 1):(i + 2)]
        if nxt_str.isupper() == True and temp_str.islower() == True:
            str1 = insert_space(str1, i + 1)
        elif nxt_str.islower() == True and temp_str.isupper() == True:
            str1 = insert_space(str1, i + 1)
    return str1


def id_def(list1):
    # this function picks out that variables in the id sentences of the
    # definition

    list2 = []
    has_plural = False
    for i in range(len(list1[0])):
        if os(list1[0][i]) and "=" in list1[0][i]:
            str1 = list1[0][i]
            g = str1.find("=")
            var = str1[1:g]
            wrd = str1[g+1:-1]
            if isvariable(var):
                if wrd == 'plural form':
                    has_plural = True
                if not isvariable(wrd):
                    list2.append([var,wrd])
    return [list2,has_plural]

def is_atomic(list1, atomic_relations):

    relat = [9,15,19,23,27,31]
    must_be_blank = [2,3,4,6,7,10,11,12,16,17,20,21,24,25,28,29,32,33]
    noun = [5,14,18,26,30,34]
    for i in range(len(list1)):
        str1 = list1[i]
        if str1 != None:
            if i in relat:
                if str1 in atomic_relations:
                    pass
                else:
                    pass
            elif i in must_be_blank:
                return False
            elif i in noun:
                if not isvariable(str1):
                    return False
    return True


def isdefineable(list1):

    must_be_blank = [2,3,4,6,7,10,11,13,15,16,17,18,20,21,23,24,25,27,28,29,31,32,33,\
        35,36,49,50,51,52]
    must_be_variable = [5,14,22,30]

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None:
            if not isvariable(list1[i]):
                return False
    return True

def build_sent(list1):

    str1 = "("
    num = [11,1,2,47,3,4,5,66,67,35,48,59,6,8,9,48,12,10,13,14,36,60,63,49,15,16,17,18,\
           61,64,50,19,20,21,22,62,65,51,23,24,25,26,52,27,28,\
           29,30,31,32,33,34]
    for i in num:
        temp_str = list1[i]
        if temp_str != None and temp_str != "":
            if str1 == "(":
                str1 = str1 + temp_str
            else:
                str1 = str1 + " " + temp_str

    return str1 + ")"

def build_sent2(list1,bool1 = False):

    if bool1:
        g = len(list1) - 3
    else:
        g = len(list1)
    for i in range(0,g):
        if i == 0:
            str1 = list1[i]
        else:
            str1 += " " + list1[i]
    return "(" + str1 + ")"

def build_sent_list(list1):
    # this list builder does not have the ~ separated from the sentence
    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i]
        else:
            str2 = str2 + ' & ' + list1[i]
    return str2

def build_sent_list2(list1):
    # this list builder has the ~ separated from the sentence
    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i][1] + list1[i][0]
        else:
            str2 = str2 + ' & ' + list1[i][1] + list1[i][0]
    return str2

def remove_duplicates(list1,i):

    list2 = []
    j = -1
    while j < len(list1) -1:
        j += 1
        if list1[j][i] in list2:
            del list1[j]
            j -= 1
        else:
            list2.append(list1[j][i])

    return list1

def remove_duplicates2d(list1,i,j):

    list2 = []
    j = -1
    while j < len(list1) -1:
        j += 1
        list3 = [list1[j][0],list1[j][1]]
        d = findposinlist(list1[j][0],list2,0)
        if d > -1:
            if list2[d][1] == list1[j][1]:
                del list1[j]
                j -= 1
        else:
            list2.append([list1[j][0],list1[j][1]])
    return list1

def id_sent(list4,all_sent,irrel_group = []):
    # this function turns the dv_nam into a string of conjuncts

    global gen_var
    global ind_var
    global definite2

    dv_nam = []
    # this loop removes duplicates in a multidimensional list
    for i in range(len(list4)):
        if list4[i] not in dv_nam:
            dv_nam.append(list4[i])

    prop2 = None
    str2 = None
    for i in range(len(dv_nam)):
        if isvariable(dv_nam[i][0]):
            irrel_group.append(dv_nam[i][0])
            list1 = [None] * 70
            list1[5] = dv_nam[i][0]
            list1[9] = "="
            list1[14] = dv_nam[i][1]
            str1 = build_sent(list1)
            list1[0] = str1
            bool1 = check_dimension(all_sent,0,str1)
            prop1 = name_sent(str1)
            list1[42] = prop1
            if not bool1:
                all_sent.append(list1)
            if dv_nam[i][0] not in gen_var and dv_nam[i][0] not in ind_var \
                    and dv_nam[i][0] not in definite2:
                definite2.append(dv_nam[i][0])
        else:
            str1 = '(' + dv_nam[i][0] + "=" + dv_nam[i][1] + ')'
            prop1 = name_sent(str1)
        if str2 == None:
            str2 = str1
            prop2 = prop1
        else:
            str2 = str2 + ' & ' + str1
            prop2 = prop2 + " & " + prop1
    return [str2, prop2]

def cut_string(str1,str2):

    if str1.find(str2) == -1:
        return str1
    g = str1.find(str2)
    str1 = str1[:g]
    str1 = str1.strip()
    return str1

def remove_values_from_list(the_list, val):

    while val in the_list:
        the_list.remove(val)
    return the_list

def divide_sent(words, list2, idf_var,tot_sent,all_sent):

    global sn
    global impl
    redundant = words[21]
    conn = words[4]
    relations = words[6]
    nonsq = False
    for i in range(len(list2)):
        str2 = list2[i][1]
        str2 = str2.lower()
        str3 = name_sent(str2)
        tot_sent.append([list2[i][0],str2,str3,"","","",""])
        all_sent.append(str2)

    # the following changes it is necessary that if p then q to if p then it is necessary q
    modals = ['possible','necessary','impossible']
    h = copy.copy(len(all_sent))
    i = -1
    while i < h - 1:
        i += 1
        all_sent[i] = all_sent[i].strip()
        all_sent[i] = all_sent[i].split(" ")
        all_sent[i] += ["("+tot_sent[i][1]+")",tot_sent[i][2],""]
        for j in range(len(all_sent[i])):
            if all_sent[i][j] in modals:
                if all_sent[i][j+1] == 'that' and all_sent[i][j+2] == 'if':
                    old_sent = all_sent[i][-3]
                    old_p = tot_sent[i][2]
                    str1 = copy.copy(all_sent[i][j])
                    del all_sent[i][j+1]
                    del all_sent[i][j]
                    del all_sent[i][j-1]
                    del all_sent[i][j-2]
                    g = all_sent[i].index('then')
                    all_sent[i].insert(g+1,'it'+up)
                    all_sent[i].insert(g+2,'is'+ua)
                    all_sent[i].insert(g+3,str1)
                    all_sent[i].insert(g+4,'that')
                    del all_sent[i][-1]
                    del all_sent[i][-1]
                    new_sent = build_sent2(all_sent[i])
                    newp = name_sent(new_sent)
                    all_sent[i] += [new_sent,newp,""]
                    dummy = new_sentence2(old_sent,old_p,new_sent,newp,tot_sent,"modal transfer")
                    break
    str1 = ""
    bool1 = False

    for i in range(len(all_sent)):
        old_sent = all_sent[i][-3]
        old_p = all_sent[i][-2]
        str2 = ""
        for j in range(len(redundant)):
            str1 = redundant[j]
            if str1 in all_sent[i]:
                bool1 = True
                all_sent[i] = remove_values_from_list(all_sent[i],str1)
                if str2 == '':
                    str2 = str1
                else:
                    str2 += "," + str1
        if bool1:
            bool1 = False
            del all_sent[i][-1]
            del all_sent[i][-1]
            del all_sent[i][-1]
            new_sent = build_sent2(all_sent[i])
            newp = name_sent(new_sent)
            all_sent[i] += [new_sent,newp,""]
            rule = "RD " + str2
            dummy = new_sentence2(old_sent,old_p,new_sent,newp,tot_sent,rule)

    rule = "DF "
    g = len(all_sent)
    i = -1
    impl = ""
    while i < g - 1:
        i += 1
        for j in range(len(all_sent[i])):
            if all_sent[i][j] in conn:
                str4 = all_sent[i][j]
                str5 = ""
                str6 = ""
                if all_sent[i][j] == 'follow':
                    str1 = nonseq
                    impl = nonseq
                    del all_sent[i][j+1]
                    del all_sent[i][j-1]
                    del all_sent[i][j-2]
                    j -= 2
                elif all_sent[i][j] == 'then' + ua:
                    str1 = conditional
                    str5 = "ant"
                    str6 = "con"
                else:
                    str1 = implies
                    impl = implies
                old_sent = all_sent[i][-3]
                old_p = all_sent[i][-2]
                del all_sent[i][-1]
                del all_sent[i][-1]
                del all_sent[i][-1]
                ant = all_sent[i][:j]
                cons = all_sent[i][j+1:]
                ant_s = build_sent2(ant)
                cons_s = build_sent2(cons)
                antp = name_sent(ant_s)
                consp = name_sent(cons_s)
                ant += [ant_s,antp,str5]
                cons += [cons_s,consp,str6]
                all_sent.append(ant)
                all_sent.append(cons)
                rule += str4
                del all_sent[i]
                i -= 1
                new_sent = "(" + ant_s + " " + str1 + " " + cons_s + ")"
                new_p = "(" + antp + ' ' + str1 + ' ' + consp + ")"
                dummy = new_sentence2(old_sent,old_p,new_sent,new_p,tot_sent,rule)
                break

    g = len(all_sent)
    i = -1
    while i < g - 1:
        i += 1
        for j in range(len(all_sent[i])):
            if all_sent[i][j] == 'that':
                old_sent = all_sent[i][-2]
                old_p = all_sent[i][-1]
                del all_sent[i][-1]
                del all_sent[i][-1]
                del all_sent[i][-1]
                ant = all_sent[i][:j]
                cons = all_sent[i][j+1:]
                cons_s = build_sent2(cons)
                consp = name_sent(cons_s)
                del ant[0]
                ant.insert(0,consp)
                ant_s = build_sent2(ant)
                antp = name_sent(ant_s)
                if consp in idf_var:
                    idf_var.remove(consp)
                    dv_nam.append([consp, cons_s,1])
                ant += [ant_s,antp]
                cons += [cons_s,consp]
                all_sent.append(ant)
                all_sent.append(cons)
                del all_sent[i]
                i -= 1
                dummy = new_sentence2(old_sent,old_p,ant_s,antp,tot_sent,'DF that')
                break

    # g = len(all_sent)
    # i = -1
    # while i < g - 1:
    #     i += 1
    #     for j in range(len(all_sent[i])):
    #         if all_sent[i][j] == 'and':
    #             old_sent = all_sent[i][-2]
    #             old_p = all_sent[i][-1]
    #             del all_sent[i][-1]
    #             del all_sent[i][-1]
    #             ant = all_sent[i][:j]
    #             cons = all_sent[i][j+1:]
    #             cons_s = build_sent2(cons)
    #             consp = name_sent(cons_s)
    #             del ant[0]
    #             ant.insert(0,consp)
    #             ant_s = build_sent2(ant)
    #             antp = name_sent(ant_s)
    #             if consp in idf_var:
    #                 idf_var.remove(consp)
    #                 dv_nam.append([consp, cons_s,1])
    #             ant += [ant_s,antp]
    #             cons += [cons_s,consp]
    #             all_sent.append(ant)
    #             all_sent.append(cons)
    #             del all_sent[i]
    #             i -= 1
    #             dummy = new_sentence2(old_sent,old_p,ant_s,antp,tot_sent,'df that')
    #             break

    return



def rel_repl(all_sent,tot_sent,words,dv_nam,idf_var,id_num):

    relations = words[6]
    pos = words[28]
    doubles = words[31]
    bool1 = False
    for j in range(len(all_sent)):
        i = -1
        while i < len(all_sent[j])-3:
            i += 1
            if "," in all_sent[j][i]:
                has_comma = True
                str3 = all_sent[j][i]
                str3 = str3.replace(",","")
            else:
                str3 = all_sent[j][i]
                has_comma = False
            if str3 == 'spies':
                bb = 8
            bool2 = check_dimension(doubles,0,str3)
            bool3 = False
            if bool2:
                str4 = all_sent[j][i] + " " + all_sent[j][i+1]
                bool3 = check_dimension(doubles,1,str4)
                if bool3:
                    str3 += " " + all_sent[j][i+1]
            str2 = findinlist(str3,relations,0,1)
            if str2 != None:
                bool1 = True
                g = findposinlist(str3,dv_nam,0)
                if g == -1:
                    dv_nam.append([str3,str2])
                if has_comma:
                    all_sent[j][i] = str2 + ","
                else:
                    all_sent[j][i] = str2
            if bool3:
                if str2 != None:
                    del all_sent[j][i+1]
                else:
                    i += 1
        old_sent = all_sent[j][-3]
        oldp = all_sent[j][-2]
        old_type = all_sent[j][-1]
        new_sent = build_sent2(all_sent[j],True)
        newp = name_sent(new_sent)
        all_sent[j][-3] = new_sent
        all_sent[j][-2] = newp
        all_sent[j][-1] = old_type
        dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"SUB",id_num)
        all_sent[j] = categorize_words(words,all_sent[j],idf_var,all_sent,0,True)

#here we change not a into no and other synonyms
    num = [8,49,50,51,52]
    cat = ['many'+un,'any'+un]
    for i in range(len(all_sent)):
        old_sent = all_sent[i][0]
        oldp = all_sent[i][42]
        bool2 = False
        for j in num:
            bool1 = False
            if all_sent[i][j] == "not" or all_sent[i][j] == neg:
                if j == 8:
                    if all_sent[i][10] == "a":
                        all_sent[i][10] = 'no' + us
                        bool1 = True
                        bool2 = True
                    elif all_sent[i][10] in cat:
                        all_sent[i][10] = 'no'
                        bool1 = True
                        bool2 = True
                elif j == 49:
                    if all_sent[i][18] == "a":
                        all_sent[i][18] = 'no' + us
                        bool1 = True
                        bool2 = True
                    elif all_sent[i][18] in cat:
                        all_sent[i][18] = 'no'
                        bool1 = True
                        bool2 = True
                elif j == 50:
                    if all_sent[i][22] == "a":
                        all_sent[i][22] = 'no' + us
                        bool1 = True
                        bool2 = True
                    elif all_sent[i][22] in cat:
                        all_sent[i][22] = 'no'
                        bool1 = True
                        bool2 = True
                elif j == 51:
                    if all_sent[i][26] == "a":
                        all_sent[i][26] = 'no' + us
                        bool1 = True
                        bool2 = True
                    elif all_sent[i][26] in cat:
                        all_sent[i][26] = 'no'
                        bool1 = True
                        bool2 = True
                elif j == 52:
                    if all_sent[i][30] == "a":
                        all_sent[i][30] = 'no' + us
                        bool1 = True
                        bool2 = True
                    elif all_sent[i][30] in cat:
                        all_sent[i][30] = 'no'
                        bool1 = True
                        bool2 = True
                if bool1:
                    all_sent[i][j] = None
                    bool1 = False
        if bool2:
            new_sent = build_sent(all_sent[i])
            newp = name_sent(new_sent)
            dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"ND","")
            bool2 = False
            all_sent[i][0] = new_sent
            all_sent[i][42] = newp

    return


def new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,rule,anc1="",conn = iff,anc2="",anc3=""):

    global sn
    str5 = old_sent + " " + conn + " " + new_sent
    str6 = oldp + " " + conn + " " + newp
    bool1 = check_dimension(tot_sent,1,str5)
    if not bool1:
        sn += 1
        if sn == 8:
            bb = 7
        tot_sent.append([sn,str5,str6,"",rule,anc1,anc2,anc3])

def build_app(list1):

    str1 = list1[0]
    for i in range(1,len(list1)):
        str1 += ", " + list1[i]
    return str1

def division(tot_sent, all_sent,words,kind):

    global anaphora
    num = [35,36]
    univ = ['all','any','no','no'+us]
    list2 = []
    p = len(all_sent) -2
    g = 0
    pos = words[28]

    for k in range(0,5):
        if k == 4:
            num = [15,19]
        elif k == 1:
            num = [66]
        elif k == 2:
            num = [4,13,17,21,25,33]
        elif k == 3:
            num = [59,60,61,62]
        if kind == 3:
            m = p
        else:
            m = -1
        while m < len(all_sent) -1:
            m += 1
            if k == 4 and m == 13:
                bb = 8
            old_sent = all_sent[m][0]
            oldp = all_sent[m][42]
            for i in num:
                list1 = [None] * 70
                if k == 0 and (kind == 1 or kind == 3) and all_sent[m][i] != None:
                    rule = "CIA"
                    if i == 35:
                        j = 5
                    elif i == 36:
                        j = 14
                    elif i == 37:
                        j = 18
                    elif i == 38:
                        j = 22
                    str1 = all_sent[m][j]
                    all_sent[m][j] = all_sent[m][i]
                    list1[14] = str1
                    list1[5] = all_sent[m][i]
                    list1[9] = "IG"
                    all_sent[m][i] = None
                    if kind == 3:
                        list2.append(list1)
                        all_sent.append(list1)
                        g += 1
                    else:
                        dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,1)
                elif k == 1 and all_sent[m][i] != None and kind == 0:
                    all_sent[m][66] = None
                    list1 = [None] * 70
                    list1[5] = all_sent[m][67]
                    all_sent[m][67] = None
                    rule = "DE and" + uc
                    for i in range(6,20):
                        list1[i] = all_sent[m][i]
                    dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,1)
                elif k == 2 and all_sent[m][i] != None:
                    if i == 13:
                        n = 10
                        r = 9
                    else:
                        r = i-2
                        n = i-1
                    if scope_uni(all_sent,m,i) and (kind == 0 or kind == 3):
                        rule = 'ADJ E'
                        if all_sent[m][8] != None or all_sent[m][12] != None:
                            str7 = "~"
                            all_sent[m][8] = None
                            all_sent[m][12] = None
                        else:
                            str7 = None
                        list1[8] = str7
                        list1[3] = all_sent[m][n]
                        if all_sent[m][r] != "IG":
                            list1[5] = all_sent[m][i+1]
                        else:
                            list1[5] = all_sent[m][5]
                        list1[9] = "IA"
                        list1[14] = all_sent[m][i]
                        all_sent[m][i] = None
                        if kind == 0:
                            dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,1)
                        else:
                            g += 1
                            list2.append(list1)
                            all_sent.append(list1)

                elif k == 4 and kind == 1 and all_sent[m][i] != None:
            #right now the only relation we have found that divides by making the object
            # the new subject is AS
                    genre = 1
                    if kind == 1:
                        bb = 8

                    str2 = findinlist(all_sent[m][i],pos,0,2)
                    if (all_sent[m][i] == 'AS'):
                        rule = "RDB"
                        a = 14
                        if all_sent[m][i] == 'AS':
                            anaphora.append(all_sent[m][5])
                    elif str2 == 'o':
                        rule = "RDC"
                        list3 = [None] * 70
                        a = 14
                        list3[8] = all_sent[m][8]
                        list3[3] = all_sent[m][10]
                        list3[5] = all_sent[m][5]
                        list3[9] = all_sent[m][i]
                        list3[10] = all_sent[m][16]
                        list3[14] = all_sent[m][18]
                        list2.append(list3)
                        all_sent.append(list3)
                        g += 1
                        genre = 2
                        str4 = build_sent(list3)
                        str4p = name_sent(str4)
                        list3[0] = str4
                        list3[42] = str4p
                    else:
                        rule = "RDA"
                        a = 5
                    if i == 15:
                        d = 16
                        c = 18
                    elif i == 19:
                        d = 20
                        c = 22
                    # elif i == 23:
                    #     a = 22
                    #     c = 26
                    # elif i == 31:
                    #     a = 30
                    #     c = 34
                    if all_sent[m][9] == "EX":
                        bb = 8
                    list1[8] = all_sent[m][8]
                    list1[3] = all_sent[m][3]
                    list1[5] = all_sent[m][a]
                    list1[9] = all_sent[m][i]
                    list1[10] = all_sent[m][d]
                    list1[14] = all_sent[m][c]
                    if genre == 1:
                        all_sent[m][i] = None
                        all_sent[m][c] = None
                        all_sent[m][d] = None
                        all_sent[m][8] = None
                    if genre == 1:
                        dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,1)
                    elif genre == 2:
                        dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,2,list3)
                        genre = 1
                elif k == 3 and kind == 1 and all_sent[m][i] != None:
                    rule = "DE " + all_sent[m][i]
                    dummy = rel_pro(i,m,all_sent,list1)
                    dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,1)
    if kind == 3:
        list2.append(all_sent[p+1])
        for i in range(0,g+1):
            del all_sent[-1]
        return list2
    else:
        return

def scope_uni(all_sent,m,i):

    comma = all_sent[m][39]
    univ = ['all','any','no','no'+us,'a','many'+up,'many'+uo]
    if comma == None:
        comma = 70
    if all_sent[m][59] != None and all_sent[m][3] in univ:
        if i < comma and i > 3:
            return False
    elif all_sent[m][60] != None and all_sent[m][10] in univ:
        if i < comma and i > 10:
            return False
    elif all_sent[m][61] != None and all_sent[m][17] in univ:
        if i < comma and i > 17:
            return False
    elif all_sent[m][62] != None and all_sent[m][21] in univ:
        if i < comma and i > 21:
            return False
    if i == 13 and all_sent[m][10] in univ:
        return False
    elif all_sent[m][i-1] in univ:
        return False
    return True

def rel_pro(i,m,all_sent,list1,new_var=""):

    subjrp = ['who','which','that'+us]
    objrp = ['who'+uo,'that'+uo,'which'+uo]
    comma = all_sent[m][39]

    if new_var != "":
        for i in range(59,63):
            if all_sent[m][i] != None:
                break

    if all_sent[m][i] in subjrp:
        srp = True
    else:
        srp = False
    if i == 59:
        if new_var == "":
            list1[5] = all_sent[m][5]
        else:
            list1[5] = new_var
        list1[8] = all_sent[m][8]
        list1[9] = all_sent[m][9]
        list1[14] = all_sent[m][14]
        all_sent[m][8] = all_sent[m][49]
        all_sent[m][9] = all_sent[m][15]
        all_sent[m][14] = all_sent[m][18]
        all_sent[m][60] = all_sent[m][61]
        all_sent[m][63] = all_sent[m][64]
        all_sent[m][15] = all_sent[m][19]
        all_sent[m][18] = all_sent[m][22]
        all_sent[m][59] = None
        all_sent[m][22] = None
        all_sent[m][64] = None
        all_sent[m][19] = None
        all_sent[m][22] = None
        all_sent[m][61] = None
    elif i == 60 and srp:
        if new_var == "":
            list1[5] = all_sent[m][14]
        else:
            list1[5] = new_var
        list1[8] = all_sent[m][49]
        list1[9] = all_sent[m][15]
        list1[14] = all_sent[m][18]
        list1[15] = all_sent[m][19]
        list1[18] = all_sent[m][22]
        all_sent[m][i] = None
        all_sent[m][49] = None
        all_sent[m][15] = None
        all_sent[m][18] = None
        all_sent[m][19] = None
        all_sent[m][22] = None
    elif i == 61:
        a = 22
        c = 26
    elif i == 62:
        a = 30
        c = 34




def new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,tot_sent,kind=1,list3=[]):

    if list1[8] == neg:
        list1[8] = "~"
    if kind == 2:
        sent2 = build_sent(list3)
        sent2p = name_sent(sent2)
    else:
        sent2 = build_sent(all_sent[m])
        sent2p = name_sent(sent2)
        all_sent[m][0] = sent2
        all_sent[m][42] = sent2p
    sent1 = build_sent(list1)
    sent1p = name_sent(sent1)
    list1[0] = sent1
    list1[42] = sent1p
    list2 = copy.deepcopy(list1)
    all_sent.append(list2)
    if kind == 1 or kind == 2:
        new_sent = "(" + sent1 + " & " + sent2 + ")"
        newp = "(" + sent1p + " & " + sent2p + ")"
        if kind == 1:
            conn = iff
        else:
            conn = conditional
    else:
        new_sent = sent1
        newp = sent1p
        conn = conditional
    dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,rule,"",conn)

def mult_defnd(list1,def_num):

    for i in range(len(list1)):
        num = list1[i][0]
        if def_num + '12' == num:
            return True
    return False

def indefiniendum(def_num,sent_num,multiple):

    if multiple:
        if sent_num[:-1] == def_num + '1':
            return True
        else:
            return False
    else:
        if sent_num == def_num + '1':
            return True
        else:
            return False

def in_dv(list1,dv_nam):

    if not list1[9] == "=":
        return False
    else:
        bool1 = check_dimension(dv_nam,1,list1[14])
        return bool1

def prop_type(paren_num,gparen_num,paren_conn,gparen_conn,sent_num,def_con):

    global sn
    str2 = None
    str1 = ""
    if paren_conn == '&' and gparen_conn == xorr:
    # formerly known as dc
        str1 = "d"
        str2 = str(sn) + paren_num
    elif paren_conn == xorr and gparen_conn == iff:
        str1 = 'd'
    if def_con == conditional:
        str1 = "cj"
    elif paren_conn == "&" and gparen_conn == conditional:
        if paren_num[-2] == "1":
            str1 = "an"
            str2 = str(sn) + paren_num
        else:
            str1 = "cn"
            str2 = str(sn) + paren_num
    elif paren_conn == conditional:
        if sent_num[-1] == "1":
            str1 = "an"
            str2 = str(sn) + paren_num
        else:
            str1 = "cn"
            str2 = str(sn) + paren_num
    elif paren_conn == iff:
        str1 = 'bic'
    elif paren_conn == "&" and gparen_conn == iff:
        str1 = 'bic'
    elif paren_conn == xorr and gparen_conn == conditional:
        easygui.msgbox("you have not coded for this sentence type yet")

    return str1

def add_sent(subj,relat,obj,conn_type):

    list1 = [None] * 70
    list1[5] = subj
    list1[9] = relat
    list1[14] = obj
    str1 = build_sent(list1)
    list1[0] = str1
    str1p = name_sent(str1)
    list1[42] = str1p
    list1[40] = False
    list1[53] = conn_type
    return list1

def add_sent2(all_sent,m,k,o,new_var2,words):

    global idf_var
    num2 = [11,1,2,47,3,4,5,66,67,35,48,59,6,8,9,48,12,10,13,14,36,60,63,49,15,16,17,18,\
           61,64,50,19,20,21,22,62,65,51,23,24,25,26,52,27,28,\
           29,30,31,32,33,34]
    num = [15,19,23,27,31]
    #list2 is the con sent, list1 is the ant sent
    comma = all_sent[m][39]
    list1 = []
    bool1 = False
    for j in num2:
        if j == k:
            bool1 = True
        if bool1:
            if comma != None:
                if j == comma + 1:
                    break
            if all_sent[m][j] != None and all_sent[m][j] != "":
                list1.append(all_sent[m][j])
                all_sent[m][j] = None
    list1 = categorize_words(words,list1,idf_var,all_sent,2,False)
    all_sent.append(list1)
    list3 = []
    list3 = division("",all_sent,words,3)
    list2 = []
    if o == 5:
        list2.append(new_var2)
    for i in num2:
        if all_sent[m][i] != None and all_sent[m][i] != "":
            list2.append(all_sent[m][i])
    if o != 5:
        list2 = categorize_words(words,list2,idf_var,all_sent,2,False,new_var2,o)
    else:
        list2 = categorize_words(words,list2,idf_var,all_sent,2,False)
    news = build_sent(list2)
    newp = name_sent(news)
    list2[0] = news
    list2[42] = newp
    all_sent[m] = list2
    all_sent.append(list2)
    for i in range(len(list3)):
        new_sent = build_sent(list3[i])
        newp = name_sent(new_sent)
        list3[i][0] = new_sent
        list3[i][42] = newp
        list3[i][53] = 'an'
        list3[i][40] = False
    return list3

def abb_change(list5, already_checked,all_sent,def_sent,i,match_dv,rename,j,def_con,\
               new_match = [],second=False):

    global never_used
    cap = False

    for t in range(len(all_sent)):
        no_match = False
        if t not in already_checked:
            for u in list5:
                if all_sent[t][u] == def_sent[i][u]:
                    pass
                elif u == 8 and (all_sent[t][53] == 'cn' or def_sent[i][53] == 'cn' or def_con == conditional):
                    cap = True
                    pass
                else:
                    if u == 9:
                        already_checked.append(t)
                    no_match = True
                    break

            if not no_match:
                match_dv.append([def_sent[i][j],all_sent[t][j]])
                #cap is for a denied consequent sentence
                str2 = "(" + def_sent[i][j] + mini_c + all_sent[t][j] + ")"
                rename.append(str2)
                if cap:
                    # if def_con == conditional:
                        # print "def_con"
                    strcn = "{" + all_sent[t][0] + "}"
                    str3 = build_sent(def_sent[i])
                    str3 = strcn + ", " + str3
                    rename.append(str3)
                if second:
                    for s in range(len(match_dv)):
                        if match_dv[s][0] == def_sent[i][j]:
                            never_used.append(match_dv[s][1])
                            break
                    new_match.append([def_sent[i][j],all_sent[t][j]])
                def_sent[i][j] = all_sent[t][j]
                return False
    return True

def abb_change2(match_dv,def_sent,i,idf_var,temp_match,j,gen_var,cnnan):

    match_dv.append([def_sent[i][j], idf_var[0]])
    temp_match.append([def_sent[i][j], idf_var[0]])
    def_sent[i][j] = idf_var[0]
    gen_var.append(idf_var[0])
    cnnan.append(idf_var[0])
    del idf_var[0]

def qadj(all_sent,m,j,new_var,kind=0):

    adj_var = all_sent[m][j-1]
    all_sent[m][j-1] = None
    list1 = [None] * 70
    list1[5] = new_var
    list1[9] = "IA"
    list1[14] = adj_var
    new_sent = build_sent(list1)
    newp = name_sent(new_sent)
    list1[0] = new_sent
    list1[42] = newp
    list1[55] = new_var
    all_sent.append(list1)
    if kind == 1:
        list1[53] = 'an'
        list1[40] = False
    return list1

def def_rn(defined,definition, definiendum,e, tot_sent,  dv_nam, idf_var, \
           words,all_sent,m,rep_rel,kind = "",k=0,circ = ""):
    # this function renames the variables in a definition

    global sn, plural_c, definite2,definite,anaphora,ind_var,gen_var
    b = time.time()
    #this is for those determinatives which have negations in their definitions where
    #the sentences has an R variable
    identical_det = ["only","anything_except","anyone_except","no","many"+um,"many"+un,\
        "no" + us]
    if definiendum == "mind":
        bb = 7
    new_idf = []
    if definiendum in identical_det:
        ident_det = True
    else:
        ident_det = False
    match_dv = []
    new_var = []
    str1 = copy.copy(definition)
    #if bool1 is false then there is a series of conjuncts that need to be removed from
    # the definition
    if kind == "" or kind == 'R':
        def_info = find_sentences(definition,True)
    else:
        def_info = find_sentences(definition)
    x = findposinlist(definiendum,rep_rel,0)
    if x > -1:
        rr_var = rep_rel[x][1]
    else:
        rr_var = 0
    def_loc = def_info[7]
    def_num = def_info[4][def_loc][0]
    ld = len(def_num)
    conn_type = all_sent[m][53]
    list1 = []
    cnnan = []
    list1 = id_def(def_info)
    dv = list1[0]
    adj_sent = []
    has_plural = list1[1]
    an_var = []
    odef = all_sent[m][0]
    con_var = all_sent[m][56]
    # we now must match the definite variables in the definition to the definite variables
    # already assigned
    list1 = []
    syn_det = ["no_one_except","any"+un]
    for i in range(len(dv)):
        temp_str = dv[i][1]
        for j in range(len(dv_nam)):
            dvn_temp = dv_nam[j][1]
            if dvn_temp == temp_str:
                telist7 = [dv[i][0],dv_nam[j][0]]
                if telist7 not in match_dv:
                    match_dv.append(telist7)
                    break
        else:
            if dv[i][0] not in idf_var:
                telist7 = [dv[i][0], idf_var[0]]
                match_dv.append(telist7)
                list1.append([idf_var[0], temp_str])
                new_var.append(idf_var[0])
                del idf_var[0]
            else:
                telist7 = [dv[i][0], temp_str]
                list1.append(telist7)
                match_dv.append([dv[i][0],dv[i][0]])
                idf_var.remove(dv[i][0])
    dv_nam += list1

    if kind == 'pronoun':

        str1 = findinlist(definiendum,dv_nam,1,0)
        if str1 == None:
            all_sent[m][k] = idf_var[0]
            match_dv.append(["c",idf_var[0]])
            dv_nam.append([idf_var[0],definiendum])
            new_var.append(idf_var[0])
            del idf_var[0]
        else:
            all_sent[m][k] = str1
            match_dv.append(["c'",str1])
        # when constructing definitions of personal pronouns or of determinatives the object of the IG relation
        # must be b and the subject must be z
    elif kind == 'determinative':
        all_sent[m][k] = ""
        if k == 10:
            j = 14
        else:
            j = k + 2
        ovar = all_sent[m][j]
        match_dv.append(["b",all_sent[m][j]])
        if definiendum == 'the' or definiendum == 'that'+ud:
            str1 = all_sent[m][j]
            str3 = findinlist(str1,dv_nam,0,1)
            str2 = findinlist(str3,definite,1,0)
            if str2 == None:
                match_dv.append(["z",idf_var[0]])
                definite.append([idf_var[0],str3])
                all_sent[m][j] = idf_var[0]
                new_var.append(idf_var[0])
                new_var2 = idf_var[0]
                del idf_var[0]
            else:
                all_sent[m][j] = str2
                match_dv.append(["z'",str2])
                new_var2 = str2
        elif definiendum not in syn_det:
            new_var2 = idf_var[0]
            all_sent[m][j] = idf_var[0]
            match_dv.append(["z",idf_var[0]])
            new_var.append(idf_var[0])
            del idf_var[0]
        if all_sent[m][45] == 1:
            str5 = findinlist(ovar,dv_nam,0,1)
            if str5 in tagged_nouns:
                tagged_nouns2.append([all_sent[m][j],str5])

    sdefinition = def_info[0][def_loc]
    def_sent = []
    rename = []
    skel_string = def_info[5]
    exception = []
    not_many = False
    first_in_def = [def_num+"1",def_num + "11"]
    temp_plural = []
    heir_num = []
    spec_var = ['y','x','w']
    rule_found = False
    univ = ['any','all','no','no' + us]
    idfq = ['a','many'+up,'many'+us,'many'+ud,'many'+uo,"a"+ua]
    sent_uniq1 = []
    cnn_type = None
    bool1 = False
    def_con = ""

    if definiendum in univ or definiendum in idfq:
        if definiendum in idfq or definiendum in univ:
            adj_var = None
            if all_sent[m][j-1] != None:
                list1 = qadj(all_sent,m,j,new_var2,0)
                sent_uniq1.append(list1)
        if definiendum in univ:

            if k == 3 and all_sent[m][59] != None:
                bool1 = True
                all_sent[m][59] = None
            elif k == 10 and all_sent[m][60] != None:
                bool1 = True
                all_sent[m][60] = None
            elif k == 16 and all_sent[m][61] != None:
                bool1 = True
                all_sent[m][61] = None
            elif k == 20 and all_sent[m][62] != None:
                bool1 = True
                all_sent[m][62] = None
            if bool1:
                list1 = add_sent2(all_sent,m,k,j,new_var2,words)
                for i in range(len(list1)):
                    sent_uniq1.append(list1[i])
            if all_sent[m][15] != None:
                new_relat = all_sent[m][9]
                new_obj = all_sent[m][14]
                all_sent[m][9] = all_sent[m][15]
                all_sent[m][15] = None
                all_sent[m][10] = all_sent[m][16]
                all_sent[m][16] = None
                all_sent[m][14] = all_sent[m][18]
                all_sent[m][18] = None
                # existence redundancy is done here
                if new_relat != 'EX':
                    list1 = add_sent(new_var2,new_relat,new_obj,"an")
                    sent_uniq1.append(list1)

    #as we loop through the sentences they must be in the definition which is the point of n
    for i in range(len(def_info[0])):
        if i == 4:
            bb = 8
        n = def_info[4][i][0][:ld]
        if kind == "determinative" or kind == "pronoun" or kind == 'AS':
            rule = "DE " + definiendum
            rule_found = True
        if def_info[4][i][1] == iff and not rule_found:
            rule = "DF " + definiendum
            rule_found = True
            def_con = iff
        elif def_info[4][i][1] == conditional and not rule_found:
            def_con = conditional
            rule = "NC " + definiendum
            rule_found = True
        if os(def_info[3][i]) == True and n == def_num:
            temp_str = space_words(def_info[3][i])
            temp_str = temp_str.replace("(","")
            temp_str = temp_str.replace(")","")
            telist7 = categorize_words(words,temp_str,idf_var,all_sent)
            bool1 = False
            bool2 = False


            if kind == 'AS' and telist7[9] == 'R':
                telist7[9] = anaphora[0]
                if i == 6:
                    telist7[5] = anaphora[1]
                str1 = build_sent(telist7)
            elif kind != "" and kind != 'R' and telist7[9] == "R":
                n = len(def_sent)
                exception.append(n)
                if telist7[3] != None:
                    temp_det = telist7[3]
                    has_detrm = True
                else:
                    has_detrm = False
                if ident_det:
                    neg1 = telist7[8]
                    if telist7[5] == 'b':
                        bool1 = True
                    if telist7[5] == 'z':
                        bool2 = True
                str2 = ''
                for p in range(2,70):
        # if the variable in the original definition is z,y,x,w then that must
        # go into the new definition in its proper place
                    if p == 14:
                        bb = 8
                    if telist7[p] in spec_var:
                        str2 = idf_var[0]
                        spec_var.remove(telist7[p])
                        match_dv.append([telist7[p],str2])
                        del idf_var[0]
                    if p == j and str2 != "" and str2 != None:
                        telist7[p] = str2
                    else:
                        telist7[p] = all_sent[m][p]
        #not many is the one negated determinative which is defined in this way and its
        #negation is removed in the definiens
                if definiendum == 'not' + ui + ' ' + 'many' + ud:
                    if all_sent[m][8] == 'not' + ui and def_info[4][i][0] not in first_in_def:
                        telist7[8] = ""
                        telist7[k] = 'exactly_one'
                        str2 = findinlist(ovar,plural_c,0,1)
                        telist7[j] = str2
                    if def_info[4][i][0] in first_in_def:
                        telist7[j] = ovar
                        if all_sent[m][47] == 'not' + ui:
                            telist7[47] = 'not' + ui
                        elif all_sent[m][8] == 'not' + ui:
                            telist7[8] = 'not' + ui
                            telist7[47] = None
                            not_many = True
        #just in case the list has a tagged noun
                telist7[45] = all_sent[m][45]
        # for the determinatives which have negations in their definition then we need
        # to do something special
                if ident_det:
                    if j == 5 or j == 14:
                        telist7[8] = neg1
                    elif j == 18 and neg1 == "~":
                        telist7[49] = neg1
                    elif j == 22 and neg1 == "~":
                        telist7[50] = neg1
                    elif j == 26 and neg1 == "~":
                        telist7[51] = neg1
        # the determinatives which have an identity statement in them behave differently
        # these are 'only' and 'anything except'
                if bool1:
                    telist7[j] = ovar
                if has_detrm:
                    telist7[k] = temp_det
                if definiendum == 'everything_except' + up and i == 13:
                    telist7[8] = "~"
                    if 'y' in idf_var:
                        telist7[j] = 'y'
                        match_dv.append(['y','y'])
                    else:
                        telist7[j] = idf_var[0]
                        match_dv.append(['y',idf_var[0]])
                        new_var.append(idf_var[0])
                        del idf_var[0]
                if (definiendum == 'all' and i == 4) or (definiendum == 'only' + up and i==9):
                    telist7[j] = 'd'
                    exception.remove(n)
                if definiendum == 'any' + un and i == 2:
                    telist7[8] = ""
                    telist7[10] = "no"
                if bool2:
                    str2 = findinlist("z",match_dv,0,1)
                    telist7[j] = str2
        # if the sentence is first then we must restor the definiendum to it
                if def_info[4][i][0] in first_in_def and not not_many:
                    telist7[k] = definiendum
        # what this does is it puts the original variable back into the definiendum
                    if kind != "pronoun":
                        telist7[j] = ovar
                str1 = build_sent(telist7)
                sdefinition = sdefinition.replace(def_info[3][i],str1)
            else:
                str1 = build_sent(telist7)
    # if a def sent is part of the defiendum then it does not have to be added to the all
    # all_sent list
            if def_info[4][i][0][:-1] in first_in_def or def_info[4][i][0] in first_in_def:
                telist7[40] = True
            else:
                telist7[40] = False

            telist7[0] = str1
            sent_num = def_info[4][i][0]
            paren_num = def_info[4][i][0][:-1]
            gparen_num = def_info[4][i][0][:-2]
            paren_conn = findinlist(paren_num,def_info[4],0,1)
            gparen_conn = findinlist(gparen_num,def_info[4],0,1)
            if paren_conn == None:
                bb = 7
    #we have to distinguish between the connective type of the parent string
    #and the connective type of the child string, the type of the parent string should
    #not change

            cnn_type = prop_type(paren_num,gparen_num,paren_conn,gparen_conn,sent_num,def_con)
            telist7[53] = cnn_type
            if conn_type == "ant" or conn_type == "con":
                telist7[53] = conn_type
            # if conn_type == "" or conn_type == None or conn_type == "cj":
            #     telist7[53] = cnn_type
            # else:
            #     telist7[53] = conn_type
            if cnn_type == 'an':
                list5 = []
                num2 = [5,14,22,29]
                for x in telist7[46]:
                    if x in num2:
                        list5.append(telist7[x])
                list6 = copy.deepcopy(list5)
                an_var.append([paren_num,list6])

            telist7[54] = str(sn) + "." + str(paren_num)
            telist7[44] = def_info[6][i][1]
            heir_num.append(def_info[4][i][0])
            for s in range(0,3):
                telist7.append(None)
            def_sent.append(telist7)

    # the purpose of this is that the subject of the definiendum must match the subject
    # of the osent to be defined.  if its a relation then the object must also match
    if (kind == "" or kind == "R" or kind == 'AS'):
        mvar = []
        bool1 = mult_defnd(def_info[4],def_num)
        for i in range(len(def_sent)):
            bool2 = indefiniendum(def_num,heir_num[i],bool1)
            if bool2:
                if heir_num[i] in first_in_def:
                    match_dv.append([def_sent[i][5],all_sent[m][5]])
                    if kind == "R" or kind == 'AS':
                        match_dv.append([def_sent[i][14],all_sent[m][14]])
                        if bool1:
                            break
                else:
                    relat = def_sent[i][9]
                    oobj = def_sent[i][14]
                    nobj = findinlist(oobj,match_dv,0,1)
                    for j in range(len(all_sent)):
                        if all_sent[j][9] == relat and all_sent[j][14] == nobj:
                            match_dv.append([def_sent[i][5],all_sent[j][5]])
                            break


    #if the definiendum is many-o then its object variable needs to be matched
    if definiendum == 'many' + uo:
        match_dv.append(['c',all_sent[m][14]])

    num = [5,14]
    num2 = [5,14,18,22,26,30,34]
    num3 = [9,14,8]
    num4 = [9,5,8]
    # the point of the exception list is that we do not change certain sentences in the
    # definiens if we are analyzing a pronoun or determinative
    str2 = ""
    already_checked2 = []
    cap = False
    unmatched = []
    temp_match = []

    if definiendum == 'doglike':
        bb = 8

    for i in range(len(def_sent)):
        if i not in exception:
            for j in num2:
                temp_str = def_sent[i][j]
                isvar = isvariable(temp_str)
                if isvar:
                    if j > 17:
                        bb = 8
                    if temp_str != None:
                        str3 = findinlist(temp_str,match_dv,0,1)
                        if str3 != None and temp_str != str3:
                            already_checked2.append([i,j])
                            def_sent[i][j] = str3
                            str2 = '(' + temp_str + mini_c + str3 + ')'
                            if str2 not in rename and str2 != "":
                                rename.append(str2)
                            str2 = ""
                        elif temp_str == str3:
                            already_checked2.append([i,j])
                            pass
                        elif def_sent[i][j] == rr_var:
                            dummy = abb_change2(match_dv,def_sent,i,idf_var,temp_match,j,gen_var,cnnan)
                        else:
                            # here we check to see if it has a plural form
                            if j == 14 and def_sent[i][9] == 'OF':
                                g = findposinlist(def_sent[i][5],plural_c,1)
                                if g > -1:
                                    match_dv.append([def_sent[i][j],plural_c[g][0]])
                                    def_sent[i][j] = plural_c[g][0]
                            else:
                                already_checked = []
                                if j == 5:
                                    list5 = num3
                                else:
                                    list5 = num4
                                no_match = abb_change(list5, already_checked,all_sent,\
                                    def_sent,i,match_dv,rename,j,def_con)
                                if not no_match and j == 14 and unmatched != []:
                                    dummy = abb_change(num3, already_checked,all_sent,\
                                        def_sent,i,match_dv,rename,j,def_con)
                                elif no_match:
                                    unmatched.append([i,j])
    if unmatched != []:
        new_match = []
        unmatched2 = []
        already_checked = []
        for k in range(len(unmatched)):
            i = unmatched[k][0]
            j = unmatched[k][1]
            if j == 5:
                no_match = abb_change(num3, already_checked,all_sent,\
                    def_sent,i,match_dv,rename,5,def_con,new_match,True)
                if no_match:
                    temp_str = def_sent[i][j]
                    str3 = findinlist(temp_str,match_dv,0,1)
                    if str3 != None and temp_str != str3:
                        def_sent[i][j] = str3
                        str2 = '(' + temp_str + mini_c + str3 + ')'
                        if str2 not in rename and str2 != "":
                            rename.append(str2)
                        str2 = ""
                    elif temp_str == str3:
                        pass
                    else:
                        dummy = abb_change2(match_dv,def_sent,i,idf_var,temp_match,j,gen_var,cnnan)
                        unmatched2.append([i,j])
            else:
                temp_str = def_sent[i][j]
                str3 = findinlist(temp_str,match_dv,0,1)
                if str3 != None and temp_str != str3:
                    def_sent[i][j] = str3
                    str2 = '(' + temp_str + mini_c + str3 + ')'
                    if str2 not in rename and str2 != "":
                        rename.append(str2)
                    str2 = ""
                else:
                    dummy = abb_change2(match_dv,def_sent,i,idf_var,temp_match,j,gen_var,cnnan)
                    unmatched2.append([i,j])

        if unmatched2 != [] and new_match != []:
            for k in range(len(unmatched2)):
                i = unmatched[k][0]
                j = unmatched[k][1]
                temp_str = def_sent[i][j]
                str3 = findinlist(temp_str,temp_match,1,0)
                if str3 != None:
                    str4 = findinlist(str3,new_match,0,1)
                    if str4 != None:
                        str2 = '(' + str3 + mini_c + str4 + ')'
                        if str2 not in rename and str2 != "":
                            rename.append(str2)
                            str2 = ""
                    if def_sent[i][j] in gen_var:
                        gen_var.remove(def_sent[i][j])
                        cnnan.remove(def_sent[i][j])
                        never_used.append(def_sent[i][j])
                    def_sent[i][j] = str4

    # definition = sdefinition
    # we now replace the skel string with the new sentences, to get the true definition
    skel_string = def_info[5]
    skel_wid = def_info[8]
    skel_string2 = def_info[5]
    first_sent_found = False

    for i in range(len(def_sent)):
        str2 = build_sent(def_sent[i])
        def_sent[i][0] = str2
    #because the definiendum for universal quantifiers is somewhat hard to get we
    # we just use the original sentence to be defined
    # also right now we are removing the negative sign from the o sent, though
    # in the future things might be more complicated than this
        if def_sent[i][40] and kind != "AS" and not first_sent_found:
            str2 = odef
            first_sent_found = True
            if kind != 'determinative' and kind != 'pronoun':
                str2 = str2.replace("~","")
        elif kind == "AS":
            str2 = def_sent[i][0]
        skel_string = skel_string.replace(def_sent[i][44], str2)
        str1 = name_sent(str2)
        def_sent[i][42] = str1
        skel_string2 = skel_string2.replace(def_sent[i][44], str1)

    if sent_uniq1 != []:
        str4 = ""
        str4p = ""
        for i in range(len(sent_uniq1)):
            if str4 == "":
                str4 = sent_uniq1[i][0]
                str4p = sent_uniq1[i][42]
            else:
                str4 += " & " + sent_uniq1[i][0]
                str4p += " & " + sent_uniq1[i][42]
            def_sent.append(sent_uniq1[i])
        str4 += " & "
        str4p += " & "
        g = skel_string.find(iff)
        h = skel_string.find(conditional)
        if h == -1:
            skel_string = skel_string[:g+4] + str4 + skel_string[g+4:]
        else:
            skel_string = skel_string[:g+4] + "(" + str4 + skel_string[g+4:h-1] + ") " + skel_string[h:]

        g = skel_string2.find(iff)
        h = skel_string2.find(conditional)
        if h == -1:
            skel_string2 = skel_string2[:g+4] + str4p + skel_string2[g+4:]
        else:
            skel_string2 = skel_string2[:g+4] + "(" + str4p + skel_string2[g+4:h-1] + ")" + skel_string2[h-1:]

    str3 = skel_string2
    if kind == "" or kind == 'R':
        sn += 1
        definition = remove_outer_paren(definition)
        if rename != []:
            d = findposinlist(definiendum,defined,0)
            if d == -1:
                tot_sent.append([sn, definition, "","", rule])
                defined.append([definiendum,sn])
                anc1 = sn
            else:
                anc1 = defined[d][1]
            str2 = build_sent_list(rename)
            if rule[:2] != "DE":
                rule = "SUB"
            else:
                rule += " " + definiendum
            if str2 != None:
                sn += 1
                tot_sent.append([sn, str2, "","", 'RN'])
            bool1 = check_dimension(tot_sent,1,skel_string)
            if not bool1:
                sn += 1
                tot_sent.append([sn, skel_string, str3,"", rule,anc1,sn-1])
        else:
            print "no variable change in definition"
            d = findposinlist(definiendum,defined,0)
            if d == -1:
                tot_sent.append([sn, definition, skel_string2,"", rule])
                defined.append([definiendum,sn])
            else:
                defined.append([definiendum,sn])

    else:
        bool1 = check_dimension(tot_sent,1,skel_string)
        if not bool1:
            sn += 1
            tot_sent.append([sn, skel_string, str3,"", rule])
    # right now we're simply deleting the first sentence of the def_sent since
    # we have already defined that but in the future when we work with more
    # complex definienda we will have to change this
    #end3
    if an_var != []:
        for i in range(len(an_var)):
            for y in range(len(an_var[i][1])):
                for j in range(len(match_dv)):
                    if an_var[i][1][y] == match_dv[j][0]:
                        an_var[i][1][y] = match_dv[j][1]
                        break

    if definiendum == 'any':
        bb = 8
    list1 = []
    num = [5,14,15,18,26,30]
    def_var = findinlist("definite",dv_nam,1,0)
    indef = findinlist("indefinite",dv_nam,1,0)
    gen = findinlist("general",dv_nam,1,0)

    for i in range(len(def_sent)):
        if def_sent[i][9] == 'IA':
            if def_sent[i][14] == def_var:
                if def_sent[i][5] not in definite2:
                    definite2.append(def_sent[i][5])
            elif def_sent[i][14] == indef:
                if def_sent[i][5] not in ind_var:
                    ind_var.append(def_sent[i][5])
            elif def_sent[i][14] == gen:
                if def_sent[i][5] not in gen_var:
                    gen_var.append(def_sent[i][5])
                    cnnan.append(def_sent[i][5])

        # the indefinite articles have special negations
        if def_sent[i][9] == "S":
            bb = 8
        for n in num:
            if def_sent[i][n] in new_var:
                new_var.remove(def_sent[i][n])
        if definiendum == "no"+us or definiendum == "no" or definiendum == "many" + un:
            if not def_sent[i][40] and def_sent[i][8] == "~":
                def_sent[i] = not_a(def_sent[i],k,tot_sent)

        if def_sent[i][53] == 'cn' or def_sent[i][53] == 'con':
            for j in range(len(an_var)):
                if def_sent[i][54] == an_var[j][0]:
                    def_sent[i][56] = an_var[j][1]
        elif (def_sent[i][53] == "an" or def_sent[i][53] == "ant") \
                and not def_sent[i][40]:
            list1.append(def_sent[i])
        if con_var != None:
            def_sent[i][56] = con_var
        bool1 = check_dimension(all_sent,0,def_sent[i][0])

        bool2 = in_dv(def_sent[i],dv_nam)
        # it used to be that the sentence had to not have a plural which means
        # that 41 had to be false
        if bool1 and def_sent[i][53] == 'cn':
            for u in range(len(all_sent)):
                if all_sent[u][0] == def_sent[i][0]:
                    all_sent[u][53] = 'cn'
                    break
        elif bool1 == False and bool2 == False and def_sent[i][40] == False:
            def_sent[i][43] = circ
            if cnnan != []:
                bb = 8
            def_sent[i][55] = cnnan
            all_sent.append(def_sent[i])
    if len(list1) > 1:
        str1 = ant_var(list1)
        if len(list1) > 1:
            list1[0][58] = str1
        else:
            list1[58] = str1
        ant_cond.append(list1)
    elif len(list1) == 1:
        list1[0][58] = list1[0][5]
        ant_cond.append(list1)

    c = time.time()
    d = c - b
    return

def ant_var(list1):

    list2 = []
    num = [5,14,18,22,26,30,34]
    num2 = [5,14]
    for i in range(len(list1)):
        for j in num2:
            if list1[i][j] != None:
                list2.append(list1[i][j])

    for i in range(len(list2)):
        g = list2.count(list2[i])
        if g > 1:
            return list2[i]

    easygui.msgbox('your method for finding the antecedent variable is not working')

def not_a(list1,k,tot_sent):

    num = [3,10,16,20,24]
    list2 = copy.deepcopy(list1)
    for i in num:
        if i > k and (list1[i] == "a" or list1[i] == "a" + ua):
            if list1[i] == "a":
                rule = "DE not a"
            else:
                rule = "DE not a" + ua
            list2[i] = 'no'+us
            list2[8] = None
            str1 = build_sent(list2)
            str1p = name_sent(str1)
            list2[0] = str1
            list2[42] = str1p
            dummy = new_sentence2(list1[0],list1[42],str1,str1p,tot_sent,"DE not a")
    return list2

def plurals(tot_sent, all_sent, words, dv_nam, idf_var):

    global plural_c
    global sn
    c = time.time()
    all_sent = remove_duplicates(all_sent,0)
    pnouns = words[26]
    bool1 = False
    for m in range(len(dv_nam)):
        pluralf = dv_nam[m][1]
        g = findposinlist(pluralf,pnouns,0)
        if g > -1:
            if not bool1:
                new_abb = idf_var[0]
                del idf_var[0]
                dv_nam.append([new_abb,"plural form"])
                bool1 = True
            str1 = dv_nam[m][0]
            singular = findinlist(pluralf,pnouns,0,1)
            singa = findinlist(singular,dv_nam,1,0)
            if singa == None:
                singa = idf_var[0]
                del idf_var[0]
                dv_nam.append([singa,singular])
            defndm = "(" + str1 + "=" + pluralf + ")"
            sent1 = "(" + str1 + "OF" + singa + ")"
            list1 = [None] * 70
            list1[0] = sent1
            list1[5] = str1
            list1[9] = "OF"
            list1[14] = singa
            plural_c.append([singa,str1])
            sent2 = "(" + str1 + "IG" + new_abb + ")"
            defp = name_sent(defndm)
            sent1p = name_sent(sent1)
            list1[42] = sent1p
            all_sent.append(list1)
            sent2p = name_sent(sent2)
            tot = defndm + " " + iff + " (" + sent1 + " & " + sent2 + ")"
            totp = defp + " " + iff + " (" + sent1p + " & " + sent2p + ")"
            sn += 1
            tot_sent.append([sn,tot,totp,"","DE " + pluralf,"","",""])
    b = time.time()
    d = b - c
    return

def categorize_words(words,str2,idf_var,all_sent,kind=1,first=False,snoun="",snum=""):

    global sn
    global anaphora
    global never_used
    has_plural = False
    bool1 = False

    if kind == 0:
        list1 = str2
        osent = str2[-3]
        prp = str2[-2]
        g = len(list1) - 3
        sent_type = str2[-1]
    elif kind == 1:
        osent = copy.copy(str2)
        list1 = str2.split(' ')
        g = len(list1)
        prp = None
        sent_type = ''
    elif kind == 2:
        list1 = str2
        g = len(list1)
        prp = None
        sent_type = ''
        osent = None

    list1_cat = [None] * 70
    relation_type = 0
    list2 = []
    list3 = []
    posp = words[28]
    doubles = words[31]
    has_comma = ""

    i = -1
    while i < g -1:
        i += 1
        if "," in list1[i]:
            list1[i] = list1[i].replace(",","")
            has_comma = list1[i]
        word = list1[i]
        if word == 'rewarded':
            bb = 8
        bool4 = check_dimension(doubles,0,word)
        bool3 = False
        if bool4:
            str4 = word + " " + list1[i+1]
            if "," in str4:
                str4 = str4.replace(",","")
                has_comma = str4
            bool3 = check_dimension(doubles,1,str4)
            if bool3:
                word = str4
                if has_comma != "":
                    has_comma = word
        if word == 'it':
            #this means that the subject of the previous sentences obtains the anaphor
            #to which it refers
            all_sent[len(all_sent)-2][57] = 3
        if isvariable(word):
            pos = 'n'
            if word in idf_var:
                idf_var.remove(word)
                never_used.append(word)
            if kind == 2:
                str1 = findinlist(word,dv_nam,0,1)
                str3 = findinlist(str1,posp,0,1)
                if str3 == "a":
                    pos = 'a'
        elif word == "~":
            pos = 'm'
        elif word == ne:
            pos = 'r'
        elif word == 'not' + ui:
            pos = 'm'
            word = neg
        elif isinstance(word,int):
            pos = 'n'
        else:
            pos = findinlist(word,posp,0,1)
        if word == 'plural form':
            has_plural = True

        #determined nouns occupy the noun position
        if pos == 'w':
            pos = 'n'

        if word == 'cold':
            bb = 7
        if word == ' ':
            pass
        elif (pos == 'd' or pos == 'q') and relation_type == 0:
            list1_cat[3] = word
            list2.append(3)
        elif pos == 'a' and relation_type == 0:
            list1_cat[4] = word
            list2.append(4)
        elif pos == 'm' and list1_cat[3] == None and list1_cat[5] == None and relation_type == 0:
            list1_cat[47] = word
            list2.append(47)
        elif (pos == 'n' or pos == 'p') and relation_type == 0 and list1_cat[5] == None:
            list1_cat[5] = word
            list2.append(5)
        elif (pos == 'n' or pos == 'p') and relation_type == 0 and list1_cat[5] == None:
            list1_cat[5] = word
            list2.append(5)
        elif pos == 'c' and relation_type == 0 and list1_cat[5] != None:
            list1_cat[66] = word
            list2.append(66)
        elif (pos == 'n' or pos == 'p') and relation_type == 0 and list1_cat[66] != None:
            list1_cat[67] = word
            list2.append(67)
        elif pos == 'n' and relation_type == 0 and list1_cat[5] != None:
            list1_cat[35] = word
            list2.append(35)
        elif pos == 'u' and relation_type == 0 and list1_cat[5] != None:
            list1_cat[59] = word
            list2.append(59)
        elif pos == 'b' and relation_type == 0:
            list1_cat[7] = word
            list2.append(7)
        elif pos == 'm' and relation_type == 0:
            list1_cat[8] = word
            list2.append(8)
        elif pos == 'r' and relation_type == 0:
            list1_cat[9] = word
            list2.append(9)
            relation_type = 1
            if list1_cat[5] == None and list1_cat[4] != None:
                list1_cat[5] = list1_cat[4]
                list1_cat[4] = None
                list2.remove(4)
                list2.append(5)
            if snoun != "" and snum == 14:
                list1_cat[14] = snoun
        elif (pos == 'd' or pos == 'q') and relation_type == 1:
            list1_cat[10] = word
            list2.append(10)
        # this line of code must be first because if the word is an adjective
        # and the relation is IA then it must go in slot 14
        elif pos == 'm' and relation_type == 1 and list1_cat[14] == None and \
            list1_cat[60] == None:
            list1_cat[12] = word
            list2.append(12)
        elif pos == 'a' and relation_type == 1 and list1_cat[9] == 'IA':
            list1_cat[14] = word
            list2.append(14)
        elif pos == 'a' and relation_type == 1:
            list1_cat[13] = word
            list2.append(13)
        elif (pos == 'n' or pos == 'p') and relation_type == 1 and list1_cat[14] == None:
            list1_cat[14] = word
            list2.append(14)
        elif pos == 'n' and relation_type == 1 and list1_cat[14] != None:
            list1_cat[36] = word
            list2.append(36)
        elif pos == 'e' and relation_type == 1:
            list1_cat[48] = word
            list2.append(48)
        elif pos == 'u' and relation_type == 1 and list1_cat[14] != None:
            list1_cat[60] = word
            list2.append(60)
        elif pos == 'n' and relation_type == 1 and list1_cat[60] != None:
            list1_cat[63] = word
            list2.append(63)
        elif pos == 'm' and relation_type == 1:
            list1_cat[49] = word
            list2.append(49)
        elif pos == 'r' and relation_type == 1:
            list1_cat[15] = word
            relation_type = 2
            list2.append(15)
            if snoun != "" and snum == 18:
                list1_cat[18] = snoun
        elif (pos == 'd' or pos == 'q') and relation_type == 2:
            list1_cat[16] = word
            list2.append(16)
        elif pos == 'a' and relation_type == 2 and list1_cat[15] == 'IA':
            list1_cat[18] = word
            relation_type = 2
            list2.append(18)
        elif pos == 'a' and relation_type == 2:
            list1_cat[17] = word
            list2.append(17)
        elif (pos == 'n' or pos == 'p') and relation_type == 2 and list1_cat[18] == None:
            list1_cat[18] = word
            list2.append(18)
        elif pos == 'u' and relation_type == 2 and list1_cat[18] != None:
            list1_cat[61] = word
            list2.append(61)
        elif pos == 'n' and relation_type == 2 and list1_cat[60] != None:
            list1_cat[64] = word
            list2.append(64)
        elif pos == 'm' and relation_type == 2:
            list1_cat[50] = word
            list2.append(50)
        elif pos == 'r' and relation_type == 2:
            relation_type = 3
            list1_cat[19] = word
            list2.append(19)
            if snoun != "" and snum == 22:
                list1_cat[22] = snoun
        elif (pos == 'd' or pos == 'q') and relation_type == 3:
            list1_cat[20] = word
            list2.append(20)
        elif pos == 'a' and relation_type == 3 and list1_cat[19] == 'IA':
            list1_cat[22] = word
            relation_type = 3
            list2.append(22)
        elif pos == 'a' and relation_type == 3:
            list1_cat[21] = word
            list2.append(21)
        elif (pos == 'n' or pos == 'p') and relation_type == 3 and list1_cat[22] == None:
            list1_cat[22] = word
            list2.append(22)
        elif pos == 'u' and relation_type == 3 and list1_cat[22] != None:
            list1_cat[62] = word
            list2.append(62)
        elif pos == 'n' and relation_type == 3 and list1_cat[60] != None:
            list1_cat[65] = word
            list2.append(65)
        elif pos == 'm' and relation_type == 3:
            list1_cat[51] = word
            list2.append(51)
        elif pos == 'r' and relation_type == 3:
            relation_type = 4
            list1_cat[23] = word
            list2.append(23)
            if snoun != "" and snum == 26:
                list1_cat[26] = snoun
        elif (pos == 'd' or pos == 'q') and relation_type == 4:
            list1_cat[24] = word
            list2.append(24)
        elif pos == 'a' and relation_type == 4 and list1_cat[23] == 'IA':
            list1_cat[26] = word
            relation_type = 4
            list2.append(26)
        elif pos == 'a' and relation_type == 4:
            list1_cat[25] = word
            list2.append(25)
        elif (pos == 'n' or pos == 'p') and relation_type == 4:
            list1_cat[26] = word
            list2.append(26)
        elif pos == 'm' and relation_type == 4:
            list1_cat[52] = word
            list2.append(52)
        elif pos == 'r' and relation_type == 4:
            relation_type = 5
            list1_cat[27] = word
            list2.append(27)
            if snoun != "" and snum == 30:
                list1_cat[30] = snoun
        elif (pos == 'd' or pos == 'q') and relation_type == 5:
            list1_cat[28] = word
            list2.append(28)
        elif pos == 'a' and relation_type == 5 and list1_cat[27] == 'IA':
            list1_cat[29] = word
            relation_type = 5
            list2.append(29)
        elif pos == 'a' and relation_type == 5:
            list1_cat[29] = word
            list2.append(29)
        elif (pos == 'n' or pos == 'p') and relation_type == 5:
            list1_cat[30] = word
            list2.append(30)
        elif pos == 'r' and relation_type == 5:
            relation_type = 6
            list1_cat[31] = word
            list2.append(31)
            if snoun != "" and snum == 34:
                list1_cat[34] = snoun
        elif (pos == 'd' or pos == 'q') and relation_type == 6:
            list1_cat[32] = word
            list2.append(32)
        elif pos == 'a' and relation_type == 6:
            list1_cat[33] = word
            list2.append(33)
        elif (pos == 'n' or pos == 'p') and relation_type == 6:
            list1_cat[34] = word
            list2.append(34)
        elif pos == 'b':
            list1_cat[7] = word
        elif pos == 'm':
            if relation_type == None:
                list1_cat[8] = word
                list2.append(8)
            elif relation_type == 'r':
                list1_cat[13] = word
                list2.append(13)
        else:
            easygui.msgbox('you did not categorize the word ' + word)
        if word in anaphoric_relations and first:
                anaphora = []
                anaphora.append(list1_cat[9])
        if has_comma != "":
            for j in range(0,69):
                if list1_cat[j] == has_comma:
                    list1_cat[39] = j
                    has_comma = ""
                    break
        if bool3:
            del list1[i+1]
            g -= 1

    list2.sort()
    list1_cat[46] = list2
    list1_cat[42] = prp
    list1_cat[0] = osent
    list1_cat[41] = has_plural
    list1_cat[53] = sent_type
    return list1_cat

def build_sent_name(prop_name):
    str1 = ''
    str2 = ''
    list1 = []

    for i in range(len(prop_name)):
        if i == 24:
            bb = 8
        str3 = remove_outer_paren(prop_name[i][2])
        str3 = str3.replace("~","")
        str1 = '(' + prop_name[i][0] + mini_e + str3 + ')'
        if len(str2) == 0 and len(str1) > 57:
            list1.append(str1)
        elif (len(str2) + len(str1)) > 57:
            list1.append(str2)
            str2 = str1
            if i + 1 == len(prop_name):
                list1.append(str2)
        elif (len(str2) + len(str1)) <= 57:
            if len(str2) == 0:
                str2 = str1
            else:
                str2 = str2 + ' & ' + str1
            if i + 1 == len(prop_name):
                list1.append(str2)
    return list1



def syn(tot_sent, all_sent, words):

    global sn
    doubles = words[31]
    bool1 = False
    synon = words[14]
    syn_pairs = words[13]
    m = -1
    while m < len(all_sent) -1:
        m += 1
        old_sent = all_sent[m][-3]
        oldp = all_sent[m][-2]
        sent_type = all_sent[m][-1]
        anc1 = ""
        anc2 = ""
        anc3 = ""
        anc4 = ""
        i = -1
        while i < len(all_sent[m])-4:
            i += 1
            str1 = all_sent[m][i]
            bool2 = check_dimension(doubles,0,str1)
            bool3 = False
            if bool2:
                str4 = all_sent[m][i] + " " + all_sent[m][i+1]
                bool3 = check_dimension(doubles,1,str4)
                if bool3:
                    str1 += " " + all_sent[m][i+1]
            if i == 9:
                bb = 7
            if str1 in synon:
                for j in range(len(syn_pairs)):
                    if str1 == syn_pairs[j][0]:
                        bool1 = True
                        rule = 'DE ' + str1
                        str5 = syn_pairs[j][2]
                        if " " in syn_pairs[j][1]:
                            str6 = syn_pairs[j][1]
                            g = str6.find(" ")
                            str3 = str6[:g]
                            str4 = str6[g+1:]
                            list2 = copy.deepcopy(all_sent[m])
                            list2[i] = str4
                            list2.insert(i,str3)
                            all_sent[m] = list2
                        else:
                            all_sent[m][i] = syn_pairs[j][1]
                        u = findinlist(str5,tot_sent,1,0,True)
                        if u == -1:
                            bool1 = True
                            str5v = name_sent(syn_pairs[j][2])
                            s = findinlist(str5,tot_sent,1,0)
                            sn += 1
                            if s == None:
                                s = sn
                            if anc1 == "":
                                anc1 = s
                            elif anc2 == "" and s != anc1:
                                w = copy.copy(s)
                                anc2 = w
                            elif anc3 == "" and s != anc1 and s != anc2:
                                u = copy.copy(s)
                                anc3 = u
                            tot_sent.append([sn, str5, "","", rule])
                        else:
                            s = tot_sent[u][0]
                            str5v = name_sent(syn_pairs[j][2])
                            if anc1 == "":
                                t = copy.copy(s)
                                anc1 = t
                            elif anc2 == "" and s != anc1:
                                w = copy.copy(s)
                                anc2 = w
                            elif anc3 == "" and s != anc1 and s != anc2:
                                u = copy.copy(s)
                                anc3 = u
                        if bool3:
                            del all_sent[m][i+1]
                        break


        if bool1:
            new_sent = build_sent2(all_sent[m],True)
            newp = name_sent(new_sent)
            all_sent[m][-3] = new_sent
            all_sent[m][-2] = newp
            all_sent[m][-1] = sent_type
            dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"SUB",anc1,iff,anc2,anc3)
            bool1 = False
    return

def print_sent_full(test_sent,p,tot_prop_name):
    global result_data
    global excel
    # p = 30
    b = time.time()
    p += 2
    for i in range(len(test_sent)):
        # if i == 2:
        #     break
        for j in range(len(test_sent[i])):
            if len(test_sent[i][j]) == 7:
                test_sent[i][j].append("")
            elif len(test_sent[i][j]) == 6:
                test_sent[i][j].append("")
                test_sent[i][j].append("")
            elif len(test_sent[i][j]) == 5:
                test_sent[i][j].append("")
                test_sent[i][j].append("")
                test_sent[i][j].append("")
            elif len(test_sent[i][j]) == 4:
                bb = 7
            if test_sent[i][j][7] != "" and test_sent[i][j][7] != None:
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5]) + ',' +\
                       str(test_sent[i][j][6]) + ',' + str(test_sent[i][j][7])
            elif test_sent[i][j][6] != "" and test_sent[i][j][6] != None:
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5]) + ',' + str(test_sent[i][j][6])
            elif test_sent[i][j][5] != None and test_sent[i][j][5] != "":
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5])
            elif test_sent[i][j][4] != None and test_sent[i][j][4] != "":
                str1 = test_sent[i][j][4]
            else:
                str1 = ""
            if j == 0:
                str1 = i
            if excel:
                w4.cell(row=p,column=2).value = test_sent[i][j][0]
                w4.cell(row=p,column=3).value = test_sent[i][j][1]
                w4.cell(row=p,column=4).value = str1
            else:
                result_data['text_'+str(p-1)+'_1']=test_sent[i][j][0]
                result_data['text_'+str(p-1)+'_2']=test_sent[i][j][1]
                result_data['text_'+str(p-1)+'_3']= str1

            p += 1
        p += 1

        list1 = build_sent_name(tot_prop_name[i])
        for j in range(len(list1)):
            if excel:
                w4.cell(row=p,column=3).value = list1[j]
            else:
                result_data['text_'+str(p)+'_2']=list1[j]
            p += 1
        p += 1

        # for j in range(len(test_sent[i])):
        #     # if j == 2:
        #     #     break
        #     if test_sent[i][j][2] != "":
        #         if test_sent[i][j][3] == "~" and not os(test_sent[i][j][2]):
        #             str1 = "(" + test_sent[i][j][2] + ")"
        #         else:
        #             str1 = test_sent[i][j][2]
        #         w4.cell(row=p,column=2).value = test_sent[i][j][0]
        #         w4.cell(row=p,column=3).value = test_sent[i][j][3] + str1
        #         p += 1
        # p += 1
        # w4.cell(row=p,column=3).value = 'irrelevant conjuncts:'
        # str3 = build_sent_list(irrel_conj)
        # w4.cell(row=p,column=3).value = str3
        # p += 1

        bool1 = False
        if tot_prop_sent != []:
            for j in range(len(tot_prop_sent[i])):
                if j == 8:
                    bb = 7
                if not bool1 and tot_prop_sent[i][j][4] != "":
                    if excel:
                        w4.cell(row=p,column=3).value = "____________________"
                    else:
                        result_data['text_'+str(p)+'_2']="____________________"

                    bool1 = True
                    p += 1
                if excel:
                    w4.cell(row=p,column=2).value = tot_prop_sent[i][j][0]
                    w4.cell(row=p,column=3).value = tot_prop_sent[i][j][2] + tot_prop_sent[i][j][1]
                else:
                    result_data['text_'+str(p)+'_1']=tot_prop_sent[i][j][0]
                    result_data['text_'+str(p)+'_2']=tot_prop_sent[i][j][2] + tot_prop_sent[i][j][1]

                str2 = ""
                if len(tot_prop_sent[i][j]) == 5:
                    str2 = tot_prop_sent[i][j][3] + " " + str(tot_prop_sent[i][j][4])
                elif len(tot_prop_sent[i][j]) == 6:
                    str2 = tot_prop_sent[i][j][3] + " " + str(tot_prop_sent[i][j][4]) + \
                    "," + str(tot_prop_sent[i][j][5])
                else:
                    if tot_prop_sent[i][j][4] != "":
                        str2 = tot_prop_sent[i][j][3] + " " + str(tot_prop_sent[i][j][4])
                    if tot_prop_sent[i][j][5] != "" and tot_prop_sent[i][j][5] != None:
                        str2 += "," + str(tot_prop_sent[i][j][5])
                    if tot_prop_sent[i][j][6] != "" and tot_prop_sent[i][j][6] != None:
                        str2 += "," + str(tot_prop_sent[i][j][6])
                    if len(tot_prop_sent[i][j]) > 7:
                        if tot_prop_sent[i][j][7] != "" and tot_prop_sent[i][j][7] != None:
                            str2 += "," + str(tot_prop_sent[i][j][7])



                if excel:
                    w4.cell(row=p,column=4).value = str2
                else:
                    result_data['text_'+str(p)+'_3']= str2
                p += 1
        p += 3
    c = time.time()
    g = c - b
    return



def build_dict(str1):

    global excel
    list1 = []
    list4 = []
    word_type = []
    detm = []
    relat = []
    srelat = []
    trelat = []
    atomic_relations = []
    det = []
    adj = []
    adv = []
    noun = []
    cor = []
    lcon = []
    subo = []
    synon = []
    redundant = []
    aux = []
    atomic_relata = []
    negg = []
    dnoun = []
    det_pairs = []
    syn_pairs = []
    particles = []
    relations = []
    relations2 = []
    definitions = []
    really_atomic = []
    pronouns = []
    poss_pronouns = []
    doubles = []
    triples = []
    plurals = []
    neg_det = []
    pos = []
    category = ['r','s','t']
    almost_done = False
    i = 0
    for row in ws:
        i += 1
        if excel:
            s = row[0].value
            str1 = row[1].value
            word = row[2].value
        else:
            s = row.extra
            str1 = row.type
            word = row.word

        if word == None and almost_done:
            break
        if word == None:
            almost_done = True
        else:
            almost_done = False
        if str1 != None:
            str1.strip()
            if word == 'speed limit':
                bb = 7
            str5 = copy.copy(str1)

            if isinstance(word,(int,long)):
                word = str(word)
            word = word.strip()
            if excel:
                str3 = row[3].value
                defin = row[4].value
            else:
                str3 = row.rel
                defin = row.definition

            if defin == 'redundant':
                redundant.append(word)
            if defin != None and defin.find("E.g.") == -1 and defin.find("EXP") == -1 \
                    and defin != 'redundant':
                if word != None:
                    word = word.strip()
                if str3 != None:
                    str3 = str3.strip()
                if str1 == None:
                    easygui.msgbox("you did not state the part of speech for " + word)
                atom = copy.copy(str1)
                str5 = str1[0:1]

                if "(" in word:
                    y = word.find("(")
                    word = word[:y-1]
                if " " in word:
                    m = word.count(" ")
                    if m == 1:
                        word1 = copy.copy(word)
                        y = word1.find(" ")
                        word1 = word1[:y]
                        doubles.append([word1,word])
                    if m == 2:
                        word1 = copy.copy(word)
                        y = word1.find(" ")
                        word1 = word1[:y]
                        triples.append([word1,word])

                atom = atom[1:2]
                str8 = copy.copy(str1)
                str8 = str1[2:3]
                if len(str1) > 4:
                    str9 = str1[4:5]
                else:
                    str9 = None

                if str5 == 'r' and atom != 's':
                    pos.append([str3,str5,str9])
                else:
                    pos.append([word,str5,str9])

                if len(str1) > 3:
                    if str1[3] == 's':
                        srelat.append(word)
                    elif str1[3] == 't':
                        trelat.append(word)
                if (atom == 'a' or atom == 'b') and str5 == "r":
                    if str3 == None:
                        bb = 8
                    if str3 not in atomic_relations:
                        atomic_relations.append(str3)
                if atom == 'b':
                    really_atomic.append(str3)
                if str5 == 'a':
                    adj.append(word)
                elif str5 == 'b':
                    aux.append(word)
                elif str5 == 'c':
                    cor.append(word)
                elif str5 == 'd':
                    detm.append(word)
                    det.append([word,atom,defin])
                elif str5 == 'r':
                    relat.append([word,str3])
                elif str5 == 'e':
                    adv.append(word)
                elif str5 == 'l' and atom == 'b':
                    lcon.append(word)
                elif str5 == 'm':
                    negg.append(word)
                elif str5 == 'n':
                    noun.append(word)
                elif str5 == 'p':
                    pronouns.append(word)
                elif str5 == 'q':
                    poss_pronouns.append(word)
                elif str5 == 'u':
                    subo.append(word)
                elif str5 == 'w':
                    dnoun.append(word)
                if atom == 'a':
                    atomic_relata.append(word)
        # in the database the definition of plural must be written as 'plural of'
                elif atom == "m":
                    str6 = defin[10:]
                    plurals.append([word,str6])
                elif atom == 'q':
                    particles.append(word)
                if atom == 'p' or atom == 'd':
                    if atom == 'p':
                        atom = 7
                    elif atom == 'd':
                        atom = 5
                    list1a = [word, atom]
                elif atom == 'c':
                    anaphoric_relations.append(str3)
                elif atom == 's':
                    str6 = defin[defin.find("=")+1:-1]
                    str6 = str6.strip()
                    str7 = defin[1:defin.find("=")]
                    str7 = str7.strip()
                    list3a = [str7, str6, defin]
                    syn_pairs.append(list3a)
                    synon.append(str7)

                if atom != 'a' and atom != 'm' and defin != "artificial" and defin != 'redundant'\
                    and defin != "postponed" and atom != 'b':
                    if str5 in category:
                        definitions.append([str3,defin,str5,atom,str8,str9])
                    else:
                        definitions.append([word,defin,str5,atom,str8,str9])

    syn_pairs.sort()
    # relations.sort()
    # relations2.sort()
    words = [adj, cor, detm, adv, lcon, noun, relat, srelat, trelat, subo,\
        aux, negg, dnoun,syn_pairs,synon,det, definitions, det_pairs, relations, \
             relations2, particles, redundant,atomic_relations, atomic_relata, \
             pronouns,poss_pronouns,plurals,neg_det,pos,really_atomic,\
             anaphoric_relations,doubles,triples]

    return words


def findinlist(str1, list1, i, j, bool1 = False):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
            if str1 == list1[d][i]:
                str2 = list1[d][j]
                if bool1:
                    return d
                else:
                    return str2
    if bool1:
        return -1
    else:
        return None

def findposinlist(str1, list1,i):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            return d
    else:
        return -1

def findin1dlist(str1,list1):

    for i in range(len(list1)):
        if str1 == list1[i]:
            return i

def isatomic(list1):

    global words
    atomic_relations = words[22]
    num = [5,14]
    if not list1[9] in atomic_relations:
        return False
    for i in num:
        if not isvariable(list1[i]):
            return False
    num = [3,4,6,7,10,11,12,13,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32, \
           33,34,35,47,48,49,50,51,52,53,54]
    for i in num:
        if list1[i] != "" and list1[i] != None:
            return False
    return True

def ismultidim(list1):

    if type(list1[0]) is list:
        return True
    else:
        return False


def cat_atoms(j,i,list,members,basic_objects,str1,bo2,words,consq,rel):

    global dv_nam
    global gen_var

    atomic_relations = words[22]
    subj = list[i][5]
    relat = list[i][9]
    obj = list[i][14]
    sent = list[i][0]

    #here we determine if the property is irrelevant for determining identity
    bool1 = False
    spec_prop = ["definite","indefinite",'general','particular']
    if list[i][8] == None:
        list[i][8] = ""
    num = [5,14,18,22]
    if subj == "t" and relat == 'LV' and obj == 'v' and j == 5:
        bb = 8

    bool3 = False
    if relat == "IA" and j == 5:
        str6 = findinlist(list[i][14],dv_nam,0,1)
        if str6 in spec_prop:
            bool3 = True

    str5 = ""
    str9 = ""
    if bool3:
        str5 = 'sp'
    elif rel == 'ir':
        str5 = "ir"
    elif relat == 'IG' and j == 5 and check_dimension(dv_nam,0,list[i][14]):
        str5 = "ir"
    elif j == 14 and relat == 'AN':
        str5 = 'ir'
    elif list[i][53] == 'cn':
        str5 = 'ir'
        str9 = "cn"
        consq.append(sent)
    elif relat not in atomic_relations:
        str5 = 'ir'
    elif relat == 'IG' and subj in gen_var and obj in gen_var:
        str5 = 'ir'
    else:
        for k in num:
            if list[i][k] in gen_var or list[i][k] == None:
                str5 = 'ir'
            else:
                str5 = ""
                break

    if list[i][14] == None:
        str3 = ""
    else:
        str3 = list[i][14]
    if str9 == "cn":
        tilde = ""
        sent = sent.replace("~","")
        sent = sent.replace("  "," ")
    else:
        tilde = list[i][8]


    if j == 5:
        if list[i][14] == None:
            str7 = ""
        elif list[i][14] in gen_var:
            str7 = "@"
        else:
            str7 = list[i][14]
        members.append([list[i][j],str1,sent,tilde + list[i][9] \
                + str7,"",str5,list[i],list[i][14]])
    else:
        if list[i][5] in gen_var:
            str7 = "@"
        else:
            str7 = list[i][5]
        members.append([list[i][j],str1,sent,str7 + tilde \
                + list[i][9],"",str5,list[i],list[i][5]])

    if str1 != "":
        if [list[i][j],str1] not in basic_objects:
            basic_objects.append([list[i][j],str1])
            bo2.append(list[i])
    return

def axioms(list1,bo2,disjuncts,tot_sent):

    already_done = []
    global dv_nam
    global sn
    added = False
    used_ax = []
    list2 = extract_list(list1,0)
    for i in range(len(list2)):
        str1 = list2[i]
        if str1 not in already_done:
            already_done.append(str1)
            g = list2.count(str1)
            if g > 1:
                conjuncts = []
                list3 = []
                for m in range(i,len(bo2)):
                    if list1[m][0] == str1:
                        prop = bo2[m][42]
                        prop = prop.replace("~","")
                        if prop not in disjuncts:
                            if bo2[m][5] == str1:
                                conjuncts.append([bo2[m],5])
                            elif bo2[m][14] == str1:
                                conjuncts.append([bo2[m],14])
                        else:
                            if bo2[m][5] == str1:
                                list3.append([bo2[m],5])
                            elif bo2[m][14] == str1:
                                list3.append([bo2[m],14])

                if len(conjuncts) == 1 and len(list3) >= 1:
                    for k in range(len(conjuncts)):
                        for j in range(len(list3)):
                            added = True
                            pos1 = conjuncts[k][1]
                            pos2 = list3[j][1]
                            rel1 = conjuncts[k][0][9]
                            rel2 = list3[j][0][9]
                            sub1 = conjuncts[k][0][5]
                            obj1 = conjuncts[k][0][14]
                            sub2 = list3[j][0][5]
                            obj2 = list3[j][0][14]
                            osec_sent = list3[j][0][0]
                            dummy = axioms2(pos1,pos2,rel1,rel2,sub1,obj1,sub2,obj2,osec_sent,tot_sent,used_ax)

                elif len(conjuncts) == 2:
                    added = True
                    pos1 = conjuncts[0][1]
                    pos2 = conjuncts[1][1]
                    rel1 = conjuncts[0][0][9]
                    rel2 = conjuncts[1][0][9]
                    sub1 = conjuncts[0][0][5]
                    obj1 = conjuncts[0][0][14]
                    sub2 = conjuncts[1][0][5]
                    obj2 = conjuncts[1][0][14]
                    osec_sent = conjuncts[1][0][0]
                    dummy = axioms2(pos1,pos2,rel1,rel2,sub1,obj1,sub2,obj2,osec_sent,tot_sent,used_ax)

                elif len(conjuncts) > 2:
                    y = 0
                    for n in range(y,g-1):
                        y += 1
                        h = y
                        while h < g:
                            added = True
                            j = h
                            h += 1
                            k = n
                            pos1 = conjuncts[k][1]
                            pos2 = conjuncts[j][1]
                            rel1 = conjuncts[k][0][9]
                            rel2 = conjuncts[j][0][9]
                            sub1 = conjuncts[k][0][5]
                            obj1 = conjuncts[k][0][14]
                            sub2 = conjuncts[j][0][5]
                            obj2 = conjuncts[j][0][14]
                            osec_sent = conjuncts[j][0][0]
                            dummy = axioms2(pos1,pos2,rel1,rel2,sub1,obj1,sub2,obj2,osec_sent,tot_sent,used_ax)
    return added

def axioms2(pos1,pos2,rel1,rel2,sub1,obj1,sub2,obj2,osec_sent,tot_sent,used_ax):

    global dv_nam
    global idf_var
    global sn

    rn_list = []
    thing_con = findinlist('thing',dv_nam,1,0)
    if thing_con == None:
        thing_con = idf_var[0]
        for i in range(len(tot_sent)):
            if tot_sent[i][4] == "ID":
                tot_sent[i][2] += " & (" + thing_con + "= thing)"
                break
        del idf_var[0]
    else:
        rn1 = "(" + "e" + mini_c + thing_con + ")"
        rn_list.append(rn1)
    if sub1 != 'b':
        rn1 = "(b" + mini_c + sub1 + ")"
        rn_list.append(rn1)
    if obj1 != 'c':
        rn1 = "(c" + mini_c + obj1 + ")"
        rn_list.append(rn1)
    thing_int = idf_var[0]
    del idf_var[0]
    new_var = idf_var[0]
    del idf_var[0]
    if thing_int != 'd':
        rn1 = "(d" + mini_c + thing_int + ")"
        rn_list.append(rn1)
    if pos1 == 5 and pos2 == 5:
        thing_var = obj2
        oax = "(((b" + rel1 + "c) & (dIGe)) " + conditional + \
        " (b~" + rel2 + "d)) & (e=thing)"
        sent3 = "(" + sub1 + "~" + rel2 + thing_int + ")"
        sent6 = "(" + sub1 + rel2 + new_var + ")"
        oax_name = "AX." + rel1 + "." + rel2 + "." + "ss"
    elif pos1 == 5 and pos2 == 14:
        thing_var = sub2
        oax = "(((b" + rel1 + "c) & (dIGe)) " + conditional + \
        " (d~" + rel2 + "b)) & (e=thing)"
        sent3 = "(" + thing_int + "~" + rel2 + sub1 + ")"
        sent6 = "(" + new_var + rel2 + sub1 + ")"
        oax_name = "AX." + rel1 + "." + rel2 + "." + "so"
    elif pos1 == 14 and pos2 == 5:
        thing_var = obj2
        oax = "(((b" + rel1 + "c) & (dIGe)) " + conditional + \
        " (c~" + rel2 + "d)) & (e=thing)"
        sent3 = "(" + obj1 + "~" + rel2 + thing_int + ")"
        sent6 = "(" + sub2 + rel2 + new_var + ")"
        oax_name = "AX." + rel1 + "." + rel2 + "." + "os"
    elif pos1 == 14 and pos2 == 14:
        thing_var = sub2
        oax = "(((b" + rel1 + "c) & (dIGe)) " + conditional + \
        " (d~" + rel2 + "c)) & (e=thing)"
        sent3 = "(" + thing_int + "~" + rel2 + obj1 + ")"
        sent6 = "(" + new_var + rel2 + obj1 + ")"
        oax_name = "AX." + rel1 + "." + rel2 + "." + "oo"

    sent1 = "(" + sub1 + rel1 + obj1 + ")"
    sent3a = sent3.replace("~","")
    sent5 = "(" + sub2 + rel2 + obj2 + ")"
    sent2 = "(" + thing_int + "IG" + thing_con + ")"
    nax = "(" + sent1 + " & " + sent2 + ") " + conditional \
        + " " + sent3
    rename = build_sent_list(rn_list)
    ax_enti = "(" + thing_var + "IG" + thing_con + ")"
    subst1 = "(" + thing_var + mini_c + new_var + ")"
    subst2 = "(" + new_var + mini_c + thing_int + ")"
    sent4 = "(" + thing_var + "IG" + thing_con + ")"
    subst4 = sent5 + " " + conditional + " " + sent6
    subst4a = sent6 + " " + conditional + " " + sent3a
    sent1p = name_sent(sent1)
    sent2p = name_sent(sent2)
    sent3p = name_sent(sent3)
    sent3ap = name_sent(sent3a)
    sent6p = name_sent(sent6)
    sent4p = name_sent(sent4)
    sent5p = name_sent(sent5)
    naxp = "(" + sent1p + " & " + sent2p + ") " + conditional \
        + " " + sent3p
    subst4p = sent5p + " " + conditional + " " + sent6p
    subst4ap = sent6p + " " + conditional + " " + sent3ap
    d = findposinlist(oax,used_ax,0)
    if d > -1:
        e = used_ax[d][1]
    else:
        e = sn+1
        sn += 1
        used_ax.append([oax, sn])
        tot_sent.append([sn,oax,"","",oax_name,"",""])
    sn += 1
    tot_sent.append([sn,rename,"","","RN","",""])
    sn += 1
    tot_sent.append([sn,nax,naxp,"","SUB",e,sn-1])
    sn += 1
    tot_sent.append([sn,ax_enti,"","","AY ENT","",""])
    sn += 1
    tot_sent.append([sn,sent2,sent2p,"","AY ENT","",""])
    sn += 1
    tot_sent.append([sn,subst1,"","","OS",sn-1,sn-2])
    sn += 1
    tot_sent.append([sn,subst2,"","","OS",sn-2,sn-3])
    sn += 1
    tot_sent.append([sn,subst4,subst4p,"","SUB",sn-2,""])
    sn += 1
    tot_sent.append([sn,subst4a,subst4ap,"","SUB",sn-2,""])

# (d'HWv) & (wIGd')
# ((bHWc) & (dIGe)) > (d~IGb) & (e=thing)
# (b>d) & (c>v) & (d>f) & (e>g)
# (d' HW v) & (fIGe) > (f~IGd'))
# (wIGe)
# (fIGe)
# (wIGe) , (wIGd') , (fIGe)
# (f>g)
# (w=g)
# (f~IGd') > (w~IGd')

def find_group(str1,all_sent,subj,basic_objects):

    global gen_var
    pair1 = ["integer",'NUMBER']
    pair2 = ["",""]
    synonyms = [pair1,pair2]
    exceptions = ['this'+un,'that'+un]
    if subj == 't':
        bb = 8

    str3 = None
    str4 = None
    list1 = []
    for i in range(len(basic_objects)):
        if basic_objects[i][0] == subj:
            str4 = basic_objects[i][1]
            if str4 not in list1:
                list1.append(str4)

    for i in range(len(all_sent)):
        if all_sent[i][9] == "IG" and all_sent[i][8] != "~" and all_sent[i][5] == subj:
            str2 = all_sent[i][14]
            str3 = findinlist(str2,dv_nam,0,1)
            if str3 in exceptions:
                str3 = None
            if str3 != None:
                d = findposinlist(str3,synonyms,0)
                if d > -1:
                    str3 = synonyms[d][1]
                str3 = str3.upper()
                if str3 not in list1:
                    list1.append(str3)
            elif all_sent[i][14] not in gen_var:
                str3 = all_sent[i][14].upper()
                str3 = str3 + " THINGS"
                list1.append(str3)

    if str3 == None and str4 == None and list1 == []:
        return None
    else:
        list2 = [subj,list1]
        return list2

def simple_id(tot_sent,all_sent,identities):

    num = [5,14,18,22]
    for i in range(len(identities)):
        str1 = identities[i][0][0]
        str2 = identities[i][0][1]
        for j in range(len(all_sent)):
            for k in num:
                if all_sent[j][k] == None and (k == 18 or k == 22):
                    break
                if all_sent[j][k] == str1 or all_sent[j][k] == str2:
                    list1 = copy.deepcopy(all_sent[j])
                    old_sent = all_sent[j][0]
                    oldp = all_sent[j][42]
                    if all_sent[j][k] == str1:
                        list1[k] = str2
                    else:
                        list1[k] = str1
                    new_sent = build_sent(list1)
                    newp = name_sent(new_sent)
                    list1[0] = new_sent
                    list1[42] = newp
                    dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"SUB")
                    all_sent.append(list1)



def identity(all_sent,tot_sent,basic_objects,words,candd,conditionals,\
    prop_sent,prop_name,id_num,identities,idf_var):

    global sn,psent,impl,definite2,ind_var,gen_var,idf_var2,never_used,simple_id
    atomic = words[29]
    members = []
    non_id = []
    used_var = []
    consq = []
    irrel_group = []
    for i in range(len(idf_var2)):
        if idf_var2[i] not in idf_var:
            used_var.append(idf_var2[i])
        else:
            if i > 26:
                break

    if identities != []:
        dummy = simple_id(tot_sent,all_sent,identities)
    dv_list = id_sent(dv_nam,all_sent,irrel_group)
    tot_sent.append(["","","","","","",""])
    nw = []
    bo2 = []



    num = [5,14,18,22]
    listfa = []
    j = -1
    while j < (len(all_sent)) -1:
        j += 1
        relat = all_sent[j][9]
        if all_sent[j][9] == ne:
            non_id.append([all_sent[j][0],all_sent[j][5],all_sent[j][14]])
        else:
            for p in num:
                if all_sent[j][p] != None and isvariable(all_sent[j][p]):
                    if j == 5 and p == 14:
                        bb = 8
                    rel = ""
                    if relat == "A" or (relat == 'T' and p == 14):
                        kind = 'MOMENT'
                    elif relat == 'I' and p == 5:
                        kind = 'RELATIONSHIP'
                    elif relat == 'AB' or relat == 'LF' or relat == 'AB' or (relat == 'S' and p == 14):
                        kind = 'POINT'
                    elif relat == 'AF' or (relat == 'AN' and p == 14):
                        kind = 'NUMBER'
                    elif relat == 'AI' and p == 5 or (relat == 'TK' and p == 14):
                        kind = 'THOUGHT'
                        rel = "ir"
                    elif relat == 'AI' and p == 14:
                        kind = 'IMAGINATION'
                    elif relat == "IG" and p == 14:
                        kind = "GROUP"
                    elif relat == "H" and p == 14:
                        kind = "PROPERTY"
                    elif relat == "IA" and p == 14:
                        kind = "PROPERTY"
                    elif relat == 'P' and p == 14:
                        kind = 'POSSIBLE WORLD'
                    elif (relat == 'AI' and p == 5) or (relat == 'DS' and p == 14):
                        kind = 'POSSIBLE RELATIONSHIP'
                    elif relat == 'AW' and p == 5:
                        kind = 'WORD'
                    elif relat == 'HW' and p == 5:
                        kind = 'NATURAL WHOLE'
                        nw.append(all_sent[j][5])
                    elif (relat == 'TK' or relat == 'DS') and p == 5:
                        kind = 'MIND'
                    elif relat == "S" and p == 5:
                        kind = 'MATTER'
                    elif relat == "SS" and p == 14:
                        kind = 'SENSORIUM'
                    elif relat == "SS" and p == 5:
                        kind = 'SENSATION'
                    else:
                        kind = ""
                    if len(members) > 30:
                        bb = 7
                    if all_sent[j][8] == "~":
                        kind = ""
                    dummy = cat_atoms(p,j,all_sent,members,basic_objects,kind,bo2,words,consq,rel)


    mem_var = []
    basic_objects2 = copy.deepcopy(basic_objects)
    basic_objects = sorted(basic_objects, key = operator.itemgetter(1,0))
    has2groups = []
    #if something is both matter and a natural whole then it is a natural whole
    #remove blanks in the second dimension
    i = -1
    while i < len(basic_objects) -1:
        i += 1
        if basic_objects[i][0] in nw and basic_objects[i][1] == 'MATTER':
            del basic_objects[i]
            i -= 1
    for i in range(len(members)):
        if i == 13:
            bb = 8
        if members[i][0] == 'v':
            pp = 7
        if members[i][4] == 'general':
            members[i][2] = members[i][2] + "*"
        if members[i][1] == "":
            group = find_group(members[i][2],all_sent,members[i][0],basic_objects)
            if group == None:
                members[i][1] = 'THING'
            else:
                if len(group[1]) > 1:
                    if group not in has2groups:
                        has2groups.append(group)
                group = group[1][0]
                members[i][1] = group
                # basic_objects.append([members[i][0],group])
        elif members[i][1] == 'MATTER' and members[i][0] in nw:
            members[i][1] = 'NATURAL WHOLE'
    #if something is both matter and material whole then it belongs to the former group

    for i in range(len(members)):
        if members[i][0] == 'u':
            pp = 7
        p = findposinlist(members[i][0],has2groups,0)
        if p > -1:
            members[i][1] = has2groups[p][1][0]

    members = sorted(members, key = operator.itemgetter(1,0))
    list4 = []
    list5 = []
    list6 = []
    list9 = []
    irrel = []
    senti = []
    senti2 = []
    gv_corr = []
    spec_prop = []
    member_prop = []
    str2 = ""
    str4 = ""
    dis_con = False
    #fill member_prop list

    if members != []:
        for i in range(len(members)):
            str1 = members[i][0]
            str3 = members[i][1]
            isirel = members[i][5]
            if str1 == 'p':
                bb = 7
            if str1 != str2 and str2 != "":
                senti = senti + senti2
                list5.append([str2,members[i-1][1],list4,irrel,spec_prop,senti])
                sn += 1
                member_prop.append([str2,members[i-1][1],list6,list4,list9,irrel,gv_corr,sn])
                list4 = []
                list6 = []
                list9 = []
                irrel = []
                senti = []
                senti2 = []
                gv_corr = []
                spec_prop = []
                if "@" in members[i][3]:
                    gv_corr.append([members[i][3],members[i][7]])
                if members[i][1] != "":
                    if isirel == "":
                        senti.append(members[i][6])
                        list4.append(members[i][2])
                        list6.append(members[i][3])
                    elif isirel == 'ir' or isirel == 'cn':
                        senti2.append(members[i][6])
                        irrel.append(members[i][2])
                        list9.append(members[i][3])
                    elif isirel == 'sp':
                        spec_prop.append(members[i][2])
            else:
                if "@" in members[i][3]:
                    gv_corr.append([members[i][3],members[i][7]])
                if members[i][1] != "":
                    if isirel == "":
                        senti.append(members[i][6])
                        list4.append(members[i][2])
                        list6.append(members[i][3])
                    elif isirel == 'ir' or isirel == 'cn':
                        senti2.append(members[i][6])
                        irrel.append(members[i][2])
                        list9.append(members[i][3])
                    elif isirel == 'sp':
                        spec_prop.append(members[i][2])
            str2 = str1

        if members[i][1] != "":
            sn += 1
            senti = senti + senti2
            list5.append([str2,members[i-1][1],list4,irrel,spec_prop,senti])
            member_prop.append([str2,members[i-1][1],list6,list4,list9,irrel,gv_corr,sn])
        if has2groups != []:
            for i in range(len(has2groups)):
                str1 = has2groups[i][0]
                list3 = []
                for m in range(len(list5)):
                    if list5[m][0] == str1:
                        break
                for j in range(1,len(has2groups[i][1])):
                    str2 = has2groups[i][1][j].upper()
                    list5.append([str1,str2,list5[m][2],list5[m][3],list5[m][4],list5[m][5]])
                    sn += 1
                    member_prop.append([str1,str2,member_prop[m][2],member_prop[m][3],\
                        member_prop[m][4],member_prop[m][5],member_prop[m][6],sn])
        list5 = sorted(list5, key = operator.itemgetter(1,0))
        member_prop = sorted(member_prop, key = operator.itemgetter(1,0))
        dummy = missing_variables(tot_sent,used_var,list5,never_used,all_sent,irrel)
        str3 = list5[0][1]
        tot_sent.append(["",str3,"","","","",""])
        non_id2 = copy.deepcopy(non_id)
        b = 0
        for i in range(len(list5)):
            str1 = list5[i][0] + " "
            str4 = list5[i][1]
            if i == 11:
                bb = 7
            if str1 == 'n ':
                bb = 8
            for k in range(2,5):
                str2 = ""
                str5 = ""
                if len(list5[i][2]) == 0 and k == 2:
                    if list5[i][0] in gen_var:
                        str2 = "[general]"
                    elif list5[i][0] in definite2:
                        str2 = "[definite]"
                    elif list5[i][0] in ind_var:
                        str2 = "[indefinite]"
                elif k == 4 and list5[i][4] != []:
                    for s in range(len(list5[i][k])):
                        if list5[i][k][s] != "":
                            if str5 == "":
                                str5 =  " [" + list5[i][k][s]
                            else:
                                str5 += " & " + list5[i][k][s]
                    str2 += str5 + "]"
                else:
                    if list5[i][k] != []:
                        for s in range(len(list5[i][k])):
                            if list5[i][k][s] != "":
                                if str2 == "":
                                    str2 = list5[i][k][s]
                                else:
                                    str2 += " & " + list5[i][k][s]
                        if k == 3:
                            str2 = " {" + str2 + "}"
                if k == 2 and str2 == "" and str1 in gen_var:
                    str1 += "[general]"
                if str2 != "":
                    str1 += str2

            if str4 != str3:
                tot_sent.append(["",str4,"","","","",""])
            str3 = str4
            b += 1
            str7 = str(b) + "a"
            if non_id2 != []:
                not_id = isnotid(list5[i][0],non_id2,tot_sent)
            tot_sent.append([member_prop[i][7],str1,"","","","",""])

        str1 = ""

    tot_sent.append(["","","","","","",""])
    tot_sent.insert(id_num-1,[id_num,dv_list[0],dv_list[1],"",'ID'])
    added = identity2(member_prop,tot_sent,list5,\
                  all_sent,identities,idf_var,irrel_group,words)
    hh = len(tot_sent)
    list4 = []
    negat = []
    sent = []
    disjuncts = []
    last_num = sn
    for i in range(len(tot_sent)):
        if tot_sent[i][2] != "":
            sent.append([tot_sent[i][0],tot_sent[i][2]])
            negat.append(tot_sent[i][3])
    # run statement logic
    consistent = plan(sent, prop_sent, candd, conditionals, prop_name,disjuncts,1,negat)

    g = len(tot_sent)
    rearrange = False
    if consistent:
        added = axioms(basic_objects2,bo2,disjuncts,tot_sent)
        if added:
            dummy = rearrange2(g,tot_sent,prop_sent,conditionals,candd,last_num)
            sn = prop_sent[-1][0]
            consistent = statement_logic(prop_sent,conditionals,candd,disjuncts,2)
        if consistent and impl != nonseq:
            print "False"

    if impl != "":
        for i in range(len(prop_sent)):
            if prop_sent[i][1] == top or prop_sent[i][1] == bottom:
                anc1 = prop_sent[i][0]
                break
        if consistent:
            for i in range(len(prop_sent)):
                if len(prop_sent[i][1]) < 3 and prop_sent[i][1] != top and prop_sent[i][1] != bottom:
                    if i == 0:
                        str1 = prop_sent[i][2] + prop_sent[i][1]
                    else:
                        str1 += ", " + prop_sent[i][2] + prop_sent[i][1]
            sn += 1
            prop_sent.append([sn,str1,"","","","","","",""])
            candd = sorted(candd, key = operator.itemgetter(1,0))
            for i in range(len(candd)):
                if len(candd[i][1]) < 3:
                    if i == 0:
                        str1 = candd[i][2] + candd[i][1]
                    else:
                        str1 += " " + candd[i][2] + candd[i][1]
            sn += 1
            prop_sent.append([sn,str1,"","","","","","",""])
            sn += 1
            prop_sent.append([sn,top,"","","","","","",""])

        if not consistent and impl == implies:
            str1 = bottom + " & " + bottom
            str2 = top
        if consistent and impl == implies:
            str1 = bottom + " & " + top
            str2 = bottom
        if not consistent and impl == nonseq:
            str1 = top + " & " + bottom
            str2 = bottom
        if consistent and impl == nonseq:
            str1 = top + " & " + top
            str2 = top
        sn += 1
        prop_sent.append([sn,str1,"","&I",anc1,sn-1,"","","",""])
        sn += 1
        prop_sent.append([sn,str2,"",str2 + "I",sn-1,"","","",""])


    i = -1
    list2 = []
    list1 = []
    list5 = []
    bool1 = False
    rn_used = False
    for i in range(0,len(tot_sent)):
        if i == 9:
            bb = 7
        str1 = tot_sent[i][4][:2]
        str2 = tot_sent[i][4][:2]
        if tot_sent[i][4] == "" and not bool1:
            list6 = copy.deepcopy(tot_sent[i])
            list5.append(list6)
        elif tot_sent[i][4] == "ID":
            list6 = copy.deepcopy(tot_sent[i])
            list5.append(list6)
            bool1 = True
        elif 'DF' == str1 or 'NC' == str1 or 'AX' == str1 \
             or 'RN' == str2:
            list4 = copy.deepcopy(tot_sent[i])
            list1.append(list4)
            rn_used = True
        elif bool1:
            if str1 == 'DE':
                tot_sent[i][4] = "DF " + tot_sent[i][4][3:]
            elif str1 == 'AY':
                tot_sent[i][4] = "AX ENT"
            list3 = copy.deepcopy(tot_sent[i])
            list2.append(list3)
    if rn_used:
        tot_sent = []
        for i in range(len(list5)):
            tot_sent.append(list5[i])
        for i in range(len(list1)):
            tot_sent.append(list1[i])
        tot_sent.append(["","","","","",""])
        for j in range(len(list2)):
            tot_sent.append(list2[j])

        list1 = []
        j = 0
        for i in range(len(tot_sent)):
            if (i > 3 and tot_sent[i][0] != "") or i <= 3:
                j += 1
                list1.append([tot_sent[i][0],j])
                tot_sent[i][0] = j

        for i in range(len(prop_sent)):
            if i == 53:
                bb = 8
            d = findinlist(prop_sent[i][0],list1,0,1)
            if d != None:
                prop_sent[i][0] = d
            for j in range(4,8):
                if prop_sent[i][j] == None or prop_sent[i][j] == "":
                    break
                d = findinlist(prop_sent[i][j],list1,0,1)
                if d != None:
                    prop_sent[i][j] = d

        for i in range(len(tot_sent)):
            for j in range(0,5):
                if len(tot_sent[i]) > 5 + j:
                    if tot_sent[i][j+5] != "":
                        g = findinlist(tot_sent[i][j+5],list1,0,1)
                        tot_sent[i][j+5] = g
                else:
                    break

        # for i in range(len(prop_sent)):
        #     if prop_sent[i][0] < last_num:
        #         prop_sent[i][0] = findinlist(prop_sent[i][0],list1,0,1)
        #     for j in range(0,5):
        #         if len(prop_sent[i]) > 7:
        #             if prop_sent[i][j+4] != None and prop_sent[i][j+4] != "":
        #                 g = findinlist(prop_sent[i][j+4],list1,0,1)
        #                 if g != None:
        #                     prop_sent[i][j+4] = g
        #         else:
        #             break

    if not consistent:
        list1 = []
        for i in range(len(prop_sent)-1,0,-1):
            if prop_sent[i][4] == "":
                break
            if (i != len(prop_sent)-1 and prop_sent[i][0] in list1) or i == len(prop_sent)-1:
                for j in range(4,8):
                    if prop_sent[i][j] == None or prop_sent[i][j] == "":
                        break
                    list1.append(prop_sent[i][j])
        list1.append(prop_sent[-1][0])
        list1.sort()
        str1 = str(list1[0])
        start = tot_sent[-1][0]
        for i in range(1,len(list1)):
            if list1[i] > start:
                break
            str1 += " " + str(list1[i])
        tot_sent.append(["","relevant sentences: " + str1,"","","","",""])
        bool1 = False
        list3 = []
        list2 = []
        for i in range(len(prop_sent)):
            if prop_sent[i][0] <= start:
                list2.append(prop_sent[i])
                list3.append([prop_sent[i][0],prop_sent[i][0]])
            elif prop_sent[i][0] in list1:
                if not bool1:
                    bool1 = True
                    j = list3[-1][0]
                j += 1
                list3.append([prop_sent[i][0],j])
                prop_sent[i][0] = j
                list2.append(prop_sent[i])
                anc1 = prop_sent[i][4]
                nanc1 = findinlist(anc1,list3,0,1)
                list2[-1][4] = nanc1
                if prop_sent[i][5] != "" and prop_sent[i][5] != None:
                    anc1 = prop_sent[i][5]
                    nanc1 = findinlist(anc1,list3,0,1)
                    list2[-1][5] = nanc1
                    if prop_sent[i][6] != "" and prop_sent[i][6] != None:
                        anc1 = prop_sent[i][6]
                        nanc1 = findinlist(anc1,list3,0,1)
                        list2[-1][6] = nanc1
                        if prop_sent[i][7] != "" and prop_sent[i][7] != None:
                            anc1 = prop_sent[i][7]
                            nanc1 = findinlist(anc1,list3,0,1)
                            list2[-1][7] = nanc1
        prop_sent = list2
    tot_prop_sent.append(prop_sent)

    # end4
    return tot_sent


def missing_variables(tot_sent,used_var,list5,never_used,all_sent,irrel):

    global definite2
    global ind_var
    global gen_var

    var = [definite2,ind_var,gen_var]
    tot_var = definite2 + ind_var + gen_var
    tot_var2 = []
    var_name = ["definite abbreviations:","indefinite abbreviations:","general abbreviations:"]
    for i in range(0,3):
        str1 = ""
        if var[i] != []:
            for j in range(len(var[i])):
                tot_var2.append(var[i][j])
                str1 += " " + var[i][j]
            str1 = var_name[i] + str1
            tot_sent.append(["",str1,"","","","","",""])
    str1 = ""
    for i in range(len(used_var)):
        if used_var[i] not in tot_var2 and used_var[i] not in never_used:
            str1 += " " + used_var[i]
    if str1 != "":
        tot_sent.append(["","variable type missing: " + str1,"","","","","",""])

    str1 = ""
    list2 = []
    for i in range(len(list5)):
        list2.append(list5[i][0])
    for i in range(len(tot_var)):
        if i == 10:
            bb = 8
        if tot_var[i] not in list2 and tot_var[i] not in irrel and \
            tot_var[i] not in never_used:
            str1 += " " + tot_var[i]
    if str1 != "":
        tot_sent.append(["","missing variables:" + str1,"","","","","",""])



def rearrange2(g,tot_sent,prop_sent,conditionals,candd,last_num):

    list4 = []
    list3 = []
    h = last_num
    start = copy.copy(h)

    for i in range(g,len(tot_sent)):
        if i == 51:
            bb = 8
        h += 1
        list3.append([h,tot_sent[i][0]])
        tot_sent[i][0] = h
        if tot_sent[i][5] != "":
            m = tot_sent[i][5]
            n = findinlist(m,list3,1,0)
            tot_sent[i][5] = n
            if tot_sent[i][6] != "":
                m = tot_sent[i][6]
                n = findinlist(m,list3,1,0)
                tot_sent[i][6] = n

        if tot_sent[i][2] != "":
            list4.append([tot_sent[i][0],tot_sent[i][2],tot_sent[i][3],"","","",""])
            if not os(tot_sent[i][2]):
                list1 = mainconn(tot_sent[i][2])
                list2 = prepare_iff_elim(tot_sent[i][2],list1[0],list1[1],tot_sent[i][0])
                conditionals.insert(0,list2)
            else:
                list5 = [tot_sent[i][0],tot_sent[i][2],tot_sent[i][3]]
                candd.insert(0,list5)


    b = list3[-1][0]
    bool1 = False
    i = -1
    while i < len(prop_sent)-1:
        i += 1
        if prop_sent[i][0] > start:
            bool1 = True
            for j in range(len(list4)):
                prop_sent.insert(i+j,list4[j])
            break
    if bool1:
        k = i + j + 1
        list3 = []
        for i in range(k,len(prop_sent)):
            b += 1
            list3.append([prop_sent[i][0],b])
            prop_sent[i][0] = b
            for j in range(4,len(prop_sent[i])):
                if prop_sent[i][j] != "" and prop_sent[i][j] != None:
                    if findinlist(prop_sent[i][j],list3,0,1) != None:
                        prop_sent[i][j] = findinlist(prop_sent[i][j],list3,0,1)
                elif prop_sent[i][j] == None:
                    break

        for i in range(len(candd)):
            if candd[i][0] > last_num:
                candd[i][0] = findinlist(candd[i][1],prop_sent,1,0)

        for i in range(len(conditionals)):
            if conditionals[i][2] > last_num:
                conditionals[i][2] = findinlist(conditionals[i][4],prop_sent,1,0)

    return

def isnotid(str1,non_id2,tot_sent):

    for y in range(len(non_id2)):
        if str1 in non_id2[y][0]:
            tot_sent.append(["",non_id2[y][0],"","","","",""])
            del non_id2[y]
            return True
    return False

def new_id(str1,str2,tot_sent,identities,anc1,anc2,done_os):

    global sn
    list3 = [str1,str2]
    list3.sort()
    if str1 == 'r' and str2 == 'x':
        bb = 8
    str3 = "(" + str1 + "=" + str2 + ")"
    if list3 not in done_os:
        done_os.append(list3)
        identities.append([list3,sn+1,anc1,anc2])
        sn += 1
        tot_sent.append([sn,str3,"","","LL",anc1,anc2,""])
        return None
    else:
        d = findinlist(str3,tot_sent,1,0)
        return d

def one_way(str1,str2,done_os,tot_sent,instant,anc1,anc2,new_abb):

    global sn
    list1 = [str2,str1]

    if list1 not in done_os:
        done_os.append([str2,str1])
        str3 = "(" + str2 + mini_c + new_abb + ")"
        str4 = "(" + new_abb + mini_c + str1 + ")"
        instant.append([str2,new_abb,sn+1,anc1,anc2])
        instant.append([new_abb,str1,sn+2,anc1,anc2])
        sn += 1
        tot_sent.append([sn,str3,"","","OS",anc1,anc2,""])
        sn += 1
        tot_sent.append([sn,str4,"","","OS",anc1,anc2,""])

def identity2(member_prop,tot_sent,list5,all_sent,identities4,\
              idf_var,irrel_group,words):

    global definite2, gen_var, ind_var,sn
    atomic_rel = words[29]
    undefined = words[22]
    st = time.time()
    list7 = []
    list8 = []
    instant = []
    temp_id = []
    done_os = []
    newid = []
    new_id3 = []
    newos = []
    newos3 = []
    identities = []
    potential = []
    potential2 = []
    pid = []
    pot_num = []
    pos = []
    tpos2 = []
    tpotential2 = []
    tpid = []
    tpid2 = []
    tpos = []
    k = 0
    str1 = ""
    bool1 = False
    for i in range(1,len(list5)):
        if i == 14:
            bb = 8
        if list5[i][1] != list5[i-1][1]:
            if i - k > 1:
                list7.append([k,i-1])
                if bool1:
                    str1 += " " + list5[i-1][1]
                    list8.append([k,i-1])
                    bool1 = False
            k = i
        if list5[i][0] in gen_var:
            bool1 = True
    if i - k > 0:
        list7.append([k,i])
        if bool1:
            list8.append([k,len(list5)-1])

    tot_sent.append(["","groups used " + str1,"","","","",""])



    # simple matrix
    # y = 0
    # for n in range(y,g-1):
    #     y += 1
    #     h = y
    #     while h < g:
    #         h += 1
    #         list4.append([n,h-1])

    for i in range(len(list8)):
        y = list8[i][0]
        g = list8[i][1]
        y -= 1
        for n in range(y,g-1):
            y += 1
            h = y
            while h < list8[i][1]:
                h += 1
                str1 = member_prop[y][0]
                str2 = member_prop[h][0]
                list1 = [str1,str2]
                list1.sort()
                potential.append(list1)

    already_done = []
    for i in range(len(list8)):
        y = list8[i][0]
        g = list8[i][1]
        y -= 1
        for n in range(y,g-1):
            y += 1
            str1 = member_prop[y][0]
            anc1 = member_prop[y][7]
            h = y
            while h < list8[i][1]:
            # while h < g:
                h += 1
                str2 = member_prop[h][0]
                anc2 = member_prop[h][7]
                alread_id = already_identical(identities4,member_prop[h][0],member_prop[y][0])
                qsent = ""
                rsent = ""
                tpos = []
                tpid = []
                newid = []
                temp_newos = []
                newos3 = []
                if str1 == 'o' or str2 == "p":
                    bb = 8
                if h == 7 and y == 8:
                    bb = 8
                if member_prop[y][1] == 'GROUP' or member_prop[h][1] == 'PROPERTY':
                    if (str1 in irrel_group or str2 in irrel_group):
                        bool2 = False
                    else:
                        bool2 = True
                else:
                    bool2 = True

                list4 = [str1,str2]
                list4.sort()
                if list4 in already_done:
                    bool1 = False
                else:
                    bool1 = True
                if not alread_id and bool1 and bool2:
                    new_var_used = False
                    already_done.append(list4)
                    if str1 in gen_var or str2 in gen_var:
                        match_found = "n"

                        if len(member_prop[y][2]) == 0 and len(member_prop[h][2]) == 0:
                            match_found = "y"
                            small = y
                            large = h
                            if str1 in gen_var and str2 in gen_var:
                                g = 11
                            elif str1 in gen_var:
                                g = 12
                            elif str2 in gen_var:
                                g = 21
                        elif len(member_prop[y][2]) > len(member_prop[h][2]) and str2 in gen_var:
                            large = y
                            small = h
                            g = 21
                            if len(member_prop[h][2]) == 0:
                                match_found = "y"
                        elif len(member_prop[y][2]) < len(member_prop[h][2]) and str1 in gen_var:
                            small = y
                            large = h
                            bool1 = True
                            g = 12
                            if len(member_prop[y][2]) == 0:
                                match_found = "y"
                        elif (len(member_prop[y][2]) == len(member_prop[h][2])) and str1 in gen_var \
                            and str2 in gen_var:
                            small = y
                            large = h
                            g = 11
                        elif (len(member_prop[y][2]) == len(member_prop[h][2])) and str1 in gen_var:
                            small = y
                            large = h
                            g = 12
                        elif (len(member_prop[y][2]) == len(member_prop[h][2])) and str2 in gen_var:
                            small = y
                            large = h
                            g = 21

                        if small == 6 and large == 8:
                            bb = 8
                        slist = member_prop[small][2] + member_prop[small][4]
                        slist_sent = member_prop[small][3] + member_prop[small][5]
                        llist = member_prop[large][2] + member_prop[large][4]
                        llist_sent = member_prop[large][3] + member_prop[large][5]
                        irrel = member_prop[small][5] + member_prop[large][5]

                        status = ""
                        if match_found == 'n':
                            match_found = 'y'
                            for q in range(len(member_prop[small][2])):
                                for r in range(len(member_prop[large][2])):
                                    temp_str1 = member_prop[small][2][q]
                                    temp_str2 = member_prop[large][2][r]
                                    temp_str1a = member_prop[small][3][q]
                                    temp_str2a = member_prop[large][3][r]
                                    bool1 = equality(member_prop,small,large,q,r,list5,potential,potential2,g)
                                    if bool1[0] == 'n' and status == 'm':
                                        bb = 8
                                    if bool1[0] == 'm':
                                        bb = 8


                                    if bool1[0] == 'y' or bool1[0] == 'm' or bool1[0] == 'u':
                                        if status == "":
                                            status = bool1[0]
                                        elif status == 'm':
                                            status = 'm'
                                        elif status == 'u':
                                            pass
                                        else:
                                            status = bool1[0]


                                        qsent = member_prop[small][3][q]
                                        rsent = member_prop[large][3][r]
                                        slist.remove(member_prop[small][2][q])
                                        llist.remove(member_prop[large][2][r])
                                        slist_sent.remove(qsent)
                                        llist_sent.remove(rsent)
                                        list11 = [qsent,rsent]
                                        list11.sort()
                                        list14 = [str1,str2]
                                        list14.sort()
                                        if g == 11 and status == 'm':
                                            if newid != []:
                                                tpid = tpid + newid
                                                newid = []
                                            if tpid == []:
                                                tpid.append([list14,[small,large]])
                                            for x in range(len(bool1[1])):
                                                if bool1[1][x] not in tpid2:
                                                    tpid2.append(bool1[1][x])

                                        if g == 11:
                                            newid.append(list11)
                                        elif status == 'm':
                                            if tpos == [] and newos != []:
                                                newos3 = []
                                            if tpos == []:
                                                tpos.append([list14,[small,large]])
                                            for x in range(len(bool1[1])):
                                                if bool1[1][x] not in tpos2:
                                                    tpos2.append(bool1[1][x])

                                        else:
                                            if list11 not in newos3:
                                                newos3.append(list11)
                                                if not new_var_used:
                                                    new_abb = idf_var[0]
                                                    new_var_used = True
                                                temp_newos.append(list11)
                                                # nq_sent = qsent.replace(member_prop[small][0],new_abb)
                                                # nr_sent = rsent.replace(member_prop[large][0],new_abb)
                                                # if bool1[0] == 'u':
                                                #     temp_newos.append([rsent,nr_sent,"","u",bool1[1]])
                                                #     temp_newos.append([nr_sent,qsent,"","u",bool1[1]])
                                                # else:
                                                #     temp_newos.append([rsent,nr_sent,"","",""])
                                                #     temp_newos.append([nr_sent,qsent,"","",""])
                                        break
                                else:
                                    match_found = 'n'
                                    break
                        #identity list
                        if match_found == 'y' or match_found == 'm':
                            match_found = "n"

                            if status == 'u':
                                bb = 8

                            if g != 11:
                                del idf_var[0]
                            if status == 'm':
                                bb = 8
                                if g == 11:
                                    potential2.append([tpid[0][0],tpid2,g,"ui",tpid[0][1]])
                                    potential.remove(tpid[0][0])
                                else:

                                    potential2.append([tpos[0][0],tpos2,g,"uo",tpos[0][1]])
                                    potential.remove(tpos[0][0])
                            else:

                                if g == 11 and status == 'y':
                                    dummy = new_id(str1,str2,tot_sent,identities,anc1,anc2,done_os)
                                if newid != []:
                                    for q in range(len(newid)):
                                        new_id3.append([newid[q][0],newid[q][1],sn])
                                if temp_newos != []:
                                    if g == 12:
                                        dummy = one_way(str1,str2,done_os,tot_sent,instant,anc1,anc2,new_abb)
                                    elif g == 21:
                                        dummy = one_way(str2,str1,done_os,tot_sent,instant,anc1,anc2,new_abb)
                                    anc3 = findinlist(new_abb,instant,1,2)
                                    anc4 = findinlist(new_abb,instant,0,2)
                                    for q in range(len(temp_newos)):
                                        qsent = temp_newos[q][0]
                                        rsent = temp_newos[q][1]
                                        nq_sent = qsent.replace(member_prop[small][0],new_abb)
                                        nr_sent = rsent.replace(member_prop[large][0],new_abb)
                                        newos.append([rsent,nr_sent,anc3,"",""])
                                        newos.append([nr_sent,qsent,anc4,"",""])
                                if temp_id != []:
                                    for q in range(len(temp_id)):
                                        dummy = new_id(temp_id[q][0],temp_id[q][1],tot_sent,identities,anc1,anc2,done_os)
                                for q in range(len(slist)):
                                    for r in range(len(llist)):
                                        tempq = slist[q]
                                        tempr = llist[r]
                                        qsent = slist_sent[q]
                                        rsent = llist_sent[r]
                                        if slist_sent[q] in irrel or llist_sent[r] in irrel:
                                            tempq = tempq.replace("~","")
                                            tempr = tempr.replace("~","")
                                            if qsent not in irrel and "~" in qsent:
                                                rsent = rsent[:2] + " ~" + rsent[2:]
                                            elif rsent not in irrel and "~" in rsent:
                                                qsent = qsent[:2] + " ~" + qsent[2:]
                                        if tempq == tempr:
                                            list11 = [qsent,rsent]
                                            list11.sort()
                                            if g == 11:
                                                str5 = new_id(str1,str2,tot_sent,identities,anc1,anc2,done_os)
                                                if str5 != None:
                                                    anc1 = str5
                                                else:
                                                    anc1 = sn
                                                new_id3.append([list11[0],list11[1],anc1])
                                            else:
                                                if list11 not in newos3:
                                                    newos3.append(list11)
                                                    if not new_var_used:
                                                        new_abb = idf_var[0]
                                                        del idf_var[0]
                                                        new_var_used = True
                                                    nq_sent = qsent.replace(member_prop[small][0],new_abb)
                                                    nr_sent = rsent.replace(member_prop[large][0],new_abb)
                                                    if g == 12:
                                                        dummy = one_way(str1,str2,done_os,tot_sent,instant,anc1,anc2,new_abb)
                                                    elif g == 21:
                                                        dummy = one_way(str2,str1,done_os,tot_sent,instant,anc1,anc2,new_abb)
                                                    anc3 = findinlist(new_abb,instant,1,2)
                                                    anc4 = findinlist(new_abb,instant,0,2)
                                                    newos.append([rsent,nr_sent,anc3,"",""])
                                                    newos.append([nr_sent,qsent,anc4,"",""])
                                            if "@" in slist[q]:
                                                bb = 8
                                                list3 = member_prop[small][6]
                                                str3 = findinlist(slist[q],list3,0,1)
                                                list6 = member_prop[large][6]
                                                str4 = findinlist(llist[r],list6,0,1)
                                                if g == 11:
                                                    dummy = new_id(str3,str4,tot_sent,identities,member_prop[small][7],member_prop[large][7],done_os)
                                                else:
                                                    print "problem"

                            newid = []
                            temp_id = []
                        else:
                            newid = []
                            temp_id = []
                            tpos = []


    new_id3 = remove_duplicates2d(new_id3,0,1)
    if potential2 != []:
        list4 = []
        for i in range(len(potential2)):
            if potential2[i][3] == 'co' or potential2[i][3] == 'ci':
                if potential2[i][3] == 'ci':
                    print 'new identity'
                list1 = potential2[i][0]
                list4.append([list1[0],idf_var[0]])
                list3 = potential2[i][4]
                small = potential2[i][4][0]
                large = potential2[i][4][1]
                str1 = potential2[i][0][0]
                str2 = potential2[i][0][1]
                j = -1
                #this is different from the above where the second dimension is 3 instead of 5
                slist_sent = list5[small][5]
                llist_sent = list5[large][5]
                slist = member_prop[small][2] + member_prop[small][4]
                llist = member_prop[large][2] + member_prop[large][4]
                irrel += member_prop[small][5] + member_prop[large][5]
                while j < len(newos) -1:
                    j += 1
                    if newos[j][4] == list3:
                        ostr1 = newos[j][1]
                        nstr1 = ostr1.replace(str2,idf_var[0])
                        newos[j][1] = nstr1
                        newos[j+1][0] = nstr1
                        dummy = one_way(str1,str2,done_os,tot_sent,instant,"","",idf_var[0])
                        new_abb = idf_var[0]
                        del idf_var[0]
                        sostr = newos[j][0]
                        lostr = newos[j+1][1]
                        j += 1
                        dummy = special_remove(sostr,slist_sent,slist)
                        dummy = special_remove(lostr,llist_sent,llist)

                for q in range(len(slist)):
                    for r in range(len(llist)):
                        tempq = slist[q]
                        tempr = llist[r]
                        qsent = slist_sent[q][0]
                        rsent = llist_sent[r][0]
                        if slist_sent[q][0] in irrel or llist_sent[r][0] in irrel:
                            tempq = tempq.replace("~","")
                            tempr = tempr.replace("~","")
                            if qsent not in irrel and "~" in qsent:
                                rsent = rsent[:2] + " ~" + rsent[2:]
                            elif rsent not in irrel and "~" in rsent:
                                qsent = qsent[:2] + " ~" + qsent[2:]
                        if tempq == tempr:
                            list11 = [qsent,rsent]
                            list11.sort()
                            if g == 11:
                                str5 = new_id(str1,str2,tot_sent,identities,anc1,anc2,done_os)
                                if str5 != None:
                                    anc1 = str5
                                else:
                                    anc1 = sn
                                new_id3.append([list11[0],list11[1],anc1])
                                # newid.append(list11)
                            else:
                                if list11 not in newos3:
                                    newos3.append(list11)
                                    new_abb = idf_var[0]
                                    del idf_var[0]
                                    if g == 12:
                                        dummy = one_way(str1,str2,done_os,tot_sent,instant,anc1,anc2,new_abb)
                                    elif g == 21:
                                        dummy = one_way(str2,str1,done_os,tot_sent,instant,anc1,anc2,new_abb)
                                    nq_sent = qsent.replace(member_prop[small][0],new_abb)
                                    nr_sent = rsent.replace(member_prop[large][0],new_abb)
                                    newos.append([rsent,nr_sent,sn+1])
                                    newos.append([nr_sent,qsent,sn+2])


    for i in range(len(new_id3)):
        sent1p = name_sent(new_id3[i][0])
        sent2p = name_sent(new_id3[i][1])
        dummy = new_sentence2(new_id3[i][0],sent1p,new_id3[i][1],sent2p,tot_sent,"SUB",new_id3[i][2])

    for i in range(len(newos)):
        sent1p = name_sent(newos[i][0])
        sent2p = name_sent(newos[i][1])
        dummy = new_sentence2(newos[i][0],sent1p,newos[i][1],sent2p,tot_sent,"SUB",newos[i][2],conditional)

    bb = 8
    acc_rel = []
    bool1 = False
    for i in range(len(all_sent)):
        if all_sent[i][9] in undefined and all_sent[i][9] not in atomic_rel:
            if all_sent[i][8] == "~" and not bool1:
                bool1 = True
            acc_rel.append([all_sent[i][9],all_sent[i]])
    acc_rel.sort()

    list8 = []
    k = 0
    if len(acc_rel) > 1 and bool1:
        bb = 8

        for i in range(1,len(acc_rel)):
            if i == 14:
                bb = 8
            if acc_rel[i][0] != acc_rel[i-1][0]:
                if i - k > 1:
                    list8.append(acc_rel[i][1])
                    bool1 = False
                k = i
        if i - k > 0:
            list8.append(acc_rel[i][1])
            if i == 1:
                list8.append(acc_rel[0][1])

        num = [5,14,18,22]
        if list8 != [] and instant != []:
            for i in range(len(list8)):
                if not list8[i][5] in gen_var or not list8[i][14] in gen_var:
                    j = -1
                    while j < (len(instant))-2:
                        j += 1
                        k = -1
                        while k < 4:
                            k += 1
                            m = num[k]
                            if list8[i][m] == None and m == 18:
                                break
                            if list8[i][m] == instant[i][0]:
                                if m == 5:
                                    bool1 = False
                                    obj = list8[i][14]
                                    anc2 = ""
                                    anc4 = ""
                                    objects = []
                                    for p in range(len(instant)):
                                        if obj == instant[p][0]:
                                            bool1 = True
                                            objects.append(p)
                                    if objects != []:
                                        for q in objects:
                                            anc2 = instant[q][2]
                                            anc4 = instant[q+1][2]
                                            anc1 = instant[j][2]
                                            anc3 = instant[j+1][2]
                                            list1 = copy.deepcopy(list8[i])
                                            list2 = copy.deepcopy(list8[i])
                                            list1[m] = instant[j][1]
                                            list2[m] = instant[j+1][1]
                                            list1[14] = instant[q][1]
                                            list2[14] = instant[q+1][1]
                                            old_sent = list8[i][0]
                                            oldp = list8[i][42]
                                            new_sent = build_sent(list1)
                                            newp = name_sent(new_sent)
                                            new_sent2 = build_sent(list2)
                                            newp2 = name_sent(new_sent2)
                                            bool2 = isrel(newp2,all_sent)
                                            if bool2:
                                                # print 'accidental relation inferred'
                                                dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"SUB",anc1,conditional,anc2)
                                                dummy = new_sentence2(new_sent,newp,new_sent2,newp2,tot_sent,"SUB",anc3,conditional,anc4)
                                                j + 1
                                                if m == 5:
                                                    k + 1
                                    else:
                                        pass

    bb = 8
    en = time.time()
    # print en-st
    #end5

def isrel(str1,all_sent):
    #if the sentence has not already appeared since we are dealing with nonatoms, then
    #we do not need to infer it
    str1 = str1.replace("~","")
    for i in range(len(all_sent)):
        str2 = all_sent[i][42]
        str2 = str2.replace("~","")
        if str2 == str1:
            return True
    return False


def equality(member_prop,small,large,q,r,list5,potential,potential2,genre):

    num = [5,14,15,18,19,22]
    svar = member_prop[small][0]
    lvar = member_prop[large][0]
    if svar == 'r' and lvar == 'v':
        bb = 8
    list2 = []
    global gen_var, definite2,ind_var
    status = 'y'
    if member_prop[small][2][q] == member_prop[large][2][r]:
        return ['y',[]]
    else:
        sparts = list5[small][5]
        lparts = list5[large][5]
        ssent = member_prop[small][3][q]
        lsent = member_prop[large][3][r]
        d = findposinlist(ssent,sparts,0)
        e = findposinlist(lsent,lparts,0)
        smem = sparts[d]
        lmem = lparts[e]
        list3 = [svar,lvar]
        list3.sort()
        potential2_used = []
        if svar == smem[5]:
            b = 14
            c = 5
        else:
            b = 5
            c = 14

        if smem[8] + smem[9] != lmem[8] + lmem[9]:
            return ['n',[]]
        elif smem[8] + smem[9] == lmem[8] + lmem[9] and lmem[9] != "=" and (smem[b] == lmem[c] or smem[c] == lmem[b]):
            #this is for sentences of the form cAFd & dAFe
            return ['n',[]]
        else:
            for i in num:
                list1 = [smem[i],lmem[i]]
                list1.sort()
                if smem[i] == None and lmem[i] == None:
                    break
                elif smem[i] == lmem[i] or (smem[i] == svar and lmem[i] == lvar):
                    pass
                elif (smem[i] in gen_var or lmem[i] in gen_var) and list1 in potential:
                    status = 'm'
                    list2.append(list1)
                else:
                    if potential2 != []:
                        bool1 = False
                        for y in range(len(potential2)):
                            if bool1:
                                break
                            for z in range(len(potential2[y][1])):
                                if list3 == potential2[y][1][z]:
                                    list4 = potential2[y][0]
                                    if list4 == list1:
                                        status = 'y'
                                        bool1 = True
                                        potential2_used.append(y)
                                        list2 = potential2[y][4]
                                        if list2 == ['q','w']:
                                            bb = 8
                                        break
                        else:
                            if not bool1:
                                status = 'n'
                            else:
                                status = 'u'

                    else:
                        return ["n",[]]


            for k in range(len(potential2_used)):
                if genre == 11:
                    potential2[k][3] = 'ci'
                else:
                    potential2[k][3] = 'co'

            return [status,list2]


def special_remove(str1,list1,list2):

    str1 = str1.replace(" ","")
    for i in range(len(list1)):
        str2 = list1[i][0]
        str2 = str2.replace(" ","")
        if str2 == str1:
            del list1[i]
            del list2[i]
            return

def already_identical(identities,str1,str2):

    for i in range(len(identities)):
        if str1 in identities[i][0] and str2 in identities[i][0]:
            return True
    return False

def arentident(jobj, iobj,non_id):

    for i in range(len(non_id)):
        if (jobj in non_id[i][0] and iobj in non_id[i][0]):
            return True
    return False

def new_sentence(tot_sent,  old_list, list1, list2, list3, quant, rule,conn = iff,anc1 = "",anc2 = ""):

    global prop_name,psent,sn
    if old_list[0] == None:
        old_sent = build_sent(old_list)
        old_prop = findinlist(old_sent2, prop_name, 1,0)
    else:
        old_sent = old_list[0]
        old_prop = old_list[42]

    if list1[0] == None:
        str1 = build_sent(list1)
    else:
        str1 = list1[0]
    str1v = name_sent(str1)
    list1[0] = str1
    list1[42] = str1v
    if quant == 2:
        str2 = build_sent(list2)
        str2v = name_sent(str2)
    if list3 != "":
        str3 = build_sent(list3)
        str3v = name_sent(str3)
        list3[0] = str3
        easygui.msgbox('you have not coded for three new sentences yet')
    if quant == 1:
        str1 = old_sent + ' ' + conn + ' ' + str1
        str1v = old_prop + ' ' + conn + ' ' + str1v
    elif quant == 2:
        str1 = '(' + old_sent + ' & ' + str2 + ') ' + conditional + ' ' + str1
        str1v = '(' + old_prop + ' & ' + str2v + ') ' + conditional + ' ' + str1v
    elif quant == 3:
        str1 = '(' + old_sent + ' & ' + str2 + " & " + str3 + ') ' + conditional + ' ' + str1
        str1v = '(' + old_prop + ' & ' + str2v + "& " + str3v + ') ' + conditional + ' ' + str1v

    g = findinlist(str1,tot_sent,2,0,True)
    if g == -1 and quant != 3:
        sn += 1
        tot_sent.append([sn, str1, str1v, "", rule, anc1, anc2])




def check_dimension(list1, i, str1,bool1 = False,k=0):

    if not bool1:
        for j in range(len(list1)):
            if list1[j][i] == str1:
                return True
        return False
    else:
        for j in range(len(list1)):
            if j != k:
                if list1[j][i] == str1:
                    return True
        return False

#################################################

## The following functions are for statement logic


def name_conditional(list1,nat_logic = True):

    # when using this function the list1 must be outputted from find_sentences
    # and you must use the 0th element since those sentences do not have negation signs
    global prop_var
    global prop_name
    skel_string = list1[5]
    list2 = []

    for i in range(1, len(list1[6])):
        if list1[6][i] != None and list1[6][i] != "":
            str3 = list1[0][i]
            str3 = remove_outer_paren(str3)
            str4 = copy.copy(str3)
            str4 = str4.replace(" ","")
            str1 = findinlist(str4,prop_name,1,0)
            if str1 != None:
                str2 = list1[1][i] + str1
                list2.append(str1)
                skel_string = skel_string.replace(list1[6][i][1], str2)
            else:
                str2 = prop_var[0]
                list2.append(str2)
                del prop_var[0]
                prop_name.append([str2, str4,str3])
                str2 = list1[1][i] + str2
                skel_string = skel_string.replace(list1[6][i][1], str2)
    return [skel_string, list2]

def tilde_removal(str1):


    str4 = ""
    if str1.find("~") > -1:
        str4 = "~"
        str1 = str1.replace("~","")
    return [str1,str4]

def tilde_removal2(str1):

    j = 0
    if str1[:2] == "~(":
        for i in range(len(str1)):
            str2 = str1[i:i+1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i > 0:
                if i == len(str1) -1:
                    str1 = str1[1:]
                    str4 = "~"
                    return [str1, str4]
                else:
                    return [str1,""]
    elif str1[0] == "~" and str1[1].islower() and os(str1):
        str1 = str1[1:]
        str4 = "~"
    else:
        str4 = ""
    return [str1,str4]

def simple_sent_name(str1):

    str1 = remove_outer_paren(str1)
    str2 = findinlist(str1,prop_name,1,0)
    if str2 == None:
        ostring = str1.replace(" ","")
        str2 = prop_var[0]
        del prop_var[0]
    prop_name.append([str2, ostring,str1])
    return str2

def is_conjunction(str1):

    if str1.count(idisj) == 0 and str1.count(iff) == 0 and \
        str1.count(conditional) == 0 and str1.count('&') > 0 and str1.count(xorr) == 0:
        return True
    else:
        return False


def get_conjuncts (str1, bool1 = False):
    # remove outparent if true
    global subscripts
    arr1 = []
    if bool1:
        str1 = remove_outer_paren(str1)

    j = 0
    k = 1
    for i in range (len(str1)):
        str2 = str1[i:i+1]
        str5 = str1[i+1:i+2]
        if i > 0:
            str4 = str1[i-1:i]
        else:
            str4 = ""
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 1 and str2 == "(" and str4 != "~":
            k = i
        elif j == 1 and str4 == "~" and str2 == "(":
            k = i - 1

        if (j==0 and str2 == ")") or (j == 0 and str2.islower()):
            if str2 == ")":
                str3 = str1[k:i+1]
            else:
                if str1[i-1:i] == "~" and str5 not in subscripts:
                    str3 = "~" + str2
                elif str1[i-1:i] != "~" and str5 not in subscripts:
                    str3 = str2
                elif str1[i-1:i] != "~" and str5 in subscripts:
                    str3 = str2 + str5
                elif str1[i-1:i] == "~" and str5 in subscripts:
                    str3 = "~" + str2 + str5
            k = 1
            str3 = str3.strip()
            arr1.append(str3)

    return arr1


def prepare_iff_elim(str2, mainc, s, num = ""):

    global sn
    list7 = [""] * 39
    if num == "":
        list7[2] = sn + 1
    else:
        list7[2] =  num
    list7[4] = str2
    list7[5] = ""
    j = 0
    str2 = remove_outer_paren(str2)
    if mainc == iff:
        list7[3] = "e"
    else:
        list7[3] = "c"

    str8 = str2[: s - 1]
    str8 = str8.strip()
    str9 = str2[s+1:]
    str9 = str9.strip()
    bool1 = True
    bool2 = True
    list8 = mainconn(str8)
    list2 = tilde_removal2(str8)
    if list8[0] != "&" or (list2[1] == "~" and list8[0] != ""):
        list2[0] = remove_outer_paren(list2[0])
        list7[0] = [list2[0],list2[1]]
        bool1 = False
    list8 = mainconn(str9)
    list2 = tilde_removal2(str9)
    if list8[0] != "&" or (list2[1] == "~" and list8[0] != ""):
        list2[0] = remove_outer_paren(list2[0])
        list7[1] = [list2[0],list2[1]]
        bool2 = False
    list4 = [bool1, bool2]
    for j in range(0, 2):
        if list4[j]:
            if j == 0:
                str7 = str8
            else:
                str7 = str9
            list2 = mainconn(str7)
            if list2[0] == '&':
                list3 = get_conjuncts(str7, True)
                list6 = []
                for k in range(len(list3)):
                    list5 = tilde_removal2(list3[k])
                    list5[0] = remove_outer_paren(list5[0])
                    list6.append([list5[0],list5[1]])
                if j == 0:
                    list7[0] = list6
                    list7[6] = str8
                else:
                    list7[1] = list6
                    list7[7] = str9

    return list7

def islist(list1):

    if list1[1] == '~' or list1[1] == "":
        return False
    else:
        return True

def new_prop(prop_sent, str1, ng, asp, anc1, anc2, anc3 = None, anc4 = None, \
             is_premise = False, num = "",ostring = ""):

    if ng == None:
        ng = ""
    global sn
    list1 = [None] * 15

    if ng == "":
        str1 = remove_outer_paren(str1)
    str2 = findinlist(str1,prop_sent,1,2)
    if str2 == ng:
        return True
    elif str2 == None:
        if is_premise:
            # sn = num
            prop_sent.append([num, str1, ng, "", "", "", None, None, ostring, None, None, \
            None, None, None, None])
        else:
            sn += 1
            prop_sent.append([sn, str1, ng, asp, anc1, anc2, None, None, None, None, None, \
            None, None, None, None])
        return True
    elif str2 != ng:
        sn += 1
        prop_sent.append([sn, str1, ng, asp, anc1, anc2, None, None, None, None, None, \
        None, None, None, None])
        anc2 = findinlist(str1,prop_sent,1,0)
        sn += 1
        if not os(str1):
            str1 = "(" + str1 + ")"
        str1 = str1 + " & " + "~" + str1
        prop_sent.append([sn, str1, "", "&I", sn-1, anc2, None, None, None, None, None, \
        None, None, None, None])
        str1 = bottom
        sn += 1
        prop_sent.append([sn, str1, "", bottom + "I", sn-1, None, None, None, None, None, None, \
        None, None, None, None])
        return False

def many_cond(candd, conditionals, kind, asp, anc2, f, g, r):

    list1 = conditionals[g][f]
    del list1[0]
    # this list allows us to keep track of the ancestors even if we don't detach a
    # a part of the conditionals on the first time
    if conditionals[g][8] == "":
        conditionals[g][8] = [candd[r][0]]
    else:
        conditionals[g][8].append(candd[r][0])

    if f == 1:
        h = 7
    else:
        h = 6
    list2 = []
    list2.append([candd[r][1],candd[r][2]])
    if list1 == []:
        cjct = conditionals[g][h]
        dummy = new_prop_sent("", kind, asp, "",anc2, \
                    conditionals,g, conditionals[g][8], cjct)
        if dummy == False:
            return False
        else:
            return True
    j = -1
    while j < len(list1)-1:
        j += 1
        str2 = list1[j][0]
        neg2 = list1[j][1]
        for i in range(len(candd)):
            if i != r:
                str1 = candd[i][1]
                ng = candd[i][2]
                anc1 = candd[i][0]
                if str1 == str2 and ng == neg2:
                    del list1[j]
                    j -= 1
                    list2.append([str1, ng])
                    conditionals[g][8].append(anc1)
                    if list1 == []:
                        cjct = conditionals[g][h]
                        dummy = new_prop_sent("", kind, asp, "",anc2, \
                                    conditionals,g, conditionals[g][8], cjct)
                        if dummy == False:
                            return False
                        else:
                            return True
                    else:
                        conditionals[g][8] == None
                        break
                elif str1 == str2 and ng != neg2:
            # the point of having blank returns is because if it returns true
            # then we need to subtract the conditional counter, here g, by 1
                    conditionals[g][8] = ""
                    return ""
    conditionals[g][f] = list1
    conditionals[g][8] = ""
    return ""

def modus_ponens(conditionals, candd, prop_sent):

    global sn
    r = -1
    while r < len(candd) -1:
        if conditionals == []:
            return True
        r += 1
        # if r == 3:
        #     bb = 7
        str1 = candd[r][1]
        str8 = candd[r][2]
        anc1 = candd[r][0]
        temp1 = copy.copy(str1)
        temp1 = temp1.replace(" ","")
        temp1 = remove_outer_paren(temp1)
        g = -1
        bool1 = False
        while g < len(conditionals) -1:
            g += 1
            if g == 2 and r==16:
                bb = 7
            if conditionals[g][0] == "":
                break
            str12 = conditionals[g][3]
            if str12 != 'd':
                anc2 = conditionals[g][2]
                if conditionals[g][6] == "":
                    aconjunction = ""
                    temp_ant = conditionals[g][0][0]
                    temp_nega = conditionals[g][0][1]
                else:
                    aconjunction = conditionals[g][6]
                    temp_ant = conditionals[g][0][0][0]
                    temp_nega = conditionals[g][0][0][1]
                if conditionals[g][7] == "":
                    cconjunction = ""
                    temp_con = conditionals[g][1][0]
                    temp_negc = conditionals[g][1][1]
                else:
                    cconjunction = conditionals[g][7]
                    temp_con = conditionals[g][1][0][0]
                    temp_negc = conditionals[g][1][0][1]
                temp_ant = temp_ant.replace(" ","")
                temp_con = temp_con.replace(" ","")
                temp_ant = remove_outer_paren(temp_ant)
                temp_con = remove_outer_paren(temp_con)

                for f in range(0,2):
                    if f == 0 and temp1 == temp_ant:
                        if str8 == temp_nega:
                            if str12 == 'c':
                                str13 = "MP"
                            else:
                                str13 = "EE"
                            if aconjunction != "" :
                                dummy = many_cond(candd, conditionals, "con", str13, \
                                                  anc2, f, g, r)
                                if dummy == False:
                                    return False
                                elif dummy:
                                    del conditionals[g]
                                    g -= 1
                                    break
                            else:
                 # con indicates that the consequent of the conditional is to be detached
                 #                if len(conditionals) <= 4:
                 #                    bb = 7
                                dummy = new_prop_sent("", "con", \
                                    str13, anc1, anc2,conditionals,g)
                                if dummy == False:
                                    return False
                                del conditionals[g]
                                g -= 1
                                break
                        elif str8 != temp_nega and str12 == 'e':
                            dummy = new_prop_sent("~", "con", \
                                        "EN", anc1, anc2, conditionals,g)
                            if dummy == False:
                                return False
                            del conditionals[g]
                            g -= 1
                            break
                    elif f == 1 and temp1 == temp_con:
                        if str8 == temp_negc and str12 == 'e':
                            if cconjunction == "":
                                dummy = new_prop_sent("", "ant", "EE", \
                                    anc1, anc2, conditionals,g)
                                if dummy == False:
                                    return False
                                del conditionals[g]
                                g -= 1
                                break
                            else:
                                dummy = many_cond(candd, conditionals, "ant", "EE", \
                                                  anc2, f, g, r)
                                if dummy == False:
                                    return False
                                elif dummy:
                                    del conditionals[g]
                                    g -= 1
                                    break
                        elif str8 != temp_negc:
                            if str12 == 'c':
                                str13 = "MT"
                            else:
                                str13 = "EN"
                            dummy = new_prop_sent("~", "ant", \
                                        str13, anc1, anc2, conditionals,g)
                            if dummy == False:
                                return False
                            del conditionals[g]
                            g -= 1
                            break
                    elif f == 0 and temp1 != temp_ant and \
                                    aconjunction != "" and str12 == 'e':
                        s = -1
                        if conditionals != []:
                            while s < len(conditionals[g][0]) -1:
                                s += 1
                                if temp1 == conditionals[g][0][s][0] and \
                                    str8 != conditionals[g][0][s][1]:
                                    dummy = new_prop_sent("~", "con", \
                                    "EN", anc1, anc2, conditionals,g)
                                    if dummy == False:
                                        return False
                                    del conditionals[g]
                                    g -= 1
                                    break

                    elif f == 1 and temp1 != temp_con and cconjunction != "":
                        s = -1
                        if conditionals != []:
                            while s < len(conditionals[g][1]) -1:
                                s += 1
                                if temp1 == conditionals[g][1][s][0] and \
                                    str8 != conditionals[g][1][s][1]:
                                    if str12 == 'e':
                                        str13 = "EN"
                                    else:
                                        str13 = "MT"
                                    dummy = new_prop_sent("~", "ant", \
                                    str13, anc1, anc2, conditionals,g)
                                    if dummy == False:
                                        return False
                                    del conditionals[g]
                                    g -= 1
                                    break
    return True

def disjunction_heirarchy(conditionals, str5,d,new_disj = False):

    global prop_name
    global sn

    str5 = enclose(str5)
    list1 = find_sentences(str5, False)
    mainc = list1[4][0][1]

    for i in range(len(list1[0])):
        list2 = unenclose(list1[0][i])
        list1[0][i] = list2[0]

    list2 = [""] * 39
    n = 7
    if mainc == xorr:
        list2[3] = 'x'
    else:
         list2[3] = 'd'
    if conditionals == [] or new_disj:
        list2[2] = sn
    else:
        list2[2] = conditionals[d][2]
    list2[5] = ""
    list2[4] = list1[0][0]# fix this
    sentences = []

    for i in range(len(list1[0])):
        if os(list1[0][i]):
            siblings = []
            list3 = [None] * 9
            n += 1
            # str1 = findinlist(list1[0][i],prop_name,1,0)
            str2 = list1[4][i][0][:-1]
            g = findinlist(str2,list1[4],0,1,True)
            parent = list1[0][g]
            if list1[4][g][1] == "&":
                list3[2] = 'c'
            elif list1[4][g][1] == xorr:
                list3[2] = 'x'
            else:
                list3[2] = 'd'
            if len(str2) > 1:
                str3 = list1[4][i][0][:-2]
                g = findinlist(str3,list1[4],0,1,True)
                gparent = list1[0][g]
            else:
                gparent = parent
            list3[1] = list1[4][i][0]
            list3[5] = parent
            list3[6] = gparent
            list3[0] = [list1[0][i], list1[1][i]]
            #fix this
            b = parent.count(xorr)
            c = parent.count(idisj)
            if c > 1 or b > 1:
                list3[7] = 2
            else:
                list3[7] = 1
            sent_num = list1[4][i][0]
            m = len(sent_num)
            for j in range(len(list1[4])):
                if len(list1[4][j][0]) == m and list1[4][j][0][:-1] == str2 \
                    and j != i:
                    siblings.append([list1[0][j],list1[1][j]]) # fix this
            list3[4] = siblings
            list2[n] = list3
            sentences.append(list3[0][0])
            if list3[0][0] not in rel_conj:
                rel_conj.append(list3[0][0])

    if conditionals == [] or new_disj:
        list2[38] = sentences
        conditionals.append(list2)
    else:
        list2[38] = sentences
        conditionals[d] = list2

def proper_spacing(str1):

    str1 = str1.replace(" ","")
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional," " + conditional + " ")
    str1 = str1.replace(idisj," " + idisj + " ")
    str1 = str1.replace(xorr," " + xorr + " ")
    str1 = str1.replace("&"," & ")
    return str1

def iff_elim(prop_sent, conditionals,kind):

    new_sent = False
    no_contr = True
    for d in range(len(conditionals)):
        if conditionals[d][0] != "":
            list1 = ["",""]
            ng = conditionals[d][5]
            if conditionals[d][3] == "d" or conditionals[d][3] == 'n':
                list1 = [conditionals[d][4], ""]
            else:
                if conditionals[d][6] == "":
                    list1[0] = conditionals[d][0][0]
                else:
                    list1[0] = conditionals[d][6]
                if conditionals[d][7] == "":
                    list1[1] = conditionals[d][1][0]
                else:
                    list1[1] = conditionals[d][7]
            for s in range(0,2):
                if list1[s].find(iff) > -1:
                    new_sent = True
                    str1 = list1[s]
                    anc1 = conditionals[d][2]
                    old_str = copy.copy(str1)
                    str1 = str1.replace(" ","")
                    str1 = str1.replace("&"," & ")
                    str1 = str1.replace(iff , " " + iff + " ")
                    str1 = remove_outer_paren(str1)
                    list5 = mainconn(str1)
                    mc = list5[0]
                    bool4 = False
                    if str1[:2] == "~(":
                        bool4 = True
                        str1 = str1[2:len(str1) - 1]
                    r = str1.count(iff)
                    i = -1
                    for o in range(0,r):
                        k = 0
                        while i < len(str1) - 1:
                            i += 1
                            str2 = str1[i:i+1]
                            if str2 == iff:
                                p = copy.copy(i)
                                t = copy.copy(i)
                                bool5 = False
                                while k != -1:
                                    i -= 1
                                    str3 = str1[i:i+1]
                                    if str3 == "(":
                                        k -= 1
                                    elif str3 == ")":
                                        k += 1
                                        bool5 = True
                                    if k == -1:
                                        m = i + 1
                                    elif k == 0 and bool5:
                                        m = i
                                        break
                                bool3 = False
                                k = 0
                                while p < len(str1):
                                    p += 1
                                    str3 = str1[p:p+1]
                                    if str3 == "(":
                                        bool3 = True
                                        k -= 1
                                    elif str3 == ")":
                                        k += 1
                                    if bool3 and k == 0:
                                        p += 1
                                        break
                                    elif k == 1:
                                        break
                                ant = str1[m:t]
                                ant = ant.strip()
                                con = str1[t+1:p]
                                con = con.strip()
                                new1 = "(" + ant + " " + conditional + " " + con + ")"
                                new2 = "(" + con + " " + conditional + " " + \
                                       ant + ")"
                                if mc != "&":
                                    if r > 1 and o + 1 != r:
                                        new3 = "(" + new1 + ' & ' + new2 + ")*"
                                    else:
                                        new3 = "(" + new1 + ' & ' + new2 + ")"
                                else:
                                    if r > 1 and o + 1 != r:
                                        new3 = new1 + ' & ' + new2 + '*'
                                    else:
                                        new3 = new1 + ' & ' + new2
                                if m != 0 or p != len(str1):
                                    replacee = str1[m-1:p+1]
                                    new3 = str1.replace(replacee, new3)
                                str1 = new3
                                if r > 1 and o + 1 != r:
                                    i = str1.find("*")
                                    str1 = str1.replace("*","")
                                break
                    new3 = str1
                    new3 = proper_spacing(new3)
                    if bool4:
                        new3 = "~(" + new3 + ")"
                    else:
                        new3 = "(" + new3 + ")"
                    conditionals[d][4] = conditionals[d][4].replace(old_str,new3)
                    if conditionals[d][3] == 'c' or conditionals[d][3] == 'e':
                        if s == 0:
                            conditionals[d][0] = None
                            conditionals[d][6] = new3
                        else:
                            conditionals[d][1] = None
                            conditionals[d][7] = new3

            if conditionals[d][3] == "e":
                conditionals[d][3] = "c"
                anc1 = conditionals[d][2]
                list1 = [""] * 39
                list1[5] = ""
                if conditionals[d][6] == "":
                    str1 = conditionals[d][0][0]
                    if not os(str1):
                        str1 = "(" + str1 + ")"
                    ng1 = conditionals[d][0][1]
                else:
                    str1 = conditionals[d][6]
                    ng1 = ""
                if conditionals[d][7] == "":
                    str4 = conditionals[d][1][0]
                    if not os(str4):
                        str4 = "(" + str4 + ")"
                    ng4 = conditionals[d][1][1]
                else:
                    str4 = conditionals[d][7]
                    ng4 = ""

                list1[0] = [str4,ng4]
                list1[1] = [str1, ng1]
                str5 =  ng1 + str1 + " " + conditional + " " + ng4 + str4
                str6 = ng4 + str4 + " " + conditional + " " + ng1 + str1
                g = copy.copy(sn+1)
                if conditionals[d][5] == "" or conditionals[d][5] == None:
                    str3 = "(" + str5 + ") & (" + str6 + ")"
                    str7 = iff + "E"
                    no_contr = new_prop(prop_sent, str3, ng, str7, anc1, None, None, None)

                    if not no_contr:
                        return False
                    conditionals[d][2] = sn+1
                    conditionals[d][0] = [str1, ng1]
                    conditionals[d][1] = [str4, ng4]
                    no_contr = new_prop(prop_sent,str5,"","&E",g,"")
                    if not no_contr:
                        return False
                    list1[2] = sn + 1
                    no_contr = new_prop(prop_sent,str6,"","&E",g,"")
                    if not no_contr:
                        return False
                    conditionals[d][4] = str5
                    list1[4] = str6
                    conditionals.append(list1)
                else:
                    str3 = "(" + str5 + ") & (" + str6 + ")"
                    no_contr = new_prop(prop_sent,str3,"~",iff + "E",g,"")
                    if not no_contr:
                        return False
                    conditionals[d][3] = 'd'
                    conditionals[d][4] = str3
            elif new_sent:
                no_contr = new_prop(prop_sent,conditionals[d][4],"",iff + "E",anc1,"")
                if not no_contr:
                    return False
    return True

def material_implication(prop_sent, conditionals,kind):

    for d in range(len(conditionals)):
        if d == 3:
            bb = 7
        if conditionals[d][0] != "":
            str1 = conditionals[d][4]
            ng = conditionals[d][5]
            if str1.find(conditional) > -1:
                anc1 = conditionals[d][2]
                i = -1
                q = -1
                s = 0
                r = str1.count(conditional)
                while s < r:
                    i += 1
                    q += 1
                    str2 = str1[i:i+1]
                    if str2 == conditional:
                        s += 1
                        j = i
                        k = 0
                        m = 0
                        bool1 = False
                        while j != 0:
                            j -= 1
                            m += 1
                            if m > 200:
                                easygui.msgbox("in the material implication function \
                                you are caught in an infinite loop")
                                return
                            str3= str1[j:j+1]
                            if str3.islower() and k == 0:
                                bool1 = True
                            elif str3 == ")":
                                bool1 = True
                                k += 1
                            elif j == 0 and str3 != "(":
                                k = 0
                                bool1 = True
                            elif str3 == "(":
                                bool1 = True
                                k -= 1

                            if bool1 and k == 0:
                                str4 = str1[0:j]
                                str5 = str1[i+2:]
                                str7 = str1[j:i]
                                str1 = str4 + "~" + str7 + idisj + " " + str5
                                i += 1
                                break

                str1 = bad_paren(str1)
                no_contr = new_prop(prop_sent, str1, ng, conditional + "E", anc1,"")
                if not no_contr:
                    return False
                if str1.find("~~") > -1:
                    str1 = str1.replace("~~","")
                    g = copy.copy(sn)
                    no_contr = new_prop(prop_sent,str1, ng, "~~E", g,"")
                    if not no_contr:
                        return False
                conditionals[d][2] = sn
                conditionals[d][4] = str1
                conditionals[d][3] = "d"
                conditionals[d][5] = ng

    return True

            # search for double negatives

def bad_paren(str1):

    if str1.find("(") == -1:
        return str1
    # we first must get rid of strings of the following form ((p) & s)
    for i in range(len(str1)):
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        if i > 1:
            str4 = str1[i-2:i-1]
        else:
            str4 = ""
        str5 = str1[i+1:i+2]
        if str2.islower() and str3 == "(" and str5 == ")":
            str1 = str1[:i-1] + str1[i:i+1] + str1[i+2:]
        elif str2.islower() and str4 == "(" and str3 == "~" and str5 == ")":
            str1 = str1[:i-2] + str1[i-1:i+1] + str1[i+2:]

    str1 = enclose(str1)
    list1 = find_sentences(str1, False)
    for i in range(len(list1[3])):
        list2 = unenclose(list1[3][i])
        list1[3][i] = list2[0]
    mstr = list1[3][0]
    for i in range(1, len(list1[3])):
        if list1[4][i][1] != "":
            mc = list1[4][i][1]
            ostr = list1[3][i]
            str2 = list1[4][i][0][:-1]
            prcnt = findinlist(str2,list1[4],0,1,False)
            if mc == prcnt:
                nstr = remove_outer_paren(ostr)
                nstr = remove_outer_paren(nstr)
                mstr = mstr.replace(ostr, nstr)
    return mstr

def demorgan(prop_sent, conditionals, candd,kind,one_sent = False, str8 = "",anc1a = "",rule = ""):

    d = -1
    temp_bool = True
    rop = False
    if one_sent:
        d = len(conditionals) - 2

    while d < len(conditionals) -1:
        d += 1
        if one_sent:
            str1 = str8
        else:
            str1 = conditionals[d][5] + conditionals[d][4]
        if one_sent:
            rop = True
        else:
            if conditionals[d][5] == "~":
                rop = True
        if str1.find("~(") > -1:
            if one_sent:
                anc1 = anc1a
            else:
                anc1 = conditionals[d][2]
            r = str1.count("~(")
            s = 0
            i = -1
            while s < r:
                i += 1
                if i > 200:
                    easygui.msgbox("you are caught in an infinite loop in the \
                    demorgan function")
                    return
                bool2 = False
                str2 = str1[i:i+2]
                if str2 == "~(":
                    s += 1
                    bool2 = True
                    if i > 0:
                        str4 = str1[0:i]
                    else:
                        str4 = ""
                    str5 = str1[i+1:len(str1)+1]
                    str1 = str4 + str5
                    j = i
                    m = 0
                    k = 0
                    bool3 = False
                    while j < len(str1):
                        m += 1
                        if m > 200:
                            easygui.msgbox("you are caught in an infinite loop in the \
                            demorgan function")
                            return
                        bool1 = False
                        str3 = str1[j:j+1]
                        if j-1 > 0:
                            str6 = str1[j-1:j+1]
                        else:
                            str6 = ""

                        if str6 == "~(" and bool3 == False and bool2 == False:
                            k = 0
                            k += 1
                            s += 1
                            # bool3 means that python is searching only for those letters and connectives
                            # which are on the same level of parentheses
                            bool3 = True
                            str4 = str1[0:j]
                            str5 = str1[j:]
                            str1 = str4 + "~" + str5
                            j += 2
                            str3 = str1[j:j+1]


                        bool2 = False

                        if str3.islower() and bool3 == False:
                            str4= str1[0:j]
                            str5 = str1[j+1:(len(str1)+1)]
                            str1 = str4 + "~" + str3 + str5
                            j += 2
                            bool1 = True
                        elif bool3 == False:
                            if str3 == "(":
                                k += 1
                                j += 1
                                bool1 = True
                            elif str3 == ")":
                                k -= 1
                                j += 1
                                bool1 = True
                            elif str3 == "&":
                                str4 = str1[0:j]
                                str5 = str1[(j + 1):(len(str1) + 1)]
                                str1 = str4 + idisj+ str5
                                j += 1
                            elif str3 == idisj or str3 == xorr:
                                str4 = str1[0:j]
                                str5 = str1[(j+1):(len(str1)+1)]
                                str1 = str4 + "&" + str5
                                j += 1
                            else:
                                j += 1

                        elif str3 == ")":
                            j += 1
                            k -= 1
                        elif str3 == "(":
                            k += 1
                            j += 1
                        else:
                            j += 1

                        if k == 0 and bool3:
                            bool3 = False
                        if k == 0 and bool1 == True:
                            i = j
                            break

        # rop means remove outer paren, this is necessary if the original sentence was negated
            if rop:
                str1 = remove_outer_paren(str1)
            list1 = [None] * 15
            if str1.find("~~") > -1:
                str2 = copy.copy(str1)
                str2 = str2.replace("~~","")
                str2 = bad_paren(str2)
            else:
                str1 = bad_paren(str1)
            no_contr = new_prop(prop_sent,str1,"","~(E",anc1,"")
            if not no_contr:
                return False
            if str1.find("~~") > -1:
                str1 = str2
                anc1 = copy.copy(sn)
                no_contr = new_prop(prop_sent,str2,"","~~E",anc1,"")
                if not no_contr:
                    return False
            list2 = mainconn(str1)

            if list2[0] == "&":
                if not one_sent:
                    del conditionals[d]
                list3 = get_conjuncts(str1,True)
                anc1 = copy.copy(sn)
                for i in range(len(list3)):
                    list4 = tilde_removal2(list3[i])
                    no_contr = new_prop(prop_sent,list4[0],list4[1],"&E",anc1,"")
                    if not no_contr:
                        return False
                    list2 = mainconn(list3[i])
                    if list2[0] == idisj or list2[0] == xorr:
                        # add in more nones if it turns out that I need them
                        if one_sent:
                            dummy = disjunction_heirarchy(conditionals,str1,0)
                        else:
                            list5 = [""] * 39
                            list5[2] = sn
                            list5[4] = list4[0]
                            list5[5] = list4[1]
                            conditionals.append(list5)
                    else:
                        candd.append([sn,list4[0],list4[1]])
            else:
                if one_sent:
                    dummy = disjunction_heirarchy(conditionals,str1,0,True)
                    return
                else:
                    conditionals[d][2] = sn
                    conditionals[d][4] = str1
    return True

def unenclose(str1):

    i = -1
    global subscripts
    list1 = []
    while i < len(str1)-1:
        i += 1
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        str4 = str1[i+1:i+2]
        if str2.islower() and str3 != "~" and str4 not in subscripts:
            str1 = str1[:i-1] + str2 + str1[i+2:]
            list1.append(str2)
        elif str2.islower() and str3 == "~" and str4 not in subscripts:
            str1 = str1[:i-2] + str3 + str2 + str1[i+2:]
            list1.append(str2)
        if str2.islower() and str3 != "~" and str4 in subscripts:
            str1 = str1[:i-1] + str2 + str4 + str1[i+3:]
            list1.append(str2 + str4)
        elif str2.islower() and str3 == "~" and str4 in subscripts:
            str1 = str1[:i-2] + str3 + str2 + str4 + str1[i+3:]
            list1.append(str2 + str4)
    return [str1,list1]

def new_disjunct(str1, ng, n, prop_sent, conditionals, candd, anc1, anc2, \
            anc3 = None, anc4=None, kind = 0, rule = ""):

    global sn
    list2 = mainconn(str1)
    if kind == 1:
        del conditionals[n]
        dummy = new_prop(prop_sent, str1, ng, "&I", \
            anc1, anc2, anc3, anc4)
        return dummy
    elif kind == 2:
        dummy = new_prop(prop_sent, str1, ng, "&I", \
            anc1, anc2, anc3, anc4)
        return dummy
    else:
        if os(str1):
            del conditionals[n]
            str1 = remove_outer_paren(str1)
            list1 = tilde_removal2(str1)
            str1 = list1[0]
            dummy = new_prop(prop_sent, str1, list1[1], rule + "E", \
            anc1, anc2)
            candd.append([sn,str1,list1[1]])
            return dummy
        elif list2[0] == "&":
            del conditionals[n]
            str1 = remove_outer_paren(str1)
            dummy = new_prop(prop_sent, str1, ng, rule + "E", \
            anc1, anc2)
            g = copy.copy(sn)
            list3 = get_conjuncts(str1)
            for i in range(len(list3)):
                list4 = tilde_removal2(list3[i])
                dummy = new_prop(prop_sent, list4[0], list4[1], "&E", g,"")
                if dummy == False:
                    return dummy
                if list3[i].find(idisj) > -1:
                    dummy = disjunction_heirarchy(conditionals, list4[0],n, True)
                else:
                    candd.append([sn,list4[0],list4[1]])
            return True
        else:
            dummy = new_prop(prop_sent, str1, ng, idisj + "E", \
            anc1, anc2)
            if dummy == False:
                return dummy
            if ng == "~":
                str1 = ng + str1
            else:
                str1 = remove_outer_paren(str1)
            dummy = disjunction_heirarchy(conditionals, str1,n, False)
            conditionals[n][2] = sn
            return True

def xorr_elim(conditionals,n,i,parent,grandparent,whole_d,candd,prop_sent,anc1,anc2,kind=0):

    str9 = ""
    de_mor = False
    if kind == 0:
        for r in range(len(conditionals[n][i][4])):
            if r != i:
                if not os(conditionals[n][i][4][r][0]):
                    de_mor = True
                if str9 == "":
                    str9 += "~" + conditionals[n][i][4][r][1] + conditionals[n][i][4][r][0]
                else:
                    str9 += " & ~" + conditionals[n][i][4][r][1] + conditionals[n][i][4][r][0]
    else:
        grandp2 = copy.copy(grandparent)
        grandp2 = grandp2.replace(parent,"")
        for r in range(8,38):
            if conditionals[n][r] == "":
                break
            if conditionals[n][r][0][0] in grandp2:
                if str9 == "":
                    str9 += "~" + conditionals[n][r][0][1] + conditionals[n][r][0][0]
                else:
                    str9 += " & ~" + conditionals[n][r][0][1] + conditionals[n][r][0][0]
    g = copy.copy(sn)
    if parent != grandparent:
        str9 = remove_outer_paren(str9)
        if grandparent == whole_d:
            mc = mainconn(str9)
            if mc[0] == '&':
                consistent = xorr_elim2(str9,prop_sent,conditionals,candd,anc1,anc2)
                if consistent == False:
                    return consistent
            else:
                list4 = tilde_removal(str9)
                consistent = new_prop(prop_sent, list4[0],list4[1], xorr + "E", \
                anc1, anc2)
                if consistent == False:
                    return consistent
        else:
            str9 = "(" + str9 + ")"
            if kind == 0:
                str9 = grandparent.replace(parent,str9)
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                dummy = new_prop(prop_sent,str9,"",xorr + "E",anc1,anc2)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~","")
                    dummy = new_prop(prop_sent,str9,"","~~E",sn,"")

            else:
                str9 = whole_d.replace(grandparent,str9)
                str9 = bad_paren(str9)
                dummy = new_prop(prop_sent,str9,"",xorr + "E",anc1,anc2)
                g = copy.copy(sn)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~","")
                    consistent = new_prop(prop_sent,str9,"","~~E",g,"")
                    if consistent == False:
                        return consistent
                dummy = disjunction_heirarchy(conditionals, str9,n, True)
                del conditionals[n]
            if de_mor:
                consistent = demorgan(prop_sent,conditionals,candd,"",True,str9,sn,xorr +"E")
                if consistent == False:
                    return consistent
            else:
                if str9.find(idisj) > -1 or str9.find(xorr) > -1:
                    dummy = disjunction_heirarchy(conditionals,str9,n,True)
                consistent = True
    else:
        #this does not account for the case where the parent == grandparent but
        # grandparent does not == whole d
        consistent = xorr_elim2(str9,prop_sent,conditionals,candd,anc1,anc2)

    return consistent

def xorr_elim2(str9,prop_sent,conditionals,candd,anc1,anc2):

    str9 = bad_paren(str9)
    consistent = new_prop(prop_sent, str9, "", xorr + "E", \
    anc1, anc2)
    if consistent == False:
        return False
    if str9.find("~~") > -1:
        str9 = str9.replace("~~","")
        dummy = new_prop(prop_sent,str9,"","~~E",sn,"")
        if dummy == False:
            return dummy
    list3 = get_conjuncts(str9)
    g = copy.copy(sn)
    for b in range(len(list3)):
        list4 = tilde_removal2(list3[b])
        list4[0] = remove_outer_paren(list4[0])
        dummy = new_prop(prop_sent, list4[0], list4[1], "&E", g,"")
        if dummy == False:
            return dummy
        if not os(list3[b]):
            if list4[1] == "~":
                consistent = demorgan(prop_sent,conditionals,\
                candd,"",True,list3[b],sn,"&E")
                if consistent == False:
                    return False
            else:
                dummy = disjunction_heirarchy(conditionals, list4[0],n, True)
        else:
            candd.append([sn,list4[0],list4[1]])
    return True

def disjunction_elimination(prop_sent, conditionals, candd, kind = ""):

    bool1 = False
    bool2 = False
    global sn
    global rel_conj

    for i in range(len(conditionals)):
        if conditionals[i][8] == "":
            dummy = disjunction_heirarchy(conditionals, conditionals[i][4],i)
    i = -1
    conjt = copy.deepcopy(candd)
    if kind == 2:
        list1 = []
        rel_conj = finddisj(conditionals,list1,1)

    while i < len(conjt) -1:
        i += 1
        if conjt[i][1] not in rel_conj:
            del conjt[i]
            i -= 1
    d = -1
    while d < len(conjt) -1:
        d += 1
        str2 = conjt[d][2]
        conj = conjt[d][1]
        if conj == 'q':
            bb = 7
        if d == 9:
            bb = 7
        anc1 = conjt[d][0]
        n = -1
        while n < len(conditionals) -1:
            if bool1:
                bool1 = False
                d = -1
                break
            n += 1
            i = 7
            while conditionals != []:
                if bool2:
                    bool2 = False
                    break
                i += 1
                if conj not in conditionals[n][38]:
                    break
                else:
                    if conditionals[n][i] == "":
                        break
                    whole_d = conditionals[n][4]
                    anc2 = conditionals[n][2]
                    str3 = conditionals[n][i][0][0]
                    # 'pos or neg'
                    str4 = conditionals[n][i][0][1]
                    # 'disjunct or conjunct
                    str5 = conditionals[n][i][2]
                    # 'disjunct number
                    str6 = conditionals[n][i][1]

                    if conj == str3:
                        grandparent = conditionals[n][i][6]
                        parent = conditionals[n][i][5]
                        parent2 = copy.copy(parent)
                        parent3 = copy.copy(parent)
                        str7 = " " + idisj + " "
                        str7a = " " + xorr + " "
                        if str2 == str4 and str5 == "d":
                            # 'if the disjuncts are not embedded within a conjunct then the disjunction
                            # is simply deleted

                            del conditionals[n]
                            if parent != grandparent:
                                conj = str2 + conj
                                str8 = whole_d.replace(parent, conj)
                                dummy = disjunction_heirarchy(conditionals, str8,n)
                            bool1 = True
                            n = -1
                            break

                        elif str2 == str4 and str5 == "x":

                            consistent = xorr_elim(conditionals,n,i,parent,grandparent,whole_d,candd,prop_sent,anc1,anc2)
                            if consistent == False:
                                return False
                            del conditionals[n]
                            bool2 = True
                            bool1 = True
                            d = -1
                        elif str2 == str4 and str5 == "c":
                            list2 = []
                            list2.append([conj,str2])
                            anc3 = ""
                            anc4 = ""
                            list1 = conditionals[n][i][4]
                            f = -1
                            while f < len(list1) -1:
                                mc = mainconn(grandparent)
                                f += 1
                                for e in range(len(candd)):
                                    anc5 = candd[e][0]
            # since it's too hard to program, if the sibling is a disjunct then we just
            # ignore this
                                    if list1[f][0].find(idisj) > -1 or list1[f][0].find(xorr) > -1:
                                        break
                                    else:
                                        if candd[e][1] == list1[f][0]:
                                            if candd[e][2] == list1[f][1]:
                                                list2.append([list1[f][0],list1[f][1]])
                                                if len(list2) == 2:
                                                    anc3 = anc5
                                                elif len(list2) == 3:
                                                    anc4 = anc5
                                                del list1[f]
                                                if list1 == []:
                                                    str3 = build_sent_list2(list2)
                                                    if mc[0] == xorr:
                                                        new_prop(prop_sent,str3,"","&I",anc1,anc3,anc4)
                                                        consistent = xorr_elim(conditionals,n,i,parent,\
                                                        grandparent,whole_d,candd,prop_sent,sn,anc2,1)
                                                        if consistent == False:
                                                            return False
                                                    else:
                                            # if the conjunct is not embedded within another conjunct
                                            # then the disjunct is simply deleted
                                                        if whole_d == grandparent:
                                                            dummy = new_disjunct(str3,"",n, prop_sent,\
                                                            conditionals,candd, anc1, anc3, anc4, anc5, 1)
                                                        else:
                                                            str8 = whole_d.replace(grandparent, parent2)
                                                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                                                str8 = bad_paren(str8)
                                                            dummy = new_disjunct(str3,"",n, prop_sent,\
                                                            conditionals,candd, anc1, "", anc3, anc4, 2)
                                                            dummy = new_disjunct(str8,"",n, prop_sent,\
                                                            conditionals,candd, sn-1, anc2)
                                                            if dummy == False:
                                                                return False
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break
                                                else:
                                                    f -= 1
                                                    break

                                            elif candd[e][2] != list1[f][1]:
                                                mc = mainconn(grandparent)
                                                if mc[0] == idisj:
                                                    rule = idisj
                                                    str7 = " " + idisj + " "
                                                else:
                                                    str7 = " " + xorr + " "
                                                    rule = xorr
                                                r = grandparent.find(parent)
                                                if r > 1:
                                                    parent = str7 + parent
                                                else:
                                                    parent = parent + str7
                                                anc1 = candd[e][0]
                                                str9 = grandparent.replace(parent, "")
                                                str8 = whole_d.replace(grandparent, str9)
                                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                                    str8 = bad_paren(str8)
                                                dummy = new_disjunct(str8,"",n, prop_sent,\
                                                        conditionals,candd, anc1, anc2,None,None,0,rule)
                                                if dummy == False:
                                                    return False
                                                else:
                                                    list1 = []
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break

                        elif str2 != str4 and str5 == "c":
                            mc = mainconn(conditionals[n][i][6])
                            if mc[0] == idisj:
                                str6 = str7 + parent
                                rule = idisj
                            else:
                                rule = xorr
                                str6 = str7a + parent
                            if grandparent.find(str6) > -1:
                                parent = str6
                            else:
                                parent = parent + str7
                            str9 = grandparent.replace(parent, "")
                            str8 = whole_d.replace(grandparent, str9)
                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                str8 = bad_paren(str8)
                            dummy = new_disjunct(str8,"",n,prop_sent, conditionals, candd,anc1, anc2, None, None,0,rule)
                            if dummy == False:
                                return False
                            bool1 = True
                            n = -1
                            break

                        elif str2 != str4 and (str5 == "d" or str5 == 'x'):
                            # if the disjunct is a triple disjunct then enter below
                            if str5 == 'd':
                                rule = idisj
                            else:
                                rule = xorr
                            if conditionals[n][i][7] > 1:
                                str6 = str4 + str3 + " " + rule + " "
                                if parent.find(str6) > -1:
                                    str5 = str6
                                else:
                                    str5 = " " + rule + " " + str4 + str3
                                str9 = parent.replace(str5, "")
                                str8 = whole_d.replace(parent, str9)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                dummy = new_disjunct(str8,"",n,prop_sent,conditionals,\
                                    candd, anc1,anc2,None,None,0,rule)
                                if dummy == False:
                                    return False
                                bool1 = True
                                n = -1
                                break

                            else:
                                str3 = conditionals[n][i][4][0][0]
                                str4 = conditionals[n][i][4][0][1]
                                str5 = str4 + str3
                                str8 = whole_d.replace(parent, str5)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                dummy = new_disjunct(str8,"",n,prop_sent,conditionals,\
                                    candd, anc1,anc2,None,None,0,rule)
                                if dummy == False:
                                    return False
                                bool1 = True
                                n = -1
                                break
    return True

def extract_list(list1,d):

    list2 = []
    for i in range(len(list1)):
        list2.append(list1[i][d])
    return list2

def statement_logic(prop_sent, conditionals, candd, disjuncts,kind="", conc="", impl=""):

    global sn
    b = time.time()
    consistent = modus_ponens(conditionals, candd, prop_sent)
    if consistent == False:
        return False
    consistent = iff_elim(prop_sent,conditionals,kind)
    if consistent == False:
        return False
    consistent = material_implication(prop_sent, conditionals,kind)
    if consistent == False:
        return False
    consistent = demorgan(prop_sent, conditionals, candd,kind)
    if consistent == False:
        return False
    consistent = disjunction_elimination(prop_sent,conditionals,candd,kind)
    if consistent == False:
        return False
    if kind == 1:
        dummy = finddisj(conditionals,disjuncts)
    c = time.time()
    d = c - b
    return True

def finddisj(conditionals,disjuncts, cate=""):

    if cate == 1:
        disjuncts = []

    for i in range(len(conditionals)):
        for j in range(8,36):
            if type(conditionals[i][j]) is list:
                if conditionals[i][j][0][0] not in disjuncts:
                    disjuncts.append(conditionals[i][j][0][0])
            else:
                break
    if cate == 1:
        return disjuncts

def add_outer_paren(str1):

    str1 = remove_outer_paren(str1)
    return "(" + str1 + ")"

def new_prop_sent(ng, kind, asp, anc1, anc2, conditionals,g,list3 =[], cjct = ""):

    global prop_sent
    global candd
    global sn
    global impl

    if kind == 'con':
        h = 1
        e = 7
    else:
        h = 0
        e = 6
    bool1 = False
    if list3 != []:
        sn += 1
        ancc = [4,5,6,7]
        list4 = [None] * 15
        for i in range(len(list3)):
            if i == 4:
                break
            list4[ancc[i]] = list3[i]
        list4[0] = sn
        list4[1] = cjct
        list4[2] = ""
        list4[3] = "&I"
        anc1 = sn
        prop_sent.append(list4)

    if conditionals[g][e] == "":
        str1 = conditionals[g][h][0]
        list2 = mainconn(str1)

        if implies in conditionals[g][h][0] or nonseq in conditionals[g][h][0]:
            sn += 1
            prop_sent.append([sn,str1, "",asp,anc1,anc2,"","",""])
            g = str1.find(implies)
            asp = "DE " + implies
            if g == -1:
                g = str1.find(nonseq)
                asp = "DE " + nonseq
            str2 = str1[:g]
            str3 = str1[g+1:]
            str2 = str2.strip()
            str3 = str3.strip()
            if impl == implies:
                str4 = str2 + " & ~" + str3 + " & " + bottom
                str5 = bottom
            else:
                str4 = str2 + " & ~" + str3 + " & " + top
                str5 = top
            sn += 1
            prop_sent.append([sn,str4, "",asp,sn-1,"","","",""])
            sn += 1
            prop_sent.append([sn,str2, "","&E",sn-1,"","","",""])
            candd.append([sn,str2,""])
            sn += 1
            prop_sent.append([sn,str3, "~","&E",sn-2,"","","",""])
            candd.append([sn,str3,"~"])
            sn += 1
            prop_sent.append([sn,str5, "","&E",sn-3,"","","",""])
            return

        # here we take care of double negatives
        if ng == "~" and conditionals[g][h][1] == "~":
            sn += 1
            prop_sent.append([sn, "~" + str1, "~", asp, anc1, anc2, None, None, None, None, None,None, None, None, None])
            g = copy.copy(sn)
            dummy = new_prop(prop_sent,str1,"","~~E",g, None)
            if dummy == False:
                return False
            ng = ""
            bool1 = True
        elif (ng == "" and conditionals[g][h][1] == "~") or \
                (ng == "~" and conditionals[g][h][1] == ""):
            ng = "~"
        if list2[0] != "" and ng != "~":
            str1 = remove_outer_paren(str1)
        elif list2[0] != "" and ng == "~":
            str1 = add_outer_paren(str1)
        if bool1 == False:
            dummy = new_prop(prop_sent,str1,ng,asp,anc1, anc2)
            if dummy == False:
                return False

        candd.append([sn, str1, ng])
        if (list2[0] == iff or list2[0] == conditional) and ng != "~":
            list3 = prepare_iff_elim(str1, list2[0],list2[1],sn)
            conditionals.append(list3)
        elif list2[0] == iff or list2[0] == conditional or list2[0] == idisj or list2[0] == xorr:
            list5 = [""] * 39
            list5[2] = sn
            list5[4] = str1
            list5[5] = ng
            list5[3] = 'd'
            conditionals.append(list5)
    else:
        str1 = conditionals[g][e]
        if ng == "~":
            list5 = [""] * 39
            list5[2] = sn + 1
            list5[4] = str1
            list5[5] = "~"
            list5[3] = 'd'
            conditionals.append(list5)
            dummy = new_prop(prop_sent, str1, ng, asp, anc1, anc2)
            if dummy == False:
                return False
        else:
            list1 = conditionals[g][h]
            str1 = remove_outer_paren(str1)
            dummy = new_prop(prop_sent, str1, ng, asp, anc1, anc2)
            if dummy == False:
                return False
            anc1 = copy.copy(sn)
            for i in range(len(list1)):
                list2 = mainconn(list1[i][0])
                dummy = new_prop(prop_sent, list1[i][0], list1[i][1], "&E", anc1, "")
                if dummy == False:
                    return False
                candd.append([sn, list1[i][0], list1[i][1]])
                if (list2[0] == conditional or list2[0] == iff) and list1[i][1] != "~":
                    list4 = prepare_iff_elim(list1[i][0], list2[0],list2[1],sn)
                    conditionals.append(list4)
                elif list2[0] != "":
                    list5 = [""] * 39
                    if list2[0] == conditional:
                        list5[3] = 'c'
                    elif list2[0] == iff:
                        list5[3] = 'e'
                    else:
                        list5[3] = 'd'
                    list5[2] = sn
                    list5[4] = list1[i][0]
                    list5[5] = list1[i][1]
                    conditionals.append(list5)
    return True

def oc(str1):
    # this function determines if there is only one connective which is not &
    # it is used to weed out conditionals from the candd list
    list1 = [conditional, idisj, iff,xorr]
    j = 0
    for i in range(len(str1)):
        str2 = str1[i:i+1]
        if str2 in list1:
            j += 1
    if j == 1:
        return True
    else:
        return False

def plan(sent, prop_sent, candd, conditionals, prop_name, disjuncts, kind = '',negat=[]):

    global conc
    global sn
    global rel_conj
    conj_elim = []
    temp_conditionals = []
    list4 = []
    str3 = ""
    str1 = ""
    conc = ""
    impl = ""
    qq = 0
    nat_logic = False
    # if the first sentence is just one letter then we're using statement logic
    # if sent[0][1].find("(") > -1:
    #     for j in range(len(sent[0][1])):
    #         str2 = sent[0][1][j:j+1]
    #         str3 = sent[0][1][j+1:j+2]
    #         if (str2.islower() or str2.isupper()) and (str3.islower() or str3.isupper()):
    #             nat_logic = True
    #         if j > 8:
    #             break


    for i in range(len(sent)):
        # if i == 23:
        #     bb = 7
        g = sent[i].count('(')
        h = sent[i].count(')')
        if g != h:
            easygui.msgbox('wrong number of parentheses in sentence:' + sent[i])
            return "stop"
        if nat_logic == False:
            sent[i][1] = enclose(sent[i][1])
        if sent[i][1].find("!") > -1:
            qq += 1
        else:
            if sent[i][1].count("(") != sent[i][1].count(")"):
                easygui.msgbox("line " + str(sent[i][0]) + " does not have the right number \
                                             of parentheses" )
            if nat_logic == False:
                sent[i][1] = remove_outer_paren(sent[i][1])
            ostring = copy.copy(sent[i][1])
            nstring = proper_spacing(ostring)

            if os(sent[i][1]):
                if kind == "":
                    list3 = tilde_removal(nstring)
                    list4 = tilde_removal(ostring)
                    str2 = list3[0]
                    ng = list3[1]
                else:
                    ng = negat[i]
                    str2 = sent[i][1]
                if nat_logic:
                    list3[0] = simple_sent_name(list3[0],list4[0])
                candd.append([sent[i][0], str2, ng])
            else:
                list1 = find_sentences(ostring, True)
                if not nat_logic:
                    list6 = unenclose(list1[0][0])
                    sent[i][1] = list6[0]
                else:
                    list6 = name_conditional(list1,nat_logic)
                    list6[0] = proper_spacing(list6[0])
                str2 = list6[0]
                if kind == "":
                    list4 = tilde_removal2(str2)
                    ng = list4[1]
                    str2 = list4[0]
                else:
                    ng = negat[i]
                list2 = mainconn(str2)
                if list2[0] == idisj or list2[0] == xorr:
                    if oc(str2):
                        candd.append([nstring, str2,ng])
                        # rel_conj = rel_conj + list6[1]
                    list5 = [""] * 39
                    list5[2] = sent[i][0]
                    list5[3] = 'd'
                    list5[4] = str2
                    list5[5] = ng
                    conditionals.append(list5)
                elif list1[4][0][1] != "&":
                    str3 = ""
                    # rel_conj = rel_conj + list6[1]
                    if list1[4][0][1] != idisj and ng == "" and list1[4][0][1] != xorr:
                        list7 = prepare_iff_elim(str2, list2[0], list2[1],sent[i][0])
                    else:
                        list7 = [""] * 39
                        list7[2] = sent[i][0]
                        list7[3] = 'd'
                        list7[4] = str2
                        list7[5] = ng
                    conditionals.append(list7)
                    if oc(str2):
                        candd.append([sent[i][0], str2,ng])
                else:
                    list3 = get_conjuncts(list6[0])
                    for j in range(len(list3)):
                        list5 = tilde_removal2(list3[j])
                        if os(list3[j]):
                            conj_elim.append([sent[i][0],list5[0],list5[1]])
                        else:
                            temp_conditionals.append([sent[i][0], list5[0],list5[1]])

            no_contr = new_prop(prop_sent,str2,ng,"",None,None,None,None,True,sent[i][0],ostring)
            if not no_contr:
                return False

    if conj_elim != []:
        for i in range(len(conj_elim)):
            no_contr = new_prop(prop_sent,conj_elim[i][1],conj_elim[i][2],"&E",\
                        conj_elim[i][0],None)
            if not no_contr:
                return False
            candd.append([sn, conj_elim[i][1], conj_elim[i][2]])

    if temp_conditionals != []:
        for i in range(len(temp_conditionals)):
            if nat_logic:
                list1 = find_sentences(temp_conditionals[i][1], True)
                list6 = name_conditional(list1)
                str2 = list6[0]
            else:
                str2 = temp_conditionals[i][1]
            ng = temp_conditionals[i][2]
            list2 = mainconn(str2)

            if list2[0] != idisj and ng == "" and list2[0] != "&" and list2[0] != xorr:
                list7 = prepare_iff_elim(str2, list2[0], list2[1],sent[i][0])
            else:
                list7 = [""] * 39
                list7[2] = sn + 1
                list7[4] = str2
                list7[3] = 'd'
            if ng == "~":
                list7[5] = "~"
            # rel_conj = rel_conj + list6[1]
            conditionals.append(list7)
            if oc(str2):
                candd.append([sn+1, str2,ng])
            no_contr = new_prop(prop_sent, str2,ng,"&E",temp_conditionals[i][0], None)
            if not no_contr:
                return False

    consistent = statement_logic(prop_sent, conditionals, candd,disjuncts,1,conc,impl)
    return consistent


def populate_sentences(p):
    global result_data
    global excel
    bool1 = False
    bool2 = False
    first_sent = False
    sent = []
    test_sent = []
    g = 0

    if not excel:

        for row in w4:
            p += 1
            if row[1] == '' and bool2 == True:
                break
            elif row[1] == 'stop':
                break
            elif row[1] == '':
                test_sent.append(sent)
                sent = []
                g = 0
                bool1 = False
                bool2 = True
                first_sent = False
            elif row[1].find("!!!") > -1:
                bool1 = True
                test_sent.append(sent)
                sent = []
            elif row[1] != '' and bool1 == False:
                str2 = row[1]
                g += 1
                if len(sent) == 0:
                    if str2.find(bottom) > -1:
                        tv = 'co'
                        str2 = str2[2:]
                    else:
                        tv = 'ta'
                else:
                    tv = ""
                str2.strip()
                sent.append([g, str2,'',tv])
                if not first_sent:
                    result_data['text_'+str(p-2)+'_1']=len(test_sent)
                first_sent = True
                bool2 = False
    else:
        for row in w4.rows:
            p += 1
            try:
                if row[2].value == None and bool2 == True:
                    break
                elif row[2].value == 'stop':
                    break
                elif row[2].value == None:
                    test_sent.append(sent)
                    sent = []
                    g = 0
                    bool1 = False
                    bool2 = True
                    first_sent = False
                elif row[2].value.find("!!!") > -1:
                    bool1 = True
                    test_sent.append(sent)
                    sent = []
                elif row[2].value != None and bool1 == False:
                    str2 = row[2].value
                    g += 1
                    if len(sent) == 0:
                        if str2.find(bottom) > -1:
                            tv = 'co'
                            str2 = str2[2:]
                        else:
                            tv = 'ta'
                    else:
                        tv = ""
                    str2.strip()
                    sent.append([g, str2,row[0].value,tv])
                    if not first_sent:
                         w4.cell(row=p-1,column=2).value = len(test_sent)
                    first_sent = True
                    bool2 = False
            except AttributeError:
                bb = 8


    return [test_sent,p]

def repeat_relations(str1):

    list1 = ["group","x"]
    list2 = ["member",'z']
    list3 = ["no"+us,"y"]
    list4 = ['no',"y"]
    list5 = ['any',"y"]

    final_list = [list1,list2,list3,list4,list5]
    return final_list

def get_result(post_data):
    global w4, result_data,p
    if not excel:
        result_data = dict(post_data.iterlists())
        w4=[]
        index=0
        while  True:
            row = (post_data["text_"+str(index)+"_1"],post_data["text_"+str(index)+"_2"],post_data["text_"+str(index)+"_3"])
            w4.append(row)
            if row[1] == "stop":
                break
            index+=1
        w4=tuple(w4)

    global prop_name,plural_c,anaphora,definite, prop_var, ind_var
    global ant_cond,conditionals,candd,rel_conj,conc,prop_sent,sn,impl
    global tagged_nouns,tagged_nouns2,dv_nam,basic_objects,idf_var,p
    global gen_var, definite2

    list1 = populate_sentences(p)
    test_sent = list1[0]
    p = list1[1]
    words = build_dict('hey')
    st = time.time()
    rep_rel = repeat_relations('hey')

    for k in range(len(test_sent)):
        print k
        if k == 2:
            bb = 7
        prop_name = []
        tot_sent = []
        all_sent = []
        never_used = []
        plural_c = []
        anaphora = ""
        impl = ""
        definite = []
        definite2 = []
        gen_var = []
        ant_cond = []
        conditionals = []
        candd = []
        ind_var = []
        rel_conj = []
        conc = []
        psent = []
        identities = []
        prop_sent = []
        tagged_nouns = []
        tagged_nouns2 = []
        dv_nam = []
        basic_objects = []
        prop_var = copy.deepcopy(prop_var4)
        idf_var = copy.deepcopy(idf_var2)
        id_num = test_sent[k][-1][0] + 1
        sn = id_num
        dummy = divide_sent(words, test_sent[k], idf_var,tot_sent,all_sent)
        dummy = syn(tot_sent, all_sent, words)
        dummy = rel_repl(all_sent,tot_sent,words,dv_nam,idf_var,id_num)
        dummy = word_sub(idf_var,dv_nam, tot_sent, all_sent,words,id_num)
        dummy = plurals(tot_sent,all_sent,words,dv_nam, idf_var)
        dummy = define(tot_sent, all_sent,idf_var, dv_nam, words,rep_rel,identities)
        tot_sent = identity(all_sent,tot_sent,basic_objects,words,candd,\
                conditionals,prop_sent,prop_name,id_num,identities,idf_var)
        test_sent[k] = tot_sent
        tot_prop_name.append(prop_name)
    en = time.time()
    g = (en-st)/(k+1)
    print g
    dummy = print_sent_full(test_sent,p,tot_prop_name)
    if not excel:
        return result_data

if excel:
    dummy = get_result('hey')
    st = time.time()
    wb4.save('inference engine.xlsx')
    en = time.time()
    print en-st
#end1