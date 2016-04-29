from openpyxl import load_workbook
import easygui
import timeit
import copy
import time
import operator


cond_r = unichr(8835)
top = unichr(8868)
bottom = unichr(8869)
neg = unichr(172)
iff = unichr(8801)
mini_c = unichr(8658)
mini_e = unichr(8703)
implies = unichr(8866)
conditional = unichr(8594)
nonseq = unichr(8876)
xorr = unichr(8891)
idisj = unichr(8744)
cj = unichr(8896)
disj = unichr(8855)
equi = unichr(8660)

## for rajiv - you're going to have to code for these, don't worry about l1,l2,l3,l7
## for output I need three buttons - one for delete everything, another for delete the previous output, a third to run
## the program
#"-c" = uc
#"-d" = ud



ne = u"\u2260"
l1 = u"\u2081"
l2 = u"\u2082"
l3 = u"\u2083"
l7 = u"\u2087"
l8 = u"\u2088"
ua = u"\u1d43"
ub = u"\u1d47"
uc = u"\u1d9c"
ud = u"\u1d48"
ue = u"\u1d49"
uf = u"\u1da0"
ug = u"\u1d4d"
ui = u"\u2071"
uk = u"\u1d4f"
um = u"\u1d50"
un = u"\u207f"
uo = u"\u1d52"
up = u"\u1d56"
ut = u"\u1d57"
uv = u"\u1d5b"
uu = u"\u1d58"
uw = u"\u02b7"
uy = u"\u02b8"
uj = u"\u02B2"
ul = u"\u02E1"
ur = u"\u02b3"
us = u"\u02e2"
uh = u"\u02b0"


pp = 7
#
# >> 8835
# ta^ 8868
# co^ 8869
# ; 172
# <> 8801
# c^ 8658
# # 8703
# i^ 8866
# > 8594
# nf^ 8876
# ed^ 8891
# + 8744
# && 8896
# @ 8855
# if^ 8660



def remove_outer_paren(str1):

    if str1 == "":
        return ""
    elif str1.count(")") == 0:
        return str1

    j = 0
    # on very rare occasions we will encounter strings of the following form ((p))
    if str1[0] != "(" and str1[-1] != ")":
        return str1
    if str1[:2] == "((" and str1[-2:] == "))":
        d = 2
    else:
        d = 1

    for k in range(0, d):
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]

    return str1


def remove_redundant_paren(str1):
    j = 0
    str2 = str1[:2]
    str3 = str1[2:]
    if str2 == '((' and str3 == '))':

        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 1 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
    return str1

def mainconn(str1):
    # str_main = str_main.replace(implies,"&")
    # str_main = str_main.replace(nonseq, "&")
    # str_main = str_main.replace("~","")
    # str_main = str_main.replace(" ", "")

    ostring = copy.copy(str1)
    if os(str1):
        return ["", 0]
    if str1.find("&") < -1 and str1.find(idisj) < -1 and str1.find(iff) < -1 and str1.find(conditional) < -1 and \
            str1.find(implies) < -1 and str1.find(nonseq) < -1 and str1.find(xorr) < -1:
        return ["", 0]

    str3 = str1
    bool1 = False

    if str1[0] == "~":
        str1 = str1[1:]
        bool1 = True

    # list1 = tilde_removal2(str1)
    # if list1[1] == "~":
    #     str1 = str1[1:]
    #     bool1 = True

    j = 0
    bool2 = False
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
            bool2 = True


    j = -1

    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == conditional:
            f = -1
        if str2 == idisj or str2 == "&" or str2 == iff or str2 == implies or \
                str2 == nonseq or str2 == conditional or str2 == xorr:
            if str3 != str2 and str3 != "":
                j = j + 1

            str3 = str2

    if j == -1:
        i = ostring.find(str3)
        return [str3, i]
    k = -1
    j = -1
    while True:
        k += 1
        if k > 150:
            break
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]

            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1

            if j == -1 and (str2 == idisj or str2 == "&" or str2 == iff or str2 == implies
                            or str2 == nonseq or str2 == conditional or str2 == xorr):
                if bool1 and bool2:
                    return [str2, i+2]
                elif bool2 or bool1:
                    return [str2, i+1]
                else:
                    return [str2, i]
        else:
            str1 = str1[1:-1]

def isvariable(str3,kind=""):

    bool2 = True
    if str3 == None:
        return False
    if str3.find(l1) > -1:
        isidef = True
    elif str3.find(l2) > -1:
        isidef = True
    elif str3.find("'") > -1:
        isidef = False
    elif str3.find(l7) > -1:
        isidef = False
    else:
        if len(str3) == 1:
            isidef = True
        else:
            isidef = False

    if str3 == 'a':
        return False
    try:
        if str3 != "":
            str3 = str3.replace(l1, "")
            str3 = str3.replace(l2, "")
            str3 = str3.replace("'", "")
            str3 = str3.replace(neg, "")
            if len(str3) == 1 and str3.islower():
                bool2 = True
            else:
                bool2 = False
    except AttributeError:
        easygui.msgbox('error in isvariable function')
    if kind == "":
        return bool2
    else:
        return isidef

def os(str1):

    cnx = [xorr, iff, idisj, conditional, implies, nonseq, "&"]
    for i in range(0, len(cnx)):
        if str1.find(cnx[i]) > -1:
            os = False
            return os
    os = True
    return os

def enclose(str1):

    i = -1
    global subscripts
    while i < len(str1) -1:
        i += 1
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        str4 = str1[i+1:i+2]
        if str2.islower() and str4 in subscripts:
            if str3 == "~":
                str1 = str1[:i-1] + "(~" + str2 + str4 + ")" + str1[i+2:]
            else:
                str1 = str1[:i] + "(" + str2 + str4 + ")" + str1[i+2:]
            i += 4
        elif str2.islower():
            if str3 == "~":
                str1 = str1[:i-1] + "(~" + str2 + ")" + str1[i+1:]
            else:
                str1 = str1[:i] + "(" + str2 + ")" + str1[i+1:]
            i += 3
    return str1


def find_sentences(instring, cut_skel = False):

    if instring == None:
        return

    g = instring.count('(')
    h = instring.count(')')
    if g != h:
        easygui.msgbox('wrong number of parentheses in sentence:' + instring)
        return "stop"
    marker = False
    il = -1
    total = -1
    c = -1
    neg_value = []
    str1 = ""
    sent1 = []
    sent_type2 = []
    wneg = []
    output = [None] * 9
    # the skel name list names each single sentence after a greek letter, even if
    # the same sentence appears twice it obtains a different name on the second
    # appearance
    skel_nam = []
    sent_num = []
    if instring.find("~(") > -1:
        instring = instring.replace("~(", "(!")
    if instring.find(implies) > -1:
        str2 = implies
    elif instring.find(nonseq) > -1:
        str2 = nonseq
    str3 = mainconn(instring)
    str4 = str3[0]
    f = str3[1]
    id_num = []
    id_num.append(["1",str4,f])
    sent_num.append([1, '1', instring, str4,f])
    prtnum = 1
    skel_string = instring
    p = 947
    connectives = ["&", idisj, iff, conditional, nonseq, implies,xorr]
    arr1 = []
    mini_c2 = mini_c + neg
    instring2 = copy.copy(instring)
    instring = instring.strip()
    prt = copy.copy(instring)
    list1 = mainconn(instring)
    grandparent_type = list1[0]
    more_num = [unichr(945 + x) for x in range(24)]
    # prop_var = [unichr(97 + t) for t in range(26)]

    if os(instring) == False:
        temp_string = mainconn(instring)

        if instring.find(implies) > -1:
            str1 = implies
        elif instring.find(nonseq) > -1:
            str1 = nonseq
        else:
            if temp_string == iff:
                str1 = "bicond"
            elif temp_string == conditional:
                str1 = "cond"
            elif temp_string == "&":
                str1 = "cj"

        sent1.append(instring)
        neg_value.append("")
        sent_type2.append(str1)
        wneg.append(instring)
        skel_nam.append(None)

    j = 0
    n = 0
    for i in range(0, len(instring)):
        str1 = instring[i:(i + 1)]
        for o in connectives:
            if str1 == o:
                j += 1

    while n < j + 1:

        il += 1
        if il > 15:
            break

        e = 0
        l = len(instring)
        x = -1
        while x < l - 1:
            x += 1
            temp_string = instring[x: x + 1]
            if instring[x: x + 1] == "(":

                if marker == False:
                    z = x
                    marker = True

                total += 1
            elif instring[x: x + 1] == ")":
                total -= 1
                if total == -1:
                    marker = False
                    e += 1
                    c += 1

                    temp_sent = instring[z: x + 1]
                    otemp_sent = copy.copy(temp_sent)

                    if (len(instring) - len(temp_sent)) > 2:

                        if temp_sent in prt:
                            prtnum = findinlist(prt,sent_num,2,1,False)
                            numb = str(prtnum) + "1"
                        else:
                            prtnum = ""
                            for bb in range(len(sent_num)-1,-1,-1):
                                str3 = sent_num[bb][2]

                                if temp_sent in sent_num[bb][2]:
                                    prtnum = sent_num[bb][1]
                                    break
                            # if prtnum == "":
                            #     easygui.msgbox('your sentences are not numbered properly')
                            g = len(prtnum) + 1
                            f = 0

                            for bb in range(len(sent_num)-1,-1,-1):
                                temp_sn = sent_num[bb][1]
                                h = temp_sn[:g-1]
                                hh = sent_num[bb][0]
                                if g > sent_num[bb][0]:
                                    break
                                if sent_num[bb][0] == g and temp_sn[:g-1] == prtnum:
                                    f += 1

                            f += 1
                            if f < 10:
                                numb = prtnum + str(f)
                            else:
                                numb = prtnum + more_num[f-10]

                        if n == 7:
                            pp = 7

                        prt = temp_sent
                        temp_mc = mainconn(temp_sent)
                        mc = temp_mc[0]
                        str3 = temp_mc[0]
                        g = temp_mc[1]
                        mc_num = temp_mc[1]
                        sent_num.append([len(numb), numb, temp_sent, str3,g])


                        if os(temp_sent):
                            # n counts the number of single sentences
                            n += 1
                            if temp_sent.find("~") > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace("~", "")

                            elif temp_sent.find(mini_c2) > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace(mini_c2, mini_c)

                            elif temp_sent.find(mini_c2) == -1 and temp_sent.find(mini_c) > -1:
                                neg_value.append("")
                            elif temp_sent.find("~") == -1:
                                neg_value.append("")
                            else:
                                break  # stop
                        else:
                            neg_value.append("")

                        sent1.append(temp_sent)
                        wneg.append(otemp_sent)
                        id_num.append([numb,mc,mc_num])

                        if os(otemp_sent):
                            p += 1
                            skel_nam.append([otemp_sent, unichr(p)])
                        else:
                            skel_nam.append(None)
                    else:
                        instring = instring[1:len(instring) - 1]
                        l = len(instring)
                        x = -1
                        c -= - 1
                        e -= - 1

        total = -1
        marker = False
        w = -1

        if n < j + 1:
            if len(sent1) > w:
                while w + 1 < len(sent1):
                    if w == 13:
                        pp = 7
                    w += 1
                    str21 = sent1[w]
                    if not os(str21) and w != 0:
                        if str21 not in arr1:
                            instring = str21
                            arr1.append(instring)
                            break

    for i in range(len(sent1)):
        temp_string = sent1[i]
        if temp_string.find("(!") > -1:
            sent1[i] = sent1[i].replace("(!", "~(")
            wneg[i] = wneg[i].replace("(!", "~(")
    m = 0
    bool1 = False
    wid = copy.copy(instring2)
    for i in range(len(id_num)):
        if (id_num[i][1] == iff or id_num[i][1] == conditional) \
                and not bool1:
            str1 = wneg[i]
            bool1 = True
            m = i
            break
    for i in range(len(wneg)):
        if os(wneg[i]):
            str1 = str1.replace(skel_nam[i][0],skel_nam[i][1])
            wid = wid.replace(skel_nam[i][0],skel_nam[i][1])

    skel_wid = wid
    skel_string = str1
    skel_string = remove_outer_paren(skel_string)
    if skel_string.find("(!") > -1:
        skel_string = skel_string.replace("(!","~(")
        wid = wid.replace("(!","~(")

    output[0] = sent1
    output[1] = neg_value
    output[2] = sent_type2
    output[3] = wneg
    output[4] = id_num
    output[5] = skel_string
    output[6] = skel_nam
    # m is the first sentence to be used when changing variables in a definition
    output[7] = m
    output[8] = skel_wid

    return output

def add_to_dv(dv_nam,all_sent,m,k,def_var,str2):


    if isvariable(str2) == False:
        str3 = findinlist(str2, dv_nam,1,0)
        if str3 == None:
            telist7 = [def_var[0], str2]
            all_sent[m][k] = def_var[0]
            del def_var[0]
            dv_nam.append(telist7)
        else:
            all_sent[m][k] = str3


def word_sub(def_var, dv_nam, tot_sent, all_sent, words,id_num):

    all_sent = remove_duplicates(all_sent,0)
    relations = words[18]
    relations2 = words[19]
    pronouns = words[24]
    num = [4, 5, 13, 14, 17, 18, 22, 26, 30,34,35,36,51,52]
    num2 = [9,15,19,23,27,31,49]
    global sn
    m = -1
    while m < len(all_sent) -1:
        m += 1
        if all_sent[m][47] == "no word sub":
            return

        bool1 = False
        list4 = copy.deepcopy(all_sent[m][46])
        old_sent = all_sent[m][0]
        oldp = all_sent[m][42]
        # old_list = copy.deepcopy(all_sent[m])
        for i in range(len(all_sent[m][46])):
            k = all_sent[m][46][i]
            str2 = all_sent[m][k]
            if k == 8 or k == 12:
                bool1 = True
                str5 = findinlist(str2,words[16],0,1)
                all_sent[m][k] = str5
            if k in num and all_sent[m][45] != k:
                bool1 = True
                if str2 != None and str2 not in pronouns and str2 != 'there':
                    dummy = add_to_dv(dv_nam,all_sent,m,k,def_var,str2)
                    list4.remove(k)
            elif k in num2:
                list4.remove(k)
                bool1 = True
                h = findin1dlist(str2,relations2)
                str5 = relations[h]
                if str5 != None:
                    dv_nam.append([str2,str5])
                    all_sent[m][k] = str5
                else:
                    easygui.msgbox('you did not code your relations properly')

        if bool1:
            new_sent = build_sent(all_sent[m])
            newp = name_sent(new_sent)
            all_sent[m][0] = new_sent
            all_sent[m][42] = newp
            dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"isub",id_num)
            all_sent[m][46] = list4
            bool1 = False
    return

def assigned_var(str1, dv_nam, def_var):

    bool1 = False
    for i in range(len(dv_nam)):
        if dv_nam[i][1] == str1:
            bool1 = True
            return dv_nam[i][0]

    if bool1 == False:
        str2 = def_var[0]
        list1 = [str2, str1]
        dv_nam.append(list1)
        del def_var[0]
        return str2

def there(all_sent,m,tot_sent,def_sent):

    global sn
    old_sent = all_sent[m][0]
    oldp = all_sent[m][42]
    all_sent[m][5] = all_sent[m][14]
    all_sent[m][3] = all_sent[m][10]
    all_sent[m][4] = all_sent[m][13]
    all_sent[m][14] = None
    all_sent[m][10] = None
    all_sent[m][13] = None
    new_sent = build_sent(all_sent[m])
    newp = name_sent(new_sent)
    all_sent[m][0] = new_sent
    def_sent.append(new_sent)
    all_sent[m][42] = newp
    bool1 = check_dimension(tot_sent,1,new_sent)
    if not bool1:
        dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"df there")


def define(tot_sent, all_sent, def_var, idf_var, dv_nam,words):

    all_sent = remove_duplicates(all_sent,0)
    num = [3,5,10,14,16,18,20,22,24,26,28,30,32,34]
    pronouns = words[24]
    determinative = words[2]
    definitions = words[16]
    posp = words[28]
    atomic_relations = words[22]
    atomic_relata = words[23]
    def_relat = ['IA','IG','=','H']
    global sn
    bool1 = False
    def_sent = []

    for i in range(len(dv_nam)):
        if i == 3:
            bb = 7
        g = findposinlist(dv_nam[i][1],definitions,0)
        if definitions[g][3] == 'u':
            list1 = [None] * 55
            list1[5] = dv_nam[i][0]
            list1[9] = '='
            list1[14] = dv_nam[i][1]
            str1 = build_sent(list1)
            str2 = name_sent(str1)
            list1[0] = str1
            list1[42] = str2
            list1[41] = 1
            all_sent.append(list1)

    m = -1
    g = (len(all_sent))
    while m < g - 1:
        m += 1
        if all_sent[m][45] > 2:
            all_sent.append(all_sent[m])
            del all_sent[m]
            m -= 1
            g -= 1
    for d in range(0,3):
        if d == 1:
            num = [19]
        elif d == 2:
            num = [9,14,48]
        m = -1
        while m < len(all_sent) -1:
            m += 1
            bool2 = False
            if m == 13 and d == 2:
                bb = 7
            if m > 100:
                easygui.msgbox('in the define function you are caught in an infinite loop')
            for i in num:
                if d == 2 and m == 11 and i == 9:
                    bb = 7
                str1 = all_sent[m][i]
                relat = all_sent[m][9]
                old_sent = all_sent[m][0]
                oldp = all_sent[m][42]
                nc = False
                if d == 0:
                    if str1 == 'there' and i == 5:
                        dummy = there(all_sent,m,tot_sent,def_sent)
                        m -= 1
                        break
            #this is for those sentences whose noun was once part of a relative pronoun
                    elif i == all_sent[m][45]:
                        str3 = findinlist(all_sent[m][i],tagged_nouns2,1,0)
                        if str3 == None:
                            all_sent.append(all_sent[m])
                            del all_sent[m]
                            m -= 1
                            break
                        all_sent[m][i] = str3
                        new_sent = build_sent(all_sent[m])
                        newp = name_sent(new_sent)
                        dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"rel pro")
                        all_sent[m][0] = new_sent
                        all_sent[m][42] = newp
                        all_sent[m][45] = None
                        m -= 1
                    elif str1 in pronouns or str1 in determinative:
                        bool1 = True
                        if str1 in pronouns:
                            str2 = "pronoun"
                        else:
                            str2 = "determinative"
                #not many is the one negated determinative which changes from plural to singular
                        if str1 == 'many' + ud and (all_sent[m][8] == 'not' + ui \
                            or all_sent[m][8] == 'not' + ui):
                            str1 = 'not' + ui + " " + 'many' + ud
                        definition = findinlist(str1,definitions,0,1)
                        if all_sent[m][0] not in def_sent:
                            dummy = def_rn(definition, str1,0, tot_sent, \
                                dv_nam, def_var, idf_var,words, all_sent,m,nc,str2,i)
                            def_sent.append(all_sent[m][0])
                            del all_sent[m]
                            m -= 1
                            break
                elif d == 1 and all_sent[m][i] != 'S' and all_sent[m][i] != None:
                    dummy = space_division(all_sent,m)
                elif d == 2:
                    adverb = False
                    id = False
                    kind = ""
                    if all_sent[m][43] != 'cc':
                        if relat in def_relat and i == 14:
                            definiendum = findinlist(str1,dv_nam,0,1)
                            if definiendum in atomic_relata:
                                bool2 = False
                            else:
                                bool2 = True
                            if "'" not in all_sent[m][14]:
                                kind = 'R'
                        elif i == 48 and all_sent[m][48] != None:
                            definiendum = str1
                            bool2 = True
                            adverb = True
                            kind = 'R'
                        elif relat == '=' and all_sent[m][41] == 1:
                            id = True
                            definiendum = all_sent[m][14]
                            all_sent[m][41] = None
                        elif i == 9 and relat != 'IA' and relat != 'IG' and relat != '=' \
                            and str1 not in atomic_relations:
                            definiendum = str1
                            bool2 = True
                            kind = 'R'
                        if (bool2 and isdefineable(all_sent[m]) and definiendum != None and \
                                definiendum != '') or id:
                            g = findposinlist(definiendum,definitions,0)
                            definition = definitions[g][1]
                            if definition == 'natural':
                                definition = "(c'=" + definiendum + ") & (d'=natural_whole) & ((bIGc') " + conditional \
                                + " (bIGd'))"
                                nc = True
                            pos = definitions[g][2]
                            circ = definitions[g][3]
                            circ2 = all_sent[m][43]
                            basic_molecule = definitions[g][4]
                #this prevents us from getting caught in an infinite loop.
                            if basic_molecule == 'b' and all_sent[m][9] == 'IG':
                                break
                            if circ2 == 'c':
                                circ += circ2
                            if (relat == 'IA' and pos == 'a') or (relat == 'IG' and pos == 'n') \
                                or pos == 'r' or pos == 'e' or pos == 's' or (relat== '=' and pos == 'n') or adverb or id:
                                if definition != None and all_sent[m][0] not in def_sent:
                                    def_sent.append(all_sent[m][0])
                                    if definiendum == 'INM':
                                        bb = 7
                                    dummy = def_rn(definition, definiendum,0,tot_sent,dv_nam,def_var, idf_var,\
                                        words,all_sent,m,nc,kind,i,circ)
                                    break

# if we state that something is not a concept then we need to falisfy that
    dummy = concept(all_sent,tot_sent,dv_nam,definitions,posp)
    list1 = id_sent(dv_nam)
    tot_sent.insert(id_num-1,[id_num,list1[0],list1[1],"",'id'])
    return

def concept(all_sent,tot_sent,dv_nam,definitions,posp):

    global sn
    str1 = ""
    list1 = [None] * 55
    for i in range(len(dv_nam)):
        if dv_nam[i][1] == 'concept':
            str1 = dv_nam[i][0]
        if str1 != "":
            for j in range(len(all_sent)):
                if all_sent[j][9] == 'IG' and all_sent[j][14] == str1:
                    str2 = all_sent[j][5]
                    con = findinlist(str2,dv_nam,0,1)
                    pos = findinlist(con,posp,0,1)
                    if pos == None:
                        bb = 7
                    if con != None:
                        if pos == 'a':
                            str4 = ' IA '
                        elif pos == 'n':
                            str4 = ' IG '
                        old = "(" + str2 + ' = ' + con + ")"
                        new = "(" + idf_var[0] + str4 + str2 + ")"
                        str4 = str4.strip()
                        list1[0] = new
                        list1[5] = idf_var[0]
                        list1[9] = str4
                        list1[14] = str2
                        list1[40] = False
                        del idf_var[0]
                        oldp = name_sent(old)
                        newp = name_sent(new)
                        list1[42] = newp
                        all_sent.append(list1)
                        str3 = old + " " + conditional + " " + new
                        str3p = oldp + " " + conditional + " " + newp
                        sn += 1
                        tot_sent.append([sn,str3,str3p,"","NC concept " + con,""])
                        return

def name_sent(str1):
    global prop_name
    global prop_var

    no_space = copy.copy(str1)
    if str1.find('~') > -1:
        no_space = str1.replace("~","")
        ng = '~'
    else:
        ng = ''
    no_space = remove_outer_paren(no_space)
    no_space = no_space.replace(" ","")

    h = findinlist(no_space,prop_name,1,0)
    if h != None:
        return ng + h
    else:
        prop_name.append([prop_var[0], no_space, str1])
        str2 = prop_var[0]
        del prop_var[0]
        return ng + str2

def assign_var(str1, str2, dv_nam):

    list1 = [str2, str1]
    dv_nam.append(list1)

def determ(idf_var, def_var, all_sent, tot_sent,words,dv_nam,m):

    global sn
    # def_det contains the determinatives and their definitions
    def_det = words[15]
    detm = words[2]
    # num = [3, 10, 16, 20, 24, 28, 32]
    list1 = copy.deepcopy(all_sent[m][46])
    num = [3,10]
    for i in range(len(all_sent[m][46])):
        k = all_sent[m][46][i]
        if k in num:
            str1 = all_sent[m][k]
            if str1 in detm:
                defin = findinlist(str1,def_det,0,2)
                all_sent[m][k] = None
                list1.remove(k)
                if i == 10:
                    if "z~Rc" in defin:
                        defin = defin.replace("z~Rc","c~Rz")
                    else:
                        defin = defin.replace("zRc","cRz")
                dummy = def_rn(defin,str1,tot_sent, dv_nam,def_var,\
                               idf_var)
    all_sent[m][46] = list1


def insert_space(str, integer):
    return str[0:integer] + ' ' + str[integer:]

def space_words(str1):
    # this function place a space between variables and relations
    # so as to make it easier to categorize words in a sentence
    str1 = str1.replace("~", " ~ ")
    str1 = str1.replace("=", " = ")
    i = 0
    while i + 1 < len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        nxt_str = str1[(i + 1):(i + 2)]
        if nxt_str.isupper() == True and temp_str.islower() == True:
            str1 = insert_space(str1, i + 1)
        elif nxt_str.islower() == True and temp_str.isupper() == True:
            str1 = insert_space(str1, i + 1)
        elif temp_str == "'" and nxt_str.isupper() == True:
            str1 = insert_space(str1, i + 1)
    return str1


def id_def(list1):
    # this function picks out that variables in the id sentences of the
    # definition

    list2 = []
    has_plural = False
    for i in range(len(list1[0])):
        if os(list1[0][i]) and "=" in list1[0][i]:
            str1 = list1[0][i]
            g = str1.find("=")
            var = str1[1:g]
            wrd = str1[g+1:-1]
            if isvariable(var):
                if wrd == 'plural_form':
                    has_plural = True
                if not isvariable(wrd):
                    list2.append([var,wrd])
    return [list2,has_plural]

def is_atomic(list1, atomic_relations):

    relat = [9,15,19,23,27,31]
    must_be_blank = [2,3,4,6,7,10,11,12,16,17,20,21,24,25,28,29,32,33]
    noun = [5,14,18,26,30,34]
    for i in range(len(list1)):
        str1 = list1[i]
        if str1 != None:
            if i in relat:
                if str1 in atomic_relations:
                    pass
                else:
                    pass
            elif i in must_be_blank:
                return False
            elif i in noun:
                if not isvariable(str1):
                    return False
    return True


def isdefineable(list1):

    must_be_blank = [2,3,4,6,7,10,11,13,15,16,17,18,20,21,23,24,25,27,28,29,31,32,33,\
        35,36,49,50,51,52]
    must_be_variable = [5,14,22,30]

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None:
            if not isvariable(list1[i]):
                return False
    return True

def build_sent(list1):

    str1 = "("
    num = [37,2,47,3,4,5,35,49,50,51,52,6,7,8,9,48,12,10,13,14,36,11,15,16,17,18,19,20,21,22,23,24,25,26,27,28,\
           29,30,31,32,33,34,38]
    for i in num:
        temp_str = list1[i]
        if temp_str != None:
            if str1 == "(":
                str1 = str1 + temp_str
            else:
                str1 = str1 + " " + temp_str

    return str1 + ")"

def build_sent2(list1):

    for i in range(len(list1)):
        if i == 0:
            str1 = list1[i]
        else:
            str1 += " " + list1[i]
    return "(" + str1 + ")"

def build_sent_list(list1):
    # this list builder does not have the ~ separated from the sentence
    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i]
        else:
            str2 = str2 + ' & ' + list1[i]
    return str2

def build_sent_list2(list1):
    # this list builder has the ~ separated from the sentence
    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i][1] + list1[i][0]
        else:
            str2 = str2 + ' & ' + list1[i][1] + list1[i][0]
    return str2

def remove_duplicates(list1,i):

    list2 = []
    for j in range(len(list1)):
        if list1[j][i] in list2:
            del list1[j]
        else:
            list2.append(list1[j][i])
    return list1

def id_sent(list4):
    # this function turns the dv_nam into a string of conjuncts

    dv_nam = []
    # this loop removes duplicates in a multidimensional list
    for i in range(len(list4)):
        if list4[i] not in dv_nam:
            dv_nam.append(list4[i])

    prop2 = None
    str2 = None
    for i in range(len(dv_nam)):
        if len(dv_nam[i]) == 3:
            str3 = mini_e
        else:
            str3 = '='
        str1 = '(' + dv_nam[i][0] + str3 + dv_nam[i][1] + ')'
        prop1 = name_sent(str1)
        if str2 == None:
            str2 = str1
            prop2 = prop1
        else:
            str2 = str2 + ' & ' + str1
            prop2 = prop2 + " & " + prop1
    return [str2, prop2]

def cut_string(str1,str2):

    if str1.find(str2) == -1:
        return str1
    g = str1.find(str2)
    str1 = str1[:g]
    str1 = str1.strip()
    return str1

def remove_values_from_list(the_list, val):

    while val in the_list:
        the_list.remove(val)
    return the_list

def divide_sent(words, list2, def_var, idf_var,tot_sent,all_sent):

    global sn
    list3 = []
    redundant = words[21]
    conn = words[4]
    for i in range(len(list2)):
        str2 = list2[i][1]
        str3 = name_sent(str2)
        tot_sent.append([list2[i][0],str2,str3,"","","",""])
        list3.append(str2)

    # the following changes it is necessary that if p then q to if p then it is necessary q
    modals = ['possible','necessary','impossible']
    h = copy.copy(len(list3))
    i = -1
    while i < h - 1:
        i += 1
        list3[i] = list3[i].strip()
        list3[i] = list3[i].split(" ")
        list3[i] += ["("+tot_sent[i][1]+")",tot_sent[i][2]]
        for j in range(len(list3[i])):
            if list3[i][j] in modals:
                if list3[i][j+1] == 'that' and list3[i][j+2] == 'if':
                    old_sent = list3[i][-2]
                    old_p = tot_sent[i][2]
                    str1 = copy.copy(list3[i][j])
                    del list3[i][j+1]
                    del list3[i][j]
                    del list3[i][j-1]
                    del list3[i][j-2]
                    g = list3[i].index('then')
                    list3[i].insert(g+1,'it'+up)
                    list3[i].insert(g+2,'is'+ua)
                    list3[i].insert(g+3,str1)
                    list3[i].insert(g+4,'that')
                    del list3[i][-1]
                    del list3[i][-1]
                    new_sent = build_sent2(list3[i])
                    newp = name_sent(new_sent)
                    list3[i] += [new_sent,newp]
                    dummy = new_sentence2(old_sent,old_p,new_sent,newp,tot_sent,"modal transfer")
                    break
    str1 = ""
    str2 = ''
    bool1 = False

    for i in range(len(list3)):
        old_sent = list3[i][-2]
        old_p = list3[i][-1]
        for j in range(len(redundant)):
            str1 = redundant[j]
            if str1 in list3[i]:
                bool1 = True
                list3[i] = remove_values_from_list(list3[i],str1)
                if str2 == '':
                    str2 = str1
                else:
                    str2 += "," + str1
        if bool1:
            bool1 = False
            del list3[i][-1]
            del list3[i][-1]
            new_sent = build_sent2(list3[i])
            newp = name_sent(new_sent)
            list3[i] += [new_sent,newp]
            rule = "RD " + str2
            dummy = new_sentence2(old_sent,old_p,new_sent,newp,tot_sent,rule)

    rule = "df "
    g = len(list3)
    i = -1
    while i < g - 1:
        i += 1
        for j in range(len(list3[i])):
            if list3[i][j] in conn:
                str4 = list3[i][j]
                if list3[i][j] == 'not_follow':
                    str1 = nonseq
                else:
                    str1 = implies
                old_sent = list3[i][-2]
                old_p = list3[i][-1]
                del list3[i][-1]
                del list3[i][-1]
                ant = list3[i][:j]
                cons = list3[i][j+1:]
                ant_s = build_sent2(ant)
                cons_s = build_sent2(cons)
                antp = name_sent(ant_s)
                consp = name_sent(cons_s)
                ant += [ant_s,antp]
                cons += [cons_s,consp]
                list3.append(ant)
                list3.append(cons)
                rule += str4
                del list3[i]
                i -= 1
                new_sent = "(" + ant_s + " " + str1 + " " + cons_s + ")"
                new_p = "(" + antp + ' ' + str1 + ' ' + consp + ")"
                dummy = new_sentence2(old_sent,old_p,new_sent,new_p,tot_sent,rule)
                break

    g = len(list3)
    i = -1
    while i < g - 1:
        i += 1
        for j in range(len(list3[i])):
            if list3[i][j] == 'that':
                old_sent = list3[i][-2]
                old_p = list3[i][-1]
                del list3[i][-1]
                del list3[i][-1]
                ant = list3[i][:j]
                cons = list3[i][j+1:]
                cons_s = build_sent2(cons)
                consp = name_sent(cons_s)
                del ant[0]
                ant.insert(0,consp)
                ant_s = build_sent2(ant)
                antp = name_sent(ant_s)
                if consp in idf_var:
                    idf_var.remove(consp)
                    dv_nam.append([consp, cons_s,1])
                ant += [ant_s,antp]
                cons += [cons_s,consp]
                list3.append(ant)
                list3.append(cons)
                del list3[i]
                i -= 1
                dummy = new_sentence2(old_sent,old_p,ant_s,antp,tot_sent,'df that')
                break

    orel_pro = ['whom','who'+uo,'that'+uo]
    rel_pro = ['that' + us, 'which', 'who']
    g = len(list3)
    i = -1
    while i < g - 1:
        i += 1
        rule = 'df '
        for j in range(len(list3[i])):
            if list3[i][j] in orel_pro or list3[i][j] in rel_pro:
                rule += " " + list3[i][j]
                old_sent = list3[i][-2]
                old_p = list3[i][-1]

                del list3[i][-1]
                del list3[i][-1]
                ant = list3[i][:j]
                cons = list3[i][j+1:]
                str1 = ant[-1]
                tagged_nouns.append(str1)
                if list3[i][j] in rel_pro:
                    cons.insert(0,str1)
                cons = categorize_words(words,cons,def_var,idf_var,all_sent,True)
                if list3[i][j] in orel_pro:
                    cons[14] = str1
                    cons[45] = 14
                    list4 = cons[46]
                    list4.append(14)
                    list4.sort()
                    cons[46] = list4
                else:
                    cons[45] = 5
                    list4 = cons[46]
                    list4.append(14)
                    list4.sort()
                    cons[46] = list4
                ant = categorize_words(words,ant,def_var,idf_var,all_sent,True)
                ant_s = build_sent(ant)
                antp = name_sent(ant_s)
                ant[0] = ant_s
                ant[42] = antp
                #this means that when the object obtains its variable then we must
                #save this info and use it when we reduce the consequent
                ant[45] = 1
                cons_s = build_sent(cons)
                cons[0] = cons_s
                consp = name_sent(cons_s)
                cons[42] = consp
                list3.append(ant)
                list3.append(cons)
                del list3[i]
                i -= 1
                new_sent = "(" + ant_s + " & " + cons_s + ")"
                newp = "(" + antp + " & " + consp + ")"
                dummy = new_sentence2(old_sent,old_p,new_sent,newp,tot_sent,rule)
                break

    for i in range(len(list3)):
        if len(list3[i]) != 55:
            sent = list3[i][-2]
            sentp = list3[i][-1]
            del list3[i][-1]
            del list3[i][-1]
            list3[i] = categorize_words(words,list3[i],def_var,idf_var,all_sent,True)
            list3[i][0] = sent
            list3[i][42] = sentp
        all_sent.append(list3[i])
    return

def new_sentence2(str1,str3,str2,str4,tot_sent,rule,anc1="",conn = iff):

    global sn
    str5 = str1 + " " + conn + " " + str2
    str6 = str3 + " " + conn + " " + str4
    bool1 = check_dimension(tot_sent,1,str5)
    if not bool1:
        sn += 1
        if sn == 8:
            bb = 7
        tot_sent.append([sn,str5,str6,"",rule,anc1,""])

def division(tot_sent, all_sent):

    num = [35,36,52]
    b = len(all_sent)
    for k in range(0,3):
        if k == 1:
            num = [49,15,23,31]
        elif k == 2:
            num = [4,13,17,21,25,33]
        for m in range(0,b):
            old_sent = all_sent[m][0]
            oldp = all_sent[m][42]
            for i in num:
                if (k == 0 and all_sent[m][i] != None)  or (k ==1 and all_sent[m][i] == 'OF')\
                or (k ==2 and all_sent[m][i] != None):
                    list1 = [None] * 55
                    if k == 0:
                        rule = "CIA"
                        if i == 35:
                            j = 5
                        elif i == 36:
                            j = 14
                        elif i == 52:
                            j = 51
                        str1 = all_sent[m][j]
                        all_sent[m][j] = all_sent[m][i]
                        list1[14] = str1
                        list1[5] = all_sent[m][i]
                        list1[9] = "IG"
                        all_sent[m][i] = None
            # I'm not sure what I was thinking here so we will ignore this for now
                    elif k == 1:
                        rule = "RDA"
                        if i == 49:
                            a = 5
                            c = 51
                        elif i == 15:
                            a = 14
                            c = 18
                        elif i == 23:
                            a = 22
                            c = 26
                        elif i == 31:
                            a = 30
                            c = 34
                        relat = all_sent[m][i]
                        obj = all_sent[m][c]
                        subj = all_sent[m][a]
                        list1[5] = subj
                        list1[9] = relat
                        list1[14] = obj
                        all_sent[m][i] = None
                        all_sent[m][c] = None
                    elif k == 2:
                        rule = 'adj elim'
                        if all_sent[m][8] != None or all_sent[m][12] != None:
                            str7 = "~"
                            all_sent[m][8] = None
                            all_sent[m][12] = None
                        else:
                            str7 = ""
                        list1[8] = str7
                        list1[5] = all_sent[m][5]
                        list1[9] = "IA"
                        list1[14] = all_sent[m][i]
                        all_sent[m][i] = None
                    dummy = new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule)

    return

def new_sent_prelim(old_sent,oldp,all_sent,list1,m,rule,kind = 1):

    sent2 = build_sent(all_sent[m])
    sent2p = name_sent(sent2)
    all_sent[m][0] = sent2
    all_sent[m][42] = sent2p
    sent1 = build_sent(list1)
    sent1p = name_sent(sent1)
    list1[0] = sent1
    list1[42] = sent1p
    list2 = copy.deepcopy(list1)
    all_sent.append(list2)
    if kind == 1:
        new_sent = "(" + sent1 + " & " + sent2 + ")"
        newp = "(" + sent1p + " & " + sent2p + ")"
        conn = iff
    else:
        new_sent = sent1
        newp = sent1p
        conn = conditional
    dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,rule,"",conn)

def space_division(all_sent,m):

    list1 = [None] * 55
    list1[5] = all_sent[m][5]
    list1[9] = all_sent[m][19]
    list1[14] = all_sent[m][22]
    list1[53] = all_sent[m][53]
    all_sent[m][19] = None
    all_sent[m][22] = None
    dummy = new_sent_prelim(all_sent[m][0],all_sent[m][42],all_sent,list1,m,"RDB",1)

def mult_defnd(list1,def_num):

    for i in range(len(list1)):
        num = list1[i][0]
        if def_num + '12' == num:
            return True
    return False

def indefiniendum(def_num,sent_num,multiple):

    if multiple:
        if sent_num[:-1] == def_num + '1':
            return True
        else:
            return False
    else:
        if sent_num == def_num + '1':
            return True
        else:
            return False

def in_dv(list1,dv_nam):

    if not list1[9] == "=":
        return False
    else:
        bool1 = check_dimension(dv_nam,1,list1[14])
        return bool1

def prop_type(paren_num,gparen_num,paren_conn,gparen_conn,sent_num,sent_type):

    global sn
    str2 = None
    str1 = None #kyle Added new line here
    if sent_type == "1cj" or sent_type == None:
        if paren_conn == '&' and gparen_conn == xorr:
            str1 = "dc"
            str2 = str(sn) + paren_num
        elif paren_conn == xorr and gparen_conn == iff:
            str1 = 'd'
        elif paren_conn == "&" and gparen_conn == conditional:
            if paren_num[-2] == "1":
                str1 = "an"
            else:
                str1 = "cn"
        elif paren_conn == conditional:
            if sent_num[-1] == "1":
                str1 = "an"
            else:
                str1 = "cn"
        elif paren_conn == iff:
            str1 = '1cj'
        elif paren_conn == "&" and gparen_conn == iff:
            str1 = '1cj'
        elif paren_conn == xorr and gparen_conn == conditional:
            str1 = 'cn'
        else:
            easygui.msgbox('there is a prop type that is not coded for')
    else:
        str1 = sent_type

    return [str1,str2]

def def_rn(definition, definiendum,e, tot_sent,  dv_nam, def_var, idf_var, \
           words,all_sent,m,nc,kind = "",k=0,circ = ""):
    # this function renames the variables in a definition

    global sn
    global plural_c
    global definite

    used_idf = []
    #this is for those determinatives which have negations in their definitions where
    #the sentences has an R variable
    identical_det = ["only","anything_except","anyone_except","no","many"+um,"many"+un,\
        "no" + us]
    if definiendum == 'any':
        bb = 7
    new_idf = []
    if definiendum in identical_det:
        ident_det = True
    else:
        ident_det = False
    if nc:
        rule = 'nc ' + definiendum
    else:
        rule = 'df ' + definiendum
    match_dv = []
    sent_type = all_sent[m][53]
    str1 = copy.copy(definition)
    #if bool1 is false then there is a series of conjuncts that need to be removed from
    # the definition
    if kind == "" or kind == 'R':
        def_info = find_sentences(definition,True)
    else:
        def_info = find_sentences(definition)
    def_loc = def_info[7]
    def_num = def_info[4][def_loc][0]
    ld = len(def_num)
    list1 = []
    list1 = id_def(def_info)
    dv = list1[0]
    has_plural = list1[1]
    # we now must match the definite variables in the definition to the definite variables
    # already assigned
    list1 = []
    syn_det = ["no_one_except","any"+un]
    for i in range(len(dv)):
        temp_str = dv[i][1]
        for j in range(len(dv_nam)):
            dvn_temp = dv_nam[j][1]
            if dvn_temp == temp_str:
                telist7 = [dv[i][0],dv_nam[j][0]]
                if telist7 not in match_dv:
                    match_dv.append(telist7)
                    break
        else:
            if dv[i][0] not in def_var:
                telist7 = [dv[i][0], def_var[0]]
                match_dv.append(telist7)
                list1.append([def_var[0], temp_str])
                del def_var[0]
            else:
                telist7 = [dv[i][0], temp_str]
                list1.append(telist7)
                def_var.remove(dv[i][0])

    if kind == 'pronoun':

        str1 = findinlist(definiendum,dv_nam,1,0)
        if str1 == None:
            all_sent[m][k] = def_var[0]
            match_dv.append(["c'",def_var[0]])
            dv_nam.append([def_var[0],definiendum])
            del def_var[0]
        else:
            all_sent[m][k] = str1
            match_dv.append(["c'",str1])
        # when constructing definitions of personal pronouns or of determinatives the object of the IG relation
        # must be b and the subject must be z
    elif kind == 'determinative':
        all_sent[m][k] = ""
        if k == 10:
            j = 14
        else:
            j = k + 2
        ovar = all_sent[m][j]
        match_dv.append(["b",all_sent[m][j]])
        if definiendum == 'the':
            str1 = all_sent[m][j]
            str3 = findinlist(str1,dv_nam,0,1)
            str2 = findinlist(str3,definite,1,0)
            if str2 == None:
                match_dv.append(["z'",def_var[0]])
                definite.append([def_var[0],str3])
                all_sent[m][j] = def_var[0]
                del def_var[0]
            else:
                all_sent[m][j] = str2
                match_dv.append(["z'",str2])
        elif definiendum not in syn_det:
            match_dv.append(["z",idf_var[0]])
            match_dv.append(["y",idf_var[1]])
            match_dv.append(["x",idf_var[2]])
            match_dv.append(["w",idf_var[3]])
            all_sent[m][j] = idf_var[0]
            del idf_var[0]
            del idf_var[0]
            del idf_var[0]
            del idf_var[0]

        if all_sent[m][45] == 1:
            str5 = findinlist(ovar,dv_nam,0,1)
            if str5 in tagged_nouns:
                tagged_nouns2.append([all_sent[m][j],str5])

    sdefinition = def_info[0][def_loc]
    def_sent = []
    rename = []
    skel_string = def_info[5]
    exception = []
    not_many = False
    first_in_def = [def_num+"1",def_num + "11"]
    temp_plural = []
    heir_num = []
    spec_var = ['z','y','x','w']

    #as we loop through the sentences they must be in the definition which is the point of n
    for i in range(len(def_info[0])):
        n = def_info[4][i][0][:ld]
        if os(def_info[3][i]) == True and n == def_num:
            temp_str = space_words(def_info[3][i])
            temp_str = temp_str.replace("(","")
            temp_str = temp_str.replace(")","")
            telist7 = categorize_words(words,temp_str,def_var,idf_var,all_sent)
            bool1 = False
            bool2 = False
            if kind != "" and kind != 'R' and telist7[9] == "R":
                n = len(def_sent)
                exception.append(n)
                if telist7[3] != None:
                    temp_det = telist7[3]
                    has_detrm = True
                else:
                    has_detrm = False
                if ident_det:
                    neg1 = telist7[8]
                    neg2 = telist7[12]
                    if telist7[5] == 'b':
                        bool1 = True
                    if telist7[5] == 'z':
                        bool2 = True
                str2 = ''
                for p in range(2,36):
        # if the variable in the original definition is z,y,x,w then that must
        # go into the new definition in its proper place
                    if telist7[p] in spec_var:
                        str2 = findinlist(telist7[p],match_dv,0,1)
                    if p == j and str2 != "":
                        telist7[p] = str2
                    else:
                        telist7[p] = all_sent[m][p]
        #not many is the one negated determinative which is defined in this way and its
        #negation is removed in the definiens
                if definiendum == 'not' + ui + ' ' + 'many' + ud:
                    if all_sent[m][8] == 'not' + ui and def_info[4][i][0] not in first_in_def:
                        telist7[8] = ""
                        telist7[k] = 'exactly_one'
                        str2 = findinlist(ovar,plural_c,0,1)
                        telist7[j] = str2
                    if def_info[4][i][0] in first_in_def:
                        telist7[j] = ovar
                        if all_sent[m][47] == 'not' + ui:
                            telist7[47] = 'not' + ui
                        elif all_sent[m][8] == 'not' + ui:
                            telist7[8] = 'not' + ui
                            telist7[47] = None
                            not_many = True
        #just in case the list has a tagged noun
                telist7[45] = all_sent[m][45]
        # for the determinatives which have negations in their definition then we need
        # to do something special
                if ident_det:
                    telist7[8] = neg1
                    telist7[12] = neg2
        # the determinatives which have an identity statement in them behave differently
        # these are 'only' and 'anything except'
                if bool1:
                    telist7[j] = ovar
                if has_detrm:
                    telist7[k] = temp_det
                if definiendum == 'everything_except' + up and i == 13:
                    telist7[8] = "~"
                    if 'y' in idf_var:
                        telist7[j] = 'y'
                        match_dv.append(['y','y'])
                    else:
                        telist7[j] = idf_var[0]
                        match_dv.append(['y',idf_var[0]])
                        del idf_var[0]
                if (definiendum == 'all' and i == 4) or (definiendum == 'only' + up and i==9):
                    telist7[j] = 'd'
                    exception.remove(n)
                if definiendum == 'any' + un and i == 2:
                    telist7[8] = ""
                    telist7[10] = "no"
                if bool2:
                    str2 = findinlist("z",match_dv,0,1)
                    telist7[j] = str2
        # if the sentence is first then we must restor the definiendum to it
                if def_info[4][i][0] in first_in_def and not not_many:
                    telist7[k] = definiendum
        # what this does is it puts the original variable back into the definiendum
                    if kind != "pronoun":
                        telist7[j] = ovar
                str1 = build_sent(telist7)
                sdefinition = sdefinition.replace(def_info[3][i],str1)
            else:
                str1 = build_sent(telist7)
    # if a def sent is part of the defiendum then it does not have to be added to the all
    # all_sent list
            if def_info[4][i][0][:-1] in first_in_def or def_info[4][i][0] in first_in_def:
                telist7[40] = True
            else:
                telist7[40] = False
            telist7[0] = str1
            sent_num = def_info[4][i][0]
            paren_num = def_info[4][i][0][:-1]
            gparen_num = def_info[4][i][0][:-2]
            paren_conn = findinlist(paren_num,def_info[4],0,1)
            gparen_conn = findinlist(gparen_num,def_info[4],0,1)
            if paren_conn == None:
                bb = 7
            list2 = prop_type(paren_num,gparen_num,paren_conn,gparen_conn,sent_num,sent_type)
            telist7[53] = list2[0]
            telist7[54] = list2[1]
            telist7[44] = def_info[6][i][1]
            heir_num.append(def_info[4][i][0])
            for s in range(0,3):
                telist7.append(None)
            def_sent.append(telist7)

    # the purpose of this is that the subject of the definiendum must match the subject
    # of the osent to be defined.  if its a relation then the object must also match
    if kind == "" or kind == "R":
        mvar = []
        bool1 = mult_defnd(def_info[4],def_num)
        for i in range(len(def_sent)):
            bool2 = indefiniendum(def_num,heir_num[i],bool1)
            if bool2:
                if heir_num[i] in first_in_def:
                    match_dv.append([def_sent[i][5],all_sent[m][5]])
                    if kind == "R":
                        match_dv.append([def_sent[i][14],all_sent[m][14]])
                    # mvar.append([def_sent[i][5],all_sent[m][5]])
                    break

    #if the definiendum is many-o then its object variable needs to be matched
    if definiendum == 'many' + uo:
        match_dv.append(['c',all_sent[m][14]])


    dv_nam += list1
    num = [5,14,18,22,26,30,34]
    # the point of the exception list is that we do not change certain sentences in the
    # definiens if we are analyzing a pronoun or determinative
    str2 = ""

    for w in range(0,2):
        for i in range(len(def_sent)):
            if i not in exception:
                for j in num:
                    temp_str = def_sent[i][j]
                    if temp_str != None:
                        str3 = findinlist(temp_str,match_dv,0,1)
                        if str3 != None and temp_str != str3:
                            def_sent[i][j] = str3
                            str2 = '(' + temp_str + mini_c + str3 + ')'
                            if str2 not in rename and str2 != "":
                                rename.append(str2)
                            str2 = ""
                        elif temp_str == str3:
                            pass
                        else:
                            if isvariable(temp_str) and temp_str.find("'") == -1 and w==0:
                                bool1 = check_dimension(new_idf,0,temp_str)
                                if bool1:
                                    str1 = findinlist(temp_str,new_idf,0,1)
                                    def_sent[i][j] = str1
                                elif temp_str in used_idf:
                                    pass
                                elif temp_str not in idf_var:
                                    k = 0
                                    while k + 1 <= len(new_idf):
                                        if temp_str == new_idf[k][0]:
                                            def_sent[i][j] = new_idf[k][1]
                                            break
                                        k += 1
                                    else:
                                        telist7 = [def_sent[i][j], idf_var[0]]
                                        new_idf.append(telist7)
                                        str2 = "(" + def_sent[i][j] + mini_c + idf_var[0] + ")"
                                        def_sent[i][j] = idf_var[0]
                                        del idf_var[0]
                                        rename.append(str2)
                                        str2 = ""
                                elif temp_str in idf_var:
                                    idf_var.remove(temp_str)
                                    used_idf.append(temp_str)

        if has_plural and w == 0:
            match_dv = []
            for i in range(len(def_sent)):
    #When constructing definitions which have a plural, 'of' must be in the definiendum
                if heir_num[i][:-1] in first_in_def \
                    and def_sent[i][9] == "OF":
                    temp_plural = [def_sent[i][5],def_sent[i][14]]
                    str1 = temp_plural[0]
                    new1 = findinlist(str1,plural_c,0,1)
                    old1 = temp_plural[1]
                    if new1 == None:
                        bool1 = check_dimension(plural_c,0,temp_plural[0])
                        if bool1:
                            old1 = temp_plural[1]
                            new1 = findinlist(temp_plural[0],plural_c,0,1)
                    if new1 == None or old1 == None:
                        easygui.msgbox("your plurals are not right")
                    match_dv.append([old1,new1])
                    exception = []
        else:
            break



    # definition = sdefinition
    # we now replace the skel string with the new sentences, to get the true definition
    skel_string = def_info[5]
    skel_wid = def_info[8]
    skel_string2 = def_info[5]

    for i in range(len(def_sent)):
        if def_sent[i][9] == "OF" and has_plural:
            plural_c.append([def_sent[i][5],def_sent[i][14]])

        str2 = build_sent(def_sent[i])
        def_sent[i][0] = str2
        skel_string = skel_string.replace(def_sent[i][44], str2)
        str1 = name_sent(str2)
        def_sent[i][42] = str1
        skel_string2 = skel_string2.replace(def_sent[i][44], str1)

    str3 = skel_string2
    if kind == "" or kind == 'R':
        sn += 1
        definition = remove_outer_paren(definition)
        if rename != []:
            definition += " !"
            tot_sent.append([sn, definition, "","", rule])
            str2 = build_sent_list(rename)
            if str2 != None:
                str2 = str2 + ' !'
                sn += 1
                tot_sent.append([sn, str2, "","", 'RN'])
            bool1 = check_dimension(tot_sent,1,skel_string)
            if not bool1:
                sn += 1
                tot_sent.append([sn, skel_string, str3,"", 'esub',sn-1,sn-2])
        else:
            tot_sent.append([sn, definition, skel_string2,"", rule])
    else:
        bool1 = check_dimension(tot_sent,1,skel_string)
        if not bool1:
            sn += 1
            tot_sent.append([sn, skel_string, str3,"", rule])
    # right now we're simply deleting the first sentence of the def_sent since
    # we have already defined that but in the future when we work with more
    # complex definienda we will have to change this
    #end3
    dummy = nonidentical(def_sent)

    for i in range(len(def_sent)):
        bool1 = check_dimension(all_sent,0,def_sent[i][0])
        bool2 = in_dv(def_sent[i],dv_nam)
        # it used to be that the sentence had to not have a plural which means
        # that 41 had to be false
        if bool1 == False and bool2 == False and def_sent[i][40] == False:
            def_sent[i][43] = circ
            all_sent.append(def_sent[i])
    return


def nonidentical(list1):

    list2 = []
    num = [5,14,22,26,30,34]
    for i in range(len(list1)):
        for j in num:
            if list1[i][j] != None and isvariable(list1[i][j]):
                if list1[i][j] not in list2:
                    list2.append(list1[i][j])
    if len(list2) > 1:
        non_id.append(list2)


def plurals(tot_sent, all_sent,  m, words, dv_nam, def_var, idf_var):

    global plural_c
    all_sent = remove_duplicates(all_sent,0)
    pnouns = words[26]
    definitions = words[16]
    list1 = copy.deepcopy(all_sent[m][46])
    num = [5,14,18,22,26,30,34,36]
    bool1 = False
    used_plural = []
    g = len(all_sent)
    for m in range(0,g):
        for i in num:
            str1 = all_sent[m][i]
            str2 = findinlist(str1, dv_nam,0,1)
            if check_dimension(pnouns,0,str2) and str2 not in used_plural:
                used_plural.append(str2)
                singular = findinlist(str2,pnouns,0,1)
                pdef = "((b'=" + str2 + ") " + iff + " ((b'IGc') & (b'OFd'))) & (d'=" \
                + singular + ") & (c'=plural_form)"
                dummy = def_rn(pdef,str2,0,tot_sent,dv_nam,def_var,\
                        idf_var,words,all_sent,m,"",k)

def categorize_words(words, str2, def_var, idf_var,all_sent,islist=False):

    global sn
    # if isinstance(str2, (list,tuple)):
    #     return str2
    word_types = []
    relations = words[18]
    relations2 = words[19]
    relat = words[6]
    srelat = words[7]
    trelat = words[8]
    grelat = ['of'+ug]
    has_plural = False
    bool1 = False
    num = [0,1,2,3,4,5,6,7,8,9,10,11,12,18,24,25]

    if islist:
        list1 = str2
        osent = ""
    else:
        osent = copy.copy(str2)
        list1 = str2.split(' ')

    for i in list1:
        str9 = i
        if isvariable(str9) == True:
            word_types.append([str9, 'n'])
            if "'" in str9 and str9 in def_var:
                def_var.remove(str9)
            elif str9 in idf_var:
                idf_var.remove(str9)
        elif str9 == ' ':
            pass
        elif str9 == 'plural_form':
            has_plural = True
        elif str9 in grelat:
            word_types.append([str9, 'g'])
        elif str9 == 'not' + ui:
            word_types.append([str9,'ni'])
        elif str9 == "~":
            word_types.append([str9, 'm'])
        elif isinstance(str9,int):
            word_types.append([str9, 'n'])
        else:
            for k in range(len(num)):
                if bool1:
                    bool1 = False
                    break
                j = num[k]
                if j == 18 and str9 == "IG":
                    pp = 7
                if i in words[j]:
                    bool1 = True
                    if j == 0:
                        word_types.append([str9, 'a'])
                    elif j == 1:
                        word_types.append([str9, 'c'])
                    elif j == 2:
                        word_types.append([str9, 'd'])
                    elif j == 3:
                        word_types.append([str9, 'e'])
                    elif j == 4:
                        word_types.append([str9, 'l'])
                    elif j == 5:
                        word_types.append([str9, 'n'])
                    elif j == 6:
                        word_types.append([str9, 'r'])
                    elif j == 7:
                        word_types.append([str9, 's'])
                    elif j == 8:
                        word_types.append([str9, 't'])
                    elif j == 9:
                        word_types.append([str9, 'u'])
                    elif j == 10:
                        word_types.append([str9, 'b'])
                    elif j == 11:
                        word_types.append([str9, 'm'])
                    elif j == 12:
                        word_types.append([str9, 'd'])
                    elif j == 18:
                        h = findin1dlist(str9,relations)
                        str8 = relations2[h]
                        if str8 in relat:
                            word_types.append([str9, 'r'])
                        elif str8 in srelat:
                            word_types.append([str9, 's'])
                        elif str8 in trelat:
                            word_types.append([str9, 't'])
                        else:
                            easygui.msgbox('you did not code your relations properly')
                    elif j == 24:
                        word_types.append([str9, 'p'])
                    elif j == 25:
                        word_types.append([str9, 'q'])
            else:
                easygui.msgbox(str9 + ' is not in your dictionary')

    # a, c, d, e, l, n, r, s, t, u, b, m, q
    list1_cat = [None] * 55
    relation_type = None
    list2 = []
    list3 = []
    subo = False

    for i in range(0, len(word_types)):
        word = word_types[i][0]
        pos = word_types[i][1]
        if word == '':
            pass
        elif (pos == 'd' or pos == 'q') and relation_type == None:
            list1_cat[3] = word
            list2.append(3)
        elif pos == 'a' and relation_type == None:
            list1_cat[4] = word
            list2.append(4)
        elif pos == 'ni' and list1_cat[3] == None and list1_cat[5] == None:
            list1_cat[47] = word
            list2.append(47)
        elif (pos == 'n' or pos == 'p') and relation_type == None and list1_cat[5] == None:
            list1_cat[5] = word
            list2.append(5)
        elif pos == 'n' and relation_type == None and list1_cat[5] != None:
            list1_cat[35] = word
            list2.append(35)
        elif pos == 'n' and relation_type == 'r' and list1_cat[14] != None:
            list1_cat[36] = word
            list2.append(36)
        elif pos == 'g' and relation_type == None:
            list1_cat[49] = word
            list2.append(49)
            relation_type = 'g'
        elif pos == 'd' and relation_type == 'g':
            list1_cat[50] = word
            list2.append(50)
        elif pos == 'n' and relation_type == 'g' and list1_cat[51] == None:
            list1_cat[51] = word
            list2.append(51)
        elif pos == 'n' and relation_type == 'g' and list1_cat[51] != None:
            list1_cat[52] = word
            list2.append(52)
        elif pos == 'b' and relation_type == None:
            list1_cat[7] = word
            list2.append(7)
        elif pos == 'm' and relation_type == None:
            list1_cat[8] = word
            list2.append(8)
        elif pos == 'ni' and (list1_cat[3] != None or list1_cat[5] != None):
            list1_cat[8] = word
            list2.append(8)
        elif pos == 'm' and relation_type == "r":
            list1_cat[12] = word
            list2.append(12)
        elif pos == 'r' and (relation_type == None or relation_type == 'g'):
            list1_cat[9] = word
            list2.append(9)
            relation_type = 'r'
            if list1_cat[5] == None and list1_cat[4] != None:
                list1_cat[5] = list1_cat[4]
                list1_cat[4] = None
                list2.remove(4)
                list2.append(5)
        elif (pos == 'm') and relation_type == "r":
            list1_cat[12] = word
            list2.append(12)
        elif (pos == 'd' or pos == 'q') and relation_type == 'r':
            list1_cat[10] = word
            list2.append(10)
        # this line of code must be first because if the word is an adjective
        # and the relation is IA then it must go in slot 14
        elif pos == 'a' and relation_type == 'r' and (list1_cat[9] == 'IA' \
        or list1_cat[9] == 'is' + ua or list1_cat[9] == 'are'+ua):
            list1_cat[14] = word
            list2.append(14)
        elif pos == 'a' and relation_type == 'r':
            list1_cat[13] = word
            list2.append(13)
        elif (pos == 'n' or pos == 'p') and relation_type == 'r' and list1_cat[14] == None:
            list1_cat[14] = word
            list2.append(14)
        elif pos == 'e' and relation_type == 'r':
            list1_cat[48] = word
            list2.append(48)
        elif pos == 'r' and relation_type == 'r':
            list1_cat[15] = word
            relation_type = 'r2'
            list2.append(15)
        elif (pos == 'd' or pos == 'q') and relation_type == 'r2':
            list1_cat[16] = word
            relation_type = 'r2'
            list2.append(16)
        elif pos == 'a' and relation_type == 'r2':
            list1_cat[17] = word
            relation_type = 'r2'
            list2.append(17)
        elif (pos == 'n' or pos == 'p') and relation_type == 'r2':
            list1_cat[18] = word
            relation_type = 'r2'
            list2.append(18)
        elif pos == 's' and relation_type == None:
            relation_type = 's'
            list1_cat[9] = word
            list2.append(9)
            relation_type = 'r'
        elif pos == 's':
            relation_type = 's'
            list1_cat[19] = word
            list2.append(19)
        elif (pos == 'd' or pos == 'q') and relation_type == 's':
            list1_cat[20] = word
            list2.append(20)
        elif pos == 'a' and relation_type == 's':
            list1_cat[21] = word
            list2.append(21)
        elif (pos == 'n' or pos == 'p') and relation_type == 's':
            list1_cat[22] = word
            list2.append(22)
        elif pos == 't' and relation_type == None:
            relation_type = 'r'
            list1_cat[9] = word
            list2.append(9)
        elif pos == 't':
            relation_type = 't'
            list1_cat[27] = word
            list2.append(27)
        elif (pos == 'd' or pos == 'q') and relation_type == 't':
            list1_cat[28] = word
            list2.append(28)
        elif pos == 'a' and relation_type == 't':
            list1_cat[29] = word
            list2.append(29)
        elif (pos == 'n' or pos == 'p') and relation_type == 't':
            list1_cat[30] = word
            list2.append(30)
        elif pos == 'b':
            list1_cat[7] = word
        elif pos == 'm':
            if relation_type == None:
                list1_cat[8] = word
                list2.append(8)
            elif relation_type == 'r':
                list1_cat[13] = word
                list2.append(13)
        else:
            easygui.msgbox('you did not categorize the word ' + word)

    list2.sort()
    list1_cat[46] = list2
    list1_cat[0] = osent
    list1_cat[41] = has_plural
    return list1_cat

def build_sent_name(prop_name):
    str1 = ''
    str2 = ''
    list1 = []

    for i in range(len(prop_name)):
        str3 = remove_outer_paren(prop_name[i][2])
        str3 = str3.replace("~","")
        str1 = '(' + prop_name[i][0] + mini_e + str3 + ')'
        if len(str2) == 0 and len(str1) > 57:
            list1.append(str1)
        elif (len(str2) + len(str1)) > 57:
            list1.append(str2)
            str2 = str1
            if i + 1 == len(prop_name):
                list1.append(str2)
        elif (len(str2) + len(str1)) < 57:
            if len(str2) == 0:
                str2 = str1
            else:
                str2 = str2 + ' & ' + str1
            if i + 1 == len(prop_name):
                list1.append(str2)
    return list1



def syn(tot_sent, all_sent, words,m):

    all_sent = remove_duplicates(all_sent,0)
    global sn
    bool1 = False
    synon = words[14]
    syn_pairs = words[13]
    m = -1
    while m < len(all_sent) -1:
        m += 1
        list3 = copy.deepcopy(all_sent[m][46])
        old_sent = all_sent[m][0]
        oldp = all_sent[m][42]
        anc1 = ""
        for i in list3:
            str1 = all_sent[m][i]
            if str1 == 'case':
                bb = 7
            if str1 in synon:
                for j in range(len(syn_pairs)):
                    if str1 == syn_pairs[j][0]:
                        bool1 = True
                        rule = 'df ' + str1
                        str5 = syn_pairs[j][2]
                        all_sent[m][i] = syn_pairs[j][1]
                        u = findinlist(str5,tot_sent,1,0,True)
                        if u == -1:
                            bool1 = True
                            str5v = name_sent(syn_pairs[j][2])
                            s = findinlist(str5,tot_sent,1,0)
                            sn += 1
                            if s == None:
                                s = sn
                            if anc1 == "":
                                anc1 = str(s)
                            else:
                                anc1 += "," + str(s)
                            tot_sent.append([sn, str5, str5v,"", rule])
                        else:
                            s = tot_sent[u][0]
                            str5v = name_sent(syn_pairs[j][2])
                            if anc1 == "":
                                anc1 = str(s)
                            else:
                                anc1 += "," + str(s)

        if bool1:
            new_sent = build_sent(all_sent[m])
            newp = name_sent(new_sent)
            all_sent[m][0] = new_sent
            all_sent[m][42] = newp
            dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,"isub",anc1)
            all_sent[m][46] = list3
            bool1 = False
    return

def ax_ent(dv_nam,  tot_sent):

    global sn
    list2 = []
    bool1 = False
    for i in range(len(dv_nam)):
        if dv_nam[i][1] == 'thing':
            tvar = dv_nam[i][0]
            for j in range(len(all_sent)):
                if all_sent[j][45] == 'c' and all_sent[j][14] == tvar:
                    for k in range(len(all_sent)):
                        if all_sent[k][45] == 'a' and all_sent[k][9] == "IG":
                            bool1 = True
                            nvar = all_sent[k][5]
                            list1 = [None] * 55
                            list1[5] = nvar
                            list1[9] = "IG"
                            list1[14] = tvar
                            str1 = build_sent(list1)
                            list1[0] = str1
                            sn += 1
                            str2 = name_sent(str1)
                            tot_sent.append([sn, str1, str2, "ax ENT", str(anc1), ""])
                            list2.append(list1)
    if bool1:
        all_sent = all_sent + list2
        return all_sent

def det_nouns(tot_sent, all_sent, words,  m):

    global sn
    dnouns = words[12]
    det_pairs = words[17]
    neg_det = words[27]
    spec_class = ['everyone','anything','everything']
    num = [3, 10]
    num2 = [47,8]
    m = -1
    anc1 = ''
    while m < len(all_sent) -1:
        m += 1
        for d in range(0,2):
            list3 = copy.deepcopy(all_sent[m][46])
            bool1 = False
            rule = "df "
            old_sent = all_sent[m][0]
            oldp = all_sent[m][42]
            if d == 0:
                for i in num2:
                    if (i == 47 or i == 8) and i in list3 and all_sent[m][i] == 'not'+ui and \
                    (check_dimension(neg_det,0,all_sent[m][3]) or check_dimension(neg_det,0,all_sent[m][10])):
                        bool1 = True
                        if all_sent[m][47] != None:
                            list3.remove(47)
                            q = 3
                            r = 5
                        else:
                            list3.remove(8)
                            q = 10
                            r = 14
                        if all_sent[m][q] == 'many' + ud:
                            break
                        if all_sent[m][q] in spec_class:
                            list3.append(r)
                        str1 = findinlist(all_sent[m][q],neg_det,0,1)
                        str2 = findinlist(all_sent[m][q],neg_det,0,2)
                        defin = findinlist(all_sent[m][q],neg_det,0,3)
                        all_sent[m][i] = None
                        rule += 'not' + ui + ' ' + all_sent[m][q]
                        if str2 == None or str2 == '':
                            all_sent[m][q] = str1
                        else:
                            all_sent[m][q] = str1
                            all_sent[m][r] = str2
                        g = findinlist(defin,tot_sent,1,0)
                        if g == None:
                            sn += 1
                            str3 = name_sent(defin)
                            tot_sent.append([sn,defin,str3,"",rule,"",""])
                            anc1 = sn
                            rule = 'isub'
                        else:
                            rule = 'isub'
                            anc1 = g
                        break
            else:
                for i in range(len(all_sent[m][46])):
                    k = all_sent[m][46][i]
                    if k in num:
                        for j in range(len(det_pairs)):
                            if det_pairs[j][2] == all_sent[m][k]:
                                bool1 = True
                                if rule == 'df ':
                                    rule = 'df ' + det_pairs[j][2]
                                else:
                                    rule = rule + ", " + det_pairs[j][2]
                                if k == 3:
                                    list3.append(5)
                                    all_sent[m][3] = det_pairs[j][0]
                                    all_sent[m][5] = det_pairs[j][1]
                                else:
                                    list3.append(14)
                                    all_sent[m][10] = det_pairs[j][0]
                                    all_sent[m][14] = det_pairs[j][1]
                                break
            if bool1:
                new_sent = build_sent(all_sent[m])
                newp = name_sent(new_sent)
                all_sent[m][0] = new_sent
                all_sent[m][42] = newp
                dummy = new_sentence2(old_sent,oldp,new_sent,newp,tot_sent,rule,anc1)
                list3.sort()
                all_sent[m][46] = list3
                bool1 = False

def print_sent_full(test_sent,p,tot_prop_name):
    # p = 30
    p += 2
    for i in range(len(test_sent)):
        # if i == 2:
        #     break
        for j in range(len(test_sent[i])):
            if i == 1 and j == 10:
                bb = 7
            if len(test_sent[i][j]) == 7 and test_sent[i][j][6] != "" and test_sent[i][j][6] != None:
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5]) + ',' + str(test_sent[i][j][6])
            elif len(test_sent[i][j]) == 7:
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5])
            elif len(test_sent[i][j]) == 6:
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5])
            elif len(test_sent[i][j]) == 5:
                str1 = test_sent[i][j][4]
            else:
                str1 = ""
            if j == 0:
                str1 = i
            w4.cell(row=p,column=2).value = test_sent[i][j][0]
            w4.cell(row=p,column=3).value = test_sent[i][j][1]
            w4.cell(row=p,column=4).value = str1
            p += 1
        p += 1

        list1 = build_sent_name(tot_prop_name[i])
        for j in range(len(list1)):
            w4.cell(row=p,column=3).value = list1[j]
            p += 1
        p += 1

        for j in range(len(test_sent[i])):
            # if j == 2:
            #     break
            if test_sent[i][j][2] != "":
                if test_sent[i][j][3] == "~" and not os(test_sent[i][j][2]):
                    str1 = "(" + test_sent[i][j][2] + ")"
                else:
                    str1 = test_sent[i][j][2]
                w4.cell(row=p,column=2).value = test_sent[i][j][0]
                w4.cell(row=p,column=3).value = test_sent[i][j][3] + str1
                p += 1
        p += 1
        w4.cell(row=p,column=3).value = 'irrelevant conjuncts:'
        str3 = build_sent_list(irrel_conj)
        w4.cell(row=p,column=3).value = str3
        p += 1

        bool1 = False
        if tot_prop_sent != []:
            for j in range(len(tot_prop_sent[i])):
            # if j == 2:
            #     break
                if not bool1 and tot_prop_sent[i][j][4] != "":
                    w4.cell(row=p,column=3).value = "____________________"
                    bool1 = True
                else:
                    w4.cell(row=p,column=2).value = tot_prop_sent[i][j][0]
                    w4.cell(row=p,column=3).value = tot_prop_sent[i][j][2] + tot_prop_sent[i][j][1]
                    str2 = ""
                    if tot_prop_sent[i][j][4] != "":
                        str2 = tot_prop_sent[i][j][3] + " " + str(tot_prop_sent[i][j][4])
                    if tot_prop_sent[i][j][5] != "" and tot_prop_sent[i][j][5] != None:
                        str2 += "," + str(tot_prop_sent[i][j][5])
                    if tot_prop_sent[i][j][6] != "" and tot_prop_sent[i][j][6] != None:
                        str2 += "," + str(tot_prop_sent[i][j][6])
                    if tot_prop_sent[i][j][7] != "" and tot_prop_sent[i][j][7] != None:
                        str2 += "," + str(tot_prop_sent[i][j][7])
                    w4.cell(row=p,column=4).value = str2

                p += 1
        p += 3


def build_dict(str1):

    list1 = []
    list4 = []
    word_type = []
    detm = []
    relat = []
    srelat = []
    trelat = []
    atomic_relations = []
    det = []
    adj = []
    adv = []
    noun = []
    cor = []
    lcon = []
    subo = []
    synon = []
    redundant = []
    aux = []
    atomic_relata = []
    negg = []
    dnoun = []
    det_pairs = []
    syn_pairs = []
    particles = []
    relations = []
    relations2 = []
    definitions = []
    really_atomic = []
    pronouns = []
    poss_pronouns = []
    plurals = []
    neg_det = []
    pos = []
    category = ['r','s','t','g']
    almost_done = False
    i = 0
    for row in ws.rows:
        i += 1
        if i > 900:
            bb = 7
        # if i == 430:
        #     pp = 5
        s = row[0].value
        if s == 67:
            bb = 7
        str1 = row[1].value
        word = row[2].value
        if word == 'space':
            bb = 7
        if word == None and almost_done:
            break
        if word == None:
            almost_done = True
        else:
            almost_done = False
        if str1 != None:
            str5 = copy.copy(str1)
            if isinstance(word,(int,long)):
                word = str(word)
            str3 = row[3].value
            defin = row[4].value
            if defin == 'redundant':
                redundant.append(word)
            if defin != None and defin.find("E.g.") == -1 and defin.find("EXP") == -1 \
                    and defin != 'redundant':
                if word != None:
                    word = word.strip()
                if str3 != None:
                    str3 = str3.strip()
                if str1 == None:
                    easygui.msgbox("you did not state the part of speech for " + word)
                atom = copy.copy(str1)
                str5 = str1[0:1]
                pos.append([word,str5])
                atom = atom[1:2]
                str8 = copy.copy(str1)
                str8 = str1[2:]

                if str5 in category:
                    if str5 == 'r':
                        relat.append(word)
                    elif str5 == 's':
                        srelat.append(word)
                    elif str5 == 't':
                        trelat.append(word)
                    if atom == 'a' or atom == 'b':
                        atomic_relations.append(str3)
        #i can't remember why I did this
                        # atom = 'b'
                    if atom == 'b':
                        really_atomic.append(str3)
                    relations.append(str3)
                    relations2.append(word)

                elif str5 == 'a':
                    adj.append(word)
                elif str5 == 'b':
                    aux.append(word)
                elif str5 == 'c':
                    cor.append(word)
                elif str5 == 'd':
                    detm.append(word)
                    det.append([word,atom,defin])
                elif str5 == 'e':
                    adv.append(word)
                elif str5 == 'l' and atom == 'b':
                    lcon.append(word)
                elif str5 == 'm':
                    negg.append(word)
                elif str5 == 'n':
                    noun.append(word)
                elif str5 == 'p':
                    pronouns.append(word)
                elif str5 == 'q':
                    poss_pronouns.append(word)
                elif str5 == 'u':
                    subo.append(word)
                elif str5 == 'w':
                    dnoun.append(word)
                if atom == 'a':
                    atomic_relata.append(word)
        # in the database the definition of plural must be written as 'plural of'
                elif atom == "m":
                    str6 = defin[10:]
                    plurals.append([word,str6])
                elif atom == 'q':
                    particles.append(word)
                if atom == 'p' or atom == 'd':
                    if atom == 'p':
                        atom = 7
                    elif atom == 'd':
                        atom = 5
                    list1a = [word, atom]
                elif atom == 's':
                    str6 = defin[defin.find("=")+1:-1]
                    str6 = str6.strip()
                    str7 = defin[1:defin.find("=")]
                    str7 = str7.strip()
                    if str5 == 'w' or str5 == 'x':
                        str8 = str6[:str6.find(" ")]
                        str7 = str6[str6.find(" ")+1:]
                        if str5 == 'w':
                            det_pairs.append([str8, str7, word, defin])
                        else:
                            word = word[5:]
                            word.strip()
                            if str6.find(" ") > -1:
                                str4.strip()
                                str6.strip()
                                str4 = str6[:str6.find(" ")]
                                str6 = str6[str6.find(" ")+1:]
                            else:
                                str4 = str6
                                str6 = ''
                            neg_det.append([word,str4,str6,defin])
                    else:
                        list3a = [str7, str6, defin]
                        syn_pairs.append(list3a)
                        synon.append(str7)

                if atom != 'a' and atom != 'm' and defin != "artificial" and defin != 'redundant'\
                    and defin != "postponed" and atom != 'b':
                    if str5 in category:
                        definitions.append([str3, defin,str5,atom,str8])
                    else:
                        definitions.append([word, defin,str5,atom,str8])

    syn_pairs.sort()
    # relations.sort()
    # relations2.sort()
    words = [adj, cor, detm, adv, lcon, noun, relat, srelat, trelat, subo,\
        aux, negg, dnoun,syn_pairs,synon,det, definitions, det_pairs, relations, \
             relations2, particles, redundant,atomic_relations, atomic_relata, \
             pronouns,poss_pronouns,plurals,neg_det,pos,really_atomic]

    return words

def findinlist(str1, list1, i, j, bool1 = False):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            if bool1:
                return d
            else:
                return str2
    if bool1:
        return -1
    else:
        return None

def findposinlist(str1, list1,i):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            return d
    else:
        return -1

def findin1dlist(str1,list1):

    for i in range(len(list1)):
        if str1 == list1[i]:
            return i

def isatomic(list1):

    global words
    atomic_relations = words[22]
    num = [5,14]
    if not list1[9] in atomic_relations:
        return False
    for i in num:
        if not isvariable(list1[i]):
            return False
    num = [3,4,6,7,10,11,12,13,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32, \
           33,34,35,47,48,49,50,51,52,53,54]
    for i in num:
        if list1[i] != "" and list1[i] != None:
            return False
    return True

def ismultidim(list1):

    if type(list1[0]) is list:
        return True
    else:
        return False

def cat_atoms(j,list1,list2,basic_objects,str1,bo2,bo,bool1 = False,members = []):

    str2 = ""
    subj = list1[5]
    relat = list1[9]
    obj = list1[14]
    list3 = ['point','moment','possible world','mind','matter','property','number',\
        'group','imagination','thought','relationship','sensational space',\
        'sensation','possible relationship','sentient being']

    if relat == "" or relat == None or subj == "" or subj == None or obj == "" or obj == None:
        easygui.msgbox("you didn't shift the objects into the right position")
    if str1 == 'nn':
        str2 = "nn"
        str1 = ""
    bool2 = False

    if bool1:
        str5 = findinlist(obj,dv_nam,0,1)
        if str5 not in list3:
            str1 = str5
            bool2 = True

    if list1[54] == None:
        list1[54] = ""
    if list1[53] == None:
        list1[53] = "cjj"
    if [list1[j],str1] not in basic_objects:
        if not bool2 and str1 != "":
            basic_objects.append([list1[j],str1])
        elif bool2 and str1 != "":
            members.append([list1[j],str1])
    if [list1[j],list1[0],list1[53]] not in bo2:
        list2.append([list1[j],list1[0],list1[53],list1[54]])
        bo2.append([list1[j],str1,str2,list1[0],list1[53]])
        bo.append([list1[j],str1,str2,list1[0],subj,relat,obj,list1[53],list1[54]])
    if not bool1:
        return list2
    else:
        return members

def violates_axioms(list1):

    list2 = extract_list(list1,0)
    for i in range(len(list2)):
        str1 = list2[i]
        g = list2.count(str1)
        if g > 1:
            return str1
    return None


def sort_mixed_list(list1):

    md_list = []
    nmd_list = []
    for i in range(len(list1)):
        if ismultidim(list1[i]):
            md_list.append(list1[i])
        else:
            nmd_list.append(list1[i])

    if nmd_list != []:
        nmd_list = sorted(nmd_list, key = operator.itemgetter(4))
        if md_list != []:
            list2 = nmd_list + md_list
        else:
            list2 = nmd_list
    else:
        list2 = md_list
    return list2

def identity(all_sent,tot_sent,basic_objects,words,candd,conditionals,prop_sent,prop_name):

    # we only need to place indefinite variables into the idf_id array if we are definite
    # a sentence whose subject is definite
    # if not ismultidim(list1):
    #     list1 = ([list1,["stop","hey"]])
    negat = []
    sent = []
    # for i in range(len(tot_sent)):
    #     if tot_sent[i][2] != "":
    #         sent.append([tot_sent[i][0],tot_sent[i][2]])
    #         negat.append(tot_sent[i][3])
    # dummy = plan(sent, prop_sent, candd, conditionals, prop_name, 1,negat)
    # tot_prop_sent.append(prop_sent)

    atomic = words[29]
    moments = []
    real_prop = []
    points = []
    numbers = []
    thoughts = []
    imag = []
    nonnum = []
    posre = []
    poswo = []
    words = []
    mind = []
    sebe = []
    feat = []
    groups = []
    matter = []
    sense = []
    sensp = []
    mind = []
    members = []
    used_var = []
    used_var2 = []
    bo2 = []
    bo = []
    str1 = ''
    i = -1
    while str1 != "b" + l7:
        i += 1
        str1 = def_var2[i]
        if str1 not in def_var:
            used_var.append([str1,"d"])

    for i in range(len(idf_var2)):
        str1 = idf_var2[i]
        if str1 not in idf_var:
            used_var.append([str1,"i"])

    former_list = copy.deepcopy(all_sent)
    j = -1
    while j < (len(all_sent)) -1:
        j += 1
        if j == 9:
            bb = 7
        if all_sent[j][9] not in atomic:
            del all_sent[j]
            j -= 1
        else:
            if all_sent[j][8] == None:
                if all_sent[j][9] == "A":
                    moments = cat_atoms(5,all_sent[j],moments,basic_objects,"mm",bo2,bo)
                    moments = cat_atoms(14,all_sent[j],moments,basic_objects,"mm",bo2,bo)
                elif all_sent[j][9] == '=':
                    str1 = ''
                    str2 = ''
                elif all_sent[j][9] == "AA":
                    real_prop = cat_atoms(5,all_sent[j],real_prop,basic_objects,"rr",bo2,bo)
                elif all_sent[j][9] == "AB" or all_sent[j][9] == "LF" or all_sent[j][9] == "FR":
                    points = cat_atoms(5,all_sent[j],points,basic_objects,"pt",bo2,bo)
                    points = cat_atoms(14,all_sent[j],points,basic_objects,"pt",bo2,bo)
                elif all_sent[j][9] == "AF":
                    numbers = cat_atoms(5,all_sent[j],numbers,basic_objects,"nm",bo2,bo)
                    numbers = cat_atoms(14,all_sent[j],numbers,basic_objects,"nm",bo2,bo)
                elif all_sent[j][9] == "AI":
                    thoughts = cat_atoms(5,all_sent[j],thoughts,basic_objects,"tk",bo2,bo)
                    imag = cat_atoms(14,all_sent[j],imag,basic_objects,"im",bo2,bo)
                elif all_sent[j][9] == "AN":
                    nonnum = cat_atoms(5,all_sent[j],nonnum,basic_objects,"nn",bo2,bo)
                    numbers = cat_atoms(14,all_sent[j],numbers,basic_objects,"nm",bo2,bo)
                elif all_sent[j][9] == "AP":
                    posre = cat_atoms(5,all_sent[j],posre,basic_objects,"pr",bo2,bo)
                    poswo = cat_atoms(14,all_sent[j],poswo,basic_objects,"pw",bo2,bo)
                elif all_sent[j][9] == "AW":
                    words = cat_atoms(5,all_sent[j],words,basic_objects,"wd",bo2,bo)
                    words = cat_atoms(14,all_sent[j],words,basic_objects,"wd",bo2,bo)
                elif all_sent[j][9] == "DS":
                    mind = cat_atoms(5,all_sent[j],mind,basic_objects,"mi",bo2,bo)
                    posre = cat_atoms(14,all_sent[j],posre,basic_objects,"pr",bo2,bo)
                elif all_sent[j][9] == "HW":
                    sebe = cat_atoms(5,all_sent[j],sebe,basic_objects,"sb",bo2,bo)
                    # str2 = 'a'
                    # str2 = ['mt','mi','se']
                elif all_sent[j][9] == "HV":
                    str1 = 'b'
                    # str1 = ['ir','rr','pr','dr']
                    words = cat_atoms(14,all_sent[j],words,basic_objects,"wd",bo2,bo)
                elif all_sent[j][9] == "IA":
                    dummy = cat_atoms(5,all_sent[j],feat,basic_objects,"",bo2,bo)
                    feat = cat_atoms(14,all_sent[j],feat,basic_objects,"fe",bo2,bo)
                elif all_sent[j][9] == "IG":
                    members = cat_atoms(5,all_sent[j],groups,basic_objects,"",bo2,bo,True,members)
                    groups = cat_atoms(14,all_sent[j],groups,basic_objects,"gr",bo2,bo)
                elif all_sent[j][9] == "S":
                    matter = cat_atoms(5,all_sent[j],matter,basic_objects,"mt",bo2,bo)
                    points = cat_atoms(14,all_sent[j],points,basic_objects,"pt",bo2,bo)
                elif all_sent[j][9] == "SS":
                    sense = cat_atoms(5,all_sent[j],sense,basic_objects,"se",bo2,bo)
                    sensp = cat_atoms(14,all_sent[j],sensp,basic_objects,"sp",bo2,bo)
                elif all_sent[j][9] == "T":
                    str1 = 'c'
                    # str1 = ['mt','pr','rr','ir','dr']
                    moments = cat_atoms(14,all_sent[j],moments,basic_objects,"mm",bo2,bo)
                elif all_sent[j][9] == "TK":
                    mind = cat_atoms(5,all_sent[j],mind,basic_objects,"mi",bo2,bo)
                    thoughts = cat_atoms(14,all_sent[j],thoughts,basic_objects,"tk",bo2,bo)

    basic_objects = sorted(basic_objects, key = operator.itemgetter(1,0))
    str1 = violates_axioms(basic_objects)
    # if str1 == None and len(basic_objects) != len(used_var):
    #     bb = 7

    bo2.sort()
    bo.sort()
    for i in range(len(bo2)):
        bo[i][8]

    bo = sorted(bo, key = operator.itemgetter(0,4))
    non_id2 = []
    list3 = ['fe','gr']
    list4 = []
    list5 = []
    list6 = []
    str2 = ""
    skip_string = ""
    dis_con = False

    for i in range(len(bo)):
        if bo[i][1] not in list3:
            str1 = bo[i][0]
            if str1 == 'u':
                bb = 7
            # if str1 != skip_string:
            skip_string = ""
    #the reason for dis_con is because if the sentence is a conjunctive disjunct then
    # it will appear twice in list5
            if not dis_con or numb != bo[i][8] or str2 != str1:
                if str1 != str2 and str2 !="":
                    numb = ""
                    dis_con = False
                    obj_type = findinlist(str2,basic_objects,0,1)
                    list4 = sort_mixed_list(list4)
                    list5.append([str2,obj_type,list4])
                    list4 = []
                if bo[i][7] != 'dc':
                    list4.append([bo[i][3],bo[i][4],bo[i][5],bo[i][6],bo[i][7]])
                else:
                    kind = bo[i][1]
                    sent = bo[i][3]
                    numb = bo[i][8]
                    if str1 == 'i':
                        bb = 7
                    list6 = []
                    list6.append([sent,bo[i][4],bo[i][5],bo[i][6],bo[i][7]])
                    list7 = []
                    list7.append(sent)
                    for o in range(len(bo)):
                        if o == 19:
                            bb = 7
                        if bo[o][3] != sent and bo[o][8] == numb:
                            if bo[o][3] not in list7:
                                list6.append([bo[o][3],bo[o][4],bo[o][5],bo[o][6],bo[o][7]])
                                list7.append(bo[o][3])
                    # skip_string = str1
                    list4.append(list6)
                    dis_con = True
            str2 = str1
    list5.append([str2,obj_type,list4])
    list5 = sorted(list5, key = operator.itemgetter(1,0))
    bb = 7

    for i in range(len(list5)):
        str1 = list5[i][0]
        if str1 == 'p':
            bb = 7
        if len(list5[i][2]) > 1 or ismultidim(list5[i][2][0]):
            str2 = ""
            for j in range(len(list5[i][2])):
                if ismultidim(list5[i][2][j]):
                    str4 = ""
                    for k in range(len(list5[i][2][j])):
                        str4 += list5[i][2][j][k][0] + " & "
                    str4 = str4[:-3]
                    str4 = "(" + str4 + ") " + xorr + " "
                    str2 += str4
                else:
                    if list5[i][2][j][4] == "1cj":
                        str3 = "&"
                    elif list5[i][2][j][4] == "d":
                        str3 = xorr
                    elif list5[i][2][j][4] == "cn":
                        str3 = implies
                    elif list5[i][2][j][4] == 'an':
                        str3 = cj
                    str2 += list5[i][2][j][0] + " " + str3 + " "
            str2 = str1 + " " + str2
        else:
            str2 = str1 + " " + list5[i][2][0][0]
        tot_sent.append(["zz",str2,"","","","",""])

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    i = -1
    while i < len(tot_sent) - 1:
        i += 1
        if ('df' in tot_sent[i][4] and 'RN' in tot_sent[i+1]) or 'RN' in tot_sent[i][4]:
            list4 = copy.deepcopy(tot_sent[i])
            list1.append(list4)
        else:
            list3 = copy.deepcopy(tot_sent[i])
            list2.append(list3)
    tot_sent = []
    for i in range(len(list1)):
        tot_sent.append(list1[i])
    for j in range(len(list2)):
        tot_sent.append(list2[j])




    # for i in range(len(non_id)):
    #     str1 = ''
    #     for j in range(len(non_id[i])):
    #         if str1 == '':
    #             str1 = "(" + non_id[i][j]
    #         else:
    #             str1 += ne + non_id[i][j]
    #     str1 += ")"
    #     non_id2.append(str1)
    #     if len(str1) > 3:
    #         tot_sent.append(["xx",str1,"","","","",""])

    # str1 = ""
    # list7 = []
    # k = 0
    # for i in range(1,len(list5)):
    #     if list5[i][1] != list5[i-1][1]:
    #         j = copy.copy(i)
    #         if i - k > 1:
    #             list7.append([k,i-1])
    #         k = i
    # if i - k > 1:
    #     list7.append([k,i])
    #
    # list4 = []
    # for i in range(len(list7)):
    #     y = list7[i][0]
    #     g = list7[i][1]
    #     y -= 1
    #     for n in range(y,g-1):
    #         y += 1
    #         h = y
    #         while h < g:
    #             h += 1
    #             if len(list5[y][2]) >  len(list5[h][2]):
    #                 large = y
    #                 small = h
    #             else:
    #                 small = y
    #                 large = h
    #                 bool1 = True
    #                 for q in range(len(list5[small][2])):
    #                     if bool1:
    #                         break
    #                     for r in range(len(list5[large][2])):
    #                         large_var = list5[large][0]
    #                         small_var = list5[small][0]
    #                         not_ident = arentident(large_var,small_var,non_id,non_id2)
    #                         if not_ident:
    #                             bool1 = True
    #                             break
    #                         else:
    #                             if large_var == list5[large][r][1]:
    #                                 large_obj = list5[large][r][3]
    #


    return tot_sent


def arentident(jobj, iobj,non_id,non_id2):

    for i in range(len(non_id2)):
        if (jobj in non_id2[i] and iobj in non_id2[i]):
            return True
    return False


def group_idf(list1):

    list2 = []
    for i in range(len(list1)):
        pass

def new_sentence(tot_sent,  old_list, list1, list2, list3, quant, rule,anc1 = ""):

    global prop_name
    global sn
    old_sent = build_sent(old_list)
    ng = ""
    old_sent2 = copy.copy(old_sent)
    if old_sent2.find("~") > -1:
        old_sent2 = old_sent2.replace("~","")
        ng = "~"
    old_sent2 = old_sent2.replace(" ","")
    old_sent2 = remove_outer_paren(old_sent2)
    old_prop = findinlist(old_sent2, prop_name, 1,0)
    str1 = build_sent(list1)
    str1v = name_sent(str1)
    list1[0] = str1
    if quant == 2:
        str2 = build_sent(list2)
        str2v = name_sent(str2)
    if list3 != "":
        str3 = build_sent(list3)
        str3v = name_sent(str3)
        list3[0] = str3
        easygui.msgbox('you have not coded for three new sentences yet')
    if quant == 1:
        str1 = old_sent + ' ' + iff + ' ' + str1
        str1v = ng + old_prop + ' ' + iff + ' ' + str1v
    elif quant == 2:
        str1 = '(' + old_sent + ' & ' + str2 + ') ' + conditional + ' ' + str1
        str1v = '(' + ng + old_prop + ' & ' + str2v + ') ' + conditional + ' ' + str1v

    g = findinlist(str1,tot_sent,2,0,True)
    if g == -1:
        sn += 1
        tot_sent.append([sn, str1, str1v, "", rule, anc1, ""])



def check_dimension(list1, i, str1,bool1 = False,k=0):

    if not bool1:
        for j in range(len(list1)):
            if list1[j][i] == str1:
                return True
        return False
    else:
        for j in range(len(list1)):
            if j != k:
                if list1[j][i] == str1:
                    return True
        return False

#################################################

## The following functions are for statement logic


def name_conditional(list1,nat_logic = True):

    # when using this function the list1 must be outputted from find_sentences
    # and you must use the 0th element since those sentences do not have negation signs
    global prop_var
    global prop_name
    skel_string = list1[5]
    list2 = []

    for i in range(1, len(list1[6])):
        if list1[6][i] != None and list1[6][i] != "":
            str3 = list1[0][i]
            str3 = remove_outer_paren(str3)
            str4 = copy.copy(str3)
            str4 = str4.replace(" ","")
            str1 = findinlist(str4,prop_name,1,0)
            if str1 != None:
                str2 = list1[1][i] + str1
                list2.append(str1)
                skel_string = skel_string.replace(list1[6][i][1], str2)
            else:
                str2 = prop_var[0]
                list2.append(str2)
                del prop_var[0]
                prop_name.append([str2, str4,str3])
                str2 = list1[1][i] + str2
                skel_string = skel_string.replace(list1[6][i][1], str2)
    return [skel_string, list2]

def tilde_removal(str1):


    str4 = ""
    if str1.find("~") > -1:
        str4 = "~"
        str1 = str1.replace("~","")
    return [str1,str4]

def tilde_removal2(str1):

    j = 0
    if str1[:2] == "~(":
        for i in range(len(str1)):
            str2 = str1[i:i+1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i > 0:
                if i == len(str1) -1:
                    str1 = str1[1:]
                    str4 = "~"
                    return [str1, str4]
                else:
                    return [str1,""]
    elif str1[0] == "~" and str1[1].islower() and os(str1):
        str1 = str1[1:]
        str4 = "~"
    else:
        str4 = ""
    return [str1,str4]

def simple_sent_name(str1):

    str1 = remove_outer_paren(str1)
    str2 = findinlist(str1,prop_name,1,0)
    if str2 == None:
        ostring = str1.replace(" ","")
        str2 = prop_var[0]
        del prop_var[0]
    prop_name.append([str2, ostring,str1])
    return str2

def is_conjunction(str1):

    if str1.count(idisj) == 0 and str1.count(iff) == 0 and \
        str1.count(conditional) == 0 and str1.count('&') > 0 and str1.count(xorr) == 0:
        return True
    else:
        return False


def get_conjuncts (str1, bool1 = False):
    # remove outparent if true
    global subscripts
    arr1 = []
    if bool1:
        str1 = remove_outer_paren(str1)

    j = 0
    k = 1
    for i in range (len(str1)):
        str2 = str1[i:i+1]
        str5 = str1[i+1:i+2]
        if i > 0:
            str4 = str1[i-1:i]
        else:
            str4 = ""
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 1 and str2 == "(" and str4 != "~":
            k = i
        elif j == 1 and str4 == "~" and str2 == "(":
            k = i - 1

        if (j==0 and str2 == ")") or (j == 0 and str2.islower()):
            if str2 == ")":
                str3 = str1[k:i+1]
            else:
                if str1[i-1:i] == "~" and str5 not in subscripts:
                    str3 = "~" + str2
                elif str1[i-1:i] != "~" and str5 not in subscripts:
                    str3 = str2
                elif str1[i-1:i] != "~" and str5 in subscripts:
                    str3 = str2 + str5
                elif str1[i-1:i] == "~" and str5 in subscripts:
                    str3 = "~" + str2 + str5
            k = 1
            str3 = str3.strip()
            arr1.append(str3)

    return arr1


def prepare_iff_elim(str2, mainc, s, num = ""):

    global sn
    list7 = [""] * 39
    if num == "":
        list7[2] = sn + 1
    else:
        list7[2] =  num
    list7[4] = str2
    list7[5] = ""
    j = 0
    str2 = remove_outer_paren(str2)

    if mainc == iff:
        list7[3] = "e"
    else:
        list7[3] = "c"

    str8 = str2[: s - 1]
    str8 = str8.strip()
    str9 = str2[s+1:]
    str9 = str9.strip()
    bool1 = True
    bool2 = True
    list8 = mainconn(str8)
    list2 = tilde_removal2(str8)
    if list8[0] != "&" or (list2[1] == "~" and list8[0] != ""):
        list2[0] = remove_outer_paren(list2[0])
        list7[0] = [list2[0],list2[1]]
        bool1 = False
    list8 = mainconn(str9)
    list2 = tilde_removal2(str9)
    if list8[0] != "&" or (list2[1] == "~" and list8[0] != ""):
        list2[0] = remove_outer_paren(list2[0])
        list7[1] = [list2[0],list2[1]]
        bool2 = False
    list4 = [bool1, bool2]
    for j in range(0, 2):
        if list4[j]:
            if j == 0:
                str7 = str8
            else:
                str7 = str9
            list2 = mainconn(str7)
            if list2[0] == '&':
                list3 = get_conjuncts(str7, True)
                list6 = []
                for k in range(len(list3)):
                    list5 = tilde_removal2(list3[k])
                    list5[0] = remove_outer_paren(list5[0])
                    list6.append([list5[0],list5[1]])
                if j == 0:
                    list7[0] = list6
                    list7[6] = str8
                else:
                    list7[1] = list6
                    list7[7] = str9

    return list7

def islist(list1):

    if list1[1] == '~' or list1[1] == "":
        return False
    else:
        return True

def new_prop(prop_sent, str1, ng, asp, anc1, anc2, anc3 = None, anc4 = None, \
             is_premise = False, num = "",ostring = ""):

    if ng == None:
        ng = ""
    global sn
    list1 = [None] * 15

    if ng == "":
        str1 = remove_outer_paren(str1)
    str2 = findinlist(str1,prop_sent,1,2)
    if str2 == ng:
        return True
    elif str2 == None:
        if is_premise:
            sn = num
            prop_sent.append([num, str1, ng, "", "", "", None, None, ostring, None, None, \
            None, None, None, None])
        else:
            sn += 1
            prop_sent.append([sn, str1, ng, asp, anc1, anc2, None, None, None, None, None, \
            None, None, None, None])
        return True
    elif str2 != ng:
        sn += 1
        prop_sent.append([sn, str1, ng, asp, anc1, anc2, None, None, None, None, None, \
        None, None, None, None])
        anc2 = findinlist(str1,prop_sent,1,0)
        sn += 1
        if not os(str1):
            str1 = "(" + str1 + ")"
        str1 = str1 + " & " + "~" + str1
        prop_sent.append([sn, str1, "", "&I", sn-1, anc2, None, None, None, None, None, \
        None, None, None, None])
        str1 = bottom
        sn += 1
        prop_sent.append([sn, str1, "", bottom + "I", sn-1, None, None, None, None, None, None, \
        None, None, None, None])
        return False

def many_cond(candd, conditionals, kind, asp, anc2, f, g, r):

    list1 = conditionals[g][f]
    del list1[0]
    # this list allows us to keep track of the ancestors even if we don't detach a
    # a part of the conditionals on the first time
    if conditionals[g][8] == "":
        conditionals[g][8] = [candd[r][0]]
    else:
        conditionals[g][8].append(candd[r][0])
    if f == 1:
        h = 7
    else:
        h = 6
    list2 = []
    list2.append([candd[r][1],candd[r][2]])
    if list1 == []:
        cjct = conditionals[g][h]
        dummy = new_prop_sent("", kind, asp, "",anc2, \
                    conditionals,g, conditionals[g][8], cjct)
        if dummy == False:
            return False
        else:
            return True
    j = -1
    while j < len(list1)-1:
        j += 1
        str2 = list1[j][0]
        neg2 = list1[j][1]
        for i in range(len(candd)):
            if i != r:
                str1 = candd[i][1]
                ng = candd[i][2]
                anc1 = candd[i][0]
                if str1 == str2 and ng == neg2:
                    del list1[j]
                    j -= 1
                    list2.append([str1, ng])
                    conditionals[g][8].append(anc1)
                    if list1 == []:
                        cjct = conditionals[g][h]
                        dummy = new_prop_sent("", kind, asp, "",anc2, \
                                    conditionals,g, conditionals[g][8], cjct)
                        if dummy == False:
                            return False
                        else:
                            return True
                    else:
                        break
                elif str1 == str2 and ng != neg2:
            # the point of having blank returns is because if it returns true
            # then we need to subtract the conditional counter, here g, by 1
                    return ""
    conditionals[g][f] = list1
    return ""

def modus_ponens(conditionals, candd, prop_sent):

    global sn
    r = -1
    while r < len(candd) -1:
        if conditionals == []:
            return True
        r += 1
        # if r == 3:
        #     bb = 7
        str1 = candd[r][1]
        str8 = candd[r][2]
        anc1 = candd[r][0]
        temp1 = copy.copy(str1)
        temp1 = temp1.replace(" ","")
        temp1 = remove_outer_paren(temp1)
        g = -1
        bool1 = False
        while g < len(conditionals) -1:
            g += 1
            if g == 3 and r==1:
                bb = 7
            str12 = conditionals[g][3]
            if str12 != 'd':
                anc2 = conditionals[g][2]
                if conditionals[g][6] == "":
                    aconjunction = ""
                    temp_ant = conditionals[g][0][0]
                    temp_nega = conditionals[g][0][1]
                else:
                    aconjunction = conditionals[g][6]
                    temp_ant = conditionals[g][0][0][0]
                    temp_nega = conditionals[g][0][0][1]
                if conditionals[g][7] == "":
                    cconjunction = ""
                    temp_con = conditionals[g][1][0]
                    temp_negc = conditionals[g][1][1]
                else:
                    cconjunction = conditionals[g][7]
                    temp_con = conditionals[g][1][0][0]
                    temp_negc = conditionals[g][1][0][1]
                temp_ant = temp_ant.replace(" ","")
                temp_con = temp_con.replace(" ","")
                temp_ant = remove_outer_paren(temp_ant)
                temp_con = remove_outer_paren(temp_con)

                for f in range(0,2):
                    if f == 0 and temp1 == temp_ant:
                        if str8 == temp_nega:
                            if str12 == 'c':
                                str13 = "MP"
                            else:
                                str13 = "EE"
                            if aconjunction != "" :
                                dummy = many_cond(candd, conditionals, "con", str13, \
                                                  anc2, f, g, r)
                                if dummy == False:
                                    return False
                                elif dummy:
                                    del conditionals[g]
                                    g -= 1
                                    break
                            else:
                 # con indicates that the consequent of the conditional is to be detached
                                dummy = new_prop_sent("", "con", \
                                    str13, anc1, anc2,conditionals,g)
                                if dummy == False:
                                    return False
                                del conditionals[g]
                                g -= 1
                                break
                        elif str8 != temp_nega and str12 == 'e':
                            dummy = new_prop_sent("~", "con", \
                                        "EN", anc1, anc2, conditionals,g)
                            if dummy == False:
                                return False
                            del conditionals[g]
                            g -= 1
                            break
                    elif f == 1 and temp1 == temp_con:
                        if str8 == temp_negc and str12 == 'e':
                            if cconjunction == "":
                                dummy = new_prop_sent("", "ant", "EE", \
                                    anc1, anc2, conditionals,g)
                                if dummy == False:
                                    return False
                                del conditionals[g]
                                g -= 1
                                break
                            else:
                                dummy = many_cond(candd, conditionals, "ant", "EE", \
                                                  anc2, f, g, r)
                                if dummy == False:
                                    return False
                                elif dummy:
                                    del conditionals[g]
                                    g -= 1
                                    break
                        elif str8 != temp_negc:
                            if str12 == 'c':
                                str13 = "MT"
                            else:
                                str13 = "EN"
                            dummy = new_prop_sent("~", "ant", \
                                        str13, anc1, anc2, conditionals,g)
                            if dummy == False:
                                return False
                            del conditionals[g]
                            g -= 1
                            break
                    elif f == 0 and temp1 != temp_ant and \
                                    aconjunction != "" and str12 == 'e':
                        s = -1
                        if conditionals != []:
                            while s < len(conditionals[g][0]) -1:
                                s += 1
                                if temp1 == conditionals[g][0][s][0] and \
                                    str8 != conditionals[g][0][s][1]:
                                    dummy = new_prop_sent("~", "con", \
                                    "EN", anc1, anc2, conditionals,g)
                                    if dummy == False:
                                        return False
                                    del conditionals[g]
                                    g -= 1
                                    break

                    elif f == 1 and temp1 != temp_con and cconjunction != "":
                        s = -1
                        if conditionals != []:
                            while s < len(conditionals[g][1]) -1:
                                s += 1
                                if temp1 == conditionals[g][1][s][0] and \
                                    str8 != conditionals[g][1][s][1]:
                                    if str12 == 'e':
                                        str13 = "EN"
                                    else:
                                        str13 = "MT"
                                    dummy = new_prop_sent("~", "ant", \
                                    str13, anc1, anc2, conditionals,g)
                                    if dummy == False:
                                        return False
                                    del conditionals[g]
                                    g -= 1
                                    break
    return True

def disjunction_heirarchy(conditionals, str5,d,new_disj = False):

    global prop_name
    global sn

    str5 = enclose(str5)
    list1 = find_sentences(str5, False)
    mainc = list1[4][0][1]

    for i in range(len(list1[0])):
        list2 = unenclose(list1[0][i])
        list1[0][i] = list2[0]

    list2 = [""] * 39
    n = 7
    if mainc == xorr:
        list2[3] = 'x'
    else:
         list2[3] = 'd'
    if conditionals == [] or new_disj:
        list2[2] = sn
    else:
        list2[2] = conditionals[d][2]
    list2[5] = ""
    list2[4] = list1[0][0] # fix this

    for i in range(len(list1[0])):
        if os(list1[0][i]):
            siblings = []
            list3 = [None] * 9
            n += 1
            # str1 = findinlist(list1[0][i],prop_name,1,0)
            str2 = list1[4][i][0][:-1]
            g = findinlist(str2,list1[4],0,1,True)
            parent = list1[0][g]
            if list1[4][g][1] == "&":
                list3[2] = 'c'
            elif list1[4][g][1] == xorr:
                list3[2] = 'x'
            else:
                list3[2] = 'd'
            if len(str2) > 1:
                str3 = list1[4][i][0][:-2]
                g = findinlist(str3,list1[4],0,1,True)
                gparent = list1[0][g]
            else:
                gparent = parent
            list3[1] = list1[4][i][0]
            list3[5] = parent
            list3[6] = gparent
            list3[0] = [list1[0][i], list1[1][i]]
            #fix this
            b = parent.count(xorr)
            c = parent.count(idisj)
            if c > 1 or b > 1:
                list3[7] = 2
            else:
                list3[7] = 1
            sent_num = list1[4][i][0]
            m = len(sent_num)
            for j in range(len(list1[4])):
                if len(list1[4][j][0]) == m and list1[4][j][0][:-1] == str2 \
                    and j != i:
                    siblings.append([list1[0][j],list1[1][j]]) # fix this
            list3[4] = siblings
            list2[n] = list3
    if conditionals == [] or new_disj:
        conditionals.append(list2)
    else:
        conditionals[d] = list2

def proper_spacing(str1):

    str1 = str1.replace(" ","")
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional," " + conditional + " ")
    str1 = str1.replace(idisj," " + idisj + " ")
    str1 = str1.replace(xorr," " + xorr + " ")
    str1 = str1.replace("&"," & ")
    return str1

def iff_elim(prop_sent, conditionals):

    new_sent = False
    no_contr = True

    for d in range(len(conditionals)):
        list1 = ["",""]
        ng = conditionals[d][5]
        if conditionals[d][3] == "d" or conditionals[d][3] == 'n':
            list1 = [conditionals[d][4], ""]
        else:
            if conditionals[d][6] == "":
                list1[0] = conditionals[d][0][0]
            else:
                list1[0] = conditionals[d][6]
            if conditionals[d][7] == "":
                list1[1] = conditionals[d][1][0]
            else:
                list1[1] = conditionals[d][7]
        for s in range(0,2):
            if list1[s].find(iff) > -1:
                new_sent = True
                str1 = list1[s]
                anc1 = conditionals[d][2]
                old_str = copy.copy(str1)
                str1 = str1.replace(" ","")
                str1 = str1.replace("&"," & ")
                str1 = str1.replace(iff , " " + iff + " ")
                str1 = remove_outer_paren(str1)
                list5 = mainconn(str1)
                mc = list5[0]
                bool4 = False
                if str1[:2] == "~(":
                    bool4 = True
                    str1 = str1[2:len(str1) - 1]
                r = str1.count(iff)
                i = -1
                for o in range(0,r):
                    k = 0
                    while i < len(str1) - 1:
                        i += 1
                        str2 = str1[i:i+1]
                        if str2 == iff:
                            p = copy.copy(i)
                            t = copy.copy(i)
                            bool5 = False
                            while k != -1:
                                i -= 1
                                str3 = str1[i:i+1]
                                if str3 == "(":
                                    k -= 1
                                elif str3 == ")":
                                    k += 1
                                    bool5 = True
                                if k == -1:
                                    m = i + 1
                                elif k == 0 and bool5:
                                    m = i
                                    break
                            bool3 = False
                            k = 0
                            while p < len(str1):
                                p += 1
                                str3 = str1[p:p+1]
                                if str3 == "(":
                                    bool3 = True
                                    k -= 1
                                elif str3 == ")":
                                    k += 1
                                if bool3 and k == 0:
                                    p += 1
                                    break
                                elif k == 1:
                                    break
                            ant = str1[m:t]
                            ant = ant.strip()
                            con = str1[t+1:p]
                            con = con.strip()
                            new1 = "(" + ant + " " + conditional + " " + con + ")"
                            new2 = "(" + con + " " + conditional + " " + \
                                   ant + ")"
                            if mc != "&":
                                if r > 1 and o + 1 != r:
                                    new3 = "(" + new1 + ' & ' + new2 + ")*"
                                else:
                                    new3 = "(" + new1 + ' & ' + new2 + ")"
                            else:
                                if r > 1 and o + 1 != r:
                                    new3 = new1 + ' & ' + new2 + '*'
                                else:
                                    new3 = new1 + ' & ' + new2
                            if m != 0 or p != len(str1):
                                replacee = str1[m-1:p+1]
                                new3 = str1.replace(replacee, new3)
                            str1 = new3
                            if r > 1 and o + 1 != r:
                                i = str1.find("*")
                                str1 = str1.replace("*","")
                            break
                new3 = str1
                new3 = proper_spacing(new3)
                if bool4:
                    new3 = "~(" + new3 + ")"
                else:
                    new3 = "(" + new3 + ")"
                conditionals[d][4] = conditionals[d][4].replace(old_str,new3)
                if conditionals[d][3] == 'c' or conditionals[d][3] == 'e':
                    if s == 0:
                        conditionals[d][0] = None
                        conditionals[d][6] = new3
                    else:
                        conditionals[d][1] = None
                        conditionals[d][7] = new3

        if conditionals[d][3] == "e":
            conditionals[d][3] = "c"
            anc1 = conditionals[d][2]
            list1 = [""] * 39
            list1[5] = ""
            if conditionals[d][6] == "":
                str1 = conditionals[d][0][0]
                if not os(str1):
                    str1 = "(" + str1 + ")"
                ng1 = conditionals[d][0][1]
            else:
                str1 = conditionals[d][6]
                ng1 = ""
            if conditionals[d][7] == "":
                str4 = conditionals[d][1][0]
                if not os(str4):
                    str4 = "(" + str4 + ")"
                ng4 = conditionals[d][1][1]
            else:
                str4 = conditionals[d][7]
                ng4 = ""

            list1[0] = [str4,ng4]
            list1[1] = [str1, ng1]
            str5 =  ng1 + str1 + " " + conditional + " " + ng4 + str4
            str6 = ng4 + str4 + " " + conditional + " " + ng1 + str1
            g = copy.copy(sn+1)
            if conditionals[d][5] == "" or conditionals[d][5] == None:
                str3 = "(" + str5 + ") & (" + str6 + ")"
                str7 = iff + "E"
                no_contr = new_prop(prop_sent, str3, ng, str7, anc1, None, None, None)

                if not no_contr:
                    return False
                conditionals[d][2] = sn+1
                conditionals[d][0] = [str1, ng1]
                conditionals[d][1] = [str4, ng4]
                no_contr = new_prop(prop_sent,str5,"","&E",g,"")
                if not no_contr:
                    return False
                list1[2] = sn + 1
                no_contr = new_prop(prop_sent,str6,"","&E",g,"")
                if not no_contr:
                    return False
                conditionals[d][4] = str5
                list1[4] = str6
                conditionals.append(list1)
            else:
                str3 = "(" + str5 + ") & (" + str6 + ")"
                no_contr = new_prop(prop_sent,str3,"~",iff + "E",g,"")
                if not no_contr:
                    return False
                conditionals[d][3] = 'd'
                conditionals[d][4] = str3
        elif new_sent:
            no_contr = new_prop(prop_sent,conditionals[d][4],"",iff + "E",anc1,"")
            if not no_contr:
                return False
    return True

def material_implication(prop_sent, conditionals):

    for d in range(len(conditionals)):
        str1 = conditionals[d][4]
        ng = conditionals[d][5]
        if str1.find(conditional) > -1:
            anc1 = conditionals[d][2]
            i = -1
            q = -1
            s = 0
            r = str1.count(conditional)
            while s < r:
                i += 1
                q += 1
                str2 = str1[i:i+1]
                if str2 == conditional:
                    s += 1
                    j = i
                    k = 0
                    m = 0
                    bool1 = False
                    while j != 0:
                        j -= 1
                        m += 1
                        if m > 200:
                            easygui.msgbox("in the material implication function \
                            you are caught in an infinite loop")
                            return
                        str3= str1[j:j+1]
                        if str3.islower() and k == 0:
                            bool1 = True
                        elif str3 == ")":
                            bool1 = True
                            k += 1
                        elif j == 0 and str3 != "(":
                            k = 0
                            bool1 = True
                        elif str3 == "(":
                            bool1 = True
                            k -= 1

                        if bool1 and k == 0:
                            str4 = str1[0:j]
                            str5 = str1[i+2:]
                            str7 = str1[j:i]
                            str1 = str4 + "~" + str7 + idisj + " " + str5
                            i += 1
                            break

            str1 = bad_paren(str1)
            no_contr = new_prop(prop_sent, str1, ng, conditional + "E", anc1,"")
            if not no_contr:
                return False
            if str1.find("~~") > -1:
                str1 = str1.replace("~~","")
                g = copy.copy(sn)
                no_contr = new_prop(prop_sent,str1, ng, "~~E", g,"")
                if not no_contr:
                    return False
            conditionals[d][2] = sn
            conditionals[d][4] = str1
            conditionals[d][3] = "d"
            conditionals[d][5] = ng

    return True

            # search for double negatives

def bad_paren(str1):

    if str1.find("(") == -1:
        return str1
    # we first must get rid of strings of the following form ((p) & s)
    for i in range(len(str1)):
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        if i > 1:
            str4 = str1[i-2:i-1]
        else:
            str4 = ""
        str5 = str1[i+1:i+2]
        if str2.islower() and str3 == "(" and str5 == ")":
            str1 = str1[:i-1] + str1[i:i+1] + str1[i+2:]
        elif str2.islower() and str4 == "(" and str3 == "~" and str5 == ")":
            str1 = str1[:i-2] + str1[i-1:i+1] + str1[i+2:]

    str1 = enclose(str1)
    list1 = find_sentences(str1, False)
    for i in range(len(list1[3])):
        list2 = unenclose(list1[3][i])
        list1[3][i] = list2[0]
    mstr = list1[3][0]
    for i in range(1, len(list1[3])):
        if list1[4][i][1] != "":
            mc = list1[4][i][1]
            ostr = list1[3][i]
            str2 = list1[4][i][0][:-1]
            prcnt = findinlist(str2,list1[4],0,1,False)
            if mc == prcnt:
                nstr = remove_outer_paren(ostr)
                nstr = remove_outer_paren(nstr)
                mstr = mstr.replace(ostr, nstr)
    return mstr

def demorgan(prop_sent, conditionals, candd,one_sent = False, str8 = "",anc1a = "",rule = ""):

    d = -1
    temp_bool = True
    rop = False
    if one_sent:
        d = len(conditionals) - 2

    while d < len(conditionals) -1:
        d += 1
        if one_sent:
            str1 = str8
        else:
            str1 = conditionals[d][5] + conditionals[d][4]
        if one_sent:
            rop = True
        else:
            if conditionals[d][5] == "~":
                rop = True
        if str1.find("~(") > -1:
            if one_sent:
                anc1 = anc1a
            else:
                anc1 = conditionals[d][2]
            r = str1.count("~(")
            s = 0
            i = -1
            while s < r:
                i += 1
                if i > 200:
                    easygui.msgbox("you are caught in an infinite loop in the \
                    demorgan function")
                    return
                bool2 = False
                str2 = str1[i:i+2]
                if str2 == "~(":
                    s += 1
                    bool2 = True
                    if i > 0:
                        str4 = str1[0:i]
                    else:
                        str4 = ""
                    str5 = str1[i+1:len(str1)+1]
                    str1 = str4 + str5
                    j = i
                    m = 0
                    k = 0
                    bool3 = False
                    while j < len(str1):
                        m += 1
                        if m > 200:
                            easygui.msgbox("you are caught in an infinite loop in the \
                            demorgan function")
                            return
                        bool1 = False
                        str3 = str1[j:j+1]
                        if j-1 > 0:
                            str6 = str1[j-1:j+1]
                        else:
                            str6 = ""

                        if str6 == "~(" and bool3 == False and bool2 == False:
                            k = 0
                            k += 1
                            s += 1
                            # bool3 means that python is searching only for those letters and connectives
                            # which are on the same level of parentheses
                            bool3 = True
                            str4 = str1[0:j]
                            str5 = str1[j:]
                            str1 = str4 + "~" + str5
                            j += 2
                            str3 = str1[j:j+1]


                        bool2 = False

                        if str3.islower() and bool3 == False:
                            str4= str1[0:j]
                            str5 = str1[j+1:(len(str1)+1)]
                            str1 = str4 + "~" + str3 + str5
                            j += 2
                            bool1 = True
                        elif bool3 == False:
                            if str3 == "(":
                                k += 1
                                j += 1
                                bool1 = True
                            elif str3 == ")":
                                k -= 1
                                j += 1
                                bool1 = True
                            elif str3 == "&":
                                str4 = str1[0:j]
                                str5 = str1[(j + 1):(len(str1) + 1)]
                                str1 = str4 + idisj+ str5
                                j += 1
                            elif str3 == idisj or str3 == xorr:
                                str4 = str1[0:j]
                                str5 = str1[(j+1):(len(str1)+1)]
                                str1 = str4 + "&" + str5
                                j += 1
                            else:
                                j += 1

                        elif str3 == ")":
                            j += 1
                            k -= 1
                        elif str3 == "(":
                            k += 1
                            j += 1
                        else:
                            j += 1

                        if k == 0 and bool3:
                            bool3 = False
                        if k == 0 and bool1 == True:
                            i = j
                            break

        # rop means remove outer paren, this is necessary if the original sentence was negated
            if rop:
                str1 = remove_outer_paren(str1)
            list1 = [None] * 15
            if str1.find("~~") > -1:
                str2 = copy.copy(str1)
                str2 = str2.replace("~~","")
                str2 = bad_paren(str2)
            else:
                str1 = bad_paren(str1)
            no_contr = new_prop(prop_sent,str1,"","~(E",anc1,"")
            if not no_contr:
                return False
            if str1.find("~~") > -1:
                str1 = str2
                anc1 = copy.copy(sn)
                no_contr = new_prop(prop_sent,str2,"","~~E",anc1,"")
                if not no_contr:
                    return False
            list2 = mainconn(str1)

            if list2[0] == "&":
                if not one_sent:
                    del conditionals[d]
                list3 = get_conjuncts(str1,True)
                anc1 = copy.copy(sn)
                for i in range(len(list3)):
                    list4 = tilde_removal2(list3[i])
                    no_contr = new_prop(prop_sent,list4[0],list4[1],"&E",anc1,"")
                    if not no_contr:
                        return False
                    list2 = mainconn(list3[i])
                    if list2[0] == idisj or list2[0] == xorr:
                        # add in more nones if it turns out that I need them
                        if one_sent:
                            dummy = disjunction_heirarchy(conditionals,str1,0)
                        else:
                            list5 = [""] * 39
                            list5[2] = sn
                            list5[4] = list4[0]
                            list5[5] = list4[1]
                            conditionals.append(list5)
                    else:
                        candd.append([sn,list4[0],list4[1]])
            else:
                if one_sent:
                    dummy = disjunction_heirarchy(conditionals,str1,0,True)
                    return
                else:
                    conditionals[d][2] = sn
                    conditionals[d][4] = str1
    return True

def unenclose(str1):

    i = -1
    global subscripts
    list1 = []
    while i < len(str1)-1:
        i += 1
        str2 = str1[i:i+1]
        str3 = str1[i-1:i]
        str4 = str1[i+1:i+2]
        if str2.islower() and str3 != "~" and str4 not in subscripts:
            str1 = str1[:i-1] + str2 + str1[i+2:]
            list1.append(str2)
        elif str2.islower() and str3 == "~" and str4 not in subscripts:
            str1 = str1[:i-2] + str3 + str2 + str1[i+2:]
            list1.append(str2)
        if str2.islower() and str3 != "~" and str4 in subscripts:
            str1 = str1[:i-1] + str2 + str4 + str1[i+3:]
            list1.append(str2 + str4)
        elif str2.islower() and str3 == "~" and str4 in subscripts:
            str1 = str1[:i-2] + str3 + str2 + str4 + str1[i+3:]
            list1.append(str2 + str4)
    return [str1,list1]

def new_disjunct(str1, ng, n, prop_sent, conditionals, candd, anc1, anc2, \
            anc3 = None, anc4=None, kind = 0, rule = ""):

    global sn
    list2 = mainconn(str1)
    if kind == 1:
        del conditionals[n]
        dummy = new_prop(prop_sent, str1, ng, "&I", \
            anc1, anc2, anc3, anc4)
        return dummy
    elif kind == 2:
        dummy = new_prop(prop_sent, str1, ng, "&I", \
            anc1, anc2, anc3, anc4)
        return dummy
    else:
        if os(str1):
            del conditionals[n]
            str1 = remove_outer_paren(str1)
            list1 = tilde_removal2(str1)
            str1 = list1[0]
            dummy = new_prop(prop_sent, str1, list1[1], rule + "E", \
            anc1, anc2)
            candd.append([sn,str1,list1[1]])
            return dummy
        elif list2[0] == "&":
            del conditionals[n]
            str1 = remove_outer_paren(str1)
            dummy = new_prop(prop_sent, str1, ng, rule + "E", \
            anc1, anc2)
            g = copy.copy(sn)
            list3 = get_conjuncts(str1)
            for i in range(len(list3)):
                list4 = tilde_removal2(list3[i])
                dummy = new_prop(prop_sent, list4[0], list4[1], "&E", g,"")
                if dummy == False:
                    return dummy
                if list3[i].find(idisj) > -1:
                    dummy = disjunction_heirarchy(conditionals, list4[0],n, True)
                else:
                    candd.append([sn,list4[0],list4[1]])
            return True
        else:
            dummy = new_prop(prop_sent, str1, ng, idisj + "E", \
            anc1, anc2)
            if dummy == False:
                return dummy
            if ng == "~":
                str1 = ng + str1
            else:
                str1 = remove_outer_paren(str1)
            dummy = disjunction_heirarchy(conditionals, str1,n, False)
            conditionals[n][2] = sn
            return True

def xorr_elim(conditionals,n,i,parent,grandparent,whole_d,candd,prop_sent,anc1,anc2,kind=0):

    str9 = ""
    de_mor = False
    if kind == 0:
        for r in range(len(conditionals[n][i][4])):
            if r != i:
                if not os(conditionals[n][i][4][r][0]):
                    de_mor = True
                if str9 == "":
                    str9 += "~" + conditionals[n][i][4][r][1] + conditionals[n][i][4][r][0]
                else:
                    str9 += " & ~" + conditionals[n][i][4][r][1] + conditionals[n][i][4][r][0]
    else:
        grandp2 = copy.copy(grandparent)
        grandp2 = grandp2.replace(parent,"")
        for r in range(8,38):
            if conditionals[n][r] == "":
                break
            if conditionals[n][r][0][0] in grandp2:
                if str9 == "":
                    str9 += "~" + conditionals[n][r][0][1] + conditionals[n][r][0][0]
                else:
                    str9 += " & ~" + conditionals[n][r][0][1] + conditionals[n][r][0][0]
    g = copy.copy(sn)
    if parent != grandparent:
        str9 = remove_outer_paren(str9)
        if grandparent == whole_d:
            mc = mainconn(str9)
            if mc[0] == '&':
                consistent = xorr_elim2(str9,prop_sent,conditionals,candd,anc1,anc2)
                if consistent == False:
                    return consistent
            else:
                list4 = tilde_removal(str9)
                consistent = new_prop(prop_sent, list4[0],list4[1], xorr + "E", \
                anc1, anc2)
                if consistent == False:
                    return consistent
        else:
            str9 = "(" + str9 + ")"
            if kind == 0:
                str9 = grandparent.replace(parent,str9)
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                dummy = new_prop(prop_sent,str9,"",xorr + "E",anc1,anc2)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~","")
                    dummy = new_prop(prop_sent,str9,"","~~E",sn,"")

            else:
                str9 = whole_d.replace(grandparent,str9)
                str9 = bad_paren(str9)
                dummy = new_prop(prop_sent,str9,"",xorr + "E",anc1,anc2)
                g = copy.copy(sn)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~","")
                    consistent = new_prop(prop_sent,str9,"","~~E",g,"")
                    if consistent == False:
                        return consistent
                dummy = disjunction_heirarchy(conditionals, str9,n, True)
                del conditionals[n]
            if de_mor:
                consistent = demorgan(prop_sent,conditionals,candd,True,str9,sn,xorr +"E")
                if consistent == False:
                    return consistent
            else:
                if str9.find(idisj) > -1 or str9.find(xorr) > -1:
                    dummy = disjunction_heirarchy(conditionals,str9,n,True)
                consistent = True
    else:
        #this does not account for the case where the parent == grandparent but
        # grandparent does not == whole d
        consistent = xorr_elim2(str9,prop_sent,conditionals,candd,anc1,anc2)

    return consistent

def xorr_elim2(str9,prop_sent,conditionals,candd,anc1,anc2):

    str9 = bad_paren(str9)
    consistent = new_prop(prop_sent, str9, "", xorr + "E", \
    anc1, anc2)
    if consistent == False:
        return False
    if str9.find("~~") > -1:
        str9 = str9.replace("~~","")
        dummy = new_prop(prop_sent,str9,"","~~E",sn,"")
        if dummy == False:
            return dummy
    list3 = get_conjuncts(str9)
    g = copy.copy(sn)
    for b in range(len(list3)):
        list4 = tilde_removal2(list3[b])
        list4[0] = remove_outer_paren(list4[0])
        dummy = new_prop(prop_sent, list4[0], list4[1], "&E", g,"")
        if dummy == False:
            return dummy
        if not os(list3[b]):
            if list4[1] == "~":
                consistent = demorgan(prop_sent,conditionals,\
                candd,True,list3[b],sn,"&E")
                if consistent == False:
                    return False
            else:
                dummy = disjunction_heirarchy(conditionals, list4[0],n, True)
        else:
            candd.append([sn,list4[0],list4[1]])
    return True


def disjunction_elimination(prop_sent, conditionals, candd):

    bool1 = False
    bool2 = False
    global sn

    for i in range(len(conditionals)):
        dummy = disjunction_heirarchy(conditionals, conditionals[i][4],i)
    i = -1
    while i < len(candd) -1:
        i += 1
        if not os(candd[i][1]):
            del candd[i]
            i -= 1
    d = -1
    while d < len(candd) -1:
        d += 1
        str2 = candd[d][2]
        conj = candd[d][1]
        if conj == 'q':
            bb = 7
        if d == 9:
            bb = 7
        anc1 = candd[d][0]
        n = -1
        while n < len(conditionals) -1:
            if bool1:
                bool1 = False
                d = -1
                break
            n += 1
            i = 7
            while conditionals != []:
                if bool2:
                    bool2 = False
                    break
                i += 1
                if conditionals[n][i] == "":
                    break

                whole_d = conditionals[n][4]
                anc2 = conditionals[n][2]
                str3 = conditionals[n][i][0][0]
                # 'pos or neg'
                str4 = conditionals[n][i][0][1]
                # 'disjunct or conjunct
                str5 = conditionals[n][i][2]
                # 'disjunct number
                str6 = conditionals[n][i][1]

                if conj == str3:
                    grandparent = conditionals[n][i][6]
                    parent = conditionals[n][i][5]
                    parent2 = copy.copy(parent)
                    parent3 = copy.copy(parent)
                    str7 = " " + idisj + " "
                    str7a = " " + xorr + " "
                    if str2 == str4 and str5 == "d":
                        # 'if the disjuncts are not embedded within a conjunct then the disjunction
                        # is simply deleted

                        del conditionals[n]
                        if parent != grandparent:
                            conj = str2 + conj
                            str8 = whole_d.replace(parent, conj)
                            dummy = disjunction_heirarchy(conditionals, str8,n)
                        bool1 = True
                        n = -1
                        break

                    elif str2 == str4 and str5 == "x":

                        consistent = xorr_elim(conditionals,n,i,parent,grandparent,whole_d,candd,prop_sent,anc1,anc2)
                        if consistent == False:
                            return False
                        del conditionals[n]
                        bool2 = True
                        bool1 = True
                        d = -1
                    elif str2 == str4 and str5 == "c":
                        list2 = []
                        list2.append([conj,str2])
                        anc3 = ""
                        anc4 = ""
                        list1 = conditionals[n][i][4]
                        f = -1
                        while f < len(list1) -1:
                            mc = mainconn(grandparent)
                            f += 1
                            for e in range(len(candd)):
                                anc5 = candd[e][0]
        # since it's too hard to program, if the sibling is a disjunct then we just
        # ignore this
                                if list1[f][0].find(idisj) > -1 or list1[f][0].find(xorr) > -1:
                                    break
                                else:
                                    if candd[e][1] == list1[f][0]:
                                        if candd[e][2] == list1[f][1]:
                                            list2.append([list1[f][0],list1[f][1]])
                                            if len(list2) == 2:
                                                anc3 = anc5
                                            elif len(list2) == 3:
                                                anc4 = anc5
                                            del list1[f]
                                            if list1 == []:
                                                str3 = build_sent_list2(list2)
                                                if mc[0] == xorr:
                                                    new_prop(prop_sent,str3,"","&I",anc1,anc3,anc4)
                                                    consistent = xorr_elim(conditionals,n,i,parent,\
                                                    grandparent,whole_d,candd,prop_sent,sn,anc2,1)
                                                    if consistent == False:
                                                        return False
                                                else:
                                        # if the conjunct is not embedded within another conjunct
                                        # then the disjunct is simply deleted
                                                    if whole_d == grandparent:
                                                        dummy = new_disjunct(str3,"",n, prop_sent,\
                                                        conditionals,candd, anc1, anc3, anc4, anc5, 1)
                                                    else:
                                                        str8 = whole_d.replace(grandparent, parent2)
                                                        if str8.find("(") > -1 and str8.find(idisj) > -1:
                                                            str8 = bad_paren(str8)
                                                        dummy = new_disjunct(str3,"",n, prop_sent,\
                                                        conditionals,candd, anc1, "", anc3, anc4, 2)
                                                        dummy = new_disjunct(str8,"",n, prop_sent,\
                                                        conditionals,candd, sn-1, anc2)
                                                        if dummy == False:
                                                            return False
                                                bool1 = True
                                                bool2 = True
                                                n = 0
                                                d = -1
                                                break
                                            else:
                                                f -= 1
                                                break

                                        elif candd[e][2] != list1[f][1]:
                                            mc = mainconn(grandparent)
                                            if mc[0] == idisj:
                                                rule = idisj
                                                str7 = " " + idisj + " "
                                            else:
                                                str7 = " " + xorr + " "
                                                rule = xorr
                                            r = grandparent.find(parent)
                                            if r > 1:
                                                parent = str7 + parent
                                            else:
                                                parent = parent + str7
                                            anc1 = candd[e][0]
                                            str9 = grandparent.replace(parent, "")
                                            str8 = whole_d.replace(grandparent, str9)
                                            if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                                str8 = bad_paren(str8)
                                            dummy = new_disjunct(str8,"",n, prop_sent,\
                                                    conditionals,candd, anc1, anc2,None,None,0,rule)
                                            if dummy == False:
                                                return False
                                            else:
                                                list1 = []
                                                bool1 = True
                                                bool2 = True
                                                n = 0
                                                d = -1
                                                break

                    elif str2 != str4 and str5 == "c":
                        mc = mainconn(conditionals[n][i][6])
                        if mc[0] == idisj:
                            str6 = str7 + parent
                            rule = idisj
                        else:
                            rule = xorr
                            str6 = str7a + parent
                        if grandparent.find(str6) > -1:
                            parent = str6
                        else:
                            parent = parent + str7
                        str9 = grandparent.replace(parent, "")
                        str8 = whole_d.replace(grandparent, str9)
                        if str8.find("(") > -1 and str8.find(idisj) > -1:
                            str8 = bad_paren(str8)
                        dummy = new_disjunct(str8,"",n,prop_sent, conditionals, candd,anc1, anc2, None, None,0,rule)
                        if dummy == False:
                            return False
                        bool1 = True
                        n = -1
                        break

                    elif str2 != str4 and (str5 == "d" or str5 == 'x'):
                        # if the disjunct is a triple disjunct then enter below
                        if str5 == 'd':
                            rule = idisj
                        else:
                            rule = xorr
                        if conditionals[n][i][7] > 1:
                            str6 = str4 + str3 + " " + rule + " "
                            if parent.find(str6) > -1:
                                str5 = str6
                            else:
                                str5 = " " + rule + " " + str4 + str3
                            str9 = parent.replace(str5, "")
                            str8 = whole_d.replace(parent, str9)
                            if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                str8 = bad_paren(str8)
                            dummy = new_disjunct(str8,"",n,prop_sent,conditionals,\
                                candd, anc1,anc2,None,None,0,rule)
                            if dummy == False:
                                return False
                            bool1 = True
                            n = -1
                            break

                        else:
                            str3 = conditionals[n][i][4][0][0]
                            str4 = conditionals[n][i][4][0][1]
                            str5 = str4 + str3
                            str8 = whole_d.replace(parent, str5)
                            if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                str8 = bad_paren(str8)
                            dummy = new_disjunct(str8,"",n,prop_sent,conditionals,\
                                candd, anc1,anc2,None,None,0,rule)
                            if dummy == False:
                                return False
                            bool1 = True
                            n = -1
                            break
    return True

def extract_list(list1,d):

    list2 = []
    for i in range(len(list1)):
        list2.append(list1[i][d])
    return list2

def build_concl(prop_sent, conditionals):

    list2 = []
    list1 = []

    for i in range(len(prop_sent)):
        if os(prop_sent[i][1]):
            list1.append([prop_sent[i][1], prop_sent[i][2]])
    list2 = copy.deepcopy(list1)
    list1.sort()
    str1 = build_sent_list2(list2)
    str2 = build_sent_list2(list1)

    if conditionals != []:
        noncj = extract_list(conditionals,4)
        for i in range(len(noncj)):
            noncj[i] = "(" + noncj[i] + ")"
        str3 = build_sent_list(noncj)
        str1 = str1 + " & " + str3
        str2 = str2 + " & " + str3
    return [str1,str2]

def impl_elim2(conc, impl,prop_sent,ng,i):

    global sn
    str1 = conc[0][2] + conc[0][1] + " " + impl + " " + conc[0][2] + conc[0][1]
    dummy = new_prop(prop_sent,str1,"","&E",sn,"")
    if impl == implies:
        str1 = "~(" + conc[0][2] + conc[0][1] + " & " + "~" + conc[0][2] + conc[0][1] + ")"
    else:
        str1 = ng + conc[0][1] + " & " + "~" + conc[0][2] + conc[0][1]
    dummy = new_prop(prop_sent,str1,"",impl + "E",sn,"")
    if conc[0][2] == "~":
        str1 = "~(" + ng + conc[0][1] + " & " + conc[0][1] + ")"
        dummy = new_prop(prop_sent,str1,"","~~E",sn,"")

def impl_elim3(conditionals, prop_sent, impl,conc):

    global sn
    list1 = build_concl(prop_sent, conditionals)
    str1 = list1[0] + " " + impl + " " + conc[0][2] + conc[0][1]
    dummy = new_prop(prop_sent,str1,"",impl + "I",conc[0][0],"")
    str1 = list1[1] + " " + impl + " " + conc[0][2] + conc[0][1]
    dummy = new_prop(prop_sent,str1,"","alpha",sn,"")
    if impl == nonseq:
        # i need to eliminate the conditionals that the negation of consequent eliminates
        str1 = list1[1] + " & ~" + conc[0][2] + conc[0][1]
        dummy = new_prop(prop_sent,str1,"",nonseq + "E",sn,"")
        if conc[0][2] == "~":
            str1 = list1[1] + " & " + conc[0][1]
            dummy = new_prop(prop_sent,str1,"","~~E",sn,"")

def implication_elimination(prop_sent, conc, impl, conditionals):

    cncl = conc[0][1]
    ng = conc[0][2]
    global sn

    for j in range(len(prop_sent)):
        if prop_sent[j][1] == cncl and ng == prop_sent[j][2] and impl == implies:
            dummy = impl_elim3(conditionals,prop_sent,impl,conc)
            dummy = impl_elim2(conc,impl,prop_sent,conc[0][2],j)
            dummy = new_prop(prop_sent,"~" + bottom,"",bottom + "I",sn,"")
            dummy = new_prop(prop_sent,top,"",bottom + "E",sn,"")
            return
        elif prop_sent[j][1] == cncl and ng != prop_sent[j][2] and impl == implies:
            dummy = impl_elim2(conc,impl,prop_sent,prop_sent[j][2],j)
            dummy = new_prop(prop_sent,bottom,"",bottom + "I",sn,"")
            return
        elif prop_sent[j][1] == cncl and ng == prop_sent[j][2] and impl == nonseq:
            dummy = impl_elim2(conc,impl,prop_sent,conc[0][2],j)
            dummy = new_prop(prop_sent,bottom,"",bottom + "I",sn,"")
            return
        elif prop_sent[j][1] == cncl and ng != prop_sent[j][2] and impl == nonseq:
            dummy = impl_elim2(conc,impl,prop_sent,prop_sent[j][2],j)
            dummy = new_prop(prop_sent,top,"",top + "I",sn,"")
            return
    else:
        dummy = impl_elim3(conditionals,prop_sent, impl, conc)
        dummy = new_prop(prop_sent,top,"",top + "I",sn,"")

def statement_logic(prop_sent, conditionals, candd, conc="", impl=""):

    global sn
    consistent = modus_ponens(conditionals, candd, prop_sent)
    if consistent == False:
        return False
    consistent = iff_elim(prop_sent,conditionals)
    if consistent == False:
        return False
    consistent = material_implication(prop_sent, conditionals)
    if consistent == False:
        return False
    consistent = demorgan(prop_sent, conditionals, candd)
    if consistent == False:
        return False
    consistent = disjunction_elimination(prop_sent,conditionals,candd)
    if consistent == False:
        return False
    if conc != "":
        consistent = implication_elimination(prop_sent,conc,impl, conditionals)
        if consistent == False:
            return False
    return True

def add_outer_paren(str1):

    str1 = remove_outer_paren(str1)
    return "(" + str1 + ")"

def new_prop_sent(ng, kind, asp, anc1, anc2, conditionals,g,list3 =[], cjct = ""):

    global prop_sent
    global candd
    global sn

    if kind == 'con':
        h = 1
        e = 7
    else:
        h = 0
        e = 6
    bool1 = False
    if list3 != []:
        sn += 1
        ancc = [4,5,6,7]
        list4 = [None] * 15
        for i in range(len(list3)):
            if i == 4:
                break
            list4[ancc[i]] = list3[i]
        list4[0] = sn
        list4[1] = cjct
        list4[2] = ""
        list4[3] = "&I"
        anc1 = sn
        prop_sent.append(list4)

    if conditionals[g][e] == "":
        str1 = conditionals[g][h][0]
        list2 = mainconn(str1)

        # here we take care of double negatives
        if ng == "~" and conditionals[g][h][1] == "~":
            sn += 1
            prop_sent.append([sn, "~" + str1, "~", asp, anc1, anc2, None, None, None, None, None,None, None, None, None])
            g = copy.copy(sn)
            dummy = new_prop(prop_sent,str1,"","~~E",g, None)
            if dummy == False:
                return False
            ng = ""
            bool1 = True
        elif (ng == "" and conditionals[g][h][1] == "~") or \
                (ng == "~" and conditionals[g][h][1] == ""):
            ng = "~"
        if list2[0] != "" and ng != "~":
            str1 = remove_outer_paren(str1)
        elif list2[0] != "" and ng == "~":
            str1 = add_outer_paren(str1)
        if bool1 == False:
            dummy = new_prop(prop_sent,str1,ng,asp,anc1, anc2)
            if dummy == False:
                return False

        candd.append([sn, str1, ng])
        if (list2[0] == iff or list2[0] == conditional) and ng != "~":
            list3 = prepare_iff_elim(str1, list2[0],list2[1],sn)
            conditionals.append(list3)
        elif list2[0] == iff or list2[0] == conditional or list2[0] == idisj or list2[0] == xorr:
            list5 = [""] * 39
            list5[2] = sn
            list5[4] = str1
            list5[5] = ng
            list5[3] = 'd'
            conditionals.append(list5)
    else:
        str1 = conditionals[g][e]
        if ng == "~":
            list5 = [""] * 39
            list5[2] = sn + 1
            list5[4] = str1
            list5[5] = "~"
            list5[3] = 'd'
            conditionals.append(list5)
            dummy = new_prop(prop_sent, str1, ng, asp, anc1, anc2)
            if dummy == False:
                return False
        else:
            list1 = conditionals[g][h]
            str1 = remove_outer_paren(str1)
            dummy = new_prop(prop_sent, str1, ng, asp, anc1, anc2)
            if dummy == False:
                return False
            anc1 = copy.copy(sn)
            for i in range(len(list1)):
                list2 = mainconn(list1[i][0])
                dummy = new_prop(prop_sent, list1[i][0], list1[i][1], "&E", anc1, "")
                if dummy == False:
                    return False
                candd.append([sn, list1[i][0], list1[i][1]])
                if (list2[0] == conditional or list2[0] == iff) and list1[i][1] != "~":
                    list4 = prepare_iff_elim(list1[i][0], list2[0],list2[1],sn)
                    conditionals.append(list4)
                elif list2[0] != "":
                    list5 = [""] * 39
                    if list2[0] == conditional:
                        list5[3] = 'c'
                    elif list2[0] == iff:
                        list5[3] = 'e'
                    else:
                        list5[3] = 'd'
                    list5[2] = sn
                    list5[4] = list1[i][0]
                    list5[5] = list1[i][1]
                    conditionals.append(list5)
    return True

def oc(str1):
    # this function determines if there is only one connective which is not &
    # it is used to weed out conditionals from the candd list
    list1 = [conditional, idisj, iff,xorr]
    j = 0
    for i in range(len(str1)):
        str2 = str1[i:i+1]
        if str2 in list1:
            j += 1
    if j == 1:
        return True
    else:
        return False

def plan(sent, prop_sent, candd, conditionals, prop_name, kind = '',negat=[]):

    global conc
    global sn
    global rel_conj
    global irrel_conj
    conj_elim = []
    temp_conditionals = []
    list4 = []
    str3 = ""
    str1 = ""
    conc = ""
    impl = ""
    qq = 0
    nat_logic = False
    # if the first sentence is just one letter then we're using statement logic
    # if sent[0][1].find("(") > -1:
    #     for j in range(len(sent[0][1])):
    #         str2 = sent[0][1][j:j+1]
    #         str3 = sent[0][1][j+1:j+2]
    #         if (str2.islower() or str2.isupper()) and (str3.islower() or str3.isupper()):
    #             nat_logic = True
    #         if j > 8:
    #             break


    for i in range(len(sent)):
        # if i == 23:
        #     bb = 7
        g = sent[i].count('(')
        h = sent[i].count(')')
        if g != h:
            easygui.msgbox('wrong number of parentheses in sentence:' + sent[i])
            return "stop"
        if nat_logic == False:
            sent[i][1] = enclose(sent[i][1])
        if sent[i][1].find("!") > -1:
            qq += 1
        else:
            if sent[i][1].count("(") != sent[i][1].count(")"):
                easygui.msgbox("line " + sent[i][0] + " does not have the right number \
                                             of parentheses" )
            if nat_logic == False:
                sent[i][1] = remove_outer_paren(sent[i][1])
            ostring = copy.copy(sent[i][1])
            nstring = proper_spacing(ostring)
            if sent[i][1].find(implies) > -1 or sent[i][1].find(nonseq) > -1:

                if sent[i][1].find(implies) > -1:
                    impl = implies
                else:
                    impl = nonseq

                str1 = nstring[1:]
                if str1.find("&") > -1:
                    list1 = str1.split("&")
                    list4 = find_sentences(str1, True)
                    list6 = name_conditional(list4)
                    str5 = list6[0]
                    str2 = impl + ' ' + str5
                    for i in range(len(list1)):
                        list3 = simple_sent_name(list1[i], True)
                        conc.append([sent[i][0], list3[0], list3[1]])
                else:
                    list3 = tilde_removal(str1)
                    list4 = tilde_removal(ostring)
                    str1 = simple_sent_name(list3[0],list4[0])

                    str2 = impl + ' ' + list3[1] + str1
                    ng = ""
                    conc.append([sent[i][0], str1, list3[1]])

            elif os(sent[i][1]):
                if kind == "":
                    list3 = tilde_removal(nstring)
                    list4 = tilde_removal(ostring)
                    str2 = list3[0]
                    ng = list3[1]
                else:
                    ng = negat[i]
                    str2 = sent[i][1]
                if nat_logic:
                    list3[0] = simple_sent_name(list3[0],list4[0])
                candd.append([sent[i][0], str2, ng])
            else:
                list1 = find_sentences(ostring, True)
                if not nat_logic:
                    list6 = unenclose(list1[0][0])
                    sent[i][1] = list6[0]
                else:
                    list6 = name_conditional(list1,nat_logic)
                    list6[0] = proper_spacing(list6[0])
                str2 = list6[0]
                if kind == "":
                    list4 = tilde_removal2(str2)
                    ng = list4[1]
                    str2 = list4[0]
                else:
                    ng = negat[i]
                list2 = mainconn(str2)
                if list2[0] == idisj or list2[0] == xorr:
                    if oc(str2):
                        candd.append([nstring, str2,ng])
                        rel_conj = rel_conj + list6[1]
                    list5 = [""] * 39
                    list5[2] = sent[i][0]
                    list5[3] = 'd'
                    list5[4] = str2
                    list5[5] = ng
                    conditionals.append(list5)
                elif list1[4][0][1] != "&":
                    str3 = ""
                    rel_conj = rel_conj + list6[1]
                    if list1[4][0][1] != idisj and ng == "" and list1[4][0][1] != xorr:
                        list7 = prepare_iff_elim(str2, list2[0], list2[1],sent[i][0])
                    else:
                        list7 = [""] * 39
                        list7[2] = sent[i][0]
                        list7[3] = 'd'
                        list7[4] = str2
                        list7[5] = ng
                    conditionals.append(list7)
                    if oc(str2):
                        candd.append([sent[i][0], str2,ng])
                else:
                    list3 = get_conjuncts(list6[0])
                    for j in range(len(list3)):
                        list5 = tilde_removal2(list3[j])
                        if os(list3[j]):
                            conj_elim.append([sent[i][0],list5[0],list5[1]])
                        else:
                            temp_conditionals.append([sent[i][0], list5[0],list5[1]])

            no_contr = new_prop(prop_sent,str2,ng,"",None,None,None,None,True,sent[i][0],ostring)
            if not no_contr:
                return False

    if conj_elim != []:
        for i in range(len(conj_elim)):
            no_contr = new_prop(prop_sent,conj_elim[i][1],conj_elim[i][2],"&E",\
                        conj_elim[i][0],None)
            if not no_contr:
                return False
            candd.append([sn, conj_elim[i][1], conj_elim[i][2]])

    if temp_conditionals != []:
        for i in range(len(temp_conditionals)):
            if nat_logic:
                list1 = find_sentences(temp_conditionals[i][1], True)
                list6 = name_conditional(list1)
                str2 = list6[0]
            else:
                str2 = temp_conditionals[i][1]
            ng = temp_conditionals[i][2]
            list2 = mainconn(str2)

            if list2[0] != idisj and ng == "" and list2[0] != "&" and list2[0] != xorr:
                list7 = prepare_iff_elim(str2, list2[0], list2[1],sent[i][0])
            else:
                list7 = [""] * 39
                list7[2] = sn + 1
                list7[4] = str2
                list7[3] = 'd'
            if ng == "~":
                list7[5] = "~"
            rel_conj = rel_conj + list6[1]
            conditionals.append(list7)
            if oc(str2):
                candd.append([sn+1, str2,ng])
            no_contr = new_prop(prop_sent, str2,ng,"&E",temp_conditionals[i][0], None)
            if not no_contr:
                return False


    # we now weed out the irrelevant conjuncts from the relevant
    # fix this later
    # i = -1
    # if len(candd) > 0:
    #     while i < len(candd) -1:
    #         i += 1
    #         if candd[i][1] not in rel_conj and os(candd[i][1]):
    #             irrel_conj.append(candd[i][1])
    #             d = findinlist(candd[i][1],prop_sent,1,True)
    #             del candd[i]
    #             i -= 1

    consistent = statement_logic(prop_sent, conditionals, candd,conc,impl)

    return consistent

def populate_sentences(p):

    bool1 = False
    bool2 = False
    first_sent = False
    sent = []
    test_sent = []
    g = 0
    for row in w4.rows:
        p += 1
        if row[2].value == None and bool2 == True:
            break
        elif row[2].value == None:
            test_sent.append(sent)
            sent = []
            g = 0
            bool1 = False
            bool2 = True
            first_sent = False
        elif row[2].value.find("!!!") > -1:
            bool1 = True
            test_sent.append(sent)
            sent = []
        elif row[2].value != None and bool1 == False:
            str2 = row[2].value
            g += 1
            if len(sent) == 0:
                if str2.find(bottom) > -1:
                    tv = 'co'
                    str2 = str2[2:]
                else:
                    tv = 'ta'
            else:
                tv = ""
            str2.strip()
            sent.append([g, str2,row[0].value,tv])
            if not first_sent:
                 w4.cell(row=p-1,column=2).value = len(test_sent)
            first_sent = True
            bool2 = False
    return [test_sent,p]

subscripts = [l1,l2,l7]
tot_prop_name = []
tot_prop_sent = []
prop_var4 = [unichr(97 + t) for t in range(26)]
prop_var2 = [unichr(97 + t) + u"\u2081" for t in range(26)]
prop_var3 = [unichr(97 + t) + u"\u2082" for t in range(26)]
prop_var5 = [unichr(97 + t) + u"\u2083" for t in range(26)]
prop_var6 = [unichr(97 + t) + u"\u2084" for t in range(26)]
prop_var4 = prop_var4 + prop_var2 + prop_var3 + prop_var5 + prop_var6
idf_var2 = [unichr(122 - t) for t in range(25)]
idf_var3 = [unichr(122 - t) + l1 for t in range(25)]
idf_var4 = [unichr(122 - t) + l2 for t in range(25)]
idf_var2 = idf_var2 + idf_var3 + idf_var4
def_var2 = [unichr(98 + t) + "'" for t in range(25)]
def_var3 = [unichr(98 + t) + l7 for t in range(25)]
def_var2 = def_var2 + def_var3
wb4 = load_workbook('test9.xlsx')
w4 = wb4.worksheets[0]
## rajiv put this in mysql
wb = load_workbook('define2.xlsx')
ws = wb.worksheets[1]
# wb2.get_sheet_names()
# ws = wb2.get_sheet_by_name('words')
p = 1

## rajiv change this function
list1 = populate_sentences(p)
test_sent = list1[0]
p = list1[1]
words = build_dict('hey')
# #begin2
st = time.time()
for k in range(len(test_sent)):
# for k in range(0,2):
    print k
    if k == 1:
        bb = 7
    prop_name = []
    tot_sent = []
    all_sent = []
    plural_c = []
    definite = []
    idf_id = [[""]]
    used_idf = []
    conditionals = []
    candd = []
    rel_conj = []
    irrel_conj = []
    conc = []
    prop_sent = []
    tagged_nouns = []
    tagged_nouns2 = []
    dv_nam = []
    non_id = []
    basic_objects = []
    t_value = test_sent[k][0][3]
    prop_var = copy.deepcopy(prop_var4)
    idf_var = copy.deepcopy(idf_var2)
    def_var = copy.deepcopy(def_var2)
    id_num = test_sent[k][-1][0] + 1
    sn = id_num
    dummy = divide_sent(words, test_sent[k], def_var, idf_var,tot_sent,all_sent)
    dummy = det_nouns(tot_sent, all_sent,words,0)
    dummy = syn(tot_sent, all_sent, words,0)
    dummy = word_sub(def_var, dv_nam, tot_sent, all_sent,words,id_num)
    dummy = division(tot_sent, all_sent)
    dummy = plurals(tot_sent,all_sent,0,words,dv_nam,def_var, idf_var)
    dummy = define(tot_sent, all_sent, def_var,idf_var, dv_nam, words)
    tot_sent = identity(all_sent,tot_sent,basic_objects,words,candd,conditionals,prop_sent,prop_name)
    # dummy = temp_id(tot_sent,idf_id,non_id)
    #dummy = id_relat(tot_sent, all_sent)
    test_sent[k] = tot_sent
    tot_prop_name.append(prop_name)
en = time.time()
g = (en-st)/(k+1)
print g
## rajiv change this function
dummy = print_sent_full(test_sent,p,tot_prop_name)
wb4.save('test9.xlsx')


# end1