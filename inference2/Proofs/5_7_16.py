from openpyxl import load_workbook
from collections import Counter
import copy
import time
import operator
import sys
from ex_dict_new import large_dict
from claims_new import pop_sent
from pprint import pprint
import collections
from start_and_stop import info

# checked coverage up to line 5575

# averaged .076 on 5.22 (proof type 'n'), .039 definitions, .004 statement logic
# averaged .059 on 5.22 proof type 'n', .023 definitions, .004 statement
# but just prior to that the speed was .066
# time spent in instantiation is .029

# on 6/8 time spent in instantiation = .009, .014
# on 6/10 time spent in instantiation = .018, definitions = .031, total .074

# on 6/26 average .026 (up to instantiation) definitions: .019
# trial 2, .020, .025, trial 3: same as 2, total 2.562

# with a lot of globals: 2.831, 3.330, 2.934

# 7/2 average .050, definitions .029, statement .0015, instantiation .010

#7/19 average .0275 3 trials using old dictionary


tot_tim = time.time()

strt, stp, print_to_doc, get_words_used, order, nonlinear = info()

excel = False
mysql = False
django2 = False
normal_proof = True
wb4 = load_workbook('/Users/kylefoley/Desktop/inference engine/temp_proof.xlsx')
w4 = wb4.worksheets[0]

if get_words_used:
    wb5 = load_workbook('/Users/kylefoley/Desktop/inference engine/dictionary4.xlsx')
    ws = wb5.worksheets[0]

if mysql:
    from inference2.models import Define3, Archives, Input
    from inference2 import views
if mysql:
    import os

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    print(BASE_DIR)
    sys.path.append(BASE_DIR)
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "inference_engine2.settings")
    import django

    django.setup()
    from inference2 import views
    from inference2.models import Define3, Archives, Input

total_sent = []
all_sent = []
attach_sent = []
detach_sent = []
prop_name = []
prop_var = []
variables = []
abbreviations = []
words_used = []
dictionary = {}
definite_assignments = {}
propositional_constants = {}
already_defined = []
variable_type = []

tot_prop_name = []
result_data = {}
time1 = 0
st_log_time = 0
inst_tim = 0
instan_used = 0  # the number of times the instan function is used
instan_time = 0  # measures the time used in instantiation
lemmas_used = 0
time_spent_in_lemma_function = 0
time_spent_reducing = 0
time_spent_defining = 0
build_sent_slots_time = 0
build_sent_slots_counter = 0

cond_r = chr(8835)
const = "\u2102"  # consistency
top = chr(8868)
bottom = chr(8869)
neg = chr(172)
idd = chr(8781)  # translation symbol
iff = chr(8801)
mini_c = chr(8658)
mini_e = chr(8703)
implies = chr(8866)
conditional = chr(8594)
nonseq = chr(8876)
xorr = chr(8891)
idisj = chr(8744)
cj = chr(8896)
aid = chr(8776)
disj = chr(8855)
equi = chr(8660)
ne = "\u2260"  # not equal

l1 = "\u2081"
l2 = "\u2082"
l3 = "\u2083"
l4 = "\u2084"
l5 = "\u2085"
l6 = "\u2086"
l7 = "\u2087"
ua = "\u1d43"
ub = "\u1d47"
uc = "\u1d9c"
ud = "\u1d48"
ue = "\u1d49"
uf = "\u1da0"
ug = "\u1d4d"
ui = "\u2071"
uk = "\u1d4f"
um = "\u1d50"
un = "\u207f"
uo = "\u1d52"
up = "\u1d56"
ut = "\u1d57"
uv = "\u1d5b"
uu = "\u1d58"
uw = "\u02b7"
uy = "\u02b8"
uj = "\u02B2"
ul = "\u02E1"
ur = "\u02b3"
us = "\u02e2"
uh = "\u02b0"

prop_var4 = [chr(97 + t) for t in range(26)]
prop_var2 = [chr(97 + t) + "\u2081" for t in range(26)]
prop_var3 = [chr(97 + t) + "\u2082" for t in range(26)]
prop_var5 = [chr(97 + t) + "\u2083" for t in range(26)]
prop_var6 = [chr(97 + t) + "\u2084" for t in range(26)]
prop_var7 = [chr(97 + t) + "\u2085" for t in range(26)]
prop_var8 = [chr(97 + t) + "\u2086" for t in range(26)]
prop_var9 = [chr(97 + t) + "\u2087" for t in range(26)]
prop_var4 = prop_var4 + prop_var2 + prop_var3 + prop_var5 + prop_var6 + prop_var7 + prop_var8 + prop_var9
variables2 = [chr(122 - t) for t in range(25)]
variables2.remove("i")
variables2.remove("l")
variables3 = [chr(122 - t) + l1 for t in range(25)]
variables4 = [chr(122 - t) + l2 for t in range(25)]
variables2 = variables2 + variables3 + variables4

subscripts = [l1, l2, l3, l4]
alpha = chr(945)
beta = chr(946)


#
# >> 8835
# ta^ 8868
# co^ 8869
# nt+ 172
# x^ 8801
# c^ 8658
# # 8703
# i^ 8866
# t^ 8594
# nf^ 8876
# ed^ 8891
# v+ 8744
# && 8896
# @ 8855
# if^ 8660


def tran_str(str1, type3):
    list2 = []
    str2 = ""
    if 'co^' in str1:
        str1 = str1.replace('co^ ', "")
        str2 = 'co'

    if "|" in str1:
        for i in range(len(str1)):
            if str1[i:i + 1] == "|":
                str3 = str1[i + 1:i + 2]
                str4 = get_super(str3)
                str1 = str1[:i] + str4 + str1[i + 2:]
                bb = 8

    if type3 == 3:

        if "t^" in str1:
            str1 = str1.replace("t^", conditional)
        if "nt+" in str1:
            str1 = str1.replace("nt+", neg)
        if "zzz" in str1:
            str1 = str1.replace("zzz", ne)
        if "x^" in str1:
            str1 = str1.replace("x^", iff)
        if "b^" in str1:
            str1 = str1.replace("b^", mini_e)
        if "c^" in str1:
            str1 = str1.replace("c^", mini_c)
        if "ed^" in str1:
            str1 = str1.replace("ed^", xorr)
        if "v+" in str1:
            str1 = str1.replace("v+", idisj)
    if type3 == 1:
        list2 = str1.split(" % ")
    else:
        list2 = str1

    return [list2, str2]


def get_super(str1):
    if str1 == "a":
        return "\u1d43"
    elif str1 == "b":
        return "\u1d47"
    elif str1 == "c":
        return "\u1d9c"
    elif str1 == "d":
        return "\u1d48"
    elif str1 == "e":
        return "\u1d49"
    elif str1 == "f":
        return "\u1da0"
    elif str1 == "g":
        return "\u1d4d"
    elif str1 == "h":
        return "\u02b0"
    elif str1 == "i":
        return "\u2071"
    elif str1 == "j":
        return "\u02B2"
    elif str1 == "k":
        return "\u1d4f"

        # ua = u"\u1d43"
    # ub = u"\u1d47"
    # uc = u"\u1d9c"
    # ud = u"\u1d48"
    # ue = u"\u1d49"
    # uf = u"\u1da0"
    # ug = u"\u1d4d"
    # uh = u"\u02b0"
    # ui = u"\u2071"
    # uj = u"\u02B2"
    # uk = u"\u1d4f"
    # ul = u"\u02E1"
    # um = u"\u1d50"
    # un = u"\u207f"
    # uo = u"\u1d52"
    # up = u"\u1d56"
    # ur = u"\u02b3"
    # us = u"\u02e2"
    # ut = u"\u1d57"
    # uu = u"\u1d58"
    # uv = u"\u1d5b"
    # uw = u"\u02b7"
    # uy = u"\u02b8"

    elif str1 == "l":
        return "\u02E1"
    elif str1 == "m":
        return "\u1d50"
    elif str1 == "n":
        return "\u207f"
    elif str1 == "o":
        return "\u1d52"
    elif str1 == "p":
        return "\u1d56"
    elif str1 == "r":
        return "\u02b3"
    elif str1 == "s":
        return "\u02e2"
    elif str1 == "t":
        return "\u1d57"
    elif str1 == "u":
        return "\u1d58"
    elif str1 == "v":
        return "\u1d5b"
    elif str1 == "w":
        return "\u02b7"
    elif str1 == "y":
        return "\u02b8"


def remove_outer_paren(str1, bool1=False):
    if str1 == "":
        return ""
    elif str1.count(")") == 0:
        if not bool1:
            return str1
        else:
            return False

    j = 0
    # on very rare occasions we will encounter strings of the following form ((p))
    if str1[0] != "(" and str1[-1] != ")":
        if not bool1:
            return str1
        else:
            return True
    if str1[:2] == "((" and str1[-2:] == "))":
        d = 2
    else:
        d = 1

    for k in range(0, d):
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
                if bool1:
                    return True
    if not bool1:
        return str1
    else:
        return False


def remove_redundant_paren(str1):
    j = 0
    str2 = str1[:2]
    str3 = str1[2:]
    if str2 == '((' and str3 == '))':

        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 1 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
    return str1


def mainconn(str1):
    ostring = copy.copy(str1)
    if os(str1):
        return ["", 0]
    if str1.find("&") < -1 and str1.find(idisj) < -1 and str1.find(iff) < -1 and str1.find(conditional) < -1 and \
                    str1.find(implies) < -1 and str1.find(nonseq) < -1 and str1.find(xorr) < -1:
        return ["", 0]

    str3 = str1
    bool1 = False

    if str1[0] == "~":
        str1 = str1[1:]
        bool1 = True

    j = 0
    bool2 = False
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
            bool2 = True

    j = -1
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == conditional:
            f = -1
        if str2 == idisj or str2 == "&" or str2 == iff or str2 == implies or \
                        str2 == nonseq or str2 == conditional or str2 == xorr:
            if str3 != str2 and str3 != "":
                j = j + 1

            str3 = str2

    if j == -1:
        i = ostring.find(str3)
        return [str3, i]
    k = -1
    j = -1
    while True:
        k += 1
        if k > 150:
            break
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]

            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1

            if j == -1 and (str2 == idisj or str2 == "&" or str2 == iff or str2 == implies
                            or str2 == nonseq or str2 == conditional or str2 == xorr):
                if bool1 and bool2:
                    return [str2, i + 2]
                elif bool2 or bool1:
                    return [str2, i + 1]
                else:
                    return [str2, i]
        else:
            str1 = str1[1:-1]


def isvariable(str3, kind=""):
    bool2 = True
    if str3 == None or str3 == "":
        return False

    if str3 == 'a':
        return False
    elif str3 == 'i':
        return False

    if str3 != "":
        str3 = str3.replace(l1, "")
        str3 = str3.replace(l2, "")
        str3 = str3.replace(neg, "")
        if len(str3) == 1 and str3.islower():
            bool2 = True
        else:
            bool2 = False

    if kind == "":
        return bool2


def os(str1):
    cnx = [xorr, iff, idisj, conditional, implies, nonseq, "&"]
    for i in range(0, len(cnx)):
        if str1.find(cnx[i]) > -1:
            os = False
            return os
    os = True
    return os


def enclose(str1):
    i = -1
    global subscripts
    while i < len(str1) - 1:
        i += 1
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        str4 = str1[i + 1:i + 2]
        if str2.islower() and str4 in subscripts:
            if str3 == "~":
                str1 = str1[:i - 1] + "(~" + str2 + str4 + ")" + str1[i + 2:]
            else:
                str1 = str1[:i] + "(" + str2 + str4 + ")" + str1[i + 2:]
            i += 4
        elif str2.islower():
            if str3 == "~":
                str1 = str1[:i - 1] + "(~" + str2 + ")" + str1[i + 1:]
            else:
                str1 = str1[:i] + "(" + str2 + ")" + str1[i + 1:]
            i += 3
    return str1


def is_natural_language(sentence, i):
    # this determines if the sentence is natural or an abbreviation
    if sentence[i + 1] == "~":
        if sentence[i + 3] in subscripts and sentence[i + 4] == ")":
            return True
        elif sentence[i + 3] == ")":
            return True
    else:
        if sentence[i + 2] in subscripts and sentence[i + 3] == ")":
            return True
        elif sentence[i + 2] == ")":
            return True
    return False


def find_sentences(sentence):
    global subscripts
    if sentence == None: g = 4 / 0
    if os(sentence): g = 4 / 0
    g = sentence.count('(')
    h = sentence.count(')')
    if g != h:
        print('wrong number of parentheses in sentence:' + sentence)
        g = 4 / 0
    if sentence.startswith("(~g)"):
        bb = 8
    marker = False
    il = -1
    total = -1
    c = -1
    neg_value = []
    str1 = ""
    sent1 = []
    sent_type2 = []
    wneg = []
    output = [None] * 9
    # the skel name list names each single sentence after a greek letter, even if
    # the same sentence appears twice it obtains a different name on the second
    # appearance
    skel_nam = []
    sent_num = []
    if sentence.find("~(") > -1:
        sentence = sentence.replace("~(", "(!")
    if sentence.find(implies) > -1:
        str2 = implies
    elif sentence.find(nonseq) > -1:
        str2 = nonseq
    str3 = mainconn(sentence)
    sentence = sentence.strip()
    str4 = str3[0]
    f = str3[1]
    id_num = []

    id_num.append(["1", str4, f])
    sent_num.append([1, '1', sentence, str4, f])
    str21 = ""
    p = 947
    greek_english_dict = {}
    unenclose_at_end = False
    connectives = ["&", idisj, iff, conditional, nonseq, implies, xorr]
    arr1 = []
    mini_c2 = mini_c + neg
    prt = copy.copy(sentence)
    more_num = [chr(945 + x) for x in range(24)]
    temp_string = mainconn(sentence)
    if sentence.find(implies) > -1:
        str1 = implies
    elif sentence.find(nonseq) > -1:
        str1 = nonseq
    else:
        if temp_string == iff:
            str1 = "bicond"
        elif temp_string == conditional:
            str1 = "cond"
        elif temp_string == "&":
            str1 = "cj"

    sent1.append(sentence)
    neg_value.append("")
    sent_type2.append(str1)
    wneg.append(sentence)
    skel_nam.append(None)

    j = 0
    n = 0
    for i in range(0, len(sentence)):
        str1 = sentence[i:(i + 1)]
        for o in connectives:
            if str1 == o:
                j += 1

    while n < j + 1:

        il += 1
        if il > 15:
            break

        e = 0
        l = len(sentence)
        x = -1
        while x < l - 1:
            x += 1
            temp_string = sentence[x:x + 1]
            if sentence[x:x + 1] == "(":
                if not unenclose_at_end:
                    unenclose_at_end = is_natural_language(sentence, x)

                if marker == False:
                    z = x
                    marker = True

                total += 1
            elif sentence[x: x + 1] == ")":
                total -= 1
                if total == -1:
                    marker = False
                    e += 1
                    c += 1

                    temp_sent = sentence[z: x + 1]
                    if temp_sent == '(bIc)':
                        pp = 7
                    otemp_sent = copy.copy(temp_sent)

                    if (len(sentence) - len(temp_sent)) > 2:
                        if temp_sent in prt and temp_sent in str21 and prt != str21:
                            prtnum = findinlist(str21, sent_num, 2, 1)
                            numb = prtnum + "1"
                        elif temp_sent in prt:
                            prtnum = findinlist(prt, sent_num, 2, 1)
                            numb = prtnum + "1"
                        else:
                            prtnum = ""
                            for bb in range(len(sent_num) - 1, -1, -1):
                                str3 = sent_num[bb][2]
                                if os(temp_sent) and str21 != "":
                                    if str21 == str3 and temp_sent in sent_num[bb][2]:
                                        # this is for those basic molecules for which the same sentence appears
                                        # in the definition several times
                                        prtnum = sent_num[bb][1]
                                        break
                                else:
                                    if temp_sent in sent_num[bb][2]:
                                        prtnum = sent_num[bb][1]
                                        break
                            # if prtnum == "":
                            #     easygui.msgbox('your sentences are not numbered properly')
                            g = len(prtnum) + 1
                            f = 0

                            for bb in range(len(sent_num) - 1, -1, -1):
                                temp_sn = sent_num[bb][1]
                                h = temp_sn[:g - 1]
                                hh = sent_num[bb][0]
                                if g > sent_num[bb][0]:
                                    break
                                if sent_num[bb][0] == g and temp_sn[:g - 1] == prtnum:
                                    f += 1

                            f += 1
                            if f < 10:
                                numb = prtnum + str(f)
                            else:
                                numb = prtnum + more_num[f - 10]

                        prt = temp_sent
                        temp_mc = mainconn(temp_sent)
                        mc = temp_mc[0]
                        str3 = temp_mc[0]
                        g = temp_mc[1]
                        mc_num = temp_mc[1]
                        sent_num.append([len(numb), numb, temp_sent, str3, g])

                        if os(temp_sent):

                            # n counts the number of single sentences
                            n += 1
                            if temp_sent.find("~") > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace("~", "")

                            elif temp_sent.find(mini_c2) > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace(mini_c2, mini_c)

                            elif temp_sent.find(mini_c2) == -1 and temp_sent.find(mini_c) > -1:
                                neg_value.append("")
                            elif temp_sent.find("~") == -1:
                                neg_value.append("")
                            else:
                                break  # stop
                        else:
                            neg_value.append("")

                        sent1.append(temp_sent)
                        wneg.append(otemp_sent)
                        id_num.append([numb, mc, mc_num])

                        if os(otemp_sent):
                            if otemp_sent in greek_english_dict:
                                skel_nam.append(greek_english_dict.get(otemp_sent))
                            else:
                                p += 1
                                greek_english_dict.update({otemp_sent: chr(p)})
                                skel_nam.append(chr(p))
                        else:
                            skel_nam.append(None)
                    else:
                        sentence = sentence[1:len(sentence) - 1]
                        l = len(sentence)
                        x = -1
                        c -= - 1
                        e -= - 1

        total = -1
        marker = False
        w = -1

        if n < j + 1:
            if len(sent1) > w:
                while w + 1 < len(sent1):
                    if w == 13:
                        pp = 7
                    w += 1
                    str21 = sent1[w]
                    if not os(str21) and w != 0:
                        if str21 not in arr1:
                            sentence = str21
                            arr1.append(sentence)
                            break

    for i in range(len(sent1)):
        temp_string = sent1[i]
        if temp_string.find("(!") > -1:
            sent1[i] = sent1[i].replace("(!", "~(")
            wneg[i] = wneg[i].replace("(!", "~(")

    if unenclose_at_end:
        for i in range(len(sent1)):
            sent1[i] = unenclose(sent1[i])
            wneg[i] = unenclose(wneg[i])

    output[0] = sent1
    output[1] = neg_value
    output[2] = sent_type2
    output[3] = wneg
    output[4] = id_num
    output[6] = translate_to_greek(skel_nam, wneg, id_num)
    output[5] = skel_nam[0]

    return output


def translate_to_greek(skel_nam, wneg, id_num):
    for i in range(len(skel_nam)):
        if skel_nam[i] == None:
            to_be_translated = wneg[i]
            for j in range(len(skel_nam) - 1, -1, -1):
                if skel_nam[j] != None:
                    if wneg[j] in to_be_translated and id_num[j][1] == "":
                        to_be_translated = to_be_translated.replace(wneg[j], skel_nam[j])
            skel_nam[i] = to_be_translated

    return skel_nam




def remove_duplicates(list1, i):
    list2 = []
    j = -1
    while j < len(list1) - 1:
        j += 1
        if list1[j][i] in list2:
            del list1[j]
            j -= 1
            print ("remove duplicates used")
        else:
            list2.append(list1[j][i])

    return list1


def set_default_list(str1, list1):
    print ('make sure this does not mess up the abbreviations global')
    g = 4 / 0
    str2 = findinlist(str1, list1, 1, 0)
    if str2 == None:
        str2 = variables[0]
        del variables[0]
        list1.append([str2, str1])

    return str2


def step_two():
    global sn, all_sent, time_spent_reducing
    aa = time.time()

    all_sent = remove_duplicates(all_sent, 0)

    all_sent = define_irregular_terms(all_sent)

    time_spent_reducing += (time.time() - aa)

    define_regular_terms(all_sent)

    add_necessary_conditions_for_concept()


def get_decision_procedure(list1):
    list2 = []
    for sent in list1:
        for i in sent[46]:
            list2.append(i)

    return list2

def sort_decisions(list1):
    for lists in list1:
        lists[46].sort()


def define_irregular_terms(list1, type=""):
    # the code for avoiding the circularity of defining 'i' is found in the add to all sent function

    do_not_define_again = []
    decision_procedure = get_decision_procedure(list1)
    decision_procedure.sort()
    sort_decisions(list1)
    m = -1
    while decision_procedure != []:
        m += 1
        while list1[m][46] != []:
            category = list1[m][46][0]
            i = list1[m][45][category][0]
            decision_procedure.remove(category)
            del list1[m][46][0]
            del list1[m][45][category][0]

            if not lies_wi_scope_of_univ_quant(list1[m], i) and \
                            list1[m][42] not in do_not_define_again and \
                    do_not_define_i_again(list1[m][i], do_not_define_again):
                if list1[m][i] == "i":
                    do_not_define_again.append("aaa")
                else:
                    do_not_define_again.append(list1[m][42])
                antecedent = copy.deepcopy(list1[m])
                consequent = copy.deepcopy(list1[m])
                _ = determine_which_function_to_use(consequent,
                                                   i,
                                                   category,
                                                   decision_procedure,
                                                   type)
                consequent, rule, decision_procedure = _
                prepare_irregular_att_sent(antecedent, consequent, rule, list1, type)
                del list1[m]
                m -= 1
                break

    return list1


def do_not_define_i_again(str1, do_not_define_again):
    if str1 == 'i' and "aaa" in do_not_define_again:
        return False
    else:
        return True


def prepare_irregular_att_sent(antecedent, consequent, rule, list1, type=""):
    if rule == None or (type == 'universal' and rule != "RDC"):
        return

    if type != "universal":
        for sent in consequent:
            list1.append(sent)

    if len(consequent) == 1:
        prepare_att_sent_1_sent(antecedent, rule, iff, consequent)
    elif len(consequent) == 2:
        prepare_att_sent_2_sent(antecedent, iff, consequent, rule)
    elif len(consequent) == 3:
        prepare_att_sent_3_sent(antecedent, iff, consequent, rule)
    elif len(consequent) == 4:
        prepare_att_sent_4_sent(antecedent, consequent, iff, rule)


def determine_which_function_to_use(list1, i, j, decision_procedure, type=""):
    if j == 1:
        consequent, rule = change_variables(list1, i)
    elif j == 3:
        consequent, rule = eliminate_common_name_possessives(list1, i)
    elif j == 4:
        consequent, rule = eliminate_proper_name_possessives3(list1, i)
    elif j == 5:
        consequent, rule = eliminate_and_coordinator(list1, i)
    elif j == 6:
        consequent, rule = eliminate_adjectives(list1, i)
    elif j == 7:
        consequent, rule = eliminate_concept_instance_apposition(list1, i)
    elif j == 8:
        consequent, rule = eliminate_parenthetical_phrase(list1, i)
    elif j == 9:
        consequent, rule, decision_procedure = eliminate_relative_pronouns(list1, i, decision_procedure)
    elif j == 10:
        consequent, rule = eliminate_that_subordinator(list1, i)
    elif j == 11:
        consequent, rule = eliminate_as(list1, i)
    elif j == 12:
        consequent, rule = divide_relations2(list1, type)
    elif j == 13:
        consequent, rule = divide_relations1(list1, i)
    elif j == 14:
        consequent, rule = eliminate_there1(list1)
    elif j == 15:
        consequent, rule = eliminate_universals(list1, i)

    return consequent, rule, decision_procedure


# aaa




#
# def eliminate_possessives(list1):
#     m = -1
#     while m < (len(list1)) - 1:
#         m += 1
#         for i in list1[45][3]:
#             if not lies_wi_scope_of_univ_quant(m, i):
#                 eliminate_possessives2(m, i)
#                 del list1
#                 m -= 1
#                 break
#
#
# def eliminate_possessives2(list1, i):
#     ant_sent_parts = copy.deepcopy(list1)
#     list1 = [None] * 80
#     str1 = list1[i][0]
#     list1[5] = str1
#     list1[9] = "OWN"
#     if i == 69:
#         str2 = list1[5]
#     elif i == 70:
#         str2 = list1[14]
#     list1[14] = str2
#     list1[i] = None
#     con_parts = copy.deepcopy(build_sent2(list1))
#     con_parts2 = copy.deepcopy(build_sent2(list1))
#     list1.append(con_parts2)
#     prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, "PNE", "e")


def eliminate_possessive_nouns(list1, i):
    list7 = [None] * 80
    str1 = list1[i][0]
    list7[5] = str1
    list7[9] = "OWN"
    if i == 69:
        str2 = list1[5]
    elif i == 70:
        str2 = list1[14]
    list7[14] = str2
    list1[i] = None
    con_parts = copy.deepcopy(build_sent2(list1))
    con_parts2 = copy.deepcopy(build_sent2(list7))
    consequent = [con_parts, con_parts2]

    return consequent, "PNE"


def eliminate_proper_name_possessives3(list1, i):
    if i == 69:
        concept_position = 5
    elif i == 70:
        concept_position = 14
    con_parts1 = copy.deepcopy(list1)
    concept = list1[concept_position]

    new_object = variables[0]
    del variables[0]

    con_parts1[concept_position] = new_object
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(new_object, "", "I", concept)
    con_parts3 = build_sent1(new_object, "", "J", "definite")
    append_to_all_sent(list1, [con_parts2, con_parts3])

    prepare_att_sent_3_sent(ant_parts, connective, consequent, "PNP")

    ant_parts = copy.deepcopy(con_parts1)
    possessor = con_parts1[i][:-2]
    con_parts1[i] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(possessor, "", "OWN", new_object)
    consequent = [con_parts1, con_parts2]

    return consequent, rule


    # if concept[-2:] is "'s":
    #     apostrophe_s = "'s"
    #     concept = concept[:-2]
    #     rule = "PNE"


def eliminate_common_name_possessives(list1, i):
    # if "the" is followed by a common name possessive then it cannot be eliminated

    if i == 69:
        concept_position = 5
        determ_position = 3
    elif i == 70:
        concept_position = 14
        determ_position = 10
    con_parts1 = copy.deepcopy(list1)
    concept = list1[concept_position]

    new_object = set_default_list(concept, abbreviations)
    possessor_concept = list1[i][:-2]
    new_possessor = set_default_list(possessor_concept, abbreviations)

    con_parts1[concept_position] = new_object
    con_parts1[i] = new_possessor + "'s"
    con_parts1[determ_position] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(new_object, "", "I", concept)
    con_parts3 = build_sent1(new_possessor, "", "I", possessor_concept)
    append_to_all_sent(list1, [con_parts2, con_parts3])

    prepare_att_sent_3_sent(ant_parts, connective, consequent, "DF the")

    ant_parts = copy.deepcopy(con_parts1)
    possessor = con_parts1[i][:-2]
    con_parts1[i] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(possessor, "", "OWN", new_object)
    consequent = [con_parts1, con_parts2]

    return consequent, rule


def eliminate_and_coordinator(list1, i):
    # this seperates a sentence with an 'and' coordinator into two


    list1[66] = None
    list7 = [None] * 80
    list7[5] = list1[67]
    list1[67] = None
    for i in range(6, 20):
        list7[i] = list1[i]
    list7 = new_categories(list7, True)
    list1 = build_sent2(list1)
    con_parts = copy.deepcopy(list7)
    con_parts2 = copy.deepcopy(list1)
    consequent = [con_parts, con_parts2]

    return consequent, "DE and" + uc


def eliminate_adjectives(list1, i):
    list7 = [None] * 80
    slots = [5, 9, 14]

    if i == 13:
        r = 9
    else:
        r = i - 2
    if list1[8] != None:
        str7 = "~"
        list1[8] = None
    else:
        str7 = None
    if str7 != None:
        list7[8] = str7
    if list1[r] != "I":
        list7[5] = list1[i + 1]
    else:
        list7[5] = list1[5]
    list7[9] = "J"
    list7[14] = list1[i]
    list1[i] = None
    list7[46] = []

    con_parts = build_sent2(list7)
    con_parts2 = build_sent2(list1)
    consequent = [con_parts, con_parts2]

    return consequent, "ADJ E"


def eliminate_concept_instance_apposition(list1, i):
    list7 = [None] * 80
    if i == 35:
        j = 5
    elif i == 36:
        j = 14
    elif i == 37:
        j = 18
    elif i == 38:
        j = 22
    str1 = list1[j]
    list1[j] = list1[i]
    list7[14] = str1
    list7[5] = list1[i]
    list7[9] = "I"
    list1[i] = None
    con_parts = copy.deepcopy(build_sent2(list7))
    list1 = build_sent2(list1)
    con_parts2 = copy.deepcopy(list1)
    consequent = [con_parts, con_parts2]

    return consequent, "CIA"


def eliminate_parenthetical_phrase(list1, i):
    sub_begin = list1[57][0]
    sub_end = list1[57][1]
    rule = "DE " + list1[i]
    list1[i] = None
    list2 = [None] * 80
    list2[3] = list1[5]
    k = 3
    for j in allowable_slots():
        if list1[j] != None:
            absolute_pos = allowable_slots().index(j)
            if absolute_pos > sub_begin and absolute_pos < sub_end:
                k += 1
                list2[k] = list1[j]
                list1[j] = None
            if absolute_pos > sub_end:
                break

    list1 = restore_original_sent(list1)
    list2 = categorize_words(list2)

    return [list1, list2], rule


# ddd
def eliminate_relative_pronouns(list1, i, decision_procedure):
    rule = "DE " + list1[i]
    con_parts1 = copy.deepcopy(list1)
    con_parts2 = [None] * 80
    map_nouns_to_relative_pronouns = {60: 14, 61: 18, 62: 22}
    j = map_nouns_to_relative_pronouns.get(i)
    relation_positions = [15, 19, 23, 27, 31]
    con_parts1[i] = None
    con_parts2[3] = con_parts1[j]
    k = 3
    started = False
    for j in allowable_slots():
        if j == 49:
            bb = 8
        if j == i:
            started = True
        if started and con_parts1[j] != None:
            k += 1
            if j != i:
                con_parts2[k] = con_parts1[j]
                con_parts1[j] = None
            if j in relation_positions:
                # the following must be changed if the decision procudure changes in
                # the get_used_slots function
                if j in con_parts1[45][13]: con_parts1[45][13].remove(j)
                if 13 in con_parts1[46]: con_parts1[46].remove(13)

    con_parts2 = categorize_words(con_parts2)
    con_parts1 = build_sent2(con_parts1)
    consequent = [con_parts1, con_parts2]
    decision_procedure = con_parts1[46] + con_parts2[46]
    return consequent, rule, decision_procedure


def eliminate_that_subordinator(list1, i):
    num = allowable_slots()
    list7 = [None] * 80
    bool1 = False
    list3 = copy.deepcopy(list1)

    k = 2
    for j in num:
        if j == i:
            bool1 = True
        if list3[j] != None and bool1 and j != i:
            k += 1
            list7[k] = list1[j]
            list3[j] = None

    list2 = categorize_words(list7)
    list2[0] = remove_outer_paren(list2[0])
    str3 = list2[0].replace(" ", "")
    g = abbreviations[1].get(str3)
    if g == None:

        for z in range(len(variables)):
            if variables[z] in prop_var:
                new_var = variables[z]
                del variables[z]
                prop_var.remove(new_var)
                break
        abbreviations[0].update({new_var: str3})
        abbreviations[1].update({str3: new_var})
    else:
        new_var = g

    list3[i] = None
    if i == 7:
        if list3[5] == 'it':
            list3[5] = new_var
        else:
            list3[14] = new_var
    elif i == 60:
        list3[5] = new_var

    elif i == 61:
        if list3[14] == 'it':
            list3[14] = new_var
        else:
            list3[18] = new_var

    elif i == 62:
        if list3[18] == 'it':
            list3[62] = new_var
        else:
            list3[22] = new_var

    list3 = build_sent2(list3)
    list3 = new_categories(list3, True)
    con_parts = copy.deepcopy(list3)
    consequent = [con_parts]

    return consequent, "DE that"


def eliminate_as(list1, i):
    con_parts1 = [None] * 80
    con_parts1[5] = list1[5]
    con_parts1[9] = list1[9]
    con_parts1[14] = list1[14]
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = [None] * 80
    con_parts2[5] = list1[18]
    con_parts2[9] = list1[9]
    con_parts2[14] = list1[14]
    con_parts2 = build_sent2(con_parts2)
    consequent = [con_parts1, con_parts2]

    return consequent, "DE AS"


def divide_relations1(list1, i):
    # b R c S d = b R c & b S d
    # b R c S d T e = b R c & b S d T e
    list2 = [None] * 80
    slots = []
    dict1 = {5: 5, 49: 8, 15: 9, 18: 14, 19: 15, 22: 18}
    for k, v in dict1.items():
        if list1[k] != None:
            list2[v] = list1[k]  # todo we need to revise list45 and 46 here
            if k != 5:
                list1[k] = None

    list2[46] = []
    list1 = build_sent2(list1)
    list2 = build_sent2(list2)
    consequent = [list1, list2]

    return consequent, "RDA"


def divide_relations2(list1, type=""):
    # (b R c S d) = (b S d) & (c S d)
    list2 = build_sent1(list1[5], "", list1[15], list1[18])
    list3 = build_sent1(list1[14], "", list1[15], list1[18])
    list2[46] = []
    list3[46] = []

    consequent = [list2, list3]

    return consequent, "RDC"


def eliminate_there1(list1):
    con_parts = [None] * 80
    con_parts[9] = 'EX'
    con_parts[5] = list1[14]
    con_parts[56] = [5, 9]
    con_parts[54] = False
    con_parts = build_sent2(con_parts)
    consequent = [con_parts]

    return consequent, "DE there"


# ddd

def eliminate_universals(list1, i):
    fir_antecedent = copy.deepcopy(list1)
    rule = "DE " + list1[i]
    class_sent = get_class_sent(list1, i)

    if has_scope_over_subclause(list1, i):
        antecedent, consequent = extract_words_from_subclause(list1, i, class_sent[5])
        antecedent = define_irregular_terms([antecedent], "universal")
    elif has_scope_over_adjective(list1, i):
        antecedent = turn_adj_into_sent(i, m)
    elif list1[i] == 'no' and list1[10] == 'a' and i == 3:
        antecedent, consequent = eliminate_no_w_indef_obj(class_sent, list1)
    else:
        consequent = get_the_simple_consequent(list1, i, class_sent[5])
        antecedent = []
    antecedent.append(class_sent)
    for sent in antecedent:
        all_sent.append(sent)
    all_sent.append(consequent)

    prepare_att_sent_univ(fir_antecedent, antecedent, consequent, rule)

    return None, None


def get_the_simple_consequent(list1, i, new_var):
    j = 14 if i == 10 else i + 2
    list1[j] = new_var
    if list1[i] == 'no':
        list1[8] = "~"
    list1[i] = None

    return build_sent2(list1)


def eliminate_no_w_indef_obj(class_sent, list1):
    list1[3] = None
    consequent = build_sent1(class_sent[5], "~", list1[9], variables[0])
    sec_antecedent = build_sent1(variables[0], "", "I", list1[14])
    del variables[0]

    return [sec_antecedent], consequent


def has_scope_over_adjective(list1, i):
    k = 13 if i == 10 else i + 1
    if i == 3 and list1[k] != None:
        return True
    return False


def prepare_att_sent_univ(fir_antecedent, antecedent, consequent, rule):
    global sn
    list1 = [""] * 50
    list2 = []
    list42 = []
    bare_sent = []
    str2 = ""
    str3 = ""
    greek = ""
    n = 933
    consequent[44] = chr(933)
    bare_sent.append(consequent[1])
    consequent[68] = "12"
    consequent[53] = 'q'
    heir_num = "11"
    for i, sent in enumerate(antecedent):
        list2.append([sent[1], sent[2]])
        list42.append([sent[72], sent[2]])
        bare_sent.append(sent[1])
        if str2 == "":
            str2 = sent[0]
            str3 = sent[42]
            greek = chr(n + i + 1)
            sent[44] = chr(n + i + 1)
        else:
            str2 += " & " + sent[0]
            str3 += " & " + sent[42]
            greek += " & " + chr(n + i + 1)
            sent[44] = chr(n + i + 1)
        sent[68] = heir_num + str(i + 1)
        if len(antecedent) > 1:
            sent[53] = "ca"
        else:
            sent[53] = "a"
    if len(antecedent) > 1:
        str2 = "(" + str2 + ")"
        str3 = "(" + str3 + ")"
        greek = "(" + greek + ")"

    list1[0] = list2
    list1[1] = [[consequent[1], consequent[2]]]
    list1[3] = 'c'
    list1[4] = str3 + " " + conditional + " " + consequent[42]
    list1[7] = [str3, ""]
    list1[8] = [consequent[1], consequent[2]]
    list1[34] = antecedent
    list1[35] = [consequent]
    list1[37] = str2 + " " + conditional + " " + consequent[0]
    list1[38] = bare_sent
    list1[40] = [str2, ""]
    list1[41] = [consequent[72], ""]
    list1[42] = list42
    list1[43] = [[consequent[72], consequent[2]]]
    list1[47] = greek + " " + conditional + " " + chr(933)

    sn += 1
    fir_antecedent[44] = chr(932)
    list5 = [""] * 50
    list5[0] = [[fir_antecedent[1], fir_antecedent[2]]]
    list5[1] = [[list1[4], ""]]
    list5[2] = sn
    list5[3] = "e"
    list5[4] = fir_antecedent[42] + " " + iff + " (" + list1[4] + ")"
    list5[7] = [fir_antecedent[1], fir_antecedent[2]]
    list5[8] = [list1[4], ""]
    list5[34] = [fir_antecedent]
    list5[37] = fir_antecedent[0] + " " + iff + " (" + list1[37] + ")"
    bare_sent.append(fir_antecedent[1])
    list5[38] = bare_sent
    list5[39] = [list1]
    list5[40] = [fir_antecedent[72], fir_antecedent[2]]
    list5[41] = [list1[37], ""]
    list5[42] = [[fir_antecedent[0], ""]]
    list5[43] = [[list1[37], ""]]
    list5[47] = fir_antecedent[44] + " " + iff + " (" + list1[47] + ")"
    attach_sent.append(list5)
    add_to_total_sent(sn, list5[37], list5[4], "", rule)


def get_class_sent(list1, i):
    general_variables = copy.deepcopy(variable_type[0])
    class_pos = 14 if i == 10 else i + 2
    list2 = build_sent1(variables[0], "", "I", list1[class_pos])
    general_variables.append(variables[0])
    variable_type[0] = general_variables
    del variables[0]

    return list2


#
# def turn_adj_into_sent(i,m):
#
#     pass



def has_scope_over_subclause(list1, i):

    if i == 3 and (list1[59] != None or list1[15] != None):
        bool1 = True
    elif i == 10 and list1[60] != None:
        bool1 = True
    elif i == 16 and list1[61] != None:
        bool1 = True
    elif i == 20 and list1[62] != None:
        bool1 = True
    else:
        bool1 = False
    return bool1


def get_last_relation(list1):
    relational_positions = [31, 27, 23, 19, 15]
    for pos in relational_positions:
        if list1[pos] != None:
            return pos

def extract_words_from_subclause(list1, i, new_var):
    # this is the sentence that will be inserted into the antecedent in the
    # definition of 'every' or 'no', in the future it should output
    # a set of lists, not just 1

    comma = get_last_relation(list1) if list1[39] == None else list1[39]
    new_sent = [None] * 80
    new_var_loc = 14 if i == 5 else i + 2
    list1[new_var_loc] = new_var
    list1[i] = None
    for num in [59, 60, 61, 62]:
        if list1[num] != None:
            list1[num] = None

    k = 2
    for n in allowable_slots():
        if list1[n] != None:
            if comma == n:
                break
            k += 1
            new_sent[k] = list1[n]
            if n != new_var_loc:
                list1[n] = None

    new_sent = categorize_words(new_sent)
    list1 = restore_original_sent(list1)

    return new_sent, list1


# def extract_words_from_subclause(list1, i, new_var):
#     # this is the sentence that will be inserted into the antecedent in the
#     # definition of 'every' or 'no', in the future it should output
#     # a set of lists, not just 1
#
#
#     if list1[55] == [0,0]:
#         comma = get_last_relation(list1)
#     else:
#         comma = list1[57][1]
#
#     new_sent = [None] * 80
#     new_var_loc = 14 if i == 5 else i + 2
#     list1[new_var_loc] = new_var
#     list1[i] = None
#     for num in [59, 60, 61, 62]:
#         if list1[num] != None:
#             list1[num] = None
#             list1[56].remove(num)
#
#     k = 2
#     n = -1
#     for j in allowable_slots():
#         if list1[j] != None and list1[j] != "":
#
#             n += 1
#             j = list1[56][n]
#             if comma == j:
#                 break
#             k += 1
#             new_sent[k] = list1[j]
#             if j != new_var_loc:
#                 list1[j] = None
#                 list1[56].remove(j)
#                 n -= 1
#     new_sent = categorize_words(new_sent)
#     list1 = restore_original_sent(list1)
#
#     return new_sent, list1


def restore_original_sent(list1):
    k = 2
    list2 = [None] * 80
    for i in allowable_slots():
        if list1[i] != None:
            k += 1
            list2[k] = list1[i]
    list2 = categorize_words(list2)

    return list2


def insert_item(_dict, pos, obj):

    lambda _dict, obj, pos: {k: v for k, v in
                             (list(_dict.items())[:pos] + list(obj.items()) + list(_dict.items())[pos:])}
    return _dict


def append_to_all_sent(list1, list2):
    for sent in list2:
        list1.append(sent)


def get_negative_position(i):
    # when we define many we have to put the negation sign in a weird position if
    # many is in an unusual place
    dict1 = {5: 8, 14: 8, 63: 49, 18: 49, 64: 50, 22: 50, 65: 51}
    j = dict1.get(i)
    return j


def define_regular_terms(list1):
    dictionary[6] = use_rarely_defined_word()
    do_not_define_again = []
    m = -1
    while m < len(list1) - 1:
        m += 1
        if isdefineable(list1[m]) and list1[m][42] not in do_not_define_again:
            do_not_define_again.append(list1[m][42])
            change_variables(list1[m], 0)


def get_definiendum(list1, i, antecedent):
    if antecedent != "":
        pass

    special_relations = ["I", "J", "H"]
    if i != 0:
        definiendum = list1[i]
        defining_abbreviation = [variables[0]]
        del variables[0]
    elif list1[9] == "=":
        definiendum = list1[14]
        defining_abbreviation = [list1[5]]
    elif list1[9] in special_relations:
        definiendum = abbreviations[0].get(list1[14])
        defining_abbreviation = [list1[5]]
    else:
        definiendum = list1[9]
        if list1[14] != None:
            defining_abbreviation = [list1[5], list1[14]]
        else:
            defining_abbreviation = [list1[5]]

    return definiendum, defining_abbreviation


def change_variables(list1, def_loc, antecedent=""):
    global time_spent_defining
    aa = time.time()

    definiendum, defining_abbreviations = get_definiendum(list1, def_loc, antecedent)

    if definiendum == None or definiendum in dictionary[6]:
        return

    if definiendum == 'non_whole':
        bb = 8

    definition = dictionary[1].get(definiendum)

    if definition == None:
        return

    def_info2 = find_sentences(definition)

    def_info = copy.deepcopy(def_info2)

    constant_map, temp_prop_const = get_abbreviations_from_definition(def_info)

    def_info = eliminate_conjuncts_from_definition(def_info)

    _ = get_new_sent(def_info, defining_abbreviations, def_loc, list1, definiendum)

    def_abbrev_dict, r_sent_loc, new_sentences = _

    total_dict = {**def_abbrev_dict, **constant_map}

    _ = replace_constants(total_dict, temp_prop_const, new_sentences)

    new_sentences, unfill_positions, prop_unfill = _

    _ = replace_indefinite_variables(new_sentences, unfill_positions, list1[1])

    new_sentences, indefinite_dict, rn_type = _

    total_dict = {**total_dict, **indefinite_dict}

    _ = replace_propositional_constants(temp_prop_const, prop_unfill, new_sentences, total_dict)

    new_sentences, old_prop_new_prop = _

    new_sentences = replace_r_sent(total_dict, r_sent_loc, new_sentences, list1, def_loc)

    new_sentences = [build_sent2(sent) for sent in new_sentences]

    new_sentences = add_first_sent_to_def_sent(new_sentences, list1, r_sent_loc)

    rename = build_rename_sent2(constant_map, def_abbrev_dict, old_prop_new_prop, indefinite_dict, rn_type)

    rule = get_rule(definiendum, r_sent_loc, def_info[0], rename)

    add_to_attach_sent(def_info, new_sentences, definition, r_sent_loc, rename, rule)

    add_def_sent_to_all_sent(definiendum, new_sentences)

    time_spent_defining += (time.time() - aa)

    if r_sent_loc != [] or definiendum == 'i':
        return list1, None


# bbb
def get_rule(definiendum, r_sent_loc, def_info, rename):
    if r_sent_loc != [] or rename == "":
        rule = "DE "
    else:
        rule = "DF "
    if def_info[4][0][1] == "":
        if def_info[4][1][1] == conditional:
            rule = "NC "
    elif def_info[4][0][1] == conditional:
        rule = "NC "

    rule += definiendum

    return rule


def replace_propositional_constants(temp_prop_const, prop_unfill, new_sentences, total_dict):
    if prop_unfill == []:
        return new_sentences, {}

    old_prop_to_new_prop = {}
    for k, v in temp_prop_const.items():
        for num in [5, 14, 18]:
            if v[num] != None:
                var = total_dict.get(v[num])
                if var != None:
                    v[num] = var
                else:
                    if v[num] in variables[0]:
                        variables.remove(v[num])
                        total_dict.update({variables[0]: variables[0]})
                    else:
                        total_dict.update({v[num]: variables[0]})
                        v[num] = variables[0]
                        del variables[0]

    for k, v in temp_prop_const.items():
        old_key = k
        v = build_sent2(v)
        old_prop_to_new_prop.update({old_key: v[1]})

    for num in prop_unfill:
        i, j = num[0], num[1]
        new_sentences[i][j] = old_prop_to_new_prop.get(new_sentences[i][j])
        assert new_sentences[i][j] != None

    return new_sentences, old_prop_to_new_prop


def add_to_attach_sent(def_info, new_sentences, definition, r_sent_loc, rename, rule):
    for i in range(len(def_info)):
        list1 = prepare_attach_sent(def_info[i], new_sentences, r_sent_loc)
        if list1[45] == "append to attach_sent list":
            attach_sent.append(list1)
        if i == 0:
            add_definitions_to_total_sent(list1, rule, rename, r_sent_loc, definition)
        if def_info[i][2] == 'eliminate as conjunct':
            list1[46] = 'eliminate as conjunct'


def add_definitions_to_total_sent(temp_attach_sent, rule, rename, r_sent_loc, definition):
    if r_sent_loc != [] or rename == "":
        add_to_total_sent(temp_attach_sent[2], temp_attach_sent[37], temp_attach_sent[4], "", rule)
    else:
        num = temp_attach_sent[2]
        add_to_total_sent(num - 2, definition, "", "", rule)
        add_to_total_sent(num - 1, rename, "", "", "RN")
        add_to_total_sent(num, temp_attach_sent[37], temp_attach_sent[4], "", "SUB")


def build_rename_sent2(constant_map, def_abbrev_dict, old_prop_to_new_prop, indefinite_dict, rn_type):
    rename_sent = ""
    for k, v in def_abbrev_dict.items():
        if k != v:
            if rename_sent == "":
                rename_sent += "(" + k + mini_c + v + ")"
            else:
                rename_sent += " & (" + k + mini_c + v + ")"

    for k, v in constant_map.items():
        if k != v:
            if rename_sent == "":
                rename_sent += "(" + k + idd + v + ")"
            else:
                rename_sent += " & (" + k + idd + v + ")"

    for k, v in indefinite_dict.items():

        if k != v:
            rn = rn_type.get(v)
            if rename_sent == "":
                rename_sent += "(" + k + idd + v + ")" + rn
            else:
                rename_sent += " & (" + k + idd + v + ")" + rn

    for k, v in old_prop_to_new_prop.items():
        if k != v:
            if rename_sent == "":
                rename_sent += "(" + k + idd + v + ")"
            else:
                rename_sent += " & (" + k + idd + v + ")"

    return rename_sent


def add_def_sent_to_all_sent(definiendum, new_sentences):
    if definiendum == 'i':
        b = 0
        new_sentences[1][46].remove(1)
        new_sentences[1] = remove_i_from_45(new_sentences[1])
    else:
        b = 1

    for i in range(b, len(new_sentences)):
        # we cannot add the first_sent to the all sent list since it is already in there
        # if the definiendum is 'i' then we need to add it since it will later be deleted
        if not isinmdlist(new_sentences[i][1], all_sent, 1) or definiendum == "i":
            all_sent.append(new_sentences[i])


def add_first_sent_to_def_sent(defin_sent, first_sent, r_sent_loc):
    new_sent = copy.deepcopy(first_sent)
    if r_sent_loc == []:
        if new_sent[8] == "~":
            new_sent[8] = None
            new_sent = build_sent2(new_sent)

    for i in range(len(defin_sent)):
        if defin_sent[i][68][1] == "1":
            greek_name = defin_sent[i][44]
            del defin_sent[i]
            break
    new_sent[44] = greek_name
    defin_sent.insert(0, new_sent)

    return defin_sent


def replace_r_sent(total_dict, r_sent_location, new_sentences, list1, def_loc):
    # modify this if you add on new non-word info to the all_sent list
    # right now the only words that have two r sentences in their definition are 'many' and 'only'
    # since the only thing that distinguishes these r sentences is the negation sign we just
    # hard code that the first r sentence is positive and the second is negative
    if r_sent_location == []:
        return new_sentences
    determinative_positions = [3, 10, 16, 20, 24, 28, 32]
    if def_loc in determinative_positions:
        new_var_loc = 14 if def_loc == 10 else def_loc + 2
    else:
        new_var_loc = def_loc

    for j, location in enumerate(r_sent_location):
        r_sent = [None] * 80
        for i in [44, 45, 46, 56, 68]:
            r_sent[i] = new_sentences[location][i]
        for i in [45, 46, 56, 57]:
            r_sent[i] = list1[i]

        if j == 0: new_var = list(total_dict.values())[0]
        # not that the variable of the second R sentence must always be the third member
        # of the total_dict
        if j == 1: new_var = list(total_dict.values())[2]
        r_sent[new_var_loc] = new_var

        for i in allowable_slots():
            if list1[i] != None:
                if i != new_var_loc and i != def_loc:
                    r_sent[i] = list1[i]
        if j == 1:
            r_sent[8] = "~"

        new_sentences[location] = r_sent

    return new_sentences


def replace_indefinite_variables(new_sentences, unfill_positions, current_sent):
    dict1 = {14: 5, 5: 14}
    indefinite_dict = {}
    rn_type = {}
    k = -1
    while k < len(unfill_positions) - 1:
        k += 1
        i = unfill_positions[k][0]
        j = unfill_positions[k][1]
        m = dict1.get(j)
        if m == None:
            return
        new_sentences[i][j]

        for sent in all_sent:
            if sent[1] != current_sent:
                if sent[2] == new_sentences[i][2] and sent[9] == new_sentences[i][9] \
                        and sent[m] == new_sentences[i][m]:
                    n = dict1.get(m)
                    indefinite_dict.update({new_sentences[i][j]: sent[n]})
                    rn_type.update({sent[n]: l1})
                    new_sentences[i][j] = sent[n]
                    del unfill_positions[k]
                    k -= 1
                    break

    _ = replace_indefinite_variables2(indefinite_dict, new_sentences, unfill_positions, rn_type)

    new_sentences, indefinite_dict, rn_type = _

    return new_sentences, indefinite_dict, rn_type


def replace_indefinite_variables2(indefinite_dict, new_sentences, unfill_positions, rn_type):
    if unfill_positions == []:
        return new_sentences, indefinite_dict, rn_type

    for pos in unfill_positions:
        i = pos[0]
        j = pos[1]
        new_var = indefinite_dict.get(new_sentences[i][j])
        if new_var != None:
            new_sentences[i][j] = new_var
        else:
            new_var = variables[0]
            del variables[0]
            variable_type[1].append(new_var)
            indefinite_dict.update({new_sentences[i][j]: new_var})
            rn_type.update({new_var: l2})
            new_sentences[i][j] = new_var

    return new_sentences, indefinite_dict, rn_type


def replace_constants(total_dict, temp_prop_const, new_sentences):
    unfill_positions = []
    prop_unfill = []
    j = -1
    for sent in new_sentences:
        j += 1
        if sent[68][1:] != "1" and sent[9] != "R" and sent[68][1:] != "11":
            for i in [5,14,18,22]:
                if sent[i] != None:
                    new_var = total_dict.get(sent[i])
                    if new_var != None:
                        sent[i] = new_var
                    elif sent[i] in temp_prop_const.keys():
                        prop_unfill.append([j, i])
                    else:
                        unfill_positions.append([j, i])

    return new_sentences, unfill_positions, prop_unfill


def get_new_sent(def_info, defining_abbreviations, def_loc, list1, definiendum):
    new_sentences = []
    r_sent_location = []
    do_not_add = []
    temp_dict1 = {}
    temp_dict2 = {}
    n = -1
    o = 0
    for j in range(len(def_info[0][3])):

        if os(def_info[0][3][j]):
            sent = copy.deepcopy(def_info[0][3][j])
            sent = categorize_words(space_words(sent))
            if sent[42] not in do_not_add:
                n += 1
                do_not_add.append(sent[42])
                new_sentences.append(sent)
                sent[68] = def_info[0][4][j][0]
                sent[44] = def_info[0][6][j]

                if def_info[0][4][j][0][1:] == "1" or def_info[0][4][j][0][1:] == "11":
                    def_abbrev_dict, k = map_defining_abbreviations(sent,
                                                                    defining_abbreviations,
                                                                    def_loc,
                                                                    list1,
                                                                    definiendum)
                elif def_info[0][4][j][0][1:] == "12":
                    map_double_definienda(sent, defining_abbreviations, def_abbrev_dict)
                elif sent[9] == "R":
                    o += 1
                    r_sent_location.append(n)
                    # the following key must be first in the dictionary because
                    # later we use the first entry in the dictionary for a special purpose
                    if definiendum == 'the':
                        concept = list(def_abbrev_dict.values())[0]
                        instance = definite_assignments.get(concept)
                        if instance == "":
                            definite_assignments[concept] = defining_abbreviations[0]
                            temp_dict1.update({sent[k]: defining_abbreviations[0]})
                        else:
                            temp_dict1.update({sent[k]: instance})
                        def_abbrev_dict = {**temp_dict1, **def_abbrev_dict}
                    elif o == 1:
                        temp_dict1.update({sent[k]: defining_abbreviations[0]})
                        def_abbrev_dict = {**temp_dict1, **def_abbrev_dict}
                    elif o == 2:
                        # for those definitions which have two r sentences
                        # the second r sentence needs another variable
                        # not that the following must be the third variable
                        def_abbrev_dict.update({sent[k]: variables[0]})
                        del variables[0]

    return def_abbrev_dict, r_sent_location, new_sentences


def map_double_definienda(sent, defining_abbreviations, def_abbrev_dict):
    for sentence in all_sent:
        if sent[9] == sentence[9] and sentence[5] == defining_abbreviations[0]:
            def_abbrev_dict.update({sent[14]: sentence[14]})
            break
    else:
        print ('you failed to find the second conjunct in the definiendum')
        g = 4 / 0


def map_defining_abbreviations(sent, defining_abbreviations, def_loc, list1, definiendum):
    special_relations = ["I", "J", "=", "H"]
    def_abbrev_dict = {}
    var_loc = 0
    determinative_positions = [3, 10, 16, 20, 24, 28, 32]
    if definiendum == 'part' + up:
        bb = 8

    concept_loc = 14 if def_loc == 10 else def_loc + 2
    if def_loc != 0 and def_loc not in determinative_positions:
        var_loc = get_non_variable_location(sent, definiendum)
    elif def_loc in determinative_positions:
        var_loc = get_non_variable_det_loc(sent, definiendum)
        # this is not the real def abbreviation as in the subject position
        # later it will be changed in the get new sent function
        if definiendum == 'the':
            if list1[concept_loc] not in definite_assignments:
                definite_assignments.update({list1[concept_loc]: ""})
        def_abbrev_dict.update({sent[var_loc]: list1[concept_loc]})
    elif sent[9] in special_relations or sent[14] == None:
        def_abbrev_dict.update({sent[5]: defining_abbreviations[0]})
    else:
        def_abbrev_dict.update({sent[5]: defining_abbreviations[0],
                                sent[14]: defining_abbreviations[1]})

    return def_abbrev_dict, var_loc


def get_non_variable_location(list1, definiendum):
    # this find out what the definining abbreviation is in pronoun definitions
    # for example if the definition is (bR he) <> (bRc) & etc
    # then the defining abbreviations is c
    for i in list1[56]:
        if list1[i] == definiendum:
            return i
    print ('you failed to find location of a the definining variable')
    g = 4 / 0


def get_non_variable_det_loc(list1, definiendum):
    determinative_positions = {3: 5, 10: 14}
    for i in list1[56]:
        if list1[i] == definiendum:
            break
    else:
        print ('you failed to find location of a the definining variable')
        g = 4 / 0
    j = determinative_positions.get(i)
    return j


def get_abbreviations_from_definition(def_info):
    # this function picks out that variables in the id sentences of the
    # definition

    constants = {}
    constant_map = {}
    temp_propositional_constants = {}
    list3 = []
    for i in range(len(def_info[0])):
        if os(def_info[0][i]) and mini_e in def_info[0][i]:
            list3.append(def_info[0][i])
        elif os(def_info[0][i]) and "=" in def_info[0][i]:
            str1 = def_info[0][i]
            g = str1.find("=")
            var = str1[1:g]
            wrd = str1[g + 1:-1]
            if isvariable(var):
                if not isvariable(wrd):
                    constants.update({var: wrd})

    for k, v in constants.items():
        if v in abbreviations[1]:
            new_var = abbreviations[1].get(v)
            constant_map.update({k: new_var})
        else:
            if k in variables:
                abbreviations[0].update({k: v})
                abbreviations[1].update({v: k})
                constant_map.update({k: k})
                variables.remove(k)
            else:
                new_var = variables[0]
                del variables[0]
                abbreviations[0].update({new_var: v})
                abbreviations[1].update({v: new_var})
                constant_map.update({k: new_var})

    constant_map.update({"i": "i"})
    temp_propositional_constants = get_propositional_constants(list3, temp_propositional_constants)

    return constant_map, temp_propositional_constants


def get_propositional_constants(list3, temp_propositional_constants):
    if list3 == []:
        return {}

    for i in range(len(list3)):
        prop_con = list3[i][1]
        str2 = list3[i].replace(" ", "")
        str2 = str2[3:-1]
        list8 = space_words(str2)
        list8 = categorize_words(list8)
        temp_propositional_constants.update({prop_con: list8})

    return temp_propositional_constants


def build_sent1(subj, tvalue, relat, obj, relat2="", obj2=""):
    list1 = [None] * 80
    tvalue = " ~ " if tvalue == "~" else " "
    list1[5] = subj
    list1[8] = tvalue
    list1[9] = relat
    list1[14] = obj
    if relat2 == "":
        sent = "(" + subj + tvalue + relat + " " + obj + ")"
        sent_abs = "(" + subj + " " + relat + " " + obj + ")"
    else:
        sent = "(" + subj + tvalue + relat + " " + obj + " " + relat2 + " " + obj2 + ")"
        sent_abs = "(" + subj + " " + relat + " " + obj + " " + relat2 + " " + obj2 + ")"

    abbrev_sent = name_sent(sent_abs)
    tvalue = "~" if tvalue == " ~ " else ""
    list1[0] = sent
    list1[72] = sent_abs
    list1[1] = abbrev_sent
    list1[2] = tvalue
    list1[42] = tvalue + abbrev_sent
    list1[46] = []

    return list1


def build_sent2(list1):
    # if you revise this list then then you must also revise it in
    # the eliminate_univ_quant_subclause, extract_words_from_subclause, as well as the function 'that', as well as new_categories
    # g=1 means that it is a sentence that identifies a propositional constant, in some cases
    # the proposition itself need not be named
    # also fix list in word sub and isatomic

    str1 = "("
    for i in allowable_slots():
        if list1[i] != None and list1[i] != "" and list1[i] != " ":
            if get_words_used:
                if list1[i] not in words_used:
                    words_used.append(list1[i])
            if str1 == "(":
                str1 += list1[i]
            else:
                str1 += " " + list1[i]

    str1 += ")"
    str1p = name_sent(str1)
    list1[0] = str1
    list1[2] = "~" if "~" in str1p else ""
    list1[1] = str1p.replace("~", "") if "~" in str1p else str1p
    list1[72] = str1.replace("~", "") if "~" in str1 else str1
    list1[72] = list1[72].replace("  ", " ")
    list1[42] = str1p

    return list1


def build_temp_sent(list1):
    # The only difference between this and build_sent2 is that it does not abbreviate the
    # sentence with a single letter

    str1 = "("

    for i in allowable_slots():
        if list1[i] != None and list1[i] != "" and list1[i] != " ":
            if get_words_used:
                if list1[i] not in words_used:
                    words_used.append(list1[i])
            if str1 == "(":
                str1 += list1[i]
            else:
                str1 += " " + list1[i]

    str1 += ")"
    list1[0] = str1
    return list1


def build_sent3(list1):
    str1 = "("
    num = allowable_slots()

    for i in num:
        temp_str = list1[i]
        if temp_str != None and temp_str != "":
            if get_words_used:
                if temp_str not in words_used:
                    words_used.append(temp_str)
            if str1 == "(":
                str1 += temp_str
            else:
                str1 += " " + temp_str

    str1 += ")"

    return str1


def build_sent_slots_known(list1):
    global build_sent_slots_time, build_sent_slots_counter
    aa = time.time()
    nums = list1[56]
    str1 = "("
    str2 = "("
    tvalue = ""
    for i in nums:
        if str1 == "(":
            str1 += list1[i]

        else:
            str1 += " " + list1[i]
        if str2 == "(":
            str2 += list1[i]
        elif list1[i] != "~":
            str2 += " " + list1[i]
        else:
            tvalue = "~"
    str1 += ")"
    str2 += ")"
    list1[0] = str1
    list1[2] = tvalue
    list1[72] = str2
    list1[1] = name_sent(str2)
    list1[42] = tvalue + list1[1]

    build_sent_slots_counter += 1
    build_sent_slots_time = (time.time() - aa)

    return list1


def build_uncategorized_sent(list1):
    for i in range(3, len(list1)):
        if list1[i] == None or list1[i] == "":
            break
        if i == 3:
            str1 = list1[i]
        else:
            str1 += " " + list1[i]

    str1 = "(" + str1 + ")"
    str1p = name_sent(str1)
    list1[0] = str1
    list1[2] = ""
    list1[1] = str1p
    list1[72] = str1
    list1[42] = str1p

    return list1


def build_sent_list(list1):
    # this list builder does not have the ~ separated from the sentence

    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i]
        else:
            str2 = str2 + ' & ' + list1[i]

    return str2


def build_sent_list2(list1, j):
    # this list builder does not have the ~ separated from the sentence

    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i][j]
        else:
            str2 = str2 + ' & ' + list1[i][j]

    return str2


def use_rarely_defined_word():
    not_oft_def = copy.deepcopy(dictionary[6])
    for word in abbreviations[1].keys():
        if word in not_oft_def:
            not_oft_def.remove(word)

    return not_oft_def


def add_necessary_conditions_for_concept():
    # if we're talking about concepts in our proof then we need to add their necesssary
    # conditiona to our proof

    global sn
    list2 = []
    con_sent_parts = [None] * 80
    for abbrev, word in abbreviations[0].items():
        if word == 'concept' + un or word == 'concept' + ua:
            str1 = abbrev

            for j in range(len(all_sent)):
                if all_sent[j][9] == "I" and all_sent[j][14] == str1:
                    ant_sent_parts = copy.deepcopy(all_sent[j])
                    str2 = all_sent[j][5]
                    con = abbreviations[0].get(str2)
                    pos = dictionary[0].get(con)
                    if pos == None:
                        bb = 7
                    if con != None:
                        if con == "dog":
                            bb = 8
                        if pos == 'a':
                            str4 = "J"
                        elif pos == 'n':
                            str4 = "I"
                        b = 0
                        for k in range(len(all_sent)):
                            if all_sent[k][9] == str4 and all_sent[k][14] == str2 and \
                                            str1 != str2 and str2 not in list2:
                                str6 = all_sent[k][5]
                                list2.append(str2)
                                b += 1
                        if b > 1:
                            print('you have not coded for multiple concepts')
                        olda = "(" + "b" + ' = ' + con + ")"
                        oldc = "(" + "c " + str4 + " b" + ")"
                        if str2 != "b":
                            rn1 = "(" + "b" + mini_c + str2 + ") & (" + "c" + mini_c + str6 + ")"
                        else:
                            rn1 = "(" + "c" + mini_c + str6 + ")"
                        nat_sent_b4_sub = olda + " " + conditional + " " + oldc
                        sn += 1
                        add_to_total_sent(sn, nat_sent_b4_sub, "", "", "NC concept " + con)
                        sn += 1
                        add_to_total_sent(sn, rn1, "", "", "RN")
                        con_sent_parts[5] = str6
                        con_sent_parts[9] = str4
                        con_sent_parts[14] = str2
                        anc1 = str(sn - 1) + "," + str(sn)
                        prepare_att_sent_1_sent(ant_sent_parts, "SUB", conditional,
                                                [build_sent2(con_sent_parts)], anc1)
                        break


def name_sent(str1, bool2=False, str4=""):
    no_space = copy.copy(str1)
    if str1.find('~') > -1:
        no_space = str1.replace("~", "")
        str1 = str1.replace("~", "")
        ng = '~'
    else:
        ng = ''

    if "  " in str1:
        str1 = str1.replace("  ", " ")

    no_space = remove_outer_paren(no_space)
    no_space = no_space.replace(" ", "")

    if bool2:
        if str4 == 'something':
            no_space = no_space.replace("something", "some thing")
        elif str4 == 'anything':
            no_space = no_space.replace("anything", "any thing")
        elif str4 == 'everything':
            no_space = no_space.replace("everything", "every thing")
        elif str4 == 'anything' + ua:
            no_space = no_space.replace("anything" + ua, "a" + ua + " thing")

    h = findinlist(no_space, prop_name, 1, 0)
    if h != None:
        return ng + h
    else:
        prop_name.append([prop_var[0], no_space, str1])
        str2 = prop_var[0]
        del prop_var[0]
        return ng + str2


def insert_space(str, integer):
    return str[0:integer] + ' ' + str[integer:]


def space_words(str1):
    # this function place a space between variables and relations
    # so as to make it easier to categorize words in a sentence

    str1 = str1.replace("(", "")
    str1 = str1.replace(")", "")
    str1 = str1.replace("~", " ~ ")
    str1 = str1.replace("=", " = ")
    str1 = str1.replace(neg, " " + neg + " ")
    str1 = str1.replace(mini_e, " " + mini_e + " ")
    str1 = str1.replace(ne, " " + ne + " ")
    i = -1
    while i + 1 < len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        nxt_str = str1[(i + 1):(i + 2)]
        if nxt_str.isupper() == True and temp_str.islower() == True:
            str1 = insert_space(str1, i + 1)
        elif nxt_str.islower() == True and temp_str.isupper() == True:
            str1 = insert_space(str1, i + 1)

    list1 = prepare_categorize_words(str1)
    list1[79] = 'is in definition'

    return list1


def isdefineable(list1):
    must_be_blank = [3, 4, 6, 7, 10, 11, 13, 16, 17, 18, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33,
                     35, 36, 49, 50, 51, 52, 55]
    must_be_variable = [5, 14, 18, 22]

    if list1[9] == "=": must_be_variable.remove(14)

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None:
            if not isvariable(list1[i]) and list1[i] != 'i':
                return False
    return True


def is_standard(list1):
    must_be_blank = [3, 4, 6, 7, 10, 11, 13, 16, 17, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33,
                     35, 36, 49, 50, 51, 52, 55]
    must_be_variable = [5, 14, 18, 22]

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None and list1[i] != "":
            if not isvariable(list1[i]) and list1[i] != 'i':
                return False
    return True


def remove_duplicates2d(list1, i, h):
    list2 = []
    j = -1
    while j < len(list1) - 1:
        j += 1
        for k in range(len(list1)):
            if k != j:
                if list1[k][i] == list1[j][i] and list1[k][h] == list1[j][h]:
                    del list1[j]
                    j -= 1
                    break
        else:
            list2.append(list1[j])
    return list2


def get_position_of_identities():
    for i in range(len(total_sent)):
        if total_sent[i][4] != "":
            return i
    g = 4 / 0


def build_list_of_abbreviations():
    # this turns the abbreviations into a conjunction
    global sn
    position_of_identities = get_position_of_identities()
    # this loop removes duplicates in a multidimensional list
    str1 = ""
    add_to_total_sent("", "_________________", "__________________")
    add_to_total_sent("", "")
    total_attach_sent = []
    temp_detach_sent = []
    for lst in attach_sent:
        for sent in lst[38]:
            total_attach_sent.append(sent)

    for key, value in abbreviations[0].items():
        if get_words_used:
            if value not in words_used:
                words_used.append(value)
        str2 = "(" + key + "=" + value + ")"
        str2p = name_sent(str2)
        if str1 == "":
            str1 = str2
            str1p = str2p
        else:
            str1 += " & " + str2
            str1p += " & " + str2p
        if str2p in total_attach_sent:
            list1 = [None] * 80
            list1[0] = str2
            list1[1] = str2p
            list1[2] = ""
            list1[5] = key
            list1[8] = ""
            list1[9] = "="
            list1[14] = value
            list1[72] = str2
            list1[42] = str2p
            temp_detach_sent.append(list1)

    if temp_detach_sent != []:
        for sent in temp_detach_sent:
            sn += 1
            sent[58] = sn
            detach_sent.append(sent)
            add_to_total_sent(sn, sent[0], sent[1], "", "&E", position_of_identities + 1)

    total_sent.insert(position_of_identities, [position_of_identities + 1, str1,
                                               str1p, "", 'ID', "", "", "", ""])


def remove_values_from_list(the_list, val):
    while val in the_list:
        the_list.remove(val)
    return the_list


def step_one(sent):

    divide_sent(sent)

    eliminate_redundant_words()

    replace_determinative_nouns()

    replace_synonyms()

    word_sub()

    eliminate_negative_determiners()


def divide_sent(list2):
    global sn

    for i in range(len(list2)):
        str2 = list2[i][1]
        str2 = str2.lower()
        str3 = name_sent(str2)
        str2 = str2.strip()
        list3 = str2.split()
        sent_parts = [None] * 80
        sent_parts[0] = str2
        sent_parts[42] = str3
        sent_parts[72] = str2
        sent_parts[1] = str3
        sent_parts[2] = ""
        sent_parts[58] = list2[i][0]
        for j in range(len(list3)):
            sent_parts[j + 3] = list3[j]
            if list3[j] not in words_used:
                words_used.append(list3[j])
        list4 = copy.deepcopy(sent_parts)
        detach_sent.append(sent_parts)
        add_to_total_sent(list2[i][0], str2, str3, "", "")
        all_sent.append(list4)


def eliminate_redundant_words():
    #modify this if we start dealing with sentences longer than 41 words
    global all_sent
    bool1 = False
    for i in range(len(all_sent)):
        ant_sent_parts = copy.deepcopy(all_sent[i])
        rule = ""
        for j in range(3, 55):
            if all_sent[i][j] == None:
                break
            pos = dictionary[0].get(all_sent[i][j])
            if pos != None and pos[0] == 't':
                if all_sent[i][j] not in words_used:
                    words_used.append(all_sent[i][j])
                bool1 = True
                if rule == '':
                    rule += all_sent[i][j]
                else:
                    rule += "," + all_sent[i][j]
                del all_sent[i][j]
                # this means that sentences must be shorter than 40 words
                all_sent[i].insert(40, None)
        if bool1:
            bool1 = False
            all_sent[i] = build_uncategorized_sent(all_sent[i])
            con_parts = copy.deepcopy(all_sent[i])
            prepare_att_sent_1_sent(ant_sent_parts, "RD " + rule, iff, [con_parts])

    for i, sent in enumerate(all_sent):
        sent[79] = 'do not rebuild sentence'
        sent2 = copy.deepcopy(sent)
        all_sent[i] = categorize_words(sent)
        for j in [0, 1, 2, 42, 72]:
            all_sent[i][j] = sent2[j]


def replace_determinative_nouns():
    global sn
    m = -1
    while m < len(all_sent)-1:
        m += 1
        replacement_made = False
        while all_sent[m][45][18] != []:
            ant_sent_parts = copy.deepcopy(all_sent[m])
            i = all_sent[m][45][18][0]
            rule = "DE " + all_sent[m][1]
            synonym = dictionary[2].get(all_sent[m][i])
            determinative = synonym[:synonym.find(" ")]
            definition = dictionary[1].get(all_sent[m][i])
            noun = synonym[synonym.find(" ")+1:]
            determinative.strip()
            noun.strip()
            replacement_made = True
            j = 10 if i == 14 else i - 2
            all_sent[m][j] = determinative
            all_sent[m][i] = noun
            if determinative == 'every' or determinative == 'no':
                b = 15
            elif determinative == 'a':
                b = 1
            else:
                g = 4 / 0

            all_sent[m][56].insert(all_sent[m][56].index(i),j)
            all_sent[m][45][b].append(j)
            all_sent[m][46].append(b)
            all_sent[m][45][18].remove(i)
        if replacement_made:
            all_sent[m] = build_sent_slots_known(all_sent[m])
            con_sent_parts = copy.deepcopy(all_sent[m])
            sn += 1
            add_to_total_sent(sn, definition, "", "", rule)
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_sent_parts], "")



def replace_synonyms():
    global sn
    definitions_added = []
    m = -1
    while m < len(all_sent)-1:
        m += 1
        replacement_made = False
        while all_sent[m][45][17] != []:
            ant_sent_parts = copy.deepcopy(all_sent[m])
            i = all_sent[m][45][17][0]
            synonym = dictionary[2].get(all_sent[m][i])
            assert synonym != None
            recategorize_word(synonym, m, i)
            definition = dictionary[1].get(all_sent[m][i])
            if definition not in definitions_added:
                definitions_added.append(definition)
                sn += 1
                add_to_total_sent(sn,definition,"","","DF " + all_sent[m][i])
            replacement_made = True
            all_sent[m][i] = synonym
            all_sent[m][45][17].remove(i)
        if replacement_made:
            all_sent[m] = build_sent_slots_known(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_parts], "")

def recategorize_word(synonym, m, i):
    # because we replace a word with a synonym we need to know it decision procedure
    # for elimination
    part_of_speech_syn = dictionary[0].get(synonym)
    pos = part_of_speech_syn[0]
    sub_pos = part_of_speech_syn[1] if len(part_of_speech_syn) > 1 else ""
    sub_sub_pos = part_of_speech_syn[1] if len(part_of_speech_syn) > 1 else ""
    b = get_used_slots(i, pos, sub_pos, sub_sub_pos)
    if b != 0:
        all_sent[m][45][b].append(i)
        all_sent[m][46].append(b)

def word_sub():
    global sn
    relational_positions = [9,15,19,23,27,31]
    m = -1
    n = len(all_sent)
    while m < n - 1:
        m += 1
        replacement_made = False
        while all_sent[m][45][0] != []:
            ant_sent_parts = copy.deepcopy(all_sent[m])
            k = all_sent[m][45][0][0]
            str2 = all_sent[m][k]
            if str2 not in words_used and not str2.isupper():
                words_used.append(str2)
            if str2 == "not":
                all_sent[m][k] = "~"
                replacement_made = True
            elif k == 69 or k == 70:
                replacement_made = True
                str2 = str2[:-2]
                replace_word_w_variable(m, k, str2)
            elif k in relational_positions:
                relat = dictionary[3].get(str2)
                assert relat != None
                replacement_made = True
                abbreviations[0].update({str2:relat})
                all_sent[m][k] = relat
            else:
                replacement_made = True
                replace_word_w_variable(m, k, str2)
            all_sent[m][45][0].remove(k)

        if replacement_made:
            all_sent[m] = build_sent2(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_parts])


def replace_word_w_variable(m, k, str2):
    not_normally_defined = dictionary[6]
    if isvariable(str2) == False:
        if str2 in not_normally_defined:
            not_normally_defined.remove(str2)
        str3 = abbreviations[1].get(str2)
        if str3 == None:
            pos = dictionary[0].get(str2)
            if len(pos) > 1 and pos[1] == "u":
                list1 = build_sent1(variables[0], "", "=", str2)
                list1[46] = []
                list1[56] = [5, 9, 14]
                all_sent.append(list1)
            if k == 69 or k == 70:
                all_sent[m][k] = variables[0] + "'s"
            else:
                all_sent[m][k] = variables[0]
            abbreviations[0].update({variables[0]: str2})
            abbreviations[1].update({str2: variables[0]})
            del variables[0]
        elif k == 69 or k == 70:
            all_sent[m][k] = str3 + "'s"
        else:
            all_sent[m][k] = str3


def eliminate_negative_determiners():
    # modify this if the category number of the universals change
    # modify this if we allow for two negative determiners in a sentence

    special_determinatives = ['a', 'every', 'many' + un, 'any' + un]
    for sent in all_sent:
        if sent[45] != None:
            for j in sent[45][16]:
                make_new_sentence = True
                if j == 8 and sent[10] in special_determinatives:
                    position = 10
                elif j == 49 and sent[18] in special_determinatives:
                    position = 18
                elif j == 50 and sent[22] in special_determinatives:
                    position = 22
                elif j == 51 and sent[26] in special_determinatives:
                    position = 26
                elif j == 52 and sent[30] in special_determinatives:
                    position = 30
                elif j == 47 and sent[3] in special_determinatives:
                    position = 3
                else:
                    make_new_sentence = False
                if make_new_sentence:
                    ant_sent_parts = copy.deepcopy(sent)
                    sent[j] = None
                    sent[56].remove(j)
                    if sent[position] == 'every':
                        sent[position] = 'many' + un
                        sent[45][1].append(position)
                        sent[45][15].remove(position)
                        sent[46].append(1)
                        sent[46].remove(15)
                        rule = "DE ~ every"
                    else:
                        rule = "DE ~ " + sent[position]
                        sent[position] = 'no'
                        sent[45][15].append(position)
                        sent[45][1].remove(position)
                        sent[46].append(15)
                        sent[46].remove(1)

                    sent = build_sent_slots_known(sent)
                    con_parts = copy.deepcopy(sent)
                    prepare_att_sent_1_sent(ant_sent_parts, rule, iff, [con_parts])


def new_categories(list5, kind=False):
    list1 = [None] * 80
    if not kind:
        print ('new categories function uses list 46')
        g = 4 / 0
        num = list5[46]
    else:
        num = [47, 3, 69, 4, 55, 5, 66, 67, 35, 48, 59, 6, 8, 9, 7, 48, 12, 10, 70,
               13, 14, 36, 60, 63, 49, 15, 16, 17, 18,
               61, 64, 50, 19, 20, 21, 22, 62, 65, 51, 23, 24, 25, 26, 52, 27, 28,
               29, 30, 31, 32, 33, 34]
    i = 2
    for j in num:
        if list5[j] != None and list5[j] != "":
            i += 1
            list1[i] = list5[j]
    list3 = categorize_words(list1)
    return list3


def is_adj_definite(list1, i):
    # modify this if more indefinite determinatives are added
    indefinite_determinatives = ['a']
    determinative_position = 10 if i == 10 else i - 1
    if list1[determinative_position] in indefinite_determinatives:
        return False
    else:
        return True


def lies_wi_mainclause(word_pos, location):
    begin = location[0]
    end = location[1]
    if word_pos > begin and word_pos < end:
        return False
    else:
        return True


def relative_pronoun_lies_in_scope_univ(list1, word_pos, univ_pos, current_universal):
    bool1 = False
    absolute_univ_pos = allowable_slots().index(univ_pos)
    univ_in_main_clause = lies_wi_mainclause(absolute_univ_pos, list1[57])
    rel_pro_in_main_clause = lies_wi_mainclause(word_pos, list1[57])

    if univ_in_main_clause:
        if absolute_univ_pos < word_pos:
            bool1 = True
    elif not univ_in_main_clause and not rel_pro_in_main_clause:
        if absolute_univ_pos < word_pos:
            bool1 = True

    return bool1


def relation_lies_wi_scope_univ(list1, univ_pos, word_pos):
    absolute_univ_pos = allowable_slots().index(univ_pos)
    univ_in_main_clause = lies_wi_mainclause(absolute_univ_pos, list1[57])
    relat_in_main_clause = lies_wi_mainclause(word_pos, list1[57])
    bool1 = False

    if not univ_in_main_clause and not relat_in_main_clause:
        bool1 = True
    elif univ_in_main_clause:
        bool1 = True

    return bool1


def adjective_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos):
    # modify this is we allow sentences with more than 1 determinative

    absolute_univ_pos = allowable_slots().index(univ_pos)
    univ_in_main_clause = lies_wi_mainclause(absolute_univ_pos, list1[57])
    adj_in_main_clause = lies_wi_mainclause(word_pos, list1[57])
    is_definite = is_adj_definite(list1, i)

    bool1 = False
    if not is_definite:
        if current_universal == 'every':
            if i == 13 and univ_pos == 10:
                bool1 = True
            elif i == univ_pos + 1:
                bool1 = True
        elif current_universal == 'no':
            if univ_in_main_clause:
                if absolute_univ_pos < word_pos:
                    bool1 = True
            else:
                if not adj_in_main_clause and absolute_univ_pos < word_pos:
                    bool1 = True

    return bool1


def determ_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos):
    # modify this if we increase the number of indefinite determinatives
    indefinite_determinatives = ['a']
    bool1 = False
    univ_pos = allowable_slots().index(univ_pos)
    univ_in_sub_clause = lies_wi_mainclause(univ_pos, list1[57])
    determ_in_sub_clause = lies_wi_mainclause(word_pos, list1[57])
    if list1[i] in indefinite_determinatives and current_universal == 'no':
        if univ_in_sub_clause and determ_in_sub_clause:
            if univ_pos < word_pos:
                bool1 = True
        elif not univ_in_sub_clause and univ_pos > word_pos:
            bool1 = True

    return bool1


def lies_wi_scope_of_univ_quant(list1, i):
    # modify this if you allow for more than one universal quantifier in a sentence
    # or you allow for more than one subclause
    # or you increase the number of determinatives

    bool1 = False
    if 15 in list1[46]:

        adjective_positions = [4, 13, 17, 21, 25, 29, 33]
        determinative_positions = [3, 10, 16, 20, 24, 28, 32]
        relation_positions = [15, 19, 23, 27, 31]
        relative_pronoun_positions = [59, 60, 61, 62]
        univ_pos = list1[45][15][0]
        current_universal = list1[univ_pos]
        word_pos = allowable_slots().index(i)

        if i in adjective_positions:
            bool1 = adjective_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos)
        elif i in determinative_positions:
            bool1 = determ_lies_wi_scope_of_univ(list1, i, current_universal, word_pos, univ_pos)
        elif i in relation_positions:
            bool1 = relation_lies_wi_scope_univ(list1, univ_pos, word_pos)
        elif i in relative_pronoun_positions:
            bool1 = relative_pronoun_lies_in_scope_univ(list1, word_pos, univ_pos, current_universal)

    return bool1


def get_subclause_position(list1):
    begin = 0
    end = 0
    for i in [59, 60, 61, 62]:
        if list1[i] != None:
            begin = allowable_slots().index(i)

            break
    if begin != 0:
        if list1[39] == None:
            end = 55
        else:
            end = allowable_slots().index(list1[39])

    return [begin, end]


def prepare_att_sent_4_sent(ant_sent_parts, consequent, connective, rule):
    global sn
    con_parts1 = consequent[0]
    con_parts2 = consequent[1]
    con_parts3 = consequent[2]
    con_parts4 = consequent[3]

    sn += 1
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts1[68] = "121"
    con_parts1[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    con_parts3[68] = "123"
    con_parts3[44] = chr(952)
    con_parts4[68] = "124"
    con_parts4[44] = chr(953)
    sent_type = "e" if connective == iff else "c"

    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts1[0] + \
                      " & " + con_parts2[0] + " & " + con_parts3[0] + " & " + con_parts4[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts1[42] + \
                    " & " + con_parts2[42] + " & " + con_parts3[42] + " & " + con_parts4[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + \
                " & " + chr(952) + " & " + chr(953) + ")"
    if connective == 'e':
        ant_sent_parts[53] = 'b'
        con_parts1[53] = 'cf'
        con_parts2[53] = 'cf'
        con_parts3[53] = 'cf'
        con_parts4[54] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts1[53] = 'cq'
        con_parts2[53] = 'cq'
        con_parts3[53] = 'cq'
        con_parts4[54] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts1[1], con_parts1[2]], [con_parts2[1], con_parts2[2]],
                [con_parts3[1], con_parts3[2]], [con_parts4[1], con_parts4[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts1[42] + " & " + con_parts2[42] + \
                " & " + con_parts3[42] + " & " + con_parts4[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts1, con_parts2, con_parts3, con_parts4]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts1[1], con_parts2[1], con_parts3[1], con_parts4[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts1[0] + " & " + con_parts2[0] + " & " + con_parts3[0] + \
                 " & " + con_parts4[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts1[72], con_parts1[2]], [con_parts2[72], con_parts2[2]],
                 [con_parts3[72], con_parts3[2]], [con_parts4[72], con_parts4[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_3_sent(ant_sent_parts, connective, consequent, rule):
    # this populates the attach_sent list provided a sentence is equivalent
    # to two conjuncts

    global sn
    sn += 1
    con_parts = consequent[0]
    con_parts2 = consequent[1]
    con_parts3 = consequent[2]
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    con_parts3[68] = "123"
    con_parts3[44] = chr(952)
    sent_type = "e" if connective == iff else "c"
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + \
                      " & " + con_parts2[0] + " & " + con_parts3[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + \
                    " & " + con_parts2[42] + " & " + con_parts3[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + \
                " & " + chr(952) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
        con_parts3[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'
        con_parts3[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]],
                [con_parts3[1], con_parts3[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + " & " + con_parts3[42] + ")", ""]
    list4[34] = [copy.deepcopy(ant_sent_parts)]
    list4[35] = [copy.deepcopy(con_parts), copy.deepcopy(con_parts2), copy.deepcopy(con_parts3)]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1], con_parts3[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + " & " + con_parts3[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]],
                 [con_parts3[72], con_parts3[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_2_sent(ant_sent_parts, connective, consequent, rule):
    # this populates the attach_sent list provided a sentence is equivalent
    # to two conjuncts

    global sn
    sn += 1
    if rule == "RDC": connective = conditional
    con_parts = consequent[0]
    con_parts2 = consequent[1]
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    sent_type = "e" if connective == iff else "c"
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + " & " + con_parts2[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + " & " + con_parts2[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts, con_parts2]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_1_sent(ant_sent_parts, rule, connective, consequent, anc1=""):
    # this populates the attach_sent list provided a sentence is equivalent to one other sentence

    global sn
    sn += 1
    list4 = [""] * 50
    con_parts = consequent[0]
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "12"
    con_parts[44] = chr(950)
    sent_type = "e" if connective == iff else "c"
    new_equivalence = ant_sent_parts[0] + " " + connective + " " + con_parts[0]
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " " + con_parts[42]
    new_greek = chr(949) + " " + connective + " " + chr(950)
    if connective == iff:
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'f'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'q'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = [con_parts[1], con_parts[2]]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = [con_parts[72], con_parts[2]]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule, anc1)
    attach_sent.append(list4)


def allowable_slots2():
    num2 = [11, 47, 3, 69, 4, 55, 5, 66, 67, 35,
            48, 59, 6, 8,
            9, 7, 48, 12, 10, 70,
            13, 14, 36, 60, 63, 40, 49, 15,
            16, 17, 18,
            61, 64, 41, 50, 19, 20, 21, 22, 62, 65, 43, 51,
            23, 24, 25, 26, 52, 27, 28,
            29, 30, 31, 32, 33, 34]

    return num2


def allowable_slots():
    num2 = [11, 47, 3, 69, 4, 55, 5, 66, 67, 35, 48, 59, 6, 8, 9, 7, 48, 12, 10, 70,
            13, 14, 36, 60, 63, 49, 15,
            16, 17, 18,
            61, 64, 50, 19, 20, 21, 22, 62, 65, 51,
            23, 24, 25, 26, 52, 27, 28,
            29, 30, 31, 32, 33, 34]

    return num2


def repl_sign(str3, match_dv, match_type):
    s = findposinmd(str3, match_dv, 1)
    s = match_type[s]
    if s == 0:
        return mini_c
    else:
        return idd


def is_set_of_bic_wo_id(def_info):
    # if a definition is composed of a set of conditionals or biconditionals
    # and there is no identity statement within the definition then
    # we do not need to delete anything from the def_info list

    if def_info[4][0][1] == "&":
        for lst in def_info[4]:
            if len(lst[0]) == 2 and lst[1] == "":
                return False
            elif len(lst[0]) > 2:
                return True

    return False


def eliminate_conjuncts_from_definition(def_info):
    # any sentences which are not within either an iff or conditiional
    # are deleted

    # 1. if a definition has some detached conjuncts then these are eliminated
    # and a new greek definition is placed in def_info[5], the def_info
    # list will have one member
    # 2. if a definition is composed of sets of complex sentences then
    # the greek definition remains in def_info[5]
    # the other complex sentences are placed in sets of equivalences
    # and it is written on [45] that they are not yet detached and
    # hence cannot be used in the modus ponens function
    # 3. when we add to the total sent list, we do so by looping through the
    # sets of equivalences list and only adding those that say in [45]
    # 'detached'



    if is_set_of_bic_wo_id(def_info):
        for lst in def_info[4]:
            lst[0] = lst[0][1:]
        def_info = make_sets_of_equivalences(def_info)
        def_info = build_conjunction_of_biconditionals(def_info)
    elif def_info[4][0][1] == "&":
        i = 0
        while len(def_info[4][i][0]) < 3:
            if def_info[4][i][1] == "" and len(def_info[4][i][0]) == 2:
                del def_info[0][i]
                del def_info[1][i]
                del def_info[3][i]
                del def_info[4][i]
                del def_info[6][i]
            else:
                i += 1

        def_info = prepare_def_info_list(def_info)
    else:
        def_info = [def_info]

    return def_info


def prepare_def_info_list(def_info):
    connectives = [iff, conditional, idisj, xorr]
    # this removes the first number from the old numbers
    for lst in def_info[4]:
        lst[0] = lst[0][1:]

    # if the definition was composed of a set of conjunctions then the
    # the greek sentence needs to be amended, otherwise we need
    # the "do not add to total sent list
    if def_info[4][1][1] in connectives and def_info[4][2][1] in connectives:
        def_info = make_sets_of_equivalences(def_info)
        def_info = build_conjunction_of_biconditionals(def_info)
    else:
        del def_info[0][0]
        del def_info[1][0]
        del def_info[3][0]
        del def_info[4][0]
        del def_info[6][0]
        def_info[6][0] = remove_outer_paren(def_info[6][0])
        def_info[5] = def_info[6][0]
        def_info = [def_info]

    return def_info


def make_sets_of_equivalences(def_info):
    # if a definition is composed of a set of equivalences
    # then this function puts each complex sentence into a list

    num = [0, 1, 3, 4, 6]
    list2 = []
    for i in range(len(def_info[4])):
        if len(def_info[4][i][0]) == 1:
            list1 = [""] * 7
            for j in num:
                list1[j] = copy.deepcopy([def_info[j][i]])
            list2.append(list1)
        elif len(def_info[4][i][0]) > 1:
            list1 = [""] * 7
            for j in num:
                list1[j] = copy.deepcopy(def_info[j][i])
            for m in range(len(list2)):
                if list1[4][0].startswith(list2[m][4][0][0]):
                    for k in num:
                        list3 = list2[m][k]
                        list3.append(list1[k])
                        list2[m][k] = list3
                    list2[m][5] = list2[m][6][0]

    def_info = [def_info]
    for lst in list2:
        def_info.append(lst)

    return def_info


def build_conjunction_of_biconditionals(def_info):
    str1 = def_info[1][6][0]
    def_info[1][2] = 'eliminate as conjunct'
    for i in range(2, len(def_info)):
        str1 += " & " + def_info[i][5]
        def_info[i][2] = 'eliminate as conjunct'

    def_info[0][5] = str1
    def_info[0][6][0] = str1
    return def_info


def update_essential_properties_dictionary(first_sent, definiendum, defin_sent):
    if first_sent[9] == "I" and isinmdlist(first_sent[14], abbreviations, 0) > -1:
        essential_properties = dictionary[8]
        subject = first_sent[5]
        num = [5, 14]
        defin_sent2 = copy.deepcopy(defin_sent)
        for i in range(1, len(defin_sent2)):
            for j in num:
                if defin_sent2[i][j] == subject:
                    defin_sent2[i][j] = alpha
                elif not isinmdlist(defin_sent2[i][j], abbreviations, 0):
                    defin_sent2[i][j] = beta

        ess_prop = []
        for i in range(1, len(defin_sent2)):
            sent = build_sent3(defin_sent2[i])
            sent = sent.replace(" ", "")
            ess_prop.append(sent)

        essential_properties[definiendum] = ess_prop


def eliminate_negated_many(defin_sent, definiendum, k):
    if definiendum == "many" + un:
        for i in range(len(defin_sent)):
            if not defin_sent[i][40] and defin_sent[i][8] == "~":
                defin_sent[i] = eliminate_not_a(defin_sent[i], k)
                return defin_sent

    return defin_sent


def remove_i_from_45(sent):
    for i in [5, 14, 18, 22, 26, 30]:
        if sent[i] == 'i':
            sent[45][1].remove(i)
    return sent


def determine_what_is_conjunct(def_info):
    # modify this if definitions have v or xor as a main connective in their consequent
    ant_conj = 0
    con_conj = 0
    for i, num in enumerate(def_info[4]):
        if len(num[0]) == 2 and num[0][1] == '1':
            if num[1] == "&":
                ant_conj = 3
            else:
                ant_conj = 2
        elif len(num[0]) == 2 and num[0][1] == '2':
            if num[1] == "&":
                con_conj = 3
            else:
                con_conj = 2
        if ant_conj != 0 and con_conj != 0:
            return ant_conj, con_conj


def prepare_attach_sent(def_info, defin_sent, r_sent_loc):
    # this populates the attach sent list

    global sn
    if r_sent_loc != []:
        sn += 1
    elif def_info[2] == 'eliminate as conjunct':
        pass
    else:
        sn += 3
    list1 = [""] * 50
    list1[2] = sn
    greek_sent = def_info[5]
    list2 = translate_complex_sent(greek_sent, defin_sent)
    list1[4] = list2[1]
    list1[37] = list2[0]
    list1[47] = greek_sent
    ant_parts = []
    con_parts = []
    ant_variables = []
    con_variables = []
    ant_conjunction = ""
    con_conjunction = ""
    spec_conn = [iff, conditional, idisj, xorr]
    total_sent_in_attach_sent = []
    embed_att_sent = []
    isdisjunction = False
    isconjunction = False
    if def_info[4][0][1] == iff:
        list1[3] = "e"
    elif def_info[4][0][1] == conditional:
        list1[3] = "c"
    elif def_info[4][0][1] == idisj:
        list1[3] = "d"
        isdisjunction = True
    elif def_info[4][0][1] == xorr:
        list1[3] = "x"
        isdisjunction = True
    else:
        isconjunction = True

    if isdisjunction:
        list1[36] = copy.deepcopy(def_info)
        list1[44] = copy.deepcopy(defin_sent)
        list1[45] = "do not append to attach_sent list"
    elif isconjunction:
        list1[45] = "do not append to attach_sent list"
    else:
        ant_conj, con_conj = determine_what_is_conjunct(def_info)
        list1[45] = "append to attach_sent list"
        for i in range(1, len(def_info[0])):
            t_value = def_info[1][i]
            if def_info[4][i][1] == "":
                d = findposinmd(def_info[6][i], defin_sent, 44)  # d = index of def sent
                defin_sent[d] = ancestor_numbers(defin_sent[d], def_info[4][i][0], def_info)
                if defin_sent[d][53][-1] == "a" or defin_sent[d][53][-1] == "b":
                    ant_parts.append(copy.deepcopy(defin_sent[d]))
                    ant_parts[-1][68] = def_info[4][i][0]
                else:
                    con_parts.append(copy.deepcopy(defin_sent[d]))
                    con_parts[-1][68] = def_info[4][i][0]
                total_sent_in_attach_sent.append(defin_sent[d][1])
                t_value = defin_sent[d][2]
                defin_sent[d][68] = def_info[4][i][0]

            if def_info[4][i][0][1] == '1' and len(def_info[4][i][0]) == ant_conj:
                ant_variables.append([def_info[6][i], t_value])
            elif def_info[4][i][0][1] == '2' and len(def_info[4][i][0]) == con_conj:
                con_variables.append([def_info[6][i], t_value])

            if def_info[4][i][1] != "" and len(def_info[4][i][0]) == 2:
                if def_info[4][i][0][1] == '1' and ant_conjunction == "":
                    ant_conjunction = def_info[6][i]

                elif def_info[4][i][0][1] == '2' and con_conjunction == "":
                    con_conjunction = def_info[6][i]

            if def_info[4][i][1] in spec_conn:
                embed_att_sent.append(i)

        list1 = prepare_attach_sent2(ant_conjunction, con_conjunction, ant_variables, \
                                     con_variables, def_info, defin_sent,
                                     embed_att_sent, list1)

        list1[34] = ant_parts
        list1[35] = con_parts
        list1[38] = total_sent_in_attach_sent

    return list1


def prepare_attach_sent2(ant_conjunction, con_conjunction, ant_variables,
                         con_variables, def_info, defin_sent,
                         embed_att_sent, list1):
    global sn
    embed_info = []
    list2 = translate_list_of_sentences(ant_variables, defin_sent)
    list1[0] = list2[1]
    list1[42] = list2[0]

    if ant_conjunction != "":
        list2 = translate_complex_sent(ant_conjunction, defin_sent)
        list1[40] = [list2[0], ""]
        list1[7] = [list2[1], ""]
    else:
        list1[40] = list2[0][0]
        list1[7] = list2[1][0]

    list2 = translate_list_of_sentences(con_variables, defin_sent)
    list1[1] = list2[1]
    list1[43] = list2[0]

    if con_conjunction != "":
        list2 = translate_complex_sent(con_conjunction, defin_sent)
        list1[41] = [list2[0], ""]
        list1[8] = [list2[1], ""]
    else:
        list1[41] = list2[0][0]
        list1[8] = list2[1][0]

    if embed_att_sent != []:
        for num in embed_att_sent:
            list4 = prepare_embed_att_sent(def_info, defin_sent, num)
            embed_info.append(list4)
            sn -= 1

        list1[39] = embed_info

    return list1


def prepare_embed_att_sent(def_info, defin_sent, num):
    # this takes those attached sentences within attached sentences and
    # prepares them to be manipulated

    def_info2 = copy.deepcopy(def_info)
    sent_num = def_info[4][num][0]
    def_info2 = delete_irrel_sent(def_info2, sent_num)
    def_info2[5] = def_info2[6][0]
    pos_of_new_num = len(sent_num) - 1
    main_conn_loc = def_info2[3][0].find(def_info2[4][0][1])
    def_info2[4][0][2] = main_conn_loc
    def_info2 = renumber_embed_sent(def_info2, pos_of_new_num)
    list1 = prepare_attach_sent(def_info2, defin_sent, "EMBED")
    get_general_variables(list1)

    return list1


def renumber_embed_sent(def_info2, pos_of_new_num):
    for position in def_info2[4]:
        position[0] = position[0][pos_of_new_num:]

    return def_info2


def delete_irrel_sent(def_info2, sent_num):
    # this deletes sentences from the def_info list
    # which are not members of the embedded sentence
    # under investigation

    i = 0
    while i < len(def_info2[4]):

        if not def_info2[4][i][0].startswith(sent_num):
            del def_info2[0][i]
            del def_info2[1][i]
            del def_info2[3][i]
            del def_info2[4][i]
            del def_info2[6][i]
        else:
            i += 1

    return def_info2


def ancestor_numbers(list2, k, def_info):
    # this determines the number and connective of the ancestors of a
    # sent in the conditional
    list2[53] = None
    self_num = k[-1]
    if len(k) == 4:
        ggparen_num = k[0]
        gparen_num = k[:2]
        paren_num = k[:3]
        ggparen_conn = findinlist(ggparen_num, def_info[4], 0, 1)
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        ggparen_conn = convert_con_to_letter(ggparen_conn, gparen_num[-1])
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn + ggparen_conn

    elif len(k) == 3:
        gparen_num = k[0]
        paren_num = k[:2]
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn

    elif len(k) == 2:
        paren_num = k[0]
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn

    elif len(k) == 5:
        gggparen_num = k[0]
        ggparen_num = k[:2]
        gparen_num = k[:3]
        paren_num = k[:4]
        gggparen_conn = findinlist(gggparen_num, def_info[4], 0, 1)
        ggparen_conn = findinlist(ggparen_num, def_info[4], 0, 1)
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        gggparen_conn = convert_con_to_letter(gggparen_conn, ggparen_num[-1])
        ggparen_conn = convert_con_to_letter(ggparen_conn, gparen_num[-1])
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn + ggparen_conn + gggparen_conn

    elif len(k) == 6:
        print("you have not coded for attached sentences with 5 generations yet")
        sys.exit()

    if list2[53] == None:
        print("the number ancestor function is messed up")
        sys.exit()
    return list2


def convert_con_to_letter(str1, str2):
    # this converts a connective to a letter
    if str1 == iff and str2 == '1':
        return 'b'
    elif str1 == iff and str2 == '2':
        return 'f'
    elif str1 == conditional and str2 == '1':
        return 'a'
    elif str1 == conditional and str2 == '2':
        return 'q'
    elif str1 == xorr and str2 == '1':
        return 'x'
    elif str1 == xorr and str2 == '2':
        return 'y'
    elif str1 == idisj and str2 == '1':
        return 'd'
    elif str1 == idisj and str2 == '2':
        return 'g'
    elif str1 == idisj and str2 == '3':
        return 'd3'
    elif str1 == idisj and str2 == '4':
        return 'd4'
    elif str1 == idisj and str2 == '5':
        return 'd5'
    elif str1 == idisj and str2 == '6':
        return 'd6'


    elif str1 == "&":
        return 'c'
    else:
        print('the convert con to letter function is messed up')
        g = 4 / 0


def get_parent_connective(def_info, i):
    sent_number = def_info[4][i][0]
    for i in range(i - 1, -1, -1):
        if sent_number.startswith(def_info[4][i][0]):
            return def_info[4][i][1]
    else:
        print ('you failed to find the parent connective')
        g = 4 / 0


def translate_list_of_sentences(to_be_converted, defin_sent):
    # this converts the 0th or 1st member of the attach_sent_list into
    # sentence variables

    abbrev_sent = copy.deepcopy(to_be_converted)
    for i in range(len(to_be_converted)):
        if os(to_be_converted[i][0]):
            d = findposinmd_alert_error(to_be_converted[i][0], defin_sent, 44)
            to_be_converted[i][0] = defin_sent[d][72]
            abbrev_sent[i][0] = defin_sent[d][1]
        elif not os(to_be_converted[i][0]):
            list1 = translate_complex_sent(to_be_converted[i][0], defin_sent)
            to_be_converted[i][0] = list1[0]
            abbrev_sent[i][0] = list1[1]

    return [to_be_converted, abbrev_sent]


def translate_complex_sent(to_be_translated, defin_sent):
    # this converts a conjunction in a definition into the definition
    # with the new variables

    to_translate_abbrev = to_be_translated

    for sentence in defin_sent:
        if sentence[44] in to_be_translated:
            to_be_translated = to_be_translated.replace(sentence[44], sentence[0])
            to_translate_abbrev = to_translate_abbrev.replace(sentence[44], sentence[42])

    return [to_be_translated, to_translate_abbrev]


def variable_change(rule, match_dv):
    global sn
    for i in range(len(match_dv)):
        if match_dv[i][0] != match_dv[i][1]:
            return rule

    rule = rule[0] + "E" + rule[2:]
    sn += 1

    return rule


def eliminate_not_a(list1, k):
    num = [10, 16, 20, 24]

    for i in num:
        if i > k and (list1[i] == "a" or list1[i] == "a" + ua):
            ant_sent_parts = copy.deepcopy(list1)
            list2 = copy.deepcopy(list1)
            if list1[i] == "a":
                rule = "DE not a"
            else:
                rule = "DE not a" + ua
            list2[i] = 'every'
            con_parts = copy.deepcopy(build_sent2(list2))
            prepare_att_sent_1_sent(ant_sent_parts, rule, "", iff, [con_parts])
            all_sent.append(list1)
            return list2

    return list1


def prepare_categorize_words(str2):
    # when we prepare to categorize a word, the first three slots
    # are reserved for the sentence, the sentence letter and the tvalue

    str2 = str2.strip()
    list1 = str2.split(' ')
    list2 = [None] * 80
    j = 2
    for i in range(len(list1)):
        j += 1
        list2[j] = list1[i]

    return list2


def categorize_words(list1):
    sentence_slots = [None] * 80
    relation_type = 0
    # up to 19
    slots_used = [[], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []]
    places_used = []
    noun_list = ['n', 'p']
    the_is_of_group = ["I", "is" + ug, "are" + ug, "be" + ug, "was" + ug, "were" + ug, "am" + ug]
    the_is_of_adjective = ["J", "is" + ua, "be" + ua, "are" + ua, "was" + ua, "am" + ua, "were" + ua]
    spec_rel = the_is_of_adjective + the_is_of_group
    categories_used = []

    i = 2
    while list1[i + 1] != None:
        i += 1
        k = 0
        word = list1[i]

        if word == 'not':
            bb = 8

        i, word, has_comma = determine_if_compound_word(i, list1, word)

        part_of_speech, sub_part, sub_sub_part = get_part_of_speech(word, list1[79])

        if word != ' ' and word != "":
            insert_special_location = False
            if get_words_used:
                if word not in words_used:
                    words_used.append(word)
            if part_of_speech == "d":  # q are possessive pronouns
                if relation_type == 0:
                    k = 3
                elif relation_type == 1:
                    k = 10
                elif relation_type == 2:
                    k = 16
                elif relation_type == 3:
                    k = 20
                elif relation_type == 4:
                    k = 24
                elif relation_type == 5:
                    k = 28
                elif relation_type == 6:
                    k = 32

            elif part_of_speech == 'o' or part_of_speech == 's':
                if relation_type == 0 and sentence_slots[5] == None:
                    k = 69
                elif relation_type == 1 and sentence_slots[14] == None:
                    k = 70

            elif part_of_speech == 'a':
                if relation_type == 0:
                    k = 4
                elif relation_type == 1 and sentence_slots[9] in the_is_of_adjective:
                    k = 14
                elif relation_type == 1:
                    k = 13
                elif relation_type == 2 and sentence_slots[15] in the_is_of_adjective:
                    k = 18
                    # no decision needed here because it is defined as a concept
                elif relation_type == 2:
                    k = 17
                elif relation_type == 3 and sentence_slots[19] in the_is_of_adjective:
                    k = 22
                elif relation_type == 3:
                    k = 21
                elif relation_type == 4 and sentence_slots[23] in the_is_of_adjective:
                    k = 26
                elif relation_type == 4:
                    k = 25
                elif relation_type == 5 and sentence_slots[27] in the_is_of_adjective:
                    k = 30
                elif relation_type == 5:
                    k = 29
                elif relation_type == 6 and sentence_slots[31] in the_is_of_adjective:
                    k = 34
                elif relation_type == 6:
                    k = 33

            elif part_of_speech == 'm':
                if sentence_slots[3] == None and sentence_slots[5] == None:
                    k = 47
                elif relation_type == 0:
                    k = 8
                elif (relation_type == 1 and sentence_slots[14] == None and sentence_slots[60] == None):
                    k = 8
                    # because 'not' in this location comes after the relation we must
                    # insert to before the relation
                    insert_special_location = True
                elif relation_type == 1:
                    k = 49
                elif relation_type == 2 and sentence_slots[15] in spec_rel:
                    k = 49
                    insert_special_location = True
                elif relation_type == 2:
                    k = 50
                elif relation_type == 3 and sentence_slots[18] in spec_rel:
                    k = 50
                    insert_special_location = True
                elif relation_type == 3:
                    k = 51
                elif relation_type == 4 and sentence_slots[24] in spec_rel:
                    k = 51
                    insert_special_location = True
                elif relation_type == 4 or sentence_slots[27] in spec_rel:
                    k = 52


            elif sub_part == 't':

                if relation_type == 0 and sentence_slots[5] == None:
                    k = 5
                elif relation_type == 1 and sentence_slots[60] != None:
                    k = 63
                elif relation_type == 2 and sentence_slots[61] != None:
                    k = 64

            elif part_of_speech in noun_list:

                if relation_type == 0:
                    if sentence_slots[66] != None:
                        k = 67
                    elif sentence_slots[5] != None:
                        k = 35
                    elif sentence_slots[5] == None:
                        k = 5
                elif relation_type == 1:
                    if sentence_slots[14] == None:
                        k = 14
                    elif sentence_slots[60] != None:
                        k = 63
                    else:
                        k = 36
                elif relation_type == 2 and sentence_slots[18] == None:
                    k = 18
                elif relation_type == 2 and sentence_slots[61] != None:
                    # todo these are wrong and need to be changed
                    k = 64
                elif relation_type == 3 and sentence_slots[22] == None:
                    k = 22
                elif relation_type == 3 and sentence_slots[62] != None:
                    k = 65
                elif relation_type == 4:
                    k = 26
                elif relation_type == 5:
                    k = 30
                elif relation_type == 6:
                    k = 34

            elif part_of_speech == 'c':
                if relation_type == 0 and sentence_slots[5] != None:
                    k = 66  # uuu

            elif part_of_speech == 'u':
                if relation_type == 0 and sentence_slots[5] != None:
                    k = 59
                elif relation_type == 1 and sentence_slots[14] != None:
                    k = 60
                elif relation_type == 2 and sentence_slots[18] != None:
                    k = 61
                elif relation_type == 3 and sentence_slots[22] != None:
                    k = 62

            elif part_of_speech == 'y':

                if sentence_slots[7] == None:
                    k = 7
                elif sentence_slots[40] == None:
                    k = 40
                elif sentence_slots[41] == None:
                    k = 41
                elif sentence_slots[43] == None:
                    k = 43

            elif part_of_speech == 'r':

                if relation_type == 0:
                    k = 9
                    relation_type = 1
                elif relation_type == 1:
                    k = 15
                    relation_type = 2
                elif relation_type == 2:
                    relation_type = 3
                    k = 19
                elif relation_type == 3:
                    relation_type = 4
                    k = 23
                elif relation_type == 4:
                    relation_type = 5
                    k = 27
                elif relation_type == 5:
                    k = 31
                    relation_type = 6

            if has_comma: sentence_slots[39] = k
            assert k != 0
            sentence_slots[k] = word
            if insert_special_location:
                places_used.insert(-1,k)
            else:
                places_used.append(k)
            b = get_used_slots(k, part_of_speech, sub_part, sub_sub_part)

            slots_word_sub = [4, 5, 12, 13, 14, 17, 18, 22, 26, 30, 49, 50, 51, 52, 8, 47, 49, 50, 51, 52,
                              34, 35, 36, 63, 64, 65, 67, 69, 70, 9, 15, 19, 23, 27, 31]

            if k in slots_word_sub and b not in [1, 14]:
                slots_used[0].append(k)
            if b != 0:
                slots_used[b].append(k)
                # the categories are only used in the define_irregular_words function
                # categories about 15 are used in other functions
                if b < 16: categories_used.append(b)

    categories_used.sort()
    sentence_slots[45] = slots_used
    sentence_slots[46] = categories_used
    sentence_slots[56] = places_used
    sentence_slots[57] = get_subclause_position(sentence_slots)
    if not list1[79] == 'do not rebuild sentence':
        sentence_slots = build_sent_slots_known(sentence_slots)

    return sentence_slots


def get_used_slots(k, part_of_speech, sub_part, sub_sub_part):
    # if you change the number of the slots then you must change which slot the universal is in
    # in the lies within univ quant function
    predicative_complement_positions = [14, 18, 22, 26, 29]

    b = 0
    if sub_sub_part == 'd':
        b = 18
    elif sub_part == 's':
        b = 17
    elif (part_of_speech == 'd' or part_of_speech == 'p') \
            and sub_part != 'b' and sub_part != "i":  # determinative, pronouns, possessive pronouns
        b = 1
    elif part_of_speech == 'o':
        b = 3  # common name possessives
    elif part_of_speech == 's':
        b = 4  # proper name possessives
    elif k == 66:
        b = 5  # and
    elif part_of_speech == 'a' and k not in predicative_complement_positions:  # adjectives
        b = 6
    elif k == 35 or k == 36:  # CIA
        b = 7
    elif part_of_speech == 'u' and k == 59:
        b = 8
    elif part_of_speech == 'u':  # relative pronouns
        b = 9
    elif part_of_speech == 'y':  # that
        b = 10
    elif part_of_speech == 'r' and sub_part == 'd':  # AS
        b = 11
    elif part_of_speech == 'r' and sub_part == 'b':  # RDC
        b = 12
    elif part_of_speech == 'r' and (k == 15 or k == 19):  # RDA
    # modify this if you change this then you must also change the eliminate relative pronouns function
        b = 13
    elif part_of_speech == 'n' and sub_part == 't':  # there
        b = 14
    elif part_of_speech == 'd' and sub_part == 'b':  # universals
        b = 15
    elif part_of_speech == 'm':
        b = 16

    return b


def get_part_of_speech(word, str5):
    sub_part_of_speech = ""
    sub_sub_part = ""
    if isvariable(word):
        pos = 'n'
        if word in variables:
            variables.remove(word)
            # if the variable stands for an adjective then its part of speech
            # is adjective
        str1 = abbreviations[0].get(word)
        if str1 != None and str5 != 'is in definition':
            posp = dictionary[0].get(str1)
            pos = posp[0]
            sub_part_of_speech = posp[1] if len(posp) > 1 else ""
            sub_sub_part = posp[2] if len(posp) > 2 else ""

    else:
        posp = dictionary[0].get(word)
        if posp == None:
            print ("you misspelled " + word)
            g = 4 / 0
        pos = posp[0]
        sub_part_of_speech = posp[1] if len(posp) > 1 else ""
        sub_sub_part = posp[2] if len(posp) > 2 else ""
        if word == "~":
            pos = 'm'
        elif word == ne:
            pos = 'r'
        elif word == 'not':
            pos = 'm'
        elif word[-2:] == "'s":
            posp = dictionary[0].get(word[-2:])
            if len(posp) > 2 and posp[2] == 'n':
                pos = 's'
            else:
                pos = 'o'

    return pos, sub_part_of_speech, sub_sub_part


def determine_if_compound_word(i, list1, word):
    if "," in word:
        word = word.replace(",", "")
        return i, word, True
    has_comma = False
    double = dictionary[4].get(word)
    triple = dictionary[5].get(word)
    triple_word = ""

    if triple != None:
        if list1[i + 1] != None and list1[i + 2] != None and "," not in list1[i + 1]:
            if "," in list1[i + 2]:
                after_next_word = list1[i + 2].replace(",", "")
                has_comma = True
            else:
                after_next_word = list1[i + 2]
            next_word = list1[i + 1]
            triple_word = word + " " + after_next_word + " " + next_word
            if triple_word in triple:
                i += 2
                word = triple_word
            else:
                triple_word = ""
        else:
            triple_word = ""

    if triple_word == "" and double != None:
        if list1[i + 1] != None:
            if "," in list1[i + 1]:
                next_word = list1[i + 1].replace(",", "")
                has_comma = True
            else:
                next_word = list1[i + 1]
            double_word = word + " " + next_word
            if double_word in double:
                i += 1
                word = double_word

    return i, word, has_comma


def build_sent_name(prop_name):
    str1 = ''
    str2 = ''
    list1 = []

    for i in range(len(prop_name)):
        if i == 24:
            bb = 8
        str3 = remove_outer_paren(prop_name[i][2])
        str3 = str3.replace("~", "")
        str1 = '(' + prop_name[i][0] + mini_e + str3 + ')'
        if len(str2) == 0 and len(str1) > 57:
            list1.append(str1)
        elif (len(str2) + len(str1)) > 57:
            list1.append(str2)
            str2 = str1
            if i + 1 == len(prop_name):
                list1.append(str2)
        elif (len(str2) + len(str1)) <= 57:
            if len(str2) == 0:
                str2 = str1
            else:
                str2 = str2 + ' & ' + str1
            if i + 1 == len(prop_name):
                list1.append(str2)
    return list1


def insert_the_long_way(list1, i):
    num = [0, 1, 2, 42, 72]
    list2 = [None] * 80

    j = 2
    while j + 1 < i:
        j += 1
        list2[j] = list1[j]
    while list1[i] != None:
        list2[i + 1] = list1[i]
        i += 1
    for j in num:
        list2[j] = list1[j]
    return list2



def quick_print(list1, list2, list3, kind=0, a=0, b=0, c=0):
    j = 1
    if kind == 0:
        for i in range(len(list2)):
            j += 1
            if list1 != []:
                w4.cell(row=j, column=2).value = list1[i]
            w4.cell(row=j, column=3).value = list2[i]
            if list3 != []:
                w4.cell(row=j, column=4).value = list3[i]
    elif kind == 1:
        for i in range(len(list2)):
            j += 1
            if a != 0:
                w4.cell(row=j, column=2).value = list2[i][a]
            w4.cell(row=j, column=3).value = list2[i][b]
            if c != 0:
                w4.cell(row=j, column=4).value = list2[i][c]

    wb4.save('../temp_proof.xlsx')
    sys.exit()


def print_sent_full(test_sent, tot_prop_name, yy=""):
    global result_data

    p = 2
    determine_words_used()
    o = -1

    for i in order:
        # if i == 2:
        #     break
        for j in range(len(test_sent[i])):

            if test_sent[i][j][6] != "":
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5]) + ',' + str(test_sent[i][j][6])
            elif test_sent[i][j][5] != "":
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5])
            elif test_sent[i][j][4] != "":
                str1 = test_sent[i][j][4]
            else:
                str1 = ""
            if j == 0:
                test_sent[i][j][4] = i
            else:
                test_sent[i][j][4] = str1
            if normal_proof:
                w4.cell(row=p, column=2).value = test_sent[i][j][0]
                w4.cell(row=p, column=3).value = test_sent[i][j][3] + test_sent[i][j][1]
                w4.cell(row=p, column=4).value = test_sent[i][j][4]
            else:
                result_data['text_' + str(p - 1) + '_1'] = test_sent[i][j][0]
                result_data['text_' + str(p - 1) + '_2'] = test_sent[i][j][1]
                result_data['text_' + str(p - 1) + '_3'] = test_sent[i][j][4]

            p += 1

        p += 1
        o += 1
        list1 = build_sent_name(tot_prop_name[o])
        for j in range(len(list1)):
            if normal_proof:
                w4.cell(row=p, column=3).value = list1[j]
            else:
                result_data['text_' + str(p) + '_2'] = list1[j]
            p += 1
        p += 1
        do_not_print = False
        for j in range(len(test_sent[i])):
            if j == 8:
                bb = 8
            if test_sent[i][j - 1][4] == 'ID':
                do_not_print = True
            elif test_sent[i][j][0] == "" and test_sent[i][j][1] == "":
                do_not_print = False
            if not do_not_print and test_sent[i][j][2] != "":
                if j == 0:
                    test_sent[i][j][4] == ""
                if normal_proof:
                    w4.cell(row=p, column=2).value = test_sent[i][j][0]
                    w4.cell(row=p, column=3).value = test_sent[i][j][3] + test_sent[i][j][2]
                    w4.cell(row=p, column=4).value = test_sent[i][j][4]
                else:
                    result_data['text_' + str(p) + '_1'] = test_sent[i][j][0]
                    result_data['text_' + str(p) + '_2'] = test_sent[i][j][3] + test_sent[i][j][2]
                    result_data['text_' + str(p - 1) + '_3'] = test_sent[i][j][4]

                p += 1
        p += 3


def determine_words_used():
    if get_words_used:
        for i in range(len(words_used)):
            j = dictionary[7].get(words_used[i], 3)
            ws.cell(row=j, column=2).value = 1


def build_dict(ex_dict):
    global dictionary

    parts_of_speech = {}
    definitions = {}
    synonyms = {}
    relations = {}
    doubles = {}
    triples = {}
    only_def_if_input = []  # words that are only defined if they appear in the input sentence
    map_words_to_row = {}
    essential_properties = {}
    almost_done = False

    i = -1
    if get_words_used:
        mm = 3000
    else:
        mm = len(ex_dict)

    while i < mm - 1:
        i += 1
        if get_words_used:
            if i == 0:
                i = 5
            s = ws.cell(row=i, column=1).value
            if s == 451:
                bb = 8
            pos = ws.cell(row=i, column=3).value
            word = ws.cell(row=i, column=4).value
            if word == "true*":
                word = "true"
            if word == "false*":
                word = "false"
            next_word = ws.cell(row=i, column=4).value
            if word == "" and next_word == "":
                break
        else:
            s = 0
            pos = ex_dict[i][0]
            word = ex_dict[i][1]
            if word != None:
                word = tran_str(word, 2)
                word = word[0]

        if word == None and almost_done:
            break

        almost_done = True if word == None else False

        if pos != None and pos != "":
            if not isinstance(pos, int):
                pos = pos.strip()
            if word == 'non_whole':
                bb = 7

            if isinstance(word, int): word = str(word)

            if "(" in word:
                cc = word.index("(")
                word = word[:cc - 1]

            word = word.strip()

            if get_words_used:
                abbrev_relat = ws.cell(row=i, column=5).value
                defin = ws.cell(row=i, column=6).value
            else:
                abbrev_relat = ex_dict[i][2]
                defin = ex_dict[i][3]
                defin = tran_str(defin, 3)
                defin = defin[0]
            map_words_to_row.update({word: s})
            if abbrev_relat != "": map_words_to_row.update({abbrev_relat: s})

            parts_of_speech.update({word: pos})
            parts_of_speech.update({abbrev_relat: pos})
            fir_let = pos[0]
            sec_let = pos[1] if len(pos) > 1 else ""
            thir_let = pos[2] if len(pos) > 2 else ""
            four_let = pos[3] if len(pos) > 3 else ""
            fif_let = pos[4] if len(pos) > 4 else ""

            synonyms = update_synonyms(defin, synonyms, sec_let)
            if fir_let == 'r': relations.update({word: abbrev_relat})

            if " " in word:
                m = word.count(" ")
                if m == 1:
                    word1 = copy.copy(word)
                    y = word1.find(" ")
                    word1 = word1[:y]
                    doubles.setdefault(word1, []).append(word)
                if m == 2:
                    word1 = copy.copy(word)
                    y = word1.find(" ")
                    word1 = word1[:y]
                    triples.setdefault(word1, []).append(word)

            if sec_let == 'k':
                only_def_if_input.append(word)

            if sec_let != 'a':
                if fir_let == "r":
                    definitions.update({abbrev_relat: defin})
                else:
                    definitions.update({word: defin})
                    if fir_let == "n":
                        essential_properties.update({word: ""})

    dictionary = [parts_of_speech, definitions,synonyms, relations,
                  doubles, triples, only_def_if_input, map_words_to_row, essential_properties]


def update_synonyms(defin, synonyms, sec_let):
    if sec_let == 's':
        str6 = defin[defin.find("=") + 1:-1]
        str6 = str6.strip()
        str7 = defin[1:defin.find("=")]
        str7 = str7.strip()
        synonyms.update({str7: str6})

    return synonyms

def findinlist(str1, list1, i, j):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            return str2
    return None


def findinlist_alert_error(str1, list1, i, j):
    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            if str2 != "" and str2 != None:
                return str2
    g = 4 / 0


def findposinmd(str1, list1, p):
    # this determines the position of an element in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return i
    return -1


def findposinmd_alert_error(str1, list1, p):
    # this determines the position of an element in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return i
    g = 4 / 0


def isinmdlist(str1, list1, p):
    # this determines whether or not an element is in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return True

    return False


def find_pos_in_1d_list_int(i, list1):
    for j in range(len(list1)):
        if i == list1[j]:
            return j
    print ('the number you were looking for was not in the list')
    g = 4 / 0


def findposinlist(str1, list1, i):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the position in the list

    if str1 == 0:
        return
    str2 = copy.copy(str1)
    if str2 != None:
        str2 = str2.replace(" ", "")
    for d in range(len(list1)):
        str3 = copy.copy(list1[d][i])
        if str3 != None:
            str3 = str3.replace(" ", "")
        if str2 == str3:
            return d
    else:
        return -1


def use_identity(identities, negated_conjunction, consistent):
    if not consistent:
        return False

    if identities != []:
        num = [5, 14, 18, 22]
        remove_duplicates2d(identities, 0, 1)
        for i in range(len(identities)):
            str1 = "(" + identities[i][0] + " = " + identities[i][1] + ")"
            for j in range(len(total_sent) - 1, 0, -1):
                if str1 in total_sent[j][1]:
                    identities[i][2] = total_sent[j][0]
                    break

        for i in range(len(identities)):
            str1 = identities[i][0]
            str2 = identities[i][1]
            anc1 = identities[i][2]
            for sent in detach_sent:
                for k in num:
                    if sent[k] == str1 or sent[k] == str2:
                        ant_sent_parts = copy.deepcopy(sent)
                        con_parts = copy.deepcopy(sent)
                        if sent[k] == str1:
                            con_parts[k] = str2
                        else:
                            con_parts[k] = str1
                            con_parts = build_sent2(con_parts)
                        if con_parts[42] == "q" + l1:
                            bb = 8
                        prepare_att_sent_1_sent(ant_sent_parts, "SUB", iff, [con_parts], anc1)
                        if con_parts[5] == con_parts[14] and \
                                        con_parts[9] != "=" and con_parts[2] == "":
                            consistent = use_reflexivity1(con_parts)
                            if not consistent:
                                return consistent

        consistent = detach1("do not use modus tollens", negated_conjunction)

    return consistent


def check_reflexivity(sent):
    consistent = True

    if sent[5] == sent[14] and sent[2] == "" and sent[9] != "=":
        consistent = use_reflexivity2(sent)

    return consistent


def use_reflexivity1(sent):
    global sn
    anc1 = sent[58]
    anc2 = attach_sent[-1][2]
    sn += 1
    add_to_total_sent(sn, sent[72], sent[1], "", "EE", anc1, anc2)
    consistent = use_reflexivity2(sent)

    return consistent


def use_reflexivity2(sent):
    global sn
    new_sent = copy.deepcopy(sent)
    new_sent[8] = "~"
    new_sent = build_sent2(new_sent)
    full_conditional = sent[0] + " " + conditional + " " + new_sent[0]
    abbrev_conditional = sent[42] + " " + conditional + " " + new_sent[42]
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "IRR")
    sn += 1
    consistent = add_to_total_sent_consist(sn, new_sent[72], new_sent[1], \
                                           new_sent[2], "MP", sn - 2, sn - 1, [])

    return consistent


def step_three(truth_value):
    global instan_time, instan_used

    negated_conjunction = []

    identities = get_identities()

    build_list_of_abbreviations()

    eliminate_attached_conjuncts()

    put_nc_id_ax_df_into_list()

    consistent = detach1("do not use modus tollens", negated_conjunction)

    consistent = add_stan_sent(consistent)

    consistent = use_identity(identities, negated_conjunction, consistent)

    consistent, object_prop2 = use_basic_lemmas(consistent)

    get_variable_type()

    consistent = use_axiom_of_definition2(consistent, negated_conjunction)

    consistent = step_four(negated_conjunction, consistent)

    consistent = final_truth_value(consistent, truth_value)

    return consistent


def final_truth_value(consistent, truth_value):
    if truth_value == 'co':
        truth_value = False
    else:
        truth_value = True
    if consistent == truth_value:
        return True
    else:
        return False


def eliminate_attached_conjuncts():
    global sn
    for sent in attach_sent:
        if sent[46] == 'eliminate as conjunct':
            sn += 1
            sent[2] = sn
            add_to_total_sent(sn, sent[37], sent[4], "", "&E", sent[2])


def get_identities():
    identities = []
    i = 0
    while i < len(detach_sent):
        if detach_sent[i][2] == "ID":
            identities.append([detach_sent[i][0], detach_sent[i][1], ""])
            del detach_sent[i]
        else:
            i += 1

    return identities


def put_nc_id_ax_df_into_list():
    # this function takes out sentences proved by NC, AX, RN etc and
    # prepares to put them into new locations
    list1 = []
    list2 = []
    list5 = []
    rn_used = False
    bool1 = False

    for i in range(0, len(total_sent)):
        if i == 9:
            bb = 7
        str1 = total_sent[i][4][:2]
        str2 = total_sent[i][4][:2]
        if total_sent[i][4] == "" and not bool1:
            list6 = copy.deepcopy(total_sent[i])
            list5.append(list6)
        elif total_sent[i][4] == "ID":
            list6 = copy.deepcopy(total_sent[i])
            list5.append(list6)
            bool1 = True
        elif 'DF' == str1 or 'NC' == str1 or 'AX' == str1 \
                or 'RN' == str2:
            list4 = copy.deepcopy(total_sent[i])
            list1.append(list4)
            rn_used = True
        elif bool1:
            if str1 == 'DE':
                total_sent[i][4] = "DF " + total_sent[i][4][3:]
            elif str1 == 'AY':
                total_sent[i][4] = "AX ENT"
            list3 = copy.deepcopy(total_sent[i])
            list2.append(list3)

    if rn_used:
        dict1 = rearrange_total_sent(list5, list1, list2)
        renumber_attach_sent(dict1)


def rearrange_total_sent(list5, list1, list2):
    # this function puts the total_sent into a better order

    global total_sent
    b = time.time()
    total_sent = []
    for i in range(len(list5)):
        total_sent.append(list5[i])
    for i in range(len(list1)):
        total_sent.append(list1[i])
    add_to_total_sent("", "")
    for j in range(len(list2)):
        total_sent.append(list2[j])

    dict1 = {}
    j = 0
    for i in range(len(total_sent)):
        if (i > 3 and total_sent[i][0] != "") or i <= 3:
            j += 1
            dict1.update({total_sent[i][0]: j})
            total_sent[i][0] = j

    for i in range(len(total_sent)):
        for j in range(0, 5):
            if len(total_sent[i]) > 5 + j:
                if total_sent[i][j + 5] != "":
                    total_sent[i][j + 5] = dict1.get(total_sent[i][j + 5], None)

                break

    return dict1


def use_axiom_of_definition2(consistent, negated_conjunction):
    global sn

    if consistent and detach_sent == []:
        new_instances = []
        for sent in attach_sent:
            if sent[3] != "x" and sent[3] != 'd':
                if sent[3] == 'e':
                    b = 35 if len(sent[35]) > len(sent[34]) else 34
                else:
                    b = 34
                for subsent in sent[b]:
                    for i in [5, 14, 18, 22]:
                        if subsent[i] in variable_type[0]:
                            new_sent = copy.deepcopy(subsent)
                            new_sent2 = copy.deepcopy(subsent)
                            new_sent2[i] = alpha
                            new_sent2 = build_temp_sent(new_sent2)
                            if new_sent2[0] not in new_instances:
                                new_instances.append(new_sent2[0])
                                new_sent[i] = variables[0]
                                variable_type[1].append(variables[0])
                                del variables[0]
                                new_sent = build_sent2(new_sent)
                                detach_sent.append(new_sent)
                                sn += 1
                                add_to_total_sent(sn, new_sent[72], new_sent[1], new_sent[2], "AX DEF")

        define_regular_terms(detach_sent)
        consistent = detach1("do not use modus tollens", negated_conjunction)

    return consistent


def get_id_sent():
    # this function gets the list of identities
    for i in range(len(total_sent)):
        if total_sent[i][4] == "ID":
            return [total_sent[i][0], total_sent[i][1], "", "", "", "", ""]


def print_variables(list1):
    # this prints out the variables within the total_sent list, just above
    # where it prints the attached sentences

    ## variable type = [general, defn, indef, same_sent]

    general = list1[0]
    indef = list1[1]
    defn = list1[2]
    identities = get_id_sent()
    gen_str = ""
    ind_str = ""
    def_str = ""

    if general != []:
        gen_str = 'General Variables: '
        for i in range(len(general)):
            gen_str += general[i] + " "
    if indef != []:
        ind_str = 'Indefinite Variables: '
        for i in range(len(indef)):
            ind_str += indef[i] + " "
    if defn != []:
        def_str = 'Constants: '
        for i in range(len(defn)):
            def_str += defn[i] + " "

    j = 0
    for i in range(len(total_sent) - 1, 0, -1):
        if len(total_sent[i][1]) > 15:
            if total_sent[i][1].startswith('STANDARD ATTA'):
                total_sent.insert(i, identities)
                if gen_str != "":
                    j += 1
                    total_sent.insert(i + j, ["", gen_str, "", "", "", "", "", ""])
                if ind_str != "":
                    total_sent.insert(i + j, ["", ind_str, "", "", "", "", "", ""])
                    j += 1
                if def_str != "":
                    total_sent.insert(i + j, ["", def_str, "", "", "", "", "", ""])
                return


def renumber_attach_sent(new_numbers):
    # this gives attach_sent their proper number according to the new
    # numbering system as arrived at in the rearrange_total_sent function

    if new_numbers != {}:
        for i in range(len(attach_sent)):
            old_num = attach_sent[i][2]
            if old_num < 400:
                attach_sent[i][2] = new_numbers.get(old_num)
                if attach_sent[i][2] == None:
                    print('you renumbering of attach_sent failed')


# object_properties, abbrev, general, class, accidental properties, parts of properties
# if object is a thing then it is stated whether or not it is consequential


def get_general_variables(sent):
    # modify this if you allow for more variables beyond 14
    # modify this is you allow for more than 2 disjuncts
    ant_var = set()
    con_var = set()
    for i in (34, 35):
        for subsent in sent[i]:
            for j in [5, 14, 18]:
                if isvariable(subsent[j]) and i == 34:
                    if subsent[j] not in ant_var:
                        ant_var.add(subsent[j])
                elif isvariable(subsent[j]) and i == 35:
                    if subsent[j] not in con_var:
                        con_var.add(subsent[j])
    general_variables = ant_var.intersection(con_var)
    for var in general_variables:
        variable_type[0].append(var)
        if var in variable_type[1]:
            variable_type[1].remove(var)


def get_variable_type():
    indefinite_concept = abbreviations[1].get("indefinite")
    for abbrev in abbreviations[0].keys():
        if isvariable(abbrev):
            variable_type[2].append(abbrev)

    for sent in all_sent:
        for i in [5, 14, 18]:
            if sent[i] == 'u':
                bb = 8
            if isvariable(sent[i]) and sent[i] not in variable_type[1] \
                    and sent[i] not in variable_type[0]:
                if sent[9] == "J" and sent[14] == indefinite_concept and i == 5:
                    if sent[i] not in variable_type[1]:
                        variable_type[1].append(sent[i])
                        if sent[i] in variable_type[2]: variable_type[2].remove(sent[i])
                elif sent[i] not in variable_type[2]:
                    variable_type[2].append(sent[i])

    variable_type[2].append("i")
    print_variables(variable_type)


def step_four(negated_conjunction, consistent):
    global instan_used, instan_time

    if consistent and attach_sent != []:
        object_properties, detached_predicates = get_detached_predicates(variable_type)

        attached_predicates, object_properties = get_attached_predicates(variable_type, object_properties)

        object_properties = rearrange_object_properties(object_properties)

        object_properties = print_general_object_properties(object_properties)

        print_object_properties(object_properties)

        instantiations = determine_if_same_class(object_properties)

        ax_def_used = use_axiom_of_definition(instantiations)

        substitute_in_attach_sent(instantiations)

        print_instantiations(instantiations)

        consistent = detach1("use modus tollens", negated_conjunction)

        consistent = reuse_axiom_of_definition(consistent, ax_def_used, negated_conjunction)

    return consistent


# object, exclusive classes, variable_type, class:sent, inclusive class, accidental properties
# rrr

def reuse_axiom_of_definition(consistent, ax_def_used, negated_conjunction):
    if consistent and ax_def_used:
        consistent = step_four(negated_conjunction, consistent)

    return consistent


def use_basic_lemmas(consistent):
    global lemmas_used, time_spent_in_lemma_function
    if not consistent:
        return consistent, {}

    lemmas_used += 1
    aa = time.time()

    exclusive_classes = ['moment', 'relationship', 'point', 'number',
                         'imagination', 'concept' + un, "property" + un, 'property',
                         'possible world', 'letter', 'mind', 'matter', 'sensorium']

    num = [5, 14, 18, 22]
    object_prop2 = {}
    error = ""
    j = -1
    while error == "" and j < len(detach_sent) - 1:
        j += 1
        sent = detach_sent[j]
        for i in num:
            object = sent[i]
            if object != None and object != "":
                category = get_class(sent[9], sent, i)
                if object in object_prop2.keys():
                    properties = object_prop2.get(object)
                    eclasses = properties[0]
                    iclasses = properties[3]
                    if category != 'thing' and category not in eclasses and category not in iclasses:
                        if category in exclusive_classes:
                            # the reason why we do not put sentences of the form bIc in the eclass_sent
                            # dictionary is because we cannot use basic lemmas with those types
                            # of sentences
                            if sent[9] == "I" and i == 5:
                                eclass_sent = properties[2]
                            else:
                                eclasses.append(category)
                                eclass_sent = properties[2]
                                eclass_sent.update({category: sent})

                        else:
                            iclasses.append(category)
                            eclass_sent = properties[2]

                        properties = [eclasses, "", eclass_sent, iclasses]
                        object_prop2[object] = properties
                        if len(eclasses) > 1:
                            error = "contradiction in exclusivity"
                            break

                else:
                    if category != 'thing':
                        eclass_sent = {}
                        if sent[9] == "I" and i == 5:
                            properties = [[], "", eclass_sent, []]
                        elif category in exclusive_classes:
                            eclass_sent = {category: sent}
                            properties = [[category], "", eclass_sent, []]
                        else:
                            properties = [[], "", eclass_sent, [category]]
                        object_prop2.update({object: properties})

    if error == "contradiction in exclusivity":
        consistent = add_basic_lemmas(object, properties)

    bb = time.time()
    time_spent_in_lemma_function += (bb - aa)

    return consistent, object_prop2


def add_basic_lemmas(object, obj_properties):
    class_sent = obj_properties[2]
    it = iter(class_sent.values())
    first_sent, second_sent = next(it), next(it)

    concept_thing = abbreviations[1].get("thing")
    if concept_thing == None:
        concept_thing = add_thing_to_abbreviations()
    obj_pos1 = 5 if first_sent[5] == object else 14
    obj_pos2 = 5 if second_sent[5] == object else 14
    sec_obj_pos = 5 if obj_pos2 == 14 else 14
    second_obj = second_sent[sec_obj_pos]

    build_original_lemma(obj_pos1, obj_pos2, second_sent, first_sent)

    int_thing = variables[0]
    del variables[0]

    build_rename_sent(object, second_obj, int_thing, concept_thing)

    build_renamed_lemma(obj_pos1, obj_pos2, second_sent, first_sent, object, concept_thing, int_thing)

    infer_sec_obj_thing(concept_thing, second_obj, second_sent)

    consistent = instantiate_int_thing(int_thing, concept_thing, second_obj, first_sent, second_sent)

    return consistent


def build_original_lemma(obj_pos1, obj_pos2, second_sent, first_sent):
    global sn
    thing_sent = "(d I e)"
    thing_concept = "(e = thing)"
    if obj_pos1 == 5 and obj_pos2 == 5:
        cond1 = "(" + "b" + " " + first_sent[9] + " " + "c" + ")"
        cond2 = "(" + "b" + " ~ " + second_sent[9] + " " + "d" + ")"
        name = "SS"
    elif obj_pos1 == 5 and obj_pos2 == 14:
        cond1 = "(" + "b" + " " + first_sent[9] + " " + "c" + ")"
        cond2 = "(" + "d" + " ~ " + second_sent[9] + " " + "b" + ")"
        name = "SO"
    elif obj_pos1 == 14 and obj_pos2 == 5:
        cond1 = "(" + "c" + " " + first_sent[9] + " " + "b" + ")"
        cond2 = "(" + "b" + " ~ " + second_sent[9] + " " + "d" + ")"
        name = "OS"
    elif obj_pos1 == 14 and obj_pos2 == 14:
        cond1 = "(" + "c" + " " + first_sent[9] + " " + "b" + ")"
        cond2 = "(" + "d" + " ~ " + second_sent[9] + " " + "b" + ")"
        name = "OO"

    original_conditional = "((" + cond1 + " & " + thing_sent + ") " + implies + " " \
                           + cond2 + ") " + thing_concept

    lemma_name = "LE." + first_sent[9] + "." + second_sent[9] + "." + name
    sn += 1
    add_to_total_sent(sn, original_conditional, "", "", lemma_name)


def build_rename_sent(key, second_obj, int_thing, concept_thing):
    global sn
    sent1 = "(b" + mini_c + key + ")"
    sent2 = "(c" + mini_c + second_obj + ")"
    sent3 = "(d" + mini_c + int_thing + ")"
    thing_sent = "(e" + idd + concept_thing + ")"
    list1 = [sent1, sent2, sent3, thing_sent]
    rename_sent = " & ".join(list1)
    sn += 1
    add_to_total_sent(sn, rename_sent, "", "", "RN")


def build_renamed_lemma(obj_pos1, obj_pos2, second_sent, first_sent,
                        key, concept_thing, int_thing):
    global sn
    if obj_pos1 == 5 and obj_pos2 == 5:
        cond1 = "(" + key + " " + first_sent[9] + " " + first_sent[14] + ")"
        cond2 = "(" + key + " ~ " + second_sent[9] + " " + second_sent[14] + ")"
    elif obj_pos1 == 5 and obj_pos2 == 14:
        cond1 = "(" + key + " " + first_sent[9] + " " + first_sent[14] + ")"
        cond2 = "(" + second_sent[5] + " ~ " + second_sent[9] + " " + key + ")"
    elif obj_pos1 == 14 and obj_pos2 == 5:
        cond1 = "(" + first_sent[5] + " " + first_sent[9] + " " + key + ")"
        cond2 = "(" + key + " ~ " + second_sent[9] + " " + second_sent[14] + ")"
    elif obj_pos1 == 14 and obj_pos2 == 14:
        cond1 = "(" + first_sent[5] + " " + first_sent[9] + " " + key + ")"
        cond2 = "(" + second_sent[5] + " ~ " + second_sent[9] + " " + key + ")"
    thing_sent = "(" + int_thing + " I " + concept_thing + ")"

    cond1p = first_sent[42]
    cond2p = name_sent(cond2)
    thing_sentp = name_sent(thing_sent)
    full_conditional = "(" + cond1 + " & " + thing_sent + ") " + implies + " " + cond2
    abbrev_conditional = "(" + cond1p + " & " + thing_sentp + ") " + implies + " " + cond2p
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "SUB", sn - 2, sn - 1)


def infer_sec_obj_thing(concept_thing, second_obj, second_sent):
    global sn
    thing_sent = "(" + second_obj + " I " + concept_thing + ")"
    thing_sentp = name_sent(thing_sent)
    full_conditional = second_sent[0] + " " + implies + " " + thing_sent
    abbrev_conditional = second_sent[42] + " " + implies + " " + thing_sentp
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "LE ENT")
    sn += 1
    add_to_total_sent(sn, thing_sent, thing_sentp, "", "MP", sn - 1, second_sent[58])
    if "~" in second_sent[42]: g = 4 / 0


def instantiate_int_thing(int_thing, concept_thing, second_obj, first_sent, second_sent):
    global sn
    rename_sent = "(" + int_thing + mini_c + second_obj + ")"
    sn += 1
    add_to_total_sent(sn, rename_sent, "", "", "RN", sn - 1, sn - 3)
    sec_antec = "(" + second_obj + " I " + concept_thing + ")"
    abbrev_sec_antec = name_sent(sec_antec)
    second_sent[8] = "~"
    second_sent = build_sent2(second_sent)
    full_conditional = "(" + first_sent[0] + " & " + sec_antec + ") " + \
                       implies + " " + second_sent[0]
    abbrev_conditional = "(" + first_sent[42] + " & " + abbrev_sec_antec + ") " + \
                         implies + " " + second_sent[42]
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "SUB", sn - 4, )
    conjunction = "(" + first_sent[0] + " & " + sec_antec + ")"
    conjunctionp = "(" + first_sent[42] + " & " + abbrev_sec_antec + ")"
    if "~" in first_sent[42]: g = 4 / 0
    sn += 1
    add_to_total_sent(sn, conjunction, conjunctionp, "", "&I", first_sent[58], sn - 3)
    sn += 1
    consistent = add_to_total_sent_consist(sn, second_sent[72],
                                           second_sent[1], second_sent[2], "MP", sn - 1, sn - 2, [])
    if consistent: g = 4 / 0

    return consistent


def add_thing_to_abbreviations():
    concept_thing = variables[0]
    del variables[0]
    abbreviations[0].update({concept_thing: 'thing'})
    abbreviations[1].update({'thing': concept_thing})
    d = findposinmd("ID", total_sent, 4)
    total_sent[d][1] += "(" + concept_thing + " = thing)"

    return concept_thing


def get_detached_predicates(variable_type):
    # this makes a list of the detached definite predicates
    object_properties = []
    detached_predicates = []
    for i in range(len(detach_sent)):
        sent = detach_sent[i]
        subj = detach_sent[i][5]
        relat = detach_sent[i][9]
        obj = detach_sent[i][14]
        obj = "" if obj == None else obj
        relat2 = detach_sent[i][15]
        obj2 = detach_sent[i][18]
        obj2 = "" if obj2 == None else obj2
        relat2 = "" if relat2 == None else relat2
        t_value = detach_sent[i][8]
        t_value = "" if t_value == None else t_value
        subj_pred = alpha + t_value + relat + obj + relat2 + obj2
        sub_sent_parts = [alpha, t_value, relat, obj, relat2, obj2]
        if obj != "":
            obj_pred = subj + t_value + relat + alpha + relat2 + obj2
            obj_sent_parts = [subj, t_value, relat, alpha, relat2, obj2]
        if obj2 != "":
            obj2_pred = subj + t_value + relat + obj + relat2 + alpha
            obj2_sent_parts = [subj, t_value, relat, obj, relat2, alpha]

        absolute_predicate = subj + relat + obj + relat2 + obj2
        s_variable_kind = get_quick_variable_type(subj, variable_type)
        o_variable_kind = get_quick_variable_type(obj, variable_type)
        o2_variable_kind = get_quick_variable_type(obj2, variable_type)
        detached_predicates.append([absolute_predicate, t_value, detach_sent[i][42]])
        if relat == "I":
            kind = abbreviations[0].get(obj)
            skind = kind_exception(kind)
        else:
            skind = get_class(relat, sent, 5)

        object_properties = get_object_properties(subj,
                                                  object_properties,
                                                  s_variable_kind,
                                                  subj_pred,
                                                  skind,
                                                  "c",
                                                  "",
                                                  "",
                                                  sub_sent_parts)
        if isvariable(obj) or obj == "i":
            okind = get_class(relat, sent, 14)
            object_properties = get_object_properties(obj,
                                                      object_properties,
                                                      o_variable_kind,
                                                      obj_pred,
                                                      okind,
                                                      "c",
                                                      "",
                                                      "",
                                                      obj_sent_parts)
        if isvariable(obj2) or obj2 == "i":
            object_properties = get_object_properties(obj2,
                                                      object_properties,
                                                      o2_variable_kind,
                                                      obj2_pred,
                                                      'thing',
                                                      "c",
                                                      "",
                                                      "",
                                                      obj2_sent_parts)

    return object_properties, detached_predicates


def use_axiom_of_definition(instantiations):
    global sn
    axiom_of_definition_used = False
    for var_list in instantiations:
        if var_list[2] == "T":
            axiom_of_definition_used = True
            list1 = [None] * 80
            list1[5] = var_list[1]
            list1[9] = "I"
            list1[14] = var_list[4]
            list1 = build_sent2(list1)
            sn += 1
            list1[58] = sn
            add_to_total_sent(sn, list1[0], list1[1], "", "AX ENT")
            detach_sent.append(list1)
        elif var_list[2] == "D":
            axiom_of_definition_used = True

    return axiom_of_definition_used


def substitute_in_attach_sent(instantiations):
    # this substitutes the attached variable with the detached variables

    num2 = [34, 35, 32, 31, 30, 29]
    var_slots = [5, 14, 18]
    total_num = []
    if instantiations == []:
        return
    o = -1
    while o < len(instantiations) - 1:
        o += 1
        sub_properties = instantiations[o]
        num = sub_properties[3]
        if isinstance(num, int):
            num = [num]
        att_var = sub_properties[0]
        det_var = sub_properties[1]
        simul_sub = False
        next_att_var = 'bb'
        if o < len(instantiations) - 1:
            next_att_var = instantiations[o + 1][0]
            next_det_var = instantiations[o + 1][1]
            next_num = instantiations[o + 1][3]
            if isinstance(next_num, int):
                next_num = [next_num]
            if next_att_var != att_var and num == next_num:
                simul_sub = True
        for i in num:
            m = -1
            while attach_sent[m + 1][26] != 'new conditional from instantiation':
                m += 1
                if attach_sent[m][2] == i:
                    attach_sent[m][26] = "not new"
                    cond_sent = copy.deepcopy(attach_sent[m])
                    if i not in total_num:
                        total_num.append(i)
                    for j in num2:
                        if cond_sent[j] == []:
                            break
                        else:
                            for f in range(len(cond_sent[j])):
                                atom_cond_sent = cond_sent[j][f]
                                new_sent = copy.deepcopy(atom_cond_sent)
                                new_sent[74] = False
                                for k in var_slots:
                                    already_done = False
                                    if atom_cond_sent[k] == att_var:
                                        new_sent[k] = det_var
                                        new_sent[74] = True
                                        already_done = True
                                        if not simul_sub:
                                            break
                                    if simul_sub and not already_done:
                                        if atom_cond_sent[k] == next_att_var:
                                            new_sent[k] = next_det_var
                                            new_sent[74] = True
                                cond_sent[j][f] = new_sent
                    make_new_attach_sent(cond_sent)
        if simul_sub:
            o += 1


def findposinmdlistint(i, list1, p):
    for j in range(len(list1)):
        if list1[j][p] == i:
            return j
    else:
        return -1


def make_new_attach_sent(cond_sent):
    # this builds new strings within the conditional list

    cond_sent[26] = "new conditional from instantiation"  # this means its new and we need to print it
    prop_var_greek = cond_sent[47]
    prop_var_greek2 = prop_var_greek
    for j in [34, 35, 32, 31, 30, 29]:
        for k in range(len(cond_sent[j])):
            atom_cond_sent = cond_sent[j][k]
            if atom_cond_sent[74]:
                abs_oldp = atom_cond_sent[1]
                atom_cond_sent = build_sent2(atom_cond_sent)
                if j == 34:
                    n = 0
                    q = 42
                elif j == 35:
                    n = 1
                    q = 43
                else:
                    print("you haven't coded for this yet")
                    g = 4 / 0

                for m in range(len(cond_sent[n])):
                    if cond_sent[n][m][0] == abs_oldp:
                        cond_sent[n][m][0] = atom_cond_sent[1]
                for m in range(len(cond_sent[38])):
                    if cond_sent[38][m] == abs_oldp:
                        cond_sent[38][m] = atom_cond_sent[1]
                for m in range(len(cond_sent[q])):
                    if cond_sent[n][m][0] == atom_cond_sent[1]:
                        cond_sent[q][m][0] = atom_cond_sent[72]

                cond_sent = build_conjunction(cond_sent, q)

            prop_var_greek = prop_var_greek.replace(atom_cond_sent[44], atom_cond_sent[0])
            prop_var_greek2 = prop_var_greek2.replace(atom_cond_sent[44], atom_cond_sent[42])
    cond_sent[4] = prop_var_greek2
    cond_sent[37] = prop_var_greek
    attach_sent.append(cond_sent)


def get_ant_and_cond(str1):
    list1 = mainconn(str1)
    ant = str1[:list1[1]]
    con = str1[list1[1] + 1:]
    ant = ant.strip()
    con = con.strip()
    ant = "" if os(ant) else ant
    con = "" if os(con) else con

    return [ant, con]


def build_short_sent(list1):
    # this makes a new natural language sentence, if it is standard

    list1[14] = "" if list1[14] == None else list1[14]
    list1[15] = "" if list1[15] == None else list1[15]
    list1[18] = "" if list1[18] == None else list1[18]
    list2 = ["(", list1[5], list1[2], list1[9], list1[14], list1[15], list1[18], ")"]
    str1 = "".join(list2)
    str1p = name_sent(str1)

    return [str1, str1p]


def build_conjunction(list1, q):
    if q == 42:
        n = 7
        o = 0
    else:
        n = 8
        o = 1

    if len(list1[q]) > 1:
        str1 = ""
        str1p = ""
        for i in range(len(list1[q])):
            if str1 == "":
                str1 = list1[q][i][1] + list1[q][i][0]
                str1p = list1[o][i][1] + list1[o][i][0]
            else:
                str1 += " & " + list1[q][i][1] + list1[q][i][0]
                str1p += " & " + list1[o][i][1] + list1[o][i][0]

        list1[n] = [str1p, ""]
        list1[q - 2] = [str1, ""]
    else:
        list1[n] = list1[o][0]
        list1[q - 2] = list1[q][0]

    return list1


def print_general_object_properties(object_properties):
    # this prints up the predicates for the general variables since
    # these are markedly more difficult to print

    for i in range(len(object_properties)):

        str3 = ""
        for j in range(len(object_properties[i][3])):
            str2 = ""
            for k in range(len(object_properties[i][3][j][0])):
                str2 += object_properties[i][3][j][0][k] + " "
            if object_properties[i][1] == "agen":
                str3 += "[" + str2 + str(object_properties[i][3][j][1]) + "] "
            else:
                str3 += str2
        if object_properties[i][2][0] == 'thing' and object_properties[i][1] == 'agen':
            if object_properties[i][6] != []:
                for m in range(len(object_properties[i][6])):
                    str3 += " " + "{" + object_properties[i][6][m][0] + " " \
                            + str(object_properties[i][6][m][3]) + "}"
        object_properties[i][4] = str3

    return object_properties


def print_instantiations(instantiations):
    # this adds the instantiations to the total_sent list
    global sn

    for instantiation in instantiations:
        sn += 1
        str1 = "(" + instantiation[0] + mini_c + instantiation[1] + ")"
        if instantiation[4] == "P":
            str2 = ""
            for number in instantiation[3]:
                str2 += " " + str(number)
            str1 += " in " + str2
        add_to_total_sent(sn, str1, "", "", "IN")

    for cond in attach_sent:
        if cond[26] == "new conditional from instantiation":
            sn += 1
            cond[2] = sn
            add_to_total_sent(cond[2], cond[37], cond[4], "", "SUB")


def print_object_properties(object_properties):
    add_to_total_sent("", "")
    add_to_total_sent("", "OBJECT PREDICATES")

    for i in object_properties:
        str1 = i[0] + "  "
        for j in range(len(i[2])):
            str1 += "  " + i[2][j]
        str1 += " |"
        str1 += "  " + i[4]
        add_to_total_sent("", str1)


def determine_if_same_class(object_properties):
    # this determines whether a particular object belongs to the same class
    # as a general object
    instantiations = []
    i = -1
    while object_properties[i + 1][1] == 'agen':
        i += 1
        gen_var = object_properties[i][0]
        general_groups = object_properties[i][2]
        general_properties = object_properties[i][3]
        general_numbers = object_properties[i][5]
        if general_groups[0] == 'thing':
            done, instantiations = instantiate_things(object_properties,
                                                      instantiations, i)
            if done:
                break
        else:
            match_found = False
            for object_property in object_properties:
                if object_property[1] != 'agen':
                    for group in object_property[2]:
                        if group in general_groups:
                            partic_var = object_property[0]
                            predicates = object_property[3]
                            b = len(instantiations)
                            instantiations = instantiate2(predicates,
                                                          general_properties,
                                                          instantiations,
                                                          general_numbers,
                                                          gen_var,
                                                          partic_var,
                                                          object_properties)
                            if len(instantiations) > b:
                                match_found = True
                            break
            else:
                if not match_found:
                    instantiations = infer_member(general_properties,
                                                  general_groups,
                                                  instantiations,
                                                  gen_var,
                                                  general_numbers)

    return instantiations


def infer_member(general_properties, general_groups, instantiations, gen_var, general_numbers):
    # if a general variable's only antecedent property is that it is a member
    # of a class then we may infer that it is detached
    # but only if its consequent contradicts a detached sentence

    if len(general_properties[0][2]) == 6:
        new_objects = copy.deepcopy(general_properties[0][2])
        for i, object in enumerate(general_properties[0][2]):
            if "'" in object:
                new_objects[i] = object.replace("'", "")
            if alpha in object:
                new_objects[i] = variables[0]
                instantiations.append([gen_var, variables[0], "D", general_properties[0][1], ""])
                variable_type[1].append(variables[0])
                del variables[0]
        new_sent = build_sent1(new_objects[0], new_objects[1],
                               new_objects[2], new_objects[3], new_objects[4], new_objects[5])
        detach_sent.append(new_sent)
        add_to_total_sent(sn, new_sent[72], new_sent[1], new_sent[2], "AX DEF")
        return instantiations
    else:

        for property in general_properties:
            for sub_property in property[2]:
                new_objects = copy.deepcopy(sub_property)
                for i, object in enumerate(new_objects):
                    if "'" in object:
                        new_objects[i] = object.replace("'", "")
                    if alpha in object:
                        variable_type[1].append(variables[0])
                        new_objects[i] = variables[0]
                        instantiations.append([gen_var, variables[0], "D", property[1], ""])
                        del variables[0]
                new_sent = build_sent1(new_objects[0], new_objects[1],
                                       new_objects[2], new_objects[3], new_objects[4], new_objects[5])
                detach_sent.append(new_sent)
                add_to_total_sent(sn, new_sent[72], new_sent[1], new_sent[2], "AX DEF")

            return instantiations


def instantiate_things(object_properties, instantiations, i):
    # this determines if an indefinite thing can be instantiated
    # we only instantiate it if the particular variable contradicts
    # the general variable's consequent property

    list1 = []
    gen_var = object_properties[i][0]
    if object_properties[i][6] == []:
        return False, instantiations
    consequent = object_properties[i][6]

    for con_list in consequent:
        gen_tvalue = ""
        con_sent = con_list[0]
        if "~" in con_sent:
            con_sent = con_sent.replace("~", "")
            gen_tvalue = "~"
        if "'" in con_sent:
            con_sent = remove_prime(con_list)
        for object in object_properties:
            if object[0] == 'v':
                bb = 8
            if object[1] != "agen":
                properties = object[3]
                for property_list in properties:
                    for property in property_list[0]:
                        partic_tvalue = ""
                        if "~" in property:
                            property = property.replace("~", "")
                            partic_tvalue = "~"
                        if con_sent == property and partic_tvalue != gen_tvalue:
                            thing_var = abbreviations[1].get("thing")
                            instantiations.append([gen_var, object[0], "T", con_list[3], thing_var])
                            return True, instantiations

    return False, instantiations


def remove_sp(str1):
    return str1.replace(" ", "")


def remove_prime(con_list):
    # this remove the prime sign from an indefinite variable

    con_list[4][1] = ""
    i = -1
    for word in con_list[4]:
        i += 1
        if "'" in word:
            word = word[:-1]
            con_list[4][i] = word

    return "".join(con_list[4])


def instantiate2(predicates, general_properties, instantiations, general_numbers, gen_var, partic_var,
                 object_properties):
    # this determines if a particular object has all the predicates of the
    # general object


    sentences = []
    properties = copy.deepcopy(general_properties)
    numbers = copy.deepcopy(general_numbers)
    for j in range(len(properties)):
        if properties[j][1] not in sentences:
            temp_list = copy.deepcopy(properties[j][0])
            for k in range(len(properties[j][0])):
                gen_prop = properties[j][0][k]
                if gen_prop == "$":
                    temp_list.remove("$")
                    if temp_list == []:
                        sentences.append(properties[j][1])
                        numbers.remove(properties[j][1])
                        break
                elif properties[j][1] not in sentences:
                    indef_instant_used = False
                    if "'" in gen_prop:
                        old_gen_prop = gen_prop
                        list1 = change_indef_attach_var(properties[j],
                                                        object_properties, k, instantiations)
                        if not list1:
                            pass
                        else:
                            gen_prop = list1[0]
                            if list1[1] != 'already found':
                                numbers2 = copy.deepcopy(numbers)
                                instantiations.append([list1[1], list1[2], "I", numbers2, ""])
                            indef_instant_used = True

                    for predicate_set in predicates:
                        for predicate in predicate_set[0]:
                            if gen_prop == predicate:
                                if indef_instant_used:
                                    gen_prop = old_gen_prop
                                temp_list.remove(gen_prop)
                                if temp_list == []:
                                    sentences.append(properties[j][1])
                                    numbers.remove(properties[j][1])
                                break

    if sentences != []:
        if numbers == []:
            instantiations.append([gen_var, partic_var, "", general_numbers, ""])
        else:
            instantiations.append([gen_var, partic_var, "", sentences, "P"])

    return instantiations


def change_indef_attach_var(sent_parts, object_properties, k, instantiations):
    # This determines whether an indefinite attached variable
    # is identical to a detached abbreviation

    for att_var in sent_parts[2][k]:
        if "'" in att_var:
            att_var = att_var[:-1]
            break
    if instantiations != []:
        for var in instantiations:
            if att_var == var[0]:
                det_var = var[1]
                new_sent = indefinite_instantiation(det_var, att_var,
                                                    sent_parts[2][k])
                return [new_sent, "already found", ""]

    d = findposinmd(att_var, object_properties, 0)
    if d == -1:
        print('there is something wrong with your ')
        'object properties list'
        sys.exit()
    set_attach = set(object_properties[d][2])
    attach_sent_parts = get_sent_parts(object_properties[d][3])
    for detach_prop in object_properties:
        set_detach = set(detach_prop[2])
        list1 = set_attach.intersection(set_detach)
        if list1 == set_attach and detach_prop[0] != att_var and detach_prop[1] != 'agen':
            det_var = detach_prop[0]
            for attach_sent_part in attach_sent_parts:
                for detach_sent_part in detach_prop[3][0][2]:
                    potentially_identical = False
                    for i in range(6):
                        try:
                            if attach_sent_part[i] == detach_sent_part[i] or \
                                            attach_sent_part[i] == alpha or \
                                            detach_sent_part[i] == alpha or \
                                            attach_sent_part[i] == att_var + "'":
                                pass
                            else:
                                break
                        except:
                            bb = 8
                    else:
                        potentially_identical = True
                        break
                else:
                    return False
            if not potentially_identical:
                return False
            else:
                new_sent = indefinite_instantiation(det_var, att_var, sent_parts[2][k])
                return [new_sent, att_var, det_var]

    return False


def get_sent_parts(list1):
    # this flattens the list of sent parts in the object properties list
    properties = []
    sent_parts = []
    for i in range(len(list1)):
        for j in range(len(list1[i][0])):
            if list1[i][0][j] not in properties:
                properties.append(list1[i][0][j])
                sent_parts.append(list1[i][2][j])

    return sent_parts


def indefinite_instantiation(det_var, att_var, sent):
    # this replaces the attached variable with the detached variable
    # in the sentence
    i = -1
    for word in sent:
        i += 1
        if word == att_var + "'":
            sent[i] = det_var
            break
    str1 = "".join(sent)
    return str1


def rearrange_object_properties(object_properties):
    # rearrange the object properties of the general objects so as to make it
    # easier to instantiate with
    # it takes the list of the form [Rd, cf, 33],[Re, cd, 33],[Rg,a,34]
    # and puts into the list [[Rd,Re],33],[[Rg],34]
    i = -1
    while i < len(object_properties) - 1:
        i += 1
        if i > 10:
            bb = 8
        list2 = []
        properties = object_properties[i][3]
        j = 0
        sentence_numbers = []
        while j < len(properties) - 1:
            sent_type = properties[j][1][-1]
            sent_num = properties[j][3]
            if sent_num not in sentence_numbers:
                sentence_numbers.append(sent_num)
            if j + 1 < len(properties):
                next_sent_num = properties[j + 1][3]
                next_sent_type = properties[j + 1][1][-1]
                predicate = properties[j][0]
                sent_parts = properties[j][4]
                list1 = []
                list3 = []
                while sent_type == next_sent_type and sent_num == next_sent_num:
                    list1.append(predicate)
                    list3.append(sent_parts)
                    if j + 1 < len(properties):
                        j += 1
                        sent_type = properties[j][1][-1]
                        sent_num = properties[j][3]
                        predicate = properties[j][0]
                        sent_parts = properties[j][4]
                        if j + 1 < len(properties):
                            next_sent_type = properties[j + 1][1][-1]
                            next_sent_num = properties[j + 1][3]
                        else:
                            break
                list1.append(predicate)
                list3.append(sent_parts)
                list2.append([list1, sent_num, list3])
                j += 1
        if j < len(properties):
            list2.append([[properties[j][0]], properties[j][3], properties[j][4]])
            sent_num = properties[j][3]
            if sent_num not in sentence_numbers:
                sentence_numbers.append(sent_num)

        object_properties[i][3] = list2
        object_properties[i][5] = sentence_numbers

    return object_properties


def get_object_properties(object,
                          object_properties,
                          variable_kind,
                          property,
                          kind,
                          sent_kind,
                          sent_num,
                          cond_num,
                          sent_parts):
    if object == 'r':
        bb = 8
    uninformative_properties = ["I", "J", "H"]  # these are properties all object_properties have
    if property not in uninformative_properties:

        d = findposinmd(object, object_properties, 0)
        if kind != 'thing2' and sent_parts[2] == 'I' and variable_kind == 'agen':
            property = "$"
        if kind == 'thing2':
            kind = 'thing'  # see kind exception for explaination

        if d == -1:

            if sent_kind[-1] == 'q':
                object_properties.append([object, variable_kind, [kind],
                                          [], "", "",
                                          [[property, sent_kind, sent_num, cond_num, sent_parts]]])
            else:
                object_properties.append([object, variable_kind, [kind],
                                          [[property, sent_kind, sent_num, cond_num, sent_parts]],
                                          "",
                                          "", []])
        else:
            for i in range(len(object_properties)):
                if object_properties[i][0] == object:
                    list_kind = object_properties[i][2]
                    conseq_properties = object_properties[i][6]
                    list_properties = object_properties[i][3]
                    if kind not in list_kind and kind != "":
                        list_kind.append(kind)
                    if sent_kind[-1] == 'q':
                        list1 = [property, sent_kind, sent_num, cond_num, sent_parts]
                        conseq_properties.append(list1)
                        object_properties[i] = [object, variable_kind, list_kind,
                                                list_properties, "", "", conseq_properties]
                    else:
                        list1 = [property, sent_kind, sent_num, cond_num, sent_parts]
                        list_properties.append(list1)
                        object_properties[i] = [object, variable_kind, list_kind,
                                                list_properties, "", "", conseq_properties]
                    break

    return object_properties


def get_class(relat, sent, p):
    # this determines what class or category an object belongs to

    if relat == "A" or (relat == 'T' and p == 14):
        kind = 'moment'
    elif relat == "IR" and p == 5:
        kind = 'relationship'
    elif relat == 'AB' or relat == "L" or relat == 'AB' or (relat == 'S' and p == 14):
        kind = 'point'
    elif relat == "G" or (relat == 'N' and p == 14):
        kind = 'number'
    elif relat == "M" and p == 5 or (relat == 'TK' and p == 14):
        kind = 'relationship'
    elif relat == "M" and p == 14:
        kind = 'imagination'
    elif relat == "I" and p == 5:
        group = sent[14]
        kind = abbreviations[0].get(group)
    elif relat == "I" and p == 14:
        kind = "concept" + un
    elif relat == "H" and p == 14:
        kind = "property" + un
    elif relat == "J" and p == 14:
        kind = "property"
    elif relat == "W" and p == 5:
        kind = "whole"
    elif relat == "W" and p == 14:
        kind = 'thing'
    elif relat == 'P' and p == 14:
        kind = 'possible world'
    elif relat == "D" and p == 14:
        kind = 'relationship'
    elif relat == 'AL':
        kind = 'letter'
    elif (relat == 'TK' or relat == "D") and p == 5:
        kind = 'mind'
    elif relat == "S" and p == 5:
        kind = 'matter'
    elif relat == "O" and p == 14:
        kind = 'sensorium'
    else:
        kind = "thing"

    return kind


def get_attached_predicates(variable_type, object_properties):
    # this makes a list of the attached definite predicates
    # and also adds to the list of object properties

    num = [34, 35, 32, 31, 30, 29]
    attached_predicates = []

    for i in range(len(attach_sent)):
        sent = attach_sent[i]
        cond_num = attach_sent[i][2]  # conditional number
        for j in num:
            if attach_sent[i][j] != []:
                for k in range(len(attach_sent[i][j])):
                    subj = attach_sent[i][j][k][5]
                    obj = attach_sent[i][j][k][14]
                    obj = "" if obj == None else obj
                    relat = attach_sent[i][j][k][9]
                    relat2 = attach_sent[i][j][k][15]
                    relat2 = "" if relat2 == None else relat2
                    obj2 = attach_sent[i][j][k][18]
                    obj2 = "" if obj2 == None else obj2
                    t_value = attach_sent[i][j][k][8]
                    t_value = "" if t_value == None else t_value
                    sent_kind = attach_sent[i][j][k][53]
                    sent_num = str(cond_num) + "." + attach_sent[i][j][k][68]
                    s_variable_kind = get_quick_variable_type(subj, variable_type)
                    o_variable_kind = get_quick_variable_type(obj, variable_type)
                    o2_variable_kind = get_quick_variable_type(obj2, variable_type)
                    new_subj, new_obj, new_obj2 = [subj, obj, obj2]
                    if s_variable_kind == 'agen':
                        new_subj = alpha
                    elif s_variable_kind == 'indefinite':
                        new_subj = subj + "'"
                    if o_variable_kind == 'agen':
                        new_obj = alpha
                    elif o_variable_kind == 'indefinite':
                        new_obj = obj + "'"
                    if o2_variable_kind == 'agen':
                        new_obj2 = alpha
                    elif o2_variable_kind == 'indefinite':
                        new_obj2 = obj2 + "'"
                    absolute_predicate = new_subj + relat + new_obj + relat2 + new_obj2
                    spredicate = alpha + t_value + relat + new_obj + relat2 + new_obj2
                    opredicate = new_subj + t_value + relat + alpha + relat2 + new_obj2
                    o2predicate = new_subj + t_value + relat + new_obj + relat2 + alpha
                    sent_parts = [new_subj, t_value, relat, new_obj, relat2, new_obj2]
                    attached_predicates.append([absolute_predicate, t_value, attach_sent[i][2]])
                    if relat == "I":
                        skind = abbreviations[0].get(obj)
                        skind = kind_exception(skind)
                    else:
                        skind = get_class(relat, sent, 5)

                    object_properties = get_object_properties(subj,
                                                              object_properties,
                                                              s_variable_kind,
                                                              spredicate,
                                                              skind,
                                                              sent_kind,
                                                              sent_num,
                                                              cond_num,
                                                              sent_parts)

                    if isvariable(obj) or obj == "i":
                        okind = get_class(relat, sent, 14)
                        object_properties = get_object_properties(obj,
                                                                  object_properties,
                                                                  o_variable_kind,
                                                                  opredicate,
                                                                  okind,
                                                                  sent_kind,
                                                                  sent_num,
                                                                  cond_num,
                                                                  sent_parts)

                    if isvariable(obj2) or obj == 'i':
                        object_properties = get_object_properties(obj2,
                                                                  object_properties,
                                                                  o2_variable_kind,
                                                                  o2predicate,
                                                                  'thing',
                                                                  sent_kind,
                                                                  sent_num,
                                                                  cond_num,
                                                                  sent_parts)

            else:
                break

    object_properties = sorted(object_properties, key=operator.itemgetter(1))
    object_properties = purge_thing_from_properties(object_properties, variable_type)

    return attached_predicates, object_properties


def purge_thing_from_properties(object_properties, variable_type):
    # if an object belongs to a class more specific than 'thing' then
    # we need not state that it belongs to the class 'thing'

    for object_property in object_properties:
        if 'thing' in object_property[2] and len(object_property[2]) > 1:
            object_property[2].remove("thing")
    return object_properties


def kind_exception(str1):
    # since everything belongs to the class 'whole' or 'part' these are not
    # genuine classes
    exceptions = ['whole', 'part']
    if str1 in exceptions:
        return 'thing'
    elif str1 == None:
        return 'thing2'
    return str1
    # if str1 equals none then that means the subject belongs to an indefinite
    # concept


def get_quick_variable_type(variable, variable_type):
    # this tells us what type a certain variable is after we already know what it is

    if variable == "":
        return ""
    defn = variable_type[2]
    gen = variable_type[0]
    indef = variable_type[1]

    if variable in gen:
        return 'agen'
    elif variable in defn:
        return 'definite'
    elif variable in indef:
        return 'indefinite'


def add_stan_sent(consistent):
    # this adds new sentences to the total_sent list

    if not consistent:
        return False

    add_to_total_sent("", "")
    add_to_total_sent("", "________________")
    add_to_total_sent("", "NONSTANDARD SENTENCES")
    i = 0
    while i < len(detach_sent):
        detach_sent[i][54] = is_standard(detach_sent[i])
        if not detach_sent[i][54]:
            add_to_total_sent(detach_sent[i][58], detach_sent[i][2] + detach_sent[i][72])
            del detach_sent[i]
        else:
            i += 1

    add_to_total_sent("", "")
    add_to_total_sent("", "________________")
    add_to_total_sent("", "STANDARD ATTACHED SENTENCES")

    for sent in attach_sent:
        add_to_total_sent(sent[2], sent[37])

    add_to_total_sent("", "")
    add_to_total_sent("", "________________")
    add_to_total_sent("", "STANDARD DETACHED SENTENCES")

    for sent in detach_sent:
        add_to_total_sent(sent[58], sent[2] + sent[72])
        consistent = check_reflexivity(sent)
        if not consistent:
            return consistent

    return True


def asymmetry(all_sent, str1, str2):
    for i in range(len(all_sent)):
        if all_sent[i][46] != "x":
            if all_sent[i][5] == str1 or all_sent[i][5] == str2:
                if all_sent[i][14] == str1 or all_sent[i][14] == str2:
                    return False
    return True


def is_in_md(list1, i, str1, bool1=False, k=0):
    if not bool1:
        for j in range(len(list1)):
            if list1[j][i] == str1:
                return True
        return False
    else:
        for j in range(len(list1)):
            if j != k:
                if list1[j][i] == str1:
                    return True
        return False


def tilde_removal(str1):
    str4 = ""
    if str1.find("~") > -1:
        str4 = "~"
        str1 = str1.replace("~", "")
    return [str1, str4]


def tilde_removal2(str1):
    j = 0
    if str1[:2] == "~(":
        for i in range(len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i > 0:
                if i == len(str1) - 1:
                    str1 = str1[1:]
                    str4 = "~"
                    return [str1, str4]
                else:
                    return [str1, ""]
    elif str1[0] == "~" and str1[1].islower() and os(str1):
        str1 = str1[1:]
        str4 = "~"
    else:
        str4 = ""
    return [str1, str4]


def determine_if_all_cond_4_detach_met(g, k):
    # this loops through all the sentences in the antecedent or the consequent
    # and determines if they all have been detached

    conditions = copy.deepcopy(attach_sent[g][k])
    del conditions[0]
    ancestors = []
    sent_type = attach_sent[g][3]
    done = False
    i = 0
    while i < len(attach_sent[g][k]) - 1 and not done:
        i += 1
        for sent in detach_sent:
            temp = sent[1]

            if sent[1] == 'h' + l1:
                bb = 8
            if sent[1] == attach_sent[g][k][i][0] \
                    and sent[2] == attach_sent[g][k][i][1]:
                ancestors.append(sent[58])
                del conditions[0]
                if conditions == []:
                    output = ["", ancestors]
                    done = True
                break

            elif sent[1] == attach_sent[g][k][i][0] \
                    and sent[2] != attach_sent[g][k][i][1]:
                if (sent_type == "e" and k == 0) or k == 1:
                    done = True
                    output = ['a sentence was negated', i]
                elif sent_type == 'c' and k == 0:
                    done = True
                    output = ["", ""]
                break
        else:
            done = True
            output = ["", ""]

    return output


def detach1(str1, negated_conjunction):
    global st_log_time
    b = time.time()
    if str1 == 'do not use modus tollens':
        kind = '~MT'
    else:
        kind = 'MT'

    if attach_sent == []:
        return True
    consistent = True
    set_of_det_sent = []
    r = -1
    while consistent and r < len(detach_sent) - 1:
        r += 1
        if r == 46:
            bb = 8
        g = -1
        while consistent and g < len(attach_sent) - 1:
            g += 1
            if attach_sent[g][26] != 'not new':
                k = -1
                while consistent and k < 1:
                    k += 1
                    sent_type = attach_sent[g][3]
                    if attach_sent[g][3] == "e" or attach_sent[g][3] == "c":
                        temp_detach_sent = detach_sent[r][1]
                        det_tvalue = detach_sent[r][2]
                        temp_attach_sent = attach_sent[g][k][0][0]
                        att_tvalue = attach_sent[g][k][0][1]

                        if temp_detach_sent == 'l':
                            bb = 8
                        if temp_detach_sent == temp_attach_sent:
                            if det_tvalue == att_tvalue and (k == 0 or (k == 1 and sent_type == "e")):
                                rule = "EE" if sent_type == 'e' else "MP"
                                if len(attach_sent[g][k]) == 1:

                                    # if k = 0 and sent_type = 'c' and set_of_det_sent = []
                                    # then p & (p > q)
                                    # if k = 0 and sent_type = 'e' and set_of_det_sent = []
                                    # then p & (p <> q)
                                    # if k = 1 and sent_type = 'e' and set_of_det_sent = []
                                    # then p & (q <> p)
                                    # if k = 0 and sent_type = 'c' and set_of_det_sent != []
                                    # then p & r & ((p & r) > q)
                                    # if k = 0 and sent_type = 'e' and set_of_det_sent != []
                                    # then p & r & ((p & r) <> q)
                                    # if k = 1 and sent_type = 'e' and set_of_det_sent != []
                                    # then p & r & (q <> (p & r))

                                    consistent, g, k = detach2(k, r, g, rule,
                                                               set_of_det_sent, negated_conjunction)
                                    if g > len(attach_sent) - 1:
                                        break
                                else:
                                    output = determine_if_all_cond_4_detach_met(g, k)
                                    if output[0] == 'a sentence is negated':
                                        rule = "EN" if sent_type == 'e' else "MT"
                                        consistent, g, k = detach2(k, output[1], g, rule, [], negated_conjunction)
                                        if g > len(attach_sent) - 1:
                                            break
                                    elif output[1] != "":
                                        output[1].insert(0, detach_sent[r][58])
                                        consistent, g, k = detach2(k, r, g, rule, output[1], negated_conjunction)
                                        if g > len(attach_sent) - 1:
                                            break
                            else:
                                if k == 0:
                                    t = 27
                                else:
                                    t = 28
                                if k == 0 and sent_type == 'c':
                                    pass
                                elif kind == "MT" and attach_sent[g][t] and \
                                        ((k == 0 and sent_type == 'e') or (k == 1)):
                                    rule = "EN" if sent_type == 'e' else "MT"
                                    consistent, g, k = detach2(k, r, g, rule, set_of_det_sent, negated_conjunction)
                                    if g > len(attach_sent) - 1:
                                        break

    c = time.time()
    c = c - b
    st_log_time += c

    return consistent


def detach2(k, r, g, rule, set_of_det_sent, negated_conjunction):
    global sn

    if sn == 25:
        bb = 8

    anc1 = attach_sent[g][2]
    anc2 = detach_sent[r][58]
    sn += 1
    if sn == 28:
        bb = 8
    if k == 0:
        m = 41
        h = 43
        t = 8
        s = 1
        n = 35
    else:
        m = 40
        h = 42
        t = 7
        s = 0
        n = 34

    introduce_conjunction(set_of_det_sent)

    if attach_sent[g][t][0] == 't':
        bb = 8

    consistent = add_to_total_sent_consist(sn, attach_sent[g][m][0], attach_sent[g][t][0],
                                           attach_sent[g][m][1], rule,
                                           anc1, anc2, negated_conjunction)

    if (rule == "MT" or rule == "EN") and len(attach_sent[g][h]) > 1:
        negated_conjunction.append([attach_sent[g][s], attach_sent[g][m]], attach_sent[g][t])
        g = 4 / 0
    else:
        if len(attach_sent[g][h]) == 1:
            if consistent:
                if os(attach_sent[g][s][0][0]):
                    list3 = attach_sent[g][n][0]
                    list3[58] = sn
                    is_in_detach_sent = isinmdlist(list3[42], detach_sent, 42)
                    if not is_in_detach_sent:
                        detach_sent.append(list3)
                        if list3[1] == 'l':
                            bb = 8
                else:
                    for lst in attach_sent[g][39]:
                        lst[2] = sn
                        attach_sent.append(lst)

        else:
            consistent = eliminate_conjuncts(g, r, h, negated_conjunction)
    del attach_sent[g]
    if g + 1 == len(attach_sent):
        g -= 1
    k -= 1
    return consistent, g, k


def introduce_conjunction(set_of_det_sent):
    global sn
    full_sent = ""
    if set_of_det_sent != []:
        sn += 1
        for num in set_of_det_sent:
            d = findposinmdlistint(num, detach_sent, 58)
            if d == -1: g = 4 / 0
            if full_sent == "":
                full_sent = detach_sent[d][0]
                abbrev_sent = detach_sent[d][42]
                anc = str(detach_sent[d][58])
            else:
                full_sent += " & " + detach_sent[d][0]
                abbrev_sent += " & " + detach_sent[d][42]
                anc += "," + str(detach_sent[d][58])
        full_sent = "(" + full_sent + ")"
        abbrev_sent = "(" + abbrev_sent + ")"
        add_to_total_sent(sn, full_sent, abbrev_sent, "", "&I", anc)


def eliminate_conjuncts(g, r, h, negated_conjunction):
    # if the detached sentences are a conjunction then this function
    # places each individual conjunct into the total_sent and detach_sent list

    global sn
    num = copy.copy(sn)
    if h == 43:
        m = 41
        k = 1
        n = 35
        c = 8
    else:
        m = 40
        k = 0
        n = 34
        c = 7
    conjunct_list = attach_sent[g][k]

    for i in range(len(conjunct_list)):
        sn += 1
        consistent = add_to_total_sent_consist(sn, attach_sent[g][h][i][0], attach_sent[g][k][i][0],
                                               attach_sent[g][h][i][1], "&E", num, "", negated_conjunction)
        if consistent:
            if os(conjunct_list[i][0]):
                d = findposinmd_alert_error(conjunct_list[i][0], attach_sent[g][n], 1)
                sent_parts = attach_sent[g][n][d]
                sent_parts[58] = sn
                is_in_detach_sent = isinmdlist(sent_parts[42], detach_sent, 42)
                if not is_in_detach_sent:
                    detach_sent.append(sent_parts)
            else:
                d = findposinmd_alert_error(conjunct_list[i][0], attach_sent[g][39], 4)
                list2 = attach_sent[g][39][d]
                list2[2] = sn
                attach_sent.append(list2)
        else:
            break

    return consistent


def add_to_total_sent_consist(num, str1, str2, tvalue, rule, anc1, anc2, negated_conjunction):
    list2 = [""] * 9
    list2[0] = num
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    total_sent.append(list2)
    consistent = check_consistency(negated_conjunction)

    return consistent


def add_to_total_sent(num, str1, str2="", tvalue="", rule="", anc1="", anc2=""):
    list2 = [""] * 9
    list2[0] = num
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    total_sent.append(list2)


def check_consistency(negated_conjunction):
    new_sent_abbr = total_sent[-1][2]
    tvalue = total_sent[-1][3]
    for i in range(len(total_sent) - 2, -1, -1):
        if total_sent[i][2] == new_sent_abbr and total_sent[i][3] != tvalue:
            build_contradiction(i)
            return False

    for lst in negated_conjunction:
        for sent in lst[0]:
            if sent[0] == new_sent_abbr and sent[1] == tvalue:
                g = 4 / 0
                del sent[0]
                del sent[1]
                ancestors = sent[4]
                ancestors.append(sn)
                if lst[0] == []:
                    bb = 8
                    pass
                    #

    return True


def build_contradiction(i):
    global sn
    sn += 1
    str1 = total_sent[-1][1] + " & ~" + total_sent[i][1]
    str2 = total_sent[-1][2] + " & ~" + total_sent[i][2]
    total_sent.append([sn, str1, str2, "", "&E", total_sent[-1][0], total_sent[i][0], "", ""])
    sn += 1
    total_sent.append([sn, bottom, "", "", bottom + "I", sn - 1, "", "", ""])


def disjunction_heirarchy(str5, d, new_disj=False):
    global prop_name
    global sn, pn

    if d > len(attach_sent) - 1:
        return
    if iff in str5 or conditional in str5:
        return

    str5 = enclose(str5)
    def_info = find_sentences(str5)
    mainc = def_info[4][0][1]
    list2 = [""] * 50
    n = 7
    if mainc == xorr:
        list2[3] = 'x'
    else:
        list2[3] = 'd'
    if attach_sent == [] or new_disj:
        list2[2] = pn
    else:
        list2[2] = attach_sent[d][2]
    list2[5] = ""
    list2[4] = def_info[0][0]  # fix this
    sentences = []

    for i in range(len(def_info[0])):
        if os(def_info[0][i]):
            siblings = []
            list3 = [None] * 9
            n += 1
            # str1 = findinlist(def_info[0][i],prop_name,1,0)
            str2 = def_info[4][i][0][:-1]
            g = findposinlist(str2, def_info[4], 0)
            parent = def_info[0][g]
            if def_info[4][g][1] == "&":
                list3[2] = 'c'
            elif def_info[4][g][1] == xorr:
                list3[2] = 'x'
            else:
                list3[2] = 'd'
            if len(str2) > 1:
                str3 = def_info[4][i][0][:-2]
                g = findposinlist(str3, def_info[4], 0)
                gparent = def_info[0][g]
            else:
                gparent = parent
            list3[1] = def_info[4][i][0]
            list3[5] = parent
            list3[6] = gparent
            list3[0] = [def_info[0][i], def_info[1][i]]
            # fix this
            b = parent.count(xorr)
            c = parent.count(idisj)
            if c > 1 or b > 1:
                list3[7] = 2
            else:
                list3[7] = 1
            sent_num = def_info[4][i][0]
            m = len(sent_num)
            for j in range(len(def_info[4])):
                if len(def_info[4][j][0]) == m and def_info[4][j][0][:-1] == str2 \
                        and j != i:
                    siblings.append([def_info[0][j], def_info[1][j]])  # fix this
            list3[4] = siblings
            list2[n] = list3
            sentences.append(list3[0][0])
            if list3[0][0] not in rel_conj:
                rel_conj.append(list3[0][0])
    list2[36] = def_info

    if attach_sent == [] or new_disj:
        list2[38] = sentences
        attach_sent.append(list2)
    else:
        list2[38] = sentences
        list2[2] = attach_sent[d][2]
        list2[37] = attach_sent[d][37]
        attach_sent[d] = list2


def proper_spacing(str1):
    str1 = str1.replace(" ", "")
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional, " " + conditional + " ")
    str1 = str1.replace(idisj, " " + idisj + " ")
    str1 = str1.replace(xorr, " " + xorr + " ")
    str1 = str1.replace("&", " & ")
    return str1


def proper_spacing2(str1):
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional, " " + conditional + " ")
    str1 = str1.replace(idisj, " " + idisj + " ")
    str1 = str1.replace(xorr, " " + xorr + " ")
    str1 = str1.replace("&", " & ")
    return str1


def bad_paren(str1):
    if str1.find("(") == -1:
        return str1
    # we first must get rid of strings of the following form ((p) & s)
    for i in range(len(str1)):
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        if i > 1:
            str4 = str1[i - 2:i - 1]
        else:
            str4 = ""
        str5 = str1[i + 1:i + 2]
        if str2.islower() and str3 == "(" and str5 == ")":
            str1 = str1[:i - 1] + str1[i:i + 1] + str1[i + 2:]
        elif str2.islower() and str4 == "(" and str3 == "~" and str5 == ")":
            str1 = str1[:i - 2] + str1[i - 1:i + 1] + str1[i + 2:]

    str1 = enclose(str1)
    list1 = find_sentences(str1)
    mstr = list1[3][0]
    for i in range(1, len(list1[3])):
        if list1[4][i][1] != "":
            mc = list1[4][i][1]
            ostr = list1[3][i]
            str2 = list1[4][i][0][:-1]
            prcnt = findinlist(str2, list1[4], 0, 1)
            if mc == prcnt:
                nstr = remove_outer_paren(ostr)
                nstr = remove_outer_paren(nstr)
                mstr = mstr.replace(ostr, nstr)
    return mstr


def unenclose(str1):
    # this removes ( ) from around a sentence abbreviation
    i = -1
    if "(" not in str1:
        return str1

    while i < len(str1) - 1:
        i += 1
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        str4 = str1[i + 1:i + 2]
        if str2.islower() and str3 != "~" and str4 not in subscripts:
            str1 = str1[:i - 1] + str2 + str1[i + 2:]
        elif str2.islower() and str3 == "~" and str4 not in subscripts:
            str1 = str1[:i - 2] + str3 + str2 + str1[i + 2:]
        if str2.islower() and str3 != "~" and str4 in subscripts:
            str1 = str1[:i - 1] + str2 + str4 + str1[i + 3:]
        elif str2.islower() and str3 == "~" and str4 in subscripts:
            str1 = str1[:i - 2] + str3 + str2 + str4 + str1[i + 3:]

    return str1


def new_disjunct(str1, ng, n, candd, conjt, anc1, anc2, anc3=None, anc4=None, kind=0, rule=""):
    global sn, pn
    list2 = mainconn(str1)
    if kind == 1:
        del attach_sent[n]
        consistent = new_prop(str1, ng, "&I", anc1, anc2, anc3, anc4)
        return consistent
    elif kind == 2:
        consistent = new_prop(str1, ng, "&I", anc1, anc2, anc3, anc4)
        return consistent
    else:
        if os(str1):
            del attach_sent[n]
            str1 = remove_outer_paren(str1)
            list1 = tilde_removal2(str1)
            str1 = list1[0]
            consistent = new_prop(str1, list1[1], rule + "E", anc1, anc2)
            candd.append([pn, list1[0], list1[1]])
            conjt.append([pn, list1[0], list1[1]])
            return consistent
        elif list2[0] == "&":
            del attach_sent[n]
            str1 = remove_outer_paren(str1)
            new_prop(str1, ng, rule + "E", anc1, anc2)
            g = copy.copy(pn)
            list3 = get_conjuncts(str1)
            for i in range(len(list3)):
                list4 = tilde_removal2(list3[i])
                consistent = new_prop(list4[0], list4[1], "&E", g, "")
                if no_output == False:
                    return consistent
                if list3[i].find(idisj) > -1:
                    disjunction_heirarchy(list4[0], n, True)
                else:
                    candd.append([pn, list4[0], list4[1]])
                    conjt.append([pn, list4[0], list4[1]])
            return True
        else:
            consistent = new_prop(str1, ng, idisj + "E", anc1, anc2)
            if consistent == False:
                return consistent
            if ng == "~":
                str1 = ng + str1
            else:
                str1 = remove_outer_paren(str1)
            attach_sent[n][2] = pn
            disjunction_heirarchy(str1, n, False)
            return True


def xorr_elim(n, i, parent, grandparent, whole_d, candd, anc1, anc2, conjt, kind=0):
    str9 = ""
    de_mor = False
    if kind == 0:
        for r in range(len(attach_sent[n][i][4])):
            if r != i:
                if not os(attach_sent[n][i][4][r][0]):
                    de_mor = True
                if str9 == "":
                    str9 += "~" + attach_sent[n][i][4][r][1] + attach_sent[n][i][4][r][0]
                else:
                    str9 += " & ~" + attach_sent[n][i][4][r][1] + attach_sent[n][i][4][r][0]
    else:
        grandp2 = copy.copy(grandparent)
        grandp2 = grandp2.replace(parent, "")
        for r in range(8, 38):
            if attach_sent[n][r] == "":
                break
            if attach_sent[n][r][0][0] in grandp2:
                if str9 == "":
                    str9 += "~" + attach_sent[n][r][0][1] + attach_sent[n][r][0][0]
                else:
                    str9 += " & ~" + attach_sent[n][r][0][1] + attach_sent[n][r][0][0]
    g = copy.copy(pn)
    if parent != grandparent:
        str9 = remove_outer_paren(str9)
        if grandparent == whole_d:
            mc = mainconn(str9)
            if mc[0] == '&':
                consistent = xorr_elim2(str9, candd, conjt, anc1, anc2)
                if consistent == False:
                    return consistent
            else:
                list4 = tilde_removal(str9)
                consistent = new_prop(list4[0], list4[1], xorr + "E", anc1, anc2)
                if consistent == False:
                    return consistent
        else:
            str9 = "(" + str9 + ")"
            if kind == 0:
                str9 = grandparent.replace(parent, str9)
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~", "")
                    consistent = new_prop(str9, "", "~~E", pn, "")

            else:
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
                g = copy.copy(pn)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~", "")
                    consistent = new_prop(str9, "", "~~E", g, "")
                    if consistent == False:
                        return consistent
                disjunction_heirarchy(str9, n, True)
                del attach_sent[n]
            if de_mor:
                list1 = demorgan(all_sent, attach_sent, total_sent, detach_sent, candd, True, str9, pn)
                consistent = list1[0]
                attach_sent = list1[1]
                if consistent == False:
                    return consistent
            else:
                if str9.find(idisj) > -1 or str9.find(xorr) > -1:
                    disjunction_heirarchy(str9, n, True)
                consistent = True
    else:
        # this does not account for the case where the parent == grandparent but
        # grandparent does not == whole d
        consistent = xorr_elim2(str9, candd, conjt, anc1, anc2)
    return consistent


def xorr_elim2(str9, candd, conjt, anc1, anc2):
    str9 = bad_paren(str9)
    consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
    if consistent == False:
        return False
    if str9.find("~~") > -1:
        str9 = str9.replace("~~", "")
        consistent = new_prop(str9, "", "~~E", pn, "")
        if consistent == False:
            return consistent
    list3 = get_conjuncts(str9)
    g = copy.copy(pn)
    for b in range(len(list3)):
        list4 = tilde_removal2(list3[b])
        list4[0] = remove_outer_paren(list4[0])
        consistent = new_prop(list4[0], list4[1], "&E", g, "")
        if consistent == False:
            return consistent
        if not os(list3[b]):
            if list4[1] == "~":
                list1 = demorgan(all_sent, attach_sent, total_sent, detach_sent, candd, list3[b], pn, "&E")
                consistent = list1[0]
                attach_sent = list1[1]
                if consistent == False:
                    return False
            else:
                disjunction_heirarchy(list4[0], n, True)
        else:
            candd.append([pn, list4[0], list4[1]])
            conjt.append([pn, list4[0], list4[1]])

    return True


def disjunction_elimination(candd, kind=""):
    bool1 = False
    bool2 = False
    global sn, pn
    global rel_conj

    for i in range(len(attach_sent)):
        if attach_sent[i][8] == "":
            disjunction_heirarchy(attach_sent[i][4], i)
    i = -1
    conjt = copy.deepcopy(candd)

    while i < len(conjt) - 1:
        i += 1
        if conjt[i][1] not in rel_conj:
            del conjt[i]
            i -= 1
    d = -1
    while d < len(conjt) - 1:
        d += 1
        str2 = conjt[d][2]
        conj = conjt[d][1]
        if conj == 'q':
            bb = 7
        if d == 42:
            bb = 7
        anc1 = conjt[d][0]
        n = -1
        while n < len(attach_sent) - 1:
            if bool1:
                bool1 = False
                d = -1
                break
            n += 1
            if n == 7:
                bb = 7
            i = 7
            while attach_sent != []:
                if bool2:
                    bool2 = False
                    break
                i += 1
                if conj not in attach_sent[n][38] and iff not in attach_sent[n][4] \
                        and conditional not in attach_sent[n][4]:
                    break
                else:
                    if attach_sent[n][i] == "":
                        break
                    whole_d = attach_sent[n][4]
                    anc2 = attach_sent[n][2]
                    str3 = attach_sent[n][i][0][0]
                    # 'pos or neg'
                    str4 = attach_sent[n][i][0][1]
                    # 'disjunct or conjunct
                    str5 = attach_sent[n][i][2]
                    # 'disjunct number
                    str6 = attach_sent[n][i][1]

                    if conj == str3:
                        grandparent = attach_sent[n][i][6]
                        parent = attach_sent[n][i][5]
                        parent2 = copy.copy(parent)
                        parent3 = copy.copy(parent)
                        str7 = " " + idisj + " "
                        str7a = " " + xorr + " "
                        if str2 == str4 and str5 == "d":
                            # 'if the disjuncts are not embedded within a conjunct then the disjunction
                            # is simply deleted

                            del attach_sent[n]
                            if parent != grandparent:
                                conj = str2 + conj
                                str8 = whole_d.replace(parent, conj)
                                disjunction_heirarchy(str8, n)
                            bool1 = True
                            n = -1
                            break

                        elif str2 == str4 and str5 == "x":

                            consistent = xorr_elim(n, i, parent, grandparent, whole_d, candd, anc1, anc2)
                            if not consistent:
                                return False
                            del attach_sent[n]
                            bool2 = True
                            bool1 = True
                            d = -1
                        elif str2 == str4 and str5 == "c":
                            list2 = []
                            list2.append([conj, str2])
                            anc3 = ""
                            anc4 = ""
                            list11 = attach_sent[n][i][4]
                            f = -1
                            while f < len(list11) - 1:
                                mc = mainconn(grandparent)
                                f += 1

                                for e in range(len(candd)):
                                    anc5 = candd[e][0]
                                    # since it's too hard to program, if the sibling is a disjunct then we just
                                    # ignore this

                                    if list11[f][0].find(idisj) > -1 or list11[f][0].find(xorr) > -1:
                                        break
                                    else:
                                        if candd[e][1] == list11[f][0]:
                                            if candd[e][2] == list11[f][1]:
                                                list2.append([list11[f][0], list11[f][1]])
                                                if len(list2) == 2:
                                                    anc3 = anc5
                                                elif len(list2) == 3:
                                                    anc4 = anc5
                                                del list11[f]
                                                if list11 == []:
                                                    str3 = build_sent_list2(list2)
                                                    if mc[0] == xorr:
                                                        new_prop(str3, "", "&I", anc1, anc3, anc4)
                                                        consistent = xorr_elim(n, i, parent, grandparent, whole_d,
                                                                               candd,
                                                                               anc1, anc2, conjt, 1)
                                                        if not consistent:
                                                            return False
                                                    else:
                                                        # if the conjunct is not embedded within another conjunct
                                                        # then the disjunct is simply deleted
                                                        if whole_d == grandparent:
                                                            consistent = new_disjunct(str3, "", n,
                                                                                      candd, conjt, anc1, anc3,
                                                                                      anc4, anc5, 1)

                                                        else:
                                                            str8 = whole_d.replace(grandparent, parent2)
                                                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                                                str8 = bad_paren(str8)
                                                                consistent = new_disjunct(str3, "", n,
                                                                                          candd, conjt, anc1, "",
                                                                                          anc3, anc4, 2)

                                                            consistent = new_disjunct(str8, "", n, candd, conjt, pn - 1,
                                                                                      anc2)
                                                            if not consistent:
                                                                return False
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break
                                                else:
                                                    f -= 1
                                                    break

                                            elif candd[e][2] != list11[f][1]:
                                                mc = mainconn(grandparent)
                                                if mc[0] == idisj:
                                                    rule = idisj
                                                    str7 = " " + idisj + " "
                                                else:
                                                    str7 = " " + xorr + " "
                                                    rule = xorr
                                                r = grandparent.find(parent)
                                                if r > 1:
                                                    parent = str7 + parent
                                                else:
                                                    parent = parent + str7
                                                anc1 = candd[e][0]
                                                str9 = grandparent.replace(parent, "")
                                                str8 = whole_d.replace(grandparent, str9)
                                                if str8.find("(") > -1 and (
                                                                str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                                    str8 = bad_paren(str8)
                                                consistent = new_disjunct(str8, "", n, candd, conjt, anc1, anc2, None,
                                                                          None,
                                                                          0, rule)

                                                if not consistent:
                                                    return False
                                                else:
                                                    list11 = []
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break

                        elif str2 != str4 and str5 == "c":
                            mc = mainconn(attach_sent[n][i][6])
                            if mc[0] == idisj:
                                str6 = str7 + parent
                                rule = idisj
                            else:
                                rule = xorr
                                str6 = str7a + parent
                            if grandparent.find(str6) > -1:
                                parent = str6
                            else:
                                parent = parent + str7
                            str9 = grandparent.replace(parent, "")
                            str8 = whole_d.replace(grandparent, str9)
                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                str8 = bad_paren(str8)
                            consistent = new_disjunct(str8, "", n, candd, conjt, anc1, anc2, None, None, 0, rule)
                            if not consistent:
                                return False
                            bool1 = True
                            n = -1
                            break

                        elif str2 != str4 and (str5 == "d" or str5 == 'x'):
                            # if the disjunct is a triple disjunct then enter below
                            if str5 == 'd':
                                rule = idisj
                            else:
                                rule = xorr
                            if attach_sent[n][i][7] > 1:
                                str6 = str4 + str3 + " " + rule + " "
                                if parent.find(str6) > -1:
                                    str5 = str6
                                else:
                                    str5 = " " + rule + " " + str4 + str3
                                str9 = parent.replace(str5, "")
                                str8 = whole_d.replace(parent, str9)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                consistent = new_disjunct(str8, "", n, candd, conjt,
                                                          anc1, anc2, None, None, 0, rule)

                                if not consistent:
                                    return False
                                bool1 = True
                                n = -1
                                break

                            else:
                                str3 = attach_sent[n][i][4][0][0]  # ddd
                                str4 = attach_sent[n][i][4][0][1]
                                str5 = str4 + str3
                                str8 = whole_d.replace(parent, str5)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                consistent = new_disjunct(str8, "", n, candd,
                                                          conjt, anc1, anc2, None, None, 0, rule)
                                if not consistent:
                                    return False
                                bool1 = True
                                n = -1
                                break
    return True


def use_statement_logic(kind=""):
    global st_log_time
    b = time.time()
    list1 = detach1(kind)
    consistent = list1[0]
    attach_sent = list1[1]
    # if consistent == False:
    #     return [False, attach_sent]
    # if kind != 2:
    #     list1 = disjunction_elimination(all_sent, attach_sent, detach_sent, \
    #                                     candd, total_sent, kind)
    #     consistent = list1[0]
    #     attach_sent = list1[1]

    c = time.time()
    d = c - b
    st_log_time += d
    return consistent


def add_outer_paren(str1):
    str1 = remove_outer_paren(str1)
    return "(" + str1 + ")"

def populate_sentences(p):
    global result_data
    global excel
    bool1 = False
    bool2 = False
    bool3 = True
    first_sent = False
    sent = []
    test_sent = []
    g = 0

    if not excel:

        for row in w4:
            p += 1
            if row[1] == "" and bool2 == True and not bool3:
                break
            elif row[1] == 'stop':
                break
            elif row[1] == "" and not bool3:
                test_sent.append(sent)
                sent = []
                g = 0
                bool1 = False
                bool2 = True
                first_sent = False
            elif row[1] == "":
                pass
            elif row[1][0] == "*":
                bool3 = True
            elif row[1] != "" and bool1 == False:
                bool3 = False
                str2 = row[1]
                g += 1
                if len(sent) == 0:
                    if str2.find(bottom) > -1:
                        tv = 'co'
                        str2 = str2[2:]
                    else:
                        tv = 'ta'
                else:
                    tv = ""
                str2.strip()
                sent.append([g, str2, '', tv])
                if not first_sent:
                    result_data['text_' + str(p - 2) + '_1'] = len(test_sent)
                first_sent = True
                bool2 = False
    else:
        for row in w4.rows:
            p += 1
            if row[2].value == None and bool2 == True and not bool3:
                break
            elif row[2].value == 'stop':
                break
            elif row[2].value == None and not bool3:
                test_sent.append(sent)
                sent = []
                g = 0
                bool1 = False
                bool2 = True
                first_sent = False
            elif row[2].value == None:
                pass
            elif row[2].value[0] == "*":
                bool3 = True
            elif row[2].value != None and bool1 == False:
                bool3 = False
                str2 = row[2].value
                g += 1
                if len(sent) == 0:
                    if str2.find(bottom) > -1:
                        tv = 'co'
                        str2 = str2[2:]
                    else:
                        tv = 'ta'
                else:
                    tv = ""
                str2.strip()
                sent.append([g, str2, row[0].value, tv])
                if not first_sent:
                    w4.cell(row=p - 1, column=2).value = len(test_sent)
                first_sent = True
                bool2 = False

    return [test_sent, p]


def get_number_of_sent_to_prove(strt, stp, order, nonlinear):
    if nonlinear:
        return order
    else:
        return [x for x in range(strt, stp)]


def calculate_time_statistics(st):
    global instan_used, instan_time, lemmas_used

    if nonlinear:
        num_of_sent = len(order)
    else:
        num_of_sent = (stp - strt)

    en = time.time()
    g = (en - st) / num_of_sent
    dd = st_log_time / num_of_sent
    tot_tim2 = time.time()
    total = tot_tim2 - tot_tim
    gg = time_spent_defining / num_of_sent
    if lemmas_used == 0: lemmas_used = 1

    if instan_used != 0:
        ee = instan_time / instan_used
    else:
        ee = 0
    print("average " + str("{0:.4f}".format(g)))
    print("time used in statement logic " + str("{0:.4f}".format(dd)))
    print("time spent reducing " + str("{0:.4f}".format(time_spent_reducing / (num_of_sent))))
    print("time used in lemma function " \
          + str("{0:.5f}".format(time_spent_in_lemma_function / lemmas_used)))
    print("time used in instantiation " + str("{0:.4f}".format(ee)))
    print("time used in change variables function " + str("{0:.4f}".format(gg)))
    # print("time used building_sent " + str("{0:.7f}".format(build_sent_slots_time/build_sent_slots_counter)))
    print("total " + str("{0:.3f}".format(total)))




def get_result(post_data, archive_id=None, request=None):
    global ws, w4, result_data, order, propositional_constants
    global sn, total_sent, prop_name, variable_type
    global all_sent, attach_sent, detach_sent
    global prop_var, variables, stp, abbreviations

    if not normal_proof:
        if archive_id:
            ws = Define3.object_properties.filter(archives_id=archive_id)
        else:
            archive = Archives.object_properties.latest('archives_date')
            ws = Define3.object_properties.filter(archives_id=archive.id)

    if not normal_proof:
        result_data = dict(post_data.iterlists())
        w4 = []
        index = 0
        while True:
            row = (post_data["text_" + str(index) + "_1"], post_data["text_" + str(index) + "_2"],
                   post_data["text_" + str(index) + "_3"])
            w4.append(row)
            if row[1] == "stop":
                break
            index += 1
        w4 = tuple(w4)

    if mysql:
        if archive_id:
            tw4 = Input.object_properties.filter(archives_id=archive_id)
        else:
            archive = Archives.object_properties.latest('archives_date')
            tw4 = Input.object_properties.filter(archives_id=archive.id)
        w4 = []
        for x in tw4:
            row = (x.col1, x.col2, x.col3)
            w4.append(row)
        w4 = tuple(w4)

    list1 = pop_sent('hey')
    for i in range(len(list1)):
        for j in range(len(list1[i])):
            list2 = tran_str(list1[i][j][1], 2)
            list1[i][j][1] = list2[0]
    test_sent = list1
    p = 1
    if not get_words_used:
        ex_dict = large_dict()
    else:
        ex_dict = ""

    build_dict(ex_dict)
    if stp == 0: stp = len(test_sent)
    not_oft_def = copy.deepcopy(dictionary[6])
    order = get_number_of_sent_to_prove(strt, stp, order, nonlinear)

    st = time.time()

    if not excel and not normal_proof:
        views.progressbar_send(request, 0, 100, 0, 1)
    for k in order:
        if not excel and not normal_proof:
            views.progressbar_send(request, strt, stp, k, 1)
        if k == 27:
            bb = 7
        st1 = time.time()
        prop_name = []
        total_sent = []
        all_sent = []
        attach_sent = []
        detach_sent = []
        dictionary[6] = not_oft_def
        variable_type = [[], [], []]
        abbreviations = [{}, {}]
        propositional_constants = {}
        prop_var = copy.deepcopy(prop_var4)
        variables = copy.deepcopy(variables2)
        sn = test_sent[k][-1][0] + 1

        step_one(test_sent[k])

        step_two()

        consistent = step_three(test_sent[k][0][3])

        test_sent[k] = copy.deepcopy(total_sent)
        tot_prop_name.append(prop_name)

        if not consistent:
            print('False')
            stp = k + 1
            # break

        print(str(k) + " - " + str("{0:.3f}".format(time.time() - st1)))

    calculate_time_statistics(st)

    print_sent_full(test_sent, tot_prop_name)

    if django2:
        views.progressbar_send(request, 0, 100, 100, 2)
    if excel:
        pass  # Saved at last
    elif mysql:
        views.save_result(result_data)
    else:
        return result_data


get_result('hey')

if print_to_doc:
    wb4.save('/Users/kylefoley/Desktop/inference engine/temp_proof.xlsx')
if get_words_used:
    wb5.save('/Users/kylefoley/Desktop/inference engine/dictionary4.xlsx')