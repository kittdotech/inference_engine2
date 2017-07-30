from openpyxl import load_workbook
from collections import Counter
import copy
import time
import operator
import sys
from ex_dict_7_3 import large_dict
from claims_7_3 import pop_sent
from pprint import pprint
import collections
from start_and_stop import info

# checked coverage up to line 4777

# averaged .076 on 5.22 (proof type 'n'), .039 definitions, .004 statement logic
# averaged .059 on 5.22 proof type 'n', .023 definitions, .004 statement
# but just prior to that the speed was .066
# time spent in instantiation is .029

# on 6/8 time spent in instantiation = .009, .014
# on 6/10 time spent in instantiation = .018, definitions = .031, total .074

# on 6/26 average .026 (up to instantiation) definitions: .019
# trial 2, .020, .025, trial 3: same as 2, total 2.562

# with a lot of globals: 2.831, 3.330, 2.934

# 7/2 average .050, definitions .029, statement .0015, instantiation .010

# 7/11 time spent reducing .0098, average .047, change of variables function .025, step two .032




tot_tim = time.time()

strt, stop, print_to_doc, get_words_used, order, nonlinear = info()

excel = False
mysql = False
django2 = False
normal_proof = True
wb4 = load_workbook('/Users/kylefoley/Desktop/inference engine/temp_proof.xlsx')
w4 = wb4.worksheets[0]

if get_words_used:
    wb5 = load_workbook('/Users/kylefoley/Desktop/inference engine/dictionary4.xlsx')
    ws = wb5.worksheets[0]











########################
# new code




if mysql:
    from inference2.models import Define3, Archives, Input
    from inference2 import views
if mysql:
    import os

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    print(BASE_DIR)
    sys.path.append(BASE_DIR)
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "inference_engine2.settings")
    import django

    django.setup()
    from inference2 import views
    from inference2.models import Define3, Archives, Input

total_sent = []
all_sent = []
attach_sent = []
detach_sent = []
prop_name = []
prop_var = []
words = []
variables = []
abbreviations = []
words_used = []
def_used = []

already_defined = []

tot_prop_name = []
result_data = {}
build_sent2_time = 0
build_sent_counter = 0
time1 = 0
st_log_time = 0
inst_tim = 0
instan_used = 0  # the number of times the instan function is used
instan_time = 0  # measures the time used in instantiation
lemmas_used = 0
time_spent_in_lemma_function = 0
time_spent_reducing = 0
time_spent_defining = 0

cond_r = chr(8835)
const = "\u2102"  # consistency
top = chr(8868)
bottom = chr(8869)
neg = chr(172)
idd = chr(8781)  # translation symbol
iff = chr(8801)
mini_c = chr(8658)
mini_e = chr(8703)
implies = chr(8866)
conditional = chr(8594)
nonseq = chr(8876)
xorr = chr(8891)
idisj = chr(8744)
cj = chr(8896)
aid = chr(8776)
disj = chr(8855)
equi = chr(8660)
ne = "\u2260"  # not equal



l1 = "\u2081"
l2 = "\u2082"
l3 = "\u2083"
l4 = "\u2084"  # if you increase to l6 then change convert function
l5 = "\u2085"
ua = "\u1d43"
ub = "\u1d47"
uc = "\u1d9c"
ud = "\u1d48"
ue = "\u1d49"
uf = "\u1da0"
ug = "\u1d4d"
ui = "\u2071"
uk = "\u1d4f"
um = "\u1d50"
un = "\u207f"
uo = "\u1d52"
up = "\u1d56"
ut = "\u1d57"
uv = "\u1d5b"
uu = "\u1d58"
uw = "\u02b7"
uy = "\u02b8"
uj = "\u02B2"
ul = "\u02E1"
ur = "\u02b3"
us = "\u02e2"
uh = "\u02b0"

prop_var4 = [chr(97 + t) for t in range(26)]
prop_var2 = [chr(97 + t) + "\u2081" for t in range(26)]
prop_var3 = [chr(97 + t) + "\u2082" for t in range(26)]
prop_var5 = [chr(97 + t) + "\u2083" for t in range(26)]
prop_var6 = [chr(97 + t) + "\u2084" for t in range(26)]
prop_var7 = [chr(97 + t) + "\u2085" for t in range(26)]
prop_var4 = prop_var4 + prop_var2 + prop_var3 + prop_var5 + prop_var6 + prop_var7
variables2 = [chr(122 - t) for t in range(25)]
variables2.remove("i")
variables2.remove("l")
variables3 = [chr(122 - t) + l1 for t in range(25)]
variables4 = [chr(122 - t) + l2 for t in range(25)]
variables2 = variables2 + variables3 + variables4
p = 1
subscripts = [l1, l2, l3, l4]
alpha = chr(945)
beta = chr(946)


#
# >> 8835
# ta^ 8868
# co^ 8869
# nt+ 172
# x^ 8801
# c^ 8658
# # 8703
# i^ 8866
# t^ 8594
# nf^ 8876
# ed^ 8891
# v+ 8744
# && 8896
# @ 8855
# if^ 8660


def tran_str(str1, type3):
    list2 = []
    str2 = ""
    if 'co^' in str1:
        str1 = str1.replace('co^ ', "")
        str2 = 'co'

    if "|" in str1:
        for i in range(len(str1)):
            if str1[i:i + 1] == "|":
                str3 = str1[i + 1:i + 2]
                str4 = get_super(str3)
                str1 = str1[:i] + str4 + str1[i + 2:]
                bb = 8

    if type3 == 3:

        if "t^" in str1:
            str1 = str1.replace("t^", conditional)
        if "nt+" in str1:
            str1 = str1.replace("nt+", neg)
        if "zzz" in str1:
            str1 = str1.replace("zzz", ne)
        if "x^" in str1:
            str1 = str1.replace("x^", iff)
        if "b^" in str1:
            str1 = str1.replace("b^", mini_e)
        if "c^" in str1:
            str1 = str1.replace("c^", mini_c)
        if "ed^" in str1:
            str1 = str1.replace("ed^", xorr)
        if "v+" in str1:
            str1 = str1.replace("v+", idisj)
    if type3 == 1:
        list2 = str1.split(" % ")
    else:
        list2 = str1

    return [list2, str2]


def get_super(str1):
    if str1 == "a":
        return "\u1d43"
    elif str1 == "b":
        return "\u1d47"
    elif str1 == "c":
        return "\u1d9c"
    elif str1 == "d":
        return "\u1d48"
    elif str1 == "e":
        return "\u1d49"
    elif str1 == "f":
        return "\u1da0"
    elif str1 == "g":
        return "\u1d4d"
    elif str1 == "h":
        return "\u02b0"
    elif str1 == "i":
        return "\u2071"
    elif str1 == "j":
        return "\u02B2"
    elif str1 == "k":
        return "\u1d4f"

        # ua = u"\u1d43"
    # ub = u"\u1d47"
    # uc = u"\u1d9c"
    # ud = u"\u1d48"
    # ue = u"\u1d49"
    # uf = u"\u1da0"
    # ug = u"\u1d4d"
    # uh = u"\u02b0"
    # ui = u"\u2071"
    # uj = u"\u02B2"
    # uk = u"\u1d4f"
    # ul = u"\u02E1"
    # um = u"\u1d50"
    # un = u"\u207f"
    # uo = u"\u1d52"
    # up = u"\u1d56"
    # ur = u"\u02b3"
    # us = u"\u02e2"
    # ut = u"\u1d57"
    # uu = u"\u1d58"
    # uv = u"\u1d5b"
    # uw = u"\u02b7"
    # uy = u"\u02b8"

    elif str1 == "l":
        return "\u02E1"
    elif str1 == "m":
        return "\u1d50"
    elif str1 == "n":
        return "\u207f"
    elif str1 == "o":
        return "\u1d52"
    elif str1 == "p":
        return "\u1d56"
    elif str1 == "r":
        return "\u02b3"
    elif str1 == "s":
        return "\u02e2"
    elif str1 == "t":
        return "\u1d57"
    elif str1 == "u":
        return "\u1d58"
    elif str1 == "v":
        return "\u1d5b"
    elif str1 == "w":
        return "\u02b7"
    elif str1 == "y":
        return "\u02b8"


def remove_outer_paren(str1, bool1=False):
    if str1 == "":
        return ""
    elif str1.count(")") == 0:
        if not bool1:
            return str1
        else:
            return False

    j = 0
    # on very rare occasions we will encounter strings of the following form ((p))
    if str1[0] != "(" and str1[-1] != ")":
        if not bool1:
            return str1
        else:
            return True
    if str1[:2] == "((" and str1[-2:] == "))":
        d = 2
    else:
        d = 1

    for k in range(0, d):
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
                if bool1:
                    return True
    if not bool1:
        return str1
    else:
        return False


def remove_redundant_paren(str1):
    j = 0
    str2 = str1[:2]
    str3 = str1[2:]
    if str2 == '((' and str3 == '))':

        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 1 and i + 1 != len(str1):
                break
            elif j == 0 and i + 1 == len(str1):
                str1 = str1[1:len(str1) - 1]
    return str1


def mainconn(str1):
    ostring = copy.copy(str1)
    if os(str1):
        return ["", 0]
    if str1.find("&") < -1 and str1.find(idisj) < -1 and str1.find(iff) < -1 and str1.find(conditional) < -1 and \
                    str1.find(implies) < -1 and str1.find(nonseq) < -1 and str1.find(xorr) < -1:
        return ["", 0]

    str3 = str1
    bool1 = False

    if str1[0] == "~":
        str1 = str1[1:]
        bool1 = True

    j = 0
    bool2 = False
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 0 and i + 1 != len(str1):
            break
        elif j == 0 and i + 1 == len(str1):
            str1 = str1[1:len(str1) - 1]
            bool2 = True

    j = -1
    for i in range(0, len(str1)):
        str2 = str1[i:i + 1]
        if str2 == conditional:
            f = -1
        if str2 == idisj or str2 == "&" or str2 == iff or str2 == implies or \
                        str2 == nonseq or str2 == conditional or str2 == xorr:
            if str3 != str2 and str3 != "":
                j = j + 1

            str3 = str2

    if j == -1:
        i = ostring.find(str3)
        return [str3, i]
    k = -1
    j = -1
    while True:
        k += 1
        if k > 150:
            break
        for i in range(0, len(str1)):
            str2 = str1[i:i + 1]

            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1

            if j == -1 and (str2 == idisj or str2 == "&" or str2 == iff or str2 == implies
                            or str2 == nonseq or str2 == conditional or str2 == xorr):
                if bool1 and bool2:
                    return [str2, i + 2]
                elif bool2 or bool1:
                    return [str2, i + 1]
                else:
                    return [str2, i]
        else:
            str1 = str1[1:-1]


def isvariable(str3):
    bool2 = True
    if str3 == None or str3 == "":
        return False

    if str3 == 'a':
        return False
    elif str3 == 'i':
        return False

    if str3 != "":
        str3 = str3.replace(l1, "")
        str3 = str3.replace(l2, "")
        str3 = str3.replace(neg, "")
        if len(str3) == 1 and str3.islower():
            bool2 = True
        else:
            bool2 = False


    return bool2


def os(str1):
    cnx = [xorr, iff, idisj, conditional, implies, nonseq, "&"]
    for i in range(0, len(cnx)):
        if str1.find(cnx[i]) > -1:
            os = False
            return os
    os = True
    return os


def enclose(str1):
    i = -1
    global subscripts
    while i < len(str1) - 1:
        i += 1
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        str4 = str1[i + 1:i + 2]
        if str2.islower() and str4 in subscripts:
            if str3 == "~":
                str1 = str1[:i - 1] + "(~" + str2 + str4 + ")" + str1[i + 2:]
            else:
                str1 = str1[:i] + "(" + str2 + str4 + ")" + str1[i + 2:]
            i += 4
        elif str2.islower():
            if str3 == "~":
                str1 = str1[:i - 1] + "(~" + str2 + ")" + str1[i + 1:]
            else:
                str1 = str1[:i] + "(" + str2 + ")" + str1[i + 1:]
            i += 3
    return str1


def is_natural_language(sentence, i):
    # this determines if the sentence is natural or an abbreviation
    if sentence[i + 1] == "~":
        if sentence[i + 3] in subscripts and sentence[i + 4] == ")":
            return True
        elif sentence[i + 3] == ")":
            return True
    else:
        if sentence[i + 2] in subscripts and sentence[i + 3] == ")":
            return True
        elif sentence[i + 2] == ")":
            return True
    return False


def find_sentences(sentence):
    global subscripts
    if sentence == None: g = 4/0
    if os(sentence): g = 4 / 0
    g = sentence.count('(')
    h = sentence.count(')')
    if g != h:
        print('wrong number of parentheses in sentence:' + sentence)
        g = 4 / 0
    if sentence.startswith("(~g)"):
        bb = 8
    marker = False
    il = -1
    total = -1
    c = -1
    neg_value = []
    str1 = ""
    sent1 = []
    sent_type2 = []
    wneg = []
    output = [None] * 9
    # the skel name list names each single sentence after a greek letter, even if
    # the same sentence appears twice it obtains a different name on the second
    # appearance
    skel_nam = []
    sent_num = []
    if sentence.find("~(") > -1:
        sentence = sentence.replace("~(", "(!")
    if sentence.find(implies) > -1:
        str2 = implies
    elif sentence.find(nonseq) > -1:
        str2 = nonseq
    str3 = mainconn(sentence)
    sentence = sentence.strip()
    str4 = str3[0]
    f = str3[1]
    id_num = []

    id_num.append(["1", str4, f])
    sent_num.append([1, '1', sentence, str4, f])
    str21 = ""
    p = 947
    greek_english_dict = {}
    unenclose_at_end = False
    connectives = ["&", idisj, iff, conditional, nonseq, implies, xorr]
    arr1 = []
    mini_c2 = mini_c + neg
    prt = copy.copy(sentence)
    more_num = [chr(945 + x) for x in range(24)]
    temp_string = mainconn(sentence)
    if sentence.find(implies) > -1:
        str1 = implies
    elif sentence.find(nonseq) > -1:
        str1 = nonseq
    else:
        if temp_string == iff:
            str1 = "bicond"
        elif temp_string == conditional:
            str1 = "cond"
        elif temp_string == "&":
            str1 = "cj"

    sent1.append(sentence)
    neg_value.append("")
    sent_type2.append(str1)
    wneg.append(sentence)
    skel_nam.append(None)

    j = 0
    n = 0
    for i in range(0, len(sentence)):
        str1 = sentence[i:(i + 1)]
        for o in connectives:
            if str1 == o:
                j += 1

    while n < j + 1:

        il += 1
        if il > 15:
            break

        e = 0
        l = len(sentence)
        x = -1
        while x < l - 1:
            x += 1
            temp_string = sentence[x:x + 1]
            if sentence[x:x + 1] == "(":
                if not unenclose_at_end:
                    unenclose_at_end = is_natural_language(sentence, x)

                if marker == False:
                    z = x
                    marker = True

                total += 1
            elif sentence[x: x + 1] == ")":
                total -= 1
                if total == -1:
                    marker = False
                    e += 1
                    c += 1

                    temp_sent = sentence[z: x + 1]
                    if temp_sent == '(bIc)':
                        pp = 7
                    otemp_sent = copy.copy(temp_sent)

                    if (len(sentence) - len(temp_sent)) > 2:
                        if temp_sent in prt and temp_sent in str21 and prt != str21:
                            prtnum = findinlist(str21, sent_num, 2, 1)
                            numb = prtnum + "1"
                        elif temp_sent in prt:
                            prtnum = findinlist(prt, sent_num, 2, 1)
                            numb = prtnum + "1"
                        else:
                            prtnum = ""
                            for bb in range(len(sent_num) - 1, -1, -1):
                                str3 = sent_num[bb][2]
                                if os(temp_sent) and str21 != "":
                                    if str21 == str3 and temp_sent in sent_num[bb][2]:
                                        # this is for those basic molecules for which the same sentence appears
                                        # in the definition several times
                                        prtnum = sent_num[bb][1]
                                        break
                                else:
                                    if temp_sent in sent_num[bb][2]:
                                        prtnum = sent_num[bb][1]
                                        break
                            # if prtnum == "":
                            #     easygui.msgbox('your sentences are not numbered properly')
                            g = len(prtnum) + 1
                            f = 0

                            for bb in range(len(sent_num) - 1, -1, -1):
                                temp_sn = sent_num[bb][1]
                                h = temp_sn[:g - 1]
                                hh = sent_num[bb][0]
                                if g > sent_num[bb][0]:
                                    break
                                if sent_num[bb][0] == g and temp_sn[:g - 1] == prtnum:
                                    f += 1

                            f += 1
                            if f < 10:
                                numb = prtnum + str(f)
                            else:
                                numb = prtnum + more_num[f - 10]

                        prt = temp_sent
                        temp_mc = mainconn(temp_sent)
                        mc = temp_mc[0]
                        str3 = temp_mc[0]
                        g = temp_mc[1]
                        mc_num = temp_mc[1]
                        sent_num.append([len(numb), numb, temp_sent, str3, g])

                        if os(temp_sent):

                            # n counts the number of single sentences
                            n += 1
                            if temp_sent.find("~") > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace("~", "")

                            elif temp_sent.find(mini_c2) > -1:
                                neg_value.append("~")
                                temp_sent = temp_sent.replace(mini_c2, mini_c)

                            elif temp_sent.find(mini_c2) == -1 and temp_sent.find(mini_c) > -1:
                                neg_value.append("")
                            elif temp_sent.find("~") == -1:
                                neg_value.append("")
                            else:
                                break  # stop
                        else:
                            neg_value.append("")

                        sent1.append(temp_sent)
                        wneg.append(otemp_sent)
                        id_num.append([numb, mc, mc_num])

                        if os(otemp_sent):
                            if otemp_sent in greek_english_dict:
                                skel_nam.append(greek_english_dict.get(otemp_sent))
                            else:
                                p += 1
                                greek_english_dict.update({otemp_sent: chr(p)})
                                skel_nam.append(chr(p))
                        else:
                            skel_nam.append(None)
                    else:
                        sentence = sentence[1:len(sentence) - 1]
                        l = len(sentence)
                        x = -1
                        c -= - 1
                        e -= - 1

        total = -1
        marker = False
        w = -1

        if n < j + 1:
            if len(sent1) > w:
                while w + 1 < len(sent1):
                    if w == 13:
                        pp = 7
                    w += 1
                    str21 = sent1[w]
                    if not os(str21) and w != 0:
                        if str21 not in arr1:
                            sentence = str21
                            arr1.append(sentence)
                            break

    for i in range(len(sent1)):
        temp_string = sent1[i]
        if temp_string.find("(!") > -1:
            sent1[i] = sent1[i].replace("(!", "~(")
            wneg[i] = wneg[i].replace("(!", "~(")

    if unenclose_at_end:
        for i in range(len(sent1)):
            sent1[i] = unenclose(sent1[i])
            wneg[i] = unenclose(wneg[i])

    output[0] = sent1
    output[1] = neg_value
    output[2] = sent_type2
    output[3] = wneg
    output[4] = id_num
    output[6] = translate_to_greek(skel_nam, wneg, id_num)
    output[5] = skel_nam[0]

    return output


def translate_to_greek(skel_nam, wneg, id_num):
    for i in range(len(skel_nam)):
        if skel_nam[i] == None:
            to_be_translated = wneg[i]
            for j in range(len(skel_nam) - 1, -1, -1):
                if skel_nam[j] != None:
                    if wneg[j] in to_be_translated and id_num[j][1] == "":
                        to_be_translated = to_be_translated.replace(wneg[j], skel_nam[j])
            skel_nam[i] = to_be_translated

    return skel_nam


def add_to_dv(m, k, str2):
    if isvariable(str2) == False:
        str3 = findinlist(str2, abbreviations, 1, 0)
        if str3 == None:
            telist7 = [variables[0], str2]
            if k == 69 or k == 70:
                all_sent[m][k] = variables[0] + "'s"
            else:
                all_sent[m][k] = variables[0]
            del variables[0]
            abbreviations.append(telist7)
        elif k == 69 or k == 70:
            all_sent[m][k] = str3 + "'s"
        else:
            all_sent[m][k] = str3


def word_sub():

    global sn
    pronouns = words[24]
    num = [4, 5, 13, 14, 17, 18, 22, 26, 30, 34, 35, 36, 51, 52, 63, 64, 65, 67, 69, 70]
    num3 = [8, 12, 49, 50, 51, 52]
    other = ['there', 'it']
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        ant_sent_parts = copy.deepcopy(all_sent[m])
        if all_sent[m][47] != "no word sub":
            bool1 = False
            list4 = copy.deepcopy(all_sent[m][46])
            for i in range(len(all_sent[m][46])):
                k = all_sent[m][46][i]
                if k == 8:
                    bb = 8
                str2 = all_sent[m][k]
                if str2 != None:
                    if str2 == "~":
                        str2 = None
                    elif str2 not in def_used and not str2.isupper():
                        def_used.append(str2)
                if k in num3 and str2 != None:
                    bool1 = True
                    str5 = findinlist(str2, words[16], 0, 1)
                    if k == 12:
                        all_sent[m][8] = str5
                        all_sent[m][12] = None
                    else:
                        all_sent[m][k] = str5
                if k == 69 or k == 70:
                    str2 = str2[:-2]
                    add_to_dv(m, k, str2)
                elif k in num:
                    bool1 = True
                    if str2 != None and str2 not in pronouns and str2 not in other:
                        add_to_dv(m, k, str2)

            if bool1:
                all_sent[m] = build_sent2(all_sent[m])
                con_parts = copy.deepcopy(all_sent[m])
                prepare_att_sent_1_sent(ant_sent_parts, "SUB", "", "e", con_parts)
                all_sent[m][46] = list4


def eliminate_relative_pronouns2(m, i):

    list1 = [None] * 80
    ant_sent_parts = copy.deepcopy(all_sent[m])
    rule = "DE " + all_sent[m][i]
    list2 = eliminate_relative_pronouns3(i, m, list1)
    con_parts2 = copy.deepcopy(build_sent2(all_sent[m]))
    con_parts = copy.deepcopy(build_sent2(list2))
    all_sent.append(list2)
    prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, rule, "e")


def remove_duplicates(list1, i):

    list2 = []
    j = -1
    while j < len(list1) - 1:
        j += 1
        if list1[j][i] in list2:
            del list1[j]
            j -= 1
            print ("remove duplicates used")
        else:
            list2.append(list1[j][i])

    return list1


def step_two():

    global sn, all_sent, time_spent_reducing
    aa = time.time()
    all_sent = remove_duplicates(all_sent, 0)

    def_sent = []
    definite_assignments = []

    defined_numbers = add_abbreviations_to_all_sent()

    #eliminate_possessive_pronouns(all_sent)

    #second_reduction(all_sent)

    eliminate_pronouns(def_sent)

    eliminate_determinatives(def_sent, definite_assignments)

    #eliminate_determinatives2(def_sent)

    eliminate_proper_name_possessives(def_sent, definite_assignments)

    eliminate_and_coordinator1()

    eliminate_adjectives()

    eliminate_concept_instance_apposition()

    eliminate_relative_pronouns1()

    eliminate_that()

    eliminate_possessive_pronouns(def_sent)

    eliminate_possessives1()

    eliminate_as()

    divide_relations1()

    def_sent = eliminate_there1(def_sent)

    eliminate_universals(def_sent)

    time_spent_reducing += (time.time() - aa)

    define_relations_and_concepts(defined_numbers, def_sent)

    add_necessary_conditions_for_concept()





def add_abbreviations_to_all_sent():

    definitions = words[16]
    relations = words[6]
    numbers_def = []
    rarely_defined = copy.deepcopy(words[36])

    for i in range(len(abbreviations)):
        if i == 3:
            bb = 7
        if abbreviations[i][1] in rarely_defined:
            rarely_defined.remove(abbreviations[i][1])
        if not isinmdlist(abbreviations[i][1], relations, 1):
            g = findposinlist(abbreviations[i][1], definitions, 0)
            if g > -1:
                list1 = [None] * 80
                list1[5] = abbreviations[i][0]
                list1[9] = '='
                list1[14] = abbreviations[i][1]
                list1 = build_sent2(list1)
                list1[41] = 1
                list1[46] = [200]
                list1[56] = [200]
                all_sent.append(list1)
                # say you have the word 7 in your claim, then the code will define all numbers
                # down to 0 if you do not have the following code
                numbers_def.append(abbreviations[i][1])

    return numbers_def


def eliminate_pronouns(def_sent):

    universal = ['every', 'no']
    i_defined = False
    pronouns = copy.deepcopy(words[24])


    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        num10 = [5, 14, 18, 22, 26, 30, 34, 63, 64, 65]  # pronouns
        if 10 in all_sent[m][56]:
            for i in num10:
                if i in all_sent[m][46]:
                    str1 = all_sent[m][i]
                    if all_sent[m][0] not in def_sent and str1 not in universal \
                            and str1 in pronouns:
                        if str1 != "i" or not i_defined:
                            change_variables(str1, m, "pronoun", i)
                            if str1 != "i":
                                del all_sent[m]
                                m -= 1
                                break
                            else:
                                i_defined = True


def eliminate_pronouns2():

    pronouns = copy.deepcopy(words[24])
    m = -1
    num10 = [5, 14, 18, 22, 26, 30, 34, 63, 64, 65]
    while m < (len(all_sent)) - 1:
        m += 1
        if 10 in all_sent[m][56]:
            for i in num10:
                if all_sent[m][i] in pronouns and all_sent[m][42] not in do_not_define_again:
                    eliminate_pronouns3(m,i)
                    del all_sent[m]
                    m -= 1
                    break





def set_default_list(str1,list1):

    print ('make sure this does not mess up the abbreviations global')
    g = 4/0
    str2 = findinlist(str1, list1, 1, 0)
    if str2 == None:
        str2 = variables[0]
        del variables[0]
        list1.append([str2, str1])

    return str2


def eliminate_pronouns3(m,i):

    ant_parts = copy.deepcopy(all_sent[m])
    con_parts1 = copy.deepcopy(all_sent[m])

    if all_sent[m][i] == "i" or all_sent[m][i] == 'you':
        connective = "c"
        rule = "NC "
    else:
        connective = "e"
        rule = "DE "
    rule += all_sent[m][i]

    if all_sent[m][i] != 'i':
        new_object = set_default_list(all_sent[m][i], abbreviations)
        con_parts1[i] = new_object
    else:
        new_object = "i"
    con_parts1 = build_sent2(con_parts1)


    if all_sent[m][i] == "it" + up:
        property = set_default_list("sexless", abbreviations)
        con_parts2 = build_sent1(new_object, "", "J", property)
        prepare_att_sent_2_sent(ant_parts, con_parts1, con_parts2, rule, "e")
        append_to_all_sent([con_parts1, con_parts2])
    else:
        abbrev_person = set_default_list("person", abbreviations)

        if all_sent[m][i] == "she":
            property = set_default_list("female", abbreviations)
        elif all_sent[m][i] == 'he':
            property = set_default_list("male", abbreviations)

        con_parts2 = build_sent1(new_object, "", "J", property)
        con_parts3 = build_sent1(new_object, "", "I", abbrev_person)
        append_to_all_sent([con_parts1, con_parts2, con_parts3])
        prepare_att_sent_3_sent(ant_parts, con_parts1, con_parts2, con_parts3, rule, connective)


def eliminate_possessive_pronouns2(m,i):

    dict1 = {"my": "i", "my" + ua: "i", "your": "you", "your" + ua: "you",
             "his": "he", "her": "she",
             "its" + ua: "it"}
    pronoun = dict1.get(all_sent[m][i])
    rule = "DE " + all_sent[m][i]
    ant_parts = copy.deepcopy(all_sent[m])
    con_parts1 = copy.deepcopy(all_sent[m])
    concept_position = 14 if i == 10 else i + 2
    new_object = variables[0]
    del variables[0]
    concept = con_parts1[concept_position]
    con_parts1[concept_position] = new_object
    con_parts1[i] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(new_object,"","I",concept)

    if all_sent[m][i] == "its" + ua:
        relat = "HM"
    elif all_sent[m][i][-1] == ua:
        relat = "W"
    else:
        relat = "OWN"

    con_parts3 = build_sent1(pronoun, "", relat, new_object)
    append_to_all_sent([con_parts1, con_parts2, con_parts3])
    prepare_att_sent_3_sent(ant_parts, con_parts1, con_parts2, con_parts3, rule, "e")


def second_reduction(list1):


    determ = ['the', 'a', 'a' + ua, 'many' + un]
    pronouns = ['i','you','he','she','it'+up,]
    do_not_define_again = []
    num10 = [5, 14, 18, 22, 26, 30, 34, 63, 64, 65]
    num20 = [3, 10, 16, 20, 24, 28, 32]
    definite_assignments = {}
    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        slots = all_sent[m][46][1]
        if 20 in all_sent[m][56]:

            for i in slots:
                if not lies_wi_scope_of_univ_quant(m, i, "") and \
                    all_sent[m][42] not in do_not_define_again:

                    def_sent.append(copy.deepcopy(all_sent[m][0]))
                    definite_assignments = eliminate_determinatives3(m, i, definite_assignments)
                    del all_sent[m]
                    m -= 1





def eliminate_determinatives(def_sent, definite_assignments):

    poss_pro = words[25]
    universal = ['every', 'no']
    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 20 in all_sent[m][56]:
            num20 = [3, 10, 16, 20, 24, 28, 32]  # determiners
            for i in num20:
                if i in all_sent[m][46] and all_sent[m][i] not in poss_pro:
                    str1 = all_sent[m][i]
                    if all_sent[m][0] not in def_sent and str1 not in universal:
                        change_variables(str1, m, "determinative", i, definite_assignments)
                        del all_sent[m]
                        m -= 1
                        break

def eliminate_determinatives2(def_sent):

    determ = ['the','a','a'+ua,'many'+un]
    definite_assignments = {}
    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 20 in all_sent[m][56]:
            num20 = [3, 10, 16, 20, 24, 28, 32]
            for i in num20:
                if all_sent[m][i] in determ and not lies_wi_scope_of_univ_quant(m,i,"") and \
                        all_sent[m][0] not in def_sent:
                    def_sent.append(copy.deepcopy(all_sent[m][0]))
                    definite_assignments = eliminate_determinatives3(m,i,definite_assignments)
                    del all_sent[m]
                    m -= 1





def eliminate_determinatives3(m, i, definite_assignments):

    ant_parts = copy.deepcopy(all_sent[m])
    con_parts1 = copy.deepcopy(all_sent[m])
    con_parts1[i] = None
    concept_position = 14 if i == 10 else i+2
    concept = all_sent[m][concept_position]
    rule = "DE " + all_sent[m][i]

    if all_sent[m][i] == "many" + un:
        eliminate_many_n(m,i,ant_parts,con_parts1,concept_position,concept)
    else:
        new_object = definite_assignments.get(concept)

        if all_sent[m][i] == 'the':
            property = 'definite'
            if new_object == None:
                definite_assignments.update({concept:variables[0]})
        elif all_sent[m][i] == 'a' or all_sent[m][i] == "a"+ua:
            property = 'indefinite'

        if new_object == None:
            new_object = variables[0]
            del variables[0]

        abbreviated_property = findinlist(property, abbreviations,1,0)
        if abbreviated_property == None:
            abbreviated_property = variables[0]
            del variables[0]
            abbreviations.append([abbreviated_property,property])

        con_parts1[concept_position] = new_object
        con_parts1 = build_sent2(con_parts1)
        con_parts2 = build_sent1(new_object, "", "I", concept)
        con_parts3 = build_sent1(new_object, "", "J", abbreviated_property)
        append_to_all_sent([con_parts1,con_parts2,con_parts3])

        prepare_att_sent_3_sent(ant_parts,con_parts1,con_parts2,con_parts3,rule,"e")

    return definite_assignments


def append_to_all_sent(list1):

    for sent in list1:
        all_sent.append(sent)

def get_negative_position(i):
    # when we define many we have to put the negation sign in a weird position if
    # many is in an unusual place
    if i == 5 or i == 14:
        j = 8
    elif i == 63 or i == 18:
        j = 49
    elif i == 64 or i == 22:
        j = 50
    elif i == 65:
        j = 51


def eliminate_many_n(m,i,ant_parts,con_parts1,concept_position,concept):

    con_parts2 = copy.deepcopy(all_sent[m])
    con_parts2[i] = None
    new_object = variables[0]
    del variables[0]
    new_object2 = variables[0]
    del variables[0]

    con_parts1[concept_position] = new_object
    con_parts2[concept_position] = new_object2
    b = get_negative_position(i)
    con_parts2[b] = "~"
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent2(con_parts2)
    con_parts3 = build_sent1(new_object, "", "I", concept)
    con_parts4 = build_sent1(new_object2, "", "I", concept)
    rule = "DE " + all_sent[m][i]
    append_to_all_sent([con_parts1, con_parts2, con_parts3, con_parts4])

    prepare_att_sent_4_sent(ant_parts, con_parts1, con_parts2, con_parts3, con_parts4, rule, "e")



def build_sent1(subj, tvalue, relat, obj):

    list1 = [None] * 80
    " ~ " if tvalue == "~" else " "
    list1[5] = subj
    list1[8] = tvalue
    list1[9] = relat
    list1[14] = obj
    sent = "(" + subj + tvalue + relat + " " + obj + ")"
    sent_abs = "(" + subj + " " + relat + " " + obj + ")"
    abbrev_sent = name_sent(sent_abs)
    "~" if tvalue == " ~ " else ""
    list1[0] = sent
    list1[72] = sent_abs
    list1[1] = abbrev_sent
    list1[2] = tvalue
    list1[42] = tvalue + abbrev_sent
    list1[46] = [5,9,14]
    list1[56] = [200]

    return list1


def eliminate_proper_name_possessives(def_sent, definite_assignments):

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 30 in all_sent[m][56]:
            num30 = [69, 70]  # proper name possessive
            for i in num30:
                if i in all_sent[m][46]:
                    if all_sent[m][0] not in def_sent:
                        if i == 69:
                            i = 5
                        elif i == 70:
                            i = 14
                        change_variables("the", m, "proper name possessive", i, definite_assignments)
                        del all_sent[m]
                        m -= 1
                        break


def eliminate_proper_name_possessives2(def_sent):

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 30 in all_sent[m][56]:
            num30 = [69, 70]  # proper name possessive
            for i in num30:
                if all_sent[m][i] != None and all_sent[m][0] not in def_sent:
                    eliminate_proper_name_possessives3(m, i)
                    del all_sent[m]
                    m -= 1
                    break


def eliminate_proper_name_possessives3(m, i):

    if i == 69:
        concept_position = 5
    elif i == 70:
        concept_position = 14
    ant_parts = copy.deepcopy(all_sent[m])
    con_parts1 = copy.deepcopy(all_sent[m])
    concept = all_sent[m][concept_position]

    new_object = variables[0]
    del variables[0]

    con_parts1[concept_position] = new_object
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(new_object, "", "I", concept)
    con_parts3 = build_sent1(new_object, "", "J", "definite")
    append_to_all_sent([con_parts2, con_parts3])

    prepare_att_sent_3_sent(ant_parts, con_parts1, con_parts2, con_parts3, "PNP", "e")

    ant_parts = copy.deepcopy(con_parts1)
    possessor = con_parts1[i][:-2]
    con_parts1[i] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(possessor, "", "OWN", new_object)
    append_to_all_sent([con_parts1, con_parts2])

    prepare_att_sent_2_sent(ant_parts, con_parts1, con_parts2, "PNE", "e")


    # if concept[-2:] is "'s":
    #     apostrophe_s = "'s"
    #     concept = concept[:-2]
    #     rule = "PNE"


def eliminate_common_name_possessives(m,i):

    # if "the" is followed by a common name possessive then it cannot be eliminated

    if i == 69:
        concept_position = 5
        determ_position = 3
    elif i == 70:
        concept_position = 14
        determ_position = 10
    ant_parts = copy.deepcopy(all_sent[m])
    con_parts1 = copy.deepcopy(all_sent[m])
    concept = all_sent[m][concept_position]

    new_object = set_default_list(concept,abbreviations)
    possessor_concept = all_sent[m][i][:-2]
    new_possessor = set_default_list(possessor_concept, abbreviations)

    con_parts1[concept_position] = new_object
    con_parts1[i] = new_possessor + "'s"
    con_parts1[determ_position] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(new_object, "", "I", concept)
    con_parts3 = build_sent1(new_possessor, "", "I", possessor_concept)
    append_to_all_sent([con_parts2, con_parts3])

    prepare_att_sent_3_sent(ant_parts, con_parts1, con_parts2, con_parts3, "DF the", "e")

    ant_parts = copy.deepcopy(con_parts1)
    possessor = con_parts1[i][:-2]
    con_parts1[i] = None
    con_parts1 = build_sent2(con_parts1)
    con_parts2 = build_sent1(possessor, "", "OWN", new_object)
    append_to_all_sent([con_parts1, con_parts2])

    prepare_att_sent_2_sent(ant_parts, con_parts1, con_parts2, "PNE", "e")


def eliminate_and_coordinator1():

    # todo None is a member of compound
    compound = words[34]
    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if all_sent[m][66] != None and all_sent[m][9] not in compound:
            eliminate_and_coordinator2(m)
            del all_sent[m]
            m -= 1


def eliminate_and_coordinator2(m):
    # this seperates a sentence with an 'and' coordinator into two

    ant_sent_parts = copy.deepcopy(all_sent[m])
    all_sent[m][66] = None
    list1 = [None] * 80
    list1[5] = all_sent[m][67]
    all_sent[m][67] = None
    for i in range(6, 20):
        list1[i] = all_sent[m][i]
    list1 = new_categories(list1, True)
    all_sent[m] = build_sent2(all_sent[m])
    con_parts = copy.deepcopy(list1)
    con_parts2 = copy.deepcopy(all_sent[m])
    all_sent.append(list1)
    prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, "DE and" + uc, "e")


def eliminate_adjectives():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 50 in all_sent[m][56]:
            num50 = [4, 13, 17, 21, 25, 33]  # adjective
            for i in num50:
                if i in all_sent[m][46] and lies_wi_scope_of_univ_quant(m, i):
                    eliminate_adjectives2(m, i)
                    break  # this only works for one adjective


def eliminate_adjectives2(m, i):

    list1 = [None] * 80
    ant_sent_parts = copy.deepcopy(all_sent[m])
    if i == 13:
        n = 10
        r = 9
    else:
        r = i - 2
        n = i - 1
    if all_sent[m][8] != None or all_sent[m][12] != None:
        str7 = "~"
        all_sent[m][8] = None
        all_sent[m][12] = None
    else:
        str7 = None
    list1[8] = str7
    list1[3] = all_sent[m][n]
    if all_sent[m][r] != "I":
        list1[5] = all_sent[m][i + 1]
    else:
        list1[5] = all_sent[m][5]
    list1[9] = "J"
    list1[14] = all_sent[m][i]
    list1[44] = chr(965)
    list1[46] = [200]
    list1[56] = [200]
    all_sent[m][i] = None
    list2 = all_sent[m][46]
    list2.remove(i)
    all_sent[m][46] = list2
    all_sent[m] = new_categories(all_sent[m])
    con_parts = copy.deepcopy(build_sent2(list1))
    con_parts2 = copy.deepcopy(all_sent[m])
    all_sent.append(list1)
    prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, "ADJ E", "e")


def eliminate_concept_instance_apposition():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if (36 in all_sent[m][46] or 35 in all_sent[m][46]):
            num60 = [35, 36]  # cia
            for i in num60:
                if all_sent[m][i] != None:
                    eliminate_concept_instance_apposition2(m, i)


def eliminate_concept_instance_apposition2(m, i):

    ant_sent_parts = copy.deepcopy(all_sent[m])
    list1 = [None] * 80
    if i == 35:
        j = 5
    elif i == 36:
        j = 14
    elif i == 37:
        j = 18
    elif i == 38:
        j = 22
    str1 = all_sent[m][j]
    all_sent[m][j] = all_sent[m][i]
    list1[14] = str1
    list1[5] = all_sent[m][i]
    list1[9] = "I"
    list1[46] = [200]
    list1[56] = [200]
    all_sent[m][i] = None
    con_parts = copy.deepcopy(build_sent2(list1))
    all_sent[m] = build_sent2(all_sent[m])
    con_parts2 = copy.deepcopy(all_sent[m])
    all_sent.append(list1)
    prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, "CIA", "e")


def eliminate_relative_pronouns1():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 70 in all_sent[m][56]:
            num70 = [59, 60, 61, 62]  # relative pronouns
            for i in num70:
                if i in all_sent[m][46] and lies_wi_scope_of_univ_quant(m, i, 1) and \
                                all_sent[m][i] != 'that' + uc:
                    eliminate_relative_pronouns2(m, i)
                    break


def eliminate_that():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 80 in all_sent[m][56]:
            num80 = [62, 61, 60, 7]  # that-c
            for i in num80:
                if i in all_sent[m][46] and lies_wi_scope_of_univ_quant(m, i, 1):
                    eliminate_that2(m, i)
                    del all_sent[m]
                    m -= 1
                    break


def eliminate_possessive_pronouns(def_sent):

    possessive_pronouns = words[25]
    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 85 in all_sent[m][56]:
            num85 = [3, 10, 16, 20, 24, 28, 32]
            for i in num85:
                if all_sent[m][i] != None and all_sent[m][i] in possessive_pronouns:
                    str1 = all_sent[m][i]
                    if all_sent[m][0] not in def_sent:
                        change_variables(str1, m, "poss pro", i)
                        del all_sent[m]
                        m -= 1
                        break



def eliminate_possessives1():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 90 in all_sent[m][56]:
            num90 = [69, 70]
            for i in num90:
                if i in all_sent[m][46]:
                    eliminate_possessives2(m, i)
                    del all_sent[m]
                    m -= 1
                    break


def eliminate_as():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if all_sent[m][15] == "AS":
            con_parts1 = [None] * 80
            ant_sent_parts = copy.deepcopy(all_sent[m])
            con_parts1[5] = all_sent[m][5]
            con_parts1[9] = all_sent[m][9]
            con_parts1[14] = all_sent[m][14]
            con_parts1[46] = [5,9,14]
            con_parts1[56] = [200]
            con_parts1 = build_sent2(con_parts1)
            con_parts2 = [None] * 80
            con_parts2[5] = all_sent[m][18]
            con_parts2[9] = all_sent[m][9]
            con_parts2[14] = all_sent[m][14]
            con_parts2[46] = [5, 9, 14]
            con_parts2[56] = [200]
            con_parts2 = build_sent2(con_parts2)
            prepare_att_sent_2_sent(ant_sent_parts, con_parts1, con_parts2, "DE AS", "e")
            del all_sent[m]
            m -= 1
            all_sent.append(con_parts1)
            all_sent.append(con_parts2)


def divide_relations1():

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 100 in all_sent[m][56]:
            num100 = [15, 19]
            for i in num100:
                if i in all_sent[m][46] and uni_scope_rel(m, i):
                    divide_relations2(m, i)
                    break


def divide_relations2(m, i):

    part_of_speech = words[28]
    genre = 1
    ant_sent_parts = copy.deepcopy(all_sent[m])
    list1 = [None] * 80
    str2 = findinlist(all_sent[m][i], part_of_speech, 0, 2)
    if str2 == 'o':
        rule = "RDC"
        list3 = [None] * 80
        a = 14
        list3[8] = all_sent[m][8]
        list3[3] = all_sent[m][10]
        list3[5] = all_sent[m][5]
        list3[9] = all_sent[m][i]
        list3[10] = all_sent[m][16]
        list3[14] = all_sent[m][18]
        list3 = new_categories(list3, True)
        all_sent[m][56] = list3[56]  # todo I'm not sure if this serves a purpose
        all_sent.append(list3)
        genre = 2
        con_parts2 = copy.deepcopy(list3)
    else:
        rule = "RDA"
        a = 5

    if i == 15:
        d = 16
        c = 18
    elif i == 19:
        d = 20
        c = 22

    list1[8] = all_sent[m][8]
    list1[3] = all_sent[m][3]
    list1[5] = all_sent[m][a]
    list1[9] = all_sent[m][i]
    list1[10] = all_sent[m][d]
    list1[14] = all_sent[m][c]
    list1[46] = [200]
    list1[56] = [200]
    if 100 in all_sent[m][56]: all_sent[m][56].remove(100)
    con_parts = copy.deepcopy(build_sent2(list1))
    all_sent.append(list1)

    if genre == 1:
        list6 = all_sent[m][46]
        list6.remove(i)
        list6.remove(c)
        if all_sent[m][8] != None:
            list6.remove(8)
        all_sent[m][46] = list6
        all_sent[m][i] = None
        all_sent[m][c] = None
        all_sent[m][d] = None
        all_sent[m][8] = None
        con_parts2 = copy.deepcopy(build_sent2(all_sent[m]))

    prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, rule, "e")


def eliminate_there1(def_sent):

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 110 in all_sent[m][56]:
            num110 = [5, 63, 64]  # there
            for i in num110:
                if all_sent[m][i] == 'there':
                    def_sent = eliminate_there2(m, def_sent)
                    del all_sent[m]
                    m -= 1
                    break

    return def_sent


def eliminate_there2(m, def_sent):

    ant_sent_parts = copy.deepcopy(all_sent[m])
    list1 = copy.deepcopy(all_sent[m])
    list1[5] = all_sent[m][14]
    list1[3] = all_sent[m][10]
    list1[4] = all_sent[m][13]
    list1[14] = None
    list1[10] = None
    list1[13] = None
    con_parts = copy.deepcopy(build_sent2(list1))
    def_sent.append(con_parts[0])
    all_sent[m][46] = [200]
    all_sent[m][56] = [200]
    all_sent.append(list1)
    bool1 = is_in_md(total_sent, 1, con_parts[0])
    if not bool1:
        prepare_att_sent_1_sent(ant_sent_parts, "DE there", "", "e", con_parts)

    return def_sent


def eliminate_universals(def_sent):

    m = -1
    while m < (len(all_sent)) - 1:
        m += 1
        if 120 in all_sent[m][56]:
            num120 = [3, 10, 16, 20, 24, 28, 32]
            for i in num120:
                if i in all_sent[m][46]:
                    str1 = all_sent[m][i]
                    if all_sent[m][0] not in def_sent:
                        change_variables(str1, m, "determinative", i)
                        del all_sent[m]
                        m -= 1
                        break

#ddd

def eliminate_universals3(i,m):

    ant_parts = copy.deepcopy(all_sent[m])
    rule = "DE " + all_sent[m][i]

    if has_scope_over_subclause(i,m):
        antecedent = extract_words_from_subclause2(i, m)
    elif has_scope_over_adjective(i,m):
        antecedent = turn_adj_into_sent(i,m)
    elif all_sent[m][3] == 'no' and all_sent[m][10] == 'a':
        pass
    else:
        consequent = copy.deepcopy(all_sent[m])
        concept_position = 14 if i == 10 else i + 2
        consequent[i] = None
        concept = all_sent[m][concept_position]
        new_object = variables[0]
        del variables[0]
        antecedent = build_sent1(new_object, "", "I", concept)
        consequent[concept_position] = new_object
        consequent = build_sent2(consequent)



def reduce_subclause(subclause):

    pass






def has_scope_over_adjective(i,m):

    k = 13 if i == 10 else i + 1
    if i == 3 and all_sent[m][k] != None:
        return True
    return False

def turn_adj_into_sent(i,m):

    pass



def has_scope_over_subclause(i,m):

    if i == 3 and all_sent[m][59] != None or \
    i == 10 and all_sent[m][60] != None or \
    i == 16 and all_sent[m][61] != None or \
    i == 20 and all_sent[m][62] != None:
        return True
    return False


def extract_words_from_subclause2(i, m):
    # this is the sentence that will be inserted into the antecedent in the
    # definition of 'every' or 'no', in the future it should output
    # a set of lists, not just 1

    num2 = allowable_slots()
    comma = num2[-1] if all_sent[m][39] == None else all_sent[m][39]
    list1 = [None] * 80
    begin_here = False
    k = 2
    for j in num2:
        if j == i:
            begin_here = True
        if begin_here:
            if comma == j:
                break
            if all_sent[m][j] != None and all_sent[m][j] != "":
                k += 1
                list1[k] = all_sent[m][j]
                all_sent[m][j] = None
    list1 = categorize_words(list1)
    list1 = build_sent2(list1)

    list1[44] = chr(970)

    return list1




def eliminate_sent_wi_univ_scope4(definiendum, determ_loc, k, m, new_var2):
    univ = ['every', 'no']
    special_set = ['a', 'many' + up, 'many' + us, 'many' + ud, "a" + ua]
    bool1 = False
    univ_quant_sent = []
    if definiendum in univ or definiendum in special_set:
        if definiendum in special_set or definiendum in univ:
            if all_sent[m][determ_loc - 1] != None:
                list1 = eliminate_adjective_wi_universal_sent(m, determ_loc, new_var2, 0)
                univ_quant_sent.append(list1)
        if definiendum in univ:
            if k == 3 and all_sent[m][59] != None:
                bool1 = True
                all_sent[m][59] = None
            elif k == 10 and all_sent[m][60] != None:
                bool1 = True
                all_sent[m][60] = None
            elif k == 16 and all_sent[m][61] != None:
                bool1 = True
                all_sent[m][61] = None
            elif k == 20 and all_sent[m][62] != None:
                bool1 = True
                all_sent[m][62] = None
            if bool1:
                list2 = eliminate_univ_quant_subclause(m, k, determ_loc, new_var2)
                for i in range(len(list2)):
                    univ_quant_sent.append(list2[i])
            if all_sent[m][15] != None:
                new_relat = all_sent[m][9]
                new_obj = all_sent[m][14]
                all_sent[m][9] = all_sent[m][15]
                all_sent[m][15] = None
                all_sent[m][10] = all_sent[m][16]
                all_sent[m][16] = None
                all_sent[m][14] = all_sent[m][18]
                all_sent[m][18] = None
                if new_relat != 'EX':
                    list1 = eliminate_adjective_wi_universal_sent2(new_var2, new_relat, new_obj)
                    univ_quant_sent.append(list1)

    return univ_quant_sent

def define_relations_and_concepts(numbers_def, def_sent):

    uniq_obj = words[37]
    atomic_relata = words[23]
    atomic_relations = words[22]
    def_relat = ["J", "I", '=', 'H']
    definitions = words[16]
    not_oft_def = use_rarely_defined_word()

    num130 = [9, 14]
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        for i in num130:
            adverb = False
            id = False
            kind = ""
            bool2 = False
            relat = all_sent[m][9]
            str1 = all_sent[m][i]
            if relat == "=" and all_sent[m][14] == 'time':
                bb = 8
            if relat == 'B':
                bb = 8
            if m == 13 and i == 9:
                bb = 8

            if all_sent[m][43] != 'cc':
                if relat in def_relat and i == 14:
                    definiendum = findinlist(str1, abbreviations, 0, 1)
                    if definiendum in atomic_relata:
                        bool2 = False
                    else:
                        bool2 = True

                elif i == 48 and all_sent[m][48] != None:
                    definiendum = str1
                    bool2 = True
                    adverb = True
                    kind = 'R'
                elif relat == '=' and all_sent[m][41] == 1:
                    id = True
                    bool2 = True
                    definiendum = all_sent[m][14]
                    all_sent[m][41] = None
                    if definiendum not in numbers_def:
                        bool2 = False

                elif relat == "=" and all_sent[m][41] != 1:
                    bool2 = False
                    list_id = [""] * 80
                    list_id[0] = all_sent[m][5]
                    list_id[1] = all_sent[m][14]
                    list_id[2] = "ID"
                    detach_sent.append(list_id)
                    # the identities are placed in the detach_sent list because
                    # I do not like passing lists around in functions
                    # they are later deleted from the list in step 3
                    all_sent[m][46] = [200]
                    all_sent[m][56] = [200]
                elif i == 9 and relat != "J" and relat != "I" and relat != '=' \
                        and str1 not in atomic_relations:
                    definiendum = str1
                    bool2 = True
                    kind = 'R'
                if (bool2 and isdefineable(all_sent[m]) and definiendum != None and \
                                definiendum != '') or id:
                    if definiendum in not_oft_def:
                        break
                    if (id and definiendum not in uniq_obj) or (definiendum == "concept" + un and id):
                        break

                    g = findposinlist(definiendum, definitions, 0)
                    definition = definitions[g][1]
                    if definiendum == 'ada':
                        bb = 8
                    if definition == 'natural':
                        definition = "(c'=" + definiendum + ") & (d'=natural_whole) & ((bIc') " + conditional \
                                     + " (bId'))"
                    pos = definitions[g][2]
                    circ = definitions[g][3]
                    circ2 = all_sent[m][43]
                    basic_molecule = definitions[g][4]
                    # this prevents us from getting caught in an infinite loop.
                    if basic_molecule == 'b' and all_sent[m][9] == "I":
                        break
                    if relat == "I" and definiendum in uniq_obj:
                        break
                    if circ2 == 'c':
                        circ += circ2
                    if (relat == "J" and pos == 'a') or (relat == "I" and pos == 'n') or (relat == 'H' and pos == 'n') \
                            or pos == 'r' or pos == 'e' or pos == 's' or (relat == '=' and pos == 'n') or adverb or id:
                        if definition != None and all_sent[m][0] not in def_sent:
                            def_sent.append(all_sent[m][0])
                            change_variables(definiendum, m, kind, i)
                            break


def use_rarely_defined_word():
    not_oft_def = copy.deepcopy(words[36])
    for word in abbreviations:
        if word[1] in not_oft_def:
            not_oft_def.remove(word[1])

    return not_oft_def


def uni_scope_rel(m, i):
    univ = ['every', 'no']
    if i == 15:
        if all_sent[m][3] in univ or all_sent[m][10] in univ:
            return False
        else:
            return True
    elif i == 19:
        if all_sent[m][3] in univ or all_sent[m][10] in univ \
                or all_sent[m][16] in univ:
            return False
        else:
            return True
    # else:
    #     return True


def add_necessary_conditions_for_concept():
    # if we're talking about concepts in our proof then we need to add their necesssary
    # conditiona to our proof

    global sn
    part_of_speech = words[28]  # part of speech
    str1 = ""
    list2 = []
    con_sent_parts = [None] * 80
    for i in range(len(abbreviations)):
        if abbreviations[i][1] == 'concept' + un or abbreviations[i][1] == 'concept' + ua:
            str1 = abbreviations[i][0]
        if str1 != "":
            for j in range(len(all_sent)):
                if all_sent[j][9] == "I" and all_sent[j][14] == str1 and all_sent[j][46] != "x" \
                        and all_sent[j][46] != "y":
                    ant_sent_parts = copy.deepcopy(all_sent[j])
                    str2 = all_sent[j][5]
                    con = findinlist(str2, abbreviations, 0, 1)
                    pos = findinlist(con, part_of_speech, 0, 1)
                    if pos == None:
                        bb = 7
                    if con != None:
                        if con == "dog":
                            bb = 8
                        if pos == 'a':
                            str4 = "J"
                        elif pos == 'n':
                            str4 = "I"
                        b = 0
                        for k in range(len(all_sent)):
                            if all_sent[k][9] == str4 and all_sent[k][14] == str2 and \
                                            str1 != str2 and str2 not in list2:
                                str6 = all_sent[k][5]
                                list2.append(str2)
                                b += 1
                        if b > 1:
                            print('you have not coded for multiple concepts')
                        olda = "(" + "b" + ' = ' + con + ")"
                        oldc = "(" + "c " + str4 + " b" + ")"
                        if str2 != "b":
                            rn1 = "(" + "b" + mini_c + str2 + ") & (" + "c" + mini_c + str6 + ")"
                        else:
                            rn1 = "(" + "c" + mini_c + str6 + ")"
                        nat_sent_b4_sub = olda + " " + conditional + " " + oldc
                        sn += 1
                        add_to_total_sent(sn, nat_sent_b4_sub, "", "", "NC concept " + con)
                        sn += 1
                        add_to_total_sent(sn, rn1, "", "", "RN")
                        con_sent_parts[5] = str6
                        con_sent_parts[9] = str4
                        con_sent_parts[14] = str2
                        anc1 = str(sn - 1) + "," + str(sn)
                        prepare_att_sent_1_sent(ant_sent_parts, "SUB", anc1, "c",
                                                build_sent2(con_sent_parts))
                        break


def name_sent(str1, bool2=False, str4=""):
    no_space = copy.copy(str1)
    if str1.find('~') > -1:
        no_space = str1.replace("~", "")
        str1 = str1.replace("~", "")
        ng = '~'
    else:
        ng = ''

    if "  " in str1:
        str1 = str1.replace("  ", " ")

    no_space = remove_outer_paren(no_space)
    no_space = no_space.replace(" ", "")

    if bool2:
        if str4 == 'something':
            no_space = no_space.replace("something", "some thing")
        elif str4 == 'anything':
            no_space = no_space.replace("anything", "any thing")
        elif str4 == 'everything':
            no_space = no_space.replace("everything", "every thing")
        elif str4 == 'anything' + ua:
            no_space = no_space.replace("anything" + ua, "a" + ua + " thing")

    h = findinlist(no_space, prop_name, 1, 0)
    if h != None:
        return ng + h
    else:
        prop_name.append([prop_var[0], no_space, str1])
        str2 = prop_var[0]
        del prop_var[0]
        return ng + str2



def insert_space(str, integer):
    return str[0:integer] + ' ' + str[integer:]


def space_words(str1):
    # this function place a space between variables and relations
    # so as to make it easier to categorize words in a sentence
    str1 = str1.replace("~", " ~ ")
    str1 = str1.replace("=", " = ")
    str1 = str1.replace(neg, " " + neg + " ")
    str1 = str1.replace(mini_e, " " + mini_e + " ")
    str1 = str1.replace(ne, " " + ne + " ")
    i = -1
    while i + 1 < len(str1):
        i += 1
        temp_str = str1[i:(i + 1)]
        nxt_str = str1[(i + 1):(i + 2)]
        if nxt_str.isupper() == True and temp_str.islower() == True:
            str1 = insert_space(str1, i + 1)
        elif nxt_str.islower() == True and temp_str.isupper() == True:
            str1 = insert_space(str1, i + 1)
    return str1


def get_abbreviations_from_definition(def_info):
    # this function picks out that variables in the id sentences of the
    # definition

    constants = []
    list3 = []
    propositional_constants = []
    for i in range(len(def_info[0])):
        if os(def_info[0][i]) and mini_e in def_info[0][i]:
            list3.append(def_info[0][i])
        elif os(def_info[0][i]) and "=" in def_info[0][i]:
            str1 = def_info[0][i]
            g = str1.find("=")
            var = str1[1:g]
            wrd = str1[g + 1:-1]
            if isvariable(var):
                if not isvariable(wrd):
                    constants.append([var, wrd])

    if list3 != []:
        propositional_constants = get_propositional_constants(list3)

    return constants, propositional_constants


def get_propositional_constants(list3):
    list4 = []
    list6 = []
    list7 = []
    propositional_constants = []
    for i in range(len(list3)):
        prop_con = list3[i][1]
        str2 = list3[i].replace(" ", "")
        str2 = str2[3:-1]
        str3 = space_words(str2)
        list8 = ["", "", ""]
        list8 = str3.split()
        list8.append(None)
        list5 = categorize_words(prepare_categorize_words(str3))
        list4.append(prop_con)
        str4 = list5[0].replace(" ", "")
        list6.append(str4)
        list7.append(list5)

    for i in range(len(list4)):
        if [list4[i], list6[i], list7[i]] not in propositional_constants:
            propositional_constants.append([list4[i], list6[i], list7[i]])

    return propositional_constants


def is_atomic(list1, atomic_relations):
    relat = [9, 15, 19, 23, 27, 31]
    must_be_blank = [2, 3, 4, 6, 7, 10, 11, 12, 16, 17, 20, 21, 24, 25, 28, 29, 32, 33]
    noun = [5, 14, 18, 26, 30, 34]
    for i in range(len(list1)):
        str1 = list1[i]
        if str1 != None:
            if i in relat:
                if str1 in atomic_relations:
                    pass
                else:
                    pass
            elif i in must_be_blank:
                return False
            elif i in noun:
                if not isvariable(str1):
                    return False
    return True


def isdefineable(list1):
    must_be_blank = [3, 4, 6, 7, 10, 11, 13, 16, 17, 18, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33,
                     35, 36, 49, 50, 51, 52, 55]
    must_be_variable = [5, 14, 18, 22]

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None:
            if not isvariable(list1[i]) and list1[i] != 'i':
                return False
    return True


def is_standard(list1):
    must_be_blank = [3, 4, 6, 7, 10, 11, 13, 16, 17, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33,
                     35, 36, 49, 50, 51, 52, 55]
    must_be_variable = [5, 14, 18, 22]

    for i in must_be_blank:
        if list1[i] != None and list1[i] != '':
            return False
    for i in must_be_variable:
        if list1[i] != None and list1[i] != "":
            if not isvariable(list1[i]) and list1[i] != 'i':
                return False
    return True


def build_sent2(list1):
    # if you revise this list then then you must also revise it in
    # the eliminate_univ_quant_subclause, extract_words_from_subclause, as well as the function 'that', as well as new_categories
    # g=1 means that it is a sentence that identifies a propositional constant, in some cases
    # the proposition itself need not be named
    # also fix list in word sub and isatomic
    global build_sent_counter, build_sent2_time
    aa = time.time()
    str1 = "("
    num = allowable_slots()

    for i in num:
        temp_str = list1[i]
        if temp_str != None and temp_str != "":
            if temp_str not in words_used:
                words_used.append(temp_str)
            if str1 == "(":
                str1 += temp_str
            else:
                str1 += " " + temp_str

    str1 += ")"
    str1p = name_sent(str1)
    if str1p == "k":
        bb = 8
    list1[0] = str1
    list1[2] = "~" if "~" in str1p else ""
    list1[1] = str1p.replace("~", "") if "~" in str1p else str1p
    list1[72] = str1.replace("~", "") if "~" in str1 else str1
    list1[72] = list1[72].replace("  ", " ")
    list1[42] = str1p
    build_sent2_time += time.time() - aa
    build_sent_counter += 1
    return list1


def build_sent3(list1):
    str1 = "("
    num = allowable_slots()

    for i in num:
        temp_str = list1[i]
        if temp_str != None and temp_str != "":
            if temp_str not in words_used:
                words_used.append(temp_str)
            if str1 == "(":
                str1 += temp_str
            else:
                str1 += " " + temp_str

    str1 += ")"

    return str1


def build_uncategorized_sent(list1):
    for i in range(3, len(list1)):
        if list1[i] == None or list1[i] == "":
            break
        if i == 3:
            str1 = list1[i]
        else:
            str1 += " " + list1[i]

    str1 = "(" + str1 + ")"
    str1p = name_sent(str1)
    list1[0] = str1
    list1[2] = ""
    list1[1] = str1p
    list1[72] = str1
    list1[42] = str1p

    return list1


def build_sent_list(list1):
    # this list builder does not have the ~ separated from the sentence

    str2 = None
    for i in range(len(list1)):
        if str2 == None:
            str2 = list1[i]
        else:
            str2 = str2 + ' & ' + list1[i]

    return str2



def remove_duplicates2d(list1, i, h):
    list2 = []
    j = -1
    while j < len(list1) - 1:
        j += 1
        for k in range(len(list1)):
            if k != j:
                if list1[k][i] == list1[j][i] and list1[k][h] == list1[j][h]:
                    del list1[j]
                    j -= 1
                    break
        else:
            list2.append(list1[j])
    return list2


def get_position_of_identities():
    for i in range(len(total_sent)):
        if total_sent[i][4] != "":
            return i
    g = 4 / 0


def build_list_of_abbreviations():
    # this turns the abbreviations into a conjunction
    global sn
    position_of_identities = get_position_of_identities()
    # this loop removes duplicates in a multidimensional list
    str1 = ""
    add_to_total_sent("", "_________________", "__________________")
    add_to_total_sent("", "")
    total_attach_sent = []
    temp_detach_sent = []
    for lst in attach_sent:
        for sent in lst[38]:
            total_attach_sent.append(sent)

    for abbrev in abbreviations:
        str2 = "(" + abbrev[0] + "=" + abbrev[1] + ")"
        str2p = name_sent(str2)
        if abbrev[1] not in words_used:
            words_used.append(abbrev[1])


        if str1 == "":
            str1 = str2
            str1p = str2p
        else:
            str1 += " & " + str2
            str1p += " & " + str2p
        if str2p in total_attach_sent:
            list1 = [None] * 80
            list1[0] = str2
            list1[1] = str2p
            list1[2] = ""
            list1[5] = abbrev[0]
            list1[8] = ""
            list1[9] = "="
            list1[14] = abbrev[1]
            list1[72] = str2
            list1[42] = str2p
            temp_detach_sent.append(list1)

    if temp_detach_sent != []:
        for sent in temp_detach_sent:
            sn += 1
            sent[58] = sn
            detach_sent.append(sent)
            add_to_total_sent(sn, sent[0], sent[1], "", "&E", position_of_identities + 1)

    total_sent.insert(position_of_identities, [position_of_identities + 1, str1,
                                               str1p, "", 'ID', "", "", "", ""])


def remove_values_from_list(the_list, val):
    while val in the_list:
        the_list.remove(val)
    return the_list


def divide_sent(list2):

    global sn

    for i in range(len(list2)):
        str2 = list2[i][1]
        str2 = str2.lower()
        str3 = name_sent(str2)
        str2 = str2.strip()
        list3 = str2.split()
        sent_parts = [None] * 80
        sent_parts[0] = str2
        sent_parts[42] = str3
        sent_parts[72] = str2
        sent_parts[1] = str3
        sent_parts[2] = ""
        sent_parts[58] = list2[i][0]
        for j in range(len(list3)):
            sent_parts[j + 3] = list3[j]
            if list3[j] not in def_used:
                def_used.append(list3[j])
        list4 = copy.deepcopy(sent_parts)
        detach_sent.append(sent_parts)
        add_to_total_sent(list2[i][0], str2, str3, "", "")
        all_sent.append(list4)


def eliminate_redundant_words():

    redundant = words[21]
    bool1 = False
    for i in range(len(all_sent)):
        ant_sent_parts = copy.deepcopy(all_sent[i])
        str2 = ""
        for j in range(len(redundant)):
            str1 = redundant[j]
            if str1 in all_sent[i]:
                if str1 not in words_used:
                    words_used.append(str1)
                bool1 = True
                all_sent[i] = remove_values_from_list(all_sent[i], str1)
                if str2 == '':
                    str2 = str1
                else:
                    str2 += "," + str1
        if bool1:
            bool1 = False
            all_sent[i] = build_uncategorized_sent(all_sent[i])
            con_parts = copy.deepcopy(all_sent[i])
            prepare_att_sent_1_sent(ant_sent_parts, "RD " + str2, "", "e", con_parts)


def replace_relations():

    relations = words[6]
    doubles = words[31]
    doubles.sort()
    for j in range(len(all_sent)):
        ant_sent_parts = copy.deepcopy(all_sent[j])
        i = 2
        while all_sent[j][i + 1] != None:
            i += 1
            if "," in all_sent[j][i]:
                has_comma = True
                str3 = all_sent[j][i]
                str3 = str3.replace(",", "")
            else:
                str3 = all_sent[j][i]
                has_comma = False
            if str3 == 'spies':
                bb = 8
            bool2 = is_in_md(doubles, 0, str3)
            bool3 = False
            if bool2 and all_sent[j][i + 1] != None:
                str4 = all_sent[j][i] + " " + all_sent[j][i + 1]
                bool3 = is_in_md(doubles, 1, str4)
                if bool3:
                    str3 += " " + all_sent[j][i + 1]
                    if str3 not in def_used:
                        def_used.append(str3)
            else:
                if str3 not in def_used:
                    def_used.append(str3)
            str2 = findinlist(str3, relations, 0, 1)

            if str2 != None:
                g = findposinlist(str3, abbreviations, 0)
                if g == -1:
                    abbreviations.append([str3, str2])
                if has_comma:
                    all_sent[j][i] = str2 + ","
                else:
                    all_sent[j][i] = str2
            if bool3:
                if str2 != None:
                    del all_sent[j][i + 1]
                else:
                    i += 1

        all_sent[j] = categorize_words(all_sent[j])
        con_parts = copy.deepcopy(all_sent[j])
        if con_parts[0] != ant_sent_parts[0]:
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", "", "e", con_parts)


def eliminate_negative_determiners():

    num = [8, 12, 49, 50, 51, 52]
    special_determinatives = ['many' + un, 'any' + un]
    for i in range(len(all_sent)):
        ant_sent_parts = copy.deepcopy(all_sent[i])
        bool2 = False
        add_120 = False
        for j in num:
            bool1 = False  # yyy
            if all_sent[i][j] == "not" or all_sent[i][j] == neg:
                if j == 8:
                    if all_sent[i][9] != ne:
                        if all_sent[i][8] == 'not':
                            all_sent[i][8] = "~"
                    if all_sent[i][10] == "a":
                        all_sent[i][10] = 'every'
                        bool2 = True
                        rule = "DE ~ a"
                        add_120 = True
                    elif all_sent[i][10] == "every":
                        all_sent[i][10] = 'many' + un
                        bool2 = True
                        bool1 = True
                        rule = "DE ~ every"
                    elif all_sent[i][10] in special_determinatives:
                        rule = "DE ~ " + all_sent[i][10]
                        all_sent[i][10] = 'every'
                        bool2 = True
                        add_120 = True
                    elif all_sent[i][9] == ne:
                        all_sent[i][9] = "="
                        bool2 = True
                        bool1 = True
                        rule = "DE ~" + ne
                elif j == 12:
                    if all_sent[i][10] == "a":
                        all_sent[i][10] = 'every'
                        bool2 = True
                        rule = "DE ~ a"
                        add_120 = True
                    elif all_sent[i][10] == "every":
                        all_sent[i][10] = 'many' + un
                        bool1 = True
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ every"
                    elif all_sent[i][10] in special_determinatives:
                        all_sent[i][10] = 'every'
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ " + all_sent[i][18]
                elif j == 49:
                    if all_sent[i][16] == "a":
                        all_sent[i][16] = 'every'
                        bool2 = True
                        rule = "DE ~ a"
                        add_120 = True
                    elif all_sent[i][16] == "every":
                        all_sent[i][16] = 'many' + un
                        bool1 = True
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ every"
                    elif all_sent[i][16] in special_determinatives:
                        all_sent[i][16] = 'every'
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ " + all_sent[i][18]
                    elif all_sent[i][15] == ne:
                        all_sent[i][15] = "="
                        bool1 = True
                        bool2 = True
                        add_120 = True
                        rule = "DE ~" + ne
                elif j == 50:
                    if all_sent[i][20] == "a":
                        all_sent[i][20] = 'every'
                        bool2 = True
                        rule = "DE ~ a"
                        add_120 = True
                    elif all_sent[i][20] == "every":
                        all_sent[i][20] = 'many' + un
                        bool1 = True
                        bool2 = True
                        rule = "DE ~ every"
                    elif all_sent[i][20] in special_determinatives:
                        all_sent[i][20] = 'every'
                        bool2 = True
                        rule = "DE ~ " + all_sent[i][22]
                elif j == 51:
                    if all_sent[i][24] == "a":
                        all_sent[i][24] = 'every'
                        bool2 = True
                        rule = "DE ~ a"
                        add_120 = True
                    elif all_sent[i][24] == "every":
                        all_sent[i][24] = 'many' + un
                        bool1 = True
                        bool2 = True
                        rule = "DE ~ every"
                        add_120 = True
                    elif all_sent[i][24] in special_determinatives:
                        all_sent[i][24] = 'every'
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ " + all_sent[i][26]
                elif j == 52:
                    if all_sent[i][28] == "a":
                        all_sent[i][28] = 'every'
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ a"
                    elif all_sent[i][28] == "every":
                        all_sent[i][28] = 'many' + un
                        bool1 = True
                        bool2 = True
                        rule = "DE ~ every"
                    elif all_sent[i][28] in special_determinatives:
                        all_sent[i][28] = 'every'
                        bool2 = True
                        add_120 = True
                        rule = "DE ~ " + all_sent[i][30]
                if bool1:
                    all_sent[i][j] = None
                    all_sent[i][46].remove(j)
                if add_120 == True:
                    all_sent[i][56].append(120)

        if bool2:
            con_parts = copy.deepcopy(build_sent2(all_sent[i]))
            prepare_att_sent_1_sent(ant_sent_parts, rule, "", "e", con_parts)
            new_categories(all_sent[i])


def new_categories(list5, kind=False):
    list1 = [None] * 80
    if not kind:
        num = list5[46]
    else:
        num = [47, 3, 69, 4, 55, 5, 66, 67, 35, 48, 59, 6, 8, 9, 7, 48, 12, 10, 70,
               13, 14, 36, 60, 63, 49, 15, 16, 17, 18,
               61, 64, 50, 19, 20, 21, 22, 62, 65, 51, 23, 24, 25, 26, 52, 27, 28,
               29, 30, 31, 32, 33, 34]
    i = 2
    for j in num:
        if list5[j] != None and list5[j] != "":
            i += 1
            list1[i] = list5[j]
    list3 = categorize_words(list1)
    return list3


def eliminate_possessives2(m, i):
    ant_sent_parts = copy.deepcopy(all_sent[m])
    list1 = [None] * 80
    str1 = all_sent[m][i][0]
    list1[5] = str1
    list1[9] = "OWN"
    if i == 69:
        str2 = all_sent[m][5]
    elif i == 70:
        str2 = all_sent[m][14]
    list1[14] = str2
    list1[46] = [200]
    list1[56] = [200]
    all_sent[m][i] = None
    con_parts = copy.deepcopy(build_sent2(all_sent[m]))
    con_parts2 = copy.deepcopy(build_sent2(list1))
    all_sent.append(con_parts2)
    prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, "PNE", "e")


def eliminate_possessive_nouns(m, n, definite_assignments, str7):

    str1 = all_sent[m][n]
    str1 = str1[:1]
    if str7 == "a":
        str2 = "indefinite"
        new_var = variables[0]
        del variables[0]
    elif str7 == "the":
        str2 = "definite"
        str9 = findinlist(str1, abbreviations, 0, 1)
        str10 = findinlist(str9, definite_assignments, 1, 0)
        if str10 == None:
            new_var = variables[0]
            del variables[0]
            # definite_list used here
            definite_assignments.append([new_var, str9])
        else:
            bb = 8

    str3 = findinlist(str2, abbreviations, 1, 0)
    all_sent[m][n] = new_var + "'s"
    list1 = [None] * 80
    list1[5] = new_var
    list1[14] = str3
    list1[9] = "J"
    list1[44] = chr(964)
    list2 = [None] * 80
    list2[5] = new_var
    list2[14] = str1
    list2[9] = "I"
    list1 = build_sent2(list1)
    list2 = build_sent2(list2)
    list2[44] = chr(965)
    list1[46] = [200]
    list1[56] = [200]
    list2[46] = [200]
    list2[56] = [200]

    return [list1, list2]


def eliminate_that2(m, i):

    num = allowable_slots()
    list1 = [None] * 80
    bool1 = False
    ant_sent_parts = copy.deepcopy(all_sent[m])
    list3 = copy.deepcopy(all_sent[m])

    k = 2
    for j in num:
        if j == i:
            bool1 = True
        if list3[j] != None and bool1 and j != i:
            k += 1
            list1[k] = all_sent[m][j]
            list3[j] = None

    list2 = categorize_words(list1)
    list2[0] = remove_outer_paren(list2[0])
    str3 = list2[0].replace(" ", "")
    g = findposinlist(str3, abbreviations, 1)
    if g == -1:
        for z in range(len(variables)):
            if variables[z] in prop_var:
                new_var = variables[z]
                del variables[z]
                prop_var.remove(new_var)
                break
        abbreviations.append([new_var, str3, 1])
    else:
        new_var = abbreviations[g][0]

    list3[i] = None
    if i == 7:
        if list3[5] == 'it':
            list3[5] = new_var
        else:
            list3[14] = new_var
    elif i == 60:
        list3[5] = new_var

    elif i == 61:
        if list3[14] == 'it':
            list3[14] = new_var
        else:
            list3[18] = new_var

    elif i == 62:
        if list3[18] == 'it':
            list3[62] = new_var
        else:
            list3[22] = new_var

    list3 = build_sent2(list3)
    list3 = new_categories(list3, True)
    con_parts = copy.deepcopy(list3)
    all_sent.append(list3)
    prepare_att_sent_1_sent(ant_sent_parts, "DE that", "", "e", con_parts)


def lies_wi_scope_of_univ_quant(m, i, kind=""):
    comma = all_sent[m][39]
    univ = ['every', 'many' + up, 'many' + uo, 'no']
    if kind == 1:
        if i == 59:
            i = 12
        elif i == 60:
            i = 16
        elif i == 61:
            i = 20
        elif i == 62:
            i = 24
    if comma == None:
        comma = 70
    if all_sent[m][59] != None and all_sent[m][3] in univ:
        if i < comma and i > 3:
            return False
    elif all_sent[m][60] != None and all_sent[m][10] in univ:
        if i < comma and i > 10:
            return False
    elif all_sent[m][61] != None and all_sent[m][17] in univ:
        if i < comma and i > 17:
            return False
    elif all_sent[m][62] != None and all_sent[m][21] in univ:
        if i < comma and i > 21:
            return False
    if i == 13 and all_sent[m][10] in univ:
        return False
    elif all_sent[m][i - 1] in univ:
        return False
    return True


def eliminate_relative_pronouns3(i, m, list1, new_var=""):

    subjrp = ['who', 'which', 'that' + us]
    objrp = ['who' + uo, 'that' + uo, 'which' + uo]

    if new_var != "":
        for i in range(59, 63):
            if all_sent[m][i] != None:
                break

    if all_sent[m][i] in subjrp:
        srp = True
    else:
        srp = False
    if i == 59:
        list2 = [[8, 49], [9, 15], [10, 16], [13, 17], [14, 18], [15, 19], [18, 22], [60, 61], [63, 64]]
        all_sent[m][59] = None
        if new_var == "":
            list1[5] = all_sent[m][5]
        else:
            list1[5] = new_var
        list1[8] = all_sent[m][8]
        list1[9] = all_sent[m][9]
        list1[14] = all_sent[m][14]

        for j in range(len(list2)):
            k = list2[j][0]
            n = list2[j][1]
            all_sent[m][k] = all_sent[m][n]
            all_sent[m][n] = None

    elif i == 60 and srp:
        list2 = [[8, 49], [9, 15], [10, 16], [13, 17], [14, 18], [15, 19], [18, 22]]
        all_sent[m][60] = None
        if new_var == "":
            list1[5] = all_sent[m][14]
        else:
            list1[5] = new_var
        for j in range(len(list2)):
            k = list2[j][0]
            n = list2[j][1]
            list1[k] = all_sent[m][n]
            all_sent[m][n] = None
    elif i == 61:
        a = 22
        c = 26
    elif i == 62:
        a = 30
        c = 34

    list1 = new_categories(list1, True)
    all_sent[m] = new_categories(all_sent[m], True)

    return list1


def prepare_att_sent_4_sent(ant_sent_parts, con_parts, con_parts2,
                            con_parts3,con_parts4, rule, sent_type):

    global sn
    sn += 1
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    con_parts3[68] = "123"
    con_parts3[44] = chr(952)
    con_parts4[68] = "124"
    con_parts4[44] = chr(953)
    connective = iff if sent_type == "e" else conditional
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + \
                      " & " + con_parts2[0] + " & " + con_parts3[0] + " & " + con_parts4[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + \
                    " & " + con_parts2[42] + " & " + con_parts3[42] + " & " + con_parts4[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + \
                " & " + chr(952) + " & " + chr(953) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
        con_parts3[53] = 'cf'
        con_parts4[54] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'
        con_parts3[53] = 'cq'
        con_parts4[54] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]],
                [con_parts3[1], con_parts3[2]], [con_parts4[1], con_parts4[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + \
                " & " + con_parts3[42] + " & " + con_parts4[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts, con_parts2, con_parts3, con_parts4]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1], con_parts3[1], con_parts4[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + " & " + con_parts3[0] +\
                 " & " + con_parts4[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]],
                [con_parts3[72], con_parts3[2]], [con_parts4[72], con_parts4[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)



def prepare_att_sent_3_sent(ant_sent_parts, con_parts, con_parts2, con_parts3, rule, sent_type):
    # this populates the attach_sent list provided a sentence is equivalent
    # to two conjuncts

    global sn
    sn += 1
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    con_parts3[68] = "123"
    con_parts3[44] = chr(952)
    connective = iff if sent_type == "e" else conditional
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + \
                      " & " + con_parts2[0] + " & " + con_parts3[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + \
                    " & " + con_parts2[42] + " & " + con_parts3[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + \
                " & " + chr(952) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
        con_parts3[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'
        con_parts3[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]],
                [con_parts3[1], con_parts3[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + " & " + con_parts3[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts, con_parts2, con_parts3]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1], con_parts3[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + " & " + con_parts3[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]],
                [con_parts3[72], con_parts3[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)



def prepare_att_sent_2_sent(ant_sent_parts, con_parts, con_parts2, rule, sent_type):
    # this populates the attach_sent list provided a sentence is equivalent
    # to two conjuncts

    global sn
    sn += 1
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "121"
    con_parts[44] = chr(950)
    con_parts2[68] = "122"
    con_parts2[44] = chr(951)
    connective = iff if sent_type == "e" else conditional
    new_equivalence = ant_sent_parts[0] + " " + connective + " (" + con_parts[0] + " & " + con_parts2[0] + ")"
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " (" + con_parts[42] + " & " + con_parts2[42] + ")"
    new_greek = chr(949) + " " + connective + " (" + chr(950) + " & " + chr(951) + ")"
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'cf'
        con_parts2[53] = 'cf'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'cq'
        con_parts2[53] = 'cq'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]], [con_parts2[1], con_parts2[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = ["(" + con_parts[42] + " & " + con_parts2[42] + ")", ""]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts, con_parts2]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1], con_parts2[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = ["(" + con_parts[0] + " & " + con_parts2[0] + ")", ""]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]], [con_parts2[72], con_parts2[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule)
    attach_sent.append(list4)


def prepare_att_sent_1_sent(ant_sent_parts, rule, anc1, sent_type, con_parts):
    # this populates the attach_sent list provided a sentence is equivalent to one other sentence

    global sn
    sn += 1
    list4 = [""] * 50
    ant_sent_parts[68] = "11"
    ant_sent_parts[44] = chr(949)
    con_parts[68] = "12"
    con_parts[44] = chr(950)
    connective = iff if sent_type == "e" else conditional
    new_equivalence = ant_sent_parts[0] + " " + connective + " " + con_parts[0]
    new_eq_abbrev = ant_sent_parts[42] + " " + connective + " " + con_parts[42]
    new_greek = chr(949) + " " + connective + " " + chr(950)
    if sent_type == 'e':
        ant_sent_parts[53] = 'b'
        con_parts[53] = 'f'
    else:
        ant_sent_parts[53] = 'a'
        con_parts[53] = 'q'

    list4[0] = [[ant_sent_parts[1], ant_sent_parts[2]]]
    list4[1] = [[con_parts[1], con_parts[2]]]
    list4[2] = sn
    list4[3] = sent_type
    list4[4] = new_eq_abbrev
    list4[7] = [ant_sent_parts[1], ant_sent_parts[2]]
    list4[8] = [con_parts[1], con_parts[2]]
    list4[34] = [ant_sent_parts]
    list4[35] = [con_parts]
    list4[37] = new_equivalence
    list4[38] = [ant_sent_parts[1], con_parts[1]]
    list4[40] = [ant_sent_parts[72], ant_sent_parts[2]]
    list4[41] = [con_parts[72], con_parts[2]]
    list4[42] = [[ant_sent_parts[72], ant_sent_parts[2]]]
    list4[43] = [[con_parts[72], con_parts[2]]]
    list4[47] = new_greek

    add_to_total_sent(sn, new_equivalence, new_eq_abbrev, "", rule, anc1)
    attach_sent.append(list4)


def is_in_abbreviations(list1):
    if not list1[9] == "=":
        return False
    else:
        bool1 = is_in_md(abbreviations, 1, list1[14])
        if bool1:
            print ("is in abbreviations")
        return bool1


def allowable_slots():

    num2 = [11, 47, 3, 69, 4, 55, 5, 66, 67, 35, 48, 59, 6, 8, 9, 7, 48, 12, 10, 70,
            13, 14, 36, 60, 63, 49, 15,
            16, 17, 18,
            61, 64, 50, 19, 20, 21, 22, 62, 65, 51,
            23, 24, 25, 26, 52, 27, 28,
            29, 30, 31, 32, 33, 34]

    return num2


def extract_words_from_subclause(m, k):
    # this is the sentence that will be inserted into the antecedent in the
    # definition of 'every' or 'no', in the future it should output
    # a set of lists, not just 1

    num2 = allowable_slots()
    n = 970
    comma = all_sent[m][39]
    list1 = [None] * 80
    bool1 = False
    i = 2
    for j in num2:
        if j == k:
            bool1 = True
        if bool1:
            if comma != None:
                if j == comma + 1:
                    break
            if all_sent[m][j] != None and all_sent[m][j] != "":
                i += 1
                list1[i] = all_sent[m][j]
                all_sent[m][j] = None
    list1 = categorize_words(list1)

    list1[44] = chr(n)
    list1[40] = False

    return [list1]


def eliminate_univ_quant_subclause(m, k, o, new_var2):
    # this eliminates a subclause which is quantified by a universal quanitifier

    list1 = extract_words_from_subclause(m, k)
    num2 = allowable_slots()

    list2 = [None] * 80
    n = 975  # this is for the greek character
    if o == 5:
        list2.append(new_var2)
    j = 2
    for i in num2:
        if (all_sent[m][i] != None and all_sent[m][i] != "") or i == o:
            j += 1
            if i == o:
                list2[j] = new_var2
            else:
                list2[j] = all_sent[m][i]
    list2 = categorize_words(list2)

    list2[44] = chr(n)
    all_sent[m] = list2
    all_sent.append(list2)

    return list1


def repl_sign(str3, match_dv, match_type):
    s = findposinmd(str3, match_dv, 1)
    s = match_type[s]
    if s == 0:
        return mini_c
    else:
        return idd


def abb_change(list5, already_checked, i, match_dv, match_type,
               rename, j, defin_sent, def_con, new_match=[],
               second=False):
    denied_consequent_sent = False
    no_match_found = True
    spec_prop = ['definite', 'indefinite']  # special properties

    for t in range(len(all_sent)):
        if t == 13:
            bb = 8
        no_match = False
        bool1 = False
        if all_sent[t][9] == "J":
            str1 = findinlist(all_sent[t][14], abbreviations, 0, 1)
            if str1 in spec_prop:
                bool1 = True
        if bool1:
            pass
        elif t not in already_checked:
            # if a variable has already been turned into all_sent[t][j] then it cannot
            # happen again since in a definition all variables stand for different things
            if not is_in_md(match_dv, 1, all_sent[t][j]):
                for u in list5:
                    if is_in_md(match_dv, 1, all_sent[t][u]) and u == i:
                        no_match = True
                    elif all_sent[t][u] == defin_sent[i][u]:
                        pass
                    elif u == 8 and (all_sent[t][53] == 'cn' or defin_sent[i][53] == 'cn' \
                                             or def_con == conditional):
                        denied_consequent_sent = True
                    else:
                        if u == 9:
                            already_checked.append(t)
                        no_match = True
                        break

                if not no_match:
                    match_dv.append([defin_sent[i][j], all_sent[t][j]])
                    # cap is for a denied consequent sentence
                    str2 = "(" + defin_sent[i][j] + idd + all_sent[t][j] + ")"
                    # print ("sub case abb change")
                    if denied_consequent_sent:
                        list3 = build_sent2(defin_sent[i])
                        str3 = list3[0]
                        str3 = str3 + l4
                        match_type.append(4)
                        rename.append(str3)
                    else:
                        str2 = str2 + l3
                        match_type.append(3)
                        rename.append(str2)
                        # eee
                    if second:
                        # print ("new match")
                        new_match.append([defin_sent[i][j], all_sent[t][j]])
                    defin_sent[i][j] = all_sent[t][j]
                    no_match_found = False
                    return no_match_found, defin_sent
    return no_match_found, defin_sent


def abb_change2(match_dv, match_type, defin_sent, i, temp_match, j, rename):

    match_dv.append([defin_sent[i][j], variables[0]])
    match_type.append(2)
    str1 = "(" + defin_sent[i][j] + idd + variables[0] + ")"
    str1 = str1 + l2
    temp_match.append([defin_sent[i][j], variables[0]])
    defin_sent[i][j] = variables[0]
    rename.append(str1)
    del variables[0]

    return match_dv, match_type, defin_sent, temp_match, rename


def eliminate_adjective_wi_universal_sent(m, j, new_var, kind=0):
    adj_var = all_sent[m][j - 1]
    all_sent[m][j - 1] = None
    list1 = [None] * 80
    list1[5] = new_var
    list1[9] = "J"
    list1[14] = adj_var
    list1 = build_sent2(list1)
    list1[46] = [200]
    list1[44] = chr(980)
    list1[56] = [200]
    all_sent.append(list1)
    list1[40] = False

    return list1


def eliminate_adjective_wi_universal_sent2(subj, relat, obj):
    list1 = [None] * 80
    list1[5] = subj
    list1[9] = relat
    list1[14] = obj
    list1 = build_sent2(list1)
    list1[40] = False
    list1[44] = chr(964)
    list1[46] = [200]
    list1[56] = [200]

    return list1


def is_set_of_bic_wo_id(def_info):
    # if a definition is composed of a set of conditionals or biconditionals
    # and there is no identity statement within the definition then
    # we do not need to delete anything from the def_info list

    if def_info[4][0][1] == "&":
        for lst in def_info[4]:
            if len(lst[0]) == 2 and lst[1] == "":
                return False
            elif len(lst[0]) > 2:
                return True

    return False


def eliminate_conjuncts_from_definition(def_info):
    # any sentences which are not within either an iff or conditiional
    # are deleted

    # 1. if a definition has some detached conjuncts then these are eliminated
    # and a new greek definition is placed in def_info[5], the def_info
    # list will have one member
    # 2. if a definition is composed of sets of complex sentences then
    # the greek definition remains in def_info[5]
    # the other complex sentences are placed in sets of equivalences
    # and it is written on [45] that they are not yet detached and
    # hence cannot be used in the modus ponens function
    # 3. when we add to the total sent list, we do so by looping through the
    # sets of equivalences list and only adding those that say in [45]
    # 'detached'



    if is_set_of_bic_wo_id(def_info):
        for lst in def_info[4]:
            lst[0] = lst[0][1:]
        def_info = make_sets_of_equivalences(def_info)
        def_info = build_conjunction_of_biconditionals(def_info)
    elif def_info[4][0][1] == "&":
        i = 0
        while len(def_info[4][i][0]) < 3:
            if def_info[4][i][1] == "" and len(def_info[4][i][0]) == 2:
                del def_info[0][i]
                del def_info[1][i]
                del def_info[3][i]
                del def_info[4][i]
                del def_info[6][i]
            else:
                i += 1

        def_info = prepare_def_info_list(def_info)
    else:
        def_info = [def_info]

    return def_info


def prepare_def_info_list(def_info):
    connectives = [iff, conditional, idisj, xorr]
    # this removes the first number from the old numbers
    for lst in def_info[4]:
        lst[0] = lst[0][1:]

    # if the definition was composed of a set of conjunctions then the
    # the greek sentence needs to be amended, otherwise we need
    # the "do not add to total sent list
    if def_info[4][1][1] in connectives and def_info[4][2][1] in connectives:
        def_info = make_sets_of_equivalences(def_info)
        def_info = build_conjunction_of_biconditionals(def_info)
    else:
        del def_info[0][0]
        del def_info[1][0]
        del def_info[3][0]
        del def_info[4][0]
        del def_info[6][0]
        def_info[6][0] = remove_outer_paren(def_info[6][0])
        def_info[5] = def_info[6][0]
        def_info = [def_info]

    return def_info


def make_sets_of_equivalences(def_info):
    # if a definition is composed of a set of equivalences
    # then this function puts each complex sentence into a list

    num = [0, 1, 3, 4, 6]
    list2 = []
    for i in range(len(def_info[4])):
        if len(def_info[4][i][0]) == 1:
            list1 = [""] * 7
            for j in num:
                list1[j] = copy.deepcopy([def_info[j][i]])
            list2.append(list1)
        elif len(def_info[4][i][0]) > 1:
            list1 = [""] * 7
            for j in num:
                list1[j] = copy.deepcopy(def_info[j][i])
            for m in range(len(list2)):
                if list1[4][0].startswith(list2[m][4][0][0]):
                    for k in num:
                        list3 = list2[m][k]
                        list3.append(list1[k])
                        list2[m][k] = list3
                    list2[m][5] = list2[m][6][0]

    def_info = [def_info]
    for lst in list2:
        def_info.append(lst)

    return def_info


def build_conjunction_of_biconditionals(def_info):
    str1 = def_info[1][6][0]
    def_info[1][2] = 'eliminate as conjunct'
    for i in range(2, len(def_info)):
        str1 += " & " + def_info[i][5]
        def_info[i][2] = 'eliminate as conjunct'

    def_info[0][5] = str1
    def_info[0][6][0] = str1
    return def_info


def change_variables(definiendum, m, kind="", k=0, definite_assignments=[]):
    # def_rn = definition rename
    # this function renames the variables in a definition
    # end0
    # match_type 0 = instantiation
    # match_type 1 = idd, constants, 2 = unused var, 3 = already has relation
    # 4 = negated consequent

    global sn, def_used, time_spent_defining
    b = time.time()
    # this is for those determinatives which have negations in their definitions where
    # the sentences has an R variable
    identical_det = ["only", "anything_except", "anyone_except", "many" + un, 'no']
    definitions = words[16]
    definition = findinlist(definiendum, definitions, 0, 1)

    if definiendum == 'moment':
        bb = 8

    if definiendum not in def_used and not definiendum.isupper():
        def_used.append(definiendum)

    if definiendum in identical_det:
        ident_det = True
    else:
        ident_det = False
    first_sent = copy.deepcopy(all_sent[m])
    if (kind == "R" or kind == "") and first_sent[8] == "~":
        first_sent[8] = ""
        first_sent = build_sent2(first_sent)

    first_sent[68] = "11"  # sometimes it might be 11 this is mainly so that it will not
    # be added back to the all sent list
    match_dv = []
    match_type = []
    new_var = []
    rule = ""
    taken_out = []
    repeat_relat = words[39]
    x = findposinlist(definiendum, repeat_relat, 0)
    if x > -1:
        rr_var = repeat_relat[x][1]
    else:
        rr_var = 0

    temp_ad = []
    ad = findposinmd(definiendum, already_defined, 0)
    if ad == -1:
        already_df = False
        def_info2 = find_sentences(definition)
        temp_ad.append(definiendum)
        temp_ad.append(def_info2)
    else:
        already_df = True
        def_info2 = already_defined[ad][1]

    def_info = copy.deepcopy(def_info2)

    const_in_def, propositional_constants = get_abbreviations_from_definition(def_info)
    def_info = eliminate_conjuncts_from_definition(def_info)
    sdefinition = def_info[0][5]

    # we now must match the definite variables in the definition to the definite variables
    # already assigned

    add_abbreviations(const_in_def, match_dv, match_type, new_var)

    determ_loc, new_var2, ovar, possessive_nouns = get_match_dv_from_determ(definiendum,
                                                                            k, kind, m,
                                                                            match_dv,
                                                                            match_type,
                                                                            new_var,
                                                                            definite_assignments)

    defin_sent = []
    rename = []
    temp_te = []
    heir_num = []
    spec_var = ['y', 'x', 'w']
    rule_found = False
    def_con = ""

    univ_quant_sent = eliminate_sent_wi_univ_scope(definiendum, determ_loc, k, m, new_var2)

    if kind == "determinative" or kind == "pronoun" or kind == 'poss pro':
        rule = "DE " + definiendum
        rule_found = True
    elif kind == "proper name possessive":
        rule = "PNP"
        kind = "determinative"
        rule_found = True

    z = -1
    for i in range(len(def_info[0][0])):
        if i == 21:
            bb = 8
        if def_info[0][4][i][1] == iff and not rule_found:
            rule = "DF " + definiendum
            rule_found = True
            def_con = iff
            if already_df and kind == "":
                break

        elif def_info[0][4][i][1] == conditional and not rule_found:
            def_con = conditional
            rule = "NC " + definiendum
            rule_found = True
            if already_df and kind == "":
                break
        if os(def_info[0][3][i]) == True:
            if not already_df:
                temp_str = space_words(def_info[0][3][i])
                temp_str = temp_str.replace("(", "")
                temp_str = temp_str.replace(")", "")
                list8 = prepare_categorize_words(temp_str)
                list8[79] = 'is in definition'
                list8 = categorize_words(list8, taken_out)


                if kind != "":
                    telist8 = copy.deepcopy(list8)
                    temp_te.append(telist8)
            else:
                z += 1
                list8 = copy.deepcopy(already_defined[ad][2][z])

            bool1 = False
            bool2 = False

            if kind != "" and kind != 'R' and list8[9] == "R":
                list8[73] = True
                if list8[3] != None:
                    temp_det = list8[3]
                    has_detrm = True
                else:
                    has_detrm = False
                if ident_det:
                    neg1 = list8[8]
                    if list8[5] == 'b':
                        bool1 = True
                    if list8[5] == 'z':
                        bool2 = True
                str2 = ''

                list8 = build_determ_definiens(determ_loc,
                        kind, list8, m, match_dv, match_type, spec_var, str2)


                # for the determinatives which have negations in their definition then we need
                # to do something special
                list1 = new_categories(list8, True)
                list8[46] = list1[46]
                list8[56] = list1[56]

                if ident_det:
                    if determ_loc == 5 or determ_loc == 14:
                        list8[8] = neg1
                    elif determ_loc == 18 and neg1 == "~":
                        list8[49] = neg1
                    elif determ_loc == 22 and neg1 == "~":
                        list8[50] = neg1
                    elif determ_loc == 26 and neg1 == "~":
                        list8[51] = neg1
                        # the determinatives which have an identity statement in them behave differently
                        # these are 'only' and 'anything except'
                if bool1:
                    list8[determ_loc] = ovar
                if has_detrm:
                    list8[k] = temp_det

                if bool2:
                    str2 = findinlist("z", match_dv, 0, 1)
                    list8[determ_loc] = str2

                list8 = build_sent2(list8)
                sdefinition = sdefinition.replace(def_info[0][6][i], list8[0])
            else:
                list8 = build_sent2(list8)
                str1 = list8[0]
            # if a def sent is part of the defiendum then it does not have to be added to the all
            # all_sent list
            if def_info[0][4][i][0][:-1] == "11" or def_info[0][4][i][0] == "111":
                list8[40] = True
            else:
                list8[40] = False

            list8[68] = def_info[0][4][i][0]
            list8[44] = def_info[0][6][i]
            heir_num.append(def_info[0][4][i][0])
            for s in range(0, 3):
                list8.append(None)
            defin_sent.append(list8)
    if not already_df:
        if kind == "":
            list9 = copy.deepcopy(defin_sent)
            temp_ad.append(list9)
            tem_he_num = copy.deepcopy(heir_num)
            temp_ad.append(tem_he_num)
        else:
            temp_ad.append(temp_te)
        list10 = copy.deepcopy(temp_ad)
        already_defined.append(list10)

    if already_df and kind == "":
        defin_sent = copy.deepcopy(already_defined[ad][2])
        heir_num = copy.deepcopy(already_defined[ad][3])

    match_dv = match_def_sub_to_all_sent_sub(defin_sent,
                                             kind, m, match_dv,
                                             match_type, definition)

    num2 = [5, 14, 18, 22, 26, 30, 34]
    num3 = [9, 14, 8]
    num4 = [9, 5, 8]
    # the point of the exception list is that we do not change certain sentences in the
    # definiens if we are analyzing a pronoun or determinative
    already_checked2 = []
    unmatched = []
    temp_match = []


    defin_sent = sorted(defin_sent, key=operator.itemgetter(68))
    prop_pos = []  # positions of propositional constants, if any

    for i in range(len(defin_sent)):
        if i == 3:
            bb = 8
        if defin_sent[i][73] == None:
            for j in num2:
                temp_str = defin_sent[i][j]
                isvar = isvariable(temp_str)
                if isvar:
                    if temp_str != None:
                        str3 = findinlist(temp_str, match_dv, 0, 1)
                        if temp_str == "i":
                            pass
                        elif str3 != None and temp_str != str3:
                            already_checked2.append([i, j])
                            defin_sent[i][j] = str3
                            str4 = repl_sign(str3, match_dv, match_type)
                            str2 = '(' + temp_str + str4 + str3 + ')'
                            # ("sub case 1")
                            if str2 not in rename and str2 != "":
                                rename.append(str2)

                        elif is_in_md(propositional_constants, 0, temp_str):
                            if defin_sent[i][2] != mini_e:
                                prop_pos.append([i, j])
                              #  print ("sub case 2")
                        elif temp_str == str3:
                            already_checked2.append([i, j])
                            # print ("sub case 3")
                        # elif defin_sent[i][j] == rr_var:
                        #     # print ("sub case 7")
                        #     match_dv, match_type, defin_sent, temp_match, rename = abb_change2(match_dv,
                        #                                                                        match_type,
                        #                                                                        defin_sent,
                        #                                                                        i,
                        #                                                                        temp_match,
                        #                                                                        j,
                        #                                                                        rename)
                        else:
                            if i == 6 and m == 17:
                                bb = 8
                            already_checked = []
                            if j == 5:
                                list5 = num3
                            else:
                                list5 = num4
                            no_match, defin_sent = abb_change(list5, already_checked,
                                                              i, match_dv, match_type, rename, j,
                                                              defin_sent, def_con)
                            if not no_match and j == 14 and unmatched != []:
                              #  print ("sub case 8")
                                no_match, defin_sent = abb_change(num3, already_checked,
                                                                  i, match_dv, match_type, rename, j,
                                                                  defin_sent, def_con)
                            elif no_match:
                                unmatched.append([i, j])
                            else:
                                # print ("sub case 9")
                                bb = 8
    if unmatched != []:
        new_match = []
        unmatched2 = []
        already_checked = []
        for k in range(len(unmatched)):
            if k == 5:
                bb = 8
            i = unmatched[k][0]
            j = unmatched[k][1]
            if j == 5:
                # sometimes an unmatched variable will appear twice and the second time
                # it needs to take the same variable as it did the first time
                str3 = findinlist(defin_sent[i][j], match_dv, 0, 1)
                if str3 != None:
                 #   print ("sub case 10")
                    defin_sent[i][j] = str3
                else:
                    # yyy
                 #   print ("sub case 11")
                    no_match, defin_sent = abb_change(num3, already_checked, i,
                                                      match_dv, match_type, rename, 5, defin_sent,
                                                      def_con, new_match, True)
                    if no_match:
                  #      print ("sub case 12")
                        abb_change2(match_dv, match_type, defin_sent, i, temp_match, j, rename)
                        unmatched2.append([i, j])
            else:
                temp_str = defin_sent[i][j]
                str3 = findinlist(temp_str, match_dv, 0, 1)
                if str3 != None and temp_str != str3:
                   # print ("sub case 13")
                    defin_sent[i][j] = str3
                    str2 = '(' + temp_str + idd + str3 + ')'
                    str2 = str2 + l2
                    if str2 not in rename and str2 != "":
                        rename.append(str2)
                else:
                    if defin_sent[i][j] not in taken_out:
                     #   print ("sub case 14")
                        abb_change2(match_dv, match_type, defin_sent, i, temp_match, j, rename)
                    unmatched2.append([i, j])

    defin_sent, rename = change_propositional_constants(defin_sent,
                                match_dv, prop_pos, propositional_constants, rename)

    rule = variable_change(rule, match_dv)


    def_info[0], defin_sent = insert_sentence_into_antecedent(def_info[0],
                                            defin_sent, univ_quant_sent)


    def_info[0], defin_sent = insert_possessive_nouns(def_info[0],
                                        defin_sent, possessive_nouns)


    defin_sent = eliminate_negated_many(defin_sent, definiendum, k)

    for sent in defin_sent:
        sent = build_sent2(sent)

    defin_sent = add_first_sent_to_def_sent(defin_sent, first_sent)

    add_to_attach_sent(def_info, defin_sent, definition, kind, rename, rule)

    add_def_sent_to_all_sent(definiendum, defin_sent)

    divide_univ_quant_sent(univ_quant_sent)

    update_essential_properties_dictionary(first_sent, definiendum, defin_sent)

    c = time.time()
    time_spent_defining += (c-b)

    # end2


def add_to_attach_sent(def_info, defin_sent, definition, kind, rename, rule):

    for i in range(len(def_info)):
        list1 = prepare_attach_sent(def_info[i], defin_sent, kind)
        if list1[45] == "append to attach_sent list":
            attach_sent.append(list1)
        if i == 0:
            add_definitions_to_total_sent(list1, rule, rename, kind, definition)
        if def_info[i][2] == 'eliminate as conjunct':
            list1[46] = 'eliminate as conjunct'


def update_essential_properties_dictionary(first_sent, definiendum, defin_sent):
    if first_sent[9] == "I" and isinmdlist(first_sent[14], abbreviations, 0) > -1:
        essential_properties = words[38]
        subject = first_sent[5]
        num = [5, 14]
        defin_sent2 = copy.deepcopy(defin_sent)
        for i in range(1, len(defin_sent2)):
            for j in num:
                if defin_sent2[i][j] == subject:
                    defin_sent2[i][j] = alpha
                elif not isinmdlist(defin_sent2[i][j], abbreviations, 0):
                    defin_sent2[i][j] = beta

        ess_prop = []
        for i in range(1, len(defin_sent2)):
            sent = build_sent3(defin_sent2[i])
            sent = sent.replace(" ", "")
            ess_prop.append(sent)

        essential_properties[definiendum] = ess_prop


def divide_univ_quant_sent(univ_quant_sent):
    # we perform this step because divide relations is the only step that we do
    # both before and after the universals are defined
    if univ_quant_sent != []:
        num100 = [15, 19]
        m = len(all_sent)-len(univ_quant_sent)
        n = len(all_sent)
        m -= 1
        while m < n-1:
            m += 1
            for i in num100:
                if i in all_sent[m][46]:
                    divide_relations2(m, i)
                    break


def change_propositional_constants(defin_sent, match_dv, prop_pos, propositional_constants, rename):
    if propositional_constants != [] and prop_pos != []:  # here we replace propositional constants
        match_dv2 = []
        num = [5, 14]
        done = []
        bool1 = False
        for i in range(len(propositional_constants)):
            for j in num:
                str1 = propositional_constants[i][2][j]
                g = findposinlist(str1, match_dv, 0)
                if g > -1:
                    bool1 = True
                    propositional_constants[i][2][j] = match_dv[g][1]
            if bool1:
                propositional_constants[i][2] = build_sent2(propositional_constants[i][2])
                str1 = propositional_constants[i][2][0]
                str1 = str1.replace(" ", "")
                str1 = remove_outer_paren(str1)
                propositional_constants[i][1] = str1
                g = findposinlist(str1, abbreviations, 1)
                str3 = propositional_constants[i][0]
                if g > -1:
                    print ("prop1")
                    str2 = abbreviations[g][0]
                    propositional_constants[i][0] = str2
                    match_dv2.append([str3, propositional_constants[i][0]])
                    str4 = "(" + str3 + idd + propositional_constants[i][0] + ")"
                    rename.append(str4)
                g = findposinlist(str3, abbreviations, 0)
                if g > -1:
                    str5 = abbreviations[g][1]
                    if str5 != str1:
                        str3 = variables[0]
                        del variables[0]
                        match_dv2.append([propositional_constants[i][0], str3])
                        str4 = "(" + propositional_constants[i][0] + idd + str3 + ")"
                        rename.append(str4)

                abbreviations.append([str3, str1, 1])
                propositional_constants[i][0] = str3
                done.append(i)

        if match_dv2 != []:
            for i in range(len(prop_pos)):
                j = prop_pos[i][0]
                o = prop_pos[i][1]
                str1 = defin_sent[j][o]
                str2 = findinlist(str1, match_dv2, 0, 1)
                defin_sent[j][o] = str2

    return defin_sent, rename




def get_match_dv_from_determ(definiendum, k, kind, m, match_dv,
                             match_type, new_var, definite_assignments):

    determ_loc = 99  # currently this is only used in the definition of many
    possessive_nouns = []
    new_var2 = ""
    syn_det = ["no_one_except", "any" + un]
    ovar = ""
    if kind == 'pronoun':
        match_type.append(9)
        if definiendum != 'i':
            str1 = findinlist(definiendum, abbreviations, 1, 0)
            if str1 == None:
                all_sent[m][k] = variables[0]
                match_dv.append(["c", variables[0]])
                abbreviations.append([variables[0], definiendum])
                new_var.append(variables[0])
                del variables[0]
            else:
                all_sent[m][k] = str1
                match_dv.append(["c'", str1])
        else:
            match_dv.append(['i', 'i'])
            # when constructing definitions of personal pronouns or of determinatives the object of the IG relation
            # must be b and the subject must be z
    elif kind == 'determinative' or kind == 'poss pro' or kind == 'proper name possessive':
        if k == 10:
            determ_loc = 14
        else:
            determ_loc = k + 2
        ovar = all_sent[m][determ_loc]
        match_type.append(9)
        if kind == "proper name possessive":
            match_dv.append(["b", all_sent[m][k]])
        elif definiendum == "its" + ua or definiendum == "its" + ub:  # its is slightly weird because it almost never exists
            # in the subject position
            match_dv.append(["c", all_sent[m][14]])
            all_sent[m][k] = ""
            match_dv.append(["b", all_sent[m][5]])
            match_type.append(9)
        else:
            all_sent[m][k] = ""
            match_dv.append(["b", all_sent[m][determ_loc]])
        if definiendum == 'the' or definiendum == 'that' + ud:
            str1 = all_sent[m][determ_loc]
            str3 = findinlist(str1, abbreviations, 0, 1)
            str2 = findinlist(str3, definite_assignments, 1, 0)
            match_type.append(9)
            if str2 == None:
                match_dv.append(["z", variables[0]])
                definite_assignments.append([variables[0], str3])
                if kind != 'proper name possessive':
                    all_sent[m][determ_loc] = variables[0]
                else:
                    all_sent[m][k] = variables[0]
                new_var.append(variables[0])
                new_var2 = variables[0]
                del variables[0]
            else:
                # definite_list used here
                all_sent[m][determ_loc] = str2
                match_dv.append(["z", str2])
                new_var2 = str2
        elif definiendum not in syn_det:
            match_type.append(9)
            new_var2 = variables[0]
            all_sent[m][determ_loc] = variables[0]
            match_dv.append(["z", variables[0]])
            new_var.append(variables[0])
            del variables[0]
        if determ_loc == 14 and all_sent[m][70] != None and kind != 'proper name possessive':
            possessive_nouns = eliminate_possessive_nouns(m, 70, definite_assignments, definiendum)
        if determ_loc == 5 and all_sent[m][69] != None and kind != 'proper name possessive':
            possessive_nouns = eliminate_possessive_nouns(m, 69, definite_assignments, definiendum)

    return determ_loc, new_var2, ovar, possessive_nouns


def add_first_sent_to_def_sent(defin_sent, first_sent):
    for i in range(len(defin_sent)):
        if defin_sent[i][68][1] == "1":
            greek_name = defin_sent[i][44]
            del defin_sent[i]
            break
    first_sent[44] = greek_name
    defin_sent.insert(0, first_sent)

    return defin_sent


def build_determ_definiens(determ_loc, kind, list8, m, match_dv, match_type, spec_var, str2):
    for p in range(2, 71):
        # if the variable in the original definition is z,y,x,w then that must
        # go into the new definition in its proper place
        if p == 46:
            bb = 8
        if list8[p] in spec_var:
            str2 = variables[0]
            spec_var.remove(list8[p])
            match_dv.append([list8[p], str2])
            match_type.append(9)
            del variables[0]
        if p == determ_loc and str2 != "" and str2 != None \
                and kind == 'determinative':
            list8[p] = str2
        elif p != 46 and p != 56:
            list8[p] = all_sent[m][p]

    return list8


def eliminate_sent_wi_univ_scope(definiendum, determ_loc, k, m, new_var2):
    univ = ['every', 'no']
    special_set = ['a', 'many' + up, 'many' + us, 'many' + ud, "a" + ua]
    bool1 = False
    univ_quant_sent = []
    if definiendum in univ or definiendum in special_set:
        if definiendum in special_set or definiendum in univ:
            if all_sent[m][determ_loc - 1] != None:
                list1 = eliminate_adjective_wi_universal_sent(m, determ_loc, new_var2, 0)
                univ_quant_sent.append(list1)
        if definiendum in univ:
            if k == 3 and all_sent[m][59] != None:
                bool1 = True
                all_sent[m][59] = None
            elif k == 10 and all_sent[m][60] != None:
                bool1 = True
                all_sent[m][60] = None
            elif k == 16 and all_sent[m][61] != None:
                bool1 = True
                all_sent[m][61] = None
            elif k == 20 and all_sent[m][62] != None:
                bool1 = True
                all_sent[m][62] = None
            if bool1:
                list2 = eliminate_univ_quant_subclause(m, k, determ_loc, new_var2)
                for i in range(len(list2)):
                    univ_quant_sent.append(list2[i])
            if all_sent[m][15] != None:
                new_relat = all_sent[m][9]
                new_obj = all_sent[m][14]
                all_sent[m][9] = all_sent[m][15]
                all_sent[m][15] = None
                all_sent[m][10] = all_sent[m][16]
                all_sent[m][16] = None
                all_sent[m][14] = all_sent[m][18]
                all_sent[m][18] = None
                if new_relat != 'EX':
                    list1 = eliminate_adjective_wi_universal_sent2(new_var2, new_relat, new_obj)
                    univ_quant_sent.append(list1)

    return univ_quant_sent


def add_abbreviations(const_in_def, match_dv, match_type, new_var):
    global abbreviations
    list1 = []
    for i in range(len(const_in_def)):
        temp_str = const_in_def[i][1]
        for j in range(len(abbreviations)):
            dvn_temp = abbreviations[j][1]
            if dvn_temp == temp_str:
                list8 = [const_in_def[i][0], abbreviations[j][0]]
                if list8 not in match_dv:
                    match_dv.append(list8)
                    match_type.append(1)
                    break
        else:
            if const_in_def[i][0] not in variables:
                list8 = [const_in_def[i][0], variables[0]]
                match_dv.append(list8)
                match_type.append(2)
                list1.append([variables[0], temp_str])
                new_var.append(variables[0])
                del variables[0]
            else:
                list8 = [const_in_def[i][0], temp_str]
                list1.append(list8)
                match_dv.append([const_in_def[i][0], const_in_def[i][0]])
                variables.remove(const_in_def[i][0])
                match_type.append(9)

    abbreviations += list1


def insert_possessive_nouns(def_info0th, defin_sent, possessive_nouns):
    # here we replace the possessive conjunction with the first sentence in the definiens

    if possessive_nouns != []:
        greek_conjunction = possessive_nouns[0][44] + " & " + possessive_nouns[1][44] + " & "

        for i in range(1, len(def_info0th[4])):
            if def_info0th[4][i][0][1] == "2" and def_info0th[4][i][1] == "&":
                str1 = def_info0th[6][i][1]
                def_info0th[6][i] = def_info0th[6][i].replace(str1, \
                                                              greek_conjunction + str1)
                def_info0th[6][0] = def_info0th[6][0].replace(str1, \
                                                              greek_conjunction + str1)
                def_info0th[5] = def_info0th[6][0]
                break

        last_num = int(def_info0th[4][-1][0])
        last_num1 = last_num + 1
        last_num2 = last_num + 2

        list1 = [possessive_nouns[0][72], possessive_nouns[0][2], possessive_nouns[0][0], \
                 [str(last_num1), "", ""], possessive_nouns[0][44]]

        num = [0, 1, 3, 4, 6]
        j = 0
        for i in num:
            def_info0th[i].append(list1[j])
            j += 1

        list1 = [possessive_nouns[1][72], possessive_nouns[1][2], possessive_nouns[1][0], \
                 [str(last_num2), "", ""], possessive_nouns[1][44]]

        j = 0
        for i in num:
            def_info0th[i].append(list1[j])
            j += 1

        defin_sent.append(possessive_nouns[0])
        defin_sent.append(possessive_nouns[1])

    return def_info0th, defin_sent


def match_def_sub_to_all_sent_sub(defin_sent, kind, m, match_dv, match_type, definition):
    if (kind == "" or kind == "R"):
        for sent in defin_sent:
            if sent[68][1] == "1" and sent[68][-1] == "1":
                match_dv.append([sent[5], all_sent[m][5]])
                match_type.append(0)
                # for definitions where I is the relation then we need not worry about the subject
                if kind == "R":
                    match_dv.append([sent[14], all_sent[m][14]])
                    match_type.append(0)
                    if len(sent[68]) == 2:
                        # if a definiendum is a conjunction then len of the sent num for the individual
                        # conjuncts must be greater than 2
                        break
            elif sent[68][1] == "1" and len(sent[68]) == 3:
                nobj = findinlist(sent[14], match_dv, 0, 1)
                # these are sentences which are in the definiendum but are not first
                # their subjects and objects must match a variable in the first sentence
                # also they have the form (bRc) & (cQd) <> (dSe)
                if nobj != None:
                    for lst in all_sent:
                        if lst[9] == sent[9] and lst[14] == nobj:
                            print (definition)
                            match_dv.append([sent[5], lst[5]])
                            match_type.append(0)
                            break

    return match_dv



def add_definitions_to_total_sent(temp_attach_sent, rule, rename, kind, definition):
    if is_non_sub_def(kind) or rename == []:
        add_to_total_sent(temp_attach_sent[2], temp_attach_sent[37], temp_attach_sent[4], "", rule)
    else:
        num = temp_attach_sent[2]
        add_to_total_sent(num - 2, definition, "", "", rule)
        add_to_total_sent(num - 1, build_sent_list(rename), "", "", "RN")
        add_to_total_sent(num, temp_attach_sent[37], temp_attach_sent[4], "", "SUB")


def is_non_sub_def(kind):
    if kind == "" or kind == "R":
        return False
    else:
        return True


def eliminate_negated_many(defin_sent, definiendum, k):

    if definiendum == "many" + un:
        for i in range(len(defin_sent)):
            if not defin_sent[i][40] and defin_sent[i][8] == "~":
                defin_sent[i] = eliminate_not_a(defin_sent[i], k)
                return defin_sent

    return defin_sent

def add_def_sent_to_all_sent(definiendum, defin_sent):

    for i in range(len(defin_sent)):
        # we cannot add the first_sent to the all sent list since it is already in there
        if not isinmdlist(defin_sent[i][1], all_sent, 1) and not \
                is_in_abbreviations(defin_sent[i]) and i != 0:
            all_sent.append(defin_sent[i])


def insert_sentence_into_antecedent(def_info0th, defin_sent, univ_quant_sent):
    # this takes some sentences which were universally quantified
    # and inserts them into the antecedent within the definiens of
    # either 'every' or 'no'
    # not that for now the only definitions which require this functionality
    # are 'every' and 'no'.  for this reason we can use the sentence number '121'
    # as a guide, but if we should acquire definitions which have a different
    # structure then this will have to change

    if univ_quant_sent == []:
        return def_info0th, defin_sent

    nat_replacer = ""
    for sent in univ_quant_sent:
        nat_replacer += " & " + sent[0]
    d = findposinmd_alert_error("121", def_info0th[4], 0)
    nat_replacee = def_info0th[3][d]
    nat_replacer = "(" + nat_replacee + nat_replacer + ")"
    new_nat = def_info0th[3][0].replace(nat_replacee, nat_replacer)
    def_info0th = find_sentences(new_nat)
    defin_sent = change_greek_in_univ_quant_sent(univ_quant_sent, defin_sent, def_info0th)

    return def_info0th, defin_sent


def change_greek_in_univ_quant_sent(univ_quant_sent, defin_sent, def_info0th):
    for sent in defin_sent:
        if sent[68] == '121':
            for i in range(len(def_info0th[3])):
                if def_info0th[4][i][0] == "1211":
                    sent[44] = def_info0th[6][i]
                    break
        elif sent[68] == "122":
            for i in range(len(def_info0th[3])):
                if def_info0th[4][i][0] == "122":
                    sent[44] = def_info0th[6][i]
                    break

    for sent in univ_quant_sent:
        for i in range(len(def_info0th[3])):
            if def_info0th[3][i] == sent[0]:
                sent[44] = def_info0th[6][i]
                defin_sent.append(sent)
                break
        else:
            g = 4 / 0
    return defin_sent


def determine_what_is_conjunct(def_info):
    #modify this if definitions have v or xor as a main connective in their consequent
    ant_conj = 0
    con_conj = 0
    for i, num in enumerate(def_info[4]):
        if len(num[0]) == 2 and num[0][1] == '1':
            if num[1] == "&":
                ant_conj = 3
            else:
                ant_conj = 2
        elif len(num[0]) == 2 and num[0][1] == '2':
            if num[1] == "&":
                con_conj = 3
            else:
                con_conj = 2
        if ant_conj != 0 and con_conj != 0:
            return ant_conj, con_conj


def prepare_attach_sent(def_info, defin_sent, kind):
    # this populates the attach sent list

    global sn
    if is_non_sub_def(kind):
        sn += 1
    elif def_info[2] == 'eliminate as conjunct':
        pass
    else:
        sn += 3
    list1 = [""] * 50
    list1[2] = sn
    greek_sent = def_info[5]
    list2 = translate_complex_sent(greek_sent, defin_sent)
    list1[4] = list2[1]
    list1[37] = list2[0]
    list1[47] = greek_sent
    ant_parts = []
    con_parts = []
    ant_variables = []
    con_variables = []
    ant_conjunction = ""
    con_conjunction = ""
    spec_conn = [iff, conditional, idisj, xorr]
    total_sent_in_attach_sent = []
    embed_att_sent = []
    isdisjunction = False
    isconjunction = False
    if def_info[4][0][1] == iff:
        list1[3] = "e"
    elif def_info[4][0][1] == conditional:
        list1[3] = "c"
    elif def_info[4][0][1] == idisj:
        list1[3] = "d"
        isdisjunction = True
    elif def_info[4][0][1] == xorr:
        list1[3] = "x"
        isdisjunction = True
    else:
        isconjunction = True

    if isdisjunction:
        list1[36] = copy.deepcopy(def_info)
        list1[44] = copy.deepcopy(defin_sent)
        list1[45] = "do not append to attach_sent list"
    elif isconjunction:
        list1[45] = "do not append to attach_sent list"
    else:
        ant_conj, con_conj = determine_what_is_conjunct(def_info)
        list1[45] = "append to attach_sent list"
        list3 = mainconn(greek_sent)
        main_conn_location = list3[1]
        for i in range(1, len(def_info[0])):
            t_value = def_info[1][i]
            d = 99
            if def_info[4][i][1] == "":
                d = findposinmd(def_info[6][i], defin_sent, 44)  # d = index of def sent
                defin_sent[d] = ancestor_numbers(defin_sent[d], def_info[4][i][0], def_info)
                if defin_sent[d][53][-1] == "a" or defin_sent[d][53][-1] == "b":
                    ant_parts.append(copy.deepcopy(defin_sent[d]))
                else:
                    con_parts.append(copy.deepcopy(defin_sent[d]))
                total_sent_in_attach_sent.append(defin_sent[d][1])
                t_value = defin_sent[d][2]
                defin_sent[d][68] = def_info[4][i][0]


            if def_info[4][i][0][1] == '1' and len(def_info[4][i][0]) == ant_conj:
                ant_variables.append([def_info[6][i], t_value])
            elif def_info[4][i][0][1] == '2' and len(def_info[4][i][0]) == con_conj:
                con_variables.append([def_info[6][i], t_value])

            if def_info[4][i][1] != "" and len(def_info[4][i][0]) == 2:
                if def_info[4][i][0][1] == '1' and ant_conjunction == "":
                    ant_conjunction = def_info[6][i]

                elif def_info[4][i][0][1] == '2' and con_conjunction == "":
                    con_conjunction = def_info[6][i]




            if def_info[4][i][1] in spec_conn:
                embed_att_sent.append(i)

        list1 = prepare_attach_sent2(ant_conjunction, con_conjunction, ant_variables, \
                                     con_variables, def_info, defin_sent,
                                     embed_att_sent, list1)

        list1[34] = ant_parts
        list1[35] = con_parts
        list1[38] = total_sent_in_attach_sent

    return list1


def put_in_0_or_1(def_info, i, defin_sent, d):
    if len(def_info[4][i][0]) == 2 and def_info[4][i][1] != "&":
        return True
    elif os(def_info[3][i]):
        if len(defin_sent[d][53]) == 2 and defin_sent[d][53][0] == 'c':
            return True
        elif len(def_info[4][i][0]) == 3 and defin_sent[d][53][0] == 'c':
            return True
        else:
            return False
    else:
        return False


def prepare_attach_sent2(ant_conjunction, con_conjunction, ant_variables,
                         con_variables, def_info, defin_sent,
                         embed_att_sent, list1):
    global sn
    embed_info = []
    list2 = translate_list_of_sentences(ant_variables, defin_sent)
    list1[0] = list2[1]
    list1[42] = list2[0]

    if ant_conjunction != "":
        list2 = translate_complex_sent(ant_conjunction, defin_sent)
        list1[40] = [list2[0], ""]
        list1[7] = [list2[1], ""]
    else:
        list1[40] = list2[0][0]
        list1[7] = list2[1][0]

    list2 = translate_list_of_sentences(con_variables, defin_sent)
    list1[1] = list2[1]
    list1[43] = list2[0]

    if con_conjunction != "":
        list2 = translate_complex_sent(con_conjunction, defin_sent)
        list1[41] = [list2[0], ""]
        list1[8] = [list2[1], ""]
    else:
        list1[41] = list2[0][0]
        list1[8] = list2[1][0]

    if embed_att_sent != []:
        for num in embed_att_sent:
            list4 = prepare_embed_att_sent(def_info, defin_sent, num)
            embed_info.append(list4)
            sn -= 1

        list1[39] = embed_info

    return list1


def prepare_embed_att_sent(def_info, defin_sent, num):
    # this takes those attached sentences within attached sentences and
    # prepares them to be manipulated

    def_info2 = copy.deepcopy(def_info)
    sent_num = def_info[4][num][0]
    def_info2 = delete_irrel_sent(def_info2, sent_num)
    def_info2[5] = def_info2[6][0]
    pos_of_new_num = len(sent_num) - 1
    main_conn_loc = def_info2[3][0].find(def_info2[4][0][1])
    def_info2[4][0][2] = main_conn_loc
    def_info2 = renumber_embed_sent(def_info2, pos_of_new_num)
    list1 = prepare_attach_sent(def_info2, defin_sent, "EMBED")

    return list1


def renumber_embed_sent(def_info2, pos_of_new_num):
    for position in def_info2[4]:
        position[0] = position[0][pos_of_new_num:]

    return def_info2


def delete_irrel_sent(def_info2, sent_num):
    # this deletes sentences from the def_info list
    # which are not members of the embedded sentence
    # under investigation

    i = 0
    while i < len(def_info2[4]):

        if not def_info2[4][i][0].startswith(sent_num):
            del def_info2[0][i]
            del def_info2[1][i]
            del def_info2[3][i]
            del def_info2[4][i]
            del def_info2[6][i]
        else:
            i += 1

    return def_info2


def ancestor_numbers(list2, k, def_info):
    # this determines the number and connective of the ancestors of a
    # sent in the conditional
    list2[53] = None
    self_num = k[-1]
    if len(k) == 4:
        ggparen_num = k[0]
        gparen_num = k[:2]
        paren_num = k[:3]
        ggparen_conn = findinlist(ggparen_num, def_info[4], 0, 1)
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        ggparen_conn = convert_con_to_letter(ggparen_conn, gparen_num[-1])
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn + ggparen_conn

    elif len(k) == 3:
        gparen_num = k[0]
        paren_num = k[:2]
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn

    elif len(k) == 2:
        paren_num = k[0]
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn

    elif len(k) == 5:
        gggparen_num = k[0]
        ggparen_num = k[:2]
        gparen_num = k[:3]
        paren_num = k[:4]
        gggparen_conn = findinlist(gggparen_num, def_info[4], 0, 1)
        ggparen_conn = findinlist(ggparen_num, def_info[4], 0, 1)
        gparen_conn = findinlist(gparen_num, def_info[4], 0, 1)
        paren_conn = findinlist(paren_num, def_info[4], 0, 1)
        gggparen_conn = convert_con_to_letter(gggparen_conn, ggparen_num[-1])
        ggparen_conn = convert_con_to_letter(ggparen_conn, gparen_num[-1])
        gparen_conn = convert_con_to_letter(gparen_conn, paren_num[-1])
        paren_conn = convert_con_to_letter(paren_conn, self_num)
        list2[53] = paren_conn + gparen_conn + ggparen_conn + gggparen_conn

    elif len(k) == 6:
        print("you have not coded for attached sentences with 5 generations yet")
        sys.exit()

    if list2[53] == None:
        print("the number ancestor function is messed up")
        sys.exit()
    return list2


def convert_con_to_letter(str1, str2):
    # this converts a connective to a letter
    if str1 == iff and str2 == '1':
        return 'b'
    elif str1 == iff and str2 == '2':
        return 'f'
    elif str1 == conditional and str2 == '1':
        return 'a'
    elif str1 == conditional and str2 == '2':
        return 'q'
    elif str1 == xorr and str2 == '1':
        return 'x'
    elif str1 == xorr and str2 == '2':
        return 'y'
    elif str1 == idisj and str2 == '1':
        return 'd'
    elif str1 == idisj and str2 == '2':
        return 'g'
    elif str1 == idisj and str2 == '3':
        return 'd3'
    elif str1 == idisj and str2 == '4':
        return 'd4'
    elif str1 == idisj and str2 == '5':
        return 'd5'
    elif str1 == idisj and str2 == '6':
        return 'd6'


    elif str1 == "&":
        return 'c'
    else:
        print('the convert con to letter function is messed up')
        g = 4 / 0


def get_parent_connective(def_info, i):
    sent_number = def_info[4][i][0]
    for i in range(i - 1, -1, -1):
        if sent_number.startswith(def_info[4][i][0]):
            return def_info[4][i][1]
    else:
        print ('you failed to find the parent connective')
        g = 4 / 0


def translate_list_of_sentences(to_be_converted, defin_sent):
    # this converts the 0th or 1st member of the attach_sent_list into
    # sentence variables

    abbrev_sent = copy.deepcopy(to_be_converted)
    for i in range(len(to_be_converted)):
        if os(to_be_converted[i][0]):
            d = findposinmd_alert_error(to_be_converted[i][0], defin_sent, 44)
            to_be_converted[i][0] = defin_sent[d][72]
            abbrev_sent[i][0] = defin_sent[d][1]
        elif not os(to_be_converted[i][0]):
            list1 = translate_complex_sent(to_be_converted[i][0], defin_sent)
            to_be_converted[i][0] = list1[0]
            abbrev_sent[i][0] = list1[1]

    return [to_be_converted, abbrev_sent]


def translate_complex_sent(to_be_translated, defin_sent):
    # this converts a conjunction in a definition into the definition
    # with the new variables

    to_translate_abbrev = to_be_translated

    for sentence in defin_sent:
        if sentence[44] in to_be_translated:
            to_be_translated = to_be_translated.replace(sentence[44], sentence[0])
            to_translate_abbrev = to_translate_abbrev.replace(sentence[44], sentence[42])

    return [to_be_translated, to_translate_abbrev]


def variable_change(rule, match_dv):

    global sn
    for i in range(len(match_dv)):
        if match_dv[i][0] != match_dv[i][1]:
            return rule

    rule = rule[0] + "E" + rule[2:]
    sn += 1

    return rule

def eliminate_not_a(list1, k):
    num = [10, 16, 20, 24]

    for i in num:
        if i > k and (list1[i] == "a" or list1[i] == "a" + ua):
            ant_sent_parts = copy.deepcopy(list1)
            list2 = copy.deepcopy(list1)
            if list1[i] == "a":
                rule = "DE not a"
            else:
                rule = "DE not a" + ua
            list2[i] = 'every'
            list1[46] = [200]
            con_parts = copy.deepcopy(build_sent2(list2))
            prepare_att_sent_1_sent(ant_sent_parts, rule, "", "e", con_parts)
            all_sent.append(list1)
            return list2

    return list1


def prepare_categorize_words(str2):
    # when we prepare to categorize a word, the first three slots
    # are reserved for the sentence, the sentence letter and the tvalue

    str2 = str2.strip()
    list1 = str2.split(' ')
    list2 = [None] * 80
    j = 2
    for i in range(len(list1)):
        j += 1
        list2[j] = list1[i]

    return list2


def categorize_words(list1, taken_out=[]):

    sentence_slots = [None] * 80
    relation_type = 0
    slots_used = []
    slots_used2 = [[],[],[],[],[],[],[],[],[]]
    slots_used3 = [[]]*12

    slots_determinatives = [3, 10, 16, 20, 24, 28, 32]
    slots_sec_reduction = [67, 5, 14, 63, 18, 64, 22, 65, 26, 30, 34, 3, 10, 16, 20, 24, 28, 32]
    slots_adjectives = [4, 13, 17, 21, 25, 29, 33]
    slots_cia = [35,36]
    slots_relative_pronouns = [59, 60, 61, 62]
    slots_subordinators = [7, 59, 60, 61, 62]
    slots_relation_division = [15,19]
    slots_word_sub = [4, 5, 13, 14, 17, 18, 22, 26, 30,
                         34, 35, 36, 51, 52, 63, 64, 65, 67, 69, 70]
    decision = [200]
    spec_rel = ["I", "J"]
    noun_list = ['n', 'p', 'it']
    predicative_complement_positions = [14, 18, 22, 26, 29]

    i = 2
    while list1[i + 1] != None:
        i += 1
        k = 0
        d = 300
        word = list1[i]

        if word == 'that'+uc:
            bb = 8

        i, word, has_comma = determine_if_compound_word(i, list1, word)

        pos = get_part_of_speech(list1, taken_out, word)

        if word != ' ' and word != "":
            if word not in words_used:
                words_used.append(word)
            if (pos == 'd' or pos == 'q'): # q are possessive pronouns
                if relation_type == 0:
                    k = 3
                elif relation_type == 1:
                    k = 10
                elif relation_type == 2:
                    k = 16
                elif relation_type == 3:
                    k = 20
                elif relation_type == 4:
                    k = 24
                elif relation_type == 5:
                    k = 28
                elif relation_type == 6:
                    k = 32
                d = decision_for_determinatives(pos, word)

            elif pos == 'o':
                if relation_type == 0 and sentence_slots[5] == None:
                    k = 69
                elif relation_type == 1 and sentence_slots[14] == None:
                    k = 70
                d = decision_for_apostrophe_s(word)

            elif pos == 'a':
                if relation_type == 0:
                    k = 4
                elif relation_type == 1 and sentence_slots[9] == "J":
                    k = 14
                elif relation_type == 1:
                    k = 13
                elif relation_type == 2 and sentence_slots[15] == "J":
                    k = 18
                    # no decision needed here because it is defined as a concept
                elif relation_type == 2:
                    k = 17
                elif relation_type == 3 and sentence_slots[19] == "J":
                    k = 22
                elif relation_type == 3:
                    k = 21
                elif relation_type == 4 and sentence_slots[23] == "J":
                    k = 26
                elif relation_type == 4:
                    k = 25
                elif relation_type == 5 and sentence_slots[27] == "J":
                    k = 30
                elif relation_type == 5:
                    k = 29
                elif relation_type == 6 and sentence_slots[31] == "J":
                    k = 34
                elif relation_type == 6:
                    k = 33
                d = 200 if k in predicative_complement_positions else 50

            elif pos == 'm':
                if sentence_slots[3] == None and sentence_slots[5] == None and relation_type == 0:
                    k = 47
                elif relation_type == 0:
                    k = 8
                elif relation_type == 1 and sentence_slots[14] == None and \
                                sentence_slots[60] == None:
                    k = 12
                elif relation_type == 1 or sentence_slots[15] in spec_rel:
                    k = 49
                elif relation_type == 2 or sentence_slots[18] in spec_rel:
                    k = 50
                elif relation_type == 3 or sentence_slots[24] in spec_rel:
                    k = 51
                elif relation_type == 4 or sentence_slots[27] in spec_rel:
                    k = 52
                d = 200

            elif word == 'there':
                d = 110
                if relation_type == 0 and sentence_slots[5] == None:
                    k = 5
                elif relation_type == 1 and sentence_slots[60] != None:
                    k = 63
                elif relation_type == 2 and sentence_slots[61] != None:
                    k = 64

            elif pos in noun_list:
                d = 10 if pos == 'p' else 200
                if relation_type == 0:
                    if sentence_slots[66] != None:
                        k = 67
                    elif sentence_slots[5] != None:
                        k = 35
                    elif sentence_slots[5] == None:
                        k = 5
                elif relation_type == 1:
                    if sentence_slots[14] == None:
                        k = 14
                    elif sentence_slots[60] != None:
                        k = 63
                        print (63)
                    else:
                        k = 36
                elif relation_type == 2 and sentence_slots[18] == None:
                    k = 18
                elif relation_type == 2 and sentence_slots[61] != None:
                    k = 64
                    print (64)
                elif relation_type == 3 and sentence_slots[22] == None:
                    k = 22
                elif relation_type == 3 and sentence_slots[62] != None:
                    k = 65
                elif relation_type == 4:
                    k = 26
                elif relation_type == 5:
                    k = 30
                elif relation_type == 6:
                    k = 34

            elif pos == 'c':
                if relation_type == 0 and sentence_slots[5] != None:
                    k = 66
                    d = 40



            elif word == 'that' + uc and sentence_slots[7] == None and sentence_slots[14] == None:  # uc
                k = 7
                d = 80

            elif pos == 'u':
                d = 80 if word == 'that' + uc else 70
                if relation_type == 0 and sentence_slots[5] != None:
                    k = 59
                    d = 70
                elif relation_type == 1 and sentence_slots[14] != None:
                    k = 60
                elif relation_type == 2 and sentence_slots[18] != None:
                    k = 61
                elif relation_type == 3 and sentence_slots[22] != None:
                    k = 62

            elif pos == 'r':
                d = 200
                if relation_type == 0:
                    k = 9
                    relation_type = 1
                elif relation_type == 1:
                    k = 15
                    d = 100
                    relation_type = 2
                elif relation_type == 2:
                    relation_type = 3
                    k = 19
                    d = 100
                elif relation_type == 3:
                    relation_type = 4
                    k = 23
                elif relation_type == 4:
                    relation_type = 5
                    k = 27
                elif relation_type == 5:
                    k = 31
                    relation_type = 6

            if has_comma:
                sentence_slots[39] = k

            if k == 0:
                print("your decision procedure for determining the placement of " + word)
                (" in sentence: did not work")
                g = 4/0

            sentence_slots[k] = word
            slots_used.append(k)
            if d == 300:
                print ("you did not state the decision procedure for " + word)
                g = 4/0
            if d != 200: decision.append(d)

            # slots_used2 = get_used_slots(slots_used2, k,d)



    slots_used.sort()
    sentence_slots[46] = slots_used
    sentence_slots[56] = decision
    sentence_slots = build_sent2(sentence_slots)
    sentence_slots[54] = isdefineable(sentence_slots)

    return sentence_slots

def get_used_slots(slots_used2, k, d):

    possessive_pronouns = slots_used2[0]
    sec_reduction = slots_used2[1]
    adjectives = slots_used2[2]
    cia = slots_used2[3]
    relative_pronouns = slots_used2[4]
    subordinators = slots_used2[5]
    relation_division = slots_used2[6]
    universals = slots_used2[7]
    word_sub = slots_used2[8]





def get_part_of_speech(list1, taken_out, word):

    posp = words[28]
    if word == "":
        return

    if isvariable(word):
        pos = 'n'
        if word in variables:
            variables.remove(word)
            # taken_out.append(word)
            # if the variable stands for an adjective then its part of speech
            # is adjective
        str1 = findinlist(word, abbreviations, 0, 1)
        if str1 != None and not list1[79] == "is in definition":
            str3 = findinlist(str1, posp, 0, 1)
            if str3 == "a":
                pos = 'a'
    elif word == "~":
        pos = 'm'
    # right now 'it' is in a unique class
    elif word == "it":
        pos = 'it'
    elif word == ne:
        pos = 'r'
    elif word == 'not':
        pos = 'm'
    elif word[-2:] == "'s":
        pos = 'o'
    else:
        pos = findinlist(word, posp, 0, 1)

    if pos == None:
        print ("you misspelled " + word)
        g = 4/0

    return pos


def decision_for_determinatives(pos, word):
    spec_det = ['every', 'many' + un, 'no']
    if pos == 'd' and word not in spec_det:
        d = 20
    elif pos == 'd' and word in spec_det:
        d = 120
    elif pos == 'q':
        d = 85

    return d


def decision_for_apostrophe_s(word):
    proper_names = words[35]
    temp_word = word[:-2]
    if temp_word in proper_names:
        d = 30
    else:
        d = 90

    return d


def determine_if_compound_word(i, list1, word):

    if "," in word:
        word = word.replace(",", "")
        return i, word, True

    doubles = words[31]
    triples = words[32]
    bool4 = is_in_md(triples, 0, word)
    bool5 = False
    has_comma = False
    # if has_comma == "": has_comma = ""
    # if list1[i + 2] != None and bool4:
    #     str4 = word + " " + list1[i + 1] + " " + list1[i + 2]
    #     bool5 = check_dimension(triples, 1, str4)
    #     if bool5:
    #         word = str4
    #         if has_comma != "":
    #             has_comma = word
    compound_will_have_comma = False
    if list1[i + 1] != None:
        next_word = list1[i + 1]
        if "," in next_word:
            next_word = next_word.replace(",", "")
            compound_will_have_comma = True
        if is_in_md(doubles, 0, word) \
                and not has_comma and is_in_md(doubles, 1, word + " " + next_word):
            i += 1
            word += " " + next_word
            if compound_will_have_comma:
                has_comma = True

    return i, word, has_comma


def build_sent_name(prop_name):
    str1 = ''
    str2 = ''
    list1 = []

    for i in range(len(prop_name)):
        if i == 24:
            bb = 8
        str3 = remove_outer_paren(prop_name[i][2])
        str3 = str3.replace("~", "")
        str1 = '(' + prop_name[i][0] + mini_e + str3 + ')'
        if len(str2) == 0 and len(str1) > 57:
            list1.append(str1)
        elif (len(str2) + len(str1)) > 57:
            list1.append(str2)
            str2 = str1
            if i + 1 == len(prop_name):
                list1.append(str2)
        elif (len(str2) + len(str1)) <= 57:
            if len(str2) == 0:
                str2 = str1
            else:
                str2 = str2 + ' & ' + str1
            if i + 1 == len(prop_name):
                list1.append(str2)
    return list1


def replace_synonyms():

    # todo nothing is equivalent to no thing
    global sn, def_used
    doubles = words[31]
    bool1 = False
    synon = words[14]
    syn_pairs = words[13]
    m = -1
    while m < len(all_sent) - 1:
        m += 1
        ant_sent_parts = copy.deepcopy(all_sent[m])
        anc1 = ""
        anc2 = ""
        anc3 = ""
        i = 2
        while all_sent[m][i + 1] != None:
            i += 1
            str1 = all_sent[m][i]
            bool2 = is_in_md(doubles, 0, str1)
            bool3 = False
            if bool2 and all_sent[m][i + 1] != None:
                str4 = all_sent[m][i] + " " + all_sent[m][i + 1]
                bool3 = is_in_md(doubles, 1, str4)
                if bool3:
                    str1 += " " + all_sent[m][i + 1]
            if i == 9:
                bb = 7
            if str1 in synon:
                if str1 not in def_used:
                    def_used.append(str1)
                for j in range(len(syn_pairs)):
                    if str1 == syn_pairs[j][0]:
                        bool1 = True
                        rule = 'DE ' + str1
                        str5 = syn_pairs[j][2]
                        if " " in syn_pairs[j][1]:
                            str6 = syn_pairs[j][1]
                            g = str6.find(" ")
                            str3 = str6[:g]
                            str4 = str6[g + 1:]
                            list2 = copy.deepcopy(all_sent[m])
                            list2[i] = str4
                            list2.insert(i, str3)
                            all_sent[m] = list2
                        else:
                            all_sent[m][i] = syn_pairs[j][1]
                        u = findposinlist(str5, total_sent, 1)
                        if u == -1:
                            bool1 = True
                            s = findinlist(str5, total_sent, 1, 0)
                            sn += 1
                            if s == None:
                                s = sn
                            if anc1 == "":
                                anc1 = s
                            elif anc2 == "" and s != anc1:
                                w = copy.copy(s)
                                anc2 = w
                            elif anc3 == "" and s != anc1 and s != anc2:
                                u = copy.copy(s)
                                anc3 = u
                            add_to_total_sent(sn, str5, "", "", rule)
                        else:
                            s = total_sent[u][0]
                            if anc1 == "":
                                t = copy.copy(s)
                                anc1 = t
                            elif anc2 == "" and s != anc1:
                                w = copy.copy(s)
                                anc2 = w
                            elif anc3 == "" and s != anc1 and s != anc2:
                                u = copy.copy(s)
                                anc3 = u
                        if bool3:
                            del all_sent[m][i + 1]
                        break

        if bool1:
            all_sent[m] = build_uncategorized_sent(all_sent[m])
            con_parts = copy.deepcopy(all_sent[m])
            if anc3 != "":
                anc1 = str(anc1) + "," + str(anc2) + "," + str(anc3)
            elif anc2 != "":
                anc1 = str(anc1) + "," + str(anc2)
            prepare_att_sent_1_sent(ant_sent_parts, "SUB", anc1, "e", con_parts)
            bool1 = False


def quick_print(list1, list2, list3, kind=0, a=0, b=0, c=0):
    j = 1
    if kind == 0:
        for i in range(len(list2)):
            j += 1
            if list1 != []:
                w4.cell(row=j, column=2).value = list1[i]
            w4.cell(row=j, column=3).value = list2[i]
            if list3 != []:
                w4.cell(row=j, column=4).value = list3[i]
    elif kind == 1:
        for i in range(len(list2)):
            j += 1
            if a != 0:
                w4.cell(row=j, column=2).value = list2[i][a]
            w4.cell(row=j, column=3).value = list2[i][b]
            if c != 0:
                w4.cell(row=j, column=4).value = list2[i][c]

    wb4.save('../temp_proof.xlsx')
    sys.exit()


def print_sent_full(test_sent, tot_prop_name, yy=""):

    global result_data

    p = 2
    determine_words_used()
    o = -1

    for i in order:
        # if i == 2:
        #     break
        for j in range(len(test_sent[i])):

            if test_sent[i][j][6] != "":
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5]) + ',' + str(test_sent[i][j][6])
            elif test_sent[i][j][5] != "":
                str1 = test_sent[i][j][4] + ' ' + str(test_sent[i][j][5])
            elif test_sent[i][j][4] != "":
                str1 = test_sent[i][j][4]
            else:
                str1 = ""
            if j == 0:
                test_sent[i][j][4] = i
            else:
                test_sent[i][j][4] = str1
            if excel or normal_proof:
                w4.cell(row=p, column=2).value = test_sent[i][j][0]
                w4.cell(row=p, column=3).value = test_sent[i][j][3] + test_sent[i][j][1]
                w4.cell(row=p, column=4).value = test_sent[i][j][4]
            else:
                result_data['text_' + str(p - 1) + '_1'] = test_sent[i][j][0]
                result_data['text_' + str(p - 1) + '_2'] = test_sent[i][j][1]
                result_data['text_' + str(p - 1) + '_3'] = test_sent[i][j][4]

            p += 1

        p += 1
        o += 1
        list1 = build_sent_name(tot_prop_name[o])
        for j in range(len(list1)):
            if excel or normal_proof:
                w4.cell(row=p, column=3).value = list1[j]
            else:
                result_data['text_' + str(p) + '_2'] = list1[j]
            p += 1
        p += 1
        do_not_print = False
        for j in range(len(test_sent[i])):
            if j == 8:
                bb = 8
            if test_sent[i][j - 1][4] == 'ID':
                do_not_print = True
            elif test_sent[i][j][0] == "" and test_sent[i][j][1] == "":
                do_not_print = False
            if not do_not_print and test_sent[i][j][2] != "":
                if j == 0:
                    test_sent[i][j][4] == ""
                if excel or normal_proof:
                    w4.cell(row=p, column=2).value = test_sent[i][j][0]
                    w4.cell(row=p, column=3).value = test_sent[i][j][3] + test_sent[i][j][2]
                    w4.cell(row=p, column=4).value = test_sent[i][j][4]
                else:
                    result_data['text_' + str(p) + '_1'] = test_sent[i][j][0]
                    result_data['text_' + str(p) + '_2'] = test_sent[i][j][3] + test_sent[i][j][2]
                    result_data['text_' + str(p - 1) + '_3'] = test_sent[i][j][4]

                p += 1
        p += 3


def determine_words_used2():

    definitions2 = words[33]
    # for i in range(len(abbreviations)):
    #     if abbreviations[i][0] not in def_used:
    #         def_used.append(abbreviations[i][1])
    if get_words_used:
        for i in range(len(def_used)):
            j = findinlist(def_used[i], definitions2, 0, 1)
            if j != None:
                # ws.cell(row=j, column=2).value = 1
                bool2 = True
                while bool2:
                    j += 1
                    word2 = ws.cell(row=j, column=3).value
                    if word2 == def_used[i]:
                        ws.cell(row=j, column=1).value = 1
                    else:
                        break


def determine_words_used():

    print (len(words_used))

    if get_words_used:
        for i in range(len(words_used)):
            j = words[33].get(words_used[i],3)
            ws.cell(row=j, column=2).value = 1


def build_dict(ex_dict):

    global words

    relat = []
    atomic_relations = []
    compound = []
    synon = []
    redundant = []
    proper_names = []
    atomic_relata = []
    not_oft_def = []  # words that are only defined if they appear in the input sentence
    uniq_obj = []  # words which have (b=julius caesar) as definiendum
    syn_pairs = []
    definitions = []
    map_words_to_row = {}
    really_atomic = []
    pronouns = []
    poss_pronouns = []
    doubles = []
    triples = []
    essential_properties = {}
    pos = []
    almost_done = False

    i = -1
    if get_words_used:
        mm = 3000
    else:
        mm = len(ex_dict)

    while i < mm - 1:
        i += 1
        if get_words_used:
            if i == 0:
                i = 5
            s = ws.cell(row=i, column=1).value
            if s == 451:
                bb = 8
            str1 = ws.cell(row=i, column=3).value
            word = ws.cell(row=i, column=4).value
            if word == "true*":
                word = "true"
            if word == "false*":
                word = "false"
            next_word = ws.cell(row=i, column=4).value
            if word == "" and next_word == "":
                break
        else:
            s = 0
            str1 = ex_dict[i][0]
            word = ex_dict[i][1]
            if word != None:
                word = tran_str(word, 2)
                word = word[0]


        if word == None and almost_done:
            break
        if word == None:
            almost_done = True
        else:
            almost_done = False
        if str1 != None and str1 != "":
            if not isinstance(str1, int):
                str1 = str1.strip()
            if word == 'non_whole':
                bb = 7

            if isinstance(word, int):
                word = str(word)

            if "(" in word:
                cc = word.index("(")
                word = word[:cc - 1]

            word = word.strip()
            if word == "<":
                bb = 8

            if get_words_used:
                abbrev_relat = ws.cell(row=i, column=5).value
                defin = ws.cell(row=i, column=6).value
            else:
                abbrev_relat = ex_dict[i][2]
                defin = ex_dict[i][3]
                defin = tran_str(defin, 3)
                defin = defin[0]
            if word == 'i':
                bb = 8
            map_words_to_row.update({word: s})
            if abbrev_relat != "":
                map_words_to_row.update({abbrev_relat: s})


            if defin == 'redundant':
                redundant.append(word)
            if defin != None and defin.find("E.g.") == -1 and defin.find("EXP") == -1 \
                    and defin != 'redundant' and word != "." and defin.find("e.g.") == -1 \
                    and str1 != "":
                if word != None:
                    word = word.strip()
                if abbrev_relat != None:
                    abbrev_relat = abbrev_relat.strip()
                if str1 == None:
                    print("you did not state the part of speech for " + word)
                    g = 4/0
                sec_let = copy.copy(str1)
                fir_let = str1[0:1]

                if "(" in word:
                    y = word.find("(")
                    word = word[:y - 1]
                if " " in word:
                    m = word.count(" ")
                    if m == 1:
                        word1 = copy.copy(word)
                        y = word1.find(" ")
                        word1 = word1[:y]
                        doubles.append([word1, word])
                    if m == 2:
                        word1 = copy.copy(word)
                        y = word1.find(" ")
                        word1 = word1[:y]
                        triples.append([word1, word])

                sec_let = sec_let[1:2]
                thir_let = str1[2:3]
                if len(str1) > 4:
                    fif_let = str1[4:5]
                else:
                    fif_let = None

                if fir_let == 'r' and sec_let != 's':
                    pos.append([abbrev_relat, fir_let, fif_let])
                else:
                    pos.append([word, fir_let, fif_let])

                if thir_let == "d":
                    compound.append(abbrev_relat)
                elif thir_let == "n":
                    proper_names.append(word)


                if (sec_let == 'a' or sec_let == 'b') and fir_let == "r":
                    if abbrev_relat == None:
                        bb = 8
                    if abbrev_relat not in atomic_relations:
                        atomic_relations.append(abbrev_relat)
                if sec_let == 'b':
                    really_atomic.append(abbrev_relat)
                elif sec_let == 'u':
                    uniq_obj.append(word)
                elif sec_let == 'k':
                    not_oft_def.append(word)


                if fir_let == 'r':
                    relat.append([word, abbrev_relat])
                elif fir_let == 'p':
                    pronouns.append(word)
                elif fir_let == 'q':
                    poss_pronouns.append(word)

                if sec_let == 'a':
                    atomic_relata.append(word)
                elif sec_let == "m":
                    str6 = defin[10:]
                if sec_let == 'p' or sec_let == 'd':
                    if sec_let == 'p':
                        sec_let = 7
                    elif sec_let == 'd':
                        sec_let = 5
                elif sec_let == 's':
                    str6 = defin[defin.find("=") + 1:-1]
                    str6 = str6.strip()
                    str7 = defin[1:defin.find("=")]
                    str7 = str7.strip()
                    list3a = [str7, str6, defin]
                    syn_pairs.append(list3a)
                    synon.append(str7)

                if sec_let != 'a' and sec_let != 'm' and defin != "artificial" and defin != 'redundant' \
                        and defin != "postponed" and sec_let != 'b':
                    if fir_let == "r":
                        definitions.append([abbrev_relat, defin, fir_let, sec_let, thir_let, fif_let, i])
                    else:
                        definitions.append([word, defin, fir_let, sec_let, thir_let, fif_let, i])
                        if fir_let == "n":
                            essential_properties.update({word: ""})

    syn_pairs.sort()
    relat = sorted(relat, key=operator.itemgetter(0))
    atomic_relata = sorted(atomic_relata, key=operator.itemgetter(0))
    repeat_relat = repeat_relations()

    words = ["", "", "", "", "", "", relat, "", "", "",
             "", "", "", syn_pairs, synon, "", definitions, "", "",
             "", "", redundant, atomic_relations, atomic_relata,
             pronouns, poss_pronouns, "", "", pos, really_atomic,
             "", doubles, triples, map_words_to_row, compound, proper_names,
             not_oft_def, uniq_obj, essential_properties, repeat_relat]


def findinlist(str1, list1, i, j):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the matching second element

    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            return str2
    return None


def findinlist_alert_error(str1, list1, i, j):
    for d in range(len(list1)):
        if str1 == list1[d][i]:
            str2 = list1[d][j]
            if str2 != "" and str2 != None:
                return str2
    g = 4 / 0


def findposinmd(str1, list1, p):
    # this determines the position of an element in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return i
    return -1


def findposinmd_alert_error(str1, list1, p):
    # this determines the position of an element in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return i
    g = 4 / 0


def isinmdlist(str1, list1, p):
    # this determines whether or not an element is in a multidimensional list
    for i in range(len(list1)):
        if list1[i][p] == str1:
            return True

    return False


def findposinlist(str1, list1, i):
    # this function takes a string, matches it to an element in the first dimension
    # of the list, then returns the position in the list

    if str1 == 0:
        return
    str2 = copy.copy(str1)
    if str2 != None:
        str2 = str2.replace(" ", "")
    for d in range(len(list1)):
        str3 = copy.copy(list1[d][i])
        if str3 != None:
            str3 = str3.replace(" ", "")
        if str2 == str3:
            return d
    else:
        return -1


def two_elements_are_in_list(list1, stri, strj, i, j):
    for k in range(len(list1)):
        if list1[k][i] == stri and list1[k][j] == strj:
            return True

    return False


def findin1dlist(str1, list1):
    for i in range(len(list1)):
        if str1 == list1[i]:
            return i


def isatomic(list1):
    atomic_relations = words[29]
    num = [5, 14]
    if not list1[9] in atomic_relations:
        return False
    for i in num:
        if not isvariable(list1[i]):
            return False
    num = [3, 4, 6, 7, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, \
           33, 34, 35, 47, 48, 49, 50, 51, 52, 53, 69, 70]
    for i in num:
        if list1[i] != "" and list1[i] != None:
            return False
    return True


def use_identity(identities, negated_conjunction, consistent):

    if not consistent:
        return False

    if identities != []:
        num = [5, 14, 18, 22]
        remove_duplicates2d(identities, 0, 1)
        for i in range(len(identities)):
            str1 = "(" + identities[i][0] + " = " + identities[i][1] + ")"
            for j in range(len(total_sent) - 1, 0, -1):
                if str1 in total_sent[j][1]:
                    identities[i][2] = total_sent[j][0]
                    break

        for i in range(len(identities)):
            str1 = identities[i][0]
            str2 = identities[i][1]
            anc1 = identities[i][2]
            for sent in detach_sent:
                for k in num:
                    if sent[k] == str1 or sent[k] == str2:
                        ant_sent_parts = copy.deepcopy(sent)
                        con_parts = copy.deepcopy(sent)
                        if sent[k] == str1:
                            con_parts[k] = str2
                        else:
                            con_parts[k] = str1
                            con_parts = build_sent2(con_parts)
                        if con_parts[42] == "q" + l1:
                            bb = 8
                        prepare_att_sent_1_sent(ant_sent_parts, "SUB", anc1, "e", con_parts)
                        if con_parts[5] == con_parts[14] and \
                                        con_parts[9] != "=" and con_parts[2] == "":
                            consistent = use_reflexivity1(con_parts)
                            if not consistent:
                                return consistent

        consistent = detach1("do not use modus tollens", negated_conjunction)

    return consistent


def check_reflexivity(sent):
    consistent = True

    if sent[5] == sent[14] and sent[2] == "" and sent[9] != "=":
        consistent = use_reflexivity2(sent)

    return consistent


def use_reflexivity1(sent):
    global sn
    anc1 = sent[58]
    anc2 = attach_sent[-1][2]
    sn += 1
    add_to_total_sent(sn, sent[72], sent[1], "", "EE", anc1, anc2)
    consistent = use_reflexivity2(sent)

    return consistent


def use_reflexivity2(sent):
    global sn
    new_sent = copy.deepcopy(sent)
    new_sent[8] = "~"
    new_sent = build_sent2(new_sent)
    full_conditional = sent[0] + " " + conditional + " " + new_sent[0]
    abbrev_conditional = sent[42] + " " + conditional + " " + new_sent[42]
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "IRR")
    sn += 1
    consistent = add_to_total_sent_consist(sn, new_sent[72], new_sent[1], \
                                           new_sent[2], "MP", sn - 2, sn - 1, [])

    return consistent


def step_three(truth_value):
    # bbb
    global instan_time, instan_used

    negated_conjunction = []

    identities = get_identities()

    build_list_of_abbreviations()

    eliminate_attached_conjuncts()

    put_nc_id_ax_df_into_list()

    consistent = detach1("do not use modus tollens", negated_conjunction)

    consistent = add_stan_sent(consistent)

    consistent = use_identity(identities, negated_conjunction, consistent)

    consistent, object_prop2 = use_basic_lemmas(consistent)

    consistent = step_four(negated_conjunction, object_prop2, consistent)

    consistent = final_truth_value(consistent, truth_value)

    return consistent


def final_truth_value(consistent, truth_value):
    if truth_value == 'co':
        truth_value = False
    else:
        truth_value = True
    if consistent == truth_value:
        return True
    else:
        return False


def determine_truth_value(consistent, impl):
    if consistent and impl != nonseq:
        return False
    elif impl == nonseq and consistent:
        return False
    return True


def eliminate_attached_conjuncts():
    global sn
    for sent in attach_sent:
        if sent[46] == 'eliminate as conjunct':
            sn += 1
            sent[2] = sn
            add_to_total_sent(sn, sent[37], sent[4], "", "&E", sent[2])


def get_identities():
    identities = []
    i = 0
    while i < len(detach_sent):
        if detach_sent[i][2] == "ID":
            identities.append([detach_sent[i][0], detach_sent[i][1], ""])
            del detach_sent[i]
        else:
            i += 1

    return identities


def put_nc_id_ax_df_into_list():
    # this function takes out sentences proved by NC, AX, RN etc and
    # prepares to put them into new locations
    list1 = []
    list2 = []
    list5 = []
    rn_used = False
    bool1 = False

    for i in range(0, len(total_sent)):
        if i == 9:
            bb = 7
        str1 = total_sent[i][4][:2]
        str2 = total_sent[i][4][:2]
        if total_sent[i][4] == "" and not bool1:
            list6 = copy.deepcopy(total_sent[i])
            list5.append(list6)
        elif total_sent[i][4] == "ID":
            list6 = copy.deepcopy(total_sent[i])
            list5.append(list6)
            bool1 = True
        elif 'DF' == str1 or 'NC' == str1 or 'AX' == str1 \
                or 'RN' == str2:
            list4 = copy.deepcopy(total_sent[i])
            list1.append(list4)
            rn_used = True
        elif bool1:
            if str1 == 'DE':
                total_sent[i][4] = "DF " + total_sent[i][4][3:]
            elif str1 == 'AY':
                total_sent[i][4] = "AX ENT"
            list3 = copy.deepcopy(total_sent[i])
            list2.append(list3)

    if rn_used:
        dict1 = rearrange_total_sent(list5, list1, list2)
        renumber_attach_sent(dict1)


def rearrange_total_sent(list5, list1, list2):
    # this function puts the total_sent into a better order

    global total_sent
    b = time.time()
    total_sent = []
    for i in range(len(list5)):
        total_sent.append(list5[i])
    for i in range(len(list1)):
        total_sent.append(list1[i])
    add_to_total_sent("", "")
    for j in range(len(list2)):
        total_sent.append(list2[j])

    dict1 = {}
    j = 0
    for i in range(len(total_sent)):
        if (i > 3 and total_sent[i][0] != "") or i <= 3:
            j += 1
            dict1.update({total_sent[i][0]: j})
            total_sent[i][0] = j

    for i in range(len(total_sent)):
        for j in range(0, 5):
            if len(total_sent[i]) > 5 + j:
                if total_sent[i][j + 5] != "":
                    total_sent[i][j + 5] = dict1.get(total_sent[i][j + 5], None)

                break

    return dict1


def get_detached_variables():
    # This categorizes all abbreviations which appear in a detached sentence

    num = [5, 14, 18, 22]
    temp_list = []
    indefinite_concept = findinlist("indefinite", abbreviations, 1, 0)
    definite_concept = findinlist("definite", abbreviations, 1, 0)
    general_concept = findinlist("general", abbreviations, 1, 0)
    defn = []
    indef = []
    general = []

    for i in range(len(detach_sent)):
        for j in num:
            if isvariable(detach_sent[i][j]):
                str1 = detach_sent[i][j]
                if str1 not in defn and str1 not in indef and str1 not in general:
                    if isinmdlist(detach_sent[i][j], abbreviations, 0):
                        defn.append(detach_sent[i][j])
                    elif detach_sent[i][9] == 'J' and detach_sent[i][14] == indefinite_concept:
                        indef.append(detach_sent[i][j])
                    elif detach_sent[i][9] == 'J' and detach_sent[i][14] == definite_concept:
                        defn.append(detach_sent[i][j])
                    elif detach_sent[i][9] == 'J' and detach_sent[i][14] == general_concept:
                        print('a general variable should not be here')
                        general.append(detach_sent[i][j])
                    else:
                        if detach_sent[i][j] not in temp_list:
                            temp_list.append(detach_sent[i][j])
            elif detach_sent[i][j] == 'i' and 'i' not in defn:
                defn.append(detach_sent[i][j])

    indef = categorize_remaining_variables(indef, defn, temp_list)

    return [general, indef, defn]


def categorize_remaining_variables(indef, defn, temp_list):
    # this places the remaining variables in the indefinite list

    for i in range(len(temp_list)):
        if temp_list[i] not in defn:
            if temp_list[i] not in indef:
                indef.append(temp_list[i])
    return indef


def categorize_variables(detached_var):
    # This determines whether non-definite variables are indefinite attached,
    # indefinite detached, general, or mixed

    potentially_general = []
    indef = detached_var[1]
    defn = detached_var[2]
    num = [5, 14, 18, 22]
    num2 = [34, 35, 32, 31, 30, 29]
    same_sent = []

    for i in range(len(attach_sent)):
        list1 = []
        for m in num2:
            if attach_sent[i][m] == "":
                break
            else:
                for j in range(len(attach_sent[i][m])):
                    for n in num:
                        str1 = attach_sent[i][m][j][n]
                        if str1 not in list1 and str1 != None:
                            list1.append(str1)
                        sent_type = attach_sent[i][m][j][53][0]
                        sent_num = str(attach_sent[i][2])
                        sibling_num = attach_sent[i][m][j][68][:-1]
                        if str1 != "" and str1 != None and str1 != "i":
                            if str1 == "i":
                                if str1 not in defn:
                                    defn.append(str1)
                            elif isinmdlist(str1, abbreviations, 0):
                                if str1 not in defn:
                                    defn.append(str1)
                            elif str1 not in defn and str1 not in indef:
                                potentially_general.append([str1, sent_type, sent_num, sibling_num])
        same_sent.append(list1)

    potentially_general = sorted(potentially_general, key=operator.itemgetter(0, 2))
    general, t_indef = variable_type(potentially_general)
    indef = quick_append(t_indef, indef)

    return [general, defn, indef, same_sent]


def quick_append(list1, list2):
    for i in range(len(list1)):
        if list1[i] not in list2:
            list2.append(list1[i])
    return list2


def variable_type(potentially_general):
    # The algorithm for determining whether or not a variable is general is very complicated
    # it must appear on both sides of a consequential connective, i.e., v -> <->
    # in the same attached sentence
    #
    if potentially_general == []:
        return [[], []]
    last_variable = ""
    last_sent_num = 0
    must_not_have = []
    general = []
    indef = []
    non_conjunctive = False
    go_to_next_variable = False
    if len(potentially_general) == 1:
        return [[], potentially_general[0][0]]

    k = -1
    for i in potentially_general:
        k += 1
        sent_type = i[1]
        sibling_num = i[3]
        sent_num = i[2]
        variable = i[0]
        if variable == 'z' + l1:
            bb = 8
        if variable == last_variable and not go_to_next_variable:
            if last_sent_num == sent_num:
                if non_conjunctive:
                    general.append(variable)
                    go_to_next_variable = True
                else:
                    if sent_type == 'c':
                        if sibling_num not in must_not_have:
                            general.append(variable)
                            go_to_next_variable = True
                            must_not_have = []
                    else:
                        general.append(variable)
                        go_to_next_variable = True
                        must_not_have = []
            else:
                must_not_have = []
                must_not_have.append(sibling_num)
                last_sent_num = sent_num
                non_conjunctive = False
        elif variable == last_variable and go_to_next_variable:
            pass
        else:
            if last_variable != "" and not go_to_next_variable:
                indef.append(potentially_general[k - 1][0])
            if sent_type == 'c':
                must_not_have.append(sibling_num)
                non_conjunctive = False
            else:
                non_conjunctive = True
            last_sent_num = sent_num
            last_variable = variable
            go_to_next_variable = False
    else:
        if variable not in general and variable not in indef:
            indef.append(variable)

    return general, indef


def get_id_sent():
    # this function gets the list of identities
    for i in range(len(total_sent)):
        if total_sent[i][4] == "ID":
            return [total_sent[i][0], total_sent[i][1], "", "", "", "", ""]


def print_variables(list1):
    # this prints out the variables within the total_sent list, just above
    # where it prints the attached sentences

    ## variable type = [general, defn, indef, same_sent]

    general = list1[0]
    indef = list1[2]
    defn = list1[1]
    identities = get_id_sent()
    gen_str = ""
    ind_str = ""
    def_str = ""

    if general != []:
        gen_str = 'General Variables: '
        for i in range(len(general)):
            gen_str += general[i] + " "
    if indef != []:
        ind_str = 'Indefinite Variables: '
        for i in range(len(indef)):
            ind_str += indef[i] + " "
    if defn != []:
        def_str = 'Constants: '
        for i in range(len(defn)):
            def_str += defn[i] + " "

    j = 0
    for i in range(len(total_sent) - 1, 0, -1):
        if len(total_sent[i][1]) > 15:
            if total_sent[i][1].startswith('STANDARD ATTA'):
                total_sent.insert(i, identities)
                if gen_str != "":
                    j += 1
                    total_sent.insert(i + j, ["", gen_str, "", "", "", "", "", ""])
                if ind_str != "":
                    total_sent.insert(i + j, ["", ind_str, "", "", "", "", "", ""])
                    j += 1
                if def_str != "":
                    total_sent.insert(i + j, ["", def_str, "", "", "", "", "", ""])
                return


def renumber_attach_sent(new_numbers):
    # this gives attach_sent their proper number according to the new
    # numbering system as arrived at in the rearrange_total_sent function

    if new_numbers != {}:
        for i in range(len(attach_sent)):
            old_num = attach_sent[i][2]
            if old_num < 400:
                attach_sent[i][2] = new_numbers.get(old_num)
                if attach_sent[i][2] == None:
                    print('you renumbering of attach_sent failed')


# object_properties, abbrev, general, class, accidental properties, parts of properties
# if object is a thing then it is stated whether or not it is consequential


def step_four(negated_conjunction, object_prop2, consistent):
    # aaa
    global instan_used, instan_time

    if consistent and attach_sent != []:
        detached_var = get_detached_variables()

        variable_type = categorize_variables(detached_var)

        print_variables(variable_type)

        object_properties, detached_predicates = get_detached_predicates(variable_type)

        attached_predicates, object_properties = get_attached_predicates(variable_type, object_properties)

        object_properties = rearrange_object_properties(object_properties)

        object_properties = print_general_object_properties(object_properties)

        print_object_properties(object_properties)

        instantiations = determine_if_same_class(object_properties)

        use_axiom_of_definition2(instantiations)

        substitute_in_attach_sent(instantiations)

        print_instantiations(instantiations)

        consistent = detach1("use modus tollens", negated_conjunction)

    return consistent


# object, exclusive classes, variable_type, class:sent, inclusive class, accidental properties
# rrr

def use_basic_lemmas(consistent):

    global lemmas_used, time_spent_in_lemma_function
    if not consistent:
        return consistent, {}

    lemmas_used += 1
    aa = time.time()

    exclusive_classes = ['moment', 'relationship', 'point', 'number',
                         'imagination', 'concept' + un, "property" + un, 'property',
                         'possible world', 'letter', 'mind', 'matter', 'sensorium']

    num = [5, 14, 18, 22]
    object_prop2 = {}
    error = ""
    j = -1
    while error == "" and j < len(detach_sent) - 1:
        j += 1
        sent = detach_sent[j]
        for i in num:
            object = sent[i]
            if object != None and object != "":
                category = get_class(sent[9], sent, i)
                if object in object_prop2.keys():
                    properties = object_prop2.get(object)
                    eclasses = properties[0]
                    iclasses = properties[3]
                    if category != 'thing' and category not in eclasses and category not in iclasses:
                        if category in exclusive_classes:
                            # the reason why we do not put sentences of the form bIc in the eclass_sent
                            # dictionary is because we cannot use basic lemmas with those types
                            # of sentences
                            if sent[9] == "I" and i == 5:
                                eclass_sent = properties[2]
                            else:
                                eclasses.append(category)
                                eclass_sent = properties[2]
                                eclass_sent.update({category: sent})

                        else:
                            iclasses.append(category)
                            eclass_sent = properties[2]

                        properties = [eclasses, "", eclass_sent, iclasses]
                        object_prop2[object] = properties
                        if len(eclasses) > 1:
                            error = "contradiction in exclusivity"
                            break

                else:
                    if category != 'thing':
                        eclass_sent = {}
                        if sent[9] == "I" and i == 5:
                            properties = [[], "", eclass_sent, []]
                        elif category in exclusive_classes:
                            eclass_sent = {category: sent}
                            properties = [[category], "", eclass_sent, []]
                        else:
                            properties = [[], "", eclass_sent, [category]]
                        object_prop2.update({object: properties})

    if error == "contradiction in exclusivity":
        consistent = add_basic_lemmas(object, properties)

    bb = time.time()
    time_spent_in_lemma_function += (bb-aa)

    return consistent, object_prop2


def add_basic_lemmas(object, obj_properties):

    class_sent = obj_properties[2]
    it = iter(class_sent.values())
    first_sent, second_sent = next(it), next(it)

    concept_thing = findinlist('thing', abbreviations, 1, 0)
    if concept_thing == None:
        concept_thing = add_thing_to_abbreviations()
    obj_pos1 = 5 if first_sent[5] == object else 14
    obj_pos2 = 5 if second_sent[5] == object else 14
    sec_obj_pos = 5 if obj_pos2 == 14 else 14
    second_obj = second_sent[sec_obj_pos]

    build_original_lemma(obj_pos1, obj_pos2, second_sent, first_sent)

    int_thing = variables[0]
    del variables[0]

    build_rename_sent(object, second_obj, int_thing, concept_thing)

    build_renamed_lemma(obj_pos1, obj_pos2, second_sent, first_sent, object, concept_thing, int_thing)

    infer_sec_obj_thing(concept_thing, second_obj, second_sent)

    consistent = instantiate_int_thing(int_thing, concept_thing, second_obj, first_sent, second_sent)

    return consistent


def build_original_lemma(obj_pos1, obj_pos2, second_sent, first_sent):
    global sn
    thing_sent = "(d I e)"
    thing_concept = "(e = thing)"
    if obj_pos1 == 5 and obj_pos2 == 5:
        cond1 = "(" + "b" + " " + first_sent[9] + " " + "c" + ")"
        cond2 = "(" + "b" + " ~ " + second_sent[9] + " " + "d" + ")"
        name = "SS"
    elif obj_pos1 == 5 and obj_pos2 == 14:
        cond1 = "(" + "b" + " " + first_sent[9] + " " + "c" + ")"
        cond2 = "(" + "d" + " ~ " + second_sent[9] + " " + "b" + ")"
        name = "SO"
    elif obj_pos1 == 14 and obj_pos2 == 5:
        cond1 = "(" + "c" + " " + first_sent[9] + " " + "b" + ")"
        cond2 = "(" + "b" + " ~ " + second_sent[9] + " " + "d" + ")"
        name = "OS"
    elif obj_pos1 == 14 and obj_pos2 == 14:
        cond1 = "(" + "c" + " " + first_sent[9] + " " + "b" + ")"
        cond2 = "(" + "d" + " ~ " + second_sent[9] + " " + "b" + ")"
        name = "OO"

    original_conditional = "((" + cond1 + " & " + thing_sent + ") " + implies + " " \
                           + cond2 + ") " + thing_concept

    lemma_name = "LE." + first_sent[9] + "." + second_sent[9] + "." + name
    sn += 1
    add_to_total_sent(sn, original_conditional, "", "", lemma_name)


def build_rename_sent(key, second_obj, int_thing, concept_thing):
    global sn
    sent1 = "(b" + mini_c + key + ")"
    sent2 = "(c" + mini_c + second_obj + ")"
    sent3 = "(d" + mini_c + int_thing + ")"
    thing_sent = "(e" + idd + concept_thing + ")"
    list1 = [sent1, sent2, sent3, thing_sent]
    rename_sent = " & ".join(list1)
    sn += 1
    add_to_total_sent(sn, rename_sent, "", "", "RN")


def build_renamed_lemma(obj_pos1, obj_pos2, second_sent, first_sent,
                        key, concept_thing, int_thing):
    global sn
    if obj_pos1 == 5 and obj_pos2 == 5:
        cond1 = "(" + key + " " + first_sent[9] + " " + first_sent[14] + ")"
        cond2 = "(" + key + " ~ " + second_sent[9] + " " + second_sent[14] + ")"
    elif obj_pos1 == 5 and obj_pos2 == 14:
        cond1 = "(" + key + " " + first_sent[9] + " " + first_sent[14] + ")"
        cond2 = "(" + second_sent[5] + " ~ " + second_sent[9] + " " + key + ")"
    elif obj_pos1 == 14 and obj_pos2 == 5:
        cond1 = "(" + first_sent[5] + " " + first_sent[9] + " " + key + ")"
        cond2 = "(" + key + " ~ " + second_sent[9] + " " + second_sent[14] + ")"
    elif obj_pos1 == 14 and obj_pos2 == 14:
        cond1 = "(" + first_sent[5] + " " + first_sent[9] + " " + key + ")"
        cond2 = "(" + second_sent[5] + " ~ " + second_sent[9] + " " + key + ")"
    thing_sent = "(" + int_thing + " I " + concept_thing + ")"

    cond1p = first_sent[42]
    cond2p = name_sent(cond2)
    thing_sentp = name_sent(thing_sent)
    full_conditional = "(" + cond1 + " & " + thing_sent + ") " + implies + " " + cond2
    abbrev_conditional = "(" + cond1p + " & " + thing_sentp + ") " + implies + " " + cond2p
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "SUB", sn - 2, sn - 1)


def infer_sec_obj_thing(concept_thing, second_obj, second_sent):
    global sn
    thing_sent = "(" + second_obj + " I " + concept_thing + ")"
    thing_sentp = name_sent(thing_sent)
    full_conditional = second_sent[0] + " " + implies + " " + thing_sent
    abbrev_conditional = second_sent[42] + " " + implies + " " + thing_sentp
    sn += 1
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "LE ENT")
    sn += 1
    add_to_total_sent(sn, thing_sent, thing_sentp, "", "MP", sn - 1, second_sent[58])
    if "~" in second_sent[42]: g = 4 / 0


def instantiate_int_thing(int_thing, concept_thing, second_obj, first_sent, second_sent):
    global sn
    rename_sent = "(" + int_thing + mini_c + second_obj + ")"
    sn += 1
    add_to_total_sent(sn, rename_sent, "", "", "RN", sn - 1, sn - 3)
    sec_antec = "(" + second_obj + " I " + concept_thing + ")"
    abbrev_sec_antec = name_sent(sec_antec)
    second_sent[8] = "~"
    second_sent = build_sent2(second_sent)
    full_conditional = "(" + first_sent[0] + " & " + sec_antec + ") " + \
                       implies + " " + second_sent[0]
    abbrev_conditional = "(" + first_sent[42] + " & " + abbrev_sec_antec + ") " + \
                         implies + " " + second_sent[42]
    add_to_total_sent(sn, full_conditional, abbrev_conditional, "", "SUB", sn - 4, )
    conjunction = "(" + first_sent[0] + " & " + sec_antec + ")"
    conjunctionp = "(" + first_sent[42] + " & " + abbrev_sec_antec + ")"
    if "~" in first_sent[42]: g = 4 / 0
    sn += 1
    add_to_total_sent(sn, conjunction, conjunctionp, "", "&I", first_sent[58], sn - 3)
    sn += 1
    consistent = add_to_total_sent_consist(sn, second_sent[72],
                                           second_sent[1], second_sent[2], "MP", sn - 1, sn - 2, [])
    if consistent: g = 4 / 0

    return consistent


def add_thing_to_abbreviations():
    concept_thing = variables[0]
    del variables[0]
    abbreviations.append([concept_thing, 'thing'])
    d = findposinmd("ID", total_sent, 4)
    total_sent[d][1] += "(" + concept_thing + " = thing)"

    return concept_thing


def get_detached_predicates2(variable_type):
    # this makes a list of the detached definite predicates
    object_properties = []
    detached_predicates = []
    for i in range(len(detach_sent)):
        sent = detach_sent[i]
        subj = detach_sent[i][5]
        relat = detach_sent[i][9]
        obj = detach_sent[i][14]
        obj = "" if obj == None else obj
        relat2 = detach_sent[i][15]
        obj2 = detach_sent[i][18]
        obj2 = "" if obj2 == None else obj2
        relat2 = "" if relat2 == None else relat2
        t_value = detach_sent[i][8]
        t_value = "" if t_value == None else t_value
        subj_pred = alpha + t_value + relat + obj + relat2 + obj2
        sub_sent_parts = [alpha, t_value, relat, obj, relat2, obj2]
        if obj != "":
            obj_pred = subj + t_value + relat + alpha + relat2 + obj2
            obj_sent_parts = [subj, t_value, relat, alpha, relat2, obj2]
        if obj2 != "":
            obj2_pred = subj + t_value + relat + obj + relat2 + alpha
            obj2_sent_parts = [subj, t_value, relat, obj, relat2, alpha]

        absolute_predicate = subj + relat + obj + relat2 + obj2
        s_variable_kind = get_quick_variable_type(subj, variable_type)
        o_variable_kind = get_quick_variable_type(obj, variable_type)
        o2_variable_kind = get_quick_variable_type(obj2, variable_type)
        detached_predicates.append([absolute_predicate, t_value, detach_sent[i][42]])
        if relat == "I":
            kind = findinlist(obj, abbreviations, 0, 1)
            skind = kind_exception(kind)
        else:
            skind = get_class(relat, sent, 5)

        object_properties = get_object_properties2(subj,
                                                   object_properties,
                                                   s_variable_kind,
                                                   subj_pred,
                                                   skind,
                                                   "c",
                                                   "",
                                                   "",
                                                   sub_sent_parts)
        if isvariable(obj) or obj == "i":
            okind = get_class(relat, sent, 14)
            object_properties = get_object_properties2(obj,
                                                       object_properties,
                                                       o_variable_kind,
                                                       obj_pred,
                                                       okind,
                                                       "c",
                                                       "",
                                                       "",
                                                       obj_sent_parts)
        if isvariable(obj2) or obj2 == "i":
            object_properties = get_object_properties2(obj2,
                                                       object_properties,
                                                       o2_variable_kind,
                                                       obj2_pred,
                                                       'thing',
                                                       "c",
                                                       "",
                                                       "",
                                                       obj2_sent_parts)

    return object_properties, detached_predicates


def get_detached_predicates(variable_type):
    # this makes a list of the detached definite predicates
    object_properties = []
    detached_predicates = []
    for i in range(len(detach_sent)):
        sent = detach_sent[i]
        subj = detach_sent[i][5]
        relat = detach_sent[i][9]
        obj = detach_sent[i][14]
        obj = "" if obj == None else obj
        relat2 = detach_sent[i][15]
        obj2 = detach_sent[i][18]
        obj2 = "" if obj2 == None else obj2
        relat2 = "" if relat2 == None else relat2
        t_value = detach_sent[i][8]
        t_value = "" if t_value == None else t_value
        subj_pred = alpha + t_value + relat + obj + relat2 + obj2
        sub_sent_parts = [alpha, t_value, relat, obj, relat2, obj2]
        if obj != "":
            obj_pred = subj + t_value + relat + alpha + relat2 + obj2
            obj_sent_parts = [subj, t_value, relat, alpha, relat2, obj2]
        if obj2 != "":
            obj2_pred = subj + t_value + relat + obj + relat2 + alpha
            obj2_sent_parts = [subj, t_value, relat, obj, relat2, alpha]

        absolute_predicate = subj + relat + obj + relat2 + obj2
        s_variable_kind = get_quick_variable_type(subj, variable_type)
        o_variable_kind = get_quick_variable_type(obj, variable_type)
        o2_variable_kind = get_quick_variable_type(obj2, variable_type)
        detached_predicates.append([absolute_predicate, t_value, detach_sent[i][42]])
        if relat == "I":
            kind = findinlist(obj, abbreviations, 0, 1)
            skind = kind_exception(kind)
        else:
            skind = get_class(relat, sent, 5)

        object_properties = get_object_properties2(subj,
                                                   object_properties,
                                                   s_variable_kind,
                                                   subj_pred,
                                                   skind,
                                                   "c",
                                                   "",
                                                   "",
                                                   sub_sent_parts)
        if isvariable(obj) or obj == "i":
            okind = get_class(relat, sent, 14)
            object_properties = get_object_properties2(obj,
                                                       object_properties,
                                                       o_variable_kind,
                                                       obj_pred,
                                                       okind,
                                                       "c",
                                                       "",
                                                       "",
                                                       obj_sent_parts)
        if isvariable(obj2) or obj2 == "i":
            object_properties = get_object_properties2(obj2,
                                                       object_properties,
                                                       o2_variable_kind,
                                                       obj2_pred,
                                                       'thing',
                                                       "c",
                                                       "",
                                                       "",
                                                       obj2_sent_parts)

    return object_properties, detached_predicates


def use_axiom_of_definition2(instantiations):
    global sn
    for var_list in instantiations:
        if var_list[2] == "D" or var_list[2] == "T":
            list1 = [None] * 80
            list1[5] = var_list[1]
            list1[9] = "I"
            list1[14] = var_list[4]
            list1 = build_sent2(list1)
            sn += 1
            list1[58] = sn
            if var_list[2] == "D":
                add_to_total_sent(sn, list1[0], list1[1], "", "AX DF")
            else:
                add_to_total_sent(sn, list1[0], list1[1], "", "AX ENT")

            detach_sent.append(list1)


def substitute_in_attach_sent(instantiations):
    # this substitutes the attached variable with the detached variables

    num2 = [34, 35, 32, 31, 30, 29]
    var_slots = [5, 14, 18]
    total_num = []
    if instantiations == []:
        return
    o = -1
    while o < len(instantiations) - 1:
        o += 1
        sub_properties = instantiations[o]
        num = sub_properties[3]
        if isinstance(num, int):
            num = [num]
        att_var = sub_properties[0]
        det_var = sub_properties[1]
        simul_sub = False
        next_att_var = 'bb'
        if o < len(instantiations) - 1:
            next_att_var = instantiations[o + 1][0]
            next_det_var = instantiations[o + 1][1]
            next_num = instantiations[o + 1][3]
            if isinstance(next_num, int):
                next_num = [next_num]
            if next_att_var != att_var and num == next_num:
                simul_sub = True
                print ('simul sub')
        for i in num:
            m = -1
            while attach_sent[m+1][26] != 'new conditional from instantiation':
                m += 1
                if attach_sent[m][2] == i:
                    attach_sent[m][26] = "not new"
                    cond_sent = copy.deepcopy(attach_sent[m])
                    if i not in total_num:
                        total_num.append(i)
                    for j in num2:
                        if cond_sent[j] == []:
                            break
                        else:
                            for f in range(len(cond_sent[j])):
                                atom_cond_sent = cond_sent[j][f]
                                new_sent = copy.deepcopy(atom_cond_sent)
                                new_sent[74] = False
                                for k in var_slots:
                                    already_done = False
                                    if atom_cond_sent[k] == att_var:
                                        new_sent[k] = det_var
                                        new_sent[74] = True
                                        already_done = True
                                        if not simul_sub:
                                            break
                                    if simul_sub and not already_done:
                                        if atom_cond_sent[k] == next_att_var:
                                            new_sent[k] = next_det_var
                                            new_sent[74] = True
                                cond_sent[j][f] = new_sent
                    make_new_attach_sent(cond_sent)
        if simul_sub:
            o+= 1


def findposinmdlistint(i, list1, p):
    for j in range(len(list1)):
        if list1[j][p] == i:
            return j
    else:
        return -1


def make_new_attach_sent(cond_sent):
    # this builds new strings within the conditional list

    cond_sent[26] = "new conditional from instantiation"  # this means its new and we need to print it
    prop_var_greek = cond_sent[47]
    prop_var_greek2 = prop_var_greek
    for j in [34, 35, 32, 31, 30, 29]:
        for k in range(len(cond_sent[j])):
            atom_cond_sent = cond_sent[j][k]
            if atom_cond_sent[74]:
                abs_oldp = atom_cond_sent[1]
                atom_cond_sent = build_sent2(atom_cond_sent)
                if j == 34:
                    n = 0
                    q = 42
                elif j == 35:
                    n = 1
                    q = 43
                else:
                    print("you haven't coded for this yet")
                    g = 4 / 0

                for m in range(len(cond_sent[n])):
                    if cond_sent[n][m][0] == abs_oldp:
                        cond_sent[n][m][0] = atom_cond_sent[1]
                for m in range(len(cond_sent[38])):
                    if cond_sent[38][m] == abs_oldp:
                        cond_sent[38][m] = atom_cond_sent[1]
                for m in range(len(cond_sent[q])):
                    if cond_sent[n][m][0] == atom_cond_sent[1]:
                        cond_sent[q][m][0] = atom_cond_sent[72]

                cond_sent = build_conjunction(cond_sent, q)

            prop_var_greek = prop_var_greek.replace(atom_cond_sent[44], atom_cond_sent[0])
            prop_var_greek2 = prop_var_greek2.replace(atom_cond_sent[44], atom_cond_sent[42])
    cond_sent[4] = prop_var_greek2
    cond_sent[37] = prop_var_greek
    attach_sent.append(cond_sent)




def make_new_attach_sent2(changed_attach_sent):
    # this builds new strings within the conditional list

    global sn
    attach_sent3 = []
    num = [34, 35, 32, 31, 30, 29]
    for i in changed_attach_sent:
        u = findposinmdlistint(i, attach_sent, 2)
        attach_sent2 = copy.deepcopy(attach_sent[u])
        attach_sent2[26] = "new conditional from instantiation"  # this means its new and we need to print it
        sn += 1
        attach_sent2[2] = sn  # finished here, need greek sentence
        prop_var_greek = attach_sent2[47]
        prop_var_greek2 = prop_var_greek
        for j in num:
            for k in range(len(attach_sent2[j])):
                atom_cond_sent = attach_sent2[j][k]
                oldp = atom_cond_sent[41]
                old_nat = atom_cond_sent[72]
                temp_oldp = "(" + oldp + ")"
                oldp_greek = atom_cond_sent[44]
                if atom_cond_sent[74]:
                    list1 = build_short_sent(atom_cond_sent)
                    atom_cond_sent[0] = list1[0]
                    atom_cond_sent[42] = list1[1]
                    prop_var_greek = prop_var_greek.replace(oldp_greek, list1[1])
                    prop_var_greek2 = prop_var_greek2.replace(oldp_greek, list1[0])
                    abs_oldp = oldp
                    abs_newp = list1[1]
                    if "~" in oldp:
                        abs_oldp = oldp.replace("~", "")
                        abs_newp = list1[1].replace("~", "")
                    if j == 34:
                        n = 0
                    elif j == 35:
                        n = 1
                    else:
                        print("you haven't coded for this yet")
                        sys.exit()
                    if isinstance(attach_sent2[n][0], list):
                        for m in range(len(attach_sent2[n])):
                            if attach_sent2[n][m][0] == abs_oldp:
                                attach_sent2[n][m][0] = abs_newp
                    else:
                        if attach_sent2[n][0] == abs_oldp:
                            attach_sent2[n][0] = abs_newp
                    for m in range(len(attach_sent2[38])):
                        if attach_sent2[38][m] == oldp:
                            attach_sent2[38][m] = list1[1]
                else:
                    prop_var_greek = prop_var_greek.replace(oldp_greek, oldp)
                    prop_var_greek2 = prop_var_greek2.replace(oldp_greek, old_nat)
        attach_sent2[4] = prop_var_greek
        attach_sent2[37] = prop_var_greek2
        list1 = get_ant_and_cond(prop_var_greek)
        attach_sent2[6] = list1[0]
        attach_sent2[7] = list1[1]
        attach_sent3.append(attach_sent2)

    return attach_sent3


def get_ant_and_cond(str1):
    list1 = mainconn(str1)
    ant = str1[:list1[1]]
    con = str1[list1[1] + 1:]
    ant = ant.strip()
    con = con.strip()
    ant = "" if os(ant) else ant
    con = "" if os(con) else con

    return [ant, con]


def get_greek(list1, str1):
    # this gets the greek letter substitute for the propositional variable

    for i in list1:
        if i != None:
            if i[0] == str1:
                return i[1]
    print("something is wrong in the get greek function")
    sys.exit()


def build_short_sent(list1):
    # this makes a new natural language sentence, if it is standard

    list1[14] = "" if list1[14] == None else list1[14]
    list1[15] = "" if list1[15] == None else list1[15]
    list1[18] = "" if list1[18] == None else list1[18]
    list2 = ["(", list1[5], list1[2], list1[9], list1[14], list1[15], list1[18], ")"]
    str1 = "".join(list2)
    str1p = name_sent(str1)

    return [str1, str1p]


def build_conjunction(list1, q):
    if q == 42:
        n = 7
        o = 0
    else:
        n = 8
        o = 1

    if len(list1[q]) > 1:
        str1 = ""
        str1p = ""
        for i in range(len(list1[q])):
            if str1 == "":
                str1 = list1[q][i][1] + list1[q][i][0]
                str1p = list1[o][i][1] + list1[o][i][0]
            else:
                str1 += " & " + list1[q][i][1] + list1[q][i][0]
                str1p += " & " + list1[o][i][1] + list1[o][i][0]

        list1[n] = [str1p, ""]
        list1[q - 2] = [str1, ""]
    else:
        list1[n] = list1[o][0]
        list1[q - 2] = list1[q][0]

    return list1


def print_general_object_properties(object_properties):
    # this prints up the predicates for the general variables since
    # these are markedly more difficult to print

    for i in range(len(object_properties)):

        str3 = ""
        for j in range(len(object_properties[i][3])):
            str2 = ""
            for k in range(len(object_properties[i][3][j][0])):
                str2 += object_properties[i][3][j][0][k] + " "
            if object_properties[i][1] == "agen":
                str3 += "[" + str2 + str(object_properties[i][3][j][1]) + "] "
            else:
                str3 += str2
        if object_properties[i][2][0] == 'thing' and object_properties[i][1] == 'agen':
            if object_properties[i][6] != []:
                for m in range(len(object_properties[i][6])):
                    str3 += " " + "{" + object_properties[i][6][m][0] + " " \
                            + str(object_properties[i][6][m][3]) + "}"
        object_properties[i][4] = str3

    return object_properties


def print_instantiations(instantiations):
    # this adds the instantiations to the total_sent list
    global sn

    for instantiation in instantiations:
        sn += 1
        str1 = "(" + instantiation[0] + mini_c + instantiation[1] + ")"
        if instantiation[4] == "P":
            str2 = ""
            for number in instantiation[3]:
                str2 += " " + str(number)
            str1 += " in " + str2
        add_to_total_sent(sn, str1, "", "", "IN")

    for cond in attach_sent:
        if cond[26] == "new conditional from instantiation":
            sn += 1
            cond[2] = sn
            add_to_total_sent(cond[2], cond[37], cond[4], "", "SUB")


def print_object_properties(object_properties):
    add_to_total_sent("", "")
    add_to_total_sent("", "OBJECT PREDICATES")

    for i in object_properties:
        str1 = i[0] + "  "
        for j in range(len(i[2])):
            str1 += "  " + i[2][j]
        str1 += " |"
        str1 += "  " + i[4]
        add_to_total_sent("", str1)


def determine_if_same_class(object_properties):
    # this determines whether a particular object belongs to the same class
    # as a general object
    instantiations = []
    i = -1
    while object_properties[i + 1][1] == 'agen':
        i += 1
        gen_var = object_properties[i][0]
        general_groups = object_properties[i][2]
        general_properties = object_properties[i][3]
        general_numbers = object_properties[i][5]
        if general_groups[0] == 'thing':
            done, instantiations = instantiate_things(object_properties,
                                                      instantiations, i)
            if done:
                break
        else:
            match_found = False
            for object_property in object_properties:
                if object_property[1] != 'agen':
                    for group in object_property[2]:
                        if group in general_groups:
                            partic_var = object_property[0]
                            predicates = object_property[3]
                            b = len(instantiations)
                            instantiations = instantiate2(predicates,
                                                          general_properties,
                                                          instantiations,
                                                          general_numbers,
                                                          gen_var,
                                                          partic_var,
                                                          object_properties)
                            if len(instantiations) > b:
                                match_found = True
                            break
            else:
                if not match_found:
                    done, instantiations = infer_member(general_properties,
                                                        general_groups, instantiations, gen_var, general_numbers)
                    if done:
                        break

    return instantiations


def infer_member(general_properties, general_groups, instantiations, gen_var, general_numbers):
    # if a general variable's only antecedent property is that it is a member
    # of a class then we may infer that it is detached
    # but only if its consequent contradicts a detached sentence

    for property in general_properties:
        if property[0][0] != "$":
            return False, instantiations
    else:
        new_var = variables[0]
        del variables[0]
        group_var = findinlist(general_groups[0], abbreviations, 1, 0)
        instantiations.append([gen_var, new_var, "D", general_numbers, group_var])
        return True, instantiations


def instantiate_things(object_properties, instantiations, i):
    # this determines if an indefinite thing can be instantiated
    # we only instantiate it if the particular variable contradicts
    # the general variable's consequent property

    list1 = []
    gen_var = object_properties[i][0]
    if object_properties[i][6] == []:
        return False, instantiations
    consequent = object_properties[i][6]

    for con_list in consequent:
        gen_tvalue = ""
        con_sent = con_list[0]
        if "~" in con_sent:
            con_sent = con_sent.replace("~", "")
            gen_tvalue = "~"
        if "'" in con_sent:
            con_sent = remove_prime(con_list)
        for object in object_properties:
            if object[0] == 'v':
                bb = 8
            if object[1] != "agen":
                properties = object[3]
                for property_list in properties:
                    for property in property_list[0]:
                        partic_tvalue = ""
                        if "~" in property:
                            property = property.replace("~", "")
                            partic_tvalue = "~"
                        if con_sent == property and partic_tvalue != gen_tvalue:
                            thing_var = findinlist("thing", abbreviations, 1, 0)
                            instantiations.append([gen_var, object[0], "T", con_list[3], thing_var])
                            return True, instantiations

    return False, instantiations


def remove_sp(str1):
    return str1.replace(" ", "")


def remove_prime(con_list):
    # this remove the prime sign from an indefinite variable

    con_list[4][1] = ""
    i = -1
    for word in con_list[4]:
        i += 1
        if "'" in word:
            word = word[:-1]
            con_list[4][i] = word

    return "".join(con_list[4])


def total_detach_prop(list1):
    list2 = []
    for set in list1:
        for prop in set[0]:
            list2.append(prop)
    return list2


def instantiate2(predicates, general_properties, instantiations, general_numbers, gen_var, partic_var,
                 object_properties):
    # this determines if a particular object has all the predicates of the
    # general object
    # bbb

    sentences = []
    properties = copy.deepcopy(general_properties)
    numbers = copy.deepcopy(general_numbers)
    for j in range(len(properties)):
        if properties[j][1] not in sentences:
            temp_list = copy.deepcopy(properties[j][0])
            for k in range(len(properties[j][0])):
                gen_prop = properties[j][0][k]
                if gen_prop == "$":
                    temp_list.remove("$")
                    if temp_list == []:
                        sentences.append(properties[j][1])
                        numbers.remove(properties[j][1])
                        break
                elif properties[j][1] not in sentences:
                    indef_instant_used = False
                    if "'" in gen_prop:
                        old_gen_prop = gen_prop
                        list1 = change_indef_attach_var(properties[j],
                                                        object_properties, k, instantiations)
                        if not list1:
                            pass
                        else:
                            gen_prop = list1[0]
                            if list1[1] != 'already found':
                                numbers2 = copy.deepcopy(numbers)
                                instantiations.append([list1[1], list1[2], "I", numbers2, ""])
                            indef_instant_used = True

                    for predicate_set in predicates:
                        for predicate in predicate_set[0]:
                            if gen_prop == predicate:
                                if indef_instant_used:
                                    gen_prop = old_gen_prop
                                temp_list.remove(gen_prop)
                                if temp_list == []:
                                    sentences.append(properties[j][1])
                                    numbers.remove(properties[j][1])
                                break

    if sentences != []:
        if numbers == []:
            instantiations.append([gen_var, partic_var, "", general_numbers, ""])
        else:
            instantiations.append([gen_var, partic_var, "", sentences, "P"])

    return instantiations


def change_indef_attach_var(sent_parts, object_properties, k, instantiations):
    # This determines whether an indefinite attached variable
    # is identical to a detached abbreviation

    for att_var in sent_parts[2][k]:
        if "'" in att_var:
            att_var = att_var[:-1]
            break
    if instantiations != []:
        for var in instantiations:
            if att_var == var[0]:
                det_var = var[1]
                new_sent = indefinite_instantiation(det_var, att_var,
                                                    sent_parts[2][k])
                return [new_sent, "already found", ""]

    d = findposinmd(att_var, object_properties, 0)
    if d == -1:
        print('there is something wrong with your ')
        'object properties list'
        sys.exit()
    set_attach = set(object_properties[d][2])
    attach_sent_parts = get_sent_parts(object_properties[d][3])
    for detach_prop in object_properties:
        set_detach = set(detach_prop[2])
        list1 = set_attach.intersection(set_detach)
        if list1 == set_attach and detach_prop[0] != att_var and detach_prop[1] != 'agen':
            det_var = detach_prop[0]
            for attach_sent_part in attach_sent_parts:
                for detach_sent_part in detach_prop[3][0][2]:
                    potentially_identical = False
                    for i in range(6):
                        try:
                            if attach_sent_part[i] == detach_sent_part[i] or \
                                            attach_sent_part[i] == alpha or \
                                            detach_sent_part[i] == alpha or \
                                            attach_sent_part[i] == att_var + "'":
                                pass
                            else:
                                break
                        except:
                            bb = 8
                    else:
                        potentially_identical = True
                        break
                else:
                    return False
            if not potentially_identical:
                return False
            else:
                new_sent = indefinite_instantiation(det_var, att_var, sent_parts[2][k])
                return [new_sent, att_var, det_var]

    return False


def get_sent_parts(list1):
    # this flattens the list of sent parts in the object properties list
    properties = []
    sent_parts = []
    for i in range(len(list1)):
        for j in range(len(list1[i][0])):
            if list1[i][0][j] not in properties:
                properties.append(list1[i][0][j])
                sent_parts.append(list1[i][2][j])

    return sent_parts


def indefinite_instantiation(det_var, att_var, sent):
    # this replaces the attached variable with the detached variable
    # in the sentence
    i = -1
    for word in sent:
        i += 1
        if word == att_var + "'":
            sent[i] = det_var
            break
    str1 = "".join(sent)
    return str1


def rearrange_object_properties(object_properties):
    # rearrange the object properties of the general objects so as to make it
    # easier to instantiate with
    # it takes the list of the form [Rd, cf, 33],[Re, cd, 33],[Rg,a,34]
    # and puts into the list [[Rd,Re],33],[[Rg],34]
    i = -1
    while i < len(object_properties) - 1:
        i += 1
        if i > 10:
            bb = 8
        list2 = []
        properties = object_properties[i][3]
        j = 0
        sentence_numbers = []
        while j < len(properties) - 1:
            sent_type = properties[j][1][-1]
            sent_num = properties[j][3]
            if sent_num not in sentence_numbers:
                sentence_numbers.append(sent_num)
            if j + 1 < len(properties):
                next_sent_num = properties[j + 1][3]
                next_sent_type = properties[j + 1][1][-1]
                predicate = properties[j][0]
                sent_parts = properties[j][4]
                list1 = []
                list3 = []
                while sent_type == next_sent_type and sent_num == next_sent_num:
                    list1.append(predicate)
                    list3.append(sent_parts)
                    if j + 1 < len(properties):
                        j += 1
                        sent_type = properties[j][1][-1]
                        sent_num = properties[j][3]
                        predicate = properties[j][0]
                        sent_parts = properties[j][4]
                        if j + 1 < len(properties):
                            next_sent_type = properties[j + 1][1][-1]
                            next_sent_num = properties[j + 1][3]
                        else:
                            break
                list1.append(predicate)
                list3.append(sent_parts)
                list2.append([list1, sent_num, list3])
                j += 1
        if j < len(properties):
            list2.append([[properties[j][0]], properties[j][3], properties[j][4]])
            sent_num = properties[j][3]
            if sent_num not in sentence_numbers:
                sentence_numbers.append(sent_num)

        object_properties[i][3] = list2
        object_properties[i][5] = sentence_numbers

    return object_properties


def get_object_properties2(str1,
                           object_properties,
                           variable_kind,
                           property,
                           kind,
                           sent_kind,
                           sent_num,
                           cond_num,
                           sent_parts):
    # this builds a list of object properties, if the variable is general
    uninformative_properties = ["I", "J", "H"]  # these are properties all object_properties have
    if property not in uninformative_properties:

        d = findposinmd(str1, object_properties, 0)
        if kind != 'thing2' and sent_parts[2] == 'I' and variable_kind == 'agen':
            property = "$"
        if kind == 'thing2':
            kind = 'thing'  # see kind exception for explaination

        if d == -1:

            if sent_kind[-1] == 'q':
                object_properties.append([str1, variable_kind, [kind],
                                          [], "", "", [[property, sent_kind, sent_num, cond_num, sent_parts]]])
            else:
                object_properties.append([str1, variable_kind, [kind], [[property,
                                                                         sent_kind, sent_num, cond_num, sent_parts]],
                                          "",
                                          "", []])
        else:
            for i in range(len(object_properties)):
                if object_properties[i][0] == str1:
                    list_kind = object_properties[i][2]
                    conseq_properties = object_properties[i][6]
                    list_properties = object_properties[i][3]
                    if kind not in list_kind and kind != "":
                        list_kind.append(kind)
                    if sent_kind[-1] == 'q':
                        list1 = [property, sent_kind, sent_num, cond_num, sent_parts]
                        conseq_properties.append(list1)
                        object_properties[i] = [str1, variable_kind, list_kind,
                                                list_properties, "", "", conseq_properties]
                    else:
                        list1 = [property, sent_kind, sent_num, cond_num, sent_parts]
                        list_properties.append(list1)
                        object_properties[i] = [str1, variable_kind, list_kind,
                                                list_properties, "", "", conseq_properties]
                    break

    return object_properties


def get_class(relat, sent, p):
    # this determines what class or category an object belongs to

    if relat == "A" or (relat == 'T' and p == 14):
        kind = 'moment'
    elif relat == "IR" and p == 5:
        kind = 'relationship'
    elif relat == 'AB' or relat == "L" or relat == 'AB' or (relat == 'S' and p == 14):
        kind = 'point'
    elif relat == "G" or (relat == 'N' and p == 14):
        kind = 'number'
    elif relat == "M" and p == 5 or (relat == 'TK' and p == 14):
        kind = 'relationship'
    elif relat == "M" and p == 14:
        kind = 'imagination'
    elif relat == "I" and p == 5:
        group = sent[14]
        kind = findinlist(group, abbreviations, 0, 1)
    elif relat == "I" and p == 14:
        kind = "concept" + un
    elif relat == "H" and p == 14:
        kind = "property" + un
    elif relat == "J" and p == 14:
        kind = "property"
    elif relat == "W" and p == 5:
        kind = "whole"
    elif relat == "W" and p == 14:
        kind = 'thing'
    elif relat == 'P' and p == 14:
        kind = 'possible world'
    elif relat == "D" and p == 14:
        kind = 'relationship'
    elif relat == 'AL':
        kind = 'letter'
    elif (relat == 'TK' or relat == "D") and p == 5:
        kind = 'mind'
    elif relat == "S" and p == 5:
        kind = 'matter'
    elif relat == "O" and p == 14:
        kind = 'sensorium'
    else:
        kind = "thing"

    return kind


def get_attached_predicates(variable_type, object_properties):
    # this makes a list of the attached definite predicates
    # and also adds to the list of object properties

    num = [34, 35, 32, 31, 30, 29]
    attached_predicates = []

    for i in range(len(attach_sent)):
        sent = attach_sent[i]
        cond_num = attach_sent[i][2]  # conditional number
        for j in num:
            if attach_sent[i][j] != []:
                for k in range(len(attach_sent[i][j])):
                    subj = attach_sent[i][j][k][5]
                    obj = attach_sent[i][j][k][14]
                    obj = "" if obj == None else obj
                    relat = attach_sent[i][j][k][9]
                    relat2 = attach_sent[i][j][k][15]
                    relat2 = "" if relat2 == None else relat2
                    obj2 = attach_sent[i][j][k][18]
                    obj2 = "" if obj2 == None else obj2
                    t_value = attach_sent[i][j][k][8]
                    t_value = "" if t_value == None else t_value
                    sent_kind = attach_sent[i][j][k][53]
                    sent_num = str(cond_num) + "." + attach_sent[i][j][k][68]
                    s_variable_kind = get_quick_variable_type(subj, variable_type)
                    o_variable_kind = get_quick_variable_type(obj, variable_type)
                    o2_variable_kind = get_quick_variable_type(obj2, variable_type)
                    new_subj, new_obj, new_obj2 = [subj, obj, obj2]
                    if s_variable_kind == 'agen':
                        new_subj = alpha
                    elif s_variable_kind == 'indefinite':
                        new_subj = subj + "'"
                    if o_variable_kind == 'agen':
                        new_obj = alpha
                    elif o_variable_kind == 'indefinite':
                        new_obj = obj + "'"
                    if o2_variable_kind == 'agen':
                        new_obj2 = alpha
                    elif o2_variable_kind == 'indefinite':
                        new_obj2 = obj2 + "'"
                    absolute_predicate = new_subj + relat + new_obj + relat2 + new_obj2
                    spredicate = alpha + t_value + relat + new_obj + relat2 + new_obj2
                    opredicate = new_subj + t_value + relat + alpha + relat2 + new_obj2
                    o2predicate = new_subj + t_value + relat + new_obj + relat2 + alpha
                    sent_parts = [new_subj, t_value, relat, new_obj, relat2, new_obj2]
                    attached_predicates.append([absolute_predicate, t_value, attach_sent[i][2]])
                    if relat == "I":
                        skind = findinlist(obj, abbreviations, 0, 1)
                        skind = kind_exception(skind)
                    else:
                        skind = get_class(relat, sent, 5)

                    object_properties = get_object_properties2(subj,
                                                               object_properties,
                                                               s_variable_kind,
                                                               spredicate,
                                                               skind,
                                                               sent_kind,
                                                               sent_num,
                                                               cond_num,
                                                               sent_parts)

                    if isvariable(obj) or obj == "i":
                        okind = get_class(relat, sent, 14)
                        object_properties = get_object_properties2(obj,
                                                                   object_properties,
                                                                   o_variable_kind,
                                                                   opredicate,
                                                                   okind,
                                                                   sent_kind,
                                                                   sent_num,
                                                                   cond_num,
                                                                   sent_parts)

                    if isvariable(obj2) or obj == 'i':
                        object_properties = get_object_properties2(obj2,
                                                                   object_properties,
                                                                   o2_variable_kind,
                                                                   o2predicate,
                                                                   'thing',
                                                                   sent_kind,
                                                                   sent_num,
                                                                   cond_num,
                                                                   sent_parts)

            else:
                break

    object_properties = sorted(object_properties, key=operator.itemgetter(1))
    object_properties = purge_thing_from_properties(object_properties)

    return attached_predicates, object_properties


def purge_thing_from_properties(object_properties):
    # if an object belongs to a class more specific than 'thing' then
    # we need not state that it belongs to the class 'thing'

    for object_property in object_properties:
        if 'thing' in object_property[2] and len(object_property[2]) > 1:
            object_property[2].remove("thing")
    return object_properties


def kind_exception(str1):
    # since everything belongs to the class 'whole' or 'part' these are not
    # genuine classes
    exceptions = ['whole', 'part']
    if str1 in exceptions:
        return 'thing'
    elif str1 == None:
        return 'thing2'
    return str1
    # if str1 equals none then that means the subject belongs to an indefinite
    # concept


def get_quick_variable_type(variable, variable_type):
    # this tells us what type a certain variable is after we already know what it is

    if variable == "":
        return ""
    defn = variable_type[1]
    gen = variable_type[0]
    indef = variable_type[2]

    if variable in gen:
        return 'agen'
    elif variable in defn:
        return 'definite'
    elif variable in indef:
        return 'indefinite'


def add_stan_sent(consistent):
    # this adds new sentences to the total_sent list

    if not consistent:
        return False

    add_to_total_sent("", "")
    add_to_total_sent("", "________________")
    add_to_total_sent("", "NONSTANDARD SENTENCES")
    i = 0
    while i < len(detach_sent):
        detach_sent[i][54] = is_standard(detach_sent[i])
        if not detach_sent[i][54]:
            add_to_total_sent(detach_sent[i][58], detach_sent[i][2] + detach_sent[i][72])
            del detach_sent[i]
        else:
            i += 1

    add_to_total_sent("", "")
    add_to_total_sent("", "________________")
    add_to_total_sent("", "STANDARD ATTACHED SENTENCES")

    for sent in attach_sent:
        add_to_total_sent(sent[2], sent[37])

    add_to_total_sent("", "")
    add_to_total_sent("", "________________")
    add_to_total_sent("", "STANDARD DETACHED SENTENCES")

    for sent in detach_sent:
        add_to_total_sent(sent[58], sent[2] + sent[72])
        consistent = check_reflexivity(sent)
        if not consistent:
            return consistent

    return True


def asymmetry(all_sent, str1, str2):
    for i in range(len(all_sent)):
        if all_sent[i][46] != "x":
            if all_sent[i][5] == str1 or all_sent[i][5] == str2:
                if all_sent[i][14] == str1 or all_sent[i][14] == str2:
                    return False
    return True


def is_in_md(list1, i, str1, bool1=False, k=0):
    if not bool1:
        for j in range(len(list1)):
            if list1[j][i] == str1:
                return True
        return False
    else:
        for j in range(len(list1)):
            if j != k:
                if list1[j][i] == str1:
                    return True
        return False


def check_dimension2d(list1, i, j, str1):
    for k in range(len(list1[i])):
        if list1[k][i][j] == str1:
            return True
    return False


def name_conditional(list1):
    # when using this function the list1 must be outputted from find_sentences
    # and you must use the 0th element since those sentences do not have negation signs
    global prop_var
    global prop_name
    skel_string = list1[5]
    list2 = []

    for i in range(1, len(list1[6])):
        if list1[6][i] != None and list1[6][i] != "":
            str3 = list1[0][i]
            str3 = remove_outer_paren(str3)
            str4 = copy.copy(str3)
            str4 = str4.replace(" ", "")
            str1 = findinlist(str4, prop_name, 1, 0)
            if str1 != None:
                str2 = list1[1][i] + str1
                list2.append(str1)
                skel_string = skel_string.replace(list1[6][i][1], str2)
            else:
                str2 = prop_var[0]
                list2.append(str2)
                del prop_var[0]
                prop_name.append([str2, str4, str3])
                str2 = list1[1][i] + str2
                skel_string = skel_string.replace(list1[6][i][1], str2)
    return [skel_string, list2]


def tilde_removal(str1):
    str4 = ""
    if str1.find("~") > -1:
        str4 = "~"
        str1 = str1.replace("~", "")
    return [str1, str4]


def tilde_removal2(str1):
    j = 0
    if str1[:2] == "~(":
        for i in range(len(str1)):
            str2 = str1[i:i + 1]
            if str2 == "(":
                j += 1
            elif str2 == ")":
                j -= 1
            if j == 0 and i > 0:
                if i == len(str1) - 1:
                    str1 = str1[1:]
                    str4 = "~"
                    return [str1, str4]
                else:
                    return [str1, ""]
    elif str1[0] == "~" and str1[1].islower() and os(str1):
        str1 = str1[1:]
        str4 = "~"
    else:
        str4 = ""
    return [str1, str4]


def get_conjuncts(str1, bool1=False):
    # remove outparent if true
    global subscripts
    arr1 = []
    if bool1:
        str1 = remove_outer_paren(str1)

    j = 0
    k = 1
    for i in range(len(str1)):
        str2 = str1[i:i + 1]
        str5 = str1[i + 1:i + 2]
        if i > 0:
            str4 = str1[i - 1:i]
        else:
            str4 = ""
        if str2 == "(":
            j += 1
        elif str2 == ")":
            j -= 1

        if j == 1 and str2 == "(" and str4 != "~":
            k = i
        elif j == 1 and str4 == "~" and str2 == "(":
            k = i - 1

        if (j == 0 and str2 == ")") or (j == 0 and str2.islower()):
            if str2 == ")":
                str3 = str1[k:i + 1]
            else:
                if str1[i - 1:i] == "~" and str5 not in subscripts:
                    str3 = "~" + str2
                elif str1[i - 1:i] != "~" and str5 not in subscripts:
                    str3 = str2
                elif str1[i - 1:i] != "~" and str5 in subscripts:
                    str3 = str2 + str5
                    str2 = str2 + str5
                elif str1[i - 1:i] == "~" and str5 in subscripts:
                    str3 = "~" + str2 + str5
                    str2 = str2 + str5

            k = 1
            str3 = str3.strip()
            arr1.append(str3)

    return arr1


def determine_if_all_cond_4_detach_met(g, k):
    # this loops through all the sentences in the antecedent or the consequent
    # and determines if they all have been detached

    conditions = copy.deepcopy(attach_sent[g][k])
    del conditions[0]
    ancestors = []
    sent_type = attach_sent[g][3]
    done = False
    i = 0
    while i < len(attach_sent[g][k]) - 1 and not done:
        i += 1
        for sent in detach_sent:
            temp = sent[1]

            if sent[1] == 'h' + l1:
                bb = 8
            if sent[1] == attach_sent[g][k][i][0] \
                    and sent[2] == attach_sent[g][k][i][1]:
                ancestors.append(sent[58])
                del conditions[0]
                if conditions == []:
                    output = ["", ancestors]
                    done = True
                break

            elif sent[1] == attach_sent[g][k][i][0] \
                    and sent[2] != attach_sent[g][k][i][1]:
                if (sent_type == "e" and k == 0) or k == 1:
                    done = True
                    output = ['a sentence was negated', i]
                elif sent_type == 'c' and k == 0:
                    done = True
                    output = ["", ""]
                break
        else:
            done = True
            output = ["", ""]

    return output


def detach1(str1, negated_conjunction):
    global st_log_time
    b = time.time()
    if str1 == 'do not use modus tollens':
        kind = '~MT'
    else:
        kind = 'MT'

    if attach_sent == []:
        return True
    consistent = True
    set_of_det_sent = []
    r = -1
    while consistent and r < len(detach_sent) - 1:
        r += 1
        if r == 46:
            bb = 8
        g = -1
        while consistent and g < len(attach_sent) - 1:
            g += 1
            if attach_sent[g][26] != "not new":
                k = -1
                while consistent and k < 1:
                    k += 1
                    sent_type = attach_sent[g][3]
                    if attach_sent[g][3] == "e" or attach_sent[g][3] == "c":
                        temp_detach_sent = detach_sent[r][1]
                        det_tvalue = detach_sent[r][2]
                        temp_attach_sent = attach_sent[g][k][0][0]
                        att_tvalue = attach_sent[g][k][0][1]

                        if temp_detach_sent == 'l':
                            bb = 8
                        if temp_detach_sent == temp_attach_sent:
                            if det_tvalue == att_tvalue and (k == 0 or (k == 1 and sent_type == "e")):
                                rule = "EE" if sent_type == 'e' else "MP"
                                if len(attach_sent[g][k]) == 1:

                                    # if k = 0 and sent_type = 'c' and set_of_det_sent = []
                                    # then p & (p > q)
                                    # if k = 0 and sent_type = 'e' and set_of_det_sent = []
                                    # then p & (p <> q)
                                    # if k = 1 and sent_type = 'e' and set_of_det_sent = []
                                    # then p & (q <> p)
                                    # if k = 0 and sent_type = 'c' and set_of_det_sent != []
                                    # then p & r & ((p & r) > q)
                                    # if k = 0 and sent_type = 'e' and set_of_det_sent != []
                                    # then p & r & ((p & r) <> q)
                                    # if k = 1 and sent_type = 'e' and set_of_det_sent != []
                                    # then p & r & (q <> (p & r))

                                    consistent, g, k = detach2(k, r, g, rule,
                                                               set_of_det_sent, negated_conjunction)
                                    if g > len(attach_sent) - 1:
                                        break
                                else:
                                    output = determine_if_all_cond_4_detach_met(g, k)
                                    if output[0] == 'a sentence is negated':
                                        rule = "EN" if sent_type == 'e' else "MT"
                                        consistent, g, k = detach2(k, output[1], g, rule, [], negated_conjunction)
                                        if g > len(attach_sent) - 1:
                                            break
                                    elif output[1] != "":
                                        output[1].insert(0, detach_sent[r][58])
                                        consistent, g, k = detach2(k, r, g, rule, output[1], negated_conjunction)
                                        if g > len(attach_sent) - 1:
                                            break
                            else:
                                if k == 0:
                                    t = 27
                                else:
                                    t = 28
                                if k == 0 and sent_type == 'c':
                                    pass
                                elif kind == "MT" and attach_sent[g][t] and \
                                        ((k == 0 and sent_type == 'e') or (k == 1)):
                                    rule = "EN" if sent_type == 'e' else "MT"
                                    consistent, g, k = detach2(k, r, g, rule, set_of_det_sent, negated_conjunction)
                                    if g > len(attach_sent) - 1:
                                        break

    c = time.time()
    c = c - b
    st_log_time += c

    return consistent


def detach2(k, r, g, rule, set_of_det_sent, negated_conjunction):
    global sn

    anc1 = attach_sent[g][2]
    anc2 = detach_sent[r][58]
    sn += 1
    if sn == 28:
        bb = 8
    if k == 0:
        m = 41
        h = 43
        t = 8
        s = 1
        n = 35
    else:
        m = 40
        h = 42
        t = 7
        s = 0
        n = 34

    introduce_conjunction(set_of_det_sent)

    if attach_sent[g][t][0] == 't':
        bb = 8

    consistent = add_to_total_sent_consist(sn, attach_sent[g][m][0], attach_sent[g][t][0],
                                           attach_sent[g][m][1], rule,
                                           anc1, anc2, negated_conjunction)

    if (rule == "MT" or rule == "EN") and len(attach_sent[g][h]) > 1:
        negated_conjunction.append([attach_sent[g][s], attach_sent[g][m]], attach_sent[g][t])
        g = 4 / 0
    else:
        if len(attach_sent[g][h]) == 1:
            if consistent:
                if os(attach_sent[g][s][0][0]):
                    list3 = attach_sent[g][n][0]
                    list3[58] = sn
                    is_in_detach_sent = isinmdlist(list3[42], detach_sent, 42)
                    if not is_in_detach_sent:
                        detach_sent.append(list3)
                        if list3[1] == 'l':
                            bb = 8
                else:
                    for lst in attach_sent[g][39]:
                        lst[2] = sn
                        attach_sent.append(lst)
        else:
            consistent = eliminate_conjuncts(g, r, h, negated_conjunction)
    del attach_sent[g]
    if g + 1 == len(attach_sent):
        g -= 1
    k = -1
    return consistent, g, k


def introduce_conjunction(set_of_det_sent):
    global sn
    full_sent = ""
    if set_of_det_sent != []:
        sn += 1
        for num in set_of_det_sent:
            d = findposinmdlistint(num, detach_sent, 58)
            if d == -1: g = 4 / 0
            if full_sent == "":
                full_sent = detach_sent[d][0]
                abbrev_sent = detach_sent[d][42]
                anc = str(detach_sent[d][58])
            else:
                full_sent += " & " + detach_sent[d][0]
                abbrev_sent += " & " + detach_sent[d][42]
                anc += "," + str(detach_sent[d][58])
        full_sent = "(" + full_sent + ")"
        abbrev_sent = "(" + abbrev_sent + ")"
        add_to_total_sent(sn, full_sent, abbrev_sent, "", "&I", anc)


def eliminate_conjuncts(g, r, h, negated_conjunction):
    # if the detached sentences are a conjunction then this function
    # places each individual conjunct into the total_sent and detach_sent list

    global sn
    num = copy.copy(sn)
    if h == 43:
        m = 41
        k = 1
        n = 35
        c = 8
    else:
        m = 40
        k = 0
        n = 34
        c = 7
    conjunct_list = attach_sent[g][k]

    for i in range(len(conjunct_list)):
        sn += 1
        consistent = add_to_total_sent_consist(sn, attach_sent[g][h][i][0], attach_sent[g][k][i][0],
                                               attach_sent[g][h][i][1], "&E", num, "", negated_conjunction)
        if consistent:
            if os(conjunct_list[i][0]):
                d = findposinmd_alert_error(conjunct_list[i][0], attach_sent[g][n], 1)
                sent_parts = attach_sent[g][n][d]
                sent_parts[58] = sn
                is_in_detach_sent = isinmdlist(sent_parts[42], detach_sent, 42)
                if not is_in_detach_sent:
                    detach_sent.append(sent_parts)
            else:
                d = findposinmd_alert_error(conjunct_list[i][0], attach_sent[g][39], 4)
                list2 = attach_sent[g][39][d]
                list2[2] = sn
                attach_sent.append(list2)
        else:
            break

    return consistent


def add_to_total_sent_consist(num, str1, str2, tvalue, rule, anc1, anc2, negated_conjunction):
    list2 = [""] * 9
    list2[0] = num
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    total_sent.append(list2)
    consistent = check_consistency(negated_conjunction)

    return consistent


def add_to_total_sent(num, str1, str2="", tvalue="", rule="", anc1="", anc2=""):
    list2 = [""] * 9
    list2[0] = num
    list2[1] = str1
    list2[2] = str2
    list2[3] = tvalue
    list2[4] = rule
    list2[5] = anc1
    list2[6] = anc2
    total_sent.append(list2)


def check_consistency(negated_conjunction):
    new_sent_abbr = total_sent[-1][2]
    tvalue = total_sent[-1][3]
    for i in range(len(total_sent) - 2, -1, -1):
        if total_sent[i][2] == new_sent_abbr and total_sent[i][3] != tvalue:
            build_contradiction(i)
            return False

    for lst in negated_conjunction:
        for sent in lst[0]:
            if sent[0] == new_sent_abbr and sent[1] == tvalue:
                g = 4 / 0
                del sent[0]
                del sent[1]
                ancestors = sent[4]
                ancestors.append(sn)
                if lst[0] == []:
                    bb = 8
                    pass
                    #

    return True


def build_contradiction(i):
    global sn
    sn += 1
    str1 = total_sent[-1][1] + " & ~" + total_sent[i][1]
    str2 = total_sent[-1][2] + " & ~" + total_sent[i][2]
    total_sent.append([sn, str1, str2, "", "&E", total_sent[-1][0], total_sent[i][0], "", ""])
    sn += 1
    total_sent.append([sn, bottom, "", "", bottom + "I", sn - 1, "", "", ""])


def disjunction_heirarchy(str5, d, new_disj=False):
    global prop_name
    global sn, pn

    if d > len(attach_sent) - 1:
        return
    if iff in str5 or conditional in str5:
        return

    str5 = enclose(str5)
    def_info = find_sentences(str5)
    mainc = def_info[4][0][1]
    list2 = [""] * 50
    n = 7
    if mainc == xorr:
        list2[3] = 'x'
    else:
        list2[3] = 'd'
    if attach_sent == [] or new_disj:
        list2[2] = pn
    else:
        list2[2] = attach_sent[d][2]
    list2[5] = ""
    list2[4] = def_info[0][0]  # fix this
    sentences = []

    for i in range(len(def_info[0])):
        if os(def_info[0][i]):
            siblings = []
            list3 = [None] * 9
            n += 1
            # str1 = findinlist(def_info[0][i],prop_name,1,0)
            str2 = def_info[4][i][0][:-1]
            g = findposinlist(str2, def_info[4], 0)
            parent = def_info[0][g]
            if def_info[4][g][1] == "&":
                list3[2] = 'c'
            elif def_info[4][g][1] == xorr:
                list3[2] = 'x'
            else:
                list3[2] = 'd'
            if len(str2) > 1:
                str3 = def_info[4][i][0][:-2]
                g = findposinlist(str3, def_info[4], 0)
                gparent = def_info[0][g]
            else:
                gparent = parent
            list3[1] = def_info[4][i][0]
            list3[5] = parent
            list3[6] = gparent
            list3[0] = [def_info[0][i], def_info[1][i]]
            # fix this
            b = parent.count(xorr)
            c = parent.count(idisj)
            if c > 1 or b > 1:
                list3[7] = 2
            else:
                list3[7] = 1
            sent_num = def_info[4][i][0]
            m = len(sent_num)
            for j in range(len(def_info[4])):
                if len(def_info[4][j][0]) == m and def_info[4][j][0][:-1] == str2 \
                        and j != i:
                    siblings.append([def_info[0][j], def_info[1][j]])  # fix this
            list3[4] = siblings
            list2[n] = list3
            sentences.append(list3[0][0])
            if list3[0][0] not in rel_conj:
                rel_conj.append(list3[0][0])
    list2[36] = def_info

    if attach_sent == [] or new_disj:
        list2[38] = sentences
        attach_sent.append(list2)
    else:
        list2[38] = sentences
        list2[2] = attach_sent[d][2]
        list2[37] = attach_sent[d][37]
        attach_sent[d] = list2


def proper_spacing(str1):
    str1 = str1.replace(" ", "")
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional, " " + conditional + " ")
    str1 = str1.replace(idisj, " " + idisj + " ")
    str1 = str1.replace(xorr, " " + xorr + " ")
    str1 = str1.replace("&", " & ")
    return str1


def proper_spacing2(str1):
    str1 = str1.replace(iff, " " + iff + " ")
    str1 = str1.replace(conditional, " " + conditional + " ")
    str1 = str1.replace(idisj, " " + idisj + " ")
    str1 = str1.replace(xorr, " " + xorr + " ")
    str1 = str1.replace("&", " & ")
    return str1


def bad_paren(str1):
    if str1.find("(") == -1:
        return str1
    # we first must get rid of strings of the following form ((p) & s)
    for i in range(len(str1)):
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        if i > 1:
            str4 = str1[i - 2:i - 1]
        else:
            str4 = ""
        str5 = str1[i + 1:i + 2]
        if str2.islower() and str3 == "(" and str5 == ")":
            str1 = str1[:i - 1] + str1[i:i + 1] + str1[i + 2:]
        elif str2.islower() and str4 == "(" and str3 == "~" and str5 == ")":
            str1 = str1[:i - 2] + str1[i - 1:i + 1] + str1[i + 2:]

    str1 = enclose(str1)
    list1 = find_sentences(str1)
    mstr = list1[3][0]
    for i in range(1, len(list1[3])):
        if list1[4][i][1] != "":
            mc = list1[4][i][1]
            ostr = list1[3][i]
            str2 = list1[4][i][0][:-1]
            prcnt = findinlist(str2, list1[4], 0, 1)
            if mc == prcnt:
                nstr = remove_outer_paren(ostr)
                nstr = remove_outer_paren(nstr)
                mstr = mstr.replace(ostr, nstr)
    return mstr


def unenclose(str1):
    # this removes ( ) from around a sentence abbreviation
    i = -1
    if "(" not in str1:
        return str1

    while i < len(str1) - 1:
        i += 1
        str2 = str1[i:i + 1]
        str3 = str1[i - 1:i]
        str4 = str1[i + 1:i + 2]
        if str2.islower() and str3 != "~" and str4 not in subscripts:
            str1 = str1[:i - 1] + str2 + str1[i + 2:]
        elif str2.islower() and str3 == "~" and str4 not in subscripts:
            str1 = str1[:i - 2] + str3 + str2 + str1[i + 2:]
        if str2.islower() and str3 != "~" and str4 in subscripts:
            str1 = str1[:i - 1] + str2 + str4 + str1[i + 3:]
        elif str2.islower() and str3 == "~" and str4 in subscripts:
            str1 = str1[:i - 2] + str3 + str2 + str4 + str1[i + 3:]

    return str1


def new_disjunct(str1, ng, n, candd, conjt, anc1, anc2, anc3=None, anc4=None, kind=0, rule=""):
    global sn, pn
    list2 = mainconn(str1)
    if kind == 1:
        del attach_sent[n]
        consistent = new_prop(str1, ng, "&I", anc1, anc2, anc3, anc4)
        return consistent
    elif kind == 2:
        consistent = new_prop(str1, ng, "&I", anc1, anc2, anc3, anc4)
        return consistent
    else:
        if os(str1):
            del attach_sent[n]
            str1 = remove_outer_paren(str1)
            list1 = tilde_removal2(str1)
            str1 = list1[0]
            consistent = new_prop(str1, list1[1], rule + "E", anc1, anc2)
            candd.append([pn, list1[0], list1[1]])
            conjt.append([pn, list1[0], list1[1]])
            return consistent
        elif list2[0] == "&":
            del attach_sent[n]
            str1 = remove_outer_paren(str1)
            new_prop(str1, ng, rule + "E", anc1, anc2)
            g = copy.copy(pn)
            list3 = get_conjuncts(str1)
            for i in range(len(list3)):
                list4 = tilde_removal2(list3[i])
                consistent = new_prop(list4[0], list4[1], "&E", g, "")
                if no_output == False:
                    return consistent
                if list3[i].find(idisj) > -1:
                    disjunction_heirarchy(list4[0], n, True)
                else:
                    candd.append([pn, list4[0], list4[1]])
                    conjt.append([pn, list4[0], list4[1]])
            return True
        else:
            consistent = new_prop(str1, ng, idisj + "E", anc1, anc2)
            if consistent == False:
                return consistent
            if ng == "~":
                str1 = ng + str1
            else:
                str1 = remove_outer_paren(str1)
            attach_sent[n][2] = pn
            disjunction_heirarchy(str1, n, False)
            return True


def xorr_elim(n, i, parent, grandparent, whole_d, candd, anc1, anc2, conjt, kind=0):
    str9 = ""
    de_mor = False
    if kind == 0:
        for r in range(len(attach_sent[n][i][4])):
            if r != i:
                if not os(attach_sent[n][i][4][r][0]):
                    de_mor = True
                if str9 == "":
                    str9 += "~" + attach_sent[n][i][4][r][1] + attach_sent[n][i][4][r][0]
                else:
                    str9 += " & ~" + attach_sent[n][i][4][r][1] + attach_sent[n][i][4][r][0]
    else:
        grandp2 = copy.copy(grandparent)
        grandp2 = grandp2.replace(parent, "")
        for r in range(8, 38):
            if attach_sent[n][r] == "":
                break
            if attach_sent[n][r][0][0] in grandp2:
                if str9 == "":
                    str9 += "~" + attach_sent[n][r][0][1] + attach_sent[n][r][0][0]
                else:
                    str9 += " & ~" + attach_sent[n][r][0][1] + attach_sent[n][r][0][0]
    g = copy.copy(pn)
    if parent != grandparent:
        str9 = remove_outer_paren(str9)
        if grandparent == whole_d:
            mc = mainconn(str9)
            if mc[0] == '&':
                consistent = xorr_elim2(str9, candd, conjt, anc1, anc2)
                if consistent == False:
                    return consistent
            else:
                list4 = tilde_removal(str9)
                consistent = new_prop(list4[0], list4[1], xorr + "E", anc1, anc2)
                if consistent == False:
                    return consistent
        else:
            str9 = "(" + str9 + ")"
            if kind == 0:
                str9 = grandparent.replace(parent, str9)
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~", "")
                    consistent = new_prop(str9, "", "~~E", pn, "")

            else:
                str9 = whole_d.replace(grandparent, str9)
                str9 = bad_paren(str9)
                consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
                g = copy.copy(pn)
                if str9.find("~~") > -1:
                    str9 = str9.replace("~~", "")
                    consistent = new_prop(str9, "", "~~E", g, "")
                    if consistent == False:
                        return consistent
                disjunction_heirarchy(str9, n, True)
                del attach_sent[n]
            if de_mor:
                list1 = demorgan(all_sent, attach_sent, total_sent, detach_sent, candd, True, str9, pn)
                consistent = list1[0]
                attach_sent = list1[1]
                if consistent == False:
                    return consistent
            else:
                if str9.find(idisj) > -1 or str9.find(xorr) > -1:
                    disjunction_heirarchy(str9, n, True)
                consistent = True
    else:
        # this does not account for the case where the parent == grandparent but
        # grandparent does not == whole d
        consistent = xorr_elim2(str9, candd, conjt, anc1, anc2)
    return consistent


def xorr_elim2(str9, candd, conjt, anc1, anc2):
    str9 = bad_paren(str9)
    consistent = new_prop(str9, "", xorr + "E", anc1, anc2)
    if consistent == False:
        return False
    if str9.find("~~") > -1:
        str9 = str9.replace("~~", "")
        consistent = new_prop(str9, "", "~~E", pn, "")
        if consistent == False:
            return consistent
    list3 = get_conjuncts(str9)
    g = copy.copy(pn)
    for b in range(len(list3)):
        list4 = tilde_removal2(list3[b])
        list4[0] = remove_outer_paren(list4[0])
        consistent = new_prop(list4[0], list4[1], "&E", g, "")
        if consistent == False:
            return consistent
        if not os(list3[b]):
            if list4[1] == "~":
                list1 = demorgan(all_sent, attach_sent, total_sent, detach_sent, candd, list3[b], pn, "&E")
                consistent = list1[0]
                attach_sent = list1[1]
                if consistent == False:
                    return False
            else:
                disjunction_heirarchy(list4[0], n, True)
        else:
            candd.append([pn, list4[0], list4[1]])
            conjt.append([pn, list4[0], list4[1]])

    return True


def disjunction_elimination(candd, kind=""):
    bool1 = False
    bool2 = False
    global sn, pn
    global rel_conj

    for i in range(len(attach_sent)):
        if attach_sent[i][8] == "":
            disjunction_heirarchy(attach_sent[i][4], i)
    i = -1
    conjt = copy.deepcopy(candd)

    while i < len(conjt) - 1:
        i += 1
        if conjt[i][1] not in rel_conj:
            del conjt[i]
            i -= 1
    d = -1
    while d < len(conjt) - 1:
        d += 1
        str2 = conjt[d][2]
        conj = conjt[d][1]
        if conj == 'q':
            bb = 7
        if d == 42:
            bb = 7
        anc1 = conjt[d][0]
        n = -1
        while n < len(attach_sent) - 1:
            if bool1:
                bool1 = False
                d = -1
                break
            n += 1
            if n == 7:
                bb = 7
            i = 7
            while attach_sent != []:
                if bool2:
                    bool2 = False
                    break
                i += 1
                if conj not in attach_sent[n][38] and iff not in attach_sent[n][4] \
                        and conditional not in attach_sent[n][4]:
                    break
                else:
                    if attach_sent[n][i] == "":
                        break
                    whole_d = attach_sent[n][4]
                    anc2 = attach_sent[n][2]
                    str3 = attach_sent[n][i][0][0]
                    # 'pos or neg'
                    str4 = attach_sent[n][i][0][1]
                    # 'disjunct or conjunct
                    str5 = attach_sent[n][i][2]
                    # 'disjunct number
                    str6 = attach_sent[n][i][1]

                    if conj == str3:
                        grandparent = attach_sent[n][i][6]
                        parent = attach_sent[n][i][5]
                        parent2 = copy.copy(parent)
                        parent3 = copy.copy(parent)
                        str7 = " " + idisj + " "
                        str7a = " " + xorr + " "
                        if str2 == str4 and str5 == "d":
                            # 'if the disjuncts are not embedded within a conjunct then the disjunction
                            # is simply deleted

                            del attach_sent[n]
                            if parent != grandparent:
                                conj = str2 + conj
                                str8 = whole_d.replace(parent, conj)
                                disjunction_heirarchy(str8, n)
                            bool1 = True
                            n = -1
                            break

                        elif str2 == str4 and str5 == "x":

                            consistent = xorr_elim(n, i, parent, grandparent, whole_d, candd, anc1, anc2)
                            if not consistent:
                                return False
                            del attach_sent[n]
                            bool2 = True
                            bool1 = True
                            d = -1
                        elif str2 == str4 and str5 == "c":
                            list2 = []
                            list2.append([conj, str2])
                            anc3 = ""
                            anc4 = ""
                            list11 = attach_sent[n][i][4]
                            f = -1
                            while f < len(list11) - 1:
                                mc = mainconn(grandparent)
                                f += 1

                                for e in range(len(candd)):
                                    anc5 = candd[e][0]
                                    # since it's too hard to program, if the sibling is a disjunct then we just
                                    # ignore this

                                    if list11[f][0].find(idisj) > -1 or list11[f][0].find(xorr) > -1:
                                        break
                                    else:
                                        if candd[e][1] == list11[f][0]:
                                            if candd[e][2] == list11[f][1]:
                                                list2.append([list11[f][0], list11[f][1]])
                                                if len(list2) == 2:
                                                    anc3 = anc5
                                                elif len(list2) == 3:
                                                    anc4 = anc5
                                                del list11[f]
                                                if list11 == []:
                                                    str3 = build_sent_list2(list2)
                                                    if mc[0] == xorr:
                                                        new_prop(str3, "", "&I", anc1, anc3, anc4)
                                                        consistent = xorr_elim(n, i, parent, grandparent, whole_d,
                                                                               candd,
                                                                               anc1, anc2, conjt, 1)
                                                        if not consistent:
                                                            return False
                                                    else:
                                                        # if the conjunct is not embedded within another conjunct
                                                        # then the disjunct is simply deleted
                                                        if whole_d == grandparent:
                                                            consistent = new_disjunct(str3, "", n,
                                                                                      candd, conjt, anc1, anc3,
                                                                                      anc4, anc5, 1)

                                                        else:
                                                            str8 = whole_d.replace(grandparent, parent2)
                                                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                                                str8 = bad_paren(str8)
                                                                consistent = new_disjunct(str3, "", n,
                                                                                          candd, conjt, anc1, "",
                                                                                          anc3, anc4, 2)

                                                            consistent = new_disjunct(str8, "", n, candd, conjt, pn - 1,
                                                                                      anc2)
                                                            if not consistent:
                                                                return False
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break
                                                else:
                                                    f -= 1
                                                    break

                                            elif candd[e][2] != list11[f][1]:
                                                mc = mainconn(grandparent)
                                                if mc[0] == idisj:
                                                    rule = idisj
                                                    str7 = " " + idisj + " "
                                                else:
                                                    str7 = " " + xorr + " "
                                                    rule = xorr
                                                r = grandparent.find(parent)
                                                if r > 1:
                                                    parent = str7 + parent
                                                else:
                                                    parent = parent + str7
                                                anc1 = candd[e][0]
                                                str9 = grandparent.replace(parent, "")
                                                str8 = whole_d.replace(grandparent, str9)
                                                if str8.find("(") > -1 and (
                                                                str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                                    str8 = bad_paren(str8)
                                                consistent = new_disjunct(str8, "", n, candd, conjt, anc1, anc2, None,
                                                                          None,
                                                                          0, rule)

                                                if not consistent:
                                                    return False
                                                else:
                                                    list11 = []
                                                    bool1 = True
                                                    bool2 = True
                                                    n = 0
                                                    d = -1
                                                    break

                        elif str2 != str4 and str5 == "c":
                            mc = mainconn(attach_sent[n][i][6])
                            if mc[0] == idisj:
                                str6 = str7 + parent
                                rule = idisj
                            else:
                                rule = xorr
                                str6 = str7a + parent
                            if grandparent.find(str6) > -1:
                                parent = str6
                            else:
                                parent = parent + str7
                            str9 = grandparent.replace(parent, "")
                            str8 = whole_d.replace(grandparent, str9)
                            if str8.find("(") > -1 and str8.find(idisj) > -1:
                                str8 = bad_paren(str8)
                            consistent = new_disjunct(str8, "", n, candd, conjt, anc1, anc2, None, None, 0, rule)
                            if not consistent:
                                return False
                            bool1 = True
                            n = -1
                            break

                        elif str2 != str4 and (str5 == "d" or str5 == 'x'):
                            # if the disjunct is a triple disjunct then enter below
                            if str5 == 'd':
                                rule = idisj
                            else:
                                rule = xorr
                            if attach_sent[n][i][7] > 1:
                                str6 = str4 + str3 + " " + rule + " "
                                if parent.find(str6) > -1:
                                    str5 = str6
                                else:
                                    str5 = " " + rule + " " + str4 + str3
                                str9 = parent.replace(str5, "")
                                str8 = whole_d.replace(parent, str9)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                consistent = new_disjunct(str8, "", n, candd, conjt,
                                                          anc1, anc2, None, None, 0, rule)

                                if not consistent:
                                    return False
                                bool1 = True
                                n = -1
                                break

                            else:
                                str3 = attach_sent[n][i][4][0][0]  # ddd
                                str4 = attach_sent[n][i][4][0][1]
                                str5 = str4 + str3
                                str8 = whole_d.replace(parent, str5)
                                if str8.find("(") > -1 and (str8.find(idisj) > -1 or str8.find(xorr) > -1):
                                    str8 = bad_paren(str8)
                                consistent = new_disjunct(str8, "", n, candd,
                                                          conjt, anc1, anc2, None, None, 0, rule)
                                if not consistent:
                                    return False
                                bool1 = True
                                n = -1
                                break
    return True


def extract_list(list1, d):
    list2 = []
    for i in range(len(list1)):
        list2.append(list1[i][d])
    return list2


def use_statement_logic(kind=""):
    global st_log_time
    b = time.time()
    list1 = detach1(kind)
    consistent = list1[0]
    attach_sent = list1[1]
    # if consistent == False:
    #     return [False, attach_sent]
    # if kind != 2:
    #     list1 = disjunction_elimination(all_sent, attach_sent, detach_sent, \
    #                                     candd, total_sent, kind)
    #     consistent = list1[0]
    #     attach_sent = list1[1]

    c = time.time()
    d = c - b
    st_log_time += d
    return consistent


def add_outer_paren(str1):
    str1 = remove_outer_paren(str1)
    return "(" + str1 + ")"


def oc(str1):
    # this function determines if there is only one connective which is not &
    # it is used to weed out attach_sent from the candd list
    list1 = [conditional, idisj, iff, xorr]
    j = 0
    for i in range(len(str1)):
        str2 = str1[i:i + 1]
        if str2 in list1:
            j += 1
    if j == 1:
        return True
    else:
        return False


def repeat_relations():
    # this is for those definitions in which the same relation is related to two different
    # general variables
    a = ["group", "x"]
    b = ["member", 'z']
    e = ['every', "y"]
    f = ['personhood', 'y']

    final_list = [a, b, e, f]
    return final_list


def populate_sentences(p):
    global result_data
    global excel
    bool1 = False
    bool2 = False
    bool3 = True
    first_sent = False
    sent = []
    test_sent = []
    g = 0

    if not excel:

        for row in w4:
            p += 1
            if row[1] == "" and bool2 == True and not bool3:
                break
            elif row[1] == 'stop':
                break
            elif row[1] == "" and not bool3:
                test_sent.append(sent)
                sent = []
                g = 0
                bool1 = False
                bool2 = True
                first_sent = False
            elif row[1] == "":
                pass
            elif row[1][0] == "*":
                bool3 = True
            elif row[1] != "" and bool1 == False:
                bool3 = False
                str2 = row[1]
                g += 1
                if len(sent) == 0:
                    if str2.find(bottom) > -1:
                        tv = 'co'
                        str2 = str2[2:]
                    else:
                        tv = 'ta'
                else:
                    tv = ""
                str2.strip()
                sent.append([g, str2, '', tv])
                if not first_sent:
                    result_data['text_' + str(p - 2) + '_1'] = len(test_sent)
                first_sent = True
                bool2 = False
    else:
        for row in w4.rows:
            p += 1
            if row[2].value == None and bool2 == True and not bool3:
                break
            elif row[2].value == 'stop':
                break
            elif row[2].value == None and not bool3:
                test_sent.append(sent)
                sent = []
                g = 0
                bool1 = False
                bool2 = True
                first_sent = False
            elif row[2].value == None:
                pass
            elif row[2].value[0] == "*":
                bool3 = True
            elif row[2].value != None and bool1 == False:
                bool3 = False
                str2 = row[2].value
                g += 1
                if len(sent) == 0:
                    if str2.find(bottom) > -1:
                        tv = 'co'
                        str2 = str2[2:]
                    else:
                        tv = 'ta'
                else:
                    tv = ""
                str2.strip()
                sent.append([g, str2, row[0].value, tv])
                if not first_sent:
                    w4.cell(row=p - 1, column=2).value = len(test_sent)
                first_sent = True
                bool2 = False

    return [test_sent, p]


def get_number_of_sent_to_prove(strt,stp,order,nonlinear):

    if nonlinear:
        return order
    else:
        return [x for x in range(strt,stp)]


def calculate_time_statistics(st):

    global instan_used, instan_time, lemmas_used

    if nonlinear:
        num_of_sent = len(order)
    else:
        num_of_sent = (stop - strt)

    en = time.time()
    g = (en - st) / num_of_sent
    dd = st_log_time / num_of_sent
    tot_tim2 = time.time()
    total = tot_tim2 - tot_tim
    gg = time_spent_defining / num_of_sent
    if lemmas_used == 0: lemmas_used = 1

    if instan_used != 0:
        ee = instan_time / instan_used
    else:
        ee = 0
    print("average " + str("{0:.4f}".format(g)))
    print("time used in statement logic " + str("{0:.4f}".format(dd)))
    print("time spent reducing " + str("{0:.4f}".format(time_spent_reducing/(num_of_sent))))
    print("time used in lemma function " \
          + str("{0:.5f}".format(time_spent_in_lemma_function/lemmas_used)))
    print("time used in instantiation " + str("{0:.4f}".format(ee)))
    print("time used in change variables function " + str("{0:.4f}".format(gg)))
    print("total " + str("{0:.3f}".format(total)))



def step_one(sent):

    divide_sent(sent)

    eliminate_redundant_words()

    replace_synonyms()

    replace_relations()

    eliminate_negative_determiners()

    word_sub()

def get_result(post_data, archive_id=None, request=None):

    global ws, w4, result_data, order
    global sn, total_sent, prop_name
    global all_sent, attach_sent, detach_sent, abbreviations
    global prop_var, variables, stop

    if not normal_proof:
        if archive_id:
            ws = Define3.object_properties.filter(archives_id=archive_id)
        else:
            archive = Archives.object_properties.latest('archives_date')
            ws = Define3.object_properties.filter(archives_id=archive.id)

    if not normal_proof:
        result_data = dict(post_data.iterlists())
        w4 = []
        index = 0
        while True:
            row = (post_data["text_" + str(index) + "_1"], post_data["text_" + str(index) + "_2"],
                   post_data["text_" + str(index) + "_3"])
            w4.append(row)
            if row[1] == "stop":
                break
            index += 1
        w4 = tuple(w4)

    if mysql:
        if archive_id:
            tw4 = Input.object_properties.filter(archives_id=archive_id)
        else:
            archive = Archives.object_properties.latest('archives_date')
            tw4 = Input.object_properties.filter(archives_id=archive.id)
        w4 = []
        for x in tw4:
            row = (x.col1, x.col2, x.col3)
            w4.append(row)
        w4 = tuple(w4)



    list1 = pop_sent('hey')
    for i in range(len(list1)):
        for j in range(len(list1[i])):
            list2 = tran_str(list1[i][j][1], 2)
            list1[i][j][1] = list2[0]
    test_sent = list1
    p = 1
    if not get_words_used:
        ex_dict = large_dict()
    else:
        ex_dict = ""

    build_dict(ex_dict)
    if stp == 0: stp = len(test_sent)
    order = get_number_of_sent_to_prove(strt, stp, order, nonlinear)

    st = time.time()
    if stp == 0:
        stp = len(test_sent)
    if not excel and not normal_proof:
        views.progressbar_send(request, 0, 100, 0, 1)
    for k in order:
        if not excel and not normal_proof:
            views.progressbar_send(request, strt, stp, k, 1)
        if k == 27:
            bb = 7
        st1 = time.time()
        prop_name = []
        total_sent = []
        all_sent = []
        attach_sent = []
        detach_sent = []
        abbreviations = []
        prop_var = copy.deepcopy(prop_var4)
        variables = copy.deepcopy(variables2)
        sn = test_sent[k][-1][0] + 1

        step_one(test_sent[k])

        step_two()

        consistent = step_three(test_sent[k][0][3])

        test_sent[k] = copy.deepcopy(total_sent)
        tot_prop_name.append(prop_name)

        if not consistent:
            print('False')
            stp = k + 1
            # break
        en1 = time.time()
        z = en1 - st1
        print(str(k) + " - " + str("{0:.3f}".format(z)))

    if stp == 0 and consistent:
        stp = k
    calculate_time_statistics(st)

    print_sent_full(test_sent, tot_prop_name)

    if django2:
        views.progressbar_send(request, 0, 100, 100, 2)
    if excel:
        pass  # Saved at last
    elif mysql:
        views.save_result(result_data)
    else:
        return result_data

get_result('hey')

if print_to_doc:
    wb4.save('/Users/kylefoley/Desktop/inference engine/temp_proof.xlsx')
if get_words_used:
    wb5.save('/Users/kylefoley/Desktop/inference engine/dictionary4.xlsx')





